# -*- coding: utf-8 -*-
"""
Log Merge Tool - No Excel / No COM final version
Target: Windows 11 / Python 3.13 / latest PySide6 / latest Nuitka

This application does not start Microsoft Excel, does not run VBA, and does not use COM.
It parses log files in Python and writes an .xlsx report using openpyxl.
"""
from __future__ import annotations

import unicodedata
import csv
import json
import fnmatch
import os
import re
import sys
import traceback
import tempfile
import zipfile
import shutil
import xml.etree.ElementTree as ET
import concurrent.futures
import multiprocessing
import threading
import time
import sqlite3
from dataclasses import dataclass, asdict
from parser_rc1 import parse_rc1_file
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Iterable, Any

from common.master_data import load_shared_site_map, shared_master_path

def startup_log_path() -> Path:
    base = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    p = Path(base) / "LogMergeTool_NoExcel"
    p.mkdir(parents=True, exist_ok=True)
    return p / "startup_error.log"


def write_startup_log(text: str) -> None:
    try:
        startup_log_path().write_text(text, encoding="utf-8")
    except Exception:
        pass


try:
    from PySide6.QtCore import Qt, QDate, QThread, Signal, QSettings, QAbstractTableModel, QModelIndex, QTimer, QRect, QItemSelectionModel, QItemSelection
    from PySide6.QtGui import QStandardItemModel, QStandardItem, QPainter, QColor, QPen, QBrush
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
        QFileDialog, QDateEdit, QCheckBox, QTextEdit, QComboBox, QMessageBox,
        QProgressBar, QGroupBox, QFormLayout, QLineEdit, QTableView, QAbstractItemView, QSplitter,
        QSpinBox, QHeaderView, QDialog, QTableWidget, QTableWidgetItem,
        QDialogButtonBox, QInputDialog, QProgressDialog, QMenu, QGridLayout,
        QFrame, QPlainTextEdit, QSizePolicy, QTabWidget,
    )
except Exception:
    write_startup_log("PySide6 import failed.\n\n" + traceback.format_exc())
    raise

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.formatting.rule import FormulaRule
except Exception:
    write_startup_log("openpyxl import failed.\n\n" + traceback.format_exc())
    raise

RELEASE_MODE = os.environ.get("INSIGHTEC_RELEASE_MODE", "0").strip().lower() in {"1", "true", "yes", "on"}

APP_TITLE = "Log Merge Tool - No Excel"
APP_VERSION = "2.0.0-rc1-commit0069"
MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# File classification is intentionally permissive so old field logs are not missed.
FILE_TYPE_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("WATERSYSTEM", re.compile(r"^WaterSystem.*\.(?:txt|log)$", re.I)),
    ("VIMEASURE", re.compile(r"^VIMeasure_.*\.txt$", re.I)),
    ("ACQUISITION", re.compile(r"^Acquisition_.*\.(?:txt|log)$", re.I)),
    ("WS", re.compile(r"^(?:20\d{2}_[A-Za-z]{3}_\d{2}_\d{2}_\d{2}_\d{2}(?:\(\d+\))?\.Log)$", re.I)),
    ("LAIS", re.compile(r"^lais.*\.(?:log|txt)$", re.I)),
    ("GESYS", re.compile(r"^gesys.*\.(?:log|txt)$", re.I)),
    ("MRSERVER", re.compile(r"^(?:mrserver|mr_server|mr-server).*\.(?:log|txt)$", re.I)),
    ("CSA", re.compile(r"^csa_brain.*\.(?:log|txt)$", re.I)),
    ("PSC", re.compile(r"^psc(?:[._-].*)?\.(?:log|txt)$", re.I)),
    ("CGA", re.compile(r"^cga_brain.*\.(?:log|txt)$", re.I)),
]


def plugins_dir() -> Path:
    p = app_dir() / "plugins"
    p.mkdir(parents=True, exist_ok=True)
    return p


def plugin_manifest_paths() -> list[Path]:
    root = plugins_dir()
    paths = []
    for p in root.iterdir():
        if p.is_dir() and (p / "manifest.json").exists():
            paths.append(p / "manifest.json")
    return sorted(paths, key=lambda x: x.parent.name.lower())


def load_file_type_plugins() -> list[dict[str, Any]]:
    plugins: list[dict[str, Any]] = []
    for manifest in plugin_manifest_paths():
        try:
            data = json.loads(manifest.read_text(encoding="utf-8"))
            if not data.get("id") or not data.get("display_name"):
                continue
            data["_path"] = str(manifest.parent)
            patterns = data.get("patterns") or data.get("file_patterns") or []
            if isinstance(patterns, str):
                patterns = [patterns]
            data["patterns"] = [str(x) for x in patterns if str(x).strip()]
            data["mode"] = data.get("mode") or ["merge", "import"]
            data["enabled"] = bool(data.get("enabled", True))
            if data["enabled"]:
                plugins.append(data)
        except Exception:
            continue
    return plugins




def plugin_manifest_by_id(plugin_id: str) -> Optional[dict[str, Any]]:
    target = str(plugin_id or "").strip().lower()
    for plg in load_file_type_plugins():
        if str(plg.get("id", "")).strip().lower() == target:
            return plg
    return None


def load_plugin_json(plugin_id: str, filename: str, fallback=None):
    plg = plugin_manifest_by_id(plugin_id)
    if not plg:
        return fallback
    try:
        path = Path(str(plg.get("_path", ""))) / filename
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        pass
    return fallback


def parse_plugin_structured_file(path: Path, source_type: str) -> Optional[list["LogRecord"]]:
    """Parse a File Type plugin whose parser.json declares structured_whitespace.

    The schema is deliberately data-driven so a new numeric log type can be
    installed through Update File Type without rebuilding the main EXE.
    """
    parser = load_plugin_json(source_type, "parser.json", {}) or {}
    if str(parser.get("format", "")).lower() != "structured_whitespace":
        return None

    lines = read_text_lines(path)
    fallback_dt = parse_filename_datetime(path)
    current_date = fallback_dt.date() if fallback_dt else None
    last_ts: Optional[datetime] = None
    header_fields = [str(x) for x in (parser.get("columns") or [])]
    header_re_text = str(parser.get("header_line_regex") or r"^;\s*Data:\s*(?P<columns>.+)$")
    data_re_text = str(parser.get("data_line_regex") or r"^(?P<time>\d{1,2}:\d{2}:\d{2}[:.]\d+)\s+(?P<values>.+)$")
    try:
        header_re = re.compile(header_re_text)
        data_re = re.compile(data_re_text)
    except Exception:
        return []
    records: list[LogRecord] = []
    time_group = str(parser.get("time_group") or "time")
    values_group = str(parser.get("values_group") or "values")
    columns_group = str(parser.get("columns_group") or "columns")
    source_name = str((plugin_manifest_by_id(source_type) or {}).get("display_name") or source_type)

    for idx, raw_line in enumerate(lines, start=1):
        line = raw_line.rstrip("\r\n")
        hm = header_re.search(line)
        if hm:
            text = hm.groupdict().get(columns_group) or (hm.group(1) if hm.groups() else "")
            detected = [x for x in re.split(r"\s+", str(text).strip()) if x]
            if detected:
                header_fields = detected
            continue
        dm = data_re.search(line.strip())
        if not dm:
            continue
        gd = dm.groupdict()
        time_text = gd.get(time_group) or ""
        values_text = gd.get(values_group) or ""
        if not time_text or not values_text:
            continue
        ts = parse_time_at_line_start(time_text, current_date, last_ts)
        if ts:
            if last_ts and ts.date() > last_ts.date():
                current_date = ts.date()
            last_ts = ts
        else:
            ts = last_ts or fallback_dt
        values = [x for x in re.split(r"\s+", values_text.strip()) if x]
        if not header_fields:
            header_fields = [f"Value{i+1}" for i in range(len(values))]
        row_data: dict[str, Any] = {}
        for i, name in enumerate(header_fields):
            if i >= len(values):
                break
            value = values[i]
            try:
                row_data[name] = float(value)
            except Exception:
                row_data[name] = value
        # Keep raw field count information for validation/debug without adding
        # it to the normal viewer columns.
        row_data["_field_count"] = len(values)
        records.append(LogRecord(
            ts, str(source_type), path.name, idx, "", source_name, "",
            json.dumps(row_data, ensure_ascii=False)
        ))
    return records

def plugin_type_patterns() -> list[tuple[str, re.Pattern[str]]]:
    out: list[tuple[str, re.Pattern[str]]] = []
    for plg in load_file_type_plugins():
        pid = str(plg.get("id", "")).strip()
        for pat in plg.get("patterns", []):
            try:
                out.append((pid, re.compile(fnmatch.translate(pat), re.I)))
            except Exception:
                pass
    return out


def validate_plugin_zip(zip_path: Path) -> tuple[bool, str, Optional[dict[str, Any]]]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = set(zf.namelist())
            if "manifest.json" not in names:
                return False, "manifest.json is missing.", None
            manifest = json.loads(zf.read("manifest.json").decode("utf-8-sig"))
            if not manifest.get("id"):
                return False, "manifest.json requires id.", None
            if not manifest.get("display_name"):
                return False, "manifest.json requires display_name.", None
            patterns = manifest.get("patterns") or manifest.get("file_patterns")
            if not patterns:
                return False, "manifest.json requires patterns or file_patterns.", None
            return True, "OK", manifest
    except Exception as exc:
        return False, str(exc), None


def install_plugin_zip(zip_path: Path) -> tuple[bool, str]:
    ok, msg, manifest = validate_plugin_zip(zip_path)
    if not ok or not manifest:
        return False, msg
    pid = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(manifest["id"]).strip())
    dest = plugins_dir() / pid
    if dest.exists():
        shutil.rmtree(dest, ignore_errors=True)
    dest.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(dest)
        return True, f"Installed plugin: {manifest.get('display_name')} ({manifest.get('version','')})"
    except Exception as exc:
        shutil.rmtree(dest, ignore_errors=True)
        return False, str(exc)


def is_review_file(name: str) -> bool:
    """Review Import target: review.out, review.out.ar, or review.out.* only."""
    n = name.strip().lower()
    return n == "review.out" or n.startswith("review.out.")

WS_FILENAME_RE = re.compile(
    r"^(?P<y>20\d{2})_(?P<mon>[A-Za-z]{3})_(?P<d>\d{2})_(?P<h>\d{2})_(?P<mi>\d{2})_(?P<s>\d{2})(?:\(\d+\))?\.Log$",
    re.I,
)

FULL_DATETIME_PATTERNS: list[re.Pattern[str]] = [
    # 2026-07-04 14:33:04.149 / 2026/07/04 14:33:04:149
    re.compile(r"(?P<y>20\d{2})[-/](?P<m>\d{1,2})[-/](?P<d>\d{1,2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?"),
    # 07/04/2026 14:33:04.149
    re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})/(?P<y>20\d{2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?"),
    # 04-Jul-2026 14:33:04.149
    re.compile(r"(?P<d>\d{1,2})[- ](?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](?P<y>20\d{2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?", re.I),
    # Jul 04 2026 14:33:04.149
    re.compile(r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ](?P<d>\d{1,2})[-/ ,]+(?P<y>20\d{2})[ T_]+(?P<h>\d{1,2}):(?P<mi>\d{1,2})(?::(?P<s>\d{1,2})(?:(?:\.|:)(?P<ms>\d{1,6}))?)?", re.I),
]

DATE_ONLY_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"(?P<y>20\d{2})[-/](?P<m>\d{1,2})[-/](?P<d>\d{1,2})"),
    re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})/(?P<y>20\d{2})"),
    re.compile(r"(?P<d>\d{1,2})[- ](?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](?P<y>20\d{2})", re.I),
    re.compile(r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ](?P<d>\d{1,2})[-/ ,]+(?P<y>20\d{2})", re.I),
]

# WS and many system logs put the record time at the very beginning of each line.
# Example: 14:33:04:149 Ent Dbg 4992 ...
TIME_AT_LINE_START_RE = re.compile(r"^\s*(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})(?:(?:\.|:)(?P<ms>\d{1,6}))?\b")

LEVEL_RE = re.compile(r"\b(ERROR|ERR|WARN(?:ING)?|INFO|DEBUG|TRACE|FATAL|ALARM)\b", re.I)

DEFAULT_CSA_RULES = [
    {"keyword": "Watch Dog Error", "message": "Watch Dog Error"},
    {"keyword": "Vacuum Failure", "message": "Vacuum Failure"},
]

DEFAULT_SITE_MAP = [
    {"serial": "4278", "site": "Yokosuka"},
    {"serial": "4259", "site": "CIM"},
    {"serial": "4257", "site": "Tsukazaki"},
    {"serial": "4087", "site": "Kyokuto"},
    {"serial": "4048", "site": "Kitano"},
    {"serial": "4083", "site": "Kashiwaba"},
    {"serial": "4078", "site": "Fukuoka"},
    {"serial": "4065", "site": "Miyagi"},
    {"serial": "4073", "site": "Atsuchi"},
]

@dataclass
class LogRecord:
    """Canonical parsed record with legacy mapping compatibility."""

    timestamp: Optional[datetime]
    source_type: str
    filename: str
    line_no: int
    level: str
    category: str
    message: str
    raw: str

    _KEY_ALIASES = {
        "_ts": "timestamp",
        "Timestamp": "timestamp",
        "timestamp": "timestamp",
        "SourceType": "source_type",
        "source_type": "source_type",
        "File": "filename",
        "Filename": "filename",
        "filename": "filename",
        "Line": "line_no",
        "LineNo": "line_no",
        "line_no": "line_no",
        "Level": "level",
        "level": "level",
        "Category": "category",
        "category": "category",
        "Message": "message",
        "message": "message",
        "Raw": "raw",
        "raw": "raw",
    }

    _DISPLAY_KEYS = (
        "_ts",
        "Timestamp",
        "SourceType",
        "File",
        "Line",
        "Level",
        "Category",
        "Message",
        "Raw",
    )

    def _resolve_key(self, key: object) -> Optional[str]:
        return self._KEY_ALIASES.get(str(key))

    def get(self, key: object, default=None):
        attribute = self._resolve_key(key)
        if attribute is None:
            return default
        return getattr(self, attribute, default)

    def __getitem__(self, key: object):
        attribute = self._resolve_key(key)
        if attribute is None:
            raise KeyError(key)
        return getattr(self, attribute)

    def __contains__(self, key: object) -> bool:
        return self._resolve_key(key) is not None

    def keys(self):
        return self._DISPLAY_KEYS

    def values(self):
        return tuple(self.get(key) for key in self._DISPLAY_KEYS)

    def items(self):
        return tuple((key, self.get(key)) for key in self._DISPLAY_KEYS)

    def __iter__(self):
        return iter(self._DISPLAY_KEYS)

    def __len__(self):
        return len(self._DISPLAY_KEYS)

    def to_dict(self) -> dict:
        return {
            "_ts": self.timestamp,
            "Timestamp": self.timestamp,
            "SourceType": self.source_type,
            "File": self.filename,
            "Line": self.line_no,
            "Level": self.level,
            "Category": self.category,
            "Message": self.message,
            "Raw": self.raw,
        }

@dataclass
class RunOptions:
    source_folder: str
    output_folder: str
    recursive: bool
    use_start: bool
    use_end: bool
    start_date: Optional[str]
    end_date: Optional[str]
    serial: str
    site: str
    include_ws: bool
    include_watersystem: bool
    include_cga: bool
    include_csa: bool
    include_mrserver: bool
    include_gesys: bool
    include_lais: bool
    include_psc: bool
    include_unknown: bool
    turbo_csv: bool = True
    summary_xlsx_only: bool = True
    fast_xlsx_mode: bool = True
    csv_summary_mode: bool = True
    skip_merged1: bool = True
    timestamp_as_text: bool = True
    disable_styles_large: bool = True
    worker_count: int = 0
    noise_enabled_merge: bool = False
    noise_exclude_output: bool = False
    noise_learning: bool = True
    selected_files: Optional[list[str]] = None
    plugin_types: Optional[list[str]] = None

@dataclass
class RunResult:
    scanned_files: int = 0
    parsed_files: int = 0
    skipped_files: int = 0
    total_records: int = 0
    included_records: int = 0
    csa_error_count: int = 0
    output_path: str = ""


def app_dir() -> Path:
    """Return the folder that contains runtime data files.

    In Nuitka onefile mode, bundled data files are extracted beside __file__,
    while sys.argv[0] points to the outer EXE path.  Using __file__ keeps
    csa_error_rules.json and site_serial_map.json readable in both source and
    built EXE modes.
    """
    try:
        return Path(__file__).resolve().parent
    except Exception:
        return Path(sys.argv[0]).resolve().parent


def load_json_file(filename: str, fallback):
    p = app_dir() / filename
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return fallback
    return fallback


def classify_file(path: Path, include_unknown: bool) -> Optional[str]:
    for typ, pat in FILE_TYPE_PATTERNS:
        if pat.match(path.name):
            return typ
    for typ, pat in plugin_type_patterns():
        if pat.match(path.name):
            return typ
    if include_unknown and path.suffix.lower() in {".log", ".txt", ".out", ".ar"}:
        return "UNKNOWN"
    return None


def parse_ws_filename_datetime(path: Path) -> Optional[datetime]:
    m = WS_FILENAME_RE.match(path.name)
    if not m:
        return None
    try:
        return datetime(
            int(m.group("y")), MONTHS[m.group("mon").lower()], int(m.group("d")),
            int(m.group("h")), int(m.group("mi")), int(m.group("s"))
        )
    except Exception:
        return None


def parse_filename_datetime(path: Path) -> Optional[datetime]:
    """Extract a best-effort date/time from a log file name.

    Used when the log body has only a time value, as seen in MRServer and
    similar logs. If the file name contains only a date, midnight is returned;
    the row-level time is then combined with this date.
    """
    ws_dt = parse_ws_filename_datetime(path)
    if ws_dt:
        return ws_dt

    name = path.name
    patterns: list[re.Pattern[str]] = [
        # MRServer example: mrserver_Thu_Jul__2_17_02_10_2026.log
        # Use the date/time embedded in the file name only as the starting date.
        re.compile(r"(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[_ -]+(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[_ -]+(?P<d>\d{1,2})[_ -]+(?P<h>\d{1,2})[_ -]+(?P<mi>\d{1,2})(?:[_ -]+(?P<s>\d{1,2}))?[_ -]+(?P<y>20\d{2})", re.I),
        re.compile(r"(?P<y>20\d{2})[_-](?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[_-](?P<d>\d{1,2})(?:[_-](?P<h>\d{1,2})[_-](?P<mi>\d{1,2})(?:[_-](?P<s>\d{1,2}))?)?", re.I),
        re.compile(r"(?P<y>20\d{2})[-_](?P<m>\d{1,2})[-_](?P<d>\d{1,2})(?:[_ T-](?P<h>\d{1,2})[-_:](?P<mi>\d{1,2})(?:[-_:](?P<s>\d{1,2}))?)?", re.I),
        re.compile(r"(?P<y>20\d{2})(?P<m>\d{2})(?P<d>\d{2})(?:[_-]?(?P<h>\d{2})(?P<mi>\d{2})(?P<s>\d{2})?)?", re.I),
        re.compile(r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[_-](?P<d>\d{1,2})[_-](?P<y>20\d{2})(?:[_-](?P<h>\d{1,2})[_-](?P<mi>\d{1,2})(?:[_-](?P<s>\d{1,2}))?)?", re.I),
    ]
    for pat in patterns:
        m = pat.search(name)
        if not m:
            continue
        try:
            gd = m.groupdict()
            year = int(gd["y"])
            month = MONTHS[gd["mon"].lower()] if gd.get("mon") else int(gd["m"])
            day = int(gd["d"])
            hour = int(gd.get("h") or 0)
            minute = int(gd.get("mi") or 0)
            second = int(gd.get("s") or 0)
            return datetime(year, month, day, hour, minute, second)
        except Exception:
            continue
    return None


def _microsecond_from_match(gd: dict[str, str]) -> int:
    ms = gd.get("ms") or ""
    if not ms:
        return 0
    return int((ms + "000000")[:6])


def _datetime_from_match(gd: dict[str, str]) -> datetime:
    year = int(gd["y"])
    month = MONTHS[gd["mon"].lower()] if gd.get("mon") else int(gd["m"])
    day = int(gd["d"])
    hour = int(gd.get("h") or 0)
    minute = int(gd.get("mi") or 0)
    second = int(gd.get("s") or 0)
    microsecond = _microsecond_from_match(gd)
    return datetime(year, month, day, hour, minute, second, microsecond)




GESYS_SECTION_DATETIME_RE = re.compile(
    r"^\s*\d+\s+\d+\s+\d+\s+(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
    r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+"
    r"(?P<d>\d{1,2})\s+"
    r"(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})\s+"
    r"(?P<y>20\d{2})\b",
    re.I,
)

LAIS_DATE_HEADER_RE = re.compile(
    r"^\s*(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
    r"(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+"
    r"(?P<d>\d{1,2})\s+"
    r"(?P<y>20\d{2})\s*$",
    re.I,
)


def parse_lais_date_header(text: str) -> Optional[date]:
    """Parse LAIS section date headers such as 'Mon Jan 26 2026'.

    LAIS logs are date-section based: a standalone weekday/month/day/year line
    defines the date for following time-led rows.  The following rows start with
    high precision times such as '16:44:32.390869'.  File date and unrelated
    embedded text must not override these section dates.
    """
    m = LAIS_DATE_HEADER_RE.match(text)
    if not m:
        return None
    try:
        gd = m.groupdict()
        return date(int(gd["y"]), MONTHS[gd["mon"].lower()], int(gd["d"]))
    except Exception:
        return None


def parse_gesys_section_datetime(text: str) -> Optional[datetime]:
    """Parse the timestamp line that starts a GESYS log section.

    GESYS records are section-based. A section timestamp appears on a line like:
        1780144501    0    1    Sat May 30 12:35:01 2026    0
    Dates inside these section headers are the authoritative record timestamp.
    File name dates and old release/build dates must not be used for GESYS rows.
    """
    m = GESYS_SECTION_DATETIME_RE.search(text)
    if not m:
        return None
    try:
        gd = m.groupdict()
        return datetime(
            int(gd["y"]),
            MONTHS[gd["mon"].lower()],
            int(gd["d"]),
            int(gd["h"]),
            int(gd["mi"]),
            int(gd["s"]),
        )
    except Exception:
        return None


def parse_datetime_from_text(text: str, fallback_date: Optional[date] = None) -> Optional[datetime]:
    """Return a full timestamp from log content.

    Priority is always the timestamp written inside the log text.  The optional
    fallback_date is used only for time-only lines, mainly WS logs whose date is
    stored in the file name but whose record timestamp is at the line head.
    """
    for pat in FULL_DATETIME_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        try:
            return _datetime_from_match(m.groupdict())
        except Exception:
            continue

    if fallback_date:
        m = TIME_AT_LINE_START_RE.search(text)
        if m:
            try:
                gd = m.groupdict()
                return datetime(
                    fallback_date.year, fallback_date.month, fallback_date.day,
                    int(gd["h"]), int(gd["mi"]), int(gd["s"]), _microsecond_from_match(gd)
                )
            except Exception:
                return None
    return None


def parse_date_only_from_text(text: str) -> Optional[date]:
    for pat in DATE_ONLY_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        try:
            gd = m.groupdict()
            year = int(gd["y"])
            month = MONTHS[gd["mon"].lower()] if gd.get("mon") else int(gd["m"])
            day = int(gd["d"])
            return date(year, month, day)
        except Exception:
            continue
    return None


def parse_time_at_line_start(text: str, current_date: Optional[date], previous_ts: Optional[datetime] = None) -> Optional[datetime]:
    if current_date is None:
        return None
    m = TIME_AT_LINE_START_RE.search(text)
    if not m:
        return None
    gd = m.groupdict()
    try:
        candidate = datetime(
            current_date.year, current_date.month, current_date.day,
            int(gd["h"]), int(gd["mi"]), int(gd["s"]), _microsecond_from_match(gd)
        )
        # Filename-date fallback logs such as MRServer often contain time-only
        # records. The file is read from top to bottom as chronological order.
        # When the record time goes backwards, treat it as a new day.
        # This intentionally does NOT use file created/modified time.
        if previous_ts and candidate < previous_ts:
            while candidate < previous_ts:
                candidate += timedelta(days=1)
        return candidate
    except Exception:
        return None



def _trusted_date_position(text: str, start_index: int) -> bool:
    """Accept dates only when they look like the record date/time, not a version date."""
    prefix = text[:start_index].strip().lower()
    if start_index <= 3:
        return True
    if prefix.endswith(("date:", "time:", "timestamp:", "date", "time", "timestamp")):
        return True
    return False


def parse_trusted_full_datetime_from_text(text: str) -> Optional[datetime]:
    for pat in FULL_DATETIME_PATTERNS:
        m = pat.search(text)
        if not m or not _trusted_date_position(text, m.start()):
            continue
        try:
            return _datetime_from_match(m.groupdict())
        except Exception:
            continue
    return None

def is_explicit_date_header(line: str) -> bool:
    """Return True only when a line is intended to define the log date.

    Do not treat dates embedded in ordinary messages such as
    "Neuro Release Reg _6.33 [v Jun 14 2022 ...]" as the record date.
    CSA/CGA/MRServer logs often have a reliable time at the beginning of each
    row and old version/build dates in the message body.
    """
    t = line.strip()
    low = t.lower()
    if not t:
        return False
    if low.startswith(("date:", "log date:", "current date:", "system date:", "start date:")):
        return True
    if re.match(r"^(date|log date|current date|system date|start date)\s*=", low):
        return True
    # Allow a line that is only a date, optionally with separators.
    if re.match(r"^(?:(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+)?(?:20\d{2}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}/\d{1,2}/20\d{2}|\d{1,2}[- ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]20\d{2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ]\d{1,2}[-/ ,]+20\d{2})\s*$", t, re.I):
        return True
    return False


def extract_content_timestamp(line: str, source_type: str, current_date: Optional[date], previous_ts: Optional[datetime]) -> tuple[Optional[datetime], Optional[date]]:
    """Extract the record timestamp from log content.

    v28 rule:
    - For time-led logs such as WS/CSA/CGA/MRServer, the time at the very
      beginning of the row is authoritative. It is combined with the current
      date, which usually comes from the file name.
    - Full date/time inside the message is used only when no row-leading time
      exists. This prevents old software/version dates in CSA/CGA messages from
      overwriting the real record timestamp.
    - Date-only text changes current_date only when it is an explicit date
      header, not an arbitrary date in the message body.
    """
    source_upper = str(source_type).upper()

    # LAIS is date-section based. A standalone line such as
    # "Mon Jan 26 2026" defines the date for following time-led rows.
    # Do not use file name dates for LAIS when the section date is present.
    if source_upper == "LAIS":
        lais_date = parse_lais_date_header(line)
        if lais_date:
            return None, lais_date

    # GESYS is section-based. Use the section timestamp line first and let
    # following lines inherit it until the next section. This prevents the
    # filename date or unrelated dates from being applied to all GESYS rows.
    if source_upper == "GESYS":
        gesys_ts = parse_gesys_section_datetime(line)
        if gesys_ts:
            return gesys_ts, gesys_ts.date()

    # First priority for time-led logs: row-leading time + current/file date.
    line_ts = parse_time_at_line_start(line, current_date, previous_ts)
    if line_ts:
        return line_ts, line_ts.date()

    # If there is no leading time, use a full date/time only when it is a
    # trusted record timestamp/header. Do not use embedded release/build dates.
    full_ts = parse_trusted_full_datetime_from_text(line)
    if full_ts:
        return full_ts, full_ts.date()

    if is_explicit_date_header(line):
        date_header = parse_date_only_from_text(line)
        if date_header:
            current_date = date_header

    return None, current_date


def parse_user_date(value: Optional[str]) -> Optional[date]:
    """Parse GUI date text into a date.

    QDateEdit may return formats such as yyyy/MM/dd, yyyy-MM-dd, or
    locale-specific strings.  This helper is intentionally small and returns a
    date object because the date range filter is inclusive by day.
    """
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    dt = parse_datetime_from_text(text)
    if dt:
        return dt.date()
    d = parse_date_only_from_text(text)
    return d

def in_date_range(ts: Optional[datetime], file_date: Optional[date], start: Optional[date], end: Optional[date]) -> bool:
    # v8: date filtering is based on the timestamp extracted from the log body.
    # file_date is kept only for backward call compatibility and is not used.
    if start is None and end is None:
        return True
    if ts is None:
        return False
    d = ts.date()
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True

def detect_level(line: str) -> str:
    m = LEVEL_RE.search(line)
    if not m:
        return ""
    return m.group(1).upper()


def detect_category(source_type: str, line: str) -> str:
    u = line.upper()
    if "ERROR" in u or "FAIL" in u or "FATAL" in u or "ALARM" in u:
        return "ERROR"
    if "WARN" in u:
        return "WARNING"
    if source_type == "WS":
        if "CIRCULATE" in u:
            return "CIRCULATE"
        if "DRAIN" in u:
            return "DRAIN"
        if "FILL" in u:
            return "FILL"
        if "PAUSE" in u:
            return "PAUSE"
    return source_type


def clean_excel_value(value):
    """Return a value safe for openpyxl worksheet cells.

    Field logs sometimes contain control characters such as SUB/ESC. Excel
    cannot store those characters in xlsx XML and openpyxl raises
    IllegalCharacterError during save. Keep the readable text and remove only
    characters that are illegal in worksheets.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = ILLEGAL_CHARACTERS_RE.sub("", value)
        except Exception:
            value = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
        return value.replace("\x00", "")
    return value


def clean_message(line: str) -> str:
    return str(clean_excel_value(line)).strip()


def viewer_message_body(text: str) -> str:
    """Message column text for Explorer only.

    Keep the raw record in the detail pane, but do not repeat the leading
    timestamp in the visible Message column.
    """
    t = str(clean_excel_value(text or "")).strip()
    t = TIME_AT_LINE_START_RE.sub("", t, count=1).strip()
    return t


def format_timestamp(ts: Optional[datetime]) -> str:
    """Format every output timestamp explicitly as yyyy/mm/dd hh:mm:ss.f.

    Missing date/time components are zero-padded by datetime/strftime.
    If the source line has no fractional seconds, append .0 so every row has
    an explicit timestamp like 2026/07/04 13:07:58.6.
    """
    if not ts:
        return ""
    base = ts.strftime("%Y/%m/%d %H:%M:%S")
    if ts.microsecond:
        frac = f"{ts.microsecond:06d}".rstrip("0")
        return f"{base}.{frac}"
    return f"{base}.0"


def record_output_headers() -> list[str]:
    return ["Timestamp", "SourceType", "File", "Line", "Level", "Category", "Message", "Raw"]


def record_to_output_row(r: "LogRecord") -> list[Any]:
    # Default: write Timestamp as Excel datetime. Writer applies display format.
    return [
        r.timestamp if r.timestamp else "",
        r.source_type, r.filename, r.line_no, r.level, r.category, r.message, r.raw
    ]


def record_to_output_row_fast(r: "LogRecord", timestamp_as_text: bool = False) -> list[Any]:
    ts_value = format_timestamp(r.timestamp) if timestamp_as_text else (r.timestamp if r.timestamp else "")
    return [ts_value, r.source_type, r.filename, r.line_no, r.level, r.category, r.message, r.raw]


def csa_hit_to_output_row(h: dict[str, Any]) -> list[Any]:
    ts = h.get("Timestamp")
    return [
        ts if isinstance(ts, datetime) else "",
        h.get("File", ""), h.get("Line", ""), h.get("Keyword", ""), h.get("Message", ""), h.get("Raw", "")
    ]


def read_text_lines(path: Path) -> list[str]:
    encodings = ("utf-8-sig", "utf-8", "cp932", "shift_jis", "latin-1")
    last_exc: Optional[Exception] = None
    for enc in encodings:
        try:
            return path.read_text(encoding=enc, errors="strict").splitlines()
        except Exception as exc:
            last_exc = exc
    try:
        return path.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        raise RuntimeError(f"Unable to read {path}: {last_exc}")


def parse_file(path: Path, source_type: str) -> list[LogRecord]:
    plugin_records = parse_plugin_structured_file(path, source_type)
    if plugin_records is not None:
        return plugin_records
    # v14 timestamp rule:
    # 1) Full date+time written in the log body is used first.
    # 2) If the body has only a time, combine it with the body date header or,
    #    when no body date exists, the date parsed from the file name.
    #    File created/modified dates are intentionally never used.
    # 3) Within one file, lines are read from top to bottom. If the line time
    #    rolls back across midnight, the date is advanced by one day.
    # 4) Lines with no timestamp inherit the previous row timestamp; if there is
    #    no previous row timestamp, the file-name datetime is used as fallback.
    fallback_dt = parse_filename_datetime(path)
    current_date: Optional[date] = fallback_dt.date() if fallback_dt else None
    records: list[LogRecord] = []
    lines = read_text_lines(path)

    # Commit0014 R4: CSA/CGA use the dedicated structured parser.
    # No silent legacy-parser fallback is allowed for these file types.
    rc1_rows = parse_rc1_file(path, source_type, lines, enable_builtin_vimeasure=True)
    if rc1_rows is not None:
        return [
            LogRecord(
                row.get("timestamp"), str(row.get("source_type", source_type)),
                str(row.get("filename", path.name)), int(row.get("line_no", 0)),
                str(row.get("level", "")), str(row.get("category", "")),
                str(row.get("message", "")), str(row.get("raw", "")),
            )
            for row in rc1_rows
        ]

    last_ts: Optional[datetime] = None
    if str(source_type).upper() == "WATERSYSTEM":
        # RC1 Commit0010: preserve the native WaterSystem table structure.
        # Native layout: Time, MainState, CoolingState, Error, then numeric columns.
        # CoolingState is intentionally omitted from structured output because the
        # current files only contain COOLING_STATE_NOT_INITIALIZED.
        header_fields: list[str] = []
        ignored_fields = {"CoolingState"}
        no_error_values = {"NO_ERROR", "NONE", "OK", "0", "NOERROR"}

        for idx, line in enumerate(lines, start=1):
            msg = clean_message(line)
            if not msg:
                continue

            ts, current_date = extract_content_timestamp(msg, source_type, current_date, last_ts)
            if ts:
                last_ts = ts
            else:
                ts = last_ts or fallback_dt
                if ts:
                    last_ts = ts

            body = TIME_AT_LINE_START_RE.sub("", msg, count=1).strip()
            if not body:
                continue

            values = body.split()

            if "MainState" in values and "Error" in values:
                header_fields = values
                continue

            if not header_fields:
                header_fields = [
                    "MainState", "CoolingState", "Error", "ChillerTemp",
                    "PrimaryFlowMeter", "AbsolutePressure", "DynamicPressure",
                    "XdTemperature", "VacuumLevel", "DOLevel", "WaterVolume",
                    "SecondaryFlowMeter", "HsCombitac", "ChillerStatus",
                    "ChillerLowLevelInd", "PressureSetPoint",
                ]

            if len(values) < 3:
                continue

            row_data: dict[str, Any] = {}
            for field_index, field_name in enumerate(header_fields):
                if field_index >= len(values):
                    break
                if field_name in ignored_fields:
                    continue
                value = values[field_index]
                if field_name in {"MainState", "Error"}:
                    row_data[field_name] = value
                else:
                    try:
                        row_data[field_name] = float(value)
                    except Exception:
                        row_data[field_name] = value

            main_state = str(row_data.get("MainState", "")).strip()
            raw_error = str(row_data.get("Error", "")).strip()
            visible_error = "" if raw_error.upper() in no_error_values else raw_error
            row_data["Error"] = visible_error

            records.append(LogRecord(
                ts,
                "WATERSYSTEM",
                path.name,
                idx,
                "ERROR" if visible_error else "",
                main_state,
                "",
                json.dumps(row_data, ensure_ascii=False),
            ))

        if not records and fallback_dt:
            records.append(LogRecord(
                fallback_dt,
                "WATERSYSTEM",
                path.name,
                0,
                "",
                "",
                "",
                json.dumps({"MainState": "", "Error": ""}, ensure_ascii=False),
            ))
        return records
    for idx, line in enumerate(lines, start=1):
        msg = clean_message(line)
        if not msg:
            continue
        ts, current_date = extract_content_timestamp(msg, source_type, current_date, last_ts)
        if ts:
            last_ts = ts
        else:
            # Multi-line continuations inherit the last timestamp. If even the
            # first line has no time, fall back to the filename date/time so the
            # row still has an explicit timestamp when possible.
            ts = last_ts or fallback_dt
            if ts:
                last_ts = ts
        level = detect_level(msg)
        category = detect_category(source_type, msg)
        records.append(LogRecord(ts, source_type, path.name, idx, level, category, msg, line.rstrip("\r\n")))
    if not records and fallback_dt:
        records.append(LogRecord(fallback_dt, source_type, path.name, 0, "", source_type, "File detected but no readable lines", ""))
    return records

def iter_files(root: Path, recursive: bool) -> Iterable[Path]:
    it = root.rglob("*") if recursive else root.glob("*")
    for p in it:
        if p.is_file():
            yield p


def enabled_types(options: RunOptions) -> set[str]:
    enabled = set()
    if options.include_ws: enabled.add("WS")
    if getattr(options, "include_watersystem", False): enabled.add("WATERSYSTEM")
    if options.include_cga: enabled.add("CGA")
    if options.include_csa: enabled.add("CSA")
    if options.include_mrserver: enabled.add("MRSERVER")
    if options.include_gesys: enabled.add("GESYS")
    if options.include_lais: enabled.add("LAIS")
    if options.include_psc: enabled.add("PSC")
    if options.include_unknown: enabled.add("UNKNOWN")
    for t in (getattr(options, "plugin_types", None) or []):
        if t:
            enabled.add(str(t))
    return enabled



PSC_DATE_HEADER_RE = re.compile(r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}\s+\d{4}$", re.I)
PSC_TIME_RE = re.compile(r"^\d{2}:\d{2}:\d{2}(?:\.\d+)?$")
PSC_PARAM_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_\[\]\.]*)(\s*=\s*)(\"[^\"]*\"|'[^']*'|[^\s;]*)")


def normalize_psc_date_header(line: str) -> str:
    parts = line.split()
    if len(parts) >= 4:
        try:
            return f"{int(parts[3]):04d}-{MONTHS[parts[1].lower()]:02d}-{int(parts[2]):02d}"
        except Exception:
            return line.strip()
    return line.strip()


def trim_trailing_punctuation(v: str) -> str:
    v = v.strip()
    while len(v) > 1 and v.endswith('.'):
        try:
            float(v[:-1])
            v = v[:-1]
        except Exception:
            break
    return v


def extract_psc_params(text: str) -> list[tuple[str, str]]:
    return [(m.group(1), trim_trailing_punctuation(m.group(3))) for m in PSC_PARAM_RE.finditer(text)]


def first_token(text: str) -> str:
    return text.strip().split()[0] if text.strip().split() else ""


def second_token(text: str) -> str:
    parts = text.strip().split()
    return parts[1] if len(parts) >= 2 else ""


def parse_psc_line(raw_line: str, current_date: str, source_file: str) -> Optional[dict[str, Any]]:
    raw_line = raw_line.strip()
    if not raw_line:
        return None
    first_space = raw_line.find(" ")
    if first_space <= 0:
        return None
    log_time = raw_line[:first_space]
    if not PSC_TIME_RE.match(log_time):
        return None
    rest = raw_line[first_space + 1:].strip()
    process_name = ""
    if rest.startswith("[") and "]" in rest:
        close = rest.find("]")
        process_name = rest[1:close]
        rest = rest[close + 1:].strip()
    full_dt = f"{current_date} {log_time}" if current_date else log_time
    module_name = ""
    function_name = ""
    raw_params = ""
    message_text = rest
    if "::" in rest:
        before, after = rest.split("::", 1)
        module_name = before.strip()
        if ":" in after:
            function_name, raw_params = after.split(":", 1)
            function_name = function_name.strip()
            raw_params = raw_params.strip()
        else:
            function_name = first_token(after)
            raw_params = after[len(function_name):].strip()
    else:
        if ":" in rest:
            module_name, temp = rest.split(":", 1)
            module_name = module_name.strip()
            temp = temp.strip()
            function_name = first_token(temp)
            raw_params = temp
        else:
            module_name = first_token(rest)
            function_name = second_token(rest)
            raw_params = rest
    params = extract_psc_params(rest)
    return {
        "SourceFile": source_file,
        "LogDate": current_date,
        "Time": log_time,
        "DateTime": full_dt,
        "Process": process_name,
        "Module": module_name,
        "Function": function_name.strip(),
        "Message": message_text,
        "ParameterText": raw_params,
        "ParamCount": len(params),
        "RawLine": raw_line,
        "Params": params,
    }


def parse_psc_file_detail(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[LogRecord]]:
    current_date = ""
    log_rows: list[dict[str, Any]] = []
    param_rows: list[dict[str, Any]] = []
    records: list[LogRecord] = []
    max_param_count = 0
    for idx, line in enumerate(read_text_lines(path), start=1):
        raw = line.strip()
        if not raw:
            continue
        if PSC_DATE_HEADER_RE.match(raw):
            current_date = normalize_psc_date_header(raw)
            continue
        row = parse_psc_line(raw, current_date, path.name)
        if not row:
            continue
        params = row.pop("Params")
        max_param_count = max(max_param_count, len(params))
        for i, (name, value) in enumerate(params, start=1):
            row[f"Param{i}"] = f"{name}={value}"
            param_rows.append({
                "SourceFile": path.name,
                "LogDate": row["LogDate"],
                "Time": row["Time"],
                "DateTime": row["DateTime"],
                "Process": row["Process"],
                "Module": row["Module"],
                "Function": row["Function"],
                "ParameterName": name,
                "ParameterValue": value,
                "Message": row["Message"],
                "RawLine": row["RawLine"],
            })
        log_rows.append(row)
        ts = parse_datetime_from_text(str(row.get("DateTime", "")))
        records.append(LogRecord(ts, "PSC", path.name, idx, detect_level(raw), detect_category("PSC", raw), row.get("Message", raw), raw))
    return log_rows, param_rows, records


REVIEW_KEY_RE = re.compile(r"(^|[\t]+| {2,})([^:\t]{1,90}?)\s*:")


def clean_control_text(s: str) -> str:
    s = s.replace("\u00a0", " ").replace("\t", " ")
    s = "".join(ch for ch in s if ord(ch) >= 32)
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def clean_review_value(s: str) -> str:
    s = s.replace("\u00a0", " ").replace("\t", " ")
    s = "".join(ch for ch in s if ord(ch) >= 32)
    return s.strip()


def is_review_section_header(s: str) -> bool:
    t = s.strip(); u = t.upper()
    if not t or ":" in t or t.startswith("-----"):
        return False
    return (u.endswith("SCREEN") or " SCREEN" in u or u in {"GENERAL INFORMATION", "USER/SYSTEM PREFERENCE STATE", "ACCELERATION", "USER CV", "USER CV SCREEN"})


def review_value_after_colon(s: str) -> str:
    return clean_review_value(s.split(":", 1)[1]) if ":" in s else ""


def normalize_review_key(s: str) -> str:
    t = clean_control_text(s)
    if t.lower().endswith(" is"):
        t = clean_control_text(t[:-3])
    if t.upper() == "DATE":
        return "date"
    if t.upper() == "TIME":
        return "time"
    if t.upper() == "ID":
        return "ID"
    return t


def extract_review_key_values(raw_line: str, section_name: str, scan_data: dict[str, Any]) -> None:
    matches = list(REVIEW_KEY_RE.finditer(raw_line))
    if not matches:
        return
    for i, m in enumerate(matches):
        raw_key = normalize_review_key(m.group(2))
        value_start = m.end()
        value_end = matches[i + 1].start() if i < len(matches) - 1 else len(raw_line)
        raw_value = clean_review_value(raw_line[value_start:value_end])
        if not raw_key:
            continue
        if raw_key in {"date", "time", "ID"}:
            scan_data[raw_key] = raw_value
        else:
            full_key = f"{clean_control_text(section_name)} - {raw_key}" if section_name else raw_key
            scan_data[full_key] = raw_value


def parse_reviewout_file(path: Path) -> list[dict[str, Any]]:
    scan_list: list[dict[str, Any]] = []
    scan_data: dict[str, Any] = {}
    section_name = ""
    for line in read_text_lines(path):
        raw_line = line.replace("\u00a0", " ").replace("\r", "")
        trimmed = raw_line.strip()
        if not trimmed:
            continue
        low = trimmed.lower()
        if low.startswith("date:"):
            if scan_data:
                scan_list.append(scan_data)
                scan_data = {}
            section_name = ""
            scan_data["date"] = review_value_after_colon(trimmed)
        elif low.startswith("time:"):
            scan_data["time"] = review_value_after_colon(trimmed)
        elif low.startswith("id") and ":" in low:
            scan_data["ID"] = review_value_after_colon(trimmed)
        elif is_review_section_header(trimmed):
            section_name = clean_control_text(trimmed)
        elif ":" in trimmed:
            extract_review_key_values(raw_line, section_name, scan_data)
    if scan_data:
        scan_list.append(scan_data)
    return scan_list


def write_generic_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], table_name: str, progress_cb=None, progress_start: int = 0, progress_total: int = 0, control=None) -> int:
    ws = wb.create_sheet(title)
    written = write_rows(ws, headers, rows, progress_cb=progress_cb, progress_start=progress_start, progress_total=progress_total, progress_label=f"Writing {title}", control=control)
    apply_table(ws, table_name)
    autosize(ws)
    return written


def import_psc_only(file_path: str, output_folder: str = "") -> str:
    path = Path(file_path)
    out_dir = Path(output_folder) if output_folder else path.parent
    log_rows, param_rows, _ = parse_psc_file_detail(path)
    if not log_rows:
        raise ValueError("No valid PSC rows were detected in the selected file.")
    wb = Workbook()
    ws = wb.active
    ws.title = "PSC_Log"
    max_param = max((int(r.get("ParamCount", 0)) for r in log_rows), default=0)
    log_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "Message", "ParameterText", "ParamCount", "RawLine"] + [f"Param{i}" for i in range(1, max_param + 1)]
    write_rows(ws, log_headers, [[r.get(h, "") for h in log_headers] for r in log_rows])
    apply_table(ws, "PSCLogTable")
    autosize(ws)
    param_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "ParameterName", "ParameterValue", "Message", "RawLine"]
    write_generic_table_sheet(wb, "PSC_Params", param_headers, [[r.get(h, "") for h in param_headers] for r in param_rows], "PSCParamTable")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"{path.stem}_PSC_Import_{stamp}.xlsx"
    wb.save(out_path)
    strip_excel_table_xml(out_path)
    return str(out_path)


def import_reviewout_only(file_path: str, output_folder: str = "") -> str:
    path = Path(file_path)
    out_dir = Path(output_folder) if output_folder else path.parent
    scan_list = parse_reviewout_file(path)
    if not scan_list:
        raise ValueError("No valid review.out scan records were detected in the selected file.")
    all_keys: list[str] = ["date", "time"]
    if any("ID" in item for item in scan_list):
        all_keys.append("ID")
    for item in scan_list:
        for k in item.keys():
            if k not in all_keys:
                all_keys.append(k)
    wb = Workbook()
    ws = wb.active
    ws.title = "ScanSummary"
    write_rows(ws, all_keys, [[item.get(h, "") for h in all_keys] for item in scan_list])
    apply_table(ws, "ReviewOutScanSummary")
    autosize(ws)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = path.name
    if "." in base:
        base = base.rsplit(".", 1)[0]
    if base.lower().endswith(".out"):
        base = base[:-4]
    out_path = out_dir / f"{base}_reviewout_Import_{stamp}.xlsx"
    wb.save(out_path)
    strip_excel_table_xml(out_path)
    return str(out_path)



def import_selected_file_only(file_path: str, output_folder: str = "") -> str:
    """Import one selected file without merge.

    This is intentionally a file-level import/export path. It does not merge
    with other logs and it applies the same timestamp parser/output writer as
    the main Search sheet so operators can inspect one file quickly.
    """
    path = Path(file_path)
    out_dir = Path(output_folder) if output_folder else path.parent
    if is_review_file(path.name):
        return import_reviewout_only(file_path, output_folder)
    typ = classify_file(path, True) or "UNKNOWN"
    if typ == "PSC":
        return import_psc_only(file_path, output_folder)
    records = parse_file(path, typ)
    if not records:
        raise ValueError("No readable rows were detected in the selected file.")
    records.sort(key=lambda r: (r.timestamp is None, r.timestamp or datetime.max, r.line_no))
    wb = Workbook()
    ws = wb.active
    ws.title = "Search"
    headers = record_output_headers()
    write_rows(ws, headers, [record_to_output_row(r) for r in records])
    apply_table(ws, "Search")
    autosize(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    activate_search_sheet(wb, ws)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", path.stem)[:80]
    out_path = out_dir / f"{safe_stem}_{typ}_Import_{stamp}.xlsx"
    wb.save(out_path)
    strip_excel_table_xml(out_path)
    return str(out_path)

def check_csa_errors(records: list[LogRecord], rules: list[dict]) -> list[dict]:
    hits: list[dict] = []
    compiled = []
    for r in rules:
        kw = str(r.get("keyword", "")).strip()
        if not kw:
            continue
        compiled.append((kw, str(r.get("message", kw)), re.compile(re.escape(kw), re.I)))
    for rec in records:
        if rec.source_type not in {"CSA", "UNKNOWN"}:
            continue
        for kw, msg, pat in compiled:
            if pat.search(rec.raw) or pat.search(rec.message):
                hits.append({
                    "Timestamp": rec.timestamp,
                    "File": rec.filename,
                    "Line": rec.line_no,
                    "Keyword": kw,
                    "Message": msg,
                    "Raw": rec.raw,
                })
    return hits


def autosize(ws, max_width: int = 70, max_scan_rows: int = 300) -> None:
    """Fast autosize.

    Large log sheets can contain many rows. Scanning every cell for width makes
    the app look frozen after the file scan reaches 100%, so only the header and
    first max_scan_rows rows are sampled.
    """
    max_row = min(ws.max_row, max_scan_rows + 1)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def apply_table(ws, name: str, max_table_rows: int = 50000) -> None:
    """Do not create Excel Table objects.

    Previous versions added openpyxl Table objects and AutoFilter together on
    multiple sheets. Excel repaired the generated workbook and removed
    /xl/tables/table*.xml. For field reliability, v13 uses worksheet
    AutoFilter only. This keeps filter drop-downs without creating table XML.
    """
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    return


def write_rows(ws, headers: list[str], rows: list[list], fast_threshold: int = 50000, progress_cb=None, progress_start: int = 0, progress_total: int = 0, progress_label: str = "Writing rows", control=None, styles: bool = True, timestamp_format: bool = True) -> int:
    """Write rows to a worksheet with row-level progress callbacks.

    openpyxl cannot report progress during the final ZIP compression/save phase,
    but it can report progress while rows are appended.  For large outputs this
    keeps the UI popup moving instead of staying at 0% until the save completes.

    Returns the number of data rows written.
    """
    headers = [clean_excel_value(h) for h in headers]
    ws.append(headers)
    row_count = len(rows)
    update_every = 1000 if row_count < 50000 else 5000
    total_for_cb = progress_total or max(1, row_count)
    for i, row in enumerate(rows, start=1):
        if control:
            control.check()
        ws.append([clean_excel_value(v) for v in row])
        if progress_cb and (i == 1 or i % update_every == 0 or i == row_count):
            progress_cb(min(progress_start + i, total_for_cb), total_for_cb, f"{progress_label}: {i}/{row_count} rows")

    if styles:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fast path for large sheets. Avoid styling every cell.
        if row_count <= fast_threshold:
            thin = Side(style="thin", color="D9E2F3")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Required output behavior: freeze at C2 and enable filters.
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Timestamp must be explicit on every row as yyyy/mm/dd hh:mm:ss.0.
    if headers and str(headers[0]).strip().lower() == "timestamp":
        ws.column_dimensions["A"].width = 24
        if timestamp_format:
            for cell in ws["A"][1:]:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy/mm/dd hh:mm:ss.0"
    return row_count


def activate_search_sheet(wb: Workbook, ws) -> None:
    """Open workbook on Search sheet with C2 selected/frozen."""
    try:
        wb.active = wb.sheetnames.index(ws.title)
    except Exception:
        pass
    try:
        ws.freeze_panes = "C2"
        ws.sheet_view.selection[0].activeCell = "C2"
        ws.sheet_view.selection[0].sqref = "C2"
    except Exception:
        pass



def strip_excel_table_xml(xlsx_path: Path) -> None:
    """Remove Excel Table parts from an xlsx as a final safety pass.

    Field Excel reported repairs in /xl/tables/table*.xml when older builds
    produced ListObject/Table XML. v15 uses worksheet AutoFilter only, but this
    sanitizer guarantees the final workbook contains no xl/tables parts even if
    a legacy code path accidentally creates them.
    """
    try:
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            return
        tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
        ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
        ET.register_namespace('', ns_main)
        ET.register_namespace('', ns_rel)
        with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                name = item.filename
                if name.startswith('xl/tables/'):
                    continue
                data = zin.read(name)
                if name == '[Content_Types].xml':
                    try:
                        root = ET.fromstring(data)
                        for elem in list(root):
                            if elem.attrib.get('PartName', '').startswith('/xl/tables/'):
                                root.remove(elem)
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except Exception:
                        pass
                elif name.startswith('xl/worksheets/_rels/') and name.endswith('.rels'):
                    try:
                        root = ET.fromstring(data)
                        for elem in list(root):
                            if '/table' in elem.attrib.get('Type', '') or elem.attrib.get('Target', '').startswith('../tables/'):
                                root.remove(elem)
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except Exception:
                        pass
                elif name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                    try:
                        root = ET.fromstring(data)
                        table_parts = root.find(f'{{{ns_main}}}tableParts')
                        if table_parts is not None:
                            root.remove(table_parts)
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except Exception:
                        pass
                zout.writestr(item, data)
        tmp_path.replace(xlsx_path)
    except Exception:
        # Do not block output creation if the safety pass itself fails.
        pass

def safe_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return format_timestamp(value)
    return str(value).replace("\x00", "")


def write_records_csv(records: list[LogRecord], csv_path: Path, log_cb=None, progress_cb=None, progress_offset: int = 0, progress_total: int = 0, control=None) -> None:
    if log_cb:
        log_cb(f"Writing fast CSV: {csv_path.name}")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    headers = record_output_headers()
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i, r in enumerate(records, start=1):
            if control and i % 1000 == 0:
                control.check()
            if progress_cb and (i % 5000 == 0 or i == len(records)):
                progress_cb(progress_offset + i, progress_total or len(records), f"Writing {csv_path.name}: {i}/{len(records)} rows")
            writer.writerow(record_to_output_row(r))


def write_dict_rows_csv(rows: list[dict[str, Any]], csv_path: Path, log_cb=None, progress_cb=None, progress_offset: int = 0, progress_total: int = 0, control=None) -> None:
    if not rows:
        return
    if log_cb:
        log_cb(f"Writing fast CSV: {csv_path.name}")
    headers: list[str] = []
    for row in rows:
        for k in row.keys():
            if k not in headers:
                headers.append(k)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i, row in enumerate(rows, start=1):
            if control and i % 1000 == 0:
                control.check()
            if progress_cb and (i % 5000 == 0 or i == len(rows)):
                progress_cb(progress_offset + i, progress_total or len(rows), f"Writing {csv_path.name}: {i}/{len(rows)} rows")
            writer.writerow([safe_csv_value(row.get(h, "")) for h in headers])




def safe_sha256(path: Path, max_bytes: int = 1024 * 1024 * 64) -> str:
    """Return a SHA256 hash for traceability. Large files are still hashed in chunks."""
    import hashlib
    try:
        h = hashlib.sha256()
        with path.open("rb") as f:
            while True:
                b = f.read(1024 * 1024)
                if not b:
                    break
                h.update(b)
        return h.hexdigest()
    except Exception:
        return ""


def make_manifest_path(output_path: Path) -> Path:
    return output_path.with_name(output_path.stem + "_manifest.json")


def write_merge_manifest(output_path: Path, options: RunOptions, result: RunResult, source_entries: list[dict[str, Any]], extra_outputs: Optional[list[str]] = None, log_cb=None) -> Path:
    """Write a JSON sidecar describing exactly which files created a merge output."""
    manifest_path = make_manifest_path(output_path)
    data = {
        "schema": "LogMergeTool.MergeManifest.v1",
        "tool_version": APP_VERSION,
        "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
        "output_file": str(output_path),
        "output_name": output_path.name,
        "mode": "merge",
        "site": options.site,
        "serial": options.serial,
        "source_folder": options.source_folder,
        "output_folder": options.output_folder,
        "recursive": bool(options.recursive),
        "date_filter": {
            "enabled": bool(options.use_start or options.use_end),
            "start": options.start_date if options.use_start else "",
            "end": options.end_date if options.use_end else "",
            "basis": "content timestamp per row; filename date fallback only when body date is unavailable; file created/modified date is never used",
        },
        "performance": {
            "turbo_csv": bool(getattr(options, "turbo_csv", False)),
            "summary_xlsx_only": bool(getattr(options, "summary_xlsx_only", False)),
            "fast_xlsx_mode": bool(getattr(options, "fast_xlsx_mode", False)),
            "csv_summary_mode": bool(getattr(options, "csv_summary_mode", False)),
            "skip_merged1": bool(getattr(options, "skip_merged1", False)),
            "timestamp_as_text": bool(getattr(options, "timestamp_as_text", False)),
            "disable_styles_large": bool(getattr(options, "disable_styles_large", False)),
            "worker_count": int(getattr(options, "worker_count", 0) or 0),
        },
        "noise_rules": {
            "enabled_in_merge": bool(getattr(options, "noise_enabled_merge", False)),
            "exclude_output": bool(getattr(options, "noise_exclude_output", False)),
            "learning": bool(getattr(options, "noise_learning", False)),
        },
        "run_result": asdict(result),
        "source_files": source_entries,
        "extra_outputs": extra_outputs or [],
        "split_support": {
            "supported": True,
            "note": "Split uses Search sheet or Search CSV rows grouped by SourceType and File. It reconstructs filtered output rows, not records removed by Noise rules.",
        },
    }
    manifest_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if log_cb:
        log_cb(f"Manifest written: {manifest_path}")
    return manifest_path


def discover_related_manifest(merge_path: Path) -> Optional[Path]:
    candidates = [
        merge_path.with_name(merge_path.stem + "_manifest.json"),
        merge_path.with_suffix(".json"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def split_merge_project(merge_file: str, output_folder: str, progress_cb=None, log_cb=None) -> str:
    """Split a merged Search sheet/CSV back into per-source reconstructed logs.

    This reconstructs what exists in the merged output. If Noise rules excluded rows
    before output, those excluded rows cannot be reconstructed from the merge file.
    """
    def log(msg: str):
        if log_cb:
            log_cb(msg)
    src = Path(merge_file)
    if not src.exists():
        raise FileNotFoundError(str(src))
    out_root = Path(output_folder) if output_folder else src.parent / (src.stem + "_split")
    out_root.mkdir(parents=True, exist_ok=True)
    rows_by_key: dict[tuple[str, str], list[str]] = {}
    count = 0

    def add_row(row: dict[str, Any]):
        nonlocal count
        typ = str(row.get("SourceType") or row.get("Type") or "UNKNOWN").strip() or "UNKNOWN"
        fname = str(row.get("File") or row.get("Filename") or f"{typ}.log").strip() or f"{typ}.log"
        raw = str(row.get("Raw") or row.get("Message") or "")
        rows_by_key.setdefault((typ, fname), []).append(raw)
        count += 1
        if progress_cb and count % 5000 == 0:
            progress_cb(count, max(count, count + 1), f"Reading merged rows: {count}")

    if src.suffix.lower() == ".csv":
        with src.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                add_row(row)
    else:
        wb = load_workbook(src, read_only=True, data_only=True)
        sheet_name = "Search" if "Search" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        headers = [str(x or "").strip() for x in next(it)]
        for vals in it:
            row = {h: vals[i] if i < len(vals) else "" for i, h in enumerate(headers)}
            add_row(row)
        wb.close()

    written_files = []
    total_groups = max(1, len(rows_by_key))
    for idx, ((typ, fname), lines) in enumerate(sorted(rows_by_key.items()), start=1):
        safe_type = re.sub(r"[^A-Za-z0-9_-]+", "_", typ) or "UNKNOWN"
        safe_name = Path(re.sub(r"[\\/:*?\"<>|]+", "_", fname)).name or f"{safe_type}.log"
        target_dir = out_root / safe_type
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / safe_name
        # If multiple groups would collide, add an index suffix.
        if target.exists():
            target = target_dir / f"{target.stem}_{idx}{target.suffix or '.log'}"
        target.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8", errors="replace")
        written_files.append({"type": typ, "file": fname, "output": str(target), "rows": len(lines)})
        if progress_cb:
            progress_cb(idx, total_groups, f"Writing {target.name}")

    manifest = discover_related_manifest(src)
    split_manifest = {
        "schema": "LogMergeTool.SplitManifest.v1",
        "tool_version": APP_VERSION,
        "created_at": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
        "source_merge_file": str(src),
        "source_manifest": str(manifest) if manifest else "",
        "output_folder": str(out_root),
        "total_rows": count,
        "files_written": written_files,
        "note": "This split recreates merged/filtered output rows. Rows not present in the merge output cannot be recovered.",
    }
    (out_root / "split_manifest.json").write_text(json.dumps(split_manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Split completed: {out_root}")
    return str(out_root)

def build_summary_only_xlsx(records: list[LogRecord], csa_hits: list[dict], result: RunResult, options: RunOptions, output_path: Path, csv_files: list[str], log_cb=None) -> None:
    def outlog(msg: str):
        if log_cb:
            log_cb(msg)
    outlog("Creating summary xlsx workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"] = "Log Merge Tool - No Excel / v13 Timestamp + C2 Freeze + Filter (No Tables)"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    summary_rows = [
        ["Version", APP_VERSION],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Output Mode", "Turbo CSV + Summary xlsx"],
        ["Source Folder", options.source_folder],
        ["Output Folder", options.output_folder],
        ["Recursive", "Yes" if options.recursive else "No"],
        ["Start Date", options.start_date if options.use_start else "Not specified"],
        ["End Date", options.end_date if options.use_end else "Not specified"],
        ["Serial Number", options.serial],
        ["Site Name", options.site],
        ["Scanned Files", result.scanned_files],
        ["Parsed Files", result.parsed_files],
        ["Skipped Files", result.skipped_files],
        ["Total Records", result.total_records],
        ["Included Records", result.included_records],
        ["CSA Error Count", result.csa_error_count],
    ]
    for r, row in enumerate(summary_rows, start=3):
        ws.cell(r, 1, row[0]).font = Font(bold=True)
        ws.cell(r, 2, row[1])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90

    def write_summary_sheet_rows(ws, headers, rows, label: str = "Writing sheet"):
        """Summary workbook writer helper.

        v24/v25 could call a local function named write_sheet_rows that only
        existed in the full-xlsx writer path.  When CSV + Summary xlsx mode was
        selected, that caused NameError during summary creation.  Keep this
        helper local and use the shared safe write_rows implementation.
        """
        outlog(label)
        return write_rows(ws, headers, rows, styles=True, timestamp_format=True)

    ws_files = wb.create_sheet("Output Files")
    write_summary_sheet_rows(ws_files, ["File", "Note"], [[f, "Open with Excel. Full merged data is here." if f.lower().endswith('.csv') else "Workbook output"] for f in csv_files], "Writing Output Files sheet")
    autosize(ws_files)

    type_counts: dict[str, int] = {}
    for r in records:
        type_counts[r.source_type] = type_counts.get(r.source_type, 0) + 1
    ws_type = wb.create_sheet("Type_Summary")
    write_summary_sheet_rows(ws_type, ["SourceType", "RecordCount"], [[k, type_counts[k]] for k in sorted(type_counts.keys())], "Writing Type Summary")
    autosize(ws_type)

    error_records = [r for r in records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"}]
    ws_errors = wb.create_sheet("Errors_Sample")
    headers = record_output_headers()
    sample = error_records[:5000]
    write_summary_sheet_rows(ws_errors, headers, [record_to_output_row(r) for r in sample], "Writing Error Sample")
    autosize(ws_errors)

    ws_csa_hits = wb.create_sheet("CSA Error Hits")
    csa_headers = ["Timestamp", "File", "Line", "Keyword", "Message", "Raw"]
    csa_rows = [csa_hit_to_output_row(h) for h in csa_hits[:5000]]
    write_summary_sheet_rows(ws_csa_hits, csa_headers, csa_rows, "Writing CSA Error Hits")
    autosize(ws_csa_hits)

    ws_rules = wb.create_sheet("CSA Error List")
    write_summary_sheet_rows(ws_rules, ["Keyword", "Message"], [[r.get("keyword", ""), r.get("message", "")] for r in load_json_file("csa_error_rules.json", DEFAULT_CSA_RULES)], "Writing CSA Error List")
    autosize(ws_rules)

    ws_lists = wb.create_sheet("Lists")
    write_summary_sheet_rows(ws_lists, ["Serial Number", "Site Name"], [[r.get("serial", ""), r.get("site", "")] for r in load_json_file("site_serial_map.json", DEFAULT_SITE_MAP)], "Writing Lists sheet")
    autosize(ws_lists)

    outlog("Saving summary xlsx file...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    strip_excel_table_xml(output_path)
    outlog("summary xlsx save completed.")

def build_output(records: list[LogRecord], csa_hits: list[dict], result: RunResult, options: RunOptions, output_path: Path, psc_log_rows: Optional[list[dict[str, Any]]] = None, psc_param_rows: Optional[list[dict[str, Any]]] = None, log_cb=None, progress_cb=None, control=None) -> None:
    def outlog(msg: str):
        if log_cb:
            log_cb(msg)
    # v24 Fast Excel Writer. For very large outputs, writing full xlsx is slow
    # because the final ZIP/save phase has no internal progress. CSV + summary
    # xlsx is therefore available as a practical high-speed path.
    turbo_threshold = 100000
    if getattr(options, "turbo_csv", True) and getattr(options, "summary_xlsx_only", True) and getattr(options, "csv_summary_mode", True) and len(records) >= turbo_threshold:
        outlog(f"Turbo CSV mode enabled for {len(records)} rows.")
        if control:
            control.check()
        sorted_records = sorted(records, key=lambda r: r.timestamp or datetime.min, reverse=True)
        base = output_path.with_suffix("")
        csv_files: list[str] = []
        search_csv = base.parent / f"{base.name}_Search.csv"
        write_total = max(1, len(sorted_records) + len(psc_log_rows or []) + len(psc_param_rows or []))
        written_offset = 0
        write_records_csv(sorted_records, search_csv, log_cb, progress_cb, written_offset, write_total, control)
        written_offset += len(sorted_records)
        csv_files.append(str(search_csv))
        error_records = [r for r in sorted_records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"}]
        if error_records:
            error_csv = base.parent / f"{base.name}_Errors.csv"
            write_records_csv(error_records, error_csv, log_cb, None, 0, 0, control)
            csv_files.append(str(error_csv))
        if psc_log_rows:
            psc_csv = base.parent / f"{base.name}_PSC_Log.csv"
            write_dict_rows_csv(psc_log_rows, psc_csv, log_cb, progress_cb, written_offset, write_total, control)
            written_offset += len(psc_log_rows)
            csv_files.append(str(psc_csv))
        if psc_param_rows:
            psc_param_csv = base.parent / f"{base.name}_PSC_Params.csv"
            write_dict_rows_csv(psc_param_rows, psc_param_csv, log_cb, progress_cb, written_offset, write_total, control)
            written_offset += len(psc_param_rows)
            csv_files.append(str(psc_param_csv))
        if control:
            control.check()
        if progress_cb:
            progress_cb(write_total, write_total, "Creating summary workbook...")
        build_summary_only_xlsx(sorted_records, csa_hits, result, options, output_path, csv_files, log_cb)
        return

    outlog("Creating workbook...")
    # Approximate progress units for xlsx creation.  The final wb.save() step is
    # a ZIP compression phase that openpyxl cannot subdivide, so reserve a small
    # tail range for it and report row-by-row progress before that phase.
    est_error_count = sum(1 for r in records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"})
    est_type_rows = 0 if len(records) > 200000 else len(records)
    est_merged_rows = 1 if len(records) > 100000 else len(records)
    write_total = max(1, len(records) + est_error_count + len(csa_hits) + est_type_rows + est_merged_rows + len(psc_log_rows or []) + len(psc_param_rows or []) + 5000)
    write_cursor = 0
    fast_xlsx_mode = bool(getattr(options, "fast_xlsx_mode", True))
    timestamp_as_text = bool(getattr(options, "timestamp_as_text", True))
    disable_styles_large = bool(getattr(options, "disable_styles_large", True))
    no_styles = bool(disable_styles_large and len(records) >= 50000)

    def write_progress(label: str):
        if progress_cb:
            progress_cb(min(write_cursor, write_total), write_total, label)

    def write_sheet_rows(ws, headers, rows, label: str):
        nonlocal write_cursor
        n = write_rows(ws, headers, rows, progress_cb=progress_cb, progress_start=write_cursor, progress_total=write_total, progress_label=label, control=control, styles=not no_styles, timestamp_format=not timestamp_as_text)
        write_cursor += max(1, n)
        write_progress(f"{label}: completed")
        return n

    def write_generic_progress(title: str, headers, rows, table_name: str):
        nonlocal write_cursor
        wsx = wb.create_sheet(title)
        n = write_rows(wsx, headers, rows, progress_cb=progress_cb, progress_start=write_cursor, progress_total=write_total, progress_label=f"Writing {title}", control=control, styles=not no_styles, timestamp_format=not timestamp_as_text)
        write_cursor += max(1, n)
        apply_table(wsx, table_name)
        autosize(wsx)
        write_progress(f"Writing {title}: completed")
        return wsx

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"] = "Log Merge Tool - No Excel"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    summary_rows = [
        ["Version", APP_VERSION],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source Folder", options.source_folder],
        ["Output Folder", options.output_folder],
        ["Recursive", "Yes" if options.recursive else "No"],
        ["Start Date", options.start_date if options.use_start else "Not specified"],
        ["End Date", options.end_date if options.use_end else "Not specified"],
        ["Serial Number", options.serial],
        ["Site Name", options.site],
        ["Scanned Files", result.scanned_files],
        ["Parsed Files", result.parsed_files],
        ["Skipped Files", result.skipped_files],
        ["Total Records", result.total_records],
        ["Included Records", result.included_records],
        ["CSA Error Count", result.csa_error_count],
    ]
    for r, row in enumerate(summary_rows, start=3):
        ws.cell(r, 1, row[0]).font = Font(bold=True)
        ws.cell(r, 2, row[1])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80

    # Search sheet: all merged records, sorted descending by timestamp.
    outlog(f"Sorting and writing merged records: {len(records)} rows")
    sorted_records = sorted(records, key=lambda r: r.timestamp or datetime.min, reverse=True)
    headers = record_output_headers()
    rows = [record_to_output_row_fast(r, timestamp_as_text=timestamp_as_text) for r in sorted_records]
    ws_search = wb.create_sheet("Search")
    outlog("Writing Search sheet rows...")
    write_sheet_rows(ws_search, headers, rows, "Writing Search sheet")
    outlog("Applying Search sheet timestamp format, C2 freeze, filter, and width settings...")
    apply_table(ws_search, "SearchTable")
    if not fast_xlsx_mode:
        autosize(ws_search)
    ws_search.column_dimensions["A"].width = 24
    ws_search.column_dimensions["G"].width = 80
    ws_search.column_dimensions["H"].width = 100
    activate_search_sheet(wb, ws_search)

    outlog("Writing PSC detail sheets...")
    # PSC sheets: merge-capable PSC log import, equivalent to ReadPSClog_LogMergeReady output.
    psc_log_rows = psc_log_rows or []
    psc_param_rows = psc_param_rows or []
    if psc_log_rows:
        max_param = max((int(r.get("ParamCount", 0)) for r in psc_log_rows), default=0)
        psc_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "Message", "ParameterText", "ParamCount", "RawLine"] + [f"Param{i}" for i in range(1, max_param + 1)]
        write_generic_progress("PSC_Log", psc_headers, [[r.get(h, "") for h in psc_headers] for r in psc_log_rows], "PSCLogTable")
    if psc_param_rows:
        psc_param_headers = ["SourceFile", "LogDate", "Time", "DateTime", "Process", "Module", "Function", "ParameterName", "ParameterValue", "Message", "RawLine"]
        write_generic_progress("PSC_Params", psc_param_headers, [[r.get(h, "") for h in psc_param_headers] for r in psc_param_rows], "PSCParamTable")

    outlog("Writing Errors sheet...")
    # Errors sheet.
    error_records = [r for r in sorted_records if r.category in {"ERROR", "WARNING"} or r.level in {"ERROR", "ERR", "WARN", "WARNING", "FATAL", "ALARM"}]
    ws_errors = wb.create_sheet("Errors")
    write_sheet_rows(ws_errors, headers, [record_to_output_row_fast(r, timestamp_as_text=timestamp_as_text) for r in error_records], "Writing Errors sheet")
    apply_table(ws_errors, "ErrorsTable")
    autosize(ws_errors)
    if ws_errors.max_row >= 2 and not no_styles:
        ws_errors.conditional_formatting.add(f"A2:H{ws_errors.max_row}", FormulaRule(formula=['OR($E2="ERROR",$E2="ERR",$E2="FATAL",$E2="ALARM",$F2="ERROR")'], fill=PatternFill("solid", fgColor="F4CCCC")))

    # CSA Error List hit sheet.
    ws_csa_hits = wb.create_sheet("CSA Error Hits")
    csa_headers = ["Timestamp", "File", "Line", "Keyword", "Message", "Raw"]
    csa_rows = [csa_hit_to_output_row(h) for h in csa_hits]
    write_rows(ws_csa_hits, csa_headers, csa_rows)
    apply_table(ws_csa_hits, "CSAErrorHits")
    autosize(ws_csa_hits)

    outlog("Writing log-type sheets...")
    # Type sheets. Search always contains all records. For very large outputs,
    # full duplicate per-type sheets can make xlsx creation appear stuck.
    if len(sorted_records) > 200000 or fast_xlsx_mode:
        ws_type_summary = wb.create_sheet("Type_Summary")
        summary = []
        for typ in ["WS", "CGA", "CSA", "PSC", "MRSERVER", "GESYS", "LAIS", "UNKNOWN"]:
            cnt = sum(1 for r in sorted_records if r.source_type == typ)
            if cnt:
                summary.append([typ, cnt, "Full records are in Search sheet"])
        write_sheet_rows(ws_type_summary, ["SourceType", "RecordCount", "Note"], summary, "Writing Type Summary")
        apply_table(ws_type_summary, "TypeSummaryTable")
        autosize(ws_type_summary)
    else:
        for typ in ["WS", "CGA", "CSA", "PSC", "MRSERVER", "GESYS", "LAIS", "UNKNOWN"]:
            typ_records = [r for r in sorted_records if r.source_type == typ]
            if not typ_records:
                continue
            ws_typ = wb.create_sheet(typ)
            write_sheet_rows(ws_typ, headers, [record_to_output_row(r) for r in typ_records], f"Writing {typ} sheet")
            apply_table(ws_typ, f"{typ}Table")
            autosize(ws_typ)

    outlog("Writing Merged_1 compatibility sheet...")
    # Merged_1 compatibility sheet. For very large outputs, duplicating the full
    # Search sheet can double the save time and file size. Keep a lightweight
    # compatibility sheet instead; Search contains the complete merged data.
    ws_merged = wb.create_sheet("Merged_1")
    if len(rows) > 100000 or bool(getattr(options, "skip_merged1", True)):
        write_sheet_rows(ws_merged, ["Notice"], [["Full merged records are in the Search sheet. Merged_1 full duplicate was skipped for performance because the output exceeded 100,000 rows."]], "Writing Merged_1 notice")
    else:
        write_sheet_rows(ws_merged, headers, rows, "Writing Merged_1 compatibility sheet")
        apply_table(ws_merged, "MergedTable")
        autosize(ws_merged)

    # Config sheets for traceability.
    ws_rules = wb.create_sheet("CSA Error List")
    write_rows(ws_rules, ["Keyword", "Message"], [[r.get("keyword", ""), r.get("message", "")] for r in load_json_file("csa_error_rules.json", DEFAULT_CSA_RULES)])
    apply_table(ws_rules, "CSAErrorList")
    autosize(ws_rules)

    ws_lists = wb.create_sheet("Lists")
    write_rows(ws_lists, ["Serial Number", "Site Name"], [[r.get("serial", ""), r.get("site", "")] for r in load_json_file("site_serial_map.json", DEFAULT_SITE_MAP)])
    apply_table(ws_lists, "SiteSerialList")
    autosize(ws_lists)

    # Ensure the user opens directly on Search with C2 freeze/filter applied.
    try:
        activate_search_sheet(wb, ws_search)
    except Exception:
        pass

    outlog("Saving xlsx file... Please wait.")
    if progress_cb:
        progress_cb(max(0, write_total - 5000), write_total, "Finalizing workbook and compressing xlsx... Excel library cannot report sub-progress in this step.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    if progress_cb:
        progress_cb(write_total - 1000, write_total, "Validating xlsx package...")
    strip_excel_table_xml(output_path)
    if progress_cb:
        progress_cb(write_total, write_total, "xlsx save completed.")
    outlog("xlsx save completed.")



def parse_candidate_file(args) -> tuple[Path, Optional[str], int, list[LogRecord], list[dict[str, Any]], list[dict[str, Any]], Optional[str]]:
    p, include_unknown = args
    typ = classify_file(p, include_unknown)
    if typ is None:
        return p, None, 0, [], [], [], None
    try:
        if typ == "PSC":
            raw_psc_log, raw_psc_params, parsed = parse_psc_file_detail(p)
        else:
            raw_psc_log, raw_psc_params = [], []
            parsed = parse_file(p, typ)
        return p, typ, len(parsed), parsed, raw_psc_log, raw_psc_params, None
    except Exception as exc:
        return p, typ, 0, [], [], [], str(exc)

def run_merge(options: RunOptions, progress_cb=None, log_cb=None, control=None) -> RunResult:
    def log(msg: str):
        if log_cb:
            log_cb(msg)

    src = Path(options.source_folder)
    out_dir = Path(options.output_folder) if options.output_folder else src
    if not src.exists():
        raise FileNotFoundError(f"Source folder not found: {src}")
    if not out_dir.exists():
        out_dir.mkdir(parents=True, exist_ok=True)

    start = date.fromisoformat(options.start_date) if options.use_start and options.start_date else None
    end = date.fromisoformat(options.end_date) if options.use_end and options.end_date else None
    if start and end and start > end:
        raise ValueError("Start Date is later than End Date.")

    enabled = enabled_types(options)
    if not enabled:
        raise ValueError("No log type is selected.")

    if getattr(options, "selected_files", None):
        files = [Path(x) for x in options.selected_files if Path(x).exists()]
        log(f"Using Smart File Discovery selected files: {len(files)}")
    else:
        files = list(iter_files(src, options.recursive))
    result = RunResult(scanned_files=len(files))
    records: list[LogRecord] = []
    psc_log_rows: list[dict[str, Any]] = []
    psc_param_rows: list[dict[str, Any]] = []
    source_entries: list[dict[str, Any]] = []

    log(f"Scanned files: {len(files)}")
    log("v15 mode: Date filter uses content timestamp first, filename date fallback only, never file created/modified date. XLSX writer removes all Excel Table XML and uses AutoFilter only.")
    # v6 Turbo parser: parse files concurrently. This is mostly I/O-bound, so
    # ThreadPoolExecutor is effective and avoids multiprocessing freeze issues in onefile EXEs.
    max_workers = int(getattr(options, "worker_count", 0) or 0)
    if max_workers <= 0:
        try:
            max_workers = min(8, max(2, (os.cpu_count() or 4)))
        except Exception:
            max_workers = 4
    log(f"Turbo parser workers: {max_workers}")
    completed = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(parse_candidate_file, (p, options.include_unknown)): p for p in files}
        for fut in concurrent.futures.as_completed(future_map):
            if control:
                control.check()
            p = future_map[fut]
            completed += 1
            if progress_cb:
                progress_cb(completed, len(files), p.name)
            try:
                p, typ, parsed_count, parsed, raw_psc_log, raw_psc_params, err = fut.result()
            except Exception as exc:
                result.skipped_files += 1
                log(f"Skipped unreadable file: {p.name} / {exc}")
                continue
            if typ is None or typ not in enabled:
                result.skipped_files += 1
                continue
            if err:
                result.skipped_files += 1
                log(f"Skipped unreadable file: {p.name} / {err}")
                continue
            result.parsed_files += 1
            before_total = result.total_records
            before_included = result.included_records
            file_first_ts = min((r.timestamp for r in parsed if r.timestamp), default=None)
            file_last_ts = max((r.timestamp for r in parsed if r.timestamp), default=None)
            for rec in parsed:
                result.total_records += 1
                if in_date_range(rec.timestamp, None, start, end):
                    records.append(rec)
                    result.included_records += 1
            try:
                stat = p.stat()
                file_size = stat.st_size
            except Exception:
                file_size = 0
            source_entries.append({
                "type": typ,
                "file_name": p.name,
                "path": str(p),
                "size_bytes": file_size,
                "sha256": safe_sha256(p),
                "first_timestamp": format_timestamp(file_first_ts),
                "last_timestamp": format_timestamp(file_last_ts),
                "rows_read": len(parsed),
                "rows_output": result.included_records - before_included,
                "noise_excluded_rows": 0,
            })
            if typ == "PSC":
                for r in raw_psc_log:
                    ts = parse_datetime_from_text(str(r.get("DateTime", "")))
                    if in_date_range(ts, None, start, end):
                        psc_log_rows.append(r)
                for r in raw_psc_params:
                    ts = parse_datetime_from_text(str(r.get("DateTime", "")))
                    if in_date_range(ts, None, start, end):
                        psc_param_rows.append(r)

    before_noise_count = len(records)
    records = apply_noise_to_records(records, options.noise_enabled_merge, options.noise_exclude_output, options.noise_learning, log)
    noise_removed_total = max(0, before_noise_count - len(records))
    if noise_removed_total:
        log(f"Noise excluded rows for manifest: {noise_removed_total}")
    rules = load_json_file("csa_error_rules.json", DEFAULT_CSA_RULES)
    csa_hits = check_csa_errors(records, rules)
    result.csa_error_count = len(csa_hits)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    site_part = re.sub(r"[^A-Za-z0-9_-]+", "_", options.site.strip()) if options.site.strip() else "NoSite"
    serial_part = re.sub(r"[^A-Za-z0-9_-]+", "_", options.serial.strip()) if options.serial.strip() else "NoSerial"
    output_path = out_dir / f"LogMerge_NoExcel_{site_part}_{serial_part}_{stamp}.xlsx"
    log(f"Scan/parse completed. Parsed files: {result.parsed_files}, included records: {result.included_records}")
    if progress_cb:
        write_total = max(1, result.included_records + len(psc_log_rows) + len(psc_param_rows))
        progress_cb(0, write_total, "Writing output files...")
    log("Now writing output files. For large folders this step can take several minutes.")
    build_output(records, csa_hits, result, options, output_path, psc_log_rows, psc_param_rows, log, progress_cb, control)
    result.output_path = str(output_path)
    try:
        # Store final output row counts after Noise filtering.
        output_counts: dict[tuple[str, str], int] = {}
        for r in records:
            output_counts[(r.source_type, r.filename)] = output_counts.get((r.source_type, r.filename), 0) + 1
        for ent in source_entries:
            ent["rows_output_after_noise"] = output_counts.get((ent.get("type", ""), ent.get("file_name", "")), 0)
        manifest_path = write_merge_manifest(output_path, options, result, source_entries, [], log)
        result.output_path = str(output_path)
        log(f"Merge manifest: {manifest_path}")
    except Exception:
        log("Manifest write failed:\n" + traceback.format_exc())
    return result


class RunControl:
    def __init__(self):
        self._cancel = threading.Event()
        self._pause = threading.Event()

    def request_cancel(self):
        self._cancel.set()

    def set_paused(self, paused: bool):
        if paused:
            self._pause.set()
        else:
            self._pause.clear()

    def check(self):
        if self._cancel.is_set():
            raise RuntimeError("Operation cancelled by user.")
        while self._pause.is_set():
            if self._cancel.is_set():
                raise RuntimeError("Operation cancelled by user.")
            time.sleep(0.15)


class MergeWorker(QThread):
    progress = Signal(int, int, str)
    log = Signal(str)
    finished_ok = Signal(object)
    failed = Signal(str)

    def __init__(self, options: RunOptions):
        super().__init__()
        self.options = options
        self.control = RunControl()

    def request_cancel(self):
        self.control.request_cancel()

    def set_paused(self, paused: bool):
        self.control.set_paused(paused)

    def run(self):
        try:
            result = run_merge(self.options, self.progress.emit, self.log.emit, self.control)
            self.finished_ok.emit(result)
        except Exception as exc:
            if "cancelled by user" in str(exc).lower():
                self.failed.emit("Operation cancelled by user.")
            else:
                self.failed.emit(traceback.format_exc())



class SimpleTaskWorker(QThread):
    finished_ok = Signal(object)
    failed = Signal(str)

    def __init__(self, func):
        super().__init__()
        self.func = func

    def run(self):
        try:
            self.finished_ok.emit(self.func())
        except Exception:
            self.failed.emit(traceback.format_exc())



# ---------------------------------------------------------------------------
# v17 Smart Noise Learning
# ---------------------------------------------------------------------------
NOISE_DB_NAME = "NoiseDB.sqlite"
NOISE_SUGGEST_LIMIT = 200
NOISE_KEYWORDS = [
    "heartbeat", "heart beat", "keepalive", "keep alive", "poll", "polling",
    "status ok", "alive", "timer", "tick", "debug trace", "periodic",
    "no change", "idle", "socket select", "refresh", "update display",
]


def runtime_data_dir() -> Path:
    base = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    d = Path(base) / "LogMergeTool_NoExcel"
    d.mkdir(parents=True, exist_ok=True)
    return d


def noise_db_path() -> Path:
    return runtime_data_dir() / NOISE_DB_NAME


def init_noise_db() -> None:
    con = sqlite3.connect(noise_db_path())
    try:
        cur = con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS noise_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pattern TEXT NOT NULL,
                match_type TEXT NOT NULL DEFAULT 'contains',
                log_type TEXT NOT NULL DEFAULT 'ANY',
                action TEXT NOT NULL DEFAULT 'hide',
                scope TEXT NOT NULL DEFAULT 'both',
                enabled INTEGER NOT NULL DEFAULT 1,
                approved INTEGER NOT NULL DEFAULT 1,
                confidence REAL NOT NULL DEFAULT 1.0,
                hit_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                notes TEXT DEFAULT ''
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS noise_samples (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_type TEXT NOT NULL,
                message TEXT NOT NULL,
                raw TEXT NOT NULL,
                label TEXT NOT NULL,
                score REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        con.commit()
    finally:
        con.close()


def _compact_pattern(text: str) -> str:
    t = re.sub(r"\s+", " ", text or "").strip()
    # Remove highly variable values to make suggested rules reusable but safe.
    t = re.sub(r"\b20\d{2}[-/]\d{1,2}[-/]\d{1,2}\b", "", t)
    t = re.sub(r"\b\d{1,2}:\d{2}:\d{2}(?:[.:]\d+)?\b", "", t)
    t = re.sub(r"\b0x[0-9a-fA-F]+\b", "", t)
    t = re.sub(r"\b\d{5,}\b", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    if len(t) > 120:
        t = t[:120].rstrip()
    return t or (text or "")[:80]



def normalize_noise_text(text: str) -> str:
    """Normalize log text for operator Noise Rule matching.

    Real logs often contain tabs, doubled spaces, control characters, or values
    split differently between Message and Raw.  Operator rules should match the
    visible text even when whitespace differs.
    """
    if text is None:
        return ""
    t = str(text)
    t = clean_excel_value(t)
    t = t.replace("\t", " ").replace("\r", " ").replace("\n", " ")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def noise_record_text(record: LogRecord) -> tuple[str, str]:
    """Return raw and normalized searchable text for Noise matching."""
    parts = [
        record.message or "",
        record.raw or "",
        viewer_message_body(record.message or ""),
        viewer_message_body(record.raw or ""),
        record.level or "",
        record.category or "",
    ]
    raw_text = "\n".join(str(x) for x in parts if x is not None)
    return raw_text, normalize_noise_text(raw_text)


def increment_noise_hits(rule_ids: Iterable[int]) -> None:
    ids = [int(x) for x in rule_ids if x is not None]
    if not ids:
        return
    con = sqlite3.connect(noise_db_path())
    try:
        now = datetime.now().isoformat(timespec="seconds")
        for rid in ids:
            con.execute("UPDATE noise_rules SET hit_count = COALESCE(hit_count,0) + 1, updated_at=? WHERE id=?", (now, rid))
        con.commit()
    finally:
        con.close()

class NoiseEngine:
    def __init__(self):
        init_noise_db()
        self.rules = self.load_rules()

    def load_rules(self) -> list[dict[str, Any]]:
        con = sqlite3.connect(noise_db_path())
        con.row_factory = sqlite3.Row
        try:
            rows = con.execute(
                "SELECT * FROM noise_rules WHERE enabled=1 AND approved=1 ORDER BY id"
            ).fetchall()
            return [dict(r) for r in rows]
        finally:
            con.close()

    def reload(self):
        self.rules = self.load_rules()

    def match_rule(self, record: LogRecord) -> Optional[dict[str, Any]]:
        raw_text, text_n = noise_record_text(record)
        typ = (record.source_type or "").upper()
        for rule in self.rules:
            rt = str(rule.get("log_type") or "ANY").upper()
            if rt not in ("ANY", typ):
                continue
            pat = str(rule.get("pattern") or "")
            if not pat:
                continue
            pat_n = normalize_noise_text(pat)
            mt = str(rule.get("match_type") or "contains").lower()
            try:
                if mt == "regex":
                    # Regex is evaluated against both original and normalized text.
                    if re.search(pat, raw_text, re.I) or (pat_n and re.search(re.escape(pat_n), text_n, re.I)):
                        return rule
                elif mt == "startswith":
                    if text_n.startswith(pat_n):
                        return rule
                elif mt == "endswith":
                    if text_n.endswith(pat_n):
                        return rule
                else:
                    # Operator expectation: visible text should match even when
                    # whitespace/control characters differ.  Use normalized contains.
                    if pat_n and pat_n in text_n:
                        return rule
            except re.error:
                continue
        return None

    def is_noise(self, record: LogRecord) -> bool:
        return self.match_rule(record) is not None

    def score(self, record: LogRecord) -> int:
        text = ((record.message or "") + " " + (record.raw or "")).lower()
        score = 0
        if any(k in text for k in NOISE_KEYWORDS):
            score += 55
        if (record.level or "").upper() in {"DBG", "DEBUG", "TRACE"}:
            score += 15
        if len((record.message or record.raw or "").strip()) < 35:
            score += 10
        if re.search(r"\b(ok|normal|idle|ready|alive)\b", text):
            score += 10
        if re.search(r"\b(error|fail|failed|warn|alarm|exception|abort|timeout)\b", text):
            score -= 45
        return max(0, min(100, score))

    def add_sample(self, record: LogRecord, label: str, score: int = 0) -> None:
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute(
                "INSERT INTO noise_samples(log_type,message,raw,label,score,created_at) VALUES(?,?,?,?,?,?)",
                (record.source_type, record.message, record.raw, label, float(score), datetime.now().isoformat(timespec="seconds")),
            )
            con.commit()
        finally:
            con.close()

    def add_rule(self, pattern: str, log_type: str = "ANY", match_type: str = "contains", action: str = "hide", scope: str = "both", notes: str = "Approved by operator") -> int:
        pattern = re.sub(r"\s+", " ", str(pattern or "")).strip()
        now = datetime.now().isoformat(timespec="seconds")
        con = sqlite3.connect(noise_db_path())
        try:
            cur = con.cursor()
            cur.execute(
                """
                INSERT INTO noise_rules(pattern,match_type,log_type,action,scope,enabled,approved,confidence,created_at,updated_at,notes)
                VALUES(?,?,?,?,?,1,1,1.0,?,?,?)
                """,
                (pattern, match_type, (log_type or "ANY").upper(), action, scope, now, now, notes),
            )
            con.commit()
            rid = int(cur.lastrowid)
        finally:
            con.close()
        self.reload()
        return rid

    def add_rule_from_record(self, record: LogRecord, match_type: str = "contains", action: str = "hide", scope: str = "both") -> int:
        pattern = _compact_pattern(record.message or record.raw)
        rid = self.add_rule(pattern, (record.source_type or "ANY").upper(), match_type, action, scope, "Approved by operator from selected row")
        self.add_sample(record, "noise-approved", 100)
        return rid

    def mark_not_noise(self, record: LogRecord) -> None:
        self.add_sample(record, "not-noise", 0)


def apply_noise_to_records(records: list[LogRecord], enabled: bool, exclude_output: bool, learning: bool, log_cb=None) -> list[LogRecord]:
    """Apply Smart Noise rules.

    v25 behavior:
    - If enabled=True, approved/enabled rules are applied to the merge result and matching rows are removed.
      This matches the operator expectation that "Enable in Merge" actually filters the output.
    - exclude_output is kept for compatibility with older projects; either flag removes matching rows.
    - Learning mode only saves suggestions and never removes records by itself.
    """
    if not (enabled or exclude_output or learning):
        return records
    engine = NoiseEngine()
    kept: list[LogRecord] = []
    removed = 0
    suggested = 0
    do_remove = bool(enabled or exclude_output)
    matched_rule_ids: list[int] = []
    for r in records:
        matched_rule = engine.match_rule(r) if do_remove else None
        if matched_rule is not None:
            rid = matched_rule.get("id")
            if rid is not None:
                matched_rule_ids.append(int(rid))
            removed += 1
            continue
        kept.append(r)
        if learning and suggested < NOISE_SUGGEST_LIMIT and not engine.is_noise(r):
            sc = engine.score(r)
            if sc >= 80:
                engine.add_sample(r, "suggest", sc)
                suggested += 1
    increment_noise_hits(matched_rule_ids)
    if log_cb:
        if removed:
            log_cb(f"Noise Filter: excluded {removed} approved noise rows from merge output.")
        elif do_remove:
            log_cb("Noise Filter: enabled, but no rows matched approved enabled rules.")
        if suggested:
            log_cb(f"Noise Learning: saved {suggested} candidate rows for operator review. No candidate was removed automatically.")
    return kept

class NoiseRuleManagerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        init_noise_db()
        self.setWindowTitle("Noise Rules Manager")
        self.resize(980, 520)
        root = QVBoxLayout(self)
        note = QLabel("Approved rules are operator-controlled. Disable or delete a rule to restore matching logs.")
        root.addWidget(note)
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(["Status", "ID", "Enabled", "LogType", "Match", "Pattern", "Action", "Scope", "Hits", "Updated"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.currentCellChanged.connect(lambda *_: self.update_toggle_button_text())
        root.addWidget(self.table)
        buttons = QHBoxLayout()
        self.refresh_btn = QPushButton("Refresh")
        self.add_btn = QPushButton("Add Manual Rule")
        self.toggle_btn = QPushButton("Enable/Disable")
        self.delete_btn = QPushButton("Delete")
        self.close_btn = QPushButton("Close")
        for b in [self.refresh_btn, self.add_btn, self.toggle_btn, self.delete_btn]:
            buttons.addWidget(b)
        buttons.addStretch(1); buttons.addWidget(self.close_btn)
        root.addLayout(buttons)
        self.refresh_btn.clicked.connect(self.load_rules)
        self.add_btn.clicked.connect(self.add_manual_rule)
        self.toggle_btn.clicked.connect(self.toggle_selected)
        self.delete_btn.clicked.connect(self.delete_selected)
        self.close_btn.clicked.connect(self.accept)
        self.load_rules()

    def selected_rule_id(self) -> Optional[int]:
        row = self.table.currentRow()
        if row < 0:
            return None
        item = self.table.item(row, 1)
        if not item:
            return None
        try:
            return int(item.text())
        except Exception:
            return None

    def load_rules(self):
        con = sqlite3.connect(noise_db_path())
        con.row_factory = sqlite3.Row
        try:
            rows = con.execute("SELECT * FROM noise_rules ORDER BY id DESC LIMIT 1000").fetchall()
        finally:
            con.close()
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            enabled_text = "Yes" if row["enabled"] else "No"
            status_text = "Enabled" if row["enabled"] else "Disabled"
            values = [
                status_text, row["id"], enabled_text, row["log_type"], row["match_type"], row["pattern"],
                row["action"], row["scope"], row["hit_count"], row["updated_at"],
            ]
            for c, val in enumerate(values):
                self.table.setItem(r, c, QTableWidgetItem(str(val)))
        self.table.resizeColumnsToContents()
        self.update_toggle_button_text()

    def update_toggle_button_text(self):
        row = self.table.currentRow()
        if row < 0:
            self.toggle_btn.setText("Enable/Disable Rule")
            return
        enabled_item = self.table.item(row, 2)
        enabled = enabled_item and enabled_item.text().lower() == "yes"
        self.toggle_btn.setText("Disable Rule" if enabled else "Enable Rule")

    def add_manual_rule(self):
        pattern, ok = QInputDialog.getText(self, "Add Manual Noise Rule", "Text to match (contains):")
        if not ok or not pattern.strip():
            return
        log_type, ok = QInputDialog.getText(self, "Log Type", "Log type (ANY/WS/CGA/CSA/MRSERVER/GESYS/LAIS/PSC/REVIEW):", text="ANY")
        if not ok:
            return
        now = datetime.now().isoformat(timespec="seconds")
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute(
                "INSERT INTO noise_rules(pattern,match_type,log_type,action,scope,enabled,approved,confidence,created_at,updated_at,notes) VALUES(?,?,?,?,?,1,1,1.0,?,?,?)",
                (pattern.strip(), "contains", (log_type or "ANY").strip().upper(), "hide", "both", now, now, "Manual rule"),
            )
            con.commit()
        finally:
            con.close()
        self.load_rules()

    def toggle_selected(self):
        rid = self.selected_rule_id()
        if rid is None:
            return
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute("UPDATE noise_rules SET enabled = CASE enabled WHEN 1 THEN 0 ELSE 1 END, updated_at=? WHERE id=?", (datetime.now().isoformat(timespec="seconds"), rid))
            con.commit()
        finally:
            con.close()
        self.load_rules()

    def delete_selected(self):
        rid = self.selected_rule_id()
        if rid is None:
            return
        if QMessageBox.question(self, "Delete Rule", f"Delete rule ID {rid}?") != QMessageBox.Yes:
            return
        con = sqlite3.connect(noise_db_path())
        try:
            con.execute("DELETE FROM noise_rules WHERE id=?", (rid,))
            con.commit()
        finally:
            con.close()
        self.load_rules()

# ---------------------------------------------------------------------------
# v16 Dual Log Viewer
# ---------------------------------------------------------------------------
VIEWER_COLUMNS = ["Timestamp", "Message", "Level", "SourceType", "File", "Line", "Category"]


def format_viewer_timestamp(ts: Optional[datetime]) -> str:
    if not ts:
        return ""
    # Display as explicit one-column timestamp. Excel-like millisecond/decisecond
    # values are kept compact but zero-padded.
    frac = int(round(ts.microsecond / 100000.0))
    if frac >= 10:
        ts = ts + timedelta(seconds=1)
        frac = 0
    return ts.strftime("%Y/%m/%d %H:%M:%S") + f".{frac}"




def parse_viewer_datetime_text(text: str, *, is_end: bool = False) -> Optional[datetime]:
    """Parse viewer time filter text.

    Supported examples:
      2026/07/04 13:07:58.6
      2026-07-04 13:07:58
      2026/07/04

    Date-only values become start-of-day for Start and end-of-day for End.
    """
    t = (text or "").strip()
    if not t:
        return None
    t = t.replace("T", " ").replace("-", "/")
    # Date only
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", t)
    if m:
        y, mo, d = map(int, m.groups())
        if is_end:
            return datetime(y, mo, d, 23, 59, 59, 999999)
        return datetime(y, mo, d, 0, 0, 0, 0)
    # Date + time. Accept fractional seconds of any length; pad/truncate to microseconds.
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2})(?:[\.:](\d{1,6}))?)?", t)
    if not m:
        raise ValueError(f"Invalid date/time: {text}")
    y, mo, d, hh, mm, ss, frac = m.groups()
    us = 0
    if frac:
        us = int((frac + "000000")[:6])
    return datetime(int(y), int(mo), int(d), int(hh), int(mm), int(ss or 0), us)

def record_to_viewer_row(r: LogRecord) -> dict[str, Any]:
    row = {
        "Timestamp": format_viewer_timestamp(r.timestamp),
        "SourceType": r.source_type,
        "File": r.filename,
        "Line": r.line_no,
        "Level": r.level,
        "Category": r.category,
        "Message": viewer_message_body(r.message),
        "Raw": clean_excel_value(r.raw),
        "_ts": r.timestamp,
    }
    # CSA/CGA structured parser fields are stored in Raw JSON so they remain
    # compatible with the existing LogRecord model.
    if r.source_type in {"CSA", "CGA"} and r.raw:
        try:
            data = json.loads(r.raw)
            if isinstance(data, dict):
                for key, value in data.items():
                    if not str(key).startswith("_"):
                        row[str(key)] = clean_excel_value(value)
                row["Type"] = str(data.get("Type", r.level))
                row["Original"] = str(data.get("Original", ""))
                row["Sub Original"] = str(data.get("Sub Original", ""))
        except Exception:
            pass

    # WaterSystem rows preserve the native structured columns. They do not use
    # a synthesized Message string.
    if r.source_type == "WATERSYSTEM" and r.raw:
        try:
            data = json.loads(r.raw)
            if isinstance(data, dict):
                for key, value in data.items():
                    row[str(key)] = clean_excel_value(value)
                row["Message"] = ""
                row["Category"] = str(data.get("MainState", ""))
                row["Level"] = "ERROR" if str(data.get("Error", "")).strip() else ""
        except Exception:
            pass

    # Structured File Type plugins keep their extracted numeric fields as real
    # columns. Viewer defaults are read from viewer_defaults.json.
    if plugin_manifest_by_id(r.source_type) and r.raw:
        try:
            data = json.loads(r.raw)
            if isinstance(data, dict):
                for key, value in data.items():
                    if not str(key).startswith("_"):
                        row[str(key)] = clean_excel_value(value)
                row["Message"] = ""
                row["Level"] = ""
        except Exception:
            pass

    return row


def review_rows_to_viewer_records(path: Path) -> list[LogRecord]:
    records: list[LogRecord] = []
    for idx, row in enumerate(parse_reviewout_file(path), start=1):
        text_dt = " ".join(str(row.get(k, "")) for k in ("date", "time") if row.get(k, ""))
        ts = parse_datetime_from_text(text_dt) if text_dt.strip() else None
        msg_parts = []
        for k, v in row.items():
            if k in {"date", "time"}:
                continue
            if v not in (None, ""):
                msg_parts.append(f"{k}={v}")
            if len(msg_parts) >= 4:
                break
        msg = "; ".join(msg_parts) if msg_parts else "Review scan row"
        records.append(LogRecord(ts, "REVIEW", path.name, idx, "", "Review", msg, json.dumps(row, ensure_ascii=False)))
    return records


class LogTableModel(QAbstractTableModel):
    """Lazy table model.

    The complete parsed row list stays in memory, but the Qt view initially
    exposes only a small window. More rows are revealed through fetchMore as the
    user scrolls. This removes the long UI stall previously seen while creating
    a model for 100k+ rows.
    """
    INITIAL_ROWS = 3000
    FETCH_ROWS = 3000

    def __init__(self, rows: Optional[list[dict[str, Any]]] = None):
        super().__init__()
        self.rows = rows or []
        self.columns = list(VIEWER_COLUMNS)
        self._visible_count = min(len(self.rows), self.INITIAL_ROWS)

    def rowCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        return self._visible_count

    def canFetchMore(self, parent=QModelIndex()):
        return (not parent.isValid()) and self._visible_count < len(self.rows)

    def fetchMore(self, parent=QModelIndex()):
        if parent.isValid() or self._visible_count >= len(self.rows):
            return
        remaining = len(self.rows) - self._visible_count
        count = min(self.FETCH_ROWS, remaining)
        first = self._visible_count
        last = first + count - 1
        self.beginInsertRows(QModelIndex(), first, last)
        self._visible_count += count
        self.endInsertRows()

    def columnCount(self, parent=QModelIndex()):
        return len(self.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or index.row() >= self._visible_count:
            return None
        row = self.rows[index.row()]
        col = self.columns[index.column()]
        if role == Qt.DisplayRole:
            value = row.get(col, "")
            if isinstance(value, float):
                return f"{value:g}"
            return str(value)
        if role == Qt.ToolTipRole:
            source_type = str(row.get("SourceType", "")).upper()
            plg = plugin_manifest_by_id(source_type)
            if source_type in {"WATERSYSTEM", "VIMEASURE", "REVIEW", "PSC"} or (plg and plg.get("hover_popup") is False):
                return None
            return str(row.get("Raw", row.get("Message", "")))
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self.columns[section]
        return str(section + 1)

    def set_rows(self, rows: list[dict[str, Any]], columns: Optional[list[str]] = None):
        self.beginResetModel()
        self.rows = rows
        self._visible_count = min(len(rows), self.INITIAL_ROWS)
        if columns is not None:
            self.columns = columns
        else:
            source_types = {str(row.get("SourceType", "")).upper() for row in rows[:200]}
            if source_types and source_types <= {"WATERSYSTEM"}:
                self.columns = [c for c in ["Timestamp", "MainState", "Error"] if any(c in row for row in rows[:50])]
            else:
                extra: list[str] = []
                seen = set(VIEWER_COLUMNS) | {"Raw", "_ts"}
                for row in rows[:500]:
                    for key in row.keys():
                        if key not in seen and key not in extra:
                            extra.append(key)
                self.columns = list(VIEWER_COLUMNS) + extra
        self.endResetModel()

    def row_at(self, row: int) -> Optional[dict[str, Any]]:
        if 0 <= row < len(self.rows):
            return self.rows[row]
        return None


class DualLogViewer(QWidget):
    """Lightweight two-pane log viewer with timestamp cross reference.

    The viewer uses QTableView + QAbstractTableModel, so it does not create
    thousands of cell widgets. This keeps scrolling reasonably light even with
    large log lists. Data is loaded per side only when requested.
    """
    SOURCES = ["Merged", "WS", "WaterSystem", "VIMeasure", "ACQUISITION", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "Review", "Custom File"]

    def __init__(self, parent_window: 'MainWindow'):
        super().__init__()
        self.parent_window = parent_window
        self.setWindowTitle("Dual Log Viewer - Cross Reference")
        self.resize(1300, 780)
        self._syncing = False
        self.left_model = LogTableModel([])
        self.right_model = LogTableModel([])
        self.left_ts: list[tuple[datetime, int]] = []
        self.right_ts: list[tuple[datetime, int]] = []
        self.left_all_rows: list[dict[str, Any]] = []
        self.right_all_rows: list[dict[str, Any]] = []
        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)

        # Global viewer controls stay in the center/top, while left/right file
        # selectors are placed directly above each pane so the operation target
        # is immediately obvious to the operator.
        top = QHBoxLayout()
        self.tolerance_spin = QSpinBox(); self.tolerance_spin.setRange(0, 3600); self.tolerance_spin.setValue(10); self.tolerance_spin.setSuffix(" sec")
        self.load_both_btn = QPushButton("Load Both")
        self.export_pair_btn = QPushButton("Export Selected Pair")
        self.apply_noise_chk = QCheckBox("Apply Noise Filter")
        self.apply_noise_chk.setChecked(False)
        self.noise_apply_mode = QComboBox(); self.noise_apply_mode.addItems(["Apply immediately after adding rule", "Apply only when Apply button is pressed"])
        self.apply_rules_btn = QPushButton("Apply Rules Now")
        self.not_noise_btn = QPushButton("Mark Selected Not Noise")
        self.manage_noise_btn = QPushButton("Manage Rules")
        self.status = QLabel("Ready")
        top.addWidget(QLabel("Jump tolerance:")); top.addWidget(self.tolerance_spin)
        top.addWidget(self.load_both_btn)
        top.addWidget(self.export_pair_btn)
        top.addWidget(self.apply_noise_chk)
        top.addWidget(self.noise_apply_mode)
        top.addWidget(self.apply_rules_btn)
        top.addWidget(self.not_noise_btn)
        top.addWidget(self.manage_noise_btn)
        for i, cb in enumerate(self.pane_visible_checks):
            cb.setChecked(i < 2)
        top.addStretch(1)
        root.addLayout(top)

        # Viewer-only time range filter. This is independent from the Merge
        # date range and only changes what is displayed/searched in Explorer.
        time_bar = QHBoxLayout()
        self.viewer_time_enable = QCheckBox("Enable Viewer Time Range")
        self.viewer_start_edit = QLineEdit(); self.viewer_start_edit.setPlaceholderText("Start: 2026/07/04 13:07:58.6")
        self.viewer_end_edit = QLineEdit(); self.viewer_end_edit.setPlaceholderText("End: 2026/07/04 13:30:00.0")
        self.apply_time_left_btn = QPushButton("Apply to Left")
        self.apply_time_right_btn = QPushButton("Apply to Right")
        self.apply_time_both_btn = QPushButton("Apply to Both")
        self.clear_time_btn = QPushButton("Clear Time Range")
        self.set_start_btn = QPushButton("Set Start from Selected")
        self.set_end_btn = QPushButton("Set End from Selected")
        time_bar.addWidget(self.viewer_time_enable)
        time_bar.addWidget(self.viewer_start_edit, 1)
        time_bar.addWidget(self.viewer_end_edit, 1)
        time_bar.addWidget(self.apply_time_left_btn)
        time_bar.addWidget(self.apply_time_right_btn)
        time_bar.addWidget(self.apply_time_both_btn)
        time_bar.addWidget(self.clear_time_btn)
        time_bar.addWidget(self.set_start_btn)
        time_bar.addWidget(self.set_end_btn)
        root.addLayout(time_bar)

        self.left_source = QComboBox(); self.left_source.addItems(self.SOURCES); self.left_source.setCurrentText("Merged")
        self.right_source = QComboBox(); self.right_source.addItems(self.SOURCES); self.right_source.setCurrentText("PSC")
        self.load_left_btn = QPushButton("Load Left")
        self.load_right_btn = QPushButton("Load Right")
        self.left_search = QLineEdit(); self.left_search.setPlaceholderText("Search left message/raw...")
        self.right_search = QLineEdit(); self.right_search.setPlaceholderText("Search right message/raw...")
        self.find_left_btn = QPushButton("Find")
        self.find_right_btn = QPushButton("Find")
        self.mark_left_noise_btn = QPushButton("Add Left to Noise Rule")
        self.mark_right_noise_btn = QPushButton("Add Right to Noise Rule")
        self.copy_left_rule_btn = QPushButton("Copy Left Rule Text")
        self.copy_right_rule_btn = QPushButton("Copy Right Rule Text")
        self.left_file_label = QLabel("File: not loaded")
        self.right_file_label = QLabel("File: not loaded")
        self.left_file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.right_file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        # Tooltips for operator guidance.
        self.left_source.setToolTip("Select the log type shown in the LEFT pane. Use Custom File to load a single file without merge.")
        self.right_source.setToolTip("Select the log type shown in the RIGHT pane. PSC/Review can be compared without running merge.")
        self.load_left_btn.setToolTip("Load only the selected LEFT-side log. A progress popup is shown while loading.")
        self.load_right_btn.setToolTip("Load only the selected RIGHT-side log. A progress popup is shown while loading.")
        self.left_search.setToolTip("Search only the LEFT pane. It searches Message and Raw text.")
        self.right_search.setToolTip("Search only the RIGHT pane. It searches Message and Raw text.")
        self.find_left_btn.setToolTip("Find next matching row in the LEFT pane.")
        self.find_right_btn.setToolTip("Find next matching row in the RIGHT pane.")
        self.noise_apply_mode.setToolTip("Choose when newly approved Noise rules are applied to the Explorer view.")
        self.apply_rules_btn.setToolTip("Reload both panes and apply the current approved Noise rules.")
        self.load_both_btn.setToolTip("Load both left and right logs. Use this after selecting the two sources to compare.")
        self.tolerance_spin.setToolTip("Maximum time difference used when jumping to the nearest row on the other side.")
        self.export_pair_btn.setToolTip("Export the currently selected left/right row pair to CSV.")
        self.apply_noise_chk.setToolTip("Hide logs matching approved Smart Noise rules in this viewer only.")
        self.mark_left_noise_btn.setToolTip("Create a Noise rule from the selected LEFT row. The pattern is editable before approval.")
        self.mark_right_noise_btn.setToolTip("Create a Noise rule from the selected RIGHT row. The pattern is editable before approval.")
        self.copy_left_rule_btn.setToolTip("Copy a compact rule candidate from the selected LEFT row to the clipboard.")
        self.copy_right_rule_btn.setToolTip("Copy a compact rule candidate from the selected RIGHT row to the clipboard.")
        self.not_noise_btn.setToolTip("Mark the selected row as Not Noise so it can be reviewed later.")
        self.manage_noise_btn.setToolTip("Open Smart Noise rule manager. Rules can be enabled, disabled, or deleted later.")
        self.viewer_time_enable.setToolTip("Enable a Viewer-only start/end timestamp filter. It does not change Merge output files.")
        self.viewer_start_edit.setToolTip("Start timestamp for Explorer display/search. Example: 2026/07/04 13:07:58.6")
        self.viewer_end_edit.setToolTip("End timestamp for Explorer display/search. Date-only input is treated as end-of-day for End.")
        self.apply_time_left_btn.setToolTip("Apply the current time range only to the LEFT pane.")
        self.apply_time_right_btn.setToolTip("Apply the current time range only to the RIGHT pane.")
        self.apply_time_both_btn.setToolTip("Apply the current time range to both panes.")
        self.clear_time_btn.setToolTip("Clear the Viewer time range and restore all loaded rows.")
        self.set_start_btn.setToolTip("Copy the selected row timestamp into the Start field.")
        self.set_end_btn.setToolTip("Copy the selected row timestamp into the End field.")

        split = QSplitter(Qt.Horizontal)
        self.main_splitter = split
        split.setHandleWidth(10)
        split.setOpaqueResize(True)
        split.setChildrenCollapsible(False)
        split.setToolTip("Drag this center divider left or right to resize both log panes. Message columns expand or shrink automatically.")
        left_pane = QWidget(); left_layout = QVBoxLayout(left_pane); left_layout.setContentsMargins(4, 4, 4, 4)
        right_pane = QWidget(); right_layout = QVBoxLayout(right_pane); right_layout.setContentsMargins(4, 4, 4, 4)

        # Compact pane header: no separate title label.
        # The log selector row itself is colored so the operator can quickly
        # distinguish Log 1 / Log 2 while keeping vertical space for messages.
        left_selector_bar = QWidget()
        left_selector_bar.setStyleSheet("background:#EAF3FF; border:1px solid #C9DDF5; border-radius:4px;")
        left_controls = QHBoxLayout(left_selector_bar)
        left_controls.setContentsMargins(6, 4, 6, 4)
        left_controls.addWidget(QLabel("Log:")); left_controls.addWidget(self.left_source, 1)
        left_controls.addWidget(self.load_left_btn)
        left_layout.addWidget(left_selector_bar)
        left_search_layout = QHBoxLayout()
        left_search_layout.addWidget(QLabel("Search:")); left_search_layout.addWidget(self.left_search, 1); left_search_layout.addWidget(self.find_left_btn)
        left_layout.addLayout(left_search_layout)
        left_layout.addWidget(self.left_file_label)
        left_noise_controls = QHBoxLayout()
        left_noise_controls.addWidget(self.mark_left_noise_btn)
        left_noise_controls.addWidget(self.copy_left_rule_btn)
        left_noise_controls.addStretch(1)
        left_layout.addLayout(left_noise_controls)

        right_selector_bar = QWidget()
        right_selector_bar.setStyleSheet("background:#EAF8EA; border:1px solid #C7E5C7; border-radius:4px;")
        right_controls = QHBoxLayout(right_selector_bar)
        right_controls.setContentsMargins(6, 4, 6, 4)
        right_controls.addWidget(QLabel("Log:")); right_controls.addWidget(self.right_source, 1)
        right_controls.addWidget(self.load_right_btn)
        right_layout.addWidget(right_selector_bar)
        right_search_layout = QHBoxLayout()
        right_search_layout.addWidget(QLabel("Search:")); right_search_layout.addWidget(self.right_search, 1); right_search_layout.addWidget(self.find_right_btn)
        right_layout.addLayout(right_search_layout)
        right_layout.addWidget(self.right_file_label)
        right_noise_controls = QHBoxLayout()
        right_noise_controls.addWidget(self.mark_right_noise_btn)
        right_noise_controls.addWidget(self.copy_right_rule_btn)
        right_noise_controls.addStretch(1)
        right_layout.addLayout(right_noise_controls)

        self.left_table = QTableView(); self.left_table.setModel(self.left_model)
        self.right_table = QTableView(); self.right_table.setModel(self.right_model)
        for tbl in (self.left_table, self.right_table):
            tbl.setSelectionBehavior(QTableView.SelectRows)
            tbl.setSelectionMode(QTableView.SingleSelection)
            tbl.setAlternatingRowColors(True)
            tbl.setSortingEnabled(False)
            tbl.verticalHeader().setDefaultSectionSize(20)
            tbl.verticalHeader().setVisible(False)
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            # Default view: Timestamp is a fixed reference at the far left,
            # while Message stretches to use the available pane width. Dragging
            # the center splitter therefore makes the message window wider/narrower.
            tbl.setColumnWidth(0, 170)   # Timestamp
            tbl.setColumnWidth(1, 620)   # Message
            tbl.setColumnWidth(2, 90)    # SourceType
            tbl.setColumnWidth(3, 200)   # File
            tbl.setColumnWidth(4, 60)    # Line
            tbl.setColumnWidth(5, 70)    # Level
            tbl.setColumnWidth(6, 100)   # Category
            tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
            tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        left_layout.addWidget(self.left_table, 1)
        right_layout.addWidget(self.right_table, 1)
        split.addWidget(left_pane)
        split.addWidget(right_pane)
        split.setSizes([650, 650])
        root.addWidget(split, 1)

        self.detail = QTextEdit(); self.detail.setReadOnly(True); self.detail.setMaximumHeight(150)
        root.addWidget(self.detail)
        root.addWidget(self.status)

        self.load_left_btn.clicked.connect(lambda: self.load_side("left"))
        self.load_right_btn.clicked.connect(lambda: self.load_side("right"))
        self.load_both_btn.clicked.connect(self.load_both)
        self.export_pair_btn.clicked.connect(self.export_selected_pair)
        self.apply_noise_chk.toggled.connect(lambda _=False: self.apply_noise_rules_now())
        self.apply_rules_btn.clicked.connect(self.apply_noise_rules_now)
        self.left_search.returnPressed.connect(lambda: self.find_next("left"))
        self.right_search.returnPressed.connect(lambda: self.find_next("right"))
        self.find_left_btn.clicked.connect(lambda: self.find_next("left"))
        self.find_right_btn.clicked.connect(lambda: self.find_next("right"))
        self.mark_left_noise_btn.clicked.connect(lambda: self.approve_selected_noise("left"))
        self.mark_right_noise_btn.clicked.connect(lambda: self.approve_selected_noise("right"))
        self.copy_left_rule_btn.clicked.connect(lambda: self.copy_rule_text("left"))
        self.copy_right_rule_btn.clicked.connect(lambda: self.copy_rule_text("right"))
        self.not_noise_btn.clicked.connect(self.mark_selected_not_noise)
        self.manage_noise_btn.clicked.connect(self.manage_noise_rules)
        self.apply_time_left_btn.clicked.connect(lambda: self.apply_view_filters("left"))
        self.apply_time_right_btn.clicked.connect(lambda: self.apply_view_filters("right"))
        self.apply_time_both_btn.clicked.connect(lambda: (self.apply_view_filters("left"), self.apply_view_filters("right")))
        self.clear_time_btn.clicked.connect(self.clear_viewer_time_range)
        self.set_start_btn.clicked.connect(lambda: self.set_time_from_selected("start"))
        self.set_end_btn.clicked.connect(lambda: self.set_time_from_selected("end"))
        self.left_table.selectionModel().selectionChanged.connect(lambda *_: self.row_selected("left"))
        self.right_table.selectionModel().selectionChanged.connect(lambda *_: self.row_selected("right"))

    def log(self, text: str):
        self.status.setText(text)
        try:
            self.parent_window.log("[Viewer] " + text)
        except Exception:
            pass

    def make_progress(self, title: str, text: str) -> QProgressDialog:
        dlg = QProgressDialog(text, "Cancel", 0, 0, self)
        dlg.setWindowTitle(title)
        dlg.setWindowModality(Qt.ApplicationModal)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setValue(0)
        return dlg

    def load_both(self):
        self.load_side("left")
        self.load_side("right")

    def current_viewer_time_range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        if not self.viewer_time_enable.isChecked():
            return None, None
        start = parse_viewer_datetime_text(self.viewer_start_edit.text(), is_end=False) if self.viewer_start_edit.text().strip() else None
        end = parse_viewer_datetime_text(self.viewer_end_edit.text(), is_end=True) if self.viewer_end_edit.text().strip() else None
        if start and end and start > end:
            raise ValueError("Viewer Start time is later than End time.")
        return start, end

    def apply_view_filters(self, side: str):
        """Apply Viewer-only time range to already loaded rows.

        This keeps the original loaded rows in memory so the operator can clear
        the filter without reloading files from disk. Search operates on the
        currently displayed rows after this filter is applied.
        """
        try:
            start, end = self.current_viewer_time_range()
        except Exception as e:
            QMessageBox.warning(self, "Viewer Time Range", str(e))
            return
        base_rows = self.left_all_rows if side == "left" else self.right_all_rows
        if start or end:
            rows = []
            for row in base_rows:
                ts = row.get("_ts")
                if not isinstance(ts, datetime):
                    continue
                if start and ts < start:
                    continue
                if end and ts > end:
                    continue
                rows.append(row)
        else:
            rows = list(base_rows)
        ts_index = [(row["_ts"], i) for i, row in enumerate(rows) if isinstance(row.get("_ts"), datetime)]
        ts_index.sort(key=lambda x: x[0])
        if side == "left":
            self.left_model.set_rows(rows)
            self.left_ts = ts_index
        else:
            self.right_model.set_rows(rows)
            self.right_ts = ts_index
        if start or end:
            self.log(f"Viewer time filter {side}: {len(rows)}/{len(base_rows)} rows")
        else:
            self.log(f"Viewer time filter cleared {side}: {len(rows)} rows")

    def clear_viewer_time_range(self):
        self.viewer_time_enable.setChecked(False)
        self.viewer_start_edit.clear()
        self.viewer_end_edit.clear()
        self.apply_view_filters("left")
        self.apply_view_filters("right")

    def set_time_from_selected(self, target: str):
        # Prefer the row with current focus; fall back to left, then right.
        for side, table, model in (("left", self.left_table, self.left_model), ("right", self.right_table, self.right_model)):
            if table.hasFocus() or table.selectionModel().selectedRows():
                indexes = table.selectionModel().selectedRows()
                if indexes:
                    row = model.row_at(indexes[0].row()) or {}
                    ts_text = row.get("Timestamp", "")
                    if ts_text:
                        if target == "start":
                            self.viewer_start_edit.setText(str(ts_text))
                        else:
                            self.viewer_end_edit.setText(str(ts_text))
                        self.viewer_time_enable.setChecked(True)
                        self.log(f"Set Viewer {target} from selected {side}: {ts_text}")
                        return
        QMessageBox.information(self, "Viewer Time Range", "Select a row first.")

    def source_to_records(self, source_name: str, progress: Optional[QProgressDialog] = None) -> list[LogRecord]:
        source_folder = Path(self.parent_window.source_edit.text().strip() or ".")
        recursive = self.parent_window.recursive_chk.isChecked()
        opt = self.parent_window.collect_options()
        if source_name == "Custom File":
            path_str, _ = QFileDialog.getOpenFileName(self, "Select log file", str(source_folder), "Log files (*.log *.txt *.out *.*);;All Files (*.*)")
            if not path_str:
                return []
            path = Path(path_str)
            if progress:
                progress.setLabelText(f"Loading custom file:\n{path.name}")
                QApplication.processEvents()
            if is_review_file(path.name):
                return review_rows_to_viewer_records(path)
            typ = classify_file(path, True) or "UNKNOWN"
            if typ == "PSC":
                _, _, recs = parse_psc_file_detail(path)
                return recs
            return parse_file(path, typ)

        if not source_folder.exists():
            QMessageBox.warning(self, "Dual Log Viewer", "Source Log Folder does not exist.")
            return []

        targets: set[str]
        if source_name == "Merged":
            targets = enabled_types(opt)
        elif source_name == "Review":
            targets = {"REVIEW"}
        else:
            targets = {source_name}

        # Build a candidate file list first so the progress popup can show a
        # useful percentage. This also avoids the previous impression that the
        # viewer had frozen while it was scanning a large folder.
        candidates: list[tuple[Path, str]] = []
        if progress:
            progress.setLabelText(f"Scanning files for {source_name} ...")
            QApplication.processEvents()
        for path in iter_files(source_folder, recursive):
            if progress and progress.wasCanceled():
                return []
            if source_name == "Review":
                if is_review_file(path.name):
                    candidates.append((path, "REVIEW"))
                continue
            typ = classify_file(path, opt.include_unknown)
            if typ and typ in targets:
                candidates.append((path, typ))

        if progress:
            progress.setRange(0, max(1, len(candidates)))
            progress.setValue(0)
            progress.setLabelText(f"Loading {source_name}: 0/{len(candidates)} files")
            QApplication.processEvents()

        records: list[LogRecord] = []
        for idx, (path, typ) in enumerate(candidates, start=1):
            if progress and progress.wasCanceled():
                self.log(f"Loading {source_name} canceled by operator.")
                break
            if progress:
                progress.setValue(idx - 1)
                progress.setLabelText(f"Loading {source_name}: {idx}/{len(candidates)}\n{path.name}")
                QApplication.processEvents()
            try:
                if typ == "REVIEW":
                    records.extend(review_rows_to_viewer_records(path))
                elif typ == "PSC":
                    _, _, recs = parse_psc_file_detail(path)
                    records.extend(recs)
                else:
                    records.extend(parse_file(path, typ))
            except Exception as e:
                records.append(LogRecord(None, typ, path.name, 0, "ERROR", "Viewer", f"Parse failed: {e}", ""))
        if progress:
            progress.setValue(len(candidates))
            progress.setLabelText(f"Filtering and sorting {len(records)} rows ...")
            QApplication.processEvents()

        # Apply the same date filter as merge, based on each row timestamp.
        # Important: this never uses file creation/modified time.
        start_date = parse_user_date(opt.start_date) if opt.use_start and opt.start_date else None
        end_date = parse_user_date(opt.end_date) if opt.use_end and opt.end_date else None
        if start_date or end_date:
            before = len(records)
            records = [r for r in records if in_date_range(r.timestamp, None, start_date, end_date)]
            self.log(f"Date filter kept {len(records)}/{before} rows for {source_name}.")
        if self.apply_noise_chk.isChecked():
            engine = NoiseEngine()
            before = len(records)
            filtered_records: list[LogRecord] = []
            matched_rule_ids: list[int] = []
            for r in records:
                matched_rule = engine.match_rule(r)
                if matched_rule is not None:
                    rid = matched_rule.get("id")
                    if rid is not None:
                        matched_rule_ids.append(int(rid))
                    continue
                filtered_records.append(r)
            records = filtered_records
            increment_noise_hits(matched_rule_ids)
            hidden = before - len(records)
            if hidden:
                self.log(f"Noise Filter hidden rows: {hidden}")
            else:
                self.log("Noise Filter: no rows matched approved enabled rules in this pane.")
        records.sort(key=lambda r: (r.timestamp is None, r.timestamp or datetime.max, r.source_type, r.filename, r.line_no))
        self.log(f"Loaded {len(records)} rows from {len(candidates)} files for {source_name}.")
        return records

    def load_side(self, side: str):
        source = self.left_source.currentText() if side == "left" else self.right_source.currentText()
        self.log(f"Loading {side}: {source} ...")
        progress = self.make_progress("Dual Log Viewer", f"Preparing {side}: {source} ...")
        progress.show()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            records = self.source_to_records(source, progress)
            if progress.wasCanceled():
                progress.close()
                self.status_label.setText("ZIP preload cancelled")
                return
            progress.setLabelText(f"Building table model for {side}: {len(records)} rows ...")
            QApplication.processEvents()
            rows = [record_to_viewer_row(r) for r in records]
            ts_index = [(row["_ts"], i) for i, row in enumerate(rows) if isinstance(row.get("_ts"), datetime)]
            ts_index.sort(key=lambda x: x[0])
            file_hint = source
            if rows:
                files = sorted({str(r.get("File", "")) for r in rows if r.get("File", "")})
                if len(files) == 1:
                    file_hint = files[0]
                elif len(files) > 1:
                    file_hint = f"{len(files)} files loaded"
            if side == "left":
                self.left_all_rows = rows
                self.left_file_label.setText(f"File: {file_hint}")
            else:
                self.right_all_rows = rows
                self.right_file_label.setText(f"File: {file_hint}")
            self.apply_view_filters(side)
            self.log(f"{side.capitalize()} loaded: {len(rows)} rows")
        except Exception:
            write_startup_log("Dual Log Viewer load_side failed.\n\n" + traceback.format_exc())
            QMessageBox.critical(self, "Dual Log Viewer", "Viewer loading failed.\n\n" + traceback.format_exc() + f"\n\nLog:\n{startup_log_path()}")
        finally:
            QApplication.restoreOverrideCursor()
            progress.close()

    def apply_noise_rules_now(self):
        """Apply approved noise rules at the operator-selected timing."""
        if self.apply_noise_chk.isChecked():
            self.log("Applying approved Noise rules to both panes...")
        else:
            self.log("Noise filter disabled. Reloading both panes...")
        self.load_both()

    def find_next(self, side: str):
        table = self.left_table if side == "left" else self.right_table
        model = self.left_model if side == "left" else self.right_model
        edit = self.left_search if side == "left" else self.right_search
        term = edit.text().strip().lower()
        if not term:
            return
        current = -1
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if indexes:
            current = indexes[0].row()
        total = model.rowCount()
        if total <= 0:
            return
        for step in range(1, total + 1):
            i = (current + step) % total
            row = model.row_at(i) or {}
            hay = " ".join(str(row.get(k, "")) for k in ("Timestamp", "SourceType", "File", "Level", "Category", "Message", "Raw")).lower()
            if term in hay:
                table.selectRow(i)
                table.scrollTo(model.index(i, 0), QTableView.PositionAtCenter)
                self.show_detail(side, row)
                self.log(f"Found in {side}: row {i + 1}/{total}")
                return
        self.log(f"No match in {side}: {term}")

    def row_selected(self, side: str):
        if self._syncing:
            return
        table = self.left_table if side == "left" else self.right_table
        model = self.left_model if side == "left" else self.right_model
        indexes = table.selectionModel().selectedRows()
        if not indexes:
            return
        row_idx = indexes[0].row()
        row = model.row_at(row_idx)
        if not row:
            return
        self.show_detail(side, row)
        ts = row.get("_ts")
        if not isinstance(ts, datetime):
            return
        self.jump_other_side(side, ts)

    def show_detail(self, side: str, row: dict[str, Any]):
        raw = row.get("Raw", "")
        msg = row.get("Message", "")
        self.detail.setPlainText(
            f"Selected {side}\n"
            f"Timestamp: {row.get('Timestamp','')}\n"
            f"SourceType: {row.get('SourceType','')}\n"
            f"File: {row.get('File','')}  Line: {row.get('Line','')}\n"
            f"Message: {msg}\n\nRaw:\n{raw}"
        )

    def jump_other_side(self, side: str, ts: datetime):
        other_ts = self.right_ts if side == "left" else self.left_ts
        other_table = self.right_table if side == "left" else self.left_table
        other_model = self.right_model if side == "left" else self.left_model
        if not other_ts:
            return
        # Binary search without importing bisect key dependency for older environments.
        lo, hi = 0, len(other_ts)
        while lo < hi:
            mid = (lo + hi) // 2
            if other_ts[mid][0] < ts:
                lo = mid + 1
            else:
                hi = mid
        candidates = []
        for pos in (lo - 1, lo, lo + 1):
            if 0 <= pos < len(other_ts):
                candidates.append(other_ts[pos])
        if not candidates:
            return
        nearest_ts, nearest_row = min(candidates, key=lambda x: abs((x[0] - ts).total_seconds()))
        delta = abs((nearest_ts - ts).total_seconds())
        tol = self.tolerance_spin.value()
        self._syncing = True
        try:
            other_table.selectRow(nearest_row)
            other_table.scrollTo(other_model.index(nearest_row, 0), QTableView.PositionAtCenter)
        finally:
            self._syncing = False
        msg = f"Cross reference: nearest row delta {delta:.3f} sec"
        if tol and delta > tol:
            msg += f" (outside ±{tol} sec)"
        self.log(msg)

    def selected_record_from_side(self, side: str) -> Optional[LogRecord]:
        table = self.left_table if side == "left" else self.right_table
        model = self.left_model if side == "left" else self.right_model
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if not indexes:
            return None
        row = model.row_at(indexes[0].row())
        if not row:
            return None
        return LogRecord(row.get("_ts"), str(row.get("SourceType", "")), str(row.get("File", "")), int(row.get("Line", 0) or 0), str(row.get("Level", "")), str(row.get("Category", "")), str(row.get("Message", "")), str(row.get("Raw", "")))

    def copy_rule_text(self, side: str):
        rec = self.selected_record_from_side(side)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        text = _compact_pattern(rec.message or rec.raw)
        QApplication.clipboard().setText(text)
        self.log(f"Copied {side} rule text to clipboard: {text}")

    def approve_selected_noise(self, side: str):
        rec = self.selected_record_from_side(side)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        engine = NoiseEngine()
        suggested = _compact_pattern(rec.message or rec.raw)
        pattern, ok = QInputDialog.getText(self, "Approve Noise Rule", "Rule text to hide (contains):", text=suggested)
        if not ok or not pattern.strip():
            return
        # Store exactly what the operator approved. Matching is normalized,
        # so small whitespace differences will not prevent filtering.
        rid = engine.add_rule(pattern.strip(), (rec.source_type or "ANY").upper(), "contains", "hide", "both", "Approved by operator from Explorer selected row")
        engine.add_sample(rec, "noise-approved", 100)
        if self.noise_apply_mode.currentIndex() == 0:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nApplying filter now.")
            self.apply_noise_rules_now()
        else:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nThe rule is saved. Press Apply Rules Now when you want to apply it to the current Explorer view.")

    def mark_selected_not_noise(self):
        rec = self.selected_record_from_side("left") or self.selected_record_from_side("right")
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        NoiseEngine().mark_not_noise(rec)
        QMessageBox.information(self, "Noise Learning", "Saved as Not Noise training sample. Approved rules were not changed.")

    def manage_noise_rules(self):
        dlg = NoiseRuleManagerDialog(self)
        dlg.exec()
        if self.noise_apply_mode.currentIndex() == 0:
            self.apply_noise_rules_now()
        else:
            self.log("Noise rules changed. Press Apply Rules Now to apply them to the current Explorer view.")

    def export_selected_pair(self):
        left_rows = self.left_table.selectionModel().selectedRows()
        right_rows = self.right_table.selectionModel().selectedRows()
        if not left_rows and not right_rows:
            QMessageBox.information(self, "Dual Log Viewer", "Select at least one row.")
            return
        out_dir = Path(self.parent_window.output_edit.text().strip() or self.parent_window.source_edit.text().strip() or str(Path.home()))
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"DualLogViewer_SelectedPair_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Side"] + VIEWER_COLUMNS + ["Raw"])
            if left_rows:
                row = self.left_model.row_at(left_rows[0].row())
                if row:
                    writer.writerow(["Left"] + [row.get(c, "") for c in VIEWER_COLUMNS] + [row.get("Raw", "")])
            if right_rows:
                row = self.right_model.row_at(right_rows[0].row())
                if row:
                    writer.writerow(["Right"] + [row.get(c, "") for c in VIEWER_COLUMNS] + [row.get("Raw", "")])
        QMessageBox.information(self, "Dual Log Viewer", f"Exported.\n\n{out_path}")



# v34 Multi Log Viewer override ------------------------------------------------
# Keep the legacy DualLogViewer above as a fallback implementation, then replace
# it with a multi-pane Explorer. MainWindow still calls DualLogViewer(self), so
# this alias preserves compatibility while adding 1/2/3/4 pane modes.
LegacyDualLogViewer = DualLogViewer

class MultiPaneLogViewer(LegacyDualLogViewer):
    SOURCES = ["Merged", "WS", "WaterSystem", "VIMeasure", "ACQUISITION", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "Review", "Custom File"]
    MAX_PANES = 4

    def __init__(self, parent_window: 'MainWindow'):
        QWidget.__init__(self)
        self.parent_window = parent_window
        self.setWindowTitle("Log Viewer / Log Explorer - 1 to 4 Logs")
        self.resize(max(900, QApplication.primaryScreen().availableGeometry().width()//3), max(520, QApplication.primaryScreen().availableGeometry().height()//3))
        self._syncing = False
        self.models: list[LogTableModel] = [LogTableModel([]) for _ in range(self.MAX_PANES)]
        self.tables: list[QTableView] = []
        self.sources: list[QComboBox] = []
        self.search_edits: list[QLineEdit] = []
        self.file_labels: list[QLabel] = []
        self.panes: list[QWidget] = []
        self.pane_visible_checks: list[QCheckBox] = []
        self.pane_column_buttons: list[QPushButton] = []
        self.pane_column_prefs: list[list[str]] = [[] for _ in range(self.MAX_PANES)]
        self.all_rows: list[list[dict[str, Any]]] = [[] for _ in range(self.MAX_PANES)]
        self.ts_indexes: list[list[tuple[datetime, int]]] = [[] for _ in range(self.MAX_PANES)]
        self.build_ui()
        self.update_view_mode()

    def side_index(self, side: Any) -> int:
        if isinstance(side, int):
            return max(0, min(self.MAX_PANES - 1, side))
        mapping = {"left": 0, "right": 1, "pane1": 0, "pane2": 1, "pane3": 2, "pane4": 3}
        return mapping.get(str(side).lower(), 0)

    def pane_name(self, idx: int) -> str:
        names = ["Left", "Right", "Log 3", "Log 4"]
        return names[idx] if 0 <= idx < len(names) else f"Log {idx + 1}"

    def visible_pane_count(self) -> int:
        text = self.mode_combo.currentText() if hasattr(self, "mode_combo") else "2 logs (Dual)"
        m = re.search(r"(\d+)", text)
        return max(1, min(self.MAX_PANES, int(m.group(1)) if m else 2))

    def build_ui(self):
        root = QVBoxLayout(self)

        top = QHBoxLayout()
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["1 log", "2 logs (Dual)", "3 logs", "4 logs"])
        self.mode_combo.setCurrentText("2 logs (Dual)")
        self.mode_combo.setToolTip("Switch Log Viewer layout between one, two, three, or four panes. Default is Dual mode.")
        self.tolerance_spin = QSpinBox(); self.tolerance_spin.setRange(0, 3600); self.tolerance_spin.setValue(10); self.tolerance_spin.setSuffix(" sec")
        self.load_visible_btn = QPushButton("LOAD LOGS")
        self.export_pair_btn = QPushButton("Export Selected Rows")
        self.apply_noise_chk = QCheckBox("Apply Noise Filter")
        self.apply_noise_chk.setChecked(False)
        self.noise_apply_mode = QComboBox(); self.noise_apply_mode.addItems(["Apply immediately after adding rule", "Apply only when Apply button is pressed"])
        self.apply_rules_btn = QPushButton("Apply Rules Now")
        self.not_noise_btn = QPushButton("Mark Selected Not Noise")
        self.manage_noise_btn = QPushButton("Manage Rules")
        self.status = QLabel("Ready")
        top.addWidget(QLabel("View mode:")); top.addWidget(self.mode_combo)
        top.addSpacing(8)
        top.addWidget(QLabel("Show:"))
        for i in range(self.MAX_PANES):
            cb = QCheckBox(str(i + 1))
            cb.setToolTip(f"Show or hide Log Pane {i + 1}. Visible panes are arranged from left to right.")
            cb.toggled.connect(self.update_view_mode)
            self.pane_visible_checks.append(cb)
            top.addWidget(cb)
        top.addSpacing(12)
        top.addWidget(QLabel("Jump tolerance:")); top.addWidget(self.tolerance_spin)
        top.addWidget(self.load_visible_btn)
        top.addWidget(self.export_pair_btn)
        top.addWidget(self.apply_noise_chk)
        top.addWidget(self.noise_apply_mode)
        top.addWidget(self.apply_rules_btn)
        top.addWidget(self.not_noise_btn)
        top.addWidget(self.manage_noise_btn)
        for i, cb in enumerate(self.pane_visible_checks):
            cb.setChecked(i < 2)
        top.addStretch(1)
        root.addLayout(top)

        time_bar = QHBoxLayout()
        self.viewer_time_enable = QCheckBox("Enable Viewer Time Range")
        self.viewer_start_edit = QLineEdit(); self.viewer_start_edit.setPlaceholderText("Start: 2026/07/04 13:07:58.6")
        self.viewer_end_edit = QLineEdit(); self.viewer_end_edit.setPlaceholderText("End: 2026/07/04 13:30:00.0")
        self.apply_time_current_btn = QPushButton("Apply to Focused")
        self.apply_time_all_btn = QPushButton("Apply to Visible")
        self.clear_time_btn = QPushButton("Clear Time Range")
        self.set_start_btn = QPushButton("Set Start from Selected")
        self.set_end_btn = QPushButton("Set End from Selected")
        time_bar.addWidget(self.viewer_time_enable)
        time_bar.addWidget(self.viewer_start_edit, 1)
        time_bar.addWidget(self.viewer_end_edit, 1)
        time_bar.addWidget(self.apply_time_current_btn)
        time_bar.addWidget(self.apply_time_all_btn)
        time_bar.addWidget(self.clear_time_btn)
        time_bar.addWidget(self.set_start_btn)
        time_bar.addWidget(self.set_end_btn)
        root.addLayout(time_bar)

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setHandleWidth(10)
        self.main_splitter.setOpaqueResize(True)
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setToolTip("Drag splitters to resize each log pane. Timestamp stays fixed and Message expands/shrinks.")

        pane_colors = ["#EAF3FF", "#EAF8EA", "#FFF4DF", "#F2EAFF"]
        default_sources = ["Merged", "PSC", "Review", "MRSERVER"]
        for i in range(self.MAX_PANES):
            pane = QWidget()
            layout = QVBoxLayout(pane); layout.setContentsMargins(4, 4, 4, 4)
            title = QLabel("")
            title.setVisible(False)

            source = QComboBox(); source.addItems(self.SOURCES); source.setCurrentText(default_sources[i])
            source.setToolTip(f"Select the log type or Custom File shown in {self.pane_name(i)} pane.")
            load_btn = QPushButton("Load This")
            load_btn.setToolTip(f"Load only {self.pane_name(i)} pane. Use Load Visible Logs to load all visible panes at once.")
            ctrl = QHBoxLayout()
            ctrl_widget = QWidget(); ctrl_widget.setStyleSheet(f"background:{pane_colors[i]}; border-radius:6px; padding:3px;")
            ctrl_inner = QHBoxLayout(ctrl_widget); ctrl_inner.setContentsMargins(6,3,6,3)
            ctrl_inner.addWidget(source, 1); ctrl_inner.addWidget(load_btn)
            ctrl.addWidget(ctrl_widget)
            layout.addLayout(ctrl)

            search = QLineEdit(); search.setPlaceholderText(f"Search {self.pane_name(i)} message/raw...")
            search.setToolTip(f"Search only {self.pane_name(i)} pane. It searches Timestamp, Message, Raw, File, Level and Category.")
            find_btn = QPushButton("Find")
            search_layout = QHBoxLayout()
            search_layout.addWidget(QLabel("Search:")); search_layout.addWidget(search, 1); search_layout.addWidget(find_btn)
            layout.addLayout(search_layout)

            file_label = QLabel("File: not loaded")
            file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            layout.addWidget(file_label)

            noise_bar = QHBoxLayout()
            add_noise = QPushButton("Add to Noise Rule")
            copy_rule = QPushButton("Copy Rule Text")
            column_btn = QPushButton("Columns...")
            add_noise.setToolTip(f"Create a Noise rule from the selected {self.pane_name(i)} row.")
            copy_rule.setToolTip(f"Copy a compact rule candidate from the selected {self.pane_name(i)} row.")
            column_btn.setToolTip("Choose which columns are visible in this pane. Default keeps Timestamp, Message and Level visible so the message stays readable.")
            noise_bar.addWidget(add_noise); noise_bar.addWidget(copy_rule); noise_bar.addWidget(column_btn); noise_bar.addStretch(1)
            layout.addLayout(noise_bar)

            table = QTableView(); table.setModel(self.models[i])
            table.setSelectionBehavior(QTableView.SelectRows)
            table.setSelectionMode(QTableView.SingleSelection)
            table.setAlternatingRowColors(True)
            table.setSortingEnabled(False)
            table.verticalHeader().setDefaultSectionSize(20)
            table.verticalHeader().setVisible(False)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            table.setColumnWidth(0, 175)
            table.setColumnWidth(1, 720)
            table.setColumnWidth(2, 70)
            table.setColumnWidth(3, 90)
            table.setColumnWidth(4, 180)
            table.setColumnWidth(5, 60)
            table.setColumnWidth(6, 100)
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            layout.addWidget(table, 1)

            self.main_splitter.addWidget(pane)
            self.panes.append(pane)
            self.sources.append(source)
            self.search_edits.append(search)
            self.file_labels.append(file_label)
            self.tables.append(table)
            self.pane_column_buttons.append(column_btn)

            load_btn.clicked.connect(lambda _=False, idx=i: self.load_pane(idx))
            find_btn.clicked.connect(lambda _=False, idx=i: self.find_next(idx))
            search.returnPressed.connect(lambda idx=i: self.find_next(idx))
            add_noise.clicked.connect(lambda _=False, idx=i: self.approve_selected_noise(idx))
            copy_rule.clicked.connect(lambda _=False, idx=i: self.copy_rule_text(idx))
            column_btn.clicked.connect(lambda _=False, idx=i: self.choose_columns(idx))
            table.clicked.connect(lambda _index, idx=i: self.row_selected(idx))

        root.addWidget(self.main_splitter, 1)
        self.detail = QTextEdit(); self.detail.setReadOnly(True); self.detail.setMaximumHeight(150)
        root.addWidget(self.detail)
        root.addWidget(self.status)

        self.mode_combo.currentIndexChanged.connect(self.update_view_mode)
        self.load_visible_btn.clicked.connect(self.load_visible)
        self.export_pair_btn.clicked.connect(self.export_selected_pair)
        self.apply_noise_chk.toggled.connect(lambda _=False: self.apply_noise_rules_now())
        self.apply_rules_btn.clicked.connect(self.apply_noise_rules_now)
        self.not_noise_btn.clicked.connect(self.mark_selected_not_noise)
        self.manage_noise_btn.clicked.connect(self.manage_noise_rules)
        self.apply_time_current_btn.clicked.connect(self.apply_time_to_focused)
        self.apply_time_all_btn.clicked.connect(self.apply_time_to_visible)
        self.clear_time_btn.clicked.connect(self.clear_viewer_time_range)
        self.set_start_btn.clicked.connect(lambda: self.set_time_from_selected("start"))
        self.set_end_btn.clicked.connect(lambda: self.set_time_from_selected("end"))

        # Backward-compatible aliases used by inherited helper code and older tests.
        self.left_model = self.models[0]; self.right_model = self.models[1]
        self.left_table = self.tables[0]; self.right_table = self.tables[1]
        self.left_source = self.sources[0]; self.right_source = self.sources[1]
        self.left_search = self.search_edits[0]; self.right_search = self.search_edits[1]
        self.left_file_label = self.file_labels[0]; self.right_file_label = self.file_labels[1]
        self.left_all_rows = self.all_rows[0]; self.right_all_rows = self.all_rows[1]
        self.left_ts = self.ts_indexes[0]; self.right_ts = self.ts_indexes[1]

    def log(self, text: str):
        self.status.setText(text)
        try:
            self.parent_window.log("[Viewer] " + text)
        except Exception:
            pass

    def visible_indices(self) -> list[int]:
        if getattr(self, "pane_visible_checks", None):
            checked = [i for i, cb in enumerate(self.pane_visible_checks) if cb.isChecked()]
            if checked:
                return checked
        return list(range(self.visible_pane_count()))

    def update_view_mode(self):
        count = self.visible_pane_count()
        # When the mode combo changes, default to the first N panes. The user can
        # then override with the small Show checkboxes.
        sender = self.sender()
        if sender is self.mode_combo and getattr(self, "pane_visible_checks", None):
            for i, cb in enumerate(self.pane_visible_checks):
                cb.blockSignals(True)
                cb.setChecked(i < count)
                cb.blockSignals(False)
        visible = self.visible_indices()
        if not visible:
            visible = list(range(count))
        for i, pane in enumerate(self.panes):
            pane.setVisible(i in visible)
        size_each = max(1, 1000 // max(1, len(visible)))
        self.main_splitter.setSizes([size_each if i in visible else 0 for i in range(self.MAX_PANES)])
        self.log(f"Viewer mode changed: {len(visible)} visible pane(s): " + ", ".join(self.pane_name(i) for i in visible))

    def load_visible(self):
        for idx in self.visible_indices():
            self.load_pane(idx)

    def load_both(self):
        self.load_visible()

    def current_viewer_time_range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        if not self.viewer_time_enable.isChecked():
            return None, None
        start = parse_viewer_datetime_text(self.viewer_start_edit.text(), is_end=False) if self.viewer_start_edit.text().strip() else None
        end = parse_viewer_datetime_text(self.viewer_end_edit.text(), is_end=True) if self.viewer_end_edit.text().strip() else None
        if start and end and start > end:
            raise ValueError("Viewer Start time is later than End time.")
        return start, end

    def default_viewer_columns_for_rows(self, rows: list[dict[str, Any]]) -> list[str]:
        raw_types = {str(r.get("SourceType", "")) for r in rows[:200] if r.get("SourceType")}
        if len(raw_types) == 1:
            only_type = next(iter(raw_types))
            defaults = load_plugin_json(only_type, "viewer_defaults.json", {}) or {}
            configured = defaults.get("default_visible_columns") or defaults.get("visible_columns")
            if isinstance(configured, list) and configured:
                return [str(c) for c in configured if any(str(c) in r for r in rows[:200])]
        sample_types = {x.upper() for x in raw_types}
        if sample_types and sample_types <= {"WATERSYSTEM"}:
            base = ["Timestamp", "MainState", "Error"]
        elif sample_types and sample_types <= {"WS"}:
            base = ["Timestamp", "Message"]
        elif sample_types and sample_types <= {"CSA"}:
            base = ["Timestamp", "Type", "Original", "Message"]
        elif sample_types and sample_types <= {"CGA"}:
            base = ["Timestamp", "Type", "Original", "Message"]
        elif sample_types and sample_types <= {"PSC", "REVIEW"}:
            base = ["Timestamp", "Parameter", "Value", "Message"]
        else:
            base = ["Timestamp", "Message", "Level"]
        return [c for c in base if any(c in r for r in rows[:200])] or base

    def all_available_columns_for_rows(self, rows: list[dict[str, Any]]) -> list[str]:
        preferred = ["Timestamp", "Type", "Original", "Message", "Num", "Sub Original", "Level", "SourceType", "Category", "File", "Line"]
        found = []
        for c in preferred:
            if any(c in r for r in rows[:200]):
                found.append(c)
        for row in rows[:1000]:
            for k in row.keys():
                if k not in found and k not in {"Raw", "_ts"}:
                    found.append(k)
        return found or preferred

    def pane_columns_for_rows(self, idx: int, rows: list[dict[str, Any]]) -> list[str]:
        prefs = self.pane_column_prefs[idx] if idx < len(self.pane_column_prefs) else []
        if prefs:
            available = set(self.all_available_columns_for_rows(rows))
            return [c for c in prefs if c in available] or self.default_viewer_columns_for_rows(rows)
        return self.default_viewer_columns_for_rows(rows)

    def apply_table_column_widths(self, idx: int):
        table = self.tables[idx]
        model = self.models[idx]
        try:
            for col_idx, col_name in enumerate(model.columns):
                if col_name == "Timestamp":
                    table.setColumnWidth(col_idx, 175)
                    table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Fixed)
                elif col_name == "Message":
                    table.setColumnWidth(col_idx, 780)
                    table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Stretch)
                elif col_name in {"Level", "Type"}:
                    table.setColumnWidth(col_idx, 65)
                    table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.ResizeToContents)
                elif col_name == "Original":
                    table.setColumnWidth(col_idx, 190)
                    table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Interactive)
                else:
                    table.setColumnWidth(col_idx, 140)
                    table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Interactive)
        except Exception:
            pass

    def choose_columns(self, idx: int):
        rows = self.all_rows[idx]
        if not rows:
            QMessageBox.information(self, "Columns", "Load a log first.")
            return
        available = self.all_available_columns_for_rows(rows)
        current = set(self.models[idx].columns or self.default_viewer_columns_for_rows(rows))
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Columns - {self.pane_name(idx)}")
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("Choose columns to show. Timestamp, Message and Level are recommended for message logs."))
        table = QTableWidget(len(available), 1)
        table.setHorizontalHeaderLabels(["Column"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for r, name in enumerate(available):
            item = QTableWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if name in current else Qt.Unchecked)
            table.setItem(r, 0, item)
        lay.addWidget(table)
        small = QHBoxLayout()
        btn_default = QPushButton("Default")
        btn_all = QPushButton("All")
        btn_none = QPushButton("None")
        small.addWidget(btn_default); small.addWidget(btn_all); small.addWidget(btn_none); small.addStretch(1)
        lay.addLayout(small)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(buttons)
        def set_state(names):
            for rr in range(table.rowCount()):
                item = table.item(rr, 0)
                item.setCheckState(Qt.Checked if item.text() in names else Qt.Unchecked)
        btn_default.clicked.connect(lambda: set_state(self.default_viewer_columns_for_rows(rows)))
        btn_all.clicked.connect(lambda: set_state(available))
        btn_none.clicked.connect(lambda: set_state([]))
        buttons.accepted.connect(dlg.accept); buttons.rejected.connect(dlg.reject)
        if dlg.exec() == QDialog.Accepted:
            cols = [table.item(r, 0).text() for r in range(table.rowCount()) if table.item(r, 0).checkState() == Qt.Checked]
            if not cols:
                cols = self.default_viewer_columns_for_rows(rows)
            self.pane_column_prefs[idx] = cols
            self.models[idx].set_rows(self.models[idx].rows, cols)
            self.apply_table_column_widths(idx)

    def apply_view_filters(self, side: Any):
        idx = self.side_index(side)
        try:
            start, end = self.current_viewer_time_range()
        except Exception as e:
            QMessageBox.warning(self, "Viewer Time Range", str(e))
            return
        base_rows = self.all_rows[idx]
        if start or end:
            rows = []
            for row in base_rows:
                ts = row.get("_ts")
                if not isinstance(ts, datetime):
                    continue
                if start and ts < start:
                    continue
                if end and ts > end:
                    continue
                rows.append(row)
        else:
            rows = list(base_rows)
        ts_index = [(row["_ts"], i) for i, row in enumerate(rows) if isinstance(row.get("_ts"), datetime)]
        ts_index.sort(key=lambda x: x[0])
        cols = self.pane_columns_for_rows(idx, rows)
        self.models[idx].set_rows(rows, cols)
        self.apply_table_column_widths(idx)
        self.ts_indexes[idx] = ts_index
        self._sync_aliases()
        label = self.pane_name(idx)
        if start or end:
            self.log(f"Viewer time filter {label}: {len(rows)}/{len(base_rows)} rows")
        else:
            self.log(f"Viewer time filter cleared {label}: {len(rows)} rows")

    def _sync_aliases(self):
        self.left_all_rows = self.all_rows[0]; self.right_all_rows = self.all_rows[1]
        self.left_ts = self.ts_indexes[0]; self.right_ts = self.ts_indexes[1]

    def apply_time_to_visible(self):
        for idx in self.visible_indices():
            self.apply_view_filters(idx)

    def apply_time_to_focused(self):
        for idx in self.visible_indices():
            if self.tables[idx].hasFocus() or self.search_edits[idx].hasFocus():
                self.apply_view_filters(idx)
                return
        self.apply_time_to_visible()

    def clear_viewer_time_range(self):
        self.viewer_time_enable.setChecked(False)
        self.viewer_start_edit.clear()
        self.viewer_end_edit.clear()
        self.apply_time_to_visible()

    def set_time_from_selected(self, target: str):
        for idx in self.visible_indices():
            indexes = self.tables[idx].selectionModel().selectedRows()
            if indexes:
                row = self.models[idx].row_at(indexes[0].row()) or {}
                ts_text = row.get("Timestamp", "")
                if ts_text:
                    if target == "start":
                        self.viewer_start_edit.setText(str(ts_text))
                    else:
                        self.viewer_end_edit.setText(str(ts_text))
                    self.viewer_time_enable.setChecked(True)
                    self.log(f"Set Viewer {target} from selected {self.pane_name(idx)}: {ts_text}")
                    return
        QMessageBox.information(self, "Viewer Time Range", "Select a row first.")

    def load_pane(self, side: Any):
        idx = self.side_index(side)
        source = self.sources[idx].currentText()
        label = self.pane_name(idx)
        self.log(f"Loading {label}: {source} ...")
        progress = self.make_progress("Log Viewer", f"Preparing {label}: {source} ...")
        progress.show()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            records = self.source_to_records(source, progress)
            if progress.wasCanceled():
                return
            progress.setLabelText(f"Building table model for {label}: {len(records)} rows ...")
            QApplication.processEvents()
            rows = [record_to_viewer_row(r) for r in records]
            file_hint = source
            if rows:
                files = sorted({str(r.get("File", "")) for r in rows if r.get("File", "")})
                if len(files) == 1:
                    file_hint = files[0]
                elif len(files) > 1:
                    file_hint = f"{len(files)} files loaded"
            self.all_rows[idx] = rows
            self.file_labels[idx].setText(f"File: {file_hint}")
            self.apply_view_filters(idx)
            self.log(f"{label} loaded: {len(rows)} rows")
        except Exception:
            write_startup_log("Multi Log Viewer load_pane failed.\n\n" + traceback.format_exc())
            QMessageBox.critical(self, "Log Viewer", "Viewer loading failed.\n\n" + traceback.format_exc() + f"\n\nLog:\n{startup_log_path()}")
        finally:
            QApplication.restoreOverrideCursor()
            progress.close()

    def load_side(self, side: str):
        self.load_pane(side)

    def apply_noise_rules_now(self):
        if self.apply_noise_chk.isChecked():
            self.log("Applying approved Noise rules to visible panes...")
        else:
            self.log("Noise filter disabled. Reloading visible panes...")
        self.load_visible()

    def find_next(self, side: Any):
        idx = self.side_index(side)
        table = self.tables[idx]
        model = self.models[idx]
        edit = self.search_edits[idx]
        term = edit.text().strip().lower()
        if not term:
            return
        current = -1
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if indexes:
            current = indexes[0].row()
        total = model.rowCount()
        if total <= 0:
            return
        for step in range(1, total + 1):
            i = (current + step) % total
            row = model.row_at(i) or {}
            hay = " ".join(str(row.get(k, "")) for k in ("Timestamp", "SourceType", "File", "Level", "Category", "Message", "Raw")).lower()
            if term in hay:
                table.selectRow(i)
                table.scrollTo(model.index(i, 0), QTableView.PositionAtCenter)
                self.show_detail(idx, row)
                self.log(f"Found in {self.pane_name(idx)}: row {i + 1}/{total}")
                return
        self.log(f"No match in {self.pane_name(idx)}: {term}")

    def row_selected(self, side: Any):
        if self._syncing:
            return
        idx = self.side_index(side)
        table = self.tables[idx]
        model = self.models[idx]
        indexes = table.selectionModel().selectedRows()
        if not indexes:
            return
        row = model.row_at(indexes[0].row())
        if not row:
            return
        self.show_detail(idx, row)
        ts = row.get("_ts")
        if isinstance(ts, datetime):
            self.jump_other_panes(idx, ts)

    def show_detail(self, side: Any, row: dict[str, Any]):
        idx = self.side_index(side)
        raw = row.get("Raw", "")
        msg = row.get("Message", "")
        self.detail.setPlainText(
            f"Selected {self.pane_name(idx)}\n"
            f"Timestamp: {row.get('Timestamp','')}\n"
            f"SourceType: {row.get('SourceType','')}\n"
            f"File: {row.get('File','')}  Line: {row.get('Line','')}\n"
            f"Message: {msg}\n\nRaw:\n{raw}"
        )

    def jump_other_panes(self, source_idx: int, ts: datetime):
        tol = self.tolerance_spin.value()
        messages = []
        self._syncing = True
        try:
            for idx in self.visible_indices():
                if idx == source_idx:
                    continue
                ts_index = self.ts_indexes[idx]
                if not ts_index:
                    continue
                lo, hi = 0, len(ts_index)
                while lo < hi:
                    mid = (lo + hi) // 2
                    if ts_index[mid][0] < ts:
                        lo = mid + 1
                    else:
                        hi = mid
                candidates = []
                for pos in (lo - 1, lo, lo + 1):
                    if 0 <= pos < len(ts_index):
                        candidates.append(ts_index[pos])
                if not candidates:
                    continue
                nearest_ts, nearest_row = min(candidates, key=lambda x: abs((x[0] - ts).total_seconds()))
                delta = abs((nearest_ts - ts).total_seconds())
                self.tables[idx].selectRow(nearest_row)
                self.tables[idx].scrollTo(self.models[idx].index(nearest_row, 0), QTableView.PositionAtCenter)
                status = f"{self.pane_name(idx)} Δ{delta:.3f}s"
                if tol and delta > tol:
                    status += f" outside ±{tol}s"
                messages.append(status)
        finally:
            self._syncing = False
        if messages:
            self.log("Cross reference: " + "; ".join(messages))

    def jump_other_side(self, side: str, ts: datetime):
        self.jump_other_panes(self.side_index(side), ts)

    def selected_record_from_pane(self, idx: int) -> Optional[LogRecord]:
        table = self.tables[idx]
        model = self.models[idx]
        indexes = table.selectionModel().selectedRows() if table.selectionModel() else []
        if not indexes:
            return None
        row = model.row_at(indexes[0].row())
        if not row:
            return None
        return LogRecord(row.get("_ts"), str(row.get("SourceType", "")), str(row.get("File", "")), int(row.get("Line", 0) or 0), str(row.get("Level", "")), str(row.get("Category", "")), str(row.get("Message", "")), str(row.get("Raw", "")))

    def selected_record_from_side(self, side: str) -> Optional[LogRecord]:
        return self.selected_record_from_pane(self.side_index(side))

    def selected_any_record(self) -> Optional[LogRecord]:
        for idx in self.visible_indices():
            rec = self.selected_record_from_pane(idx)
            if rec is not None:
                return rec
        return None

    def copy_rule_text(self, side: Any):
        idx = self.side_index(side)
        rec = self.selected_record_from_pane(idx)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        text = _compact_pattern(rec.message or rec.raw)
        QApplication.clipboard().setText(text)
        self.log(f"Copied {self.pane_name(idx)} rule text to clipboard: {text}")

    def approve_selected_noise(self, side: Any):
        idx = self.side_index(side)
        rec = self.selected_record_from_pane(idx)
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        engine = NoiseEngine()
        suggested = _compact_pattern(rec.message or rec.raw)
        pattern, ok = QInputDialog.getText(self, "Approve Noise Rule", "Rule text to hide (contains):", text=suggested)
        if not ok or not pattern.strip():
            return
        rid = engine.add_rule(pattern.strip(), (rec.source_type or "ANY").upper(), "contains", "hide", "both", "Approved by operator from Explorer selected row")
        engine.add_sample(rec, "noise-approved", 100)
        if self.noise_apply_mode.currentIndex() == 0:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nApplying filter now.")
            self.apply_noise_rules_now()
        else:
            QMessageBox.information(self, "Noise Learning", f"Approved noise rule ID {rid}.\nThe rule is saved. Press Apply Rules Now when you want to apply it to the current Explorer view.")

    def mark_selected_not_noise(self):
        rec = self.selected_any_record()
        if rec is None:
            QMessageBox.information(self, "Noise Learning", "Select a row first.")
            return
        NoiseEngine().mark_not_noise(rec)
        QMessageBox.information(self, "Noise Learning", "Saved as Not Noise training sample. Approved rules were not changed.")

    def manage_noise_rules(self):
        dlg = NoiseRuleManagerDialog(self)
        dlg.exec()
        if self.noise_apply_mode.currentIndex() == 0:
            self.apply_noise_rules_now()
        else:
            self.log("Noise rules changed. Press Apply Rules Now to apply them to the current Explorer view.")

    def export_selected_pair(self):
        out_dir = Path(self.parent_window.output_edit.text().strip() or self.parent_window.source_edit.text().strip() or str(Path.home()))
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"LogViewer_SelectedRows_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        selected = []
        for idx in self.visible_indices():
            rows = self.tables[idx].selectionModel().selectedRows()
            if rows:
                row = self.models[idx].row_at(rows[0].row())
                if row:
                    selected.append((self.pane_name(idx), row, self.models[idx].columns))
        if not selected:
            QMessageBox.information(self, "Log Viewer", "Select at least one row.")
            return
        with out_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Pane"] + VIEWER_COLUMNS + ["Raw"])
            for pane, row, _cols in selected:
                writer.writerow([pane] + [row.get(c, "") for c in VIEWER_COLUMNS] + [row.get("Raw", "")])
        QMessageBox.information(self, "Log Viewer", f"Exported.\n\n{out_path}")

# Replace the legacy class name used by MainWindow.
DualLogViewer = MultiPaneLogViewer



@dataclass
class DiscoveredFile:
    path: str
    log_type: str
    row_count: int
    start_ts: Optional[datetime]
    end_ts: Optional[datetime]
    size_bytes: int
    error: str = ""


def discover_one_file(path: Path, include_unknown: bool) -> DiscoveredFile:
    """Parse one candidate file for Smart File Discovery.

    This intentionally uses the same parser/timestamp logic as Merge so the
    preflight list and the final merge cannot disagree about date filtering.
    File created/modified timestamps are never used.
    """
    typ = "REVIEW" if is_review_file(path.name) else classify_file(path, include_unknown)
    if typ is None:
        return DiscoveredFile(str(path), "", 0, None, None, path.stat().st_size if path.exists() else 0)
    try:
        if typ == "REVIEW":
            rows = parse_review_file(path)
            timestamps = []
            for r in rows:
                text_dt = " ".join(str(r.get(k, "")) for k in ("date", "time") if r.get(k, ""))
                ts = parse_datetime_from_text(text_dt) if text_dt.strip() else None
                if ts:
                    timestamps.append(ts)
            return DiscoveredFile(str(path), typ, len(rows), min(timestamps) if timestamps else None, max(timestamps) if timestamps else None, path.stat().st_size)
        if typ == "PSC":
            psc_log, psc_params, parsed = parse_psc_file_detail(path)
            timestamps = [r.timestamp for r in parsed if r.timestamp]
            return DiscoveredFile(str(path), typ, len(parsed) + len(psc_log) + len(psc_params), min(timestamps) if timestamps else None, max(timestamps) if timestamps else None, path.stat().st_size)
        parsed = parse_file(path, typ)
        timestamps = [r.timestamp for r in parsed if r.timestamp]
        return DiscoveredFile(str(path), typ, len(parsed), min(timestamps) if timestamps else None, max(timestamps) if timestamps else None, path.stat().st_size)
    except Exception as exc:
        return DiscoveredFile(str(path), typ or "UNKNOWN", 0, None, None, path.stat().st_size if path.exists() else 0, str(exc))


class SmartFileDiscoveryDialog(QDialog):
    def __init__(self, parent, base_options: RunOptions):
        super().__init__(parent)
        self.setWindowTitle("Smart File Discovery")
        self.resize(1050, 720)
        self.base_options = base_options
        self.discovered: list[DiscoveredFile] = []
        self.filtered: list[DiscoveredFile] = []
        self.selected_files: list[str] = []
        self.selected_types: set[str] = set()
        self._build_ui()
        self._scan()

    def _build_ui(self):
        root = QVBoxLayout(self)
        info = QLabel("1) Detect log types in the selected folder. 2) Select log types and date range. 3) Confirm target files before starting.")
        info.setWordWrap(True)
        root.addWidget(info)

        type_box = QGroupBox("Detected Log Types")
        self.type_layout = QHBoxLayout(type_box)
        self.type_checks: dict[str, QCheckBox] = {}
        self.detected_all_btn = QPushButton("All")
        self.detected_clear_btn = QPushButton("Clear")
        self.detected_all_btn.setMaximumWidth(70)
        self.detected_clear_btn.setMaximumWidth(70)
        self.detected_all_btn.setToolTip("Select all detected log types. Existing file-level selections are kept when possible.")
        self.detected_clear_btn.setToolTip("Clear all detected log type checks. Use this when previous selections remain from another scan.")
        self.detected_all_btn.clicked.connect(lambda: self.set_all_detected_types(True))
        self.detected_clear_btn.clicked.connect(lambda: self.set_all_detected_types(False))
        root.addWidget(type_box)

        range_box = QGroupBox("Date Range Filter (based on each file's parsed row timestamps, not file created/modified date)")
        rg = QHBoxLayout(range_box)
        self.enable_start = QCheckBox("Start")
        self.start_dt = QLineEdit()
        self.start_dt.setPlaceholderText("YYYY/MM/DD HH:MM:SS or YYYY-MM-DD")
        self.enable_end = QCheckBox("End")
        self.end_dt = QLineEdit()
        self.end_dt.setPlaceholderText("YYYY/MM/DD HH:MM:SS or YYYY-MM-DD")
        self.apply_btn = QPushButton("Apply Type/Date Filter")
        self.apply_btn.clicked.connect(self.apply_filter)
        for w in [self.enable_start, self.start_dt, self.enable_end, self.end_dt, self.apply_btn]:
            rg.addWidget(w)
        root.addWidget(range_box)

        self.summary_label = QLabel("Ready")
        root.addWidget(self.summary_label)
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["Use", "Type", "Start", "End", "Rows", "Size MB", "File", "Error"])
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.table.setSortingEnabled(True)
        root.addWidget(self.table, stretch=1)

        btns = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All")
        self.select_all_btn.clicked.connect(lambda: self.set_all(True))
        self.clear_all_btn = QPushButton("Clear All")
        self.clear_all_btn.clicked.connect(lambda: self.set_all(False))
        self.start_btn = QPushButton("Start with Checked Files")
        self.start_btn.clicked.connect(self.accept_with_selection)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        for b in [self.select_all_btn, self.clear_all_btn, self.start_btn, self.cancel_btn]:
            btns.addWidget(b)
        root.addLayout(btns)

    def _scan(self):
        src = Path(self.base_options.source_folder)
        all_files = list(iter_files(src, self.base_options.recursive))
        progress = QProgressDialog("Scanning and parsing candidate files...", "Cancel", 0, max(1, len(all_files)), self)
        progress.setWindowTitle("Smart File Discovery")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(300)
        include_unknown = self.base_options.include_unknown
        discovered = []
        for i, path in enumerate(all_files, start=1):
            if progress.wasCanceled():
                break
            progress.setValue(i - 1)
            progress.setLabelText(f"Scanning {i}/{len(all_files)}\n{path.name}")
            QApplication.processEvents()
            meta = discover_one_file(path, include_unknown)
            if meta.log_type:
                discovered.append(meta)
        progress.setValue(len(all_files))
        self.discovered = discovered
        self.populate_type_checks()
        self.apply_filter()

    def populate_type_checks(self):
        while self.type_layout.count():
            item = self.type_layout.takeAt(0)
            w = item.widget()
            if w and w not in (self.detected_all_btn, self.detected_clear_btn):
                w.deleteLater()
        self.type_checks.clear()
        counts: dict[str, int] = {}
        for d in self.discovered:
            counts[d.log_type] = counts.get(d.log_type, 0) + 1
        preferred = ["WS", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "REVIEW", "UNKNOWN"]
        for typ in [t for t in preferred if t in counts] + sorted(t for t in counts if t not in preferred):
            chk = QCheckBox(f"{typ} ({counts[typ]})")
            chk.setChecked(False)
            chk.setToolTip(f"Include/exclude all detected {typ} files before final file confirmation.")
            chk.stateChanged.connect(self.apply_filter)
            self.type_checks[typ] = chk
            self.type_layout.addWidget(chk)
        self.type_layout.addStretch(1)
        self.type_layout.addWidget(self.detected_all_btn)
        self.type_layout.addWidget(self.detected_clear_btn)


    def set_all_detected_types(self, checked: bool):
        # Block repeated filtering until all checkboxes have been updated.
        for chk in self.type_checks.values():
            chk.blockSignals(True)
            chk.setChecked(checked)
            chk.blockSignals(False)
        self.apply_filter()

    def _range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        start = end = None
        if self.enable_start.isChecked() and self.start_dt.text().strip():
            start = parse_viewer_datetime_text(self.start_dt.text().strip(), is_end=False)
        if self.enable_end.isChecked() and self.end_dt.text().strip():
            end = parse_viewer_datetime_text(self.end_dt.text().strip(), is_end=True)
        if start and end and start > end:
            raise ValueError("Start is later than End.")
        return start, end

    def _overlaps_range(self, d: DiscoveredFile, start: Optional[datetime], end: Optional[datetime]) -> bool:
        if not start and not end:
            return True
        # Files without parsed timestamps are kept for operator decision, because
        # they may still be important or require parser improvement.
        if not d.start_ts and not d.end_ts:
            return True
        s = d.start_ts or d.end_ts
        e = d.end_ts or d.start_ts
        if start and e and e < start:
            return False
        if end and s and s > end:
            return False
        return True

    def apply_filter(self):
        try:
            start, end = self._range()
        except Exception as exc:
            QMessageBox.warning(self, "Smart File Discovery", str(exc))
            return
        enabled = {typ for typ, chk in self.type_checks.items() if chk.isChecked()}
        self.selected_types = set(enabled)
        rows = [d for d in self.discovered if d.log_type in enabled and self._overlaps_range(d, start, end)]
        rows.sort(key=lambda d: (d.start_ts or datetime.max, d.log_type, Path(d.path).name))
        self.filtered = rows
        self.populate_file_table(rows)
        total_rows = sum(d.row_count for d in rows)
        total_mb = sum(d.size_bytes for d in rows) / (1024*1024)
        full_start = min((d.start_ts for d in self.discovered if d.start_ts), default=None)
        full_end = max((d.end_ts for d in self.discovered if d.end_ts), default=None)
        range_text = f"Detected range: {format_viewer_timestamp(full_start)} - {format_viewer_timestamp(full_end)}" if full_start or full_end else "Detected range: unknown"
        self.summary_label.setText(f"{range_text} | Filtered files: {len(rows)} | Estimated rows: {total_rows:,} | Size: {total_mb:.1f} MB")

    def populate_file_table(self, rows: list[DiscoveredFile]):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for r, d in enumerate(rows):
            chk = QTableWidgetItem("")
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            chk.setCheckState(Qt.Checked)
            chk.setData(Qt.UserRole, d.path)
            self.table.setItem(r, 0, chk)
            vals = [d.log_type, format_viewer_timestamp(d.start_ts), format_viewer_timestamp(d.end_ts), f"{d.row_count:,}", f"{d.size_bytes/(1024*1024):.1f}", str(Path(d.path).name), d.error]
            for c, v in enumerate(vals, start=1):
                item = QTableWidgetItem(str(v))
                item.setToolTip(d.path if c == 6 else str(v))
                self.table.setItem(r, c, item)
        self.table.setSortingEnabled(True)

    def set_all(self, checked: bool):
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item:
                item.setCheckState(Qt.Checked if checked else Qt.Unchecked)

    def accept_with_selection(self):
        selected = []
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item and item.checkState() == Qt.Checked:
                selected.append(str(item.data(Qt.UserRole)))
        if not selected:
            QMessageBox.warning(self, "Smart File Discovery", "No file is checked.")
            return
        self.selected_files = selected
        self.accept()


class PluginManagerDialog(QDialog):
    """Install/update/remove File Type plugins without rebuilding the main EXE."""
    COLS = ["Status", "ID", "Display Name", "Version", "Mode", "Patterns", "Path"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Plugin Manager / Update File Type")
        self.resize(980, 520)
        root = QVBoxLayout(self)
        info = QLabel("Install File Type plugin ZIP files. Installed plugins are hot-reloaded and will appear in Log Types and Smart File Discovery.")
        info.setWordWrap(True)
        root.addWidget(info)

        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        root.addWidget(self.table, stretch=1)

        btns = QHBoxLayout()
        self.install_btn = QPushButton("Install / Update Plugin ZIP")
        self.install_btn.setToolTip("Select a *.plugin.zip. The manifest.json is validated, then the plugin is installed into the plugins folder.")
        self.validate_btn = QPushButton("Validate Selected")
        self.reload_btn = QPushButton("Reload")
        self.enable_btn = QPushButton("Enable")
        self.disable_btn = QPushButton("Disable")
        self.remove_btn = QPushButton("Remove")
        self.close_btn = QPushButton("Close")
        for b in [self.install_btn, self.validate_btn, self.reload_btn, self.enable_btn, self.disable_btn, self.remove_btn]:
            btns.addWidget(b)
        btns.addStretch(1)
        btns.addWidget(self.close_btn)
        root.addLayout(btns)

        self.install_btn.clicked.connect(self.install_plugin)
        self.validate_btn.clicked.connect(self.validate_selected)
        self.reload_btn.clicked.connect(self.reload)
        self.enable_btn.clicked.connect(lambda: self.set_selected_enabled(True))
        self.disable_btn.clicked.connect(lambda: self.set_selected_enabled(False))
        self.remove_btn.clicked.connect(self.remove_selected)
        self.close_btn.clicked.connect(self.accept)
        self.reload()

    def _all_manifests(self) -> list[tuple[Path, dict[str, Any]]]:
        out = []
        root = plugins_dir()
        for d in sorted(root.iterdir(), key=lambda x: x.name.lower()):
            mf = d / "manifest.json"
            if not mf.exists():
                continue
            try:
                data = json.loads(mf.read_text(encoding="utf-8"))
                out.append((mf, data))
            except Exception:
                out.append((mf, {"id": d.name, "display_name": "Invalid manifest", "enabled": False, "_invalid": True}))
        return out

    def reload(self):
        rows = self._all_manifests()
        self.table.setRowCount(0)
        for mf, data in rows:
            r = self.table.rowCount(); self.table.insertRow(r)
            enabled = bool(data.get("enabled", True)) and not data.get("_invalid")
            status = "Enabled" if enabled else ("Invalid" if data.get("_invalid") else "Disabled")
            patterns = data.get("patterns") or data.get("file_patterns") or []
            if isinstance(patterns, str): patterns = [patterns]
            vals = [
                status,
                str(data.get("id", mf.parent.name)),
                str(data.get("display_name", "")),
                str(data.get("version", "")),
                ", ".join(data.get("mode", [])) if isinstance(data.get("mode", []), list) else str(data.get("mode", "")),
                ", ".join(map(str, patterns)),
                str(mf.parent),
            ]
            for c, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if c == 0:
                    if status == "Enabled": item.setToolTip("This plugin is active and will be used by Auto Detect and Log Types.")
                    elif status == "Disabled": item.setToolTip("This plugin is installed but not used.")
                    else: item.setToolTip("This plugin manifest could not be read.")
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()

    def _selected_manifest_path(self) -> Optional[Path]:
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Plugin Manager", "Select a plugin row first.")
            return None
        item = self.table.item(row, 6)
        if not item:
            return None
        return Path(item.text()) / "manifest.json"

    def install_plugin(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Select File Type Plugin ZIP", "", "Plugin ZIP (*.zip *.plugin.zip);;All files (*.*)")
        if not fn:
            return
        z = Path(fn)
        ok, msg, manifest = validate_plugin_zip(z)
        if not ok:
            QMessageBox.critical(self, "Invalid Plugin", msg)
            return
        name = manifest.get("display_name", manifest.get("id"))
        ver = manifest.get("version", "")
        if QMessageBox.question(self, "Install Plugin", f"Install/update plugin?\n\n{name} {ver}\n\n{z.name}") != QMessageBox.Yes:
            return
        ok, msg = install_plugin_zip(z)
        if ok:
            QMessageBox.information(self, "Plugin Installed", msg)
        else:
            QMessageBox.critical(self, "Install Failed", msg)
        self.reload()

    def validate_selected(self):
        mf = self._selected_manifest_path()
        if not mf: return
        try:
            data = json.loads(mf.read_text(encoding="utf-8"))
            missing = [k for k in ["id", "display_name"] if not data.get(k)]
            pats = data.get("patterns") or data.get("file_patterns")
            if not pats: missing.append("patterns/file_patterns")
            if missing:
                QMessageBox.warning(self, "Plugin Validation", "Missing: " + ", ".join(missing))
            else:
                QMessageBox.information(self, "Plugin Validation", "OK")
        except Exception as exc:
            QMessageBox.critical(self, "Plugin Validation", str(exc))

    def set_selected_enabled(self, enabled: bool):
        mf = self._selected_manifest_path()
        if not mf: return
        try:
            data = json.loads(mf.read_text(encoding="utf-8"))
            data["enabled"] = bool(enabled)
            data["updated_at"] = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            mf.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            self.reload()
        except Exception as exc:
            QMessageBox.critical(self, "Plugin Manager", str(exc))

    def remove_selected(self):
        mf = self._selected_manifest_path()
        if not mf: return
        if QMessageBox.question(self, "Remove Plugin", f"Remove plugin folder?\n\n{mf.parent}") != QMessageBox.Yes:
            return
        try:
            shutil.rmtree(mf.parent, ignore_errors=True)
            self.reload()
        except Exception as exc:
            QMessageBox.critical(self, "Remove Failed", str(exc))


class CoreUpdateDialog(QDialog):
    """Stage a core application update ZIP safely. Actual EXE replacement is done by a generated helper BAT."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Update Tool Core")
        self.resize(760, 420)
        root = QVBoxLayout(self)
        root.addWidget(QLabel("Select a Core Update ZIP. The package is validated, backed up/staged, and an apply_update.bat helper is generated."))
        self.log_box = QTextEdit(); self.log_box.setReadOnly(True)
        root.addWidget(self.log_box, stretch=1)
        row = QHBoxLayout()
        self.select_btn = QPushButton("Select Core Update ZIP")
        self.open_stage_btn = QPushButton("Open Update Folder")
        self.close_btn = QPushButton("Close")
        row.addWidget(self.select_btn); row.addWidget(self.open_stage_btn); row.addStretch(1); row.addWidget(self.close_btn)
        root.addLayout(row)
        self.select_btn.clicked.connect(self.select_zip)
        self.open_stage_btn.clicked.connect(self.open_update_folder)
        self.close_btn.clicked.connect(self.accept)
        self.stage_root = Path(os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()) / "LogMergeTool_NoExcel" / "core_updates"
        self.stage_root.mkdir(parents=True, exist_ok=True)

    def log(self, text: str):
        self.log_box.append(text)

    def select_zip(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Select Core Update ZIP", "", "Core Update ZIP (*.zip);;All files (*.*)")
        if not fn:
            return
        z = Path(fn)
        try:
            with zipfile.ZipFile(z, "r") as zf:
                names = set(zf.namelist())
                if "manifest.json" not in names:
                    QMessageBox.critical(self, "Invalid Update", "manifest.json is missing.")
                    return
                manifest = json.loads(zf.read("manifest.json").decode("utf-8-sig"))
                if manifest.get("package_type") not in {"core_update", "tool_update"}:
                    QMessageBox.critical(self, "Invalid Update", "manifest.json package_type must be core_update.")
                    return
                ver = str(manifest.get("version", "unknown"))
                app_id = str(manifest.get("app_id", "log_merge_tool_noexcel"))
                if QMessageBox.question(self, "Stage Core Update", f"Stage core update?\n\nApp: {app_id}\nVersion: {ver}\nFile: {z.name}") != QMessageBox.Yes:
                    return
                stage = self.stage_root / re.sub(r"[^A-Za-z0-9_.-]+", "_", ver)
                if stage.exists():
                    shutil.rmtree(stage, ignore_errors=True)
                stage.mkdir(parents=True, exist_ok=True)
                zf.extractall(stage)
                (stage / "installed_from.txt").write_text(str(z), encoding="utf-8")
                bat = stage / "apply_update.bat"
                bat.write_text(self._apply_bat_text(stage), encoding="utf-8")
                self.log(f"Staged core update: {ver}")
                self.log(f"Folder: {stage}")
                self.log("Next step: close LogMergeTool, then run apply_update.bat from the staged folder.")
                QMessageBox.information(self, "Core Update Staged", f"Update staged.\n\n{stage}\n\nClose this tool, then run apply_update.bat.")
        except Exception as exc:
            QMessageBox.critical(self, "Update Failed", str(exc))

    def _apply_bat_text(self, stage: Path) -> str:
        return """@echo off
setlocal
cd /d "%~dp0"
echo LogMergeTool core update staged in this folder.
echo.
echo This helper is conservative: close LogMergeTool, back up the current install folder,
echo then copy the staged EXE/files into the install folder.
echo.
echo Future versions can replace files automatically using this same updater hook.
pause
"""

    def open_update_folder(self):
        try:
            os.startfile(str(self.stage_root))
        except Exception as exc:
            QMessageBox.critical(self, "Open Folder Failed", str(exc))

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.settings = QSettings("LogMerge", "NoExcel")
        self.worker: Optional[MergeWorker] = None
        self.progress_dialog = None
        self.import_worker = None
        self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
        self.resize(860, 680)
        self.build_ui()
        self.load_settings()

    def build_ui(self):
        root = QVBoxLayout(self)

        title = QLabel(f"{APP_TITLE}")
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        root.addWidget(title)
        subtitle = QLabel("No Microsoft Excel / No VBA / No COM. v39: Log Viewer selector layout + Smart Discovery All/Clear.")
        root.addWidget(subtitle)

        folder_box = QGroupBox("Folders")
        folder_layout = QFormLayout(folder_box)
        self.source_edit = QLineEdit()
        self.source_btn = QPushButton("Browse...")
        self.source_btn.clicked.connect(self.pick_source)
        h1 = QHBoxLayout(); h1.addWidget(self.source_edit); h1.addWidget(self.source_btn)
        folder_layout.addRow("Source Log Folder", h1)
        self.output_edit = QLineEdit()
        self.output_btn = QPushButton("Browse...")
        self.output_btn.clicked.connect(self.pick_output)
        self.output_same_btn = QPushButton("Same as Source")
        self.output_same_btn.setToolTip("Set Output Folder to the current Source Log Folder.")
        self.output_same_btn.clicked.connect(lambda: self.output_edit.setText(self.source_edit.text().strip()))
        h2 = QHBoxLayout(); h2.addWidget(self.output_edit); h2.addWidget(self.output_btn); h2.addWidget(self.output_same_btn)
        folder_layout.addRow("Output Folder", h2)
        self.recursive_chk = QCheckBox("Include all subfolders recursively")
        self.recursive_chk.setChecked(True)
        folder_layout.addRow("", self.recursive_chk)
        root.addWidget(folder_box)

        date_box = QGroupBox("Date Range")
        date_layout = QFormLayout(date_box)
        self.use_start = QCheckBox("Use Start Date")
        self.use_start.setChecked(False)
        self.start_date = QDateEdit(QDate.currentDate())
        self.start_date.setCalendarPopup(True)
        hs = QHBoxLayout(); hs.addWidget(self.use_start); hs.addWidget(self.start_date)
        date_layout.addRow("Start", hs)
        self.use_end = QCheckBox("Use End Date")
        self.use_end.setChecked(False)
        self.end_date = QDateEdit(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        he = QHBoxLayout(); he.addWidget(self.use_end); he.addWidget(self.end_date)
        date_layout.addRow("End", he)
        root.addWidget(date_box)

        meta_box = QGroupBox("Site Information")
        meta_layout = QFormLayout(meta_box)
        self.serial_combo = QComboBox(); self.serial_combo.setEditable(True)
        self.site_combo = QComboBox(); self.site_combo.setEditable(True)
        self.site_map = load_shared_site_map(app_dir() / "site_serial_map.json", DEFAULT_SITE_MAP)
        self.serial_combo.addItem(""); self.site_combo.addItem("")
        for serial in sorted({str(item.get("serial", "")).strip() for item in self.site_map if str(item.get("serial", "")).strip()}):
            self.serial_combo.addItem(serial)
        for site in sorted({str(item.get("site", "")).strip() for item in self.site_map if str(item.get("site", "")).strip()}):
            self.site_combo.addItem(site)
        self.serial_combo.setToolTip(f"Shared Service Hub master: {shared_master_path()}")
        self.site_combo.setToolTip(f"Shared Service Hub master: {shared_master_path()}")
        self.serial_combo.currentTextChanged.connect(self.sync_site_from_serial)
        self.site_combo.currentTextChanged.connect(self.sync_serial_from_site)
        meta_layout.addRow("Serial Number", self.serial_combo)
        meta_layout.addRow("Site Name", self.site_combo)
        root.addWidget(meta_box)

        type_box = QGroupBox("Log Types")
        type_layout = QHBoxLayout(type_box)
        self.type_layout = type_layout
        self.plugin_checks: dict[str, QCheckBox] = {}
        self.chk_ws = QCheckBox("WS"); self.chk_ws.setChecked(False)
        self.chk_cga = QCheckBox("CGA"); self.chk_cga.setChecked(False)
        self.chk_csa = QCheckBox("CSA"); self.chk_csa.setChecked(False)
        self.chk_mrserver = QCheckBox("MRSERVER"); self.chk_mrserver.setChecked(False)
        self.chk_gesys = QCheckBox("GESYS"); self.chk_gesys.setChecked(False)
        self.chk_lais = QCheckBox("LAIS"); self.chk_lais.setChecked(False)
        self.chk_psc = QCheckBox("PSC"); self.chk_psc.setChecked(False)
        self.chk_review = QCheckBox("Review (Import Only)"); self.chk_review.setChecked(False)
        self.chk_review.setEnabled(False)
        self.chk_review.setToolTip("Review files are not merged. Use the Import Review Only button. Target: review.out, review.out.ar, review.out.*")
        self.chk_unknown = QCheckBox("UNKNOWN .log/.txt"); self.chk_unknown.setChecked(False)
        for w in [self.chk_ws, self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_review, self.chk_unknown]:
            type_layout.addWidget(w)
        self.logtypes_all_btn = QPushButton("All")
        self.logtypes_none_btn = QPushButton("None")
        self.logtypes_all_btn.setMaximumWidth(48)
        self.logtypes_none_btn.setMaximumWidth(58)
        self.logtypes_all_btn.setToolTip("Select all merge-capable built-in log types.")
        self.logtypes_none_btn.setToolTip("Clear all log type selections.")
        self.logtypes_all_btn.clicked.connect(lambda: self.set_all_log_types(True))
        self.logtypes_none_btn.clicked.connect(lambda: self.set_all_log_types(False))
        type_layout.addStretch(1)
        type_layout.addWidget(self.logtypes_all_btn)
        type_layout.addWidget(self.logtypes_none_btn)
        self.refresh_plugin_checkboxes()
        root.addWidget(type_box)

        perf_box = QGroupBox("Performance")
        perf_layout = QHBoxLayout(perf_box)
        self.chk_turbo_csv = QCheckBox("v8 Turbo CSV for large data")
        self.chk_turbo_csv.setChecked(True)
        self.chk_summary_xlsx = QCheckBox("CSV + Summary xlsx for large data")
        self.chk_summary_xlsx.setChecked(True)
        self.chk_fast_xlsx = QCheckBox("Fast xlsx mode")
        self.chk_fast_xlsx.setChecked(True)
        self.chk_skip_merged1 = QCheckBox("Skip Merged_1 duplicate")
        self.chk_skip_merged1.setChecked(True)
        self.chk_timestamp_text = QCheckBox("Timestamp as text for speed")
        self.chk_timestamp_text.setChecked(True)
        self.chk_disable_styles = QCheckBox("No styles for large xlsx")
        self.chk_disable_styles.setChecked(True)
        self.chk_turbo_csv.setToolTip("ON: For large merged results, write full data to CSV instead of forcing a very large xlsx. This is usually the fastest and most stable option.")
        self.chk_summary_xlsx.setToolTip("ON: When CSV output is used, also create a small summary xlsx for Excel review. Recommended for large logs.")
        self.chk_fast_xlsx.setToolTip("ON: Use a lighter xlsx writer path and reduce expensive Excel features where possible.")
        self.chk_skip_merged1.setToolTip("ON: Skip the duplicate Merged_1 compatibility sheet. This reduces file size and write time.")
        self.chk_timestamp_text.setToolTip("ON: Write Timestamp as text for speed. OFF uses Excel datetime formatting, which is slower but easier for Excel date filters.")
        self.chk_disable_styles.setToolTip("ON: Disable heavy cell styling for large xlsx files. This improves speed and avoids Excel repair issues.")
        self.worker_combo = QComboBox()
        self.worker_combo.addItem("Auto", "0")
        for n in [2, 4, 6, 8, 12, 16]:
            self.worker_combo.addItem(str(n), str(n))
        self.worker_combo.setToolTip("Parser worker count. Auto is recommended unless you need to limit CPU usage.")
        perf_layout.addWidget(self.chk_turbo_csv)
        perf_layout.addWidget(self.chk_summary_xlsx)
        perf_layout.addWidget(self.chk_fast_xlsx)
        perf_layout.addWidget(self.chk_skip_merged1)
        perf_layout.addWidget(self.chk_timestamp_text)
        perf_layout.addWidget(self.chk_disable_styles)
        perf_layout.addWidget(QLabel("Parser workers"))
        perf_layout.addWidget(self.worker_combo)
        root.addWidget(perf_box)

        noise_group = QGroupBox("Smart Noise Learning / Filter")
        noise_layout = QHBoxLayout(noise_group)
        self.noise_enable_merge_chk = QCheckBox("Apply approved Noise Rules in Merge output")
        self.noise_enable_merge_chk.setChecked(True)
        self.noise_exclude_output_chk = QCheckBox("Legacy: exclude approved noise from output")
        self.noise_exclude_output_chk.setVisible(False)
        self.noise_learning_chk = QCheckBox("Learning Mode: suggest only")
        self.noise_learning_chk.setChecked(True)
        self.noise_manage_btn = QPushButton("Manage Noise Rules")
        self.noise_manage_btn.clicked.connect(self.manage_noise_rules_main)
        self.noise_manage_btn.setToolTip("Open the Noise Rules Manager to enable, disable, test, edit, or delete operator-approved rules.")
        self.noise_enable_merge_chk.setToolTip("ON: Approved and enabled Noise rules are applied during Merge. Matching rows are removed from output. Turn OFF to keep all records even if rules exist.")
        self.noise_learning_chk.setToolTip("ON: The tool may suggest noise candidates, but nothing is removed until an operator approves a rule. Recommended ON.")
        noise_layout.addWidget(self.noise_enable_merge_chk)
        noise_layout.addWidget(self.noise_exclude_output_chk)
        noise_layout.addWidget(self.noise_learning_chk)
        noise_layout.addWidget(self.noise_manage_btn)
        noise_layout.addStretch(1)
        root.addWidget(noise_group)

        btn_layout = QHBoxLayout()
        self.run_btn = QPushButton("Run Log Merge")
        self.run_btn.clicked.connect(self.run_clicked)
        self.pause_btn = QPushButton("Pause")
        self.pause_btn.clicked.connect(self.pause_clicked)
        self.pause_btn.setEnabled(False)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.cancel_clicked)
        self.cancel_btn.setEnabled(False)
        self.import_psc_btn = QPushButton("Import PSC Only")
        self.import_psc_btn.clicked.connect(self.import_psc_clicked)
        self.import_review_btn = QPushButton("Import Review Only")
        self.import_review_btn.clicked.connect(self.import_review_clicked)
        self.import_file_btn = QPushButton("Import Selected File Only")
        self.import_file_btn.clicked.connect(self.import_selected_file_clicked)
        self.export_project_btn = QPushButton("Export Project")
        self.export_project_btn.clicked.connect(self.export_project_clicked)
        self.import_project_btn = QPushButton("Import Project")
        self.import_project_btn.clicked.connect(self.import_project_clicked)
        self.split_merge_btn = QPushButton("Split Merge Project")
        self.split_merge_btn.clicked.connect(self.split_merge_clicked)
        self.split_merge_btn.setToolTip("Split a merged Search xlsx/csv back into per-source reconstructed log files. Uses Raw column grouped by SourceType and File.")
        self.open_out_btn = QPushButton("Open Output Folder")
        self.open_out_btn.clicked.connect(self.open_output_folder)
        self.update_file_type_btn = QPushButton("Update File Type")
        self.update_file_type_btn.clicked.connect(self.update_file_type_clicked)
        self.update_file_type_btn.setToolTip("Install or update a File Type plugin ZIP. New log types can be added without rebuilding the EXE.")
        self.update_tool_btn = QPushButton("Update Tool")
        self.update_tool_btn.clicked.connect(self.update_tool_clicked)
        self.update_tool_btn.setToolTip("Stage a Core Update ZIP for the tool itself. This is separate from File Type plugins.")
        self.reset_defaults_btn = QPushButton("Reset to Defaults")
        self.reset_defaults_btn.clicked.connect(self.reset_to_defaults_clicked)
        self.reset_defaults_btn.setToolTip("Reset folders, date range, site information, and log types to blank. Performance is set to the recommended fastest settings. Smart Noise checkboxes are turned ON.")
        self.viewer_btn = QPushButton("Open Dual Log Viewer / Log Explorer")
        self.viewer_btn.clicked.connect(self.open_dual_viewer)
        self.viewer_window = None
        btn_layout.addWidget(self.run_btn)
        btn_layout.addWidget(self.pause_btn)
        btn_layout.addWidget(self.cancel_btn)
        btn_layout.addWidget(self.import_psc_btn)
        btn_layout.addWidget(self.import_review_btn)
        btn_layout.addWidget(self.import_file_btn)
        btn_layout.addWidget(self.export_project_btn)
        btn_layout.addWidget(self.import_project_btn)
        btn_layout.addWidget(self.split_merge_btn)
        btn_layout.addWidget(self.open_out_btn)
        btn_layout.addWidget(self.update_file_type_btn)
        btn_layout.addWidget(self.update_tool_btn)
        btn_layout.addWidget(self.reset_defaults_btn)
        btn_layout.addWidget(self.viewer_btn)
        root.addLayout(btn_layout)

        self.progress = QProgressBar(); self.progress.setRange(0, 100)
        root.addWidget(self.progress)
        self.status_label = QLabel("Ready")
        root.addWidget(self.status_label)
        self.log_view = QTextEdit(); self.log_view.setReadOnly(True)
        root.addWidget(self.log_view, stretch=1)

    def set_all_log_types(self, checked: bool):
        for chk in [self.chk_ws, self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_unknown]:
            chk.setChecked(checked)
        for chk in getattr(self, "plugin_checks", {}).values():
            chk.setChecked(checked)

    def refresh_plugin_checkboxes(self):
        # Remove old dynamic plugin checkboxes from the layout.
        for chk in list(getattr(self, "plugin_checks", {}).values()):
            try:
                self.type_layout.removeWidget(chk)
                chk.deleteLater()
            except Exception:
                pass
        self.plugin_checks = {}
        for plg in load_file_type_plugins():
            pid = str(plg.get("id", "")).strip()
            if not pid:
                continue
            label = f"{plg.get('display_name', pid)} (Plugin)"
            chk = QCheckBox(label)
            chk.setChecked(False)
            chk.setToolTip(f"Plugin: {pid} / version {plg.get('version','')} / patterns: {', '.join(plg.get('patterns', []))}")
            self.plugin_checks[pid] = chk
            self.type_layout.addWidget(chk)

    def update_file_type_clicked(self):
        dlg = PluginManagerDialog(self)
        dlg.exec()
        self.refresh_plugin_checkboxes()
        self.log("Plugin list reloaded. Newly installed File Types are available as plugin checkboxes and in Smart File Discovery.")

    def update_tool_clicked(self):
        dlg = CoreUpdateDialog(self)
        dlg.exec()

    def log(self, msg: str):
        self.log_view.append(msg)

    def _show_progress_popup(self, title: str, label: str, minimum: int = 0, maximum: int = 100, cancellable: bool = True):
        self._close_progress_popup()
        dlg = QProgressDialog(label, "Cancel" if cancellable else None, minimum, maximum, self)
        dlg.setWindowTitle(title)
        dlg.setMinimumDuration(500)
        dlg.resize(560, 190)
        dlg.setStyleSheet("""
            QProgressDialog { background: #fbfbfd; font-size: 12pt; }
            QLabel { font-size: 12pt; padding: 6px; }
            QProgressBar { min-height: 22px; border: 1px solid #c9ced6; border-radius: 8px; text-align: center; background: #f0f2f5; }
            QProgressBar::chunk { border-radius: 8px; background: #6aa7ff; }
            QPushButton { padding: 6px 16px; border-radius: 8px; }
        """)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setWindowModality(Qt.WindowModal)
        if cancellable:
            dlg.canceled.connect(self.cancel_clicked)
        dlg.show()
        QApplication.processEvents()
        self.progress_dialog = dlg
        return dlg

    def _close_progress_popup(self):
        try:
            if self.progress_dialog is not None:
                self.progress_dialog.close()
        except Exception:
            pass
        self.progress_dialog = None

    def _run_import_task(self, title: str, label: str, func, done_message_prefix: str):
        self._show_progress_popup(title, label, 0, 0, cancellable=False)
        self.status_label.setText(label)
        self.import_worker = SimpleTaskWorker(func)
        def ok(out):
            self._close_progress_popup()
            self.import_worker = None
            self.status_label.setText(f"{done_message_prefix} completed")
            self.log(f"{done_message_prefix} completed: {out}")
            QMessageBox.information(self, APP_TITLE, f"{done_message_prefix} completed.\n\nOutput:\n{out}")
        def failed(text):
            self._close_progress_popup()
            self.import_worker = None
            self.status_label.setText(f"{done_message_prefix} failed")
            self.log(text)
            QMessageBox.critical(self, APP_TITLE, text)
        self.import_worker.finished_ok.connect(ok)
        self.import_worker.failed.connect(failed)
        self.import_worker.start()

    def pick_source(self):
        d = QFileDialog.getExistingDirectory(self, "Select Source Log Folder", self.source_edit.text() or str(Path.home()))
        if d:
            self.source_edit.setText(d)
            if not self.output_edit.text():
                self.output_edit.setText(d)

    def pick_output(self):
        d = QFileDialog.getExistingDirectory(self, "Select Output Folder", self.output_edit.text() or self.source_edit.text() or str(Path.home()))
        if d:
            self.output_edit.setText(d)

    def sync_site_from_serial(self, text: str):
        if getattr(self, "_site_master_syncing", False):
            return
        serial = text.strip()
        matches = [str(item.get("site", "")).strip() for item in self.site_map if str(item.get("serial", "")).strip() == serial]
        sites = sorted({x for x in matches if x})
        if len(sites) == 1 and self.site_combo.currentText().strip() != sites[0]:
            self._site_master_syncing = True
            try:
                self.site_combo.setCurrentText(sites[0])
            finally:
                self._site_master_syncing = False

    def sync_serial_from_site(self, text: str):
        if getattr(self, "_site_master_syncing", False):
            return
        site = text.strip()
        serials = sorted({str(item.get("serial", "")).strip() for item in self.site_map if str(item.get("site", "")).strip() == site and str(item.get("serial", "")).strip()})
        if not serials:
            return
        self._site_master_syncing = True
        try:
            current = self.serial_combo.currentText().strip()
            self.serial_combo.blockSignals(True)
            self.serial_combo.clear()
            self.serial_combo.addItem("")
            self.serial_combo.addItems(serials)
            self.serial_combo.blockSignals(False)
            if len(serials) == 1:
                self.serial_combo.setCurrentText(serials[0])
            elif current in serials:
                self.serial_combo.setCurrentText(current)
            else:
                self.serial_combo.setCurrentText("")
        finally:
            self._site_master_syncing = False

    def collect_options(self) -> RunOptions:
        return RunOptions(
            source_folder=self.source_edit.text().strip(),
            output_folder=self.output_edit.text().strip() or self.source_edit.text().strip(),
            recursive=self.recursive_chk.isChecked(),
            use_start=self.use_start.isChecked(),
            use_end=self.use_end.isChecked(),
            start_date=self.start_date.date().toString("yyyy-MM-dd"),
            end_date=self.end_date.date().toString("yyyy-MM-dd"),
            serial=self.serial_combo.currentText().strip(),
            site=self.site_combo.currentText().strip(),
            include_ws=self.chk_ws.isChecked(),
            include_watersystem=getattr(self, "chk_watersystem", QCheckBox()).isChecked(),
            include_cga=self.chk_cga.isChecked(),
            include_csa=self.chk_csa.isChecked(),
            include_mrserver=self.chk_mrserver.isChecked(),
            include_gesys=self.chk_gesys.isChecked(),
            include_lais=self.chk_lais.isChecked(),
            include_psc=self.chk_psc.isChecked(),
            include_unknown=self.chk_unknown.isChecked(),
            turbo_csv=self.chk_turbo_csv.isChecked(),
            summary_xlsx_only=self.chk_summary_xlsx.isChecked(),
            fast_xlsx_mode=self.chk_fast_xlsx.isChecked(),
            csv_summary_mode=self.chk_summary_xlsx.isChecked(),
            skip_merged1=self.chk_skip_merged1.isChecked(),
            timestamp_as_text=self.chk_timestamp_text.isChecked(),
            disable_styles_large=self.chk_disable_styles.isChecked(),
            worker_count=int(self.worker_combo.currentData() or 0),
            noise_enabled_merge=self.noise_enable_merge_chk.isChecked(),
            noise_exclude_output=(self.noise_exclude_output_chk.isChecked() or self.noise_enable_merge_chk.isChecked()),
            noise_learning=self.noise_learning_chk.isChecked(),
            plugin_types=[pid for pid, chk in getattr(self, "plugin_checks", {}).items() if chk.isChecked()],
        )

    def validate_options(self, opt: RunOptions) -> Optional[str]:
        if not opt.source_folder:
            return "Source Log Folder is required."
        if not Path(opt.source_folder).exists():
            return "Source Log Folder does not exist."
        if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
            return "Start Date is later than End Date."
        if not any([opt.include_ws, getattr(opt, "include_watersystem", False), opt.include_cga, opt.include_csa, opt.include_mrserver, opt.include_gesys, opt.include_lais, opt.include_psc, opt.include_unknown]):
            return "Select at least one log type."
        return None

    def pause_clicked(self):
        if not self.worker:
            return
        paused = self.pause_btn.text() == "Pause"
        self.worker.set_paused(paused)
        if paused:
            self.pause_btn.setText("Resume")
            self.status_label.setText("Paused")
            self.log("Paused by user.")
        else:
            self.pause_btn.setText("Pause")
            self.status_label.setText("Running...")
            self.log("Resumed by user.")

    def cancel_clicked(self):
        if not self.worker:
            return
        self.worker.request_cancel()
        self.cancel_btn.setEnabled(False)
        self.pause_btn.setEnabled(False)
        self.status_label.setText("Cancelling...")
        self.log("Cancel requested. Current file/write chunk will finish, then the operation will stop.")

    def import_psc_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select PSC log", self.source_edit.text() or str(Path.home()), "PSC Log (*.log *.txt);;All Files (*.*)")
        if not path:
            return
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        self._run_import_task(
            "PSC Import Progress",
            "Importing PSC and writing output...",
            lambda: import_psc_only(path, out_dir),
            "PSC import",
        )

    def import_review_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Review file", self.source_edit.text() or str(Path.home()), "Review files (review.out review.out.ar review.out.*);;All Files (*.*)")
        if not path:
            return
        if not is_review_file(Path(path).name):
            QMessageBox.warning(self, APP_TITLE, "Review Import target must be review.out, review.out.ar, or review.out.*")
            return
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        self._run_import_task(
            "Review Import Progress",
            "Importing review.out and writing output...",
            lambda: import_reviewout_only(path, out_dir),
            "Review import",
        )

    def import_selected_file_clicked(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select one log file to import only",
            self.source_edit.text() or str(Path.home()),
            "Log files (*.log *.txt *.out *.*);;All Files (*.*)",
        )
        if not path:
            return
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        self._run_import_task(
            "Selected File Import Progress",
            "Importing selected file and writing output...",
            lambda: import_selected_file_only(path, out_dir),
            "Selected file import",
        )

    def current_project_dict(self) -> dict[str, Any]:
        return {
            "app_version": APP_VERSION,
            "source_folder": self.source_edit.text(),
            "output_folder": self.output_edit.text(),
            "recursive": self.recursive_chk.isChecked(),
            "use_start": self.use_start.isChecked(),
            "use_end": self.use_end.isChecked(),
            "start_date": self.start_date.date().toString("yyyy-MM-dd"),
            "end_date": self.end_date.date().toString("yyyy-MM-dd"),
            "serial": self.serial_combo.currentText(),
            "site": self.site_combo.currentText(),
            "log_types": {
                "WS": self.chk_ws.isChecked(),
                "CGA": self.chk_cga.isChecked(),
                "CSA": self.chk_csa.isChecked(),
                "MRSERVER": self.chk_mrserver.isChecked(),
                "GESYS": self.chk_gesys.isChecked(),
                "LAIS": self.chk_lais.isChecked(),
                "PSC": self.chk_psc.isChecked(),
                "UNKNOWN": self.chk_unknown.isChecked(),
            },
            "performance": {
                "turbo_csv": self.chk_turbo_csv.isChecked(),
                "summary_xlsx_only": self.chk_summary_xlsx.isChecked(),
                "worker_count": self.worker_combo.currentData(),
            },
            "noise": {
                "enable_merge": self.noise_enable_merge_chk.isChecked(),
                "exclude_output": self.noise_exclude_output_chk.isChecked(),
                "learning": self.noise_learning_chk.isChecked(),
            },
        }

    def apply_project_dict(self, data: dict[str, Any]):
        self.source_edit.setText(str(data.get("source_folder", "")))
        self.output_edit.setText(str(data.get("output_folder", "")))
        self.recursive_chk.setChecked(bool(data.get("recursive", True)))
        self.use_start.setChecked(bool(data.get("use_start", False)))
        self.use_end.setChecked(bool(data.get("use_end", False)))
        try:
            self.start_date.setDate(QDate.fromString(str(data.get("start_date", "")), "yyyy-MM-dd"))
        except Exception:
            pass
        try:
            self.end_date.setDate(QDate.fromString(str(data.get("end_date", "")), "yyyy-MM-dd"))
        except Exception:
            pass
        self.serial_combo.setCurrentText(str(data.get("serial", "")))
        self.site_combo.setCurrentText(str(data.get("site", "")))
        lt = data.get("log_types", {}) or {}
        self.chk_ws.setChecked(bool(lt.get("WS", True)))
        self.chk_cga.setChecked(bool(lt.get("CGA", True)))
        self.chk_csa.setChecked(bool(lt.get("CSA", True)))
        self.chk_mrserver.setChecked(bool(lt.get("MRSERVER", True)))
        self.chk_gesys.setChecked(bool(lt.get("GESYS", True)))
        self.chk_lais.setChecked(bool(lt.get("LAIS", True)))
        self.chk_psc.setChecked(bool(lt.get("PSC", True)))
        self.chk_unknown.setChecked(bool(lt.get("UNKNOWN", False)))
        perf = data.get("performance", {}) or {}
        self.chk_turbo_csv.setChecked(bool(perf.get("turbo_csv", True)))
        self.chk_summary_xlsx.setChecked(bool(perf.get("summary_xlsx_only", True)))
        self.chk_fast_xlsx.setChecked(bool(perf.get("fast_xlsx_mode", True)))
        self.chk_skip_merged1.setChecked(bool(perf.get("skip_merged1", True)))
        self.chk_timestamp_text.setChecked(bool(perf.get("timestamp_as_text", True)))
        self.chk_disable_styles.setChecked(bool(perf.get("disable_styles_large", True)))
        wc = str(perf.get("worker_count", "0"))
        idx = self.worker_combo.findData(wc)
        if idx >= 0:
            self.worker_combo.setCurrentIndex(idx)
        noise = data.get("noise", {}) or {}
        self.noise_enable_merge_chk.setChecked(bool(noise.get("enable_merge", False)))
        self.noise_exclude_output_chk.setChecked(bool(noise.get("exclude_output", False)))
        self.noise_learning_chk.setChecked(bool(noise.get("learning", True)))

    def export_project_clicked(self):
        default_dir = self.output_edit.text().strip() or self.source_edit.text().strip() or str(Path.home())
        default_path = str(Path(default_dir) / f"LogMergeProject_{datetime.now().strftime('%Y%m%d_%H%M%S')}.lmpkg")
        out_path, _ = QFileDialog.getSaveFileName(self, "Export Project", default_path, "Log Merge Project (*.lmpkg);;Zip files (*.zip);;All Files (*.*)")
        if not out_path:
            return
        data = self.current_project_dict()
        try:
            with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as z:
                z.writestr("project.json", json.dumps(data, ensure_ascii=False, indent=2))
                dbp = noise_db_path()
                if dbp.exists():
                    z.write(dbp, "NoiseDB.sqlite")
            QMessageBox.information(self, APP_TITLE, f"Project exported.\n\n{out_path}")
        except Exception:
            text = traceback.format_exc()
            self.log(text)
            QMessageBox.critical(self, APP_TITLE, text)

    def import_project_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Project", str(Path.home()), "Log Merge Project (*.lmpkg *.zip);;All Files (*.*)")
        if not path:
            return
        try:
            with zipfile.ZipFile(path, "r") as z:
                data = json.loads(z.read("project.json").decode("utf-8"))
                self.apply_project_dict(data)
                if "NoiseDB.sqlite" in z.namelist():
                    if QMessageBox.question(self, APP_TITLE, "Import NoiseDB.sqlite from this project and replace current rules?") == QMessageBox.Yes:
                        noise_db_path().write_bytes(z.read("NoiseDB.sqlite"))
            self.save_settings()
            QMessageBox.information(self, APP_TITLE, "Project imported.")
        except Exception:
            text = traceback.format_exc()
            self.log(text)
            QMessageBox.critical(self, APP_TITLE, text)

    def split_merge_clicked(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select merged Search xlsx/csv to split",
            self.output_edit.text().strip() or self.source_edit.text().strip() or str(Path.home()),
            "Merged output (*.xlsx *.csv);;All Files (*.*)",
        )
        if not path:
            return
        default_out = str(Path(path).with_suffix("").with_name(Path(path).stem + "_split"))
        out_dir = QFileDialog.getExistingDirectory(self, "Select split output folder", default_out)
        if not out_dir:
            return
        self._run_import_task(
            "Split Merge Project Progress",
            "Splitting merged output into per-source reconstructed logs...",
            lambda: split_merge_project(path, out_dir, log_cb=self.log),
            "Split merge project",
        )

    def run_clicked(self):
        opt = self.collect_options()
        # v33 Smart File Discovery: log types can be blank before scanning.
        # First validate only the source/date basics, then let the discovery
        # dialog detect log types and let the operator choose final files.
        if not opt.source_folder:
            QMessageBox.warning(self, APP_TITLE, "Source Log Folder is required.")
            return
        if not Path(opt.source_folder).exists():
            QMessageBox.warning(self, APP_TITLE, "Source Log Folder does not exist.")
            return
        if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
            QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
            return

        dlg = SmartFileDiscoveryDialog(self, opt)
        if dlg.exec() != QDialog.Accepted:
            self.status_label.setText("Cancelled before merge")
            return

        selected_types = set(dlg.selected_types)
        opt.include_ws = "WS" in selected_types
        opt.include_watersystem = "WATERSYSTEM" in selected_types
        opt.include_cga = "CGA" in selected_types
        opt.include_csa = "CSA" in selected_types
        opt.include_mrserver = "MRSERVER" in selected_types
        opt.include_gesys = "GESYS" in selected_types
        opt.include_lais = "LAIS" in selected_types
        opt.include_psc = "PSC" in selected_types
        opt.include_unknown = "UNKNOWN" in selected_types
        opt.selected_files = dlg.selected_files
        if not enabled_types(opt):
            QMessageBox.warning(self, APP_TITLE, "No merge-capable log type is selected. Review files are Import Only.")
            return

        self.save_settings()
        self.log_view.clear()
        self.log(f"Smart File Discovery selected {len(opt.selected_files or [])} files.")
        self.progress.setValue(0)
        self.status_label.setText("Running...")
        self.run_btn.setEnabled(False)
        self.pause_btn.setEnabled(True)
        self.cancel_btn.setEnabled(True)
        self.pause_btn.setText("Pause")
        self._show_progress_popup("Log Merge Progress", "Parsing selected files...", 0, 100, cancellable=True)
        self.worker = MergeWorker(opt)
        self.worker.progress.connect(self.on_progress)
        self.worker.log.connect(self.log)
        self.worker.finished_ok.connect(self.on_finished)
        self.worker.failed.connect(self.on_failed)
        self.worker.start()

    def on_progress(self, idx: int, total: int, name: str):
        pct = max(0, min(100, int(idx * 100 / total))) if total else 0
        self.progress.setValue(pct)
        self.progress.setFormat(f"{pct}%")
        self.status_label.setText(f"{pct}%  ({idx}/{total}): {name}")
        if self.progress_dialog is not None:
            self.progress_dialog.setRange(0, 100)
            self.progress_dialog.setValue(pct)
            self.progress_dialog.setLabelText(f"{pct}%  ({idx}/{total})\n{name}")
            QApplication.processEvents()

    def on_finished(self, result: RunResult):
        self._close_progress_popup()
        self.run_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        self.progress.setValue(100)
        self.status_label.setText("Completed")
        self.log("Completed.")
        self.log(f"Output: {result.output_path}")
        if result.csa_error_count > 0:
            self.log(f"CSA errors detected: {result.csa_error_count}")
        # v40: Do not interrupt with a completion message. Open Log Viewer automatically and bring it to front.
        try:
            self.open_dual_viewer()
        except Exception:
            write_startup_log("Auto-open Log Viewer after merge failed.\n\n" + traceback.format_exc())

    def on_failed(self, text: str):
        self._close_progress_popup()
        self.run_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        self.status_label.setText("Failed")
        self.log(text)
        QMessageBox.critical(self, APP_TITLE, text)

    def open_output_folder(self):
        p = self.output_edit.text().strip() or self.source_edit.text().strip()
        if p and Path(p).exists():
            os.startfile(p)  # type: ignore[attr-defined]

    def open_dual_viewer(self):
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        self.viewer_window.show()
        self.viewer_window.raise_()
        self.viewer_window.activateWindow()

    def apply_default_settings(self):
        """Apply factory defaults without clearing saved settings directly."""
        self.source_edit.setText("")
        self.output_edit.setText("")
        self.recursive_chk.setChecked(True)
        self.use_start.setChecked(False)
        self.use_end.setChecked(False)
        self.start_date.setDate(QDate.currentDate())
        self.end_date.setDate(QDate.currentDate())
        self.serial_combo.setCurrentText("")
        self.site_combo.setCurrentText("")
        for chk in [self.chk_ws, self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_review, self.chk_unknown]:
            chk.setChecked(False)
        for chk in getattr(self, "plugin_checks", {}).values():
            chk.setChecked(False)
        # Recommended fastest/stablest defaults.
        self.chk_turbo_csv.setChecked(True)
        self.chk_summary_xlsx.setChecked(True)
        self.chk_fast_xlsx.setChecked(True)
        self.chk_skip_merged1.setChecked(True)
        self.chk_timestamp_text.setChecked(True)
        self.chk_disable_styles.setChecked(True)
        idx = self.worker_combo.findData("0")
        if idx >= 0:
            self.worker_combo.setCurrentIndex(idx)
        # Smart Noise defaults: both visible options ON.
        self.noise_enable_merge_chk.setChecked(True)
        self.noise_exclude_output_chk.setChecked(False)
        self.noise_learning_chk.setChecked(True)

    def reset_to_defaults_clicked(self):
        reply = QMessageBox.question(
            self,
            APP_TITLE,
            "Reset UI settings to defaults?\n\nFolders, date range, site information, and log type selections will be cleared.\nPerformance will use the recommended fastest settings.\nSmart Noise options will be turned ON.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.apply_default_settings()
        self.save_settings()
        self.status_label.setText("Defaults restored")
        self.log("Settings were reset to defaults.")

    def manage_noise_rules_main(self):
        dlg = NoiseRuleManagerDialog(self)
        dlg.exec()

    def load_settings(self):
        # Restore the previous user state. If no previous state exists, use factory defaults.
        self.source_edit.setText(self.settings.value("source_folder", ""))
        self.output_edit.setText(self.settings.value("output_folder", ""))
        self.recursive_chk.setChecked(self.settings.value("recursive", "true") == "true")
        self.use_start.setChecked(self.settings.value("use_start_date", "false") == "true")
        self.use_end.setChecked(self.settings.value("use_end_date", "false") == "true")
        sd = self.settings.value("start_date", "")
        ed = self.settings.value("end_date", "")
        if sd:
            self.start_date.setDate(QDate.fromString(sd, "yyyy-MM-dd"))
        if ed:
            self.end_date.setDate(QDate.fromString(ed, "yyyy-MM-dd"))
        self.serial_combo.setCurrentText(self.settings.value("serial", ""))
        self.site_combo.setCurrentText(self.settings.value("site", ""))
        self.chk_ws.setChecked(self.settings.value("log_ws", "false") == "true")
        self.chk_cga.setChecked(self.settings.value("log_cga", "false") == "true")
        self.chk_csa.setChecked(self.settings.value("log_csa", "false") == "true")
        self.chk_mrserver.setChecked(self.settings.value("log_mrserver", "false") == "true")
        self.chk_gesys.setChecked(self.settings.value("log_gesys", "false") == "true")
        self.chk_lais.setChecked(self.settings.value("log_lais", "false") == "true")
        self.chk_psc.setChecked(self.settings.value("log_psc", "false") == "true")
        self.chk_review.setChecked(self.settings.value("log_review", "false") == "true")
        self.chk_unknown.setChecked(self.settings.value("log_unknown", "false") == "true")
        saved_plugins = set(str(self.settings.value("log_plugins", "")).split(","))
        for pid, chk in getattr(self, "plugin_checks", {}).items():
            chk.setChecked(pid in saved_plugins)
        self.chk_turbo_csv.setChecked(self.settings.value("turbo_csv", "true") == "true")
        self.chk_summary_xlsx.setChecked(self.settings.value("summary_xlsx", "true") == "true")
        self.chk_fast_xlsx.setChecked(self.settings.value("fast_xlsx", "true") == "true")
        self.chk_skip_merged1.setChecked(self.settings.value("skip_merged1", "true") == "true")
        self.chk_timestamp_text.setChecked(self.settings.value("timestamp_text", "true") == "true")
        self.chk_disable_styles.setChecked(self.settings.value("disable_styles", "true") == "true")
        if hasattr(self, "noise_enable_merge_chk"):
            self.noise_enable_merge_chk.setChecked(self.settings.value("noise_enable_merge", "true") == "true")
            self.noise_exclude_output_chk.setChecked(self.settings.value("noise_exclude_output", "false") == "true")
            self.noise_learning_chk.setChecked(self.settings.value("noise_learning", "true") == "true")
        worker_value = self.settings.value("worker_count", "0")
        idx = self.worker_combo.findData(str(worker_value))
        if idx >= 0:
            self.worker_combo.setCurrentIndex(idx)

    def save_settings(self):
        self.settings.setValue("source_folder", self.source_edit.text())
        self.settings.setValue("output_folder", self.output_edit.text())
        self.settings.setValue("recursive", "true" if self.recursive_chk.isChecked() else "false")
        self.settings.setValue("use_start_date", "true" if self.use_start.isChecked() else "false")
        self.settings.setValue("use_end_date", "true" if self.use_end.isChecked() else "false")
        self.settings.setValue("start_date", self.start_date.date().toString("yyyy-MM-dd"))
        self.settings.setValue("end_date", self.end_date.date().toString("yyyy-MM-dd"))
        self.settings.setValue("noise_enable_merge", "true" if self.noise_enable_merge_chk.isChecked() else "false")
        self.settings.setValue("noise_exclude_output", "true" if self.noise_exclude_output_chk.isChecked() else "false")
        self.settings.setValue("noise_learning", "true" if self.noise_learning_chk.isChecked() else "false")
        self.settings.setValue("serial", self.serial_combo.currentText())
        self.settings.setValue("site", self.site_combo.currentText())
        self.settings.setValue("log_ws", "true" if self.chk_ws.isChecked() else "false")
        self.settings.setValue("log_cga", "true" if self.chk_cga.isChecked() else "false")
        self.settings.setValue("log_csa", "true" if self.chk_csa.isChecked() else "false")
        self.settings.setValue("log_mrserver", "true" if self.chk_mrserver.isChecked() else "false")
        self.settings.setValue("log_gesys", "true" if self.chk_gesys.isChecked() else "false")
        self.settings.setValue("log_lais", "true" if self.chk_lais.isChecked() else "false")
        self.settings.setValue("log_psc", "true" if self.chk_psc.isChecked() else "false")
        self.settings.setValue("log_review", "true" if self.chk_review.isChecked() else "false")
        self.settings.setValue("log_unknown", "true" if self.chk_unknown.isChecked() else "false")
        self.settings.setValue("log_plugins", ",".join([pid for pid, chk in getattr(self, "plugin_checks", {}).items() if chk.isChecked()]))
        self.settings.setValue("turbo_csv", "true" if self.chk_turbo_csv.isChecked() else "false")
        self.settings.setValue("summary_xlsx", "true" if self.chk_summary_xlsx.isChecked() else "false")
        self.settings.setValue("fast_xlsx", "true" if self.chk_fast_xlsx.isChecked() else "false")
        self.settings.setValue("skip_merged1", "true" if self.chk_skip_merged1.isChecked() else "false")
        self.settings.setValue("timestamp_text", "true" if self.chk_timestamp_text.isChecked() else "false")
        self.settings.setValue("disable_styles", "true" if self.chk_disable_styles.isChecked() else "false")
        self.settings.setValue("worker_count", str(self.worker_combo.currentData() or "0"))
        self.settings.sync()


# ---------------------------------------------------------------------------
# v40 Professional UI Refresh overrides
# ---------------------------------------------------------------------------
class SoftProgressDialog(QProgressDialog):
    def __init__(self, title: str, label: str, cancel_text: Optional[str], minimum: int, maximum: int, parent=None):
        super().__init__(label, cancel_text, minimum, maximum, parent)
        self.setWindowTitle(title)
        self.setMinimumDuration(250)
        self.resize(720, 260)
        self.setAutoClose(False)
        self.setAutoReset(False)
        self.setWindowModality(Qt.WindowModal)
        self.setStyleSheet("""
            QProgressDialog { background: #f7f9fc; font-size: 14pt; border-radius: 14px; }
            QLabel { font-size: 14pt; padding: 12px; color: #1f2937; }
            QProgressBar { min-height: 30px; border: 1px solid #c7d2fe; border-radius: 12px; text-align: center; background: #eef2ff; font-size: 13pt; }
            QProgressBar::chunk { border-radius: 12px; background: #60a5fa; }
            QPushButton { min-width: 96px; padding: 8px 18px; border-radius: 10px; background: #ffffff; border: 1px solid #cbd5e1; font-size: 12pt; }
            QPushButton:hover { background: #eff6ff; }
        """)

def _v40_show_progress_popup(self, title: str, label: str, minimum: int = 0, maximum: int = 100, cancellable: bool = True):
    self._close_progress_popup()
    dlg = SoftProgressDialog(title, label, "Cancel" if cancellable else None, minimum, maximum, self)
    if cancellable:
        dlg.canceled.connect(self.cancel_clicked)
    dlg.show()
    QApplication.processEvents()
    self.progress_dialog = dlg
    return dlg

def _v40_update_clicked(self):
    fn, _ = QFileDialog.getOpenFileName(self, "Select Update ZIP", "", "Update ZIP (*.zip);;All files (*.*)")
    if not fn:
        return
    z = Path(fn)
    try:
        with zipfile.ZipFile(z, "r") as zf:
            if "manifest.json" not in set(zf.namelist()):
                QMessageBox.critical(self, "Update", "manifest.json is missing.")
                return
            manifest = json.loads(zf.read("manifest.json").decode("utf-8-sig"))
            ptype = str(manifest.get("package_type", "")).lower()
            if ptype in {"core_update", "tool_update"}:
                dlg = CoreUpdateDialog(self)
                # Reuse the staging code through the dialog UI when direct staging is not exposed.
                QMessageBox.information(self, "Tool Update", "This ZIP is a Tool/Core update. Open the Tool Update dialog and select the same ZIP to stage it safely.")
                dlg.exec()
                return
            ok, msg = install_plugin_zip(z)
            if ok:
                self.refresh_plugin_checkboxes()
                QMessageBox.information(self, "File Type Update", msg)
                self.log("File Type plugin installed/reloaded: " + z.name)
            else:
                QMessageBox.critical(self, "File Type Update Failed", msg)
    except Exception as exc:
        QMessageBox.critical(self, "Update Failed", str(exc))

def _v40_import_clicked(self):
    items = ["Import selected file", "Import PSC", "Import Review"]
    item, ok = QInputDialog.getItem(self, "Import", "Select import mode", items, 0, False)
    if not ok:
        return
    if item.startswith("Import selected"):
        self.import_selected_file_clicked()
    elif item.endswith("PSC"):
        self.import_psc_clicked()
    else:
        self.import_review_clicked()

def _v40_main_build_ui(self):
    root = QVBoxLayout(self)
    self.setMinimumWidth(980)
    title = QLabel(f"{APP_TITLE}")
    title.setStyleSheet("font-size: 22px; font-weight: 700; color:#0f172a;")
    root.addWidget(title)
    subtitle = QLabel("RC1 Foundation Fix: stable Viewer layout, resizable columns, editable Copy Rule Text, shared Site/Serial master.")
    subtitle.setStyleSheet("color:#475569;")
    root.addWidget(subtitle)

    action_bar = QHBoxLayout()
    self.run_btn = QPushButton("▶ START")
    self.run_btn.setMinimumSize(150, 48)
    self.run_btn.setStyleSheet("QPushButton{font-size:16pt;font-weight:700;background:#16a34a;color:white;border-radius:12px;padding:8px 22px;} QPushButton:hover{background:#15803d;}")
    self.run_btn.clicked.connect(self.run_clicked)
    self.import_btn = QPushButton("Import")
    self.import_btn.setMinimumHeight(40)
    self.import_btn.clicked.connect(self.import_clicked)
    self.viewer_btn = QPushButton("Log Viewer")
    self.viewer_btn.setMinimumHeight(40)
    self.viewer_btn.clicked.connect(self.open_dual_viewer)
    self.update_btn = QPushButton("Update...")
    self.update_btn.setMinimumHeight(40)
    self.update_btn.setToolTip("Select either a File Type plugin ZIP or a Tool/Core update ZIP. The package type is detected from manifest.json.")
    self.update_btn.clicked.connect(self.update_clicked)
    self.pause_btn = QPushButton("Pause"); self.pause_btn.clicked.connect(self.pause_clicked); self.pause_btn.setEnabled(False)
    self.cancel_btn = QPushButton("Cancel"); self.cancel_btn.clicked.connect(self.cancel_clicked); self.cancel_btn.setEnabled(False)
    action_bar.addWidget(self.run_btn)
    action_bar.addWidget(self.import_btn)
    action_bar.addWidget(self.viewer_btn)
    action_bar.addWidget(self.update_btn)
    action_bar.addStretch(1)
    action_bar.addWidget(self.pause_btn)
    action_bar.addWidget(self.cancel_btn)
    root.addLayout(action_bar)

    folder_box = QGroupBox("Folders")
    folder_layout = QFormLayout(folder_box)
    self.source_edit = QLineEdit(); self.source_btn = QPushButton("Browse..."); self.source_btn.clicked.connect(self.pick_source)
    h1 = QHBoxLayout(); h1.addWidget(self.source_edit); h1.addWidget(self.source_btn)
    folder_layout.addRow("Source", h1)
    self.output_edit = QLineEdit(); self.output_btn = QPushButton("Browse..."); self.output_btn.clicked.connect(self.pick_output)
    self.output_same_btn = QPushButton("Same as Source"); self.output_same_btn.clicked.connect(lambda: self.output_edit.setText(self.source_edit.text().strip()))
    h2 = QHBoxLayout(); h2.addWidget(self.output_edit); h2.addWidget(self.output_btn); h2.addWidget(self.output_same_btn)
    folder_layout.addRow("Output", h2)
    self.recursive_chk = QCheckBox("Include all subfolders recursively"); self.recursive_chk.setChecked(True)
    folder_layout.addRow("", self.recursive_chk)
    root.addWidget(folder_box)

    date_box = QGroupBox("Date Range")
    date_layout = QFormLayout(date_box)
    self.use_start = QCheckBox("Use Start")
    self.start_date = QDateEdit(QDate.currentDate()); self.start_date.setCalendarPopup(True); self.start_date.setDisplayFormat("yyyy/MM/dd")
    hs = QHBoxLayout(); hs.addWidget(self.use_start); hs.addWidget(self.start_date); hs.addStretch(1)
    date_layout.addRow("Start", hs)
    self.use_end = QCheckBox("Use End")
    self.end_date = QDateEdit(QDate.currentDate()); self.end_date.setCalendarPopup(True); self.end_date.setDisplayFormat("yyyy/MM/dd")
    he = QHBoxLayout(); he.addWidget(self.use_end); he.addWidget(self.end_date); he.addStretch(1)
    date_layout.addRow("End", he)
    root.addWidget(date_box)

    meta_box = QGroupBox("Site Information")
    meta_layout = QFormLayout(meta_box)
    self.serial_combo = QComboBox(); self.serial_combo.setEditable(True)
    self.site_combo = QComboBox(); self.site_combo.setEditable(True)
    self.site_map = load_shared_site_map(app_dir() / "site_serial_map.json", DEFAULT_SITE_MAP)
    self.serial_combo.addItem(""); self.site_combo.addItem("")
    for serial in sorted({str(item.get("serial", "")).strip() for item in self.site_map if str(item.get("serial", "")).strip()}): self.serial_combo.addItem(serial)
    for site in sorted({str(item.get("site", "")).strip() for item in self.site_map if str(item.get("site", "")).strip()}): self.site_combo.addItem(site)
    self.serial_combo.setToolTip(f"Shared Service Hub master: {shared_master_path()}")
    self.site_combo.setToolTip(f"Shared Service Hub master: {shared_master_path()}")
    self.serial_combo.currentTextChanged.connect(self.sync_site_from_serial)
    self.site_combo.currentTextChanged.connect(self.sync_serial_from_site)
    meta_layout.addRow("Serial", self.serial_combo); meta_layout.addRow("Site", self.site_combo)
    root.addWidget(meta_box)

    type_box = QGroupBox("Log Types")
    type_layout = QHBoxLayout(type_box); self.type_layout = type_layout; self.plugin_checks = {}
    self.chk_ws = QCheckBox("WS"); self.chk_watersystem = QCheckBox("WaterSystem")
    self.chk_cga = QCheckBox("CGA"); self.chk_csa = QCheckBox("CSA"); self.chk_mrserver = QCheckBox("MRSERVER")
    self.chk_gesys = QCheckBox("GESYS"); self.chk_lais = QCheckBox("LAIS"); self.chk_psc = QCheckBox("PSC")
    self.chk_review = QCheckBox("Review (Import Only)"); self.chk_review.setEnabled(False)
    self.chk_unknown = QCheckBox("UNKNOWN")
    for w in [self.chk_ws, self.chk_watersystem, self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_review, self.chk_unknown]:
        w.setChecked(False); type_layout.addWidget(w)
    self.logtypes_all_btn = QPushButton("All"); self.logtypes_none_btn = QPushButton("Clear")
    self.logtypes_all_btn.setMaximumWidth(54); self.logtypes_none_btn.setMaximumWidth(64)
    self.logtypes_all_btn.clicked.connect(lambda: self.set_all_log_types(True)); self.logtypes_none_btn.clicked.connect(lambda: self.set_all_log_types(False))
    type_layout.addWidget(self.logtypes_all_btn); type_layout.addWidget(self.logtypes_none_btn); type_layout.addStretch(1)
    root.addWidget(type_box)
    self.refresh_plugin_checkboxes()

    perf_box = QGroupBox("Performance")
    perf_layout = QHBoxLayout(perf_box)
    self.chk_turbo_csv = QCheckBox("CSV for large data"); self.chk_turbo_csv.setChecked(True)
    self.chk_summary_xlsx = QCheckBox("Summary xlsx"); self.chk_summary_xlsx.setChecked(True)
    self.chk_fast_xlsx = QCheckBox("Fast xlsx"); self.chk_fast_xlsx.setChecked(True)
    self.chk_skip_merged1 = QCheckBox("Skip Merged_1"); self.chk_skip_merged1.setChecked(True)
    self.chk_timestamp_text = QCheckBox("Timestamp text"); self.chk_timestamp_text.setChecked(True)
    self.chk_disable_styles = QCheckBox("No styles"); self.chk_disable_styles.setChecked(True)
    self.worker_combo = QComboBox(); self.worker_combo.addItem("Auto", "0")
    for n in [2,4,6,8,12,16]: self.worker_combo.addItem(str(n), str(n))
    for w in [self.chk_turbo_csv,self.chk_summary_xlsx,self.chk_fast_xlsx,self.chk_skip_merged1,self.chk_timestamp_text,self.chk_disable_styles,QLabel("Workers"),self.worker_combo]: perf_layout.addWidget(w)
    root.addWidget(perf_box)

    noise_group = QGroupBox("Smart Noise")
    noise_layout = QHBoxLayout(noise_group)
    self.noise_enable_merge_chk = QCheckBox("Apply approved rules in Merge"); self.noise_enable_merge_chk.setChecked(False)
    self.noise_exclude_output_chk = QCheckBox("Legacy exclude"); self.noise_exclude_output_chk.setVisible(False)
    self.noise_learning_chk = QCheckBox("Learning Mode"); self.noise_learning_chk.setChecked(False)
    self.noise_manage_btn = QPushButton("Manage Rules"); self.noise_manage_btn.clicked.connect(self.manage_noise_rules_main)
    for w in [self.noise_enable_merge_chk,self.noise_exclude_output_chk,self.noise_learning_chk,self.noise_manage_btn]: noise_layout.addWidget(w)
    noise_layout.addStretch(1); root.addWidget(noise_group)

    project_bar = QHBoxLayout()
    self.export_project_btn = QPushButton("Export Project"); self.export_project_btn.clicked.connect(self.export_project_clicked)
    self.import_project_btn = QPushButton("Import Project"); self.import_project_btn.clicked.connect(self.import_project_clicked)
    self.split_merge_btn = QPushButton("Split Merge"); self.split_merge_btn.clicked.connect(self.split_merge_clicked)
    self.open_out_btn = QPushButton("Open Output"); self.open_out_btn.clicked.connect(self.open_output_folder)
    self.reset_defaults_btn = QPushButton("Reset Defaults"); self.reset_defaults_btn.clicked.connect(self.reset_to_defaults_clicked)
    for w in [self.export_project_btn,self.import_project_btn,self.split_merge_btn,self.open_out_btn,self.reset_defaults_btn]: project_bar.addWidget(w)
    project_bar.addStretch(1); root.addLayout(project_bar)

    # Hidden compatibility buttons for older callbacks/tests.
    self.import_psc_btn = QPushButton("Import PSC Only"); self.import_psc_btn.setVisible(False)
    self.import_review_btn = QPushButton("Import Review Only"); self.import_review_btn.setVisible(False)
    self.import_file_btn = QPushButton("Import Selected File Only"); self.import_file_btn.setVisible(False)
    self.update_file_type_btn = QPushButton("Update File Type"); self.update_file_type_btn.setVisible(False)
    self.update_tool_btn = QPushButton("Update Tool"); self.update_tool_btn.setVisible(False)

    self.progress = QProgressBar(); self.progress.setRange(0,100); root.addWidget(self.progress)
    self.status_label = QLabel("Ready"); root.addWidget(self.status_label)
    self.log_view = QTextEdit(); self.log_view.setReadOnly(True); root.addWidget(self.log_view, stretch=1)
    self.viewer_window = None

def _v40_set_all_log_types(self, checked: bool):
    for chk in [self.chk_ws, getattr(self, "chk_watersystem", None), self.chk_cga, self.chk_csa, self.chk_mrserver, self.chk_gesys, self.chk_lais, self.chk_psc, self.chk_unknown]:
        if chk is not None:
            chk.setChecked(checked)
    for chk in getattr(self, "plugin_checks", {}).values():
        chk.setChecked(checked)

def _v40_load_settings(self):
    # Keep previous folders/settings restoration, but default Smart Noise OFF for v40.
    try:
        self.source_edit.setText(self.settings.value("source_folder", ""))
        self.output_edit.setText(self.settings.value("output_folder", ""))
        self.recursive_chk.setChecked(self.settings.value("recursive", "true") == "true")
        self.use_start.setChecked(self.settings.value("use_start_date", "false") == "true")
        self.use_end.setChecked(self.settings.value("use_end_date", "false") == "true")
        sd = self.settings.value("start_date", "")
        ed = self.settings.value("end_date", "")
        if sd: self.start_date.setDate(QDate.fromString(str(sd), "yyyy-MM-dd"))
        if ed: self.end_date.setDate(QDate.fromString(str(ed), "yyyy-MM-dd"))
        self.serial_combo.setCurrentText(self.settings.value("serial", ""))
        self.site_combo.setCurrentText(self.settings.value("site", ""))
        self.chk_ws.setChecked(self.settings.value("log_ws", "false") == "true")
        self.chk_watersystem.setChecked(self.settings.value("log_watersystem", "false") == "true")
        self.chk_cga.setChecked(self.settings.value("log_cga", "false") == "true")
        self.chk_csa.setChecked(self.settings.value("log_csa", "false") == "true")
        self.chk_mrserver.setChecked(self.settings.value("log_mrserver", "false") == "true")
        self.chk_gesys.setChecked(self.settings.value("log_gesys", "false") == "true")
        self.chk_lais.setChecked(self.settings.value("log_lais", "false") == "true")
        self.chk_psc.setChecked(self.settings.value("log_psc", "false") == "true")
        self.chk_unknown.setChecked(self.settings.value("log_unknown", "false") == "true")
        self.noise_enable_merge_chk.setChecked(self.settings.value("noise_enable_merge", "false") == "true")
        self.noise_learning_chk.setChecked(self.settings.value("noise_learning", "false") == "true")
    except Exception:
        write_startup_log("v40 load_settings failed.\n\n" + traceback.format_exc())

def _v40_save_settings(self):
    self.settings.setValue("source_folder", self.source_edit.text())
    self.settings.setValue("output_folder", self.output_edit.text())
    self.settings.setValue("recursive", "true" if self.recursive_chk.isChecked() else "false")
    self.settings.setValue("use_start_date", "true" if self.use_start.isChecked() else "false")
    self.settings.setValue("use_end_date", "true" if self.use_end.isChecked() else "false")
    self.settings.setValue("start_date", self.start_date.date().toString("yyyy-MM-dd"))
    self.settings.setValue("end_date", self.end_date.date().toString("yyyy-MM-dd"))
    self.settings.setValue("serial", self.serial_combo.currentText())
    self.settings.setValue("site", self.site_combo.currentText())
    self.settings.setValue("log_ws", "true" if self.chk_ws.isChecked() else "false")
    self.settings.setValue("log_watersystem", "true" if self.chk_watersystem.isChecked() else "false")
    self.settings.setValue("log_cga", "true" if self.chk_cga.isChecked() else "false")
    self.settings.setValue("log_csa", "true" if self.chk_csa.isChecked() else "false")
    self.settings.setValue("log_mrserver", "true" if self.chk_mrserver.isChecked() else "false")
    self.settings.setValue("log_gesys", "true" if self.chk_gesys.isChecked() else "false")
    self.settings.setValue("log_lais", "true" if self.chk_lais.isChecked() else "false")
    self.settings.setValue("log_psc", "true" if self.chk_psc.isChecked() else "false")
    self.settings.setValue("log_unknown", "true" if self.chk_unknown.isChecked() else "false")
    self.settings.setValue("noise_enable_merge", "true" if self.noise_enable_merge_chk.isChecked() else "false")
    self.settings.setValue("noise_learning", "true" if self.noise_learning_chk.isChecked() else "false")
    self.settings.sync()

def _v40_open_dual_viewer(self):
    try:
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        geo = QApplication.primaryScreen().availableGeometry()
        self.viewer_window.resize(max(900, geo.width() // 3), max(520, geo.height() // 3))
        self.viewer_window.show()
        self.viewer_window.raise_()
        self.viewer_window.activateWindow()
    except Exception:
        QMessageBox.critical(self, "Log Viewer", "Viewer failed to open.\n\n" + traceback.format_exc())

# Patch MainWindow methods.
MainWindow.build_ui = _v40_main_build_ui
MainWindow._show_progress_popup = _v40_show_progress_popup
MainWindow.update_clicked = _v40_update_clicked
MainWindow.import_clicked = _v40_import_clicked
MainWindow.set_all_log_types = _v40_set_all_log_types
MainWindow.load_settings = _v40_load_settings
MainWindow.save_settings = _v40_save_settings
MainWindow.open_dual_viewer = _v40_open_dual_viewer

# Viewer source mapping for WaterSystem display label.
_old_source_to_records = MultiPaneLogViewer.source_to_records
def _v40_source_to_records(self, source, progress=None):
    if source == "WaterSystem":
        return _old_source_to_records(self, "WATERSYSTEM", progress)
    return _old_source_to_records(self, source, progress)
MultiPaneLogViewer.source_to_records = _v40_source_to_records



# v40 replacement Smart File Discovery dialog: calendar date inputs, automatic filtering,
# selection preservation, All/Clear controls, and a prominent START button.
class SmartFileDiscoveryDialog(QDialog):
    def __init__(self, parent, base_options: RunOptions):
        super().__init__(parent)
        self.setWindowTitle("Smart File Discovery")
        self.resize(1120, 760)
        self.base_options = base_options
        self.discovered: list[DiscoveredFile] = []
        self.filtered: list[DiscoveredFile] = []
        self.selected_files: list[str] = []
        self.selected_types: set[str] = set()
        self.file_selection: dict[str, bool] = {}
        self._updating = False
        self._build_ui()
        self._scan()
        try:
            screen = self.parent().windowHandle().screen() if self.parent() and self.parent().windowHandle() else self.screen()
            geo = screen.availableGeometry() if screen else QApplication.primaryScreen().availableGeometry()
            self.resize(max(760, geo.width() - 16), max(520, geo.height() - 16))
            self.move(geo.x() + 8, geo.y() + 8)
        except Exception:
            pass

    def _build_ui(self):
        root = QVBoxLayout(self)
        info = QLabel("Detected file types and files are filtered using parsed log timestamps. File created/modified dates are not used.")
        info.setWordWrap(True)
        root.addWidget(info)

        type_box = QGroupBox("Detected Log Types")
        self.type_layout = QHBoxLayout(type_box)
        self.type_checks: dict[str, QCheckBox] = {}
        self.detected_all_btn = QPushButton("All")
        self.detected_clear_btn = QPushButton("Clear")
        self.detected_all_btn.clicked.connect(lambda: self.set_all_detected_types(True))
        self.detected_clear_btn.clicked.connect(lambda: self.set_all_detected_types(False))
        root.addWidget(type_box)

        range_box = QGroupBox("Actual Data Range from Checked Files")
        rg = QHBoxLayout(range_box)
        rg.addWidget(QLabel("Start"))
        self.start_combo = QComboBox()
        self.start_combo.setMinimumWidth(245)
        rg.addWidget(self.start_combo)
        rg.addWidget(QLabel("End"))
        self.end_combo = QComboBox()
        self.end_combo.setMinimumWidth(245)
        rg.addWidget(self.end_combo)
        self.custom_date_btn = QPushButton("Custom Calendar...")
        rg.addWidget(self.custom_date_btn)
        rg.addStretch(1)
        root.addWidget(range_box)

        # Calendar controls remain available only as a Custom fallback.
        self.enable_start = QCheckBox("Start")
        self.start_date = QDateEdit(QDate.currentDate()); self.start_date.setCalendarPopup(True); self.start_date.setDisplayFormat("yyyy/MM/dd")
        self.enable_end = QCheckBox("End")
        self.end_date = QDateEdit(QDate.currentDate()); self.end_date.setCalendarPopup(True); self.end_date.setDisplayFormat("yyyy/MM/dd")
        self.enable_start.hide(); self.start_date.hide(); self.enable_end.hide(); self.end_date.hide()
        self.start_combo.currentIndexChanged.connect(self.apply_filter)
        self.end_combo.currentIndexChanged.connect(self.apply_filter)
        self.custom_date_btn.clicked.connect(self._choose_custom_date_range)

        self.summary_label = QLabel("Ready")
        root.addWidget(self.summary_label)
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["Use", "Type", "Start", "End", "Rows", "Size MB", "File", "Error"])
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.table.itemChanged.connect(self._file_item_changed)
        root.addWidget(self.table, stretch=1)

        btns = QHBoxLayout()
        self.select_all_btn = QPushButton("All Files")
        self.select_all_btn.clicked.connect(lambda: self.set_all_files(True))
        self.clear_all_btn = QPushButton("Clear Files")
        self.clear_all_btn.clicked.connect(lambda: self.set_all_files(False))
        self.start_btn = QPushButton("▶ START")
        self.start_btn.setMinimumSize(150, 46)
        self.start_btn.setStyleSheet("QPushButton{font-size:15pt;font-weight:700;background:#16a34a;color:white;border-radius:12px;padding:8px 22px;} QPushButton:hover{background:#15803d;}")
        self.start_btn.clicked.connect(self.accept_with_selection)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.start_btn)
        btns.addWidget(self.select_all_btn)
        btns.addWidget(self.clear_all_btn)
        btns.addStretch(1)
        btns.addWidget(self.cancel_btn)
        root.addLayout(btns)

    def _scan(self):
        src = Path(self.base_options.source_folder)
        all_files = list(iter_files(src, self.base_options.recursive))
        progress = SoftProgressDialog("Smart File Discovery", "Scanning and parsing candidate files...", "Cancel", 0, max(1, len(all_files)), self)
        progress.show()
        QApplication.processEvents()
        discovered = []
        scan_started = time.monotonic()
        total_files = max(1, len(all_files))
        for i, path in enumerate(all_files, start=1):
            if progress.wasCanceled():
                break
            progress.setValue(i - 1)
            elapsed = max(0.001, time.monotonic() - scan_started)
            rate = i / elapsed
            remaining = max(0.0, (total_files - i) / rate) if rate > 0 else 0.0
            progress.setLabelText(
                f"Scanning and parsing files\n"
                f"{i:,} / {len(all_files):,}  ({int(i * 100 / total_files)}%)\n"
                f"Current: {path.name}\n"
                f"Elapsed: {int(elapsed // 60):02d}:{int(elapsed % 60):02d}   "
                f"Estimated remaining: {int(remaining // 60):02d}:{int(remaining % 60):02d}"
            )
            QApplication.processEvents()
            meta = discover_one_file(path, self.base_options.include_unknown)
            if meta.log_type:
                discovered.append(meta)
                self.file_selection.setdefault(meta.path, True)
        progress.setValue(len(all_files))
        progress.close()
        self.discovered = discovered
        self.populate_type_checks()
        self._set_detected_range_defaults()
        self.apply_filter()

    def _selected_boundary_files(self) -> list[DiscoveredFile]:
        candidates = self.filtered if self.filtered else self.discovered
        selected = [d for d in candidates if self.file_selection.get(d.path, True)]
        return selected or candidates

    def _refresh_boundary_combos(self):
        files = self._selected_boundary_files()
        starts = sorted({d.start_ts for d in files if d.start_ts})
        ends = sorted({d.end_ts for d in files if d.end_ts})
        old_start = self.start_combo.currentData() if self.start_combo.count() else None
        old_end = self.end_combo.currentData() if self.end_combo.count() else None
        self.start_combo.blockSignals(True); self.end_combo.blockSignals(True)
        self.start_combo.clear(); self.end_combo.clear()
        self.start_combo.addItem("No start limit", None)
        self.end_combo.addItem("No end limit", None)
        if starts:
            self.start_combo.addItem("Earliest checked data — " + format_timestamp(starts[0]), starts[0])
            for dt in starts:
                self.start_combo.addItem(format_timestamp(dt), dt)
        if ends:
            self.end_combo.addItem("Latest checked data — " + format_timestamp(ends[-1]), ends[-1])
            for dt in ends:
                self.end_combo.addItem(format_timestamp(dt), dt)
        # Default to the full real range; retain prior boundary when possible.
        start_target = old_start if old_start in starts else (starts[0] if starts else None)
        end_target = old_end if old_end in ends else (ends[-1] if ends else None)
        start_index = self.start_combo.findData(start_target)
        end_index = self.end_combo.findData(end_target)
        self.start_combo.setCurrentIndex(max(0, start_index))
        self.end_combo.setCurrentIndex(max(0, end_index))
        self.start_combo.blockSignals(False); self.end_combo.blockSignals(False)

    def _set_detected_range_defaults(self):
        self._refresh_boundary_combos()

    def _choose_custom_date_range(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Custom Date Range")
        lay = QFormLayout(dlg)
        start_check = QCheckBox("Enable custom start")
        start_edit = QDateEdit(self.start_date.date()); start_edit.setCalendarPopup(True); start_edit.setDisplayFormat("yyyy/MM/dd")
        end_check = QCheckBox("Enable custom end")
        end_edit = QDateEdit(self.end_date.date()); end_edit.setCalendarPopup(True); end_edit.setDisplayFormat("yyyy/MM/dd")
        lay.addRow(start_check, start_edit); lay.addRow(end_check, end_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addRow(buttons); buttons.accepted.connect(dlg.accept); buttons.rejected.connect(dlg.reject)
        if dlg.exec() == QDialog.Accepted:
            self.enable_start.setChecked(start_check.isChecked()); self.start_date.setDate(start_edit.date())
            self.enable_end.setChecked(end_check.isChecked()); self.end_date.setDate(end_edit.date())
            self.start_combo.blockSignals(True); self.end_combo.blockSignals(True)
            self.start_combo.setCurrentIndex(0); self.end_combo.setCurrentIndex(0)
            self.start_combo.blockSignals(False); self.end_combo.blockSignals(False)
            self.apply_filter()

    def populate_type_checks(self):
        while self.type_layout.count():
            item = self.type_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self.type_checks.clear()
        counts: dict[str, int] = {}
        for d in self.discovered:
            counts[d.log_type] = counts.get(d.log_type, 0) + 1
        preferred = ["WS", "WATERSYSTEM", "VIMEASURE", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "REVIEW", "UNKNOWN"]
        for typ in [t for t in preferred if t in counts] + sorted(t for t in counts if t not in preferred):
            chk = QCheckBox(f"{typ} ({counts[typ]})")
            chk.setChecked(False)
            chk.stateChanged.connect(self.apply_filter)
            self.type_checks[typ] = chk
            self.type_layout.addWidget(chk)
        self.type_layout.addStretch(1)
        self.type_layout.addWidget(self.detected_all_btn)
        self.type_layout.addWidget(self.detected_clear_btn)

    def set_all_detected_types(self, checked: bool):
        for chk in self.type_checks.values():
            chk.blockSignals(True); chk.setChecked(checked); chk.blockSignals(False)
        self.apply_filter()

    def _range(self) -> tuple[Optional[datetime], Optional[datetime]]:
        start = self.start_combo.currentData() if hasattr(self, "start_combo") else None
        end = self.end_combo.currentData() if hasattr(self, "end_combo") else None
        # Custom calendar values override dropdowns only when enabled.
        if self.enable_start.isChecked():
            qd = self.start_date.date(); start = datetime(qd.year(), qd.month(), qd.day(), 0, 0, 0)
        if self.enable_end.isChecked():
            qd = self.end_date.date(); end = datetime(qd.year(), qd.month(), qd.day(), 23, 59, 59, 999999)
        if start and end and start > end:
            raise ValueError("Start is later than End.")
        return start, end

    def _overlaps_range(self, d: DiscoveredFile, start: Optional[datetime], end: Optional[datetime]) -> bool:
        if not start and not end:
            return True
        if not d.start_ts and not d.end_ts:
            return True
        s = d.start_ts or d.end_ts
        e = d.end_ts or d.start_ts
        if start and e and e < start: return False
        if end and s and s > end: return False
        return True

    def apply_filter(self):
        if self._updating: return
        try:
            start, end = self._range()
        except Exception as exc:
            QMessageBox.warning(self, "Smart File Discovery", str(exc)); return
        enabled = {typ for typ, chk in self.type_checks.items() if chk.isChecked()}
        self.selected_types = set(enabled)
        rows = [d for d in self.discovered if d.log_type in enabled and self._overlaps_range(d, start, end)]
        rows.sort(key=lambda d: (d.start_ts or datetime.max, d.log_type, Path(d.path).name))
        self.filtered = rows
        self.populate_file_table(rows)
        checked = sum(1 for d in rows if self.file_selection.get(d.path, True))
        self.summary_label.setText(f"Filtered files: {len(rows)} / Checked: {checked} / Estimated rows: {sum(d.row_count for d in rows):,}")

    def populate_file_table(self, rows: list[DiscoveredFile]):
        self._updating = True
        self.table.setSortingEnabled(False)
        self.table.blockSignals(True)
        self.table.setRowCount(len(rows))
        for r, d in enumerate(rows):
            chk = QTableWidgetItem("")
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            chk.setCheckState(Qt.Checked if self.file_selection.get(d.path, True) else Qt.Unchecked)
            chk.setData(Qt.UserRole, d.path)
            self.table.setItem(r, 0, chk)
            vals = [d.log_type, format_viewer_timestamp(d.start_ts), format_viewer_timestamp(d.end_ts), f"{d.row_count:,}", f"{d.size_bytes/(1024*1024):.1f}", str(Path(d.path).name), d.error]
            for c, v in enumerate(vals, start=1):
                item = QTableWidgetItem(v)
                item.setData(Qt.UserRole, d.path)
                self.table.setItem(r, c, item)
        self.table.blockSignals(False)
        self.table.setSortingEnabled(True)
        self._updating = False

    def _file_item_changed(self, item: QTableWidgetItem):
        if self._updating or item.column() != 0:
            return
        path = item.data(Qt.UserRole)
        if path:
            self.file_selection[str(path)] = (item.checkState() == Qt.Checked)
            self._refresh_boundary_combos()
            self.apply_filter()

    def set_all_files(self, checked: bool):
        self._updating = True
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item:
                item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
                path = item.data(Qt.UserRole)
                if path:
                    self.file_selection[str(path)] = checked
        self._updating = False
        self._refresh_boundary_combos()
        self.apply_filter()

    def accept_with_selection(self):
        files = []
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item and item.checkState() == Qt.Checked:
                path = item.data(Qt.UserRole)
                if path:
                    files.append(str(path))
        if not files:
            QMessageBox.warning(self, "Smart File Discovery", "No file is checked.")
            return
        self.selected_files = files
        self.selected_types = {d.log_type for d in self.filtered if d.path in set(files)}
        self.accept()



def show_startup_error(text: str) -> None:
    write_startup_log(text)
    try:
        _app = QApplication.instance() or QApplication(sys.argv)
        QMessageBox.critical(None, APP_TITLE, text + f"\n\nLog:\n{startup_log_path()}")
    except Exception:
        pass


_RC72_HANDOFF = None

def main():
    global _RC72_HANDOFF
    from insightec_handoff import load_handoff
    _RC72_HANDOFF = load_handoff("log_explorer")
    def excepthook(exc_type, exc, tb):
        text = "Unhandled application error.\n\n" + "".join(traceback.format_exception(exc_type, exc, tb))
        show_startup_error(text)
    sys.excepthook = excepthook

    try:
        app = QApplication(sys.argv)
        w = MainWindow()
        w.show()
        sys.exit(app.exec())
    except Exception:
        show_startup_error("Startup failed.\n\n" + traceback.format_exc())
        raise


# v41.2 Clean Integrated: the application entry point is intentionally moved
# to the very end of this file so all v41/v41.1/v41.2 override blocks are
# applied before MainWindow is instantiated.  This prevents the UI from
# starting with older classes and fixes the apparent "ancestor return" issue.

# ---------------------------------------------------------------------------
# v41 Workflow + Viewer Usability overrides
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"
# Keep PSC/Review import-only for Merge: they are selectable Log Types, but do
# not contribute rows to the merged Search output. START imports selected files
# for Viewer use.
def merge_capable_types_from_selected(selected_types: set[str]) -> set[str]:
    return {t for t in selected_types if t not in {"PSC", "REVIEW"}}

_old_v40_build_ui = MainWindow.build_ui

def _v41_build_ui(self):
    _old_v40_build_ui(self)
    try:
        self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
        # START now means Import/Load-to-Viewer; MERGE is explicit.
        self.run_btn.setText("▶ START")
        self.run_btn.setToolTip("Import selected logs and open Log Viewer. This does not create a merged Excel/CSV output.")
        try:
            self.run_btn.clicked.disconnect()
        except Exception:
            pass
        self.run_btn.clicked.connect(self.start_import_clicked)
        self.merge_btn = QPushButton("MERGE")
        self.merge_btn.setMinimumSize(120, 44)
        self.merge_btn.setStyleSheet("QPushButton{font-size:14pt;font-weight:700;background:#2563eb;color:white;border-radius:10px;padding:8px 18px;} QPushButton:hover{background:#1d4ed8;}")
        self.merge_btn.setToolTip("Create merged output from merge-capable log types. PSC and Review remain import/viewer-only.")
        self.merge_btn.clicked.connect(self.run_clicked)
        # Put MERGE next to START in the top action bar.
        lay = self.layout().itemAt(2).layout() if self.layout() and self.layout().count() > 2 else None
        if lay is not None:
            lay.insertWidget(1, self.merge_btn)
        if hasattr(self, "import_btn"):
            self.import_btn.setVisible(False)
        if hasattr(self, "chk_review"):
            self.chk_review.setText("Review")
            self.chk_review.setEnabled(True)
            self.chk_review.setToolTip("Review is handled as a normal detected log type for Import/Viewer, but is not included in merged Search output.")
        if hasattr(self, "chk_psc"):
            self.chk_psc.setToolTip("PSC is handled as a normal detected log type for Import/Viewer, but is not included in merged Search output.")
        # Button text trimming.
        for b in self.findChildren(QPushButton):
            if b.text() == "Open Output":
                b.setMinimumWidth(105)
            elif b.text() == "Reset Defaults":
                b.setMinimumWidth(115)
    except Exception:
        write_startup_log("v41 build_ui patch failed.\n\n" + traceback.format_exc())

MainWindow.build_ui = _v41_build_ui


def _v41_start_import_clicked(self):
    opt = self.collect_options()
    if not opt.source_folder:
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder is required.")
        return
    if not Path(opt.source_folder).exists():
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder does not exist.")
        return
    if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
        QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
        return
    dlg = SmartFileDiscoveryDialog(self, opt)
    if dlg.exec() != QDialog.Accepted:
        self.status_label.setText("Cancelled before import")
        return
    self.viewer_selected_files = list(dlg.selected_files)
    self.viewer_selected_types = set(dlg.selected_types)
    self.save_settings()
    self.log_view.clear()
    self.log(f"Import selected {len(self.viewer_selected_files)} files for Log Viewer.")
    # Show final preparation progress so the user does not think the app froze.
    progress = SoftProgressDialog("Import Progress", "Preparing Log Viewer...", None, 0, 0, self)
    progress.show(); QApplication.processEvents()
    try:
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        if hasattr(self.viewer_window, "refresh_available_sources"):
            self.viewer_window.refresh_available_sources()
        progress.setLabelText("Opening Log Viewer..."); QApplication.processEvents()
        self.open_dual_viewer()
    finally:
        progress.close()

MainWindow.start_import_clicked = _v41_start_import_clicked


_old_run_clicked_v40 = MainWindow.run_clicked

def _v41_run_clicked(self):
    # Same discovery flow as before, but PSC/Review are import/viewer-only.
    opt = self.collect_options()
    if not opt.source_folder:
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder is required.")
        return
    if not Path(opt.source_folder).exists():
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder does not exist.")
        return
    if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
        QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
        return
    dlg = SmartFileDiscoveryDialog(self, opt)
    if dlg.exec() != QDialog.Accepted:
        self.status_label.setText("Cancelled before merge")
        return
    selected_types = set(dlg.selected_types)
    # Preserve selected files for Viewer after merge/import.
    self.viewer_selected_files = list(dlg.selected_files)
    self.viewer_selected_types = set(selected_types)
    merge_types = merge_capable_types_from_selected(selected_types)
    opt.include_ws = "WS" in merge_types
    opt.include_watersystem = "WATERSYSTEM" in merge_types
    opt.include_cga = "CGA" in merge_types
    opt.include_csa = "CSA" in merge_types
    opt.include_mrserver = "MRSERVER" in merge_types
    opt.include_gesys = "GESYS" in merge_types
    opt.include_lais = "LAIS" in merge_types
    opt.include_psc = False
    opt.include_unknown = "UNKNOWN" in merge_types
    opt.selected_files = [f for f in dlg.selected_files if classify_file(Path(f), True) in merge_types]
    if not enabled_types(opt):
        # If operator selected only import-only files, open Viewer instead of error.
        self.log("Only import/viewer-only log types were selected. Opening Log Viewer.")
        return self.start_import_clicked()
    self.save_settings()
    self.log_view.clear()
    self.log(f"Smart File Discovery selected {len(dlg.selected_files or [])} files. Merge-capable files: {len(opt.selected_files or [])}.")
    self.progress.setValue(0)
    self.status_label.setText("Running merge...")
    self.run_btn.setEnabled(False)
    if hasattr(self, "merge_btn"):
        self.merge_btn.setEnabled(False)
    self.pause_btn.setEnabled(True)
    self.cancel_btn.setEnabled(True)
    self.pause_btn.setText("Pause")
    self._show_progress_popup("Log Merge Progress", "Parsing selected files...", 0, 100, cancellable=True)
    self.worker = MergeWorker(opt)
    self.worker.progress.connect(self.on_progress)
    self.worker.log.connect(self.log)
    self.worker.finished_ok.connect(self.on_finished)
    self.worker.failed.connect(self.on_failed)
    self.worker.start()

MainWindow.run_clicked = _v41_run_clicked

_old_on_finished_v40 = MainWindow.on_finished

def _v41_on_finished(self, result: RunResult):
    _old_on_finished_v40(self, result)
    if hasattr(self, "merge_btn"):
        self.merge_btn.setEnabled(True)

MainWindow.on_finished = _v41_on_finished

_old_on_failed_v40 = MainWindow.on_failed

def _v41_on_failed(self, text: str):
    _old_on_failed_v40(self, text)
    if hasattr(self, "merge_btn"):
        self.merge_btn.setEnabled(True)

MainWindow.on_failed = _v41_on_failed


# Viewer fixes and UI simplification.
# Foundation baseline: capture the class initializer before v41/v41.1/v41.2
# compatibility patches alter it.  A single final initializer is installed at
# the end of this file.
_FOUNDATION_MPL_BASE_INIT = MultiPaneLogViewer.__init__
_old_mpl_init = MultiPaneLogViewer.__init__
def _v41_mpl_init(self, parent_window):
    _old_mpl_init(self, parent_window)
    try:
        geo = QApplication.primaryScreen().availableGeometry()
        self.resize(max(850, int(geo.width()*0.62)), max(480, int(geo.height()*0.55)))
        if hasattr(self, "detail"):
            self.detail.setVisible(False)
            self.detail.setMaximumHeight(0)
        # No View mode dropdown. Show checkboxes control visible panes.
        if hasattr(self, "mode_combo"):
            self.mode_combo.setVisible(False)
        for lab in self.findChildren(QLabel):
            if lab.text().strip().lower().startswith("view mode"):
                lab.setVisible(False)
        self.refresh_available_sources()
        for i, table in enumerate(getattr(self, "tables", [])):
            table.doubleClicked.connect(lambda _idx, pane=i: self.set_time_from_clicked_row(pane))
        self.update_view_mode()
    except Exception:
        write_startup_log("v41 viewer init patch failed.\n\n" + traceback.format_exc())

MultiPaneLogViewer.__init__ = _v41_mpl_init


def _v41_visible_indices(self) -> list[int]:
    checked = [i for i, cb in enumerate(getattr(self, "pane_visible_checks", [])) if cb.isChecked()]
    return checked or [0, 1]
MultiPaneLogViewer.visible_indices = _v41_visible_indices


def _v41_update_view_mode(self):
    if not hasattr(self, "main_splitter") or not hasattr(self, "panes"):
        return
    visible = self.visible_indices()
    for i, pane in enumerate(self.panes):
        pane.setVisible(i in visible)
    size_each = max(1, 1000 // max(1, len(visible)))
    try:
        self.main_splitter.setSizes([size_each if i in visible else 0 for i in range(self.MAX_PANES)])
    except Exception:
        pass
    self.log(f"Viewer panes: " + ", ".join(str(i+1) for i in visible))
MultiPaneLogViewer.update_view_mode = _v41_update_view_mode


def _v41_refresh_available_sources(self):
    # If START/Discovery has selected files, only show log types that actually exist.
    parent = getattr(self, "parent_window", None)
    types = []
    files = getattr(parent, "viewer_selected_files", None) if parent is not None else None
    if files:
        found = set()
        for f in files:
            p = Path(f)
            if is_review_file(p.name):
                found.add("Review")
            else:
                typ = classify_file(p, True)
                if typ:
                    found.add("WaterSystem" if typ == "WATERSYSTEM" else typ)
        types = sorted(found, key=lambda x: ["Merged","WS","WaterSystem","CGA","CSA","MRSERVER","GESYS","LAIS","PSC","Review","UNKNOWN"].index(x) if x in ["Merged","WS","WaterSystem","CGA","CSA","MRSERVER","GESYS","LAIS","PSC","Review","UNKNOWN"] else 99)
    if not types:
        types = [s for s in self.SOURCES if s != "Custom File"]
    # Merged is available only when not using selected-files-only import, or if output exists.
    if "Merged" not in types and not files:
        types.insert(0, "Merged")
    types.append("Custom File")
    for i, combo in enumerate(getattr(self, "sources", [])):
        current = combo.currentText()
        combo.blockSignals(True)
        combo.clear(); combo.addItems(types)
        if current in types:
            combo.setCurrentText(current)
        elif i < len(types):
            combo.setCurrentIndex(min(i, len(types)-1))
        combo.blockSignals(False)

MultiPaneLogViewer.refresh_available_sources = _v41_refresh_available_sources

_old_mpl_source_to_records = MultiPaneLogViewer.source_to_records

def _v41_source_to_records(self, source_name: str, progress: Optional[QProgressDialog] = None) -> list[LogRecord]:
    parent = getattr(self, "parent_window", None)
    selected = getattr(parent, "viewer_selected_files", None) if parent is not None else None
    if selected and source_name not in {"Merged", "Custom File"}:
        wanted = "WATERSYSTEM" if source_name == "WaterSystem" else ("REVIEW" if source_name == "Review" else source_name)
        records: list[LogRecord] = []
        candidates: list[Path] = []
        for f in selected:
            p = Path(f)
            if wanted == "REVIEW":
                if is_review_file(p.name):
                    candidates.append(p)
            else:
                typ = classify_file(p, True)
                if typ == wanted:
                    candidates.append(p)
        if progress:
            progress.setRange(0, max(1, len(candidates)))
            progress.setLabelText(f"Loading {source_name}: 0/{len(candidates)} files")
            QApplication.processEvents()
        for idx, p in enumerate(candidates, 1):
            if progress and progress.wasCanceled():
                return records
            if progress:
                progress.setValue(idx-1); progress.setLabelText(f"Loading {source_name}: {idx}/{len(candidates)}\n{p.name}"); QApplication.processEvents()
            try:
                if wanted == "REVIEW":
                    records.extend(review_rows_to_viewer_records(p))
                elif wanted == "PSC":
                    _, _, recs = parse_psc_file_detail(p); records.extend(recs)
                else:
                    records.extend(parse_file(p, wanted))
            except Exception as e:
                records.append(LogRecord(None, wanted, p.name, 0, "ERROR", "Viewer", f"Parse failed: {e}", ""))
        opt = self.parent_window.collect_options()
        start_date = parse_user_date(opt.start_date) if opt.use_start and opt.start_date else None
        end_date = parse_user_date(opt.end_date) if opt.use_end and opt.end_date else None
        if start_date or end_date:
            records = [r for r in records if in_date_range(r.timestamp, None, start_date, end_date)]
        records.sort(key=lambda r: r.timestamp or datetime.min)
        return records
    return _old_mpl_source_to_records(self, source_name, progress)

MultiPaneLogViewer.source_to_records = _v41_source_to_records


def _v41_default_viewer_columns_for_rows(self, rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["Timestamp", "Message", "Level"]
    stypes = {str(r.get("SourceType", "")) for r in rows[:200]}
    if "REVIEW" in stypes or "Review" in stypes:
        preferred = ["Timestamp"] + [k for k in rows[0].keys() if k not in {"Timestamp", "Message", "Raw", "_ts"}]
        return preferred[:12] or ["Timestamp", "Level"]
    if "WATERSYSTEM" in stypes or "WaterSystem" in stypes:
        return [c for c in ["Timestamp", "MainState", "Error"] if any(c in r for r in rows[:50])]
    return [c for c in ["Timestamp", "Message", "Level"] if any(c in r for r in rows[:50])] or ["Timestamp", "Message", "Level"]

MultiPaneLogViewer.default_viewer_columns_for_rows = _v41_default_viewer_columns_for_rows


def _v41_apply_table_column_widths(self, idx: int):
    table = self.tables[idx]
    model = self.models[idx]
    try:
        for col_idx, col_name in enumerate(model.columns):
            if col_name == "Timestamp":
                table.setColumnWidth(col_idx, 148)
                table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Fixed)
            elif col_name == "Message":
                table.setColumnWidth(col_idx, 620)
                table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Stretch)
            elif col_name in {"Level", "Error"}:
                table.setColumnWidth(col_idx, 82)
                table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Interactive)
            elif col_name == "MainState":
                table.setColumnWidth(col_idx, 155)
                table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Interactive)
            elif col_name in {"ChillerTemp", "PrimaryFlowMeter", "AbsolutePressure", "DynamicPressure", "XdTemperature", "VacuumLevel", "DOLevel", "WaterVolume", "SecondaryFlowMeter", "HsCombitac", "ChillerStatus", "ChillerLowLevelInd", "PressureSetPoint"}:
                table.setColumnWidth(col_idx, 105)
                table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Interactive)
            else:
                table.setColumnWidth(col_idx, 120)
                table.horizontalHeader().setSectionResizeMode(col_idx, QHeaderView.Interactive)
    except Exception:
        pass

MultiPaneLogViewer.apply_table_column_widths = _v41_apply_table_column_widths


def _v41_set_time_from_clicked_row(self, pane_idx: int):
    try:
        indexes = self.tables[pane_idx].selectionModel().selectedRows()
        if not indexes:
            return
        row = self.models[pane_idx].row_at(indexes[0].row()) or {}
        ts = str(row.get("Timestamp", "")).strip()
        if not ts:
            return
        self.viewer_time_enable.setChecked(True)
        if not self.viewer_start_edit.text().strip():
            self.viewer_start_edit.setText(ts)
        else:
            self.viewer_end_edit.setText(ts)
        self.apply_time_to_visible()
    except Exception:
        pass

MultiPaneLogViewer.set_time_from_clicked_row = _v41_set_time_from_clicked_row

_old_mpl_make_progress = MultiPaneLogViewer.make_progress

def _v41_make_progress(self, title: str, text: str) -> QProgressDialog:
    return SoftProgressDialog(title, text, "Cancel", 0, 100, self)

MultiPaneLogViewer.make_progress = _v41_make_progress

# v41 Plugin Builder capability note: builder ZIP manifest accepts supports_merge.
# Existing builder remains backward compatible; plugins without supports_merge are treated as merge-capable unless marked import_only.

# ---------------------------------------------------------------------------
# v41.1 Stabilization patch
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"
# 1) Make Log Viewer opening robust.  In v41 the viewer could be constructed
# while UI patches were still initializing and the exception hook displayed the
# same error twice.  This version logs once, shows one concise message, and
# guarantees the window is kept on screen.
def _v411_open_dual_viewer(self):
    try:
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        if hasattr(self.viewer_window, "refresh_available_sources"):
            self.viewer_window.refresh_available_sources()
        geo = QApplication.primaryScreen().availableGeometry()
        # Keep the first viewer window comfortably inside the display.
        w = min(max(860, int(geo.width() * 0.58)), geo.width() - 80)
        h = min(max(520, int(geo.height() * 0.52)), geo.height() - 80)
        self.viewer_window.resize(w, h)
        self.viewer_window.move(geo.x() + max(20, (geo.width() - w)//2), geo.y() + max(20, (geo.height() - h)//2))
        self.viewer_window.show()
        self.viewer_window.raise_()
        self.viewer_window.activateWindow()
    except Exception:
        text = "Log Viewer open failed.\n\n" + traceback.format_exc()
        write_startup_log(text)
        QMessageBox.critical(self, APP_TITLE, text)

MainWindow.open_dual_viewer = _v411_open_dual_viewer


# 2) Make START import progress explicit until the viewer is visible.
def _v411_start_import_clicked(self):
    opt = self.collect_options()
    if not opt.source_folder:
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder is required.")
        return
    if not Path(opt.source_folder).exists():
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder does not exist.")
        return
    if opt.use_start and opt.use_end and opt.start_date and opt.end_date and opt.start_date > opt.end_date:
        QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
        return
    dlg = SmartFileDiscoveryDialog(self, opt)
    if dlg.exec() != QDialog.Accepted:
        self.status_label.setText("Cancelled before import")
        return
    self.viewer_selected_files = list(dlg.selected_files)
    self.viewer_selected_types = set(dlg.selected_types)
    self.save_settings()
    self.log_view.clear()
    self.log(f"Import selected {len(self.viewer_selected_files)} files for Log Viewer.")
    progress = SoftProgressDialog("Import Progress", "Preparing Log Viewer...", None, 0, 100, self)
    progress.setValue(5)
    progress.show(); QApplication.processEvents()
    try:
        progress.setLabelText("Building viewer source list..."); progress.setValue(35); QApplication.processEvents()
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        progress.setLabelText("Applying selected log types..."); progress.setValue(60); QApplication.processEvents()
        if hasattr(self.viewer_window, "refresh_available_sources"):
            self.viewer_window.refresh_available_sources()
        progress.setLabelText("Opening Log Viewer..."); progress.setValue(90); QApplication.processEvents()
        self.open_dual_viewer()
        progress.setValue(100); QApplication.processEvents()
    finally:
        progress.close()

MainWindow.start_import_clicked = _v411_start_import_clicked


# 3) Ensure the viewer only offers loaded log types after START/MERGE.
def _v411_refresh_available_sources(self):
    parent = getattr(self, "parent_window", None)
    files = getattr(parent, "viewer_selected_files", None) if parent is not None else None
    types = []
    if files:
        found = set()
        for f in files:
            p = Path(f)
            if is_review_file(p.name):
                found.add("Review")
            else:
                typ = classify_file(p, True)
                if typ:
                    found.add("WaterSystem" if typ == "WATERSYSTEM" else typ)
        order = ["Merged", "WS", "WaterSystem", "VIMeasure", "ACQUISITION", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "Review", "UNKNOWN"]
        types = sorted(found, key=lambda x: order.index(x) if x in order else 99)
    else:
        # Manual Log Viewer open: keep Custom File only until logs are loaded.
        types = ["Custom File"]
    if not types:
        types = ["Custom File"]
    if "Custom File" not in types:
        types.append("Custom File")
    for i, combo in enumerate(getattr(self, "sources", [])):
        current = combo.currentText()
        combo.blockSignals(True)
        combo.clear(); combo.addItems(types)
        if current in types:
            combo.setCurrentText(current)
        else:
            combo.setCurrentIndex(min(i, len(types)-1))
        combo.blockSignals(False)

MultiPaneLogViewer.refresh_available_sources = _v411_refresh_available_sources


# 4) Make viewer progress match the larger soft progress dialog.
def _v411_make_progress(self, title: str, text: str) -> QProgressDialog:
    dlg = SoftProgressDialog(title, text, "Cancel", 0, 100, self)
    dlg.resize(760, 280)
    return dlg

MultiPaneLogViewer.make_progress = _v411_make_progress


# 5) Default pane visibility and layout: Show checkboxes only, no view-mode dropdown,
#    timestamp compact, message-first display, no bottom details.
def _v411_mpl_init(self, parent_window):
    try:
        LegacyDualLogViewer.__init__(self, parent_window)
        # Re-apply MultiPaneLogViewer-specific layout if the previous monkey patches
        # have been bypassed by initialization order.  Existing attributes are reused.
        if hasattr(self, "mode_combo"):
            self.mode_combo.setVisible(False)
        for lab in self.findChildren(QLabel):
            if lab.text().strip().lower().startswith("view mode"):
                lab.setVisible(False)
        if hasattr(self, "detail"):
            self.detail.hide(); self.detail.setMaximumHeight(0)
        if hasattr(self, "pane_visible_checks"):
            for i, cb in enumerate(self.pane_visible_checks):
                cb.setChecked(i < 2)
        self.refresh_available_sources()
        geo = QApplication.primaryScreen().availableGeometry()
        self.resize(min(max(860, int(geo.width()*0.58)), geo.width()-80), min(max(520, int(geo.height()*0.52)), geo.height()-80))
        if hasattr(self, "update_view_mode"):
            self.update_view_mode()
    except Exception:
        write_startup_log("v41.1 viewer init patch failed.\n\n" + traceback.format_exc())
        raise

# Keep the already-built MultiPane __init__ if present; patch only when it still errors.
# This fallback avoids the main_splitter access race from v41 while preserving existing UI.
_old_v411_mpl_init_original = MultiPaneLogViewer.__init__
def _v411_safe_mpl_init(self, parent_window):
    try:
        _old_v411_mpl_init_original(self, parent_window)
    except AttributeError as exc:
        if "main_splitter" not in str(exc):
            raise
        _v411_mpl_init(self, parent_window)
    try:
        if hasattr(self, "mode_combo"):
            self.mode_combo.setVisible(False)
        for lab in self.findChildren(QLabel):
            if lab.text().strip().lower().startswith("view mode"):
                lab.setVisible(False)
        if hasattr(self, "detail"):
            self.detail.hide(); self.detail.setMaximumHeight(0)
        self.refresh_available_sources()
        self.update_view_mode()
    except Exception:
        write_startup_log("v41.1 viewer post-init patch failed.\n\n" + traceback.format_exc())

MultiPaneLogViewer.__init__ = _v411_safe_mpl_init


# 6) After MERGE completes, show a preparation progress while opening Viewer.
_old_v411_on_finished_prev = MainWindow.on_finished
def _v411_on_finished(self, result: RunResult):
    # Original handler already updates status and may auto-open; make sure it doesn't
    # leave the user without feedback for the viewer preparation stage.
    try:
        self._close_progress_popup()
        self.run_btn.setEnabled(True)
        if hasattr(self, "merge_btn"):
            self.merge_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        self.progress.setValue(100)
        self.status_label.setText("Completed")
        self.log("Completed.")
        self.log(f"Output: {result.output_path}")
        prep = SoftProgressDialog("Opening Log Viewer", "Preparing Log Viewer...", None, 0, 100, self)
        prep.setValue(10); prep.show(); QApplication.processEvents()
        prep.setLabelText("Refreshing loaded log list..."); prep.setValue(50); QApplication.processEvents()
        self.open_dual_viewer()
        prep.setValue(100); QApplication.processEvents()
        prep.close()
    except Exception:
        write_startup_log("v41.1 on_finished failed.\n\n" + traceback.format_exc())
        QMessageBox.critical(self, APP_TITLE, traceback.format_exc())

MainWindow.on_finished = _v411_on_finished

# ---------------------------------------------------------------------------
# v41.2 Clean Integrated patch
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"
def _v412_polish_main_window(self):
    """Final integrated UI cleanup applied after v40/v41 widgets exist."""
    try:
        self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
        # START is Import/Viewer. MERGE is explicit.
        if hasattr(self, "run_btn"):
            self.run_btn.setText("▶ START")
            self.run_btn.setMinimumSize(150, 48)
            self.run_btn.setToolTip("Import selected logs and open Log Viewer. No merged output is created.")
        if hasattr(self, "merge_btn"):
            self.merge_btn.setText("MERGE")
            self.merge_btn.setMinimumSize(120, 44)
            self.merge_btn.setToolTip("Create merged output from merge-capable log types only. PSC and Review remain import/viewer-only.")
        # Old individual import/update buttons must stay hidden.
        for name in ["import_psc_btn", "import_review_btn", "import_file_btn", "update_file_type_btn", "update_tool_btn"]:
            w = getattr(self, name, None)
            if w is not None:
                w.hide()
        # Only one Update button should be visible.
        if hasattr(self, "update_btn"):
            self.update_btn.setText("Update...")
            self.update_btn.setMinimumWidth(105)
            self.update_btn.setToolTip("Select a ZIP. The tool automatically detects Tool/Core update ZIP or File Type plugin ZIP.")
        # Review/PSC are normal selectable log types, but import-only for Merge.
        if hasattr(self, "chk_review"):
            self.chk_review.setText("Review")
            self.chk_review.setEnabled(True)
            self.chk_review.setToolTip("Review is available for Import/Viewer. It is automatically skipped from merged Search output.")
        if hasattr(self, "chk_psc"):
            self.chk_psc.setToolTip("PSC is available for Import/Viewer. It is automatically skipped from merged Search output.")
        # Noise filter default off.
        if hasattr(self, "noise_enable_merge_chk"):
            self.noise_enable_merge_chk.setChecked(False)
        if hasattr(self, "noise_learning_chk"):
            self.noise_learning_chk.setChecked(False)
        # Avoid truncated text.
        for b in self.findChildren(QPushButton):
            txt = b.text().strip()
            if txt in {"Export Project", "Import Project", "Split Merge", "Reset Defaults", "Manage Rules", "Same as Source"}:
                b.setMinimumWidth(max(b.minimumWidth(), 105))
    except Exception:
        write_startup_log("v41.2 main polish failed.\n\n" + traceback.format_exc())


_old_v412_main_init = MainWindow.__init__
def _v412_main_init(self):
    _old_v412_main_init(self)
    _v412_polish_main_window(self)
    try:
        screen = self.screen() or QApplication.primaryScreen()
        if screen is not None:
            available = screen.availableGeometry()
            frame = self.frameGeometry()
            width = min(max(self.width(), 980), max(980, available.width() - 40))
            height = min(max(self.height(), 650), max(650, available.height() - 40))
            self.resize(width, height)
            self.move(
                available.x() + max(0, (available.width() - width) // 2),
                available.y() + max(0, (available.height() - height) // 2),
            )
    except Exception:
        write_startup_log("Main window position correction failed.\n\n" + traceback.format_exc())

MainWindow.__init__ = _v412_main_init


# Replace v41.1 MERGE completion handler with a wrapper that cannot show the
# same exception twice and always provides viewer-preparation feedback.
def _v412_on_finished(self, result: RunResult):
    try:
        self._close_progress_popup()
        if hasattr(self, "run_btn"):
            self.run_btn.setEnabled(True)
        if hasattr(self, "merge_btn"):
            self.merge_btn.setEnabled(True)
        if hasattr(self, "pause_btn"):
            self.pause_btn.setEnabled(False)
        if hasattr(self, "cancel_btn"):
            self.cancel_btn.setEnabled(False)
        if hasattr(self, "progress"):
            self.progress.setValue(100)
        self.status_label.setText("Completed")
        self.log("Completed.")
        self.log(f"Output: {result.output_path}")
        prep = SoftProgressDialog("Opening Log Viewer", "Preparing Log Viewer...", None, 0, 100, self)
        prep.setValue(10); prep.show(); QApplication.processEvents()
        prep.setLabelText("Building viewer source list..."); prep.setValue(45); QApplication.processEvents()
        prep.setLabelText("Opening Log Viewer..."); prep.setValue(80); QApplication.processEvents()
        self.open_dual_viewer()
        prep.setValue(100); QApplication.processEvents()
        prep.close()
    except Exception:
        text = "Merge completed, but opening Log Viewer failed.\n\n" + traceback.format_exc()
        write_startup_log(text)
        QMessageBox.critical(self, APP_TITLE, text)

MainWindow.on_finished = _v412_on_finished


# Log Viewer cleanup: no View Mode dropdown, LOAD LOGS wording, visible-loaded
# source list only, compact timestamp and message-priority defaults.
_old_v412_mpl_init = MultiPaneLogViewer.__init__
def _v412_mpl_init(self, parent_window):
    _old_v412_mpl_init(self, parent_window)
    try:
        if hasattr(self, "mode_combo"):
            self.mode_combo.hide()
        for lab in self.findChildren(QLabel):
            if lab.text().strip().lower().startswith("view mode"):
                lab.hide()
        for btn in self.findChildren(QPushButton):
            if "Load Visible" in btn.text() or btn.text().strip().lower() in {"load visible logs", "load all visible"}:
                btn.setText("LOAD LOGS")
                btn.setMinimumHeight(34)
                btn.setMinimumWidth(110)
                btn.setStyleSheet("QPushButton{font-weight:700;background:#2563eb;color:white;border-radius:8px;padding:6px 14px;} QPushButton:hover{background:#1d4ed8;}")
        if hasattr(self, "detail"):
            self.detail.hide(); self.detail.setMaximumHeight(0)
        self.refresh_available_sources()
        self.update_view_mode()
    except Exception:
        write_startup_log("v41.2 viewer polish failed.\n\n" + traceback.format_exc())

MultiPaneLogViewer.__init__ = _v412_mpl_init


# ---------------------------------------------------------------------------
# Commit0008 Foundation: deterministic Viewer initialization
# ---------------------------------------------------------------------------
# The older compatibility functions above are retained temporarily for source
# history, but they no longer form a chained runtime path.  All Viewer creation
# now starts from the original class initializer captured before those patches.
from foundation.viewer import initialize_viewer_foundation

def _foundation_mpl_init(self, parent_window):
    initialize_viewer_foundation(
        self,
        parent_window,
        _FOUNDATION_MPL_BASE_INIT,
    )

MultiPaneLogViewer.__init__ = _foundation_mpl_init


# ---------------------------------------------------------------------------
# Commit0014: Shared Feedback Engine + single-file ZIP import
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"
def _safe_zip_member_name(name: str) -> bool:
    """Return True only for normal relative files inside a ZIP archive."""
    text = str(name or "").replace("\\", "/")
    if not text or text.endswith("/"):
        return False
    path = Path(text)
    if path.is_absolute() or ".." in path.parts:
        return False
    return True


def _zip_import_candidates(zip_path: Path) -> list[str]:
    """List one-file-import candidates without extracting the whole archive."""
    supported = {".log", ".txt", ".out", ".ar"}
    with zipfile.ZipFile(zip_path, "r") as zf:
        result = []
        for info in zf.infolist():
            name = info.filename
            if info.is_dir() or not _safe_zip_member_name(name):
                continue
            suffix = Path(name).suffix.lower()
            base = Path(name).name.lower()
            if suffix in supported or base.startswith("review.out"):
                result.append(name)
        return sorted(result, key=lambda x: x.casefold())


def extract_one_file_from_zip(zip_path: str | Path, parent: QWidget | None = None) -> tuple[Path, tempfile.TemporaryDirectory]:
    """Select and safely extract exactly one supported file from an archive.

    The returned TemporaryDirectory must be kept alive until processing ends.
    """
    archive = Path(zip_path)
    candidates = _zip_import_candidates(archive)
    if not candidates:
        raise ValueError("No supported log file was found in the selected ZIP.")
    selected = candidates[0]
    if len(candidates) > 1:
        label_map = {f"{Path(name).name}    [{name}]": name for name in candidates}
        labels = list(label_map.keys())
        item, ok = QInputDialog.getItem(
            parent,
            "Select one file from ZIP",
            "This RC1 test imports one file only. Select the file to extract:",
            labels,
            0,
            False,
        )
        if not ok or not item:
            raise RuntimeError("ZIP file selection was cancelled.")
        selected = label_map[str(item)]
    temp_holder = tempfile.TemporaryDirectory(prefix="logmerge_zip_one_")
    temp_root = Path(temp_holder.name)
    target = temp_root / Path(selected).name
    with zipfile.ZipFile(archive, "r") as zf:
        info = zf.getinfo(selected)
        if info.file_size > 1024 * 1024 * 1024:
            temp_holder.cleanup()
            raise ValueError("The selected ZIP member is larger than the 1 GB safety limit.")
        with zf.open(info, "r") as src, target.open("wb") as dst:
            shutil.copyfileobj(src, dst, length=1024 * 1024)
    return target, temp_holder


def _commit0014_import_selected_file_clicked(self):
    path, _ = QFileDialog.getOpenFileName(
        self,
        "Select one log file or ZIP to import",
        self.source_edit.text() or str(Path.home()),
        "Log or ZIP files (*.log *.txt *.out *.ar *.zip);;ZIP files (*.zip);;All Files (*.*)",
    )
    if not path:
        return
    selected_path = Path(path)
    temp_holder = None
    try:
        if selected_path.suffix.lower() == ".zip":
            selected_path, temp_holder = extract_one_file_from_zip(selected_path, self)
            self.log(f"ZIP single-file import: extracted {selected_path.name}")
        out_dir = self.output_edit.text().strip() or str(Path(path).parent)
        # Keep the temporary directory alive until the worker has consumed the file.
        def task():
            try:
                return import_selected_file_only(str(selected_path), out_dir)
            finally:
                if temp_holder is not None:
                    temp_holder.cleanup()
        self._run_import_task(
            "Selected File Import Progress",
            f"Importing {selected_path.name} and writing output...",
            task,
            "Selected file import",
        )
    except RuntimeError as exc:
        if temp_holder is not None:
            temp_holder.cleanup()
        if "cancelled" not in str(exc).lower():
            QMessageBox.warning(self, "ZIP Import", str(exc))
    except Exception as exc:
        if temp_holder is not None:
            temp_holder.cleanup()
        QMessageBox.critical(self, "ZIP Import Failed", str(exc))


MainWindow.import_selected_file_clicked = _commit0014_import_selected_file_clicked


class FeedbackDialog(QDialog):
    """Host UI for the shared InSightec Feedback Engine v1."""
    def __init__(self, host: "MainWindow"):
        super().__init__(host)
        self.host = host
        self.setWindowTitle("Send Feedback / Validation")
        self.resize(720, 560)
        self.attachments: list[str] = []
        root = QVBoxLayout(self)
        form = QFormLayout()
        self.category = QComboBox(); self.category.addItems(["Bug Report", "Improvement Request", "Validation Result", "Question"])
        self.priority = QComboBox(); self.priority.addItems(["Low", "Normal", "High", "Critical"]); self.priority.setCurrentText("Normal")
        self.repro = QComboBox(); self.repro.addItems(["Unknown", "Always", "Sometimes", "Once", "Not Applicable"])
        self.mode = QComboBox(); self.mode.addItems(["template", "outlook"])
        self.related = QLineEdit("Log Merge Tool")
        self.page = QLineEdit("Main Window")
        form.addRow("Category", self.category)
        form.addRow("Priority", self.priority)
        form.addRow("Reproducible", self.repro)
        form.addRow("Send mode", self.mode)
        form.addRow("Related tool", self.related)
        form.addRow("Current screen", self.page)
        root.addLayout(form)
        root.addWidget(QLabel("Comment"))
        self.comment = QTextEdit(); self.comment.setPlaceholderText("Describe the issue, expected behavior, and steps to reproduce.")
        root.addWidget(self.comment, 1)
        attach_row = QHBoxLayout()
        add_btn = QPushButton("Add Files..."); add_btn.clicked.connect(self.add_files)
        clear_btn = QPushButton("Clear Attachments"); clear_btn.clicked.connect(self.clear_files)
        attach_row.addWidget(add_btn); attach_row.addWidget(clear_btn); attach_row.addStretch(1)
        root.addLayout(attach_row)
        self.attach_label = QLabel("Attachments: none")
        self.attach_label.setWordWrap(True)
        root.addWidget(self.attach_label)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Save).setText("Prepare Feedback")
        buttons.accepted.connect(self.prepare_feedback)
        buttons.rejected.connect(self.reject)
        root.addWidget(buttons)

    def add_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select feedback attachments", str(Path.home()), "All Files (*.*)")
        for item in files:
            if item and item not in self.attachments:
                self.attachments.append(item)
        self.update_attachment_label()

    def clear_files(self):
        self.attachments.clear()
        self.update_attachment_label()

    def update_attachment_label(self):
        if not self.attachments:
            self.attach_label.setText("Attachments: none")
        else:
            self.attach_label.setText("Attachments: " + ", ".join(Path(x).name for x in self.attachments))

    def prepare_feedback(self):
        try:
            from feedback_engine import FeedbackEngine, FeedbackRequest, build_runtime_context
            extra = {
                "source_folder": self.host.source_edit.text().strip() if hasattr(self.host, "source_edit") else "",
                "output_folder": self.host.output_edit.text().strip() if hasattr(self.host, "output_edit") else "",
                "viewer_selected_files": len(getattr(self.host, "viewer_selected_files", []) or []),
                "installed_file_types": [p.get("display_name", p.get("id", "")) for p in load_file_type_plugins()],
            }
            context = build_runtime_context(
                application="Log Merge Tool",
                application_version=APP_VERSION,
                build="Commit0014",
                current_page=self.page.text().strip(),
                related_tool=self.related.text().strip() or "Log Merge Tool",
                tool_version=APP_VERSION,
                extra=extra,
            )
            engine = FeedbackEngine(app_dir(), log_dir=app_dir() / "logs")
            request = FeedbackRequest(
                category=self.category.currentText(),
                priority=self.priority.currentText(),
                reproducible=self.repro.currentText(),
                comment=self.comment.toPlainText(),
                mode=self.mode.currentText(),
                attachments=list(self.attachments),
                context=context,
            )
            validation = [
                f"Application started: yes",
                f"Plugin count: {len(load_file_type_plugins())}",
                f"Current screen: {context.current_page}",
            ]
            body, paths = engine.prepare(request, validation)
            if request.mode == "outlook":
                QMessageBox.information(
                    self,
                    "Feedback Prepared",
                    "The shared Feedback Engine prepared the template and manifest.\n"
                    "Outlook COM sending remains a host integration step in this RC1 test.\n\n"
                    f"Template: {paths['template']}\nManifest: {paths['manifest']}",
                )
            else:
                QMessageBox.information(
                    self,
                    "Feedback Prepared",
                    f"Feedback files were created.\n\nTemplate: {paths['template']}\nManifest: {paths['manifest']}",
                )
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Feedback Failed", str(exc))


def _commit0014_open_feedback(self):
    FeedbackDialog(self).exec()


_old_commit0014_main_init = MainWindow.__init__
def _commit0014_main_init(self):
    _old_commit0014_main_init(self)
    try:
        self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
        self.feedback_btn = QPushButton("Feedback")
        self.feedback_btn.setMinimumHeight(40)
        self.feedback_btn.setToolTip("Create an InSightec feedback template and manifest with current runtime context.")
        self.feedback_btn.clicked.connect(self.open_feedback)
        # Main action bar is the third item in the current RC1 layout.
        action_layout = self.layout().itemAt(2).layout() if self.layout() and self.layout().count() > 2 else None
        if action_layout is not None:
            insert_at = max(0, action_layout.count() - 3)
            action_layout.insertWidget(insert_at, self.feedback_btn)
        # Clarify the one-file ZIP limitation in this test build.
        if hasattr(self, "import_file_btn"):
            self.import_file_btn.setToolTip("Import one log directly, or select one supported file from a ZIP archive.")
    except Exception:
        write_startup_log("Commit0014 initialization failed.\n\n" + traceback.format_exc())


MainWindow.open_feedback = _commit0014_open_feedback
MainWindow.__init__ = _commit0014_main_init




# ---------------------------------------------------------------------------
# Commit0014 Revision 2: ZIP -> Smart File Discovery -> multi-file import
# - UNKNOWN is excluded from ZIP discovery.
# - Nested ZIP files are not expanded; only their file names are reported.
# - Recognized files may be selected and imported in multiples.
# - Selected records are preloaded into memory before the temporary folder is
#   deleted, so the Viewer remains usable after cleanup.
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"
def _safe_extract_zip_for_discovery(zip_path: str | Path) -> tuple[tempfile.TemporaryDirectory, Path, list[str]]:
    archive = Path(zip_path)
    holder = tempfile.TemporaryDirectory(prefix="logmerge_zip_discovery_")
    root = Path(holder.name)
    nested_zip_names: list[str] = []
    total_uncompressed = 0
    max_total = 4 * 1024 * 1024 * 1024
    max_single = 1024 * 1024 * 1024
    try:
        with zipfile.ZipFile(archive, "r") as zf:
            for info in zf.infolist():
                name = info.filename.replace("\\", "/")
                if info.is_dir():
                    continue
                if not _safe_zip_member_name(name):
                    raise ValueError(f"Unsafe ZIP member was rejected: {Path(name).name}")
                if Path(name).suffix.lower() == ".zip":
                    nested_zip_names.append(Path(name).name)
                    continue
                if info.file_size > max_single:
                    raise ValueError(f"ZIP member exceeds the 1 GB safety limit: {Path(name).name}")
                total_uncompressed += int(info.file_size)
                if total_uncompressed > max_total:
                    raise ValueError("ZIP exceeds the 4 GB total extraction safety limit.")
                target = root / name
                target.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(info, "r") as src, target.open("wb") as dst:
                    shutil.copyfileobj(src, dst, length=1024 * 1024)
        return holder, root, sorted(set(nested_zip_names), key=str.upper)
    except Exception:
        holder.cleanup()
        raise


def _parse_selected_zip_files_to_memory(files: list[str]) -> tuple[dict[str, list[LogRecord]], list[str]]:
    records_by_type: dict[str, list[LogRecord]] = {}
    failures: list[str] = []
    for file_path in files:
        path = Path(file_path)
        typ = classify_file(path, False)  # UNKNOWN is intentionally disabled.
        if not typ:
            continue
        try:
            if is_review_file(path.name):
                recs = review_rows_to_viewer_records(path)
                typ = "REVIEW"
            elif typ == "PSC":
                _, _, recs = parse_psc_file_detail(path)
            else:
                recs = parse_file(path, typ)
            records_by_type.setdefault(str(typ).upper(), []).extend(recs)
        except Exception as exc:
            failures.append(f"{path.name}: {exc}")
    for recs in records_by_type.values():
        recs.sort(key=lambda r: (r.timestamp is None, r.timestamp or datetime.max, r.filename, r.line_no))
    return records_by_type, failures


def _commit0014_r2_import_selected_file_clicked(self):
    path, _ = QFileDialog.getOpenFileName(
        self,
        "Select a log file or ZIP to import",
        self.source_edit.text() or str(Path.home()),
        "Log or ZIP files (*.log *.txt *.out *.ar *.zip);;ZIP files (*.zip);;All Files (*.*)",
    )
    if not path:
        return
    selected_path = Path(path)

    # Normal single-file import keeps the existing path.
    if selected_path.suffix.lower() != ".zip":
        out_dir = self.output_edit.text().strip() or str(selected_path.parent)
        self._run_import_task(
            "Selected File Import Progress",
            f"Importing {selected_path.name} and writing output...",
            lambda: import_selected_file_only(str(selected_path), out_dir),
            "Selected file import",
        )
        return

    holder = None
    try:
        holder, extract_root, nested_zips = _safe_extract_zip_for_discovery(selected_path)
        if nested_zips:
            QMessageBox.information(
                self,
                "ZIP files inside archive",
                "Nested ZIP files were not expanded. File names only:\n\n" + "\n".join(nested_zips),
            )

        opt = self.collect_options()
        opt.source_folder = str(extract_root)
        opt.recursive = True
        opt.include_unknown = False
        # Discovery must not inherit a prior UNKNOWN selection.
        try:
            self.chk_unknown.setChecked(False)
        except Exception:
            pass

        dlg = SmartFileDiscoveryDialog(self, opt)
        if dlg.exec() != QDialog.Accepted:
            self.status_label.setText("ZIP import cancelled in Smart File Discovery")
            return
        selected_files = list(dlg.selected_files or [])
        if not selected_files:
            QMessageBox.warning(self, "ZIP Import", "No recognized file was selected.")
            return

        progress = SoftProgressDialog(
            "ZIP Import Progress",
            "Reading selected ZIP files into memory...",
            "Cancel",
            0,
            max(1, len(selected_files)),
            self,
        )
        progress.show(); QApplication.processEvents()
        records_by_type: dict[str, list[LogRecord]] = {}
        failures: list[str] = []
        for index, f in enumerate(selected_files, start=1):
            if progress.wasCanceled():
                raise RuntimeError("ZIP import was cancelled.")
            progress.setValue(index - 1)
            progress.setLabelText(f"Reading {index}/{len(selected_files)}\n{Path(f).name}")
            QApplication.processEvents()
            partial, errs = _parse_selected_zip_files_to_memory([f])
            for typ, recs in partial.items():
                records_by_type.setdefault(typ, []).extend(recs)
            failures.extend(errs)
        progress.setValue(len(selected_files)); progress.close()

        if not records_by_type:
            detail = "\n".join(failures[:20])
            raise ValueError("No recognized ZIP file could be parsed." + (f"\n\n{detail}" if detail else ""))

        # Cache parsed records so the extracted files can be deleted immediately.
        self.zip_import_records_by_type = records_by_type
        self.zip_import_source_name = selected_path.name
        self.viewer_selected_types = set(records_by_type.keys())
        self.viewer_selected_files = []
        self.log(
            f"ZIP Smart Discovery import: {len(selected_files)} selected files, "
            f"{sum(len(v) for v in records_by_type.values()):,} rows, "
            f"types={', '.join(sorted(records_by_type))}"
        )
        if failures:
            self.log("ZIP parse warnings: " + " | ".join(failures[:10]))

        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        if hasattr(self.viewer_window, "refresh_available_sources"):
            self.viewer_window.refresh_available_sources()
        self.open_dual_viewer()
        self.status_label.setText(f"ZIP import ready: {len(selected_files)} files")
    except RuntimeError as exc:
        if "cancel" not in str(exc).lower():
            QMessageBox.warning(self, "ZIP Import", str(exc))
    except Exception as exc:
        QMessageBox.critical(self, "ZIP Import Failed", str(exc))
    finally:
        if holder is not None:
            holder.cleanup()
            self.log("ZIP temporary extraction folder deleted.")


MainWindow.import_selected_file_clicked = _commit0014_r2_import_selected_file_clicked


_old_r2_source_to_records = MultiPaneLogViewer.source_to_records

def _commit0014_r2_source_to_records(self, source, progress=None):
    parent = getattr(self, "parent_window", None)
    cache = getattr(parent, "zip_import_records_by_type", None) if parent is not None else None
    if cache:
        key = str(source).upper()
        aliases = {
            "WATERSYSTEM": "WATERSYSTEM",
            "WATER SYSTEM": "WATERSYSTEM",
            "REVIEW": "REVIEW",
            "VIMEASURE": "VIMEASURE",
        }
        key = aliases.get(key, key)
        if key == "MERGED":
            rows = [r for values in cache.values() for r in values]
        else:
            rows = list(cache.get(key, []))
        if rows:
            if progress:
                progress.setRange(0, 1)
                progress.setLabelText(f"Loading cached ZIP records for {source}: {len(rows):,} rows")
                progress.setValue(1)
                QApplication.processEvents()
            return rows
    return _old_r2_source_to_records(self, source, progress)


MultiPaneLogViewer.source_to_records = _commit0014_r2_source_to_records


_old_r2_refresh_sources = MultiPaneLogViewer.refresh_available_sources

def _commit0014_r2_refresh_available_sources(self):
    _old_r2_refresh_sources(self)
    parent = getattr(self, "parent_window", None)
    cache = getattr(parent, "zip_import_records_by_type", None) if parent is not None else None
    if not cache:
        return
    display_map = {
        "WATERSYSTEM": "WaterSystem",
        "VIMEASURE": "VIMeasure",
        "REVIEW": "Review",
        "VIMEASURE": "VIMeasure",
    }
    available = [display_map.get(t, t) for t in sorted(cache.keys())]
    available = ["Merged"] + available
    for combo in getattr(self, "sources", []):
        current = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        combo.addItems(available)
        if current in available:
            combo.setCurrentText(current)
        elif len(available) > 1:
            combo.setCurrentIndex(1)
        combo.blockSignals(False)


MultiPaneLogViewer.refresh_available_sources = _commit0014_r2_refresh_available_sources


_old_r2_main_init = MainWindow.__init__
def _commit0014_r2_main_init(self):
    _old_r2_main_init(self)
    self.zip_import_records_by_type = {}
    self.zip_import_source_name = ""
    try:
        if hasattr(self, "import_file_btn"):
            self.import_file_btn.setToolTip(
                "Import one normal log file, or open a ZIP in Smart File Discovery. "
                "Recognized files can be selected in multiples; UNKNOWN is excluded."
            )
    except Exception:
        pass


MainWindow.__init__ = _commit0014_r2_main_init




# Commit0016 Viewer & Investigation UX Stabilization --------------------------
def _commit0016_screen_geometry(widget):
    try:
        handle = widget.windowHandle()
        screen = handle.screen() if handle and handle.screen() else None
        if screen is None and widget.parentWidget() is not None:
            ph = widget.parentWidget().windowHandle()
            screen = ph.screen() if ph and ph.screen() else None
        if screen is None:
            screen = QApplication.screenAt(widget.mapToGlobal(widget.rect().center()))
        if screen is None:
            screen = QApplication.primaryScreen()
        return screen.availableGeometry()
    except Exception:
        return QApplication.primaryScreen().availableGeometry()


def _commit0016_fit_window(widget):
    geo = _commit0016_screen_geometry(widget)
    margin = 10
    widget.setGeometry(geo.x()+margin, geo.y()+margin, max(820, geo.width()-margin*2), max(520, geo.height()-margin*2))


_old_commit0016_init = MultiPaneLogViewer.__init__
def _commit0016_viewer_init(self, parent_window):
    _old_commit0016_init(self, parent_window)
    try:
        _commit0016_fit_window(self)
        if hasattr(self, 'detail'):
            self.detail.hide(); self.detail.setMaximumHeight(0); self.detail.setMinimumHeight(0)
        if hasattr(self, 'main_splitter'):
            self.main_splitter.setStretchFactor(0, 1)
        self.refresh_available_sources()
        for idx, table in enumerate(getattr(self, 'tables', [])):
            table.setContextMenuPolicy(Qt.CustomContextMenu)
            table.customContextMenuRequested.connect(lambda pos, pane=idx: self.show_row_context_menu(pane, pos))
    except Exception:
        write_startup_log('Commit0016 viewer init failed.\n\n'+traceback.format_exc())
MultiPaneLogViewer.__init__ = _commit0016_viewer_init


def _commit0016_refresh_available_sources(self):
    parent = getattr(self, 'parent_window', None)
    selected = getattr(parent, 'viewer_selected_files', None) if parent is not None else None
    found=[]
    if selected:
        seen=set()
        for f in selected:
            p=Path(f)
            typ='REVIEW' if is_review_file(p.name) else classify_file(p, False)
            if not typ or typ=='UNKNOWN':
                continue
            label={'WATERSYSTEM':'WaterSystem','VIMEASURE':'VIMeasure','REVIEW':'Review'}.get(typ,typ)
            if label not in seen:
                seen.add(label); found.append(label)
    if not found:
        found=[s for s in self.SOURCES if s not in {'Custom File','UNKNOWN'}]
    order=['Merged','WS','WaterSystem','CGA','CSA','MRSERVER','GESYS','LAIS','PSC','Review','VIMeasure']
    found=sorted(set(found), key=lambda x: order.index(x) if x in order else 999)
    if not selected and 'Merged' not in found:
        found.insert(0,'Merged')
    found.append('Custom File')
    for combo in getattr(self,'sources',[]):
        current=combo.currentText(); combo.blockSignals(True); combo.clear(); combo.addItems(found)
        if current in found: combo.setCurrentText(current)
        combo.blockSignals(False)
MultiPaneLogViewer.refresh_available_sources = _commit0016_refresh_available_sources


def _commit0016_default_columns(self, rows):
    if not rows:
        return ['Timestamp','Message']
    types={str(r.get('SourceType','')).upper() for r in rows[:200]}
    if types <= {'VIMEASURE'}:
        # Value-style view: Timestamp plus all detected numeric rails.
        preferred=['Timestamp']
        for key in rows[0].keys():
            if key not in {'Timestamp','SourceType','File','Line','Level','Category','Message','Raw','_ts'}:
                preferred.append(key)
        return preferred[:16]
    if types <= {'REVIEW','PSC'}:
        preferred=['Timestamp']+[k for k in rows[0].keys() if k not in {'Timestamp','Message','Raw','_ts','SourceType','File','Line'}]
        return preferred[:16]
    if types <= {'WATERSYSTEM'}:
        return [c for c in ['Timestamp','MainState','Error'] if any(c in r for r in rows[:50])]
    if types <= {'WS'}:
        return [c for c in ['Timestamp','Type','State','Num','Message'] if any(c in r for r in rows[:50])]
    if types <= {'CSA'} or types <= {'CGA'}:
        return [c for c in ['Timestamp','Type','Status','SubStatus','Message'] if any(c in r for r in rows[:50])]
    return [c for c in ['Timestamp','Message','Level'] if any(c in r for r in rows[:50])]
MultiPaneLogViewer.default_viewer_columns_for_rows = _commit0016_default_columns


def _commit0016_available_columns(self, rows):
    preferred=['Timestamp','Type','State','Num','Status','SubStatus','Message','NumericValue','Unit','MainState','Error','Level','SourceType','Category','File','Line']
    found=[c for c in preferred if any(c in r for r in rows[:300])]
    for row in rows[:1000]:
        for key in row.keys():
            if key not in found and key not in {'Raw','_ts'}:
                found.append(key)
    return found
MultiPaneLogViewer.all_available_columns_for_rows = _commit0016_available_columns


def _commit0016_apply_filters(self, side):
    idx=self.side_index(side)
    try: start,end=self.current_viewer_time_range()
    except Exception as exc:
        QMessageBox.warning(self,'Viewer Time Range',str(exc)); return
    base=list(self.all_rows[idx])
    rows=[]
    for row in base:
        ts=row.get('_ts')
        if start or end:
            if not isinstance(ts,datetime): continue
            if start and ts<start: continue
            if end and ts>end: continue
        rows.append(row)
    # Apply visible pane text filter, including structured columns.
    if hasattr(self,'foundation_filter_edits') and idx < len(self.foundation_filter_edits):
        expr=self.foundation_filter_edits[idx].text().strip()
        if expr:
            if '=' in expr:
                key,value=expr.split('=',1); key=key.strip().lower(); value=value.strip().lower()
                rows=[r for r in rows if str(next((v for k,v in r.items() if k.lower()==key), '')).lower()==value]
            else:
                term=expr.lower(); rows=[r for r in rows if term in ' '.join(str(v) for k,v in r.items() if k not in {'_ts'}).lower()]
    ts_index=[(r['_ts'],i) for i,r in enumerate(rows) if isinstance(r.get('_ts'),datetime)]; ts_index.sort(key=lambda x:x[0])
    cols=self.pane_columns_for_rows(idx,rows)
    self.models[idx].set_rows(rows,cols); self.apply_table_column_widths(idx); self.ts_indexes[idx]=ts_index; self._sync_aliases()
    self.log(f"Viewer filter {self.pane_name(idx)}: {len(rows)}/{len(base)} rows")
MultiPaneLogViewer.apply_view_filters = _commit0016_apply_filters


_old_commit0016_load_pane = MultiPaneLogViewer.load_pane
def _commit0016_load_pane(self, side):
    # Commit0043: no automatic source-specific startup filter.
    return _old_commit0016_load_pane(self, side)
MultiPaneLogViewer.load_pane = _commit0016_load_pane


def _commit0016_context_menu(self, pane, pos):
    table=self.tables[pane]; model=self.models[pane]
    idx=table.indexAt(pos)
    if not idx.isValid(): return
    table.selectRow(idx.row()); row=model.row_at(idx.row()) or {}
    menu=QMenu(table)
    def copy_text(value): QApplication.clipboard().setText(str(value or ''))
    a=menu.addAction('Copy Cell'); a.triggered.connect(lambda: copy_text(model.data(idx,Qt.DisplayRole)))
    a=menu.addAction('Copy Row'); a.triggered.connect(lambda: copy_text('\t'.join(str(row.get(c,'')) for c in model.columns)))
    a=menu.addAction('Copy Timestamp'); a.triggered.connect(lambda: copy_text(row.get('Timestamp','')))
    a=menu.addAction('Copy Message / Value'); a.triggered.connect(lambda: copy_text(row.get('Message', row.get('Value',''))))
    menu.addSeparator()
    a=menu.addAction('Copy Rule Text'); a.triggered.connect(lambda: self.copy_rule_text(pane))
    a=menu.addAction('Add to Noise Rule'); a.triggered.connect(lambda: self.approve_selected_noise(pane))
    menu.addSeparator()
    a=menu.addAction('Filter by This Value')
    def set_filter():
        if hasattr(self,'foundation_filter_edits') and pane < len(self.foundation_filter_edits):
            col=model.columns[idx.column()]; self.foundation_filter_edits[pane].setText(f"{col}={model.data(idx,Qt.DisplayRole)}"); self.apply_view_filters(pane)
    a.triggered.connect(set_filter)
    a=menu.addAction('Clear Pane Filter')
    def clear_filter():
        if hasattr(self,'foundation_filter_edits') and pane < len(self.foundation_filter_edits): self.foundation_filter_edits[pane].clear(); self.apply_view_filters(pane)
    a.triggered.connect(clear_filter)
    menu.exec(table.viewport().mapToGlobal(pos))
MultiPaneLogViewer.show_row_context_menu = _commit0016_context_menu


# ---------------------------------------------------------------------------
# Commit0017: Simplified Source Workflow + Screen-Fit Smart Discovery
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"


def _commit0017_screen_for(widget):
    try:
        parent = widget.parentWidget()
        if parent is not None and parent.windowHandle() and parent.windowHandle().screen():
            return parent.windowHandle().screen()
        if widget.windowHandle() and widget.windowHandle().screen():
            return widget.windowHandle().screen()
        return QApplication.screenAt(widget.mapToGlobal(widget.rect().center())) or QApplication.primaryScreen()
    except Exception:
        return QApplication.primaryScreen()


def _commit0017_discovery_build_ui(self):
    root = QVBoxLayout(self)
    root.setContentsMargins(8, 8, 8, 8)
    root.setSpacing(6)

    info = QLabel(
        "Detected file types and actual log timestamps are used here. "
        "The main screen no longer requires File Type or Date Range selection."
    )
    info.setWordWrap(True)
    root.addWidget(info)

    type_box = QGroupBox("Detected File Types")
    self.type_layout = QGridLayout(type_box)
    self.type_layout.setContentsMargins(8, 6, 8, 6)
    self.type_layout.setHorizontalSpacing(14)
    self.type_layout.setVerticalSpacing(4)
    self.type_checks = {}
    self.detected_all_btn = QPushButton("All")
    self.detected_clear_btn = QPushButton("Clear")
    self.detected_all_btn.clicked.connect(lambda: self.set_all_detected_types(True))
    self.detected_clear_btn.clicked.connect(lambda: self.set_all_detected_types(False))
    root.addWidget(type_box)

    range_box = QGroupBox("Actual Data Range")
    rg = QGridLayout(range_box)
    self.start_combo = QComboBox(); self.start_combo.setMinimumWidth(280)
    self.end_combo = QComboBox(); self.end_combo.setMinimumWidth(280)
    self.custom_date_btn = QPushButton("Custom Calendar...")
    rg.addWidget(QLabel("Start"), 0, 0)
    rg.addWidget(self.start_combo, 0, 1)
    rg.addWidget(QLabel("End"), 1, 0)
    rg.addWidget(self.end_combo, 1, 1)
    rg.addWidget(self.custom_date_btn, 0, 2, 2, 1)
    rg.setColumnStretch(1, 1)
    root.addWidget(range_box)

    self.enable_start = QCheckBox("Start")
    self.start_date = QDateEdit(QDate.currentDate()); self.start_date.setCalendarPopup(True); self.start_date.setDisplayFormat("yyyy/MM/dd")
    self.enable_end = QCheckBox("End")
    self.end_date = QDateEdit(QDate.currentDate()); self.end_date.setCalendarPopup(True); self.end_date.setDisplayFormat("yyyy/MM/dd")
    for widget in (self.enable_start, self.start_date, self.enable_end, self.end_date):
        widget.hide()
    self.start_combo.currentIndexChanged.connect(self.apply_filter)
    self.end_combo.currentIndexChanged.connect(self.apply_filter)
    self.custom_date_btn.clicked.connect(self._choose_custom_date_range)

    self.summary_label = QLabel("Ready")
    root.addWidget(self.summary_label)

    self.table = QTableWidget(0, 8)
    self.table.setHorizontalHeaderLabels(["Use", "Type", "Start", "End", "Rows", "Size MB", "File", "Error"])
    self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
    self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
    self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
    self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
    self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
    self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
    self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
    self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
    self.table.itemChanged.connect(self._file_item_changed)
    root.addWidget(self.table, 1)

    btns = QHBoxLayout()
    self.start_btn = QPushButton("▶ START")
    self.start_btn.setMinimumSize(145, 42)
    self.start_btn.setStyleSheet("QPushButton{font-size:14pt;font-weight:700;background:#16a34a;color:white;border-radius:10px;padding:7px 18px;} QPushButton:hover{background:#15803d;}")
    self.start_btn.clicked.connect(self.accept_with_selection)
    self.select_all_btn = QPushButton("All Files"); self.select_all_btn.clicked.connect(lambda: self.set_all_files(True))
    self.clear_all_btn = QPushButton("Clear Files"); self.clear_all_btn.clicked.connect(lambda: self.set_all_files(False))
    self.cancel_btn = QPushButton("Cancel"); self.cancel_btn.clicked.connect(self.reject)
    btns.addWidget(self.start_btn); btns.addWidget(self.select_all_btn); btns.addWidget(self.clear_all_btn)
    btns.addStretch(1); btns.addWidget(self.cancel_btn)
    root.addLayout(btns)


def _commit0017_populate_type_checks(self):
    while self.type_layout.count():
        item = self.type_layout.takeAt(0)
        widget = item.widget()
        if widget:
            widget.deleteLater()
    self.type_checks.clear()
    counts = {}
    for item in self.discovered:
        counts[item.log_type] = counts.get(item.log_type, 0) + 1
    preferred = ["WS", "WATERSYSTEM", "VIMEASURE", "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "REVIEW"]
    ordered = [name for name in preferred if name in counts] + sorted(name for name in counts if name not in preferred and name != "UNKNOWN")
    columns = 4
    for index, typ in enumerate(ordered):
        checkbox = QCheckBox(f"{typ} ({counts[typ]})")
        checkbox.setChecked(True)
        checkbox.stateChanged.connect(self.apply_filter)
        self.type_checks[typ] = checkbox
        self.type_layout.addWidget(checkbox, index // columns, index % columns)
    row = (len(ordered) + columns - 1) // columns
    self.type_layout.addWidget(self.detected_all_btn, row, 0)
    self.type_layout.addWidget(self.detected_clear_btn, row, 1)
    self.type_layout.setColumnStretch(columns, 1)


SmartFileDiscoveryDialog._build_ui = _commit0017_discovery_build_ui
SmartFileDiscoveryDialog.populate_type_checks = _commit0017_populate_type_checks

_old_commit0017_discovery_init = SmartFileDiscoveryDialog.__init__
def _commit0017_discovery_init(self, parent, base_options):
    _old_commit0017_discovery_init(self, parent, base_options)
    screen = _commit0017_screen_for(self)
    geo = screen.availableGeometry()
    margin = 12
    width = min(1050, max(760, geo.width() - margin * 2))
    height = min(820, max(560, geo.height() - margin * 2))
    x = geo.x() + max(margin, (geo.width() - width) // 2)
    y = geo.y() + max(margin, (geo.height() - height) // 2)
    self.setGeometry(x, y, width, height)
    self.setMinimumSize(min(760, width), min(560, height))
SmartFileDiscoveryDialog.__init__ = _commit0017_discovery_init


_old_commit0017_build_ui = MainWindow.build_ui
def _commit0017_build_ui(self):
    _old_commit0017_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")

    # File Type and Date selection now belong only to Smart File Discovery.
    for box in self.findChildren(QGroupBox):
        if box.title() in {"Date Range", "Log Types"}:
            box.setVisible(False)

    # Keep compatibility controls alive but neutral.
    self.use_start.setChecked(False); self.use_end.setChecked(False)
    for name in ("chk_ws", "chk_watersystem", "chk_cga", "chk_csa", "chk_mrserver", "chk_gesys", "chk_lais", "chk_psc"):
        widget = getattr(self, name, None)
        if widget is not None:
            widget.setChecked(True)
    self.chk_unknown.setChecked(False)

    folder_box = next((box for box in self.findChildren(QGroupBox) if box.title() == "Folders"), None)
    if folder_box is not None and isinstance(folder_box.layout(), QFormLayout):
        self.source_mode_combo = QComboBox()
        self.source_mode_combo.addItems(["Folder", "ZIP File", "Project"])
        self.source_mode_combo.setToolTip("Select a folder, ZIP archive, or saved project package.")
        folder_box.layout().insertRow(0, "Source Type", self.source_mode_combo)
        self.source_mode_combo.currentTextChanged.connect(self._commit0017_source_mode_changed)
        try:
            self.source_btn.clicked.disconnect()
        except Exception:
            pass
        self.source_btn.clicked.connect(self._commit0017_pick_source)
        self._commit0017_source_mode_changed("Folder")


def _commit0017_source_mode_changed(self, mode):
    if mode == "ZIP File":
        self.source_btn.setText("Browse ZIP...")
        self.recursive_chk.setVisible(False)
    elif mode == "Project":
        self.source_btn.setText("Open Project...")
        self.recursive_chk.setVisible(False)
    else:
        self.source_btn.setText("Browse Folder...")
        self.recursive_chk.setVisible(True)


def _commit0017_pick_source(self):
    mode = self.source_mode_combo.currentText() if hasattr(self, "source_mode_combo") else "Folder"
    current = self.source_edit.text().strip() or str(Path.home())
    if mode == "ZIP File":
        path, _ = QFileDialog.getOpenFileName(self, "Select Source ZIP", current, "ZIP archives (*.zip);;All Files (*.*)")
        if path:
            self.source_edit.setText(path)
            if not self.output_edit.text().strip():
                self.output_edit.setText(str(Path(path).parent))
    elif mode == "Project":
        self.import_project_clicked()
        self.source_mode_combo.setCurrentText("Folder")
    else:
        path = QFileDialog.getExistingDirectory(self, "Select Source Log Folder", current)
        if path:
            self.source_edit.setText(path)
            if not self.output_edit.text().strip():
                self.output_edit.setText(path)


MainWindow._commit0017_source_mode_changed = _commit0017_source_mode_changed
MainWindow._commit0017_pick_source = _commit0017_pick_source
MainWindow.build_ui = _commit0017_build_ui


_old_commit0017_start = MainWindow.start_import_clicked
def _commit0017_start_import(self):
    mode = self.source_mode_combo.currentText() if hasattr(self, "source_mode_combo") else "Folder"
    source = self.source_edit.text().strip()
    if mode != "ZIP File" and not source.lower().endswith(".zip"):
        return _old_commit0017_start(self)
    if not source or not Path(source).is_file():
        QMessageBox.warning(self, APP_TITLE, "Select a valid Source ZIP file.")
        return

    holder = None
    try:
        holder, extract_root, nested = _safe_extract_zip_for_discovery(source)
        if nested:
            QMessageBox.information(self, "ZIP files inside archive", "Nested ZIP files were not expanded:\n\n" + "\n".join(nested))
        options = self.collect_options()
        options.source_folder = str(extract_root)
        options.recursive = True
        options.include_unknown = False
        dialog = SmartFileDiscoveryDialog(self, options)
        if dialog.exec() != QDialog.Accepted:
            self.status_label.setText("ZIP import cancelled")
            return
        selected = list(dialog.selected_files or [])
        if not selected:
            QMessageBox.warning(self, APP_TITLE, "No recognized file was selected.")
            return
        progress = SoftProgressDialog("ZIP Import", "Reading selected files into memory...", "Cancel", 0, len(selected), self)
        progress.show(); QApplication.processEvents()
        cache = {}; failures = []
        for index, path in enumerate(selected, 1):
            if progress.wasCanceled():
                return
            progress.setValue(index - 1); progress.setLabelText(f"Reading {index}/{len(selected)}\n{Path(path).name}")
            QApplication.processEvents()
            partial, errors = _parse_selected_zip_files_to_memory([path])
            for typ, records in partial.items():
                cache.setdefault(typ, []).extend(records)
            failures.extend(errors)
        progress.setValue(len(selected)); progress.close()
        if not cache:
            raise ValueError("No selected ZIP file could be parsed." + (("\n\n" + "\n".join(failures[:20])) if failures else ""))
        self.zip_import_records_by_type = cache
        self.zip_import_source_name = Path(source).name
        self.viewer_selected_files = []
        self.viewer_selected_types = set(cache)
        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)
        self.viewer_window.refresh_available_sources()
        self.open_dual_viewer()
        self.status_label.setText(f"ZIP loaded: {len(selected)} files")
    except Exception:
        QMessageBox.critical(self, APP_TITLE, traceback.format_exc())
    finally:
        if holder is not None:
            holder.cleanup()

MainWindow.start_import_clicked = _commit0017_start_import


_old_commit0017_merge = MainWindow.run_clicked
def _commit0017_merge(self):
    mode = self.source_mode_combo.currentText() if hasattr(self, "source_mode_combo") else "Folder"
    source = self.source_edit.text().strip()
    if mode != "ZIP File" and not source.lower().endswith(".zip"):
        return _old_commit0017_merge(self)
    if not source or not Path(source).is_file():
        QMessageBox.warning(self, APP_TITLE, "Select a valid Source ZIP file.")
        return
    try:
        holder, extract_root, nested = _safe_extract_zip_for_discovery(source)
        if nested:
            QMessageBox.information(self, "ZIP files inside archive", "Nested ZIP files were not expanded:\n\n" + "\n".join(nested))
        self._commit0017_zip_holder = holder
        self._commit0017_zip_original_source = source
        self.source_edit.setText(str(extract_root))
        self.source_mode_combo.setCurrentText("Folder")
        _old_commit0017_merge(self)
    except Exception:
        holder = getattr(self, "_commit0017_zip_holder", None)
        if holder:
            holder.cleanup()
        self._commit0017_zip_holder = None
        QMessageBox.critical(self, APP_TITLE, traceback.format_exc())

MainWindow.run_clicked = _commit0017_merge


def _commit0017_cleanup_zip(self):
    holder = getattr(self, "_commit0017_zip_holder", None)
    original = getattr(self, "_commit0017_zip_original_source", "")
    if holder is not None:
        try:
            holder.cleanup()
        except Exception:
            pass
    self._commit0017_zip_holder = None
    if original:
        self.source_edit.setText(original)
        if hasattr(self, "source_mode_combo"):
            self.source_mode_combo.setCurrentText("ZIP File")

_old_commit0017_finished = MainWindow.on_finished
def _commit0017_finished(self, result):
    try:
        return _old_commit0017_finished(self, result)
    finally:
        _commit0017_cleanup_zip(self)
MainWindow.on_finished = _commit0017_finished

_old_commit0017_failed = MainWindow.on_failed
def _commit0017_failed(self, message):
    try:
        return _old_commit0017_failed(self, message)
    finally:
        _commit0017_cleanup_zip(self)
MainWindow.on_failed = _commit0017_failed



# ---------------------------------------------------------------------------
# Commit0018: clear drag-and-drop source area
# ---------------------------------------------------------------------------

class Commit0018DropZone(QFrame):
    filesDropped = Signal(list)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setObjectName("commit0018DropZone")
        self.setMinimumHeight(145)
        self.setCursor(Qt.PointingHandCursor)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 14, 18, 14)
        layout.setSpacing(7)

        self.icon_label = QLabel("▼")
        self.icon_label.setAlignment(Qt.AlignCenter)
        self.icon_label.setStyleSheet("font-size: 22px; font-weight: 700;")

        self.title_label = QLabel("Drop Folder / ZIP / Log Files Here")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: 700;")

        self.detail_label = QLabel(
            "Folder, one ZIP archive, or multiple log files\n"
            "Files are passed to Smart File Discovery before import."
        )
        self.detail_label.setAlignment(Qt.AlignCenter)
        self.detail_label.setWordWrap(True)

        button_row = QHBoxLayout()
        button_row.addStretch(1)
        self.folder_button = QPushButton("Browse Folder...")
        self.zip_button = QPushButton("Browse ZIP...")
        self.files_button = QPushButton("Browse Files...")
        button_row.addWidget(self.folder_button)
        button_row.addWidget(self.zip_button)
        button_row.addWidget(self.files_button)
        button_row.addStretch(1)

        layout.addWidget(self.icon_label)
        layout.addWidget(self.title_label)
        layout.addWidget(self.detail_label)
        layout.addLayout(button_row)

        self._set_normal_style()

    def _set_normal_style(self):
        self.setStyleSheet(
            """
            QFrame#commit0018DropZone {
                border: 2px dashed #3b82f6;
                border-radius: 12px;
                background: #eff6ff;
            }
            QFrame#commit0018DropZone QLabel {
                color: #1e3a8a;
                background: transparent;
            }
            QFrame#commit0018DropZone QPushButton {
                padding: 6px 12px;
            }
            """
        )
        self.icon_label.setText("▼")
        self.title_label.setText("Drop Folder / ZIP / Log Files Here")

    def _set_active_style(self):
        self.setStyleSheet(
            """
            QFrame#commit0018DropZone {
                border: 3px solid #16a34a;
                border-radius: 12px;
                background: #dcfce7;
            }
            QFrame#commit0018DropZone QLabel {
                color: #14532d;
                background: transparent;
            }
            QFrame#commit0018DropZone QPushButton {
                padding: 6px 12px;
            }
            """
        )
        self.icon_label.setText("↓")
        self.title_label.setText("Release to Import")

    @staticmethod
    def _local_paths(event):
        if not event.mimeData().hasUrls():
            return []
        paths = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                value = url.toLocalFile()
                if value:
                    paths.append(value)
        return paths

    def dragEnterEvent(self, event):
        paths = self._local_paths(event)
        if paths:
            event.acceptProposedAction()
            self._set_active_style()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if self._local_paths(event):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self._set_normal_style()
        event.accept()

    def dropEvent(self, event):
        paths = self._local_paths(event)
        self._set_normal_style()
        if not paths:
            event.ignore()
            return
        event.acceptProposedAction()
        self.filesDropped.emit(paths)


def _commit0018_cleanup_staged_drop(self):
    holder = getattr(self, "_commit0018_drop_holder", None)
    if holder is not None:
        try:
            holder.cleanup()
        except Exception:
            pass
    self._commit0018_drop_holder = None


def _commit0018_set_source_mode(self, mode):
    if hasattr(self, "source_mode_combo"):
        self.source_mode_combo.setCurrentText(mode)


def _commit0018_choose_folder(self):
    current = self.source_edit.text().strip() or str(Path.home())
    path = QFileDialog.getExistingDirectory(self, "Select Source Log Folder", current)
    if path:
        _commit0018_cleanup_staged_drop(self)
        _commit0018_set_source_mode(self, "Folder")
        self.source_edit.setText(path)
        self.recursive_chk.setChecked(True)
        if not self.output_edit.text().strip():
            self.output_edit.setText(path)
        self.status_label.setText("Folder selected. Press START or MERGE.")


def _commit0018_choose_zip(self):
    current = self.source_edit.text().strip() or str(Path.home())
    path, _ = QFileDialog.getOpenFileName(
        self,
        "Select Source ZIP",
        current,
        "ZIP archives (*.zip);;All Files (*.*)",
    )
    if path:
        _commit0018_cleanup_staged_drop(self)
        _commit0018_set_source_mode(self, "ZIP File")
        self.source_edit.setText(path)
        if not self.output_edit.text().strip():
            self.output_edit.setText(str(Path(path).parent))
        self.status_label.setText("ZIP selected. Press START or MERGE.")


def _commit0018_choose_files(self):
    current = self.source_edit.text().strip() or str(Path.home())
    paths, _ = QFileDialog.getOpenFileNames(
        self,
        "Select Log Files",
        current,
        "Log files (*.log *.txt *.out *.ar);;All Files (*.*)",
    )
    if paths:
        self._commit0018_handle_drop(paths)


def _commit0018_stage_files(self, paths):
    _commit0018_cleanup_staged_drop(self)
    holder = tempfile.TemporaryDirectory(prefix="logmerge_dropped_files_")
    root = Path(holder.name)

    copied = []
    used_names = set()
    for original in paths:
        source = Path(original)
        if not source.is_file():
            continue

        name = source.name
        stem = source.stem
        suffix = source.suffix
        counter = 1
        while name.lower() in used_names or (root / name).exists():
            name = f"{stem}_{counter}{suffix}"
            counter += 1

        target = root / name
        shutil.copy2(source, target)
        used_names.add(name.lower())
        copied.append(target)

    if not copied:
        holder.cleanup()
        raise ValueError("No readable file was dropped.")

    self._commit0018_drop_holder = holder
    _commit0018_set_source_mode(self, "Folder")
    self.source_edit.setText(str(root))
    self.recursive_chk.setChecked(False)

    common_parent = Path(paths[0]).parent
    if not self.output_edit.text().strip():
        self.output_edit.setText(str(common_parent))

    return copied


def _commit0018_handle_drop(self, paths):
    clean = [str(Path(value)) for value in paths if str(value).strip()]
    if not clean:
        return

    directories = [Path(value) for value in clean if Path(value).is_dir()]
    files = [Path(value) for value in clean if Path(value).is_file()]
    zips = [path for path in files if path.suffix.lower() == ".zip"]
    regular_files = [path for path in files if path.suffix.lower() != ".zip"]

    try:
        if directories:
            if len(clean) != 1:
                QMessageBox.warning(
                    self,
                    APP_TITLE,
                    "Drop one folder at a time.\n"
                    "For multiple inputs, drop the log files themselves.",
                )
                return

            _commit0018_cleanup_staged_drop(self)
            folder = directories[0]
            _commit0018_set_source_mode(self, "Folder")
            self.source_edit.setText(str(folder))
            self.recursive_chk.setChecked(True)
            if not self.output_edit.text().strip():
                self.output_edit.setText(str(folder))
            self.status_label.setText(
                f"Folder dropped: {folder.name}. Press START or MERGE."
            )
            return

        if zips:
            if len(clean) != 1:
                QMessageBox.warning(
                    self,
                    APP_TITLE,
                    "Drop one ZIP archive at a time.\n"
                    "ZIP files cannot be mixed with ordinary log files.",
                )
                return

            _commit0018_cleanup_staged_drop(self)
            archive = zips[0]
            _commit0018_set_source_mode(self, "ZIP File")
            self.source_edit.setText(str(archive))
            if not self.output_edit.text().strip():
                self.output_edit.setText(str(archive.parent))
            self.status_label.setText(
                f"ZIP dropped: {archive.name}. Press START or MERGE."
            )
            return

        if regular_files:
            copied = _commit0018_stage_files(self, regular_files)
            self.status_label.setText(
                f"{len(copied)} files dropped and staged. Press START or MERGE."
            )
            return

        QMessageBox.warning(
            self,
            APP_TITLE,
            "The dropped item is not a supported folder, ZIP, or log file.",
        )

    except Exception:
        QMessageBox.critical(self, APP_TITLE, traceback.format_exc())


_old_commit0018_build_ui = MainWindow.build_ui
def _commit0018_build_ui(self):
    _old_commit0018_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
    self.setAcceptDrops(True)

    self.commit0018_drop_zone = Commit0018DropZone(self)
    self.commit0018_drop_zone.filesDropped.connect(self._commit0018_handle_drop)
    self.commit0018_drop_zone.folder_button.clicked.connect(self._commit0018_choose_folder)
    self.commit0018_drop_zone.zip_button.clicked.connect(self._commit0018_choose_zip)
    self.commit0018_drop_zone.files_button.clicked.connect(self._commit0018_choose_files)

    root = self.layout()
    folder_box = next(
        (box for box in self.findChildren(QGroupBox) if box.title() == "Folders"),
        None,
    )
    insert_index = 3
    if folder_box is not None:
        for index in range(root.count()):
            if root.itemAt(index).widget() is folder_box:
                insert_index = index + 1
                break
    root.insertWidget(insert_index, self.commit0018_drop_zone)

    self.commit0018_drop_zone.setToolTip(
        "Drop a source folder, one ZIP archive, or multiple log files here."
    )


def _commit0018_window_drag_enter(self, event):
    if event.mimeData().hasUrls():
        event.acceptProposedAction()
        if hasattr(self, "commit0018_drop_zone"):
            self.commit0018_drop_zone._set_active_style()
    else:
        event.ignore()


def _commit0018_window_drag_leave(self, event):
    if hasattr(self, "commit0018_drop_zone"):
        self.commit0018_drop_zone._set_normal_style()
    event.accept()


def _commit0018_window_drop(self, event):
    paths = Commit0018DropZone._local_paths(event)
    if hasattr(self, "commit0018_drop_zone"):
        self.commit0018_drop_zone._set_normal_style()
    if paths:
        event.acceptProposedAction()
        self._commit0018_handle_drop(paths)
    else:
        event.ignore()


_old_commit0018_close_event = getattr(MainWindow, "closeEvent", None)
def _commit0018_close_event(self, event):
    _commit0018_cleanup_staged_drop(self)
    if _old_commit0018_close_event is not None:
        return _old_commit0018_close_event(self, event)
    event.accept()


MainWindow._commit0018_choose_folder = _commit0018_choose_folder
MainWindow._commit0018_choose_zip = _commit0018_choose_zip
MainWindow._commit0018_choose_files = _commit0018_choose_files
MainWindow._commit0018_handle_drop = _commit0018_handle_drop
MainWindow._commit0018_stage_files = _commit0018_stage_files
MainWindow.build_ui = _commit0018_build_ui
MainWindow.dragEnterEvent = _commit0018_window_drag_enter
MainWindow.dragLeaveEvent = _commit0018_window_drag_leave
MainWindow.dropEvent = _commit0018_window_drop
MainWindow.closeEvent = _commit0018_close_event


# ---------------------------------------------------------------------------
# Commit0018 R2: Output auto-selection + common same-monitor window placement
# ---------------------------------------------------------------------------

def _c18r2_output_for_paths(paths):
    clean = [Path(value).resolve() for value in paths if str(value).strip()]
    if not clean:
        return None, False

    if len(clean) == 1 and clean[0].is_dir():
        return clean[0], False

    parents = [
        item.parent if item.is_file() else item
        for item in clean
    ]
    first = parents[0]
    mixed = any(parent != first for parent in parents[1:])
    return first, mixed


def _c18r2_set_output_from_paths(self, paths, source_description="source"):
    output, mixed = _c18r2_output_for_paths(paths)
    if output is None:
        return

    self.output_edit.setText(str(output))

    if mixed:
        self.status_label.setText(
            f"{source_description} selected from multiple folders. "
            f"Output Folder was set to the first item's folder: {output}"
        )
    else:
        self.status_label.setText(
            f"{source_description} selected. Output Folder was set to: {output}"
        )


def _c18r2_screen_for(parent=None, widget=None):
    candidates = []

    if parent is not None:
        candidates.append(parent)
    if widget is not None:
        candidates.append(widget)

    for candidate in candidates:
        try:
            top = candidate.window()
            handle = top.windowHandle()
            if handle is not None and handle.screen() is not None:
                return handle.screen()
        except Exception:
            pass

    for candidate in candidates:
        try:
            point = candidate.mapToGlobal(candidate.rect().center())
            screen = QApplication.screenAt(point)
            if screen is not None:
                return screen
        except Exception:
            pass

    return QApplication.primaryScreen()


def _c18r2_clamp_rect(rect, available, margin=10):
    max_width = max(480, available.width() - margin * 2)
    max_height = max(360, available.height() - margin * 2)

    width = min(max(rect.width(), 480), max_width)
    height = min(max(rect.height(), 360), max_height)

    left = available.left() + margin
    top = available.top() + margin
    right_limit = available.right() - margin - width + 1
    bottom_limit = available.bottom() - margin - height + 1

    x = min(max(rect.x(), left), max(left, right_limit))
    y = min(max(rect.y(), top), max(top, bottom_limit))

    return QRect(x, y, width, height)


def _c18r2_place_window(
    widget,
    parent=None,
    width_ratio=0.88,
    height_ratio=0.88,
    minimum=(760, 520),
    maximum=None,
    center=True,
):
    try:
        screen = _c18r2_screen_for(parent=parent, widget=widget)
        available = screen.availableGeometry()
        margin = 10

        target_width = max(minimum[0], int(available.width() * width_ratio))
        target_height = max(minimum[1], int(available.height() * height_ratio))

        if maximum is not None:
            target_width = min(target_width, maximum[0])
            target_height = min(target_height, maximum[1])

        target_width = min(target_width, max(480, available.width() - margin * 2))
        target_height = min(target_height, max(360, available.height() - margin * 2))

        if center:
            x = available.x() + (available.width() - target_width) // 2
            y = available.y() + (available.height() - target_height) // 2
        else:
            current = widget.frameGeometry()
            x, y = current.x(), current.y()

        fitted = _c18r2_clamp_rect(
            QRect(x, y, target_width, target_height),
            available,
            margin,
        )
        widget.setGeometry(fitted)
    except Exception:
        write_startup_log(
            "Commit0018 R2 window placement failed.\n\n"
            + traceback.format_exc()
        )


def _c18r2_center_dialog(widget, parent=None, preferred_size=None):
    try:
        screen = _c18r2_screen_for(parent=parent, widget=widget)
        available = screen.availableGeometry()
        margin = 14

        if preferred_size:
            width, height = preferred_size
        else:
            hint = widget.sizeHint()
            width = max(widget.width(), hint.width())
            height = max(widget.height(), hint.height())

        width = min(max(width, 360), max(360, available.width() - margin * 2))
        height = min(max(height, 160), max(160, available.height() - margin * 2))
        x = available.x() + (available.width() - width) // 2
        y = available.y() + (available.height() - height) // 2
        widget.setGeometry(_c18r2_clamp_rect(QRect(x, y, width, height), available, margin))
    except Exception:
        pass


# --- Output Folder rules -----------------------------------------------------

_old_c18r2_choose_folder = MainWindow._commit0018_choose_folder
def _c18r2_choose_folder(self):
    current = self.source_edit.text().strip() or str(Path.home())
    path = QFileDialog.getExistingDirectory(
        self,
        "Select Source Log Folder",
        current,
    )
    if not path:
        return

    _commit0018_cleanup_staged_drop(self)
    _commit0018_set_source_mode(self, "Folder")
    self.source_edit.setText(path)
    self.recursive_chk.setChecked(True)
    _c18r2_set_output_from_paths(self, [path], "Folder")


_old_c18r2_choose_zip = MainWindow._commit0018_choose_zip
def _c18r2_choose_zip(self):
    current = self.source_edit.text().strip() or str(Path.home())
    path, _ = QFileDialog.getOpenFileName(
        self,
        "Select Source ZIP",
        current,
        "ZIP archives (*.zip);;All Files (*.*)",
    )
    if not path:
        return

    _commit0018_cleanup_staged_drop(self)
    _commit0018_set_source_mode(self, "ZIP File")
    self.source_edit.setText(path)
    _c18r2_set_output_from_paths(self, [path], "ZIP")


_old_c18r2_stage_files = MainWindow._commit0018_stage_files
def _c18r2_stage_files(self, paths):
    copied = _old_c18r2_stage_files(self, paths)
    _c18r2_set_output_from_paths(self, paths, f"{len(copied)} file(s)")
    return copied


_old_c18r2_handle_drop = MainWindow._commit0018_handle_drop
def _c18r2_handle_drop(self, paths):
    clean = [str(Path(value)) for value in paths if str(value).strip()]
    if not clean:
        return

    directories = [Path(value) for value in clean if Path(value).is_dir()]
    files = [Path(value) for value in clean if Path(value).is_file()]
    zips = [path for path in files if path.suffix.lower() == ".zip"]
    regular_files = [path for path in files if path.suffix.lower() != ".zip"]

    try:
        if directories:
            if len(clean) != 1:
                QMessageBox.warning(
                    self,
                    APP_TITLE,
                    "Drop one folder at a time.\n"
                    "For multiple inputs, drop the log files themselves.",
                )
                return

            _commit0018_cleanup_staged_drop(self)
            folder = directories[0]
            _commit0018_set_source_mode(self, "Folder")
            self.source_edit.setText(str(folder))
            self.recursive_chk.setChecked(True)
            _c18r2_set_output_from_paths(self, [folder], "Dropped folder")
            return

        if zips:
            if len(clean) != 1:
                QMessageBox.warning(
                    self,
                    APP_TITLE,
                    "Drop one ZIP archive at a time.\n"
                    "ZIP files cannot be mixed with ordinary log files.",
                )
                return

            _commit0018_cleanup_staged_drop(self)
            archive = zips[0]
            _commit0018_set_source_mode(self, "ZIP File")
            self.source_edit.setText(str(archive))
            _c18r2_set_output_from_paths(self, [archive], "Dropped ZIP")
            return

        if regular_files:
            copied = _c18r2_stage_files(self, regular_files)
            self.status_label.setText(
                f"{len(copied)} file(s) dropped and staged. "
                + self.status_label.text()
            )
            return

        QMessageBox.warning(
            self,
            APP_TITLE,
            "The dropped item is not a supported folder, ZIP, or log file.",
        )
    except Exception:
        QMessageBox.critical(self, APP_TITLE, traceback.format_exc())


MainWindow._commit0018_choose_folder = _c18r2_choose_folder
MainWindow._commit0018_choose_zip = _c18r2_choose_zip
MainWindow._commit0018_stage_files = _c18r2_stage_files
MainWindow._commit0018_handle_drop = _c18r2_handle_drop


# Reconnect drop-zone buttons because Commit0018 connected bound methods
# during UI construction.
_old_c18r2_build_ui = MainWindow.build_ui
def _c18r2_build_ui(self):
    _old_c18r2_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")

    zone = getattr(self, "commit0018_drop_zone", None)
    if zone is not None:
        for signal in (
            zone.folder_button.clicked,
            zone.zip_button.clicked,
            zone.files_button.clicked,
            zone.filesDropped,
        ):
            try:
                signal.disconnect()
            except Exception:
                pass

        zone.folder_button.clicked.connect(self._commit0018_choose_folder)
        zone.zip_button.clicked.connect(self._commit0018_choose_zip)
        zone.files_button.clicked.connect(self._commit0018_choose_files)
        zone.filesDropped.connect(self._commit0018_handle_drop)

    # Main window itself is kept on its current monitor and inside available area.
    QTimer.singleShot(
        0,
        lambda: _c18r2_place_window(
            self,
            parent=self,
            width_ratio=0.82,
            height_ratio=0.88,
            minimum=(900, 650),
            maximum=(1500, 1000),
        ),
    )


MainWindow.build_ui = _c18r2_build_ui


# --- Smart File Discovery ----------------------------------------------------

_old_c18r2_discovery_init = SmartFileDiscoveryDialog.__init__
def _c18r2_discovery_init(self, parent, base_options):
    _old_c18r2_discovery_init(self, parent, base_options)
    self.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
    QTimer.singleShot(
        0,
        lambda: _c18r2_place_window(
            self,
            parent=parent,
            width_ratio=0.78,
            height_ratio=0.82,
            minimum=(760, 560),
            maximum=(1150, 860),
        ),
    )


SmartFileDiscoveryDialog.__init__ = _c18r2_discovery_init


# --- Log Viewer --------------------------------------------------------------

_old_c18r2_viewer_init = MultiPaneLogViewer.__init__
def _c18r2_viewer_init(self, parent_window):
    _old_c18r2_viewer_init(self, parent_window)
    QTimer.singleShot(
        0,
        lambda: _c18r2_place_window(
            self,
            parent=parent_window,
            width_ratio=0.96,
            height_ratio=0.94,
            minimum=(900, 620),
        ),
    )


MultiPaneLogViewer.__init__ = _c18r2_viewer_init


# --- Investigation window ---------------------------------------------------

try:
    _old_c18r2_investigation_init = InvestigationWorkspace.__init__

    def _c18r2_investigation_init(self, parent_window, *args, **kwargs):
        _old_c18r2_investigation_init(self, parent_window, *args, **kwargs)
        QTimer.singleShot(
            0,
            lambda: _c18r2_place_window(
                self,
                parent=parent_window,
                width_ratio=0.96,
                height_ratio=0.94,
                minimum=(900, 620),
            ),
        )

    InvestigationWorkspace.__init__ = _c18r2_investigation_init
except Exception:
    pass


# --- Progress dialogs --------------------------------------------------------

_old_c18r2_make_progress = MultiPaneLogViewer.make_progress
def _c18r2_make_progress(self, title, text):
    dialog = _old_c18r2_make_progress(self, title, text)
    dialog.setWindowTitle(title)
    QTimer.singleShot(
        0,
        lambda: _c18r2_center_dialog(dialog, parent=self, preferred_size=(520, 170)),
    )
    return dialog


MultiPaneLogViewer.make_progress = _c18r2_make_progress


try:
    _old_c18r2_soft_progress_init = SoftProgressDialog.__init__

    def _c18r2_soft_progress_init(self, *args, **kwargs):
        _old_c18r2_soft_progress_init(self, *args, **kwargs)
        parent = self.parentWidget()
        QTimer.singleShot(
            0,
            lambda: _c18r2_center_dialog(
                self,
                parent=parent,
                preferred_size=(560, 190),
            ),
        )

    SoftProgressDialog.__init__ = _c18r2_soft_progress_init
except Exception:
    pass


# ---------------------------------------------------------------------------
# Commit0019: VIMeasure value view, compact viewer layout, context menu
# ---------------------------------------------------------------------------

def _c19_is_vimeasure_type(value):
    text = str(value or "").strip().upper().replace("_", "")
    return text == "VIMEASURE"


def _c19_vimeasure_records_to_value_rows(records):
    """
    Convert wide VIMeasure records into normalized value rows:
      Timestamp | Parameter | Value | Unit

    Existing raw rows remain untouched in memory. This function only creates
    a viewer projection.
    """
    output = []
    ignored = {
        "timestamp", "type", "level", "category", "status", "substatus",
        "message", "raw", "source_type", "filename", "line_no", "num",
        "parser", "unit", "parameter", "value",
    }

    for record in records or []:
        if isinstance(record, dict):
            getter = record.get
        else:
            getter = lambda name, default=None: getattr(record, name, default)

        timestamp = getter("timestamp")
        source_type = getter("source_type", getter("type", "VIMEASURE"))
        raw = getter("raw", "")
        fields = {}

        if isinstance(raw, dict):
            fields.update(raw)
        elif isinstance(raw, str) and raw.strip():
            try:
                loaded = json.loads(raw)
                if isinstance(loaded, dict):
                    fields.update(loaded)
            except Exception:
                pass

        # Also inspect plain attributes/dict values because the built-in parser
        # may store VIMeasure columns directly.
        if isinstance(record, dict):
            for key, value in record.items():
                if str(key).lower() not in ignored:
                    fields.setdefault(key, value)
        else:
            for key in dir(record):
                if key.startswith("_") or key.lower() in ignored:
                    continue
                try:
                    value = getattr(record, key)
                except Exception:
                    continue
                if callable(value):
                    continue
                fields.setdefault(key, value)

        # If parser already returned normalized rows, keep them.
        parameter = getter("parameter", None)
        value = getter("value", None)
        unit = getter("unit", "")
        if parameter not in (None, "") and value not in (None, ""):
            output.append({
                "Timestamp": timestamp,
                "Parameter": str(parameter),
                "Value": value,
                "Unit": str(unit or ""),
                "_source_type": source_type,
                "_raw_record": record,
            })
            continue

        for key, value in fields.items():
            key_text = str(key)
            if key_text.lower() in ignored:
                continue
            if value is None or value == "":
                continue

            numeric = value
            if isinstance(value, str):
                stripped = value.strip()
                try:
                    numeric = float(stripped)
                except ValueError:
                    continue

            if not isinstance(numeric, (int, float)):
                continue

            unit_guess = ""
            lower = key_text.lower()
            if lower.endswith("temp") or "temperature" in lower:
                unit_guess = "°C"
            elif lower.endswith("vi") or lower.endswith("current") or lower.endswith("i"):
                unit_guess = "A"
            elif lower.endswith("vv") or lower.endswith("voltage") or lower.endswith("v"):
                unit_guess = "V"
            elif "power" in lower:
                unit_guess = "W"

            output.append({
                "Timestamp": timestamp,
                "Parameter": key_text,
                "Value": numeric,
                "Unit": unit_guess,
                "_source_type": source_type,
                "_raw_record": record,
            })

    return output


def _c19_copy_to_clipboard(text):
    QApplication.clipboard().setText(str(text or ""))


def _c19_context_menu_for_table(owner, table, pos):
    index = table.indexAt(pos)
    if not index.isValid():
        return

    model = table.model()
    row = index.row()
    col = index.column()

    def value_at(r, c):
        idx = model.index(r, c)
        value = model.data(idx, Qt.DisplayRole)
        return "" if value is None else str(value)

    headers = []
    for c in range(model.columnCount()):
        value = model.headerData(c, Qt.Horizontal, Qt.DisplayRole)
        headers.append("" if value is None else str(value))

    cell_value = value_at(row, col)
    row_values = [value_at(row, c) for c in range(model.columnCount())]
    row_map = dict(zip(headers, row_values))

    menu = QMenu(table)

    action_copy_cell = menu.addAction("Copy Cell")
    action_copy_row = menu.addAction("Copy Row")
    action_copy_timestamp = menu.addAction("Copy Timestamp")

    message_key = next(
        (key for key in ("Message", "Value", "Parameter") if key in row_map),
        None,
    )
    action_copy_message = menu.addAction(
        "Copy Message / Value" if message_key else "Copy Displayed Value"
    )

    menu.addSeparator()
    action_filter = menu.addAction("Filter by This Value")
    action_clear_filter = menu.addAction("Clear Pane Filter")
    menu.addSeparator()
    action_copy_rule = menu.addAction("Copy Rule Text")
    action_noise = menu.addAction("Add to Noise Rule")
    action_export = menu.addAction("Export Selected Row")

    chosen = menu.exec(table.viewport().mapToGlobal(pos))
    if chosen is None:
        return

    if chosen == action_copy_cell:
        _c19_copy_to_clipboard(cell_value)
    elif chosen == action_copy_row:
        _c19_copy_to_clipboard("\t".join(row_values))
    elif chosen == action_copy_timestamp:
        _c19_copy_to_clipboard(row_map.get("Timestamp", ""))
    elif chosen == action_copy_message:
        _c19_copy_to_clipboard(
            row_map.get(message_key, cell_value) if message_key else cell_value
        )
    elif chosen == action_filter:
        if hasattr(owner, "apply_filter_from_value"):
            owner.apply_filter_from_value(headers[col], cell_value)
        elif hasattr(owner, "set_filter_value"):
            owner.set_filter_value(headers[col], cell_value)
        else:
            # Generic local proxy filtering fallback.
            proxy = table.model()
            if hasattr(proxy, "setFilterKeyColumn") and hasattr(proxy, "setFilterFixedString"):
                proxy.setFilterKeyColumn(col)
                proxy.setFilterFixedString(cell_value)
    elif chosen == action_clear_filter:
        if hasattr(owner, "clear_filter"):
            owner.clear_filter()
        else:
            proxy = table.model()
            if hasattr(proxy, "setFilterFixedString"):
                proxy.setFilterFixedString("")
    elif chosen == action_copy_rule:
        rule_text = row_map.get("Message") or row_map.get("Value") or cell_value
        _c19_copy_to_clipboard(rule_text)
    elif chosen == action_noise:
        rule_text = row_map.get("Message") or row_map.get("Value") or cell_value
        if hasattr(owner, "add_noise_rule_text"):
            owner.add_noise_rule_text(rule_text)
        elif hasattr(owner, "add_noise_rule"):
            owner.add_noise_rule(rule_text)
        else:
            _c19_copy_to_clipboard(rule_text)
            QMessageBox.information(
                owner,
                APP_TITLE,
                "Rule text was copied. The current pane does not expose a direct "
                "Add Noise Rule handler.",
            )
    elif chosen == action_export:
        path, _ = QFileDialog.getSaveFileName(
            owner,
            "Export Selected Row",
            "selected_row.tsv",
            "TSV (*.tsv);;Text (*.txt)",
        )
        if path:
            Path(path).write_text(
                "\t".join(headers) + "\n" + "\t".join(row_values) + "\n",
                encoding="utf-8",
            )


def _c19_enable_context_menu(owner, table):
    if table is None:
        return
    table.setContextMenuPolicy(Qt.CustomContextMenu)
    try:
        table.customContextMenuRequested.disconnect()
    except Exception:
        pass
    table.customContextMenuRequested.connect(
        lambda pos, o=owner, t=table: _c19_context_menu_for_table(o, t, pos)
    )


def _c19_find_tables(widget):
    tables = []
    for cls in (QTableView, QTableWidget):
        try:
            tables.extend(widget.findChildren(cls))
        except Exception:
            pass
    # preserve order, remove duplicates
    seen = set()
    result = []
    for table in tables:
        ident = id(table)
        if ident in seen:
            continue
        seen.add(ident)
        result.append(table)
    return result


def _c19_compact_viewer_layout(viewer):
    """
    Give almost all vertical space to the log/value tables.
    Details/message panes start collapsed and no artificial blank area remains.
    """
    try:
        viewer.setMinimumSize(900, 620)

        for splitter in viewer.findChildren(QSplitter):
            count = splitter.count()
            if count < 2:
                continue

            orientation = splitter.orientation()
            if orientation == Qt.Vertical:
                total = max(splitter.height(), 600)
                sizes = [max(120, total - 10 * (count - 1))] + [0] * (count - 1)
                splitter.setSizes(sizes)
                splitter.setCollapsible(0, False)
                for index in range(1, count):
                    splitter.setCollapsible(index, True)

        # Remove fixed-height/minimum-height leftovers that create large gaps.
        for table in _c19_find_tables(viewer):
            table.setMinimumHeight(220)
            table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            table.verticalHeader().setDefaultSectionSize(22)

        # Hide empty detail widgets at startup.
        for widget in viewer.findChildren(QWidget):
            name = widget.objectName().lower()
            if any(token in name for token in ("detail", "messagepanel", "message_panel")):
                if isinstance(widget, (QTextEdit, QPlainTextEdit, QTableView, QTableWidget)):
                    if not widget.isVisible():
                        continue
    except Exception:
        write_startup_log(
            "Commit0019 compact layout failed.\n\n" + traceback.format_exc()
        )


def _c19_apply_vimeasure_projection(pane):
    """
    Apply only when the pane's active file type is VIMeasure.
    The function is intentionally defensive because pane implementations differ
    between RC1 revisions.
    """
    try:
        file_type = None
        for name in ("source_type", "file_type", "current_type", "log_type"):
            value = getattr(pane, name, None)
            if value:
                file_type = value
                break

        if not _c19_is_vimeasure_type(file_type):
            # inspect loaded records as fallback
            records = getattr(pane, "records", None) or getattr(pane, "_records", None)
            if records:
                first = records[0]
                if isinstance(first, dict):
                    file_type = first.get("source_type") or first.get("type")
                else:
                    file_type = getattr(first, "source_type", None)
        if not _c19_is_vimeasure_type(file_type):
            return

        records = (
            getattr(pane, "records", None)
            or getattr(pane, "_records", None)
            or getattr(pane, "all_records", None)
            or []
        )
        value_rows = _c19_vimeasure_records_to_value_rows(records)
        if not value_rows:
            return

        table = None
        for name in ("table", "table_view", "view", "log_table"):
            candidate = getattr(pane, name, None)
            if isinstance(candidate, (QTableView, QTableWidget)):
                table = candidate
                break
        if table is None:
            tables = _c19_find_tables(pane)
            table = tables[0] if tables else None
        if table is None:
            return

        model = QStandardItemModel(table)
        model.setHorizontalHeaderLabels(["Timestamp", "Parameter", "Value", "Unit"])

        for row in value_rows:
            items = [
                QStandardItem(str(row.get("Timestamp", ""))),
                QStandardItem(str(row.get("Parameter", ""))),
                QStandardItem(str(row.get("Value", ""))),
                QStandardItem(str(row.get("Unit", ""))),
            ]
            model.appendRow(items)

        table.setModel(model)
        table.setSortingEnabled(True)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        _c19_enable_context_menu(pane, table)
    except Exception:
        write_startup_log(
            "Commit0019 VIMeasure projection failed.\n\n" + traceback.format_exc()
        )


# Apply to viewer after construction and after load/reload methods.
_old_c19_viewer_init = MultiPaneLogViewer.__init__
def _c19_viewer_init(self, *args, **kwargs):
    _old_c19_viewer_init(self, *args, **kwargs)
    QTimer.singleShot(0, lambda: _c19_compact_viewer_layout(self))
    QTimer.singleShot(
        0,
        lambda: [
            _c19_enable_context_menu(self, table)
            for table in _c19_find_tables(self)
        ],
    )
    QTimer.singleShot(
        250,
        lambda: [
            _c19_apply_vimeasure_projection(pane)
            for pane in self.findChildren(QWidget)
        ],
    )


MultiPaneLogViewer.__init__ = _c19_viewer_init


def _c19_wrap_refresh_method(cls, method_name):
    original = getattr(cls, method_name, None)
    if original is None or getattr(original, "_commit0019_wrapped", False):
        return

    def wrapped(self, *args, **kwargs):
        result = original(self, *args, **kwargs)
        QTimer.singleShot(0, lambda: _c19_compact_viewer_layout(self))
        QTimer.singleShot(
            0,
            lambda: [
                _c19_enable_context_menu(self, table)
                for table in _c19_find_tables(self)
            ],
        )
        QTimer.singleShot(
            100,
            lambda: [
                _c19_apply_vimeasure_projection(pane)
                for pane in self.findChildren(QWidget)
            ],
        )
        return result

    wrapped._commit0019_wrapped = True
    setattr(cls, method_name, wrapped)


for _method_name in (
    "load_files",
    "load_records",
    "reload",
    "refresh",
    "populate",
    "set_records",
    "apply_records",
):
    _c19_wrap_refresh_method(MultiPaneLogViewer, _method_name)


# Apply context menus to generic pane classes in the module.
for _candidate_name in (
    "ViewerPane",
    "LogViewerPane",
    "StructuredLogPane",
    "LogPane",
):
    _candidate = globals().get(_candidate_name)
    if _candidate is None:
        continue
    _old_init = getattr(_candidate, "__init__", None)
    if _old_init is None or getattr(_old_init, "_commit0019_wrapped", False):
        continue

    def _make_init(original):
        def wrapped(self, *args, **kwargs):
            original(self, *args, **kwargs)
            QTimer.singleShot(
                0,
                lambda: [
                    _c19_enable_context_menu(self, table)
                    for table in _c19_find_tables(self)
                ],
            )
            QTimer.singleShot(100, lambda: _c19_apply_vimeasure_projection(self))
        wrapped._commit0019_wrapped = True
        return wrapped

    _candidate.__init__ = _make_init(_old_init)


# ---------------------------------------------------------------------------
# Commit0019 R2: direct Viewer row-fit and right-click implementation
# ---------------------------------------------------------------------------

def _c19r2_table_value(table, row, column):
    model = table.model()
    if model is None:
        return ""
    index = model.index(row, column)
    value = model.data(index, Qt.DisplayRole)
    return "" if value is None else str(value)


def _c19r2_table_headers(table):
    model = table.model()
    if model is None:
        return []
    return [
        str(model.headerData(column, Qt.Horizontal, Qt.DisplayRole) or "")
        for column in range(model.columnCount())
    ]


def _c19r2_copy_text(value):
    QApplication.clipboard().setText(str(value or ""))


def _c19r2_apply_filter(viewer, pane_index, column, value):
    try:
        rows = viewer.all_rows[pane_index]
        headers = _c19r2_table_headers(viewer.tables[pane_index])
        if not headers or column >= len(headers):
            return

        field = headers[column]
        filtered = [
            row for row in rows
            if str(row.get(field, "")) == str(value)
        ]

        viewer.models[pane_index].set_rows(filtered)
        viewer.status.setText(
            f"Pane {pane_index + 1}: filtered {field} = {value} "
            f"({len(filtered)} rows)"
        )
        _c19r2_update_row_fit(viewer, pane_index)
    except Exception:
        write_startup_log(
            "Commit0019 R2 filter failed.\n\n" + traceback.format_exc()
        )


def _c19r2_clear_filter(viewer, pane_index):
    try:
        viewer.models[pane_index].set_rows(viewer.all_rows[pane_index])
        viewer.status.setText(
            f"Pane {pane_index + 1}: filter cleared "
            f"({len(viewer.all_rows[pane_index])} rows)"
        )
        _c19r2_update_row_fit(viewer, pane_index)
    except Exception:
        write_startup_log(
            "Commit0019 R2 clear filter failed.\n\n" + traceback.format_exc()
        )


def _c19r2_context_menu(viewer, pane_index, position):
    try:
        table = viewer.tables[pane_index]
        index = table.indexAt(position)
        if not index.isValid():
            return

        row = index.row()
        column = index.column()
        headers = _c19r2_table_headers(table)
        values = [
            _c19r2_table_value(table, row, col)
            for col in range(table.model().columnCount())
        ]
        row_map = dict(zip(headers, values))
        cell_value = values[column] if column < len(values) else ""

        menu = QMenu(table)

        copy_cell = menu.addAction("Copy Cell")
        copy_row = menu.addAction("Copy Row")
        copy_timestamp = menu.addAction("Copy Timestamp")
        copy_message = menu.addAction("Copy Message / Value")

        menu.addSeparator()

        filter_value = menu.addAction("Filter by This Value")
        clear_filter = menu.addAction("Clear Pane Filter")

        menu.addSeparator()

        copy_rule = menu.addAction("Copy Rule Text")
        add_noise = menu.addAction("Add to Noise Rule")
        export_row = menu.addAction("Export Selected Row")

        chosen = menu.exec(table.viewport().mapToGlobal(position))
        if chosen is None:
            return

        if chosen == copy_cell:
            _c19r2_copy_text(cell_value)

        elif chosen == copy_row:
            _c19r2_copy_text("\t".join(values))

        elif chosen == copy_timestamp:
            _c19r2_copy_text(row_map.get("Timestamp", ""))

        elif chosen == copy_message:
            value = (
                row_map.get("Message")
                or row_map.get("Value")
                or row_map.get("Raw")
                or cell_value
            )
            _c19r2_copy_text(value)

        elif chosen == filter_value:
            _c19r2_apply_filter(viewer, pane_index, column, cell_value)

        elif chosen == clear_filter:
            _c19r2_clear_filter(viewer, pane_index)

        elif chosen == copy_rule:
            rule_text = (
                row_map.get("Message")
                or row_map.get("Value")
                or cell_value
            )
            _c19r2_copy_text(rule_text)

        elif chosen == add_noise:
            try:
                viewer.approve_selected_noise(pane_index)
            except Exception:
                rule_text = (
                    row_map.get("Message")
                    or row_map.get("Value")
                    or cell_value
                )
                _c19r2_copy_text(rule_text)
                QMessageBox.information(
                    viewer,
                    APP_TITLE,
                    "The selected rule text was copied to the clipboard.",
                )

        elif chosen == export_row:
            output, _ = QFileDialog.getSaveFileName(
                viewer,
                "Export Selected Row",
                "selected_row.tsv",
                "TSV files (*.tsv);;Text files (*.txt)",
            )
            if output:
                Path(output).write_text(
                    "\t".join(headers) + "\n" + "\t".join(values) + "\n",
                    encoding="utf-8",
                )

    except Exception:
        write_startup_log(
            "Commit0019 R2 context menu failed.\n\n" + traceback.format_exc()
        )


def _c19r2_update_row_fit(viewer, pane_index=None):
    """
    Remove the large empty lower area for short result sets.

    Up to 25 visible rows:
      rows stretch to fill the table viewport.

    Larger result sets:
      compact fixed-height rows with normal scrolling.
    """
    try:
        indices = (
            [pane_index]
            if pane_index is not None
            else range(len(viewer.tables))
        )

        for index in indices:
            table = viewer.tables[index]
            model = table.model()
            if model is None:
                continue

            row_count = model.rowCount()
            header = table.verticalHeader()

            if 0 < row_count <= 25:
                header.setSectionResizeMode(QHeaderView.Stretch)
                table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            else:
                header.setSectionResizeMode(QHeaderView.Fixed)
                header.setDefaultSectionSize(22)
                table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

            table.setSizePolicy(
                QSizePolicy.Expanding,
                QSizePolicy.Expanding,
            )
            table.updateGeometry()
            table.viewport().update()

        if hasattr(viewer, "detail"):
            viewer.detail.hide()
            viewer.detail.setMinimumHeight(0)
            viewer.detail.setMaximumHeight(0)

        if hasattr(viewer, "main_splitter"):
            viewer.main_splitter.setSizePolicy(
                QSizePolicy.Expanding,
                QSizePolicy.Expanding,
            )

    except Exception:
        write_startup_log(
            "Commit0019 R2 row fit failed.\n\n" + traceback.format_exc()
        )


def _c19r2_install_table_handlers(viewer):
    try:
        for index, table in enumerate(viewer.tables):
            table.setContextMenuPolicy(Qt.CustomContextMenu)

            try:
                table.customContextMenuRequested.disconnect()
            except Exception:
                pass

            table.customContextMenuRequested.connect(
                lambda position, pane=index:
                    _c19r2_context_menu(viewer, pane, position)
            )

            # Also support the keyboard context-menu key / Shift+F10.
            table.setFocusPolicy(Qt.StrongFocus)

        _c19r2_update_row_fit(viewer)

    except Exception:
        write_startup_log(
            "Commit0019 R2 table-handler install failed.\n\n"
            + traceback.format_exc()
        )


_old_c19r2_viewer_init = MultiPaneLogViewer.__init__
def _c19r2_viewer_init(self, *args, **kwargs):
    _old_c19r2_viewer_init(self, *args, **kwargs)

    QTimer.singleShot(
        0,
        lambda: _c19r2_install_table_handlers(self),
    )
    QTimer.singleShot(
        300,
        lambda: _c19r2_update_row_fit(self),
    )


MultiPaneLogViewer.__init__ = _c19r2_viewer_init


def _c19r2_wrap_method(method_name):
    original = getattr(MultiPaneLogViewer, method_name, None)
    if original is None or getattr(original, "_c19r2_wrapped", False):
        return

    def wrapped(self, *args, **kwargs):
        result = original(self, *args, **kwargs)

        QTimer.singleShot(
            0,
            lambda: _c19r2_install_table_handlers(self),
        )
        QTimer.singleShot(
            100,
            lambda: _c19r2_update_row_fit(self),
        )

        return result

    wrapped._c19r2_wrapped = True
    setattr(MultiPaneLogViewer, method_name, wrapped)


for _method in (
    "load_pane",
    "load_visible",
    "apply_search",
    "apply_noise_rules_now",
    "apply_time_to_focused",
    "apply_time_to_visible",
    "clear_viewer_time_range",
    "update_view_mode",
):
    _c19r2_wrap_method(_method)


# ---------------------------------------------------------------------------
# Commit0020: viewer error fix, large-table progress, CallID cross-linking
# ---------------------------------------------------------------------------

_C20_CALL_ID_PATTERNS = [
    re.compile(r"(?i)\bcall[\s_-]*id\b\s*[:=#]?\s*[\[\(<{]?\s*([A-Za-z0-9_-]{2,64})"),
    re.compile(r"(?i)\bcallid\b\s*[:=#]?\s*[\[\(<{]?\s*([A-Za-z0-9_-]{2,64})"),
    re.compile(r"(?i)\bcase[\s_-]*id\b\s*[:=#]?\s*[\[\(<{]?\s*([A-Za-z0-9_-]{2,64})"),
]


def _c20_extract_call_id(*values):
    """
    Safely extract CallID/Call ID/Case ID from strings, dictionaries,
    nested lists, datetime-containing records, and other parsed values.
    """
    pending = list(values)
    visited = set()

    while pending:
        value = pending.pop(0)

        if value is None:
            continue

        object_id = id(value)
        if object_id in visited:
            continue
        visited.add(object_id)

        if isinstance(value, dict):
            for key, item in value.items():
                normalized = (
                    str(key)
                    .lower()
                    .replace("_", "")
                    .replace("-", "")
                    .replace(" ", "")
                )

                if normalized in {"callid", "caseid"} and item not in (None, ""):
                    return str(item).strip().strip("[](){}<>.,;:")

                pending.append(item)

            try:
                text_value = json.dumps(
                    value,
                    ensure_ascii=False,
                    default=str,
                )
            except Exception:
                text_value = str(value)

        elif isinstance(value, (list, tuple, set)):
            pending.extend(value)
            text_value = str(value)

        else:
            text_value = str(value)

        for pattern in _C20_CALL_ID_PATTERNS:
            match = pattern.search(text_value)
            if match:
                return match.group(1).strip().strip("[](){}<>.,;:")

    return ""


_old_c20_record_to_viewer_row = record_to_viewer_row
def _c20_record_to_viewer_row(record):
    row = _old_c20_record_to_viewer_row(record)

    source_type = str(getattr(record, "source_type", "") or "").upper()
    if source_type in {"WS", "CSA", "CGA", "MRSERVER"}:
        raw_value = getattr(record, "raw", "")
        raw_object = raw_value
        if isinstance(raw_value, str) and raw_value.strip():
            try:
                raw_object = json.loads(raw_value)
            except Exception:
                raw_object = raw_value

        call_id = _c20_extract_call_id(
            row,
            raw_object,
            getattr(record, "message", ""),
            row.get("Message", ""),
            row.get("Raw", ""),
        )
        if call_id:
            row["CallID"] = call_id

    return row


record_to_viewer_row = _c20_record_to_viewer_row


def _c20_build_rows_with_progress(records, progress, label):
    total = len(records)
    rows = []
    batch_size = 2000

    if progress:
        progress.setRange(0, max(1, total))
        progress.setValue(0)
        progress.setLabelText(f"Building table rows for {label}: 0/{total}")
        QApplication.processEvents()

    for index, record in enumerate(records, start=1):
        rows.append(record_to_viewer_row(record))

        if progress and (index % batch_size == 0 or index == total):
            progress.setValue(index)
            progress.setLabelText(
                f"Building table rows for {label}: {index:,}/{total:,}"
            )
            QApplication.processEvents()

            if progress.wasCanceled():
                raise RuntimeError("Viewer table build was cancelled.")

    return rows


# Fix the callback signature that caused:
# TypeError: _v41_update_view_mode() takes 1 positional argument but 2 were given.
def _c20_update_view_mode(self, *args, **kwargs):
    if not hasattr(self, "main_splitter") or not hasattr(self, "panes"):
        return

    visible = self.visible_indices()
    for index, pane in enumerate(self.panes):
        pane.setVisible(index in visible)

    size_each = max(1, 1000 // max(1, len(visible)))
    try:
        self.main_splitter.setSizes([
            size_each if index in visible else 0
            for index in range(self.MAX_PANES)
        ])
    except Exception:
        pass

    # Avoid duplicate log lines when multiple checkbox stateChanged signals fire.
    state = tuple(visible)
    if getattr(self, "_c20_last_visible_state", None) != state:
        self._c20_last_visible_state = state
        self.log("Viewer panes: " + ", ".join(str(index + 1) for index in visible))


MultiPaneLogViewer.update_view_mode = _c20_update_view_mode


def _c20_filter_rows_by_call_id(viewer, pane_index, call_id):
    rows = viewer.all_rows[pane_index]
    filtered = [
        row for row in rows
        if str(row.get("CallID", "")).strip() == str(call_id).strip()
    ]

    columns = viewer.pane_columns_for_rows(pane_index, filtered)
    if "CallID" not in columns:
        columns = ["CallID"] + columns

    viewer.models[pane_index].set_rows(filtered, columns)
    viewer.apply_table_column_widths(pane_index)
    viewer.status.setText(
        f"{viewer.pane_name(pane_index)}: CallID {call_id} "
        f"({len(filtered)} rows)"
    )


def _c20_link_call_id_across_panes(viewer, call_id):
    matched_panes = 0

    for pane_index in viewer.visible_indices():
        rows = viewer.all_rows[pane_index]
        if any(
            str(row.get("CallID", "")).strip() == str(call_id).strip()
            for row in rows
        ):
            _c20_filter_rows_by_call_id(viewer, pane_index, call_id)
            matched_panes += 1

    if matched_panes == 0:
        QMessageBox.information(
            viewer,
            "CallID Link",
            f"No visible pane contains CallID: {call_id}",
        )
    else:
        viewer.log(
            f"CallID link applied: {call_id} across {matched_panes} pane(s)"
        )


_old_c20_context_menu = _c19r2_context_menu
def _c20_context_menu(viewer, pane_index, position):
    try:
        table = viewer.tables[pane_index]
        index = table.indexAt(position)
        if not index.isValid():
            return

        row_number = index.row()
        model_row = viewer.models[pane_index].row_at(row_number) or {}
        call_id = str(model_row.get("CallID", "")).strip()

        if not call_id:
            return _old_c20_context_menu(viewer, pane_index, position)

        model = table.model()
        headers = _c19r2_table_headers(table)
        values = [
            _c19r2_table_value(table, row_number, column)
            for column in range(model.columnCount())
        ]
        row_map = dict(zip(headers, values))
        cell_value = values[index.column()] if index.column() < len(values) else ""

        menu = QMenu(table)

        copy_cell = menu.addAction("Copy Cell")
        copy_row = menu.addAction("Copy Row")
        copy_timestamp = menu.addAction("Copy Timestamp")
        copy_message = menu.addAction("Copy Message / Value")

        menu.addSeparator()

        link_call = menu.addAction(f"Link Same CallID Across Panes: {call_id}")
        filter_call = menu.addAction(f"Show Only This CallID: {call_id}")
        clear_filter = menu.addAction("Clear Pane Filter")

        menu.addSeparator()

        copy_rule = menu.addAction("Copy Rule Text")
        add_noise = menu.addAction("Add to Noise Rule")
        export_row = menu.addAction("Export Selected Row")

        chosen = menu.exec(table.viewport().mapToGlobal(position))
        if chosen is None:
            return

        if chosen == copy_cell:
            _c19r2_copy_text(cell_value)
        elif chosen == copy_row:
            _c19r2_copy_text("\t".join(values))
        elif chosen == copy_timestamp:
            _c19r2_copy_text(row_map.get("Timestamp", ""))
        elif chosen == copy_message:
            _c19r2_copy_text(
                row_map.get("Message")
                or row_map.get("Value")
                or cell_value
            )
        elif chosen == link_call:
            _c20_link_call_id_across_panes(viewer, call_id)
        elif chosen == filter_call:
            _c20_filter_rows_by_call_id(viewer, pane_index, call_id)
        elif chosen == clear_filter:
            _c19r2_clear_filter(viewer, pane_index)
        elif chosen == copy_rule:
            _c19r2_copy_text(
                row_map.get("Message")
                or row_map.get("Value")
                or cell_value
            )
        elif chosen == add_noise:
            try:
                viewer.approve_selected_noise(pane_index)
            except Exception:
                _c19r2_copy_text(
                    row_map.get("Message")
                    or row_map.get("Value")
                    or cell_value
                )
        elif chosen == export_row:
            output, _ = QFileDialog.getSaveFileName(
                viewer,
                "Export Selected Row",
                "selected_row.tsv",
                "TSV files (*.tsv);;Text files (*.txt)",
            )
            if output:
                Path(output).write_text(
                    "\t".join(headers) + "\n" + "\t".join(values) + "\n",
                    encoding="utf-8",
                )

    except Exception:
        write_startup_log(
            "Commit0020 CallID context menu failed.\n\n"
            + traceback.format_exc()
        )


_c19r2_context_menu = _c20_context_menu


# Replace the expensive one-shot list comprehension in load_pane with
# chunked conversion and visible progress updates.
_old_c20_load_pane = MultiPaneLogViewer.load_pane
def _c20_load_pane(self, side):
    index = self.side_index(side)
    source = self.sources[index].currentText()
    label = self.pane_name(index)

    self.log(f"Loading {label}: {source} ...")
    progress = self.make_progress("Log Viewer", f"Preparing {label}: {source} ...")
    progress.show()
    QApplication.setOverrideCursor(Qt.WaitCursor)

    try:
        records = self.source_to_records(source, progress)
        if progress.wasCanceled():
            return

        rows = _c20_build_rows_with_progress(records, progress, label)

        file_hint = source
        if rows:
            files = sorted({
                str(row.get("File", ""))
                for row in rows
                if row.get("File", "")
            })
            if len(files) == 1:
                file_hint = files[0]
            elif len(files) > 1:
                file_hint = f"{len(files)} files loaded"

        self.all_rows[index] = rows
        self.file_labels[index].setText(f"File: {file_hint}")

        self.apply_view_filters(index)
        self.log(f"{label} loaded: {len(rows)} rows")

    except RuntimeError as error:
        if "cancelled" in str(error).lower():
            self.log(str(error))
        else:
            write_startup_log(
                "Multi Log Viewer load_pane failed.\n\n"
                + traceback.format_exc()
            )
            QMessageBox.critical(
                self,
                "Log Viewer",
                "Viewer loading failed.\n\n"
                + traceback.format_exc()
                + f"\n\nLog:\n{startup_log_path()}",
            )
    except Exception:
        write_startup_log(
            "Multi Log Viewer load_pane failed.\n\n"
            + traceback.format_exc()
        )
        QMessageBox.critical(
            self,
            "Log Viewer",
            "Viewer loading failed.\n\n"
            + traceback.format_exc()
            + f"\n\nLog:\n{startup_log_path()}",
        )
    finally:
        QApplication.restoreOverrideCursor()
        progress.close()


MultiPaneLogViewer.load_pane = _c20_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)


# ---------------------------------------------------------------------------
# Commit0020 R2: final START routing and datetime-safe Load This override
# ---------------------------------------------------------------------------
def _c20r2_extract_call_id(*values):
    queue = list(values)
    visited = set()
    while queue:
        value = queue.pop(0)
        if value is None:
            continue
        oid = id(value)
        if oid in visited:
            continue
        visited.add(oid)
        if isinstance(value, dict):
            for key, item in value.items():
                normalized = re.sub(r"[\s_-]+", "", str(key).strip().lower())
                if normalized in {"callid", "caseid"} and item not in (None, ""):
                    return str(item).strip().strip("[](){}<>.,;:")
                queue.append(item)
            continue
        if isinstance(value, (list, tuple, set)):
            queue.extend(value)
            continue
        value_text = str(value)
        for pattern in _C20_CALL_ID_PATTERNS:
            match = pattern.search(value_text)
            if match:
                return match.group(1).strip().strip("[](){}<>.,;:")
    return ""

_c20_extract_call_id = _c20r2_extract_call_id

def _c20r2_record_to_viewer_row(record):
    row = _old_c20_record_to_viewer_row(record)
    source_type = str(getattr(record, "source_type", "") or row.get("SourceType", "") or row.get("Type", "")).upper()
    if source_type in {"WS", "CSA", "CGA", "MRSERVER"}:
        call_id = _c20r2_extract_call_id(row, getattr(record, "raw", None), getattr(record, "message", None), row.get("Message"), row.get("Raw"))
        if call_id:
            row["CallID"] = call_id
    return row
record_to_viewer_row = _c20r2_record_to_viewer_row

def _c20r2_build_rows_with_progress(records, progress, label):
    total = len(records); rows = []; batch = 1000
    if progress is not None:
        progress.setRange(0, max(1,total)); progress.setValue(0)
        progress.setLabelText(f"Building table rows for {label}: 0/{total:,}")
        QApplication.processEvents()
    for index, record in enumerate(records, 1):
        rows.append(_c20r2_record_to_viewer_row(record))
        if progress is not None and (index % batch == 0 or index == total):
            progress.setValue(index)
            progress.setLabelText(f"Building table rows for {label}: {index:,}/{total:,}")
            QApplication.processEvents()
            if progress.wasCanceled():
                raise RuntimeError("Viewer table build was cancelled.")
    return rows
_c20_build_rows_with_progress = _c20r2_build_rows_with_progress

def _c20r2_load_pane(self, side):
    index = self.side_index(side); source = self.sources[index].currentText(); label = self.pane_name(index)
    self.log(f"Loading {label}: {source} ...")
    progress = self.make_progress("Log Viewer", f"Preparing {label}: {source} ...")
    progress.show(); QApplication.setOverrideCursor(Qt.WaitCursor)
    try:
        records = self.source_to_records(source, progress)
        if progress.wasCanceled(): return
        rows = _c20r2_build_rows_with_progress(records, progress, label)
        file_hint = source
        if rows:
            files = sorted({str(row.get("File", "")) for row in rows if row.get("File", "")})
            if len(files)==1: file_hint=files[0]
            elif len(files)>1: file_hint=f"{len(files)} files loaded"
        self.all_rows[index]=rows; self.file_labels[index].setText(f"File: {file_hint}")
        self.apply_view_filters(index); self.log(f"{label} loaded: {len(rows)} rows")
    except RuntimeError as error:
        if "cancelled" in str(error).lower(): self.log(str(error))
        else:
            write_startup_log("Commit0020 R2 load_pane failed.\n\n"+traceback.format_exc())
            QMessageBox.critical(self,"Log Viewer","Viewer loading failed.\n\n"+traceback.format_exc()+f"\n\nLog:\n{startup_log_path()}")
    except Exception:
        write_startup_log("Commit0020 R2 load_pane failed.\n\n"+traceback.format_exc())
        QMessageBox.critical(self,"Log Viewer","Viewer loading failed.\n\n"+traceback.format_exc()+f"\n\nLog:\n{startup_log_path()}")
    finally:
        QApplication.restoreOverrideCursor(); progress.close()
MultiPaneLogViewer.load_pane = _c20r2_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)

# START is explicitly bound to the current discovery/import handler, never run_clicked.
def _c20r2_build_ui(self):
    _old_c20r2_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
    if hasattr(self, "run_btn"):
        try: self.run_btn.clicked.disconnect()
        except Exception: pass
        self.run_btn.setText("▶ START")
        self.run_btn.setToolTip("Open Smart File Discovery and load selected files into Viewer. Merge is not started.")
        self.run_btn.clicked.connect(lambda checked=False: self.start_import_clicked())
_old_c20r2_build_ui = MainWindow.build_ui
MainWindow.build_ui = _c20r2_build_ui


# ---------------------------------------------------------------------------
# Commit0021: Viewer-only workflow
# ---------------------------------------------------------------------------

def _c21_disable_merge_workflow(self, *args, **kwargs):
    QMessageBox.information(
        self,
        APP_TITLE,
        "Merge output is disabled in this Viewer-focused version.\n\n"
        "Use START to discover files and open them in Log Explorer.",
    )
    return None


def _c21_viewer_only_build_ui(self):
    _old_c21_viewer_only_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")

    # START is the only primary workflow.
    if hasattr(self, "run_btn"):
        try:
            self.run_btn.clicked.disconnect()
        except Exception:
            pass

        self.run_btn.setText("▶ START")
        self.run_btn.setToolTip(
            "Open Smart File Discovery, select files, and display them "
            "in Log Explorer. No Excel/CSV merge output is created."
        )
        self.run_btn.clicked.connect(
            lambda checked=False: self.start_import_clicked()
        )

    # Remove the explicit MERGE button from the UI.
    if hasattr(self, "merge_btn"):
        try:
            self.merge_btn.clicked.disconnect()
        except Exception:
            pass

        self.merge_btn.hide()
        self.merge_btn.setEnabled(False)
        self.merge_btn.setParent(None)
        self.merge_btn.deleteLater()
        self.merge_btn = None

    # Remove merge-specific project actions from the visible UI.
    for attr in (
        "split_merge_btn",
        "noise_enable_merge_chk",
        "chk_skip_merged1",
    ):
        widget = getattr(self, attr, None)
        if widget is not None:
            widget.hide()
            widget.setEnabled(False)

    # Clarify that same-type files are combined only for Viewer display.
    if hasattr(self, "status_label"):
        self.status_label.setText(
            "Ready. START combines selected files by log type for Viewer display only."
        )


_old_c21_viewer_only_build_ui = MainWindow.build_ui
MainWindow.build_ui = _c21_viewer_only_build_ui

# Even if an old shortcut or stale signal reaches run_clicked,
# no merge output can be started.
MainWindow.run_clicked = _c21_disable_merge_workflow


_old_c21_build_rows = _c20r2_build_rows_with_progress
def _c21_build_rows_with_progress(records, progress, label):
    total = len(records)
    rows = []
    batch = 1000

    if progress is not None:
        progress.setRange(0, max(1, total))
        progress.setValue(0)
        progress.setLabelText(
            f"Indexing selected {label} rows for Viewer: 0/{total:,}"
        )
        QApplication.processEvents()

    for index, record in enumerate(records, 1):
        rows.append(_c20r2_record_to_viewer_row(record))

        if progress is not None and (
            index % batch == 0 or index == total
        ):
            progress.setValue(index)
            progress.setLabelText(
                f"Indexing selected {label} rows for Viewer: "
                f"{index:,}/{total:,}"
            )
            QApplication.processEvents()

            if progress.wasCanceled():
                raise RuntimeError("Viewer indexing was cancelled.")

    return rows


_c20r2_build_rows_with_progress = _c21_build_rows_with_progress
_c20_build_rows_with_progress = _c21_build_rows_with_progress


_old_c21_refresh_sources = getattr(
    MultiPaneLogViewer,
    "refresh_available_sources",
    None,
)
def _c21_refresh_available_sources(self, *args, **kwargs):
    if callable(_old_c21_refresh_sources):
        result = _old_c21_refresh_sources(self, *args, **kwargs)
    else:
        result = None

    # "Merged" is an output concept and is not shown in Viewer-only mode.
    for combo in getattr(self, "sources", []):
        for index in range(combo.count() - 1, -1, -1):
            if combo.itemText(index).strip().upper() == "MERGED":
                combo.removeItem(index)

    return result


if callable(_old_c21_refresh_sources):
    MultiPaneLogViewer.refresh_available_sources = _c21_refresh_available_sources


# ---------------------------------------------------------------------------
# Commit0023: Viewer quick filters, CSA pipeline recovery, comprehensive eval
# ---------------------------------------------------------------------------

_C23_CSA_LINE_RE = re.compile(
    r"^\s*\d{1,2}:\d{2}:\d{2}:\d{3}\s+"
    r"(?:Inf|Err|Wrn|Sev|Dbg|Ftl)\s+\d+\s+",
    re.I,
)
_C23_CGA_HINT_RE = re.compile(r"\b(?:CGA|Cga Interface|CThreadCga)\b", re.I)
_C23_CSA_HINT_RE = re.compile(r"\b(?:CSA|Csa Interface|CThreadCsa)\b", re.I)


def _c23_content_type_probe(path):
    """Content fallback for renamed CSA/CGA/Acquisition files."""
    try:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            lines = [handle.readline() for _ in range(12)]
    except Exception:
        return None

    joined = "".join(lines)
    name = path.name.lower()

    if (
        name.startswith("acquisition_")
        or "Process <Acquisition>" in joined
        or "CAcquisitionObject" in joined
    ):
        return "ACQUISITION"

    structured = sum(1 for line in lines if _C23_CSA_LINE_RE.match(line))
    if structured >= 2:
        if _C23_CGA_HINT_RE.search(joined) or "cga" in name:
            return "CGA"
        if _C23_CSA_HINT_RE.search(joined) or "csa" in name:
            return "CSA"

    if "cga" in name:
        return "CGA"
    if "csa" in name:
        return "CSA"
    return None


_old_c23_classify_file = classify_file
def _c23_classify_file(path_or_name, include_unknown=False):
    path = Path(path_or_name)
    name = path.name

    # Spectrum Dumps are Investigation assets only.
    if re.match(r"(?i)^Spectrum_.*\.dmp_FFT$", name):
        return None

    detected = _old_c23_classify_file(path_or_name, include_unknown)
    if detected and detected != "UNKNOWN":
        return detected

    if path.exists() and path.is_file():
        fallback = _c23_content_type_probe(path)
        if fallback:
            return fallback

    return "UNKNOWN" if include_unknown else None


classify_file = _c23_classify_file


def _c23_row_severity(row):
    values = [
        row.get("Type", ""),
        row.get("Level", ""),
        row.get("Status", ""),
        row.get("Error", ""),
        row.get("Message", ""),
    ]
    text = " ".join(str(value or "") for value in values).lower()

    if any(token in text for token in ("critical", "fatal", "ftl", "severe", " sev ")):
        return "CRITICAL"
    if any(token in text for token in ("error", " err ", "exception", "failed", "failure")):
        return "ERROR"
    if any(token in text for token in ("warning", "warn", " wrn ")):
        return "WARNING"
    if any(token in text for token in ("info", " inf ")):
        return "INFO"
    return "OTHER"


def _c23_quick_filter_counts(rows):
    counts = {
        "ALL": len(rows),
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "CRITICAL": 0,
    }
    for row in rows:
        severity = _c23_row_severity(row)
        if severity in counts:
            counts[severity] += 1
    return counts


def _c23_set_quick_filter(viewer, pane_index, mode):
    viewer.quick_filter_modes[pane_index] = str(mode).upper()
    viewer.apply_view_filters(pane_index)


def _c23_update_quick_filter_labels(viewer, pane_index):
    if not hasattr(viewer, "quick_filter_buttons"):
        return
    counts = _c23_quick_filter_counts(viewer.all_rows[pane_index])
    for mode, button in viewer.quick_filter_buttons[pane_index].items():
        label = "All" if mode == "ALL" else mode.title()
        button.setText(f"{label} ({counts.get(mode, 0):,})")
        button.setChecked(viewer.quick_filter_modes[pane_index] == mode)


_old_c23_viewer_build_ui = MultiPaneLogViewer.build_ui
def _c23_viewer_build_ui(self):
    _old_c23_viewer_build_ui(self)

    self.quick_filter_modes = ["ALL"] * self.MAX_PANES
    self.quick_filter_buttons = []

    for pane_index, pane in enumerate(self.panes):
        pane_layout = pane.layout()
        bar_widget = QWidget(pane)
        bar = QHBoxLayout(bar_widget)
        bar.setContentsMargins(0, 0, 0, 0)
        bar.setSpacing(4)
        bar.addWidget(QLabel("Quick Filter:"))

        buttons = {}
        for mode in ("ALL", "ERROR", "WARNING", "INFO", "CRITICAL"):
            button = QPushButton(mode.title())
            button.setCheckable(True)
            button.setAutoExclusive(True)
            button.setToolTip(
                f"Show {mode.title()} rows in {self.pane_name(pane_index)}. "
                "This is a Viewer-only memory filter."
            )
            button.clicked.connect(
                lambda checked=False, i=pane_index, m=mode:
                    _c23_set_quick_filter(self, i, m)
            )
            buttons[mode] = button
            bar.addWidget(button)

        buttons["ALL"].setChecked(True)
        bar.addStretch(1)
        self.quick_filter_buttons.append(buttons)

        # Insert below the search controls and before the file label.
        pane_layout.insertWidget(2, bar_widget)

    # Remove every inherited automatic Error expression.
    for edit in getattr(self, "foundation_filter_edits", []):
        pass  # Commit0043 removed legacy automatic filter body

MultiPaneLogViewer.build_ui = _c23_viewer_build_ui


def _c23_manual_expression_matches(row, expression):
    expression = str(expression or "").strip()
    if not expression:
        return True

    if "=" in expression:
        column, value = expression.split("=", 1)
        return str(row.get(column.strip(), "")).lower() == value.strip().lower()

    term = expression.lower()
    return term in " ".join(str(value or "") for value in row.values()).lower()


def _c23_apply_view_filters(self, side):
    pane_index = self.side_index(side)

    try:
        start, end = self.current_viewer_time_range()
    except Exception as error:
        QMessageBox.warning(self, "Viewer Time Range", str(error))
        return

    base_rows = self.all_rows[pane_index]
    rows = []

    manual_expression = ""
    if (
        hasattr(self, "foundation_filter_edits")
        and pane_index < len(self.foundation_filter_edits)
    ):
        manual_expression = self.foundation_filter_edits[pane_index].text().strip()
        if manual_expression.lower() == "type=err":
            # Commit0023 removes the old hidden/default behavior.
            self.foundation_filter_edits[pane_index].clear()
            manual_expression = ""

    quick_mode = (
        self.quick_filter_modes[pane_index]
        if hasattr(self, "quick_filter_modes")
        else "ALL"
    )

    for row in base_rows:
        timestamp = row.get("_ts")
        if start or end:
            if not isinstance(timestamp, datetime):
                continue
            if start and timestamp < start:
                continue
            if end and timestamp > end:
                continue

        if quick_mode != "ALL" and _c23_row_severity(row) != quick_mode:
            continue

        if not _c23_manual_expression_matches(row, manual_expression):
            continue

        rows.append(row)

    columns = self.pane_columns_for_rows(pane_index, rows or base_rows)
    self.models[pane_index].set_rows(rows, columns)
    self.apply_table_column_widths(pane_index)

    timestamp_index = [
        (row["_ts"], index)
        for index, row in enumerate(rows)
        if isinstance(row.get("_ts"), datetime)
    ]
    timestamp_index.sort(key=lambda item: item[0])
    self.ts_indexes[pane_index] = timestamp_index
    self._sync_aliases()

    _c23_update_quick_filter_labels(self, pane_index)

    source = self.sources[pane_index].currentText()
    self.log(
        f"{source} pipeline — parsed/indexed: {len(base_rows):,}; "
        f"quick filter: {quick_mode}; displayed: {len(rows):,}"
    )


MultiPaneLogViewer.apply_view_filters = _c23_apply_view_filters


_old_c23_source_to_records = MultiPaneLogViewer.source_to_records
def _c23_source_to_records(self, source_name, progress=None):
    records = _old_c23_source_to_records(self, source_name, progress)

    # CSA recovery path: scan files with both filename and content detection.
    if source_name.upper() == "CSA" and not records:
        source_folder = Path(self.parent_window.source_edit.text().strip() or ".")
        recursive = self.parent_window.recursive_chk.isChecked()
        candidates = []

        if source_folder.exists():
            for path in iter_files(source_folder, recursive):
                if progress and progress.wasCanceled():
                    return []
                try:
                    detected = _c23_classify_file(path, False)
                except Exception:
                    detected = None
                if detected == "CSA":
                    candidates.append(path)

        if progress:
            progress.setRange(0, max(1, len(candidates)))
            progress.setValue(0)
            progress.setLabelText(
                f"CSA recovery parser: 0/{len(candidates)} files"
            )
            QApplication.processEvents()

        recovered = []
        for file_index, path in enumerate(candidates, 1):
            if progress and progress.wasCanceled():
                return []
            try:
                parsed = parse_file(path, "CSA")
                recovered.extend(parsed)
            except Exception:
                write_startup_log(
                    f"CSA recovery parser failed: {path}\n"
                    + traceback.format_exc()
                )

            if progress:
                progress.setValue(file_index)
                progress.setLabelText(
                    f"CSA recovery parser: {file_index}/{len(candidates)} files; "
                    f"{len(recovered):,} rows"
                )
                QApplication.processEvents()

        records = recovered
        self.log(
            f"CSA recovery — detected files: {len(candidates)}; "
            f"parsed rows: {len(records):,}"
        )

    return records


MultiPaneLogViewer.source_to_records = _c23_source_to_records


_old_c23_load_pane = MultiPaneLogViewer.load_pane
def _c23_load_pane(self, side):
    pane_index = self.side_index(side)

    # No type starts with an Error filter.
    if (
        hasattr(self, "foundation_filter_edits")
        and pane_index < len(self.foundation_filter_edits)
    ):
        edit = self.foundation_filter_edits[pane_index]

    if hasattr(self, "quick_filter_modes"):
        self.quick_filter_modes[pane_index] = "ALL"

    result = _old_c23_load_pane(self, side)
    _c23_update_quick_filter_labels(self, pane_index)

    source = self.sources[pane_index].currentText()
    if source.upper() == "CSA":
        parsed_count = len(self.all_rows[pane_index])
        displayed_count = self.models[pane_index].rowCount()
        self.log(
            f"CSA pipeline complete — indexed: {parsed_count:,}; "
            f"displayed: {displayed_count:,}"
        )
        if parsed_count == 0:
            QMessageBox.warning(
                self,
                "CSA Viewer",
                "CSA was selected but no rows were indexed.\n\n"
                "The startup log contains file detection and parser details.",
            )

    return result


MultiPaneLogViewer.load_pane = _c23_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)


_old_c23_refresh_sources = getattr(
    MultiPaneLogViewer,
    "refresh_available_sources",
    None,
)
def _c23_refresh_available_sources(self, *args, **kwargs):
    if callable(_old_c23_refresh_sources):
        result = _old_c23_refresh_sources(self, *args, **kwargs)
    else:
        result = None

    parent = getattr(self, "parent_window", None)
    selected = getattr(parent, "viewer_selected_files", []) or []
    detected = set()

    for value in selected:
        path = Path(value)
        detected_type = _c23_classify_file(path, False)
        if detected_type:
            detected.add(detected_type)

    # A selected CSA file must always make CSA available in every pane.
    if "CSA" in detected:
        for combo in self.sources:
            if combo.findText("CSA") < 0:
                combo.addItem("CSA")

    # Spectrum Dump remains excluded from normal Viewer.
    for combo in self.sources:
        for index in range(combo.count() - 1, -1, -1):
            if "SPECTRUM" in combo.itemText(index).upper():
                combo.removeItem(index)

    return result


if callable(_old_c23_refresh_sources):
    MultiPaneLogViewer.refresh_available_sources = _c23_refresh_available_sources


# ---------------------------------------------------------------------------
# Commit0024: independent Spectrum Analysis and graphical Acquisition dashboard
# ---------------------------------------------------------------------------

from foundation.investigation import InvestigationWorkspace
from foundation.spectrum_analysis import SpectrumAnalysisWidget


class StandaloneSpectrumWindow(QMainWindow):
    def __init__(self, parent=None, initial_paths=None):
        super().__init__(parent)
        self.setWindowTitle(f"Spectrum Analysis — {APP_VERSION}")
        self.resize(1380, 860)
        self.setAcceptDrops(True)

        container = QWidget(self)
        layout = QVBoxLayout(container)
        layout.setContentsMargins(6, 6, 6, 6)

        path_bar = QHBoxLayout()
        path_bar.addWidget(QLabel("Search folder:"))
        self.folder_edit = QLineEdit()
        self.folder_edit.setPlaceholderText(
            "Select a folder or drop Spectrum_*.dmp_FFT files/folders here"
        )
        browse_files = QPushButton("Select Files")
        browse_files.clicked.connect(self.browse_files)
        browse_folder = QPushButton("Select Folder")
        browse_folder.clicked.connect(self.browse_folder)
        browse_zip = QPushButton("Select ZIP")
        browse_zip.clicked.connect(self.browse_zip)
        scan = QPushButton("Scan Selected Source")
        scan.clicked.connect(self.scan_folder)
        path_bar.addWidget(self.folder_edit, 1)
        path_bar.addWidget(browse_files)
        path_bar.addWidget(browse_folder)
        path_bar.addWidget(browse_zip)
        path_bar.addWidget(scan)
        layout.addLayout(path_bar)

        self.analysis = SpectrumAnalysisWidget(
            investigation=None,
            standalone=True,
        )
        layout.addWidget(self.analysis, 1)
        self.setCentralWidget(container)

        if initial_paths:
            QTimer.singleShot(
                0,
                lambda paths=list(initial_paths): self.load_paths(paths),
            )

    def browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Spectrum Dump Files",
            self.folder_edit.text().strip() or str(Path.home()),
            "Spectrum Dumps (*.dmp_FFT *.DMP_FFT *.dmp.fft);;All Files (*.*)",
        )
        if files:
            self.folder_edit.setText(str(Path(files[0]).parent))
            self.load_paths(files)

    def browse_zip(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select ZIP Containing Spectrum Dumps",
            self.folder_edit.text().strip() or str(Path.home()),
            "ZIP Files (*.zip);;All Files (*.*)",
        )
        if path:
            self.folder_edit.setText(path)
            self.load_paths([path])

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(
            self,
            "Select Spectrum Search Folder",
            self.folder_edit.text().strip() or str(Path.home()),
        )
        if folder:
            self.folder_edit.setText(folder)
            self.scan_folder()

    def scan_folder(self):
        source = Path(self.folder_edit.text().strip())
        if not source.exists():
            QMessageBox.warning(
                self,
                "Spectrum Analysis",
                "The selected file, ZIP or folder does not exist.",
            )
            return
        self.analysis.load_paths([source])

    def load_paths(self, values):
        paths = [Path(value) for value in values]
        folders = [path for path in paths if path.is_dir()]
        if folders:
            self.folder_edit.setText(str(folders[0]))
        elif paths:
            self.folder_edit.setText(str(paths[0].parent))
        self.analysis.load_paths(paths)

    def dragEnterEvent(self, event):
        urls = event.mimeData().urls()
        if any(url.isLocalFile() for url in urls):
            event.acceptProposedAction()

    def dropEvent(self, event):
        paths = [
            Path(url.toLocalFile())
            for url in event.mimeData().urls()
            if url.isLocalFile()
        ]
        if paths:
            self.load_paths(paths)
            event.acceptProposedAction()


def _c24_open_spectrum_analysis(self, paths=None):
    window = StandaloneSpectrumWindow(
        self,
        initial_paths=paths or [],
    )
    window.setAttribute(Qt.WA_DeleteOnClose, True)
    window.show()
    window.raise_()
    window.activateWindow()

    if not hasattr(self, "_spectrum_windows"):
        self._spectrum_windows = []
    self._spectrum_windows.append(window)
    window.destroyed.connect(
        lambda *_: self._spectrum_windows.remove(window)
        if window in self._spectrum_windows else None
    )
    return window


_old_c24_build_ui = MainWindow.build_ui
def _c24_build_ui(self):
    _old_c24_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")

    spectrum_button = QPushButton("Spectrum Analysis — Standalone")
    spectrum_button.setObjectName("standaloneSpectrumButton")
    spectrum_button.setToolTip(
        "Open Spectrum Analysis independently from Log Viewer. Select or drop "
        "Spectrum files, a folder, or a ZIP package."
    )
    spectrum_button.clicked.connect(
        lambda checked=False: _c24_open_spectrum_analysis(self)
    )

    # Add to the top-level action area near Investigation/Viewer controls.
    inserted = False
    for layout_name in (
        "top_action_layout",
        "action_layout",
        "main_button_layout",
        "button_layout",
    ):
        layout = getattr(self, layout_name, None)
        if layout is not None and hasattr(layout, "addWidget"):
            layout.addWidget(spectrum_button)
            inserted = True
            break

    if not inserted:
        # MainWindow is a QWidget, not a QMainWindow. Add the button directly
        # to the existing root layout instead of calling centralWidget().
        root = self.layout()
        if root is not None:
            insert_index = 1 if root.count() >= 1 else 0
            root.insertWidget(insert_index, spectrum_button)
            inserted = True

    if not inserted:
        # Do not crash startup when an unexpected layout is used.
        spectrum_button.setParent(self)
        spectrum_button.move(12, 72)
        spectrum_button.show()

    self.spectrum_analysis_btn = spectrum_button


MainWindow.build_ui = _c24_build_ui
MainWindow.open_spectrum_analysis = _c24_open_spectrum_analysis


# ---------------------------------------------------------------------------
# Commit0030: Header right-click arbitrary column text filters
# ---------------------------------------------------------------------------

def _c30_parse_column_filter(expression):
    expression = str(expression or "").strip()
    if not expression:
        return None, None, None

    # Column~text means case-insensitive partial match.
    if "~" in expression:
        column, value = expression.split("~", 1)
        column = column.strip()
        if column:
            return "contains", column, value

    # Column=text means case-insensitive exact match.
    if "=" in expression:
        column, value = expression.split("=", 1)
        column = column.strip()
        if column:
            return "exact", column, value

    return "global", None, expression


def _c30_manual_expression_matches(row, expression):
    mode, column, value = _c30_parse_column_filter(expression)
    if mode is None:
        return True

    search_text = str(value or "").casefold()

    if mode == "contains":
        cell_text = str(row.get(column, "") or "").casefold()
        return search_text in cell_text

    if mode == "exact":
        cell_text = str(row.get(column, "") or "").casefold()
        return cell_text == search_text

    return search_text in " ".join(
        str(cell_value or "") for cell_value in row.values()
    ).casefold()


# Commit0023 Viewer filter implementation resolves this name at runtime.
_c23_manual_expression_matches = _c30_manual_expression_matches


def _c30_filter_edit(viewer, pane_index):
    edits = getattr(viewer, "foundation_filter_edits", [])
    if 0 <= pane_index < len(edits):
        return edits[pane_index]
    return None


def _c30_column_name(viewer, pane_index, logical_index):
    model = viewer.models[pane_index]
    columns = getattr(model, "columns", []) or []
    if 0 <= logical_index < len(columns):
        return str(columns[logical_index])

    header = model.headerData(
        logical_index,
        Qt.Horizontal,
        Qt.DisplayRole,
    )
    return str(header or "").strip()


def _c30_apply_header_filter(
    viewer,
    pane_index,
    column_name,
    mode,
):
    edit = _c30_filter_edit(viewer, pane_index)
    if edit is None:
        QMessageBox.warning(
            viewer,
            "Column Filter",
            "The filter field for this Viewer pane is unavailable.",
        )
        return

    current_mode, current_column, current_value = _c30_parse_column_filter(
        edit.text()
    )
    default_text = (
        str(current_value or "")
        if current_column == column_name
        else ""
    )

    title = f"Filter: {column_name}"
    if mode == "contains":
        prompt = (
            f"Show rows where {column_name} contains this text "
            "(case-insensitive):"
        )
    else:
        prompt = (
            f"Show rows where {column_name} exactly matches this text "
            "(case-insensitive):"
        )

    value, accepted = QInputDialog.getText(
        viewer,
        title,
        prompt,
        text=default_text,
    )
    if not accepted:
        return

    value = str(value)
    if not value:
        edit.clear()
    elif mode == "contains":
        edit.setText(f"{column_name}~{value}")
    else:
        edit.setText(f"{column_name}={value}")

    viewer.apply_view_filters(pane_index)


def _c30_clear_header_filter(viewer, pane_index):
    edit = _c30_filter_edit(viewer, pane_index)
    if edit is not None:
        edit.clear()
        viewer.apply_view_filters(pane_index)


def _c30_show_header_context_menu(viewer, pane_index, position):
    if not (0 <= pane_index < len(viewer.tables)):
        return

    table = viewer.tables[pane_index]
    header = table.horizontalHeader()
    logical_index = header.logicalIndexAt(position)
    if logical_index < 0:
        return

    column_name = _c30_column_name(
        viewer,
        pane_index,
        logical_index,
    )
    if not column_name:
        return

    menu = QMenu(header)
    title_action = menu.addAction(f"Column: {column_name}")
    title_action.setEnabled(False)
    menu.addSeparator()

    contains_action = menu.addAction("Filter contains...")
    exact_action = menu.addAction("Filter exact...")
    menu.addSeparator()
    clear_action = menu.addAction("Clear column filter")

    edit = _c30_filter_edit(viewer, pane_index)
    has_filter = bool(edit and edit.text().strip())
    clear_action.setEnabled(has_filter)

    selected = menu.exec(header.mapToGlobal(position))
    if selected is contains_action:
        _c30_apply_header_filter(
            viewer,
            pane_index,
            column_name,
            "contains",
        )
    elif selected is exact_action:
        _c30_apply_header_filter(
            viewer,
            pane_index,
            column_name,
            "exact",
        )
    elif selected is clear_action:
        _c30_clear_header_filter(viewer, pane_index)


def _c30_install_header_filters(viewer):
    for pane_index, table in enumerate(viewer.tables):
        header = table.horizontalHeader()

        # Avoid duplicate connections if the UI is rebuilt.
        if header.property("c30ColumnFilterInstalled"):
            continue

        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(
            lambda position, index=pane_index:
                _c30_show_header_context_menu(
                    viewer,
                    index,
                    position,
                )
        )
        header.setProperty("c30ColumnFilterInstalled", True)
        header.setToolTip(
            "Right-click a column header to filter that column "
            "by arbitrary text."
        )


_old_c30_viewer_build_ui = MultiPaneLogViewer.build_ui
def _c30_viewer_build_ui(self):
    _old_c30_viewer_build_ui(self)
    _c30_install_header_filters(self)


MultiPaneLogViewer.build_ui = _c30_viewer_build_ui

# Real final entry point: all Commit0016 and Commit0017 overrides above are active.

# ---------------------------------------------------------------------------
# Commit0032B: GESYS / LAIS / PSC / review.out detection and parser recovery
# ---------------------------------------------------------------------------

_C32B_TYPE_ALIASES = {
    "GESYS": "GESYS",
    "GESYSLOG": "GESYS",
    "GE_SYS": "GESYS",
    "LAIS": "LAIS",
    "PSC": "PSC",
    "REVIEW": "Review",
    "REVIEW.OUT": "Review",
    "REVIEWOUT": "Review",
}


def _c32b_normalize_log_type(value):
    text_value = str(value or "").strip()
    if not text_value:
        return ""
    upper = text_value.upper().replace(" ", "").replace("-", "").replace("_", "")
    return _C32B_TYPE_ALIASES.get(upper, text_value)


def _c32b_filename_type(path):
    """Reliable filename classification for the four recovered formats."""
    name = Path(path).name
    lower = name.lower()

    # review.out variants must be checked before generic extensions.
    if (
        lower == "review.out"
        or lower.startswith("review.out.")
        or lower.startswith("review_")
        or "review.out" in lower
    ):
        return "Review"

    if (
        lower.startswith("gesys")
        or "gesyslog" in lower
        or re.search(r"(^|[._-])gesys([._-]|$)", lower)
    ):
        return "GESYS"

    if (
        lower.startswith("lais")
        or re.search(r"(^|[._-])lais([._-]|$)", lower)
    ):
        return "LAIS"

    if (
        lower == "psc.log"
        or lower.startswith("psc.")
        or lower.startswith("psc_")
        or re.search(r"(^|[._-])psc([._-]|$)", lower)
    ):
        return "PSC"

    return ""


def _c32b_probe_content_type(path, sample_size=262144):
    """Content fallback when filenames were renamed or copied."""
    try:
        raw = Path(path).read_bytes()[:sample_size]
    except Exception:
        return ""

    text_sample = raw.decode("utf-8", errors="ignore")
    if not text_sample.strip():
        text_sample = raw.decode("latin-1", errors="ignore")

    low = text_sample.lower()

    # Review.out commonly contains scan/review structured lines and coil/protocol fields.
    if (
        "review.out" in low
        or ("protocol" in low and "coil" in low and "series" in low)
        or ("review" in low and "scan" in low and "patient" in low)
    ):
        return "Review"

    # PSC signatures observed in service logs.
    if (
        "[psc" in low
        or "psc::" in low
        or "power supply controller" in low
        or "psc state" in low
    ):
        return "PSC"

    # LAIS signatures.
    if (
        "[lais" in low
        or "lais::" in low
        or "laser alignment" in low
        or "lais state" in low
    ):
        return "LAIS"

    # GESYS signatures.
    if (
        "gesys" in low
        or "ge sys" in low
        or "gemr" in low
        or "system event log" in low
    ):
        return "GESYS"

    return ""


def _c32b_detect_recovered_type(path):
    return _c32b_filename_type(path) or _c32b_probe_content_type(path)


def _c32b_timestamp_from_text(line, fallback=None):
    """Extract common timestamps without discarding valid rows."""
    value = str(line or "")

    patterns = [
        # YYYY/MM/DD HH:MM:SS(.sss)
        r"(?P<y>\d{4})[/-](?P<m>\d{1,2})[/-](?P<d>\d{1,2})[ T]"
        r"(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})(?:[.,:](?P<ms>\d{1,6}))?",
        # DD/MM/YYYY HH:MM:SS(.sss)
        r"(?P<d>\d{1,2})[/-](?P<m>\d{1,2})[/-](?P<y>\d{4})[ T]"
        r"(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})(?:[.,:](?P<ms>\d{1,6}))?",
    ]

    for pattern in patterns:
        match = re.search(pattern, value)
        if not match:
            continue
        try:
            parts = match.groupdict()
            micros = (parts.get("ms") or "0")[:6].ljust(6, "0")
            return datetime(
                int(parts["y"]),
                int(parts["m"]),
                int(parts["d"]),
                int(parts["h"]),
                int(parts["mi"]),
                int(parts["s"]),
                int(micros),
            )
        except Exception:
            pass

    # Time-only records: retain fallback date where available.
    time_match = re.search(
        r"(?<!\d)(?P<h>\d{1,2}):(?P<mi>\d{2}):(?P<s>\d{2})"
        r"(?:[.,:](?P<ms>\d{1,6}))?(?!\d)",
        value,
    )
    if time_match:
        try:
            base = fallback if isinstance(fallback, datetime) else datetime(1970, 1, 1)
            parts = time_match.groupdict()
            micros = (parts.get("ms") or "0")[:6].ljust(6, "0")
            return base.replace(
                hour=int(parts["h"]),
                minute=int(parts["mi"]),
                second=int(parts["s"]),
                microsecond=int(micros),
            )
        except Exception:
            pass

    return fallback


def _c32b_make_record(source_type, path, line_no, line, timestamp=None):
    """Create the canonical record without rejecting non-empty rows."""
    message = str(line or "").rstrip("\r\n")
    level = ""

    level_match = re.search(
        r"(?i)(?:^|\s|\[)(dbg|debug|inf|info|wrn|warn|warning|err|error|critical|fatal|ext)(?:\]|\s|$)",
        message,
    )
    if level_match:
        level = level_match.group(1)

    category = ""
    source_match = re.search(r"\[([^\[\]]{1,120})\]", message)
    if source_match:
        category = source_match.group(1).strip()

    return LogRecord(
        timestamp=timestamp,
        source_type=source_type,
        filename=Path(path).name,
        line_no=int(line_no),
        level=level,
        category=category,
        message=message,
        raw=message,
    )


def _c32b_parse_text_log(path, source_type):
    """Loss-tolerant parser used as recovery fallback."""
    records = []
    current_date_time = None

    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    content = None
    for encoding in encodings:
        try:
            content = Path(path).read_text(encoding=encoding)
            break
        except Exception:
            continue

    if content is None:
        return records

    for line_no, line in enumerate(content.splitlines(), 1):
        if not line.strip():
            continue

        timestamp = _c32b_timestamp_from_text(line, current_date_time)
        if isinstance(timestamp, datetime):
            current_date_time = timestamp

        records.append(
            _c32b_make_record(
                source_type,
                path,
                line_no,
                line,
                timestamp,
            )
        )

    return records


def _c32b_parser_dispatch(path, detected_type):
    source_type = _c32b_normalize_log_type(detected_type)

    # Use any existing dedicated parser first. If it raises or returns no rows,
    # fall back to the loss-tolerant parser.
    parser_names = {
        "GESYS": (
            "parse_gesys",
            "parse_gesyslog",
            "parse_gesys_log",
        ),
        "LAIS": (
            "parse_lais",
            "parse_lais_log",
        ),
        "PSC": (
            "parse_psc",
            "parse_psc_log",
        ),
        "Review": (
            "parse_review",
            "parse_review_out",
            "parse_reviewout",
        ),
    }

    for parser_name in parser_names.get(source_type, ()):
        parser = globals().get(parser_name)
        if not callable(parser):
            continue
        try:
            parsed = parser(path)
            parsed = list(parsed or [])
            if parsed:
                return parsed
        except Exception:
            pass

    return _c32b_parse_text_log(path, source_type)


# Preserve and wrap the current detector if one is present.
_c32b_detector_names = (
    "detect_log_type",
    "detect_file_type",
    "classify_log_file",
    "guess_log_type",
)
for _c32b_name in _c32b_detector_names:
    _c32b_original = globals().get(_c32b_name)
    if callable(_c32b_original):
        def _c32b_wrapped_detector(path, _original=_c32b_original):
            recovered = _c32b_detect_recovered_type(path)
            if recovered:
                return recovered
            try:
                return _original(path)
            except Exception:
                return ""
        globals()[_c32b_name] = _c32b_wrapped_detector


# Patch parser registries found in the current source.
for _c32b_registry_name in (
    "PARSERS",
    "PARSER_MAP",
    "LOG_PARSERS",
    "FILE_TYPE_PARSERS",
):
    _c32b_registry = globals().get(_c32b_registry_name)
    if isinstance(_c32b_registry, dict):
        for _c32b_type in ("GESYS", "LAIS", "PSC", "Review"):
            _c32b_registry[_c32b_type] = (
                lambda path, detected_type=_c32b_type:
                    _c32b_parser_dispatch(path, detected_type)
            )


# Expose explicit parser functions for dynamic lookups.
def parse_gesys(path):
    return _c32b_parse_text_log(path, "GESYS")


def parse_lais(path):
    return _c32b_parse_text_log(path, "LAIS")


def parse_psc(path):
    return _c32b_parse_text_log(path, "PSC")


def parse_review_out(path):
    return _c32b_parse_text_log(path, "Review")


# Ensure normalized names are accepted wherever type lists are consulted.
for _c32b_collection_name in (
    "SUPPORTED_TYPES",
    "LOG_TYPES",
    "KNOWN_LOG_TYPES",
    "DEFAULT_LOG_TYPES",
):
    _c32b_collection = globals().get(_c32b_collection_name)
    if isinstance(_c32b_collection, list):
        for _c32b_type in ("GESYS", "LAIS", "PSC", "Review"):
            if _c32b_type not in _c32b_collection:
                _c32b_collection.append(_c32b_type)
    elif isinstance(_c32b_collection, set):
        _c32b_collection.update(("GESYS", "LAIS", "PSC", "Review"))



# ---------------------------------------------------------------------------
# Commit0032C: Header filter value dropdown for low-cardinality columns
# ---------------------------------------------------------------------------

def _c32c_column_unique_values(viewer, pane_index, column_name, limit=11):
    """Return sorted unique displayed-model values, stopping at limit.

    The dropdown is offered only when the actual distinct-value count is
    between 1 and 10. Reaching 11 values means the menu must not show it.
    """
    if not (0 <= pane_index < len(viewer.models)):
        return []

    model = viewer.models[pane_index]
    rows = getattr(model, "filtered_rows", None)
    if rows is None:
        rows = getattr(model, "rows", None)
    if rows is None:
        rows = getattr(model, "_rows", None)
    if rows is None:
        return []

    unique = {}
    for row in rows:
        try:
            if hasattr(row, "get"):
                raw_value = row.get(column_name, "")
            elif isinstance(row, dict):
                raw_value = row.get(column_name, "")
            else:
                raw_value = getattr(row, column_name, "")
        except Exception:
            raw_value = ""

        display_value = "" if raw_value is None else str(raw_value)
        comparison_key = display_value.casefold()

        if comparison_key not in unique:
            unique[comparison_key] = display_value
            if len(unique) >= limit:
                return []

    return sorted(
        unique.values(),
        key=lambda value: (value == "", value.casefold()),
    )


def _c32c_apply_selected_value(viewer, pane_index, column_name, value):
    edit = _c30_filter_edit(viewer, pane_index)
    if edit is None:
        return

    # Exact-match syntax used by the existing column filter implementation.
    edit.setText(f"{column_name}={value}")
    viewer.apply_view_filters(pane_index)


def _c32c_show_header_context_menu(viewer, pane_index, position):
    if not (0 <= pane_index < len(viewer.tables)):
        return

    table = viewer.tables[pane_index]
    header = table.horizontalHeader()
    logical_index = header.logicalIndexAt(position)
    if logical_index < 0:
        return

    column_name = _c30_column_name(
        viewer,
        pane_index,
        logical_index,
    )
    if not column_name:
        return

    menu = QMenu(header)
    title_action = menu.addAction(f"Column: {column_name}")
    title_action.setEnabled(False)
    menu.addSeparator()

    values = _c32c_column_unique_values(
        viewer,
        pane_index,
        column_name,
        11,
    )

    value_actions = {}
    if 1 <= len(values) <= 10:
        value_menu = menu.addMenu("Select value")
        for value in values:
            label = "(Blank)" if value == "" else value
            action = value_menu.addAction(label)
            value_actions[action] = value
        menu.addSeparator()

    contains_action = menu.addAction("Filter contains...")
    exact_action = menu.addAction("Filter exact...")
    menu.addSeparator()
    clear_action = menu.addAction("Clear column filter")

    edit = _c30_filter_edit(viewer, pane_index)
    has_filter = bool(edit and edit.text().strip())
    clear_action.setEnabled(has_filter)

    selected = menu.exec(header.mapToGlobal(position))

    if selected in value_actions:
        _c32c_apply_selected_value(
            viewer,
            pane_index,
            column_name,
            value_actions[selected],
        )
    elif selected is contains_action:
        _c30_apply_header_filter(
            viewer,
            pane_index,
            column_name,
            "contains",
        )
    elif selected is exact_action:
        _c30_apply_header_filter(
            viewer,
            pane_index,
            column_name,
            "exact",
        )
    elif selected is clear_action:
        _c30_clear_header_filter(viewer, pane_index)


# Replace the Commit0030 menu handler while retaining its installation path.
_c30_show_header_context_menu = _c32c_show_header_context_menu



# ---------------------------------------------------------------------------
# Commit0033: definitive GESYS/LAIS/PSC/Review Viewer path recovery
# ---------------------------------------------------------------------------

_C33_RECOVERED_TYPES = {"GESYS", "LAIS", "PSC", "REVIEW"}


def _c33_canonical_type(value):
    normalized = str(value or "").strip().upper().replace(" ", "")
    aliases = {
        "GESYSLOG": "GESYS",
        "GESYS": "GESYS",
        "LAIS": "LAIS",
        "PSC": "PSC",
        "REVIEW": "REVIEW",
        "REVIEW.OUT": "REVIEW",
        "REVIEWOUT": "REVIEW",
    }
    return aliases.get(normalized, normalized)


def _c33_classify_recovered(path):
    name = Path(path).name.lower()

    if name == "review.out" or name.startswith("review.out.") or "review.out" in name:
        return "REVIEW"
    if name.startswith("gesys") or "gesyslog" in name:
        return "GESYS"
    if name.startswith("lais") or re.search(r"(^|[._-])lais([._-]|$)", name):
        return "LAIS"
    if (
        name == "psc.log"
        or name.startswith("psc.")
        or name.startswith("psc_")
        or re.search(r"(^|[._-])psc([._-]|$)", name)
    ):
        return "PSC"
    return ""


_c33_old_classify_file = classify_file
def _c33_classify_file(path, include_unknown=False):
    recovered = _c33_classify_recovered(path)
    if recovered:
        return recovered
    result = _c33_old_classify_file(Path(path), include_unknown)
    return "REVIEW" if str(result).upper() == "REVIEW" else result


classify_file = _c33_classify_file


def _c33_parse_recovered(path, source_type):
    canonical = _c33_canonical_type(source_type)

    if canonical == "PSC":
        try:
            _, _, records = parse_psc_file_detail(Path(path))
            if records:
                return records
        except Exception:
            pass

    if canonical == "REVIEW":
        try:
            records = review_rows_to_viewer_records(Path(path))
            if records:
                return records
        except Exception:
            pass

    # GESYS and LAIS, plus fallback for PSC/Review.
    try:
        records = parse_file(Path(path), canonical)
        if records:
            return records
    except Exception:
        pass

    # Loss-tolerant final fallback preserving every non-empty line.
    records = []
    fallback = parse_filename_datetime(Path(path))
    last_ts = fallback
    for line_no, line in enumerate(read_text_lines(Path(path)), 1):
        message = str(line).strip()
        if not message:
            continue
        try:
            timestamp, _ = extract_content_timestamp(
                message,
                canonical,
                last_ts.date() if isinstance(last_ts, datetime) else None,
                last_ts,
            )
        except Exception:
            timestamp = last_ts
        if isinstance(timestamp, datetime):
            last_ts = timestamp
        records.append(
            LogRecord(
                timestamp,
                "Review" if canonical == "REVIEW" else canonical,
                Path(path).name,
                line_no,
                detect_level(message),
                "",
                message,
                message,
            )
        )
    return records


_c33_old_source_to_records = MultiPaneLogViewer.source_to_records
def _c33_source_to_records(self, source_name, progress=None):
    canonical = _c33_canonical_type(source_name)
    if canonical not in _C33_RECOVERED_TYPES:
        return _c33_old_source_to_records(self, source_name, progress)

    parent = self.parent_window
    selected = list(getattr(parent, "viewer_selected_files", []) or [])
    source_folder = Path(parent.source_edit.text().strip() or ".")
    recursive = parent.recursive_chk.isChecked()

    candidates = []
    seen = set()

    # First priority: exact files chosen in Smart File Discovery.
    for value in selected:
        path = Path(value)
        if not path.exists() or not path.is_file():
            continue
        if _c33_canonical_type(_c33_classify_file(path, True)) == canonical:
            key = str(path.resolve()).casefold()
            if key not in seen:
                seen.add(key)
                candidates.append(path)

    # Fallback: source folder scan.
    if not candidates and source_folder.exists():
        for path in iter_files(source_folder, recursive):
            if progress and progress.wasCanceled():
                return []
            if _c33_canonical_type(_c33_classify_file(path, True)) == canonical:
                key = str(path.resolve()).casefold()
                if key not in seen:
                    seen.add(key)
                    candidates.append(path)

    if progress:
        progress.setRange(0, max(1, len(candidates)))
        progress.setValue(0)
        progress.setLabelText(
            f"Loading {canonical}: 0/{len(candidates)} files"
        )
        QApplication.processEvents()

    records = []
    for index, path in enumerate(candidates, 1):
        if progress and progress.wasCanceled():
            return []
        parsed = _c33_parse_recovered(path, canonical)
        records.extend(parsed)
        if progress:
            progress.setValue(index)
            progress.setLabelText(
                f"Loading {canonical}: {index}/{len(candidates)} files; "
                f"{len(records):,} rows"
            )
            QApplication.processEvents()

    self.log(
        f"{canonical} definitive path — selected files: {len(selected)}; "
        f"matched files: {len(candidates)}; rows: {len(records):,}"
    )
    return records


MultiPaneLogViewer.source_to_records = _c33_source_to_records


_c33_old_refresh_sources = getattr(MultiPaneLogViewer, "refresh_available_sources", None)
def _c33_refresh_available_sources(self, *args, **kwargs):
    result = (
        _c33_old_refresh_sources(self, *args, **kwargs)
        if callable(_c33_old_refresh_sources)
        else None
    )

    parent = self.parent_window
    selected = list(getattr(parent, "viewer_selected_files", []) or [])
    found = {
        _c33_canonical_type(_c33_classify_file(Path(value), True))
        for value in selected
        if Path(value).exists()
    }

    labels = {
        "GESYS": "GESYS",
        "LAIS": "LAIS",
        "PSC": "PSC",
        "REVIEW": "Review",
    }
    for combo in self.sources:
        for canonical, label in labels.items():
            if canonical in found and combo.findText(label) < 0:
                combo.addItem(label)

    return result


if callable(_c33_old_refresh_sources):
    MultiPaneLogViewer.refresh_available_sources = _c33_refresh_available_sources



# ---------------------------------------------------------------------------
# Commit0034: strict CSA/CGA filename scope on every import path
# ---------------------------------------------------------------------------

def _c34_strict_brain_type(path):
    name = Path(path).name.lower()
    supported_extension = Path(path).suffix.lower() in {".log", ".txt"}

    if supported_extension and name.startswith("csa_brain"):
        return "CSA"
    if supported_extension and name.startswith("cga_brain"):
        return "CGA"
    return ""


_c34_previous_classify_file = classify_file
def _c34_classify_file(path, include_unknown=False):
    path = Path(path)
    name = path.name.lower()

    strict_type = _c34_strict_brain_type(path)
    if strict_type:
        return strict_type

    # Any other CSA/CGA-looking filename must not leak through an older,
    # broader classifier.
    if name.startswith("csa") or name.startswith("cga"):
        return "UNKNOWN" if include_unknown else None

    result = _c34_previous_classify_file(path, include_unknown)
    if str(result or "").upper() in {"CSA", "CGA"}:
        return "UNKNOWN" if include_unknown else None
    return result


classify_file = _c34_classify_file


_c34_previous_source_to_records = MultiPaneLogViewer.source_to_records
def _c34_source_to_records(self, source_name, progress=None):
    source_upper = str(source_name or "").upper()
    if source_upper not in {"CSA", "CGA"}:
        return _c34_previous_source_to_records(self, source_name, progress)

    parent = self.parent_window
    selected = list(getattr(parent, "viewer_selected_files", []) or [])
    source_folder = Path(parent.source_edit.text().strip() or ".")
    recursive = parent.recursive_chk.isChecked()

    candidates = []
    seen = set()

    for value in selected:
        path = Path(value)
        if not path.exists() or not path.is_file():
            continue
        if _c34_strict_brain_type(path) == source_upper:
            key = str(path.resolve()).casefold()
            if key not in seen:
                seen.add(key)
                candidates.append(path)

    if not candidates and source_folder.exists():
        for path in iter_files(source_folder, recursive):
            if progress and progress.wasCanceled():
                return []
            if _c34_strict_brain_type(path) == source_upper:
                key = str(path.resolve()).casefold()
                if key not in seen:
                    seen.add(key)
                    candidates.append(path)

    if progress:
        progress.setRange(0, max(1, len(candidates)))
        progress.setValue(0)
        progress.setLabelText(
            f"Loading {source_upper}: 0/{len(candidates)} strict brain files"
        )
        QApplication.processEvents()

    records = []
    for index, path in enumerate(candidates, 1):
        if progress and progress.wasCanceled():
            return []

        try:
            parsed = parse_csa_cga_file(path, source_upper)
        except Exception:
            try:
                parsed = parse_file(path, source_upper)
            except Exception:
                parsed = []

        records.extend(list(parsed or []))

        if progress:
            progress.setValue(index)
            progress.setLabelText(
                f"Loading {source_upper}: {index}/{len(candidates)} files; "
                f"{len(records):,} rows"
            )
            QApplication.processEvents()

    self.log(
        f"{source_upper} strict scope — matched files: {len(candidates)}; "
        f"rows: {len(records):,}"
    )
    return records


MultiPaneLogViewer.source_to_records = _c34_source_to_records




# ---------------------------------------------------------------------------
# Commit0043 RC1 Foundation Cleanup
# One Review parser, one Discovery path, one ZIP cache path and one Viewer
# projection. Replaces overlapping Commit0037-0042 patches.
# ---------------------------------------------------------------------------

def _f43_is_review_file(path):
    name = Path(path).name.lower()
    return (
        name == "review.out"
        or name.startswith("review.out.")
        or name.startswith("review_")
        or name.startswith("review-")
    )


def _f43_decode_review_text(path):
    try:
        raw = Path(path).read_bytes()
    except Exception:
        return ""
    if not raw:
        return ""

    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        for encoding in ("utf-16", "utf-16-le", "utf-16-be"):
            try:
                value = raw.decode(encoding)
                if value.strip():
                    return value
            except Exception:
                pass

    if raw.count(b"\x00") > max(8, len(raw) // 8):
        candidates = []
        for encoding in ("utf-16-le", "utf-16-be"):
            try:
                value = raw.decode(encoding)
                printable = sum(ch.isprintable() for ch in value)
                candidates.append((printable, value))
            except Exception:
                pass
        if candidates:
            value = max(candidates, key=lambda item: item[0])[1]
            if value.strip():
                return value

    for encoding in (
        "utf-8-sig", "utf-8", "cp932",
        "shift_jis", "cp1252", "latin-1",
    ):
        try:
            value = raw.decode(encoding)
            if value.strip():
                return value
        except Exception:
            pass

    return raw.replace(b"\x00", b"\n").decode(
        "latin-1", errors="replace"
    )


def _f43_review_lines(path):
    value = _f43_decode_review_text(path)
    value = value.replace("\r\r\n", "\n")
    value = value.replace("\r\n", "\n")
    value = value.replace("\r", "\n")
    value = value.replace("\x00", "\n")

    output = []
    for raw_line in value.splitlines():
        line = raw_line.strip().strip("\ufeff")
        if not line:
            continue
        printable = sum(ch.isprintable() for ch in line)
        if printable >= max(1, int(len(line) * 0.45)):
            output.append(line)
    return output


def _f43_review_records(path):
    """Canonical Review parser for Discovery, ZIP cache and Viewer."""
    path = Path(path)
    records = []

    try:
        scans = list(parse_reviewout_file(path) or [])
    except Exception:
        scans = []

    if scans:
        for scan_index, scan in enumerate(scans, 1):
            timestamp = parse_datetime_from_text(
                f"{scan.get('date', '')} {scan.get('time', '')}".strip()
            )
            for parameter, value in scan.items():
                if str(parameter).lower() in {"date", "time"}:
                    continue
                if value in ("", None):
                    continue
                records.append(
                    LogRecord(
                        timestamp=timestamp,
                        source_type="Review",
                        filename=path.name,
                        line_no=scan_index,
                        level="",
                        category=str(parameter),
                        message=str(value),
                        raw=json.dumps(
                            {
                                "scan": scan_index,
                                "parameter": str(parameter),
                                "value": value,
                            },
                            ensure_ascii=False,
                            default=str,
                        ),
                    )
                )
        if records:
            return records

    current_timestamp = parse_filename_datetime(path)

    for line_number, line in enumerate(_f43_review_lines(path), 1):
        parsed_timestamp = parse_datetime_from_text(line)
        if isinstance(parsed_timestamp, datetime):
            current_timestamp = parsed_timestamp

        pair = re.match(
            r"\s*([^:=]{1,120})\s*[:=]\s*(.+?)\s*$",
            line,
        )
        if pair:
            parameter = pair.group(1).strip()
            value = pair.group(2).strip()
        else:
            parameter = "Value"
            value = line.strip()

        if value:
            records.append(
                LogRecord(
                    timestamp=current_timestamp,
                    source_type="Review",
                    filename=path.name,
                    line_no=line_number,
                    level=detect_level(line),
                    category=parameter,
                    message=value,
                    raw=line,
                )
            )
    return records


def _f43_parse_type(path, detected_type):
    path = Path(path)
    canonical = str(detected_type or "").strip().upper()

    if _f43_is_review_file(path) or canonical == "REVIEW":
        return "REVIEW", _f43_review_records(path)

    if canonical == "PSC":
        _, _, records = parse_psc_file_detail(path)
        return "PSC", list(records or [])

    return canonical, list(parse_file(path, canonical) or [])


def _f43_discover_one_file(path, include_unknown):
    path = Path(path)
    detected = (
        "REVIEW"
        if _f43_is_review_file(path)
        else classify_file(path, include_unknown)
    )
    size = path.stat().st_size if path.exists() else 0

    if not detected:
        return DiscoveredFile(str(path), "", 0, None, None, size)

    try:
        canonical, records = _f43_parse_type(path, detected)
        timestamps = [
            record.timestamp for record in records
            if isinstance(getattr(record, "timestamp", None), datetime)
        ]
        return DiscoveredFile(
            str(path),
            "Review" if canonical == "REVIEW" else canonical,
            len(records),
            min(timestamps) if timestamps else None,
            max(timestamps) if timestamps else None,
            size,
        )
    except Exception as error:
        return DiscoveredFile(
            str(path),
            "Review" if str(detected).upper() == "REVIEW" else str(detected),
            0, None, None, size, str(error),
        )


discover_one_file = _f43_discover_one_file


def _f43_parse_selected_files(files):
    cache = {}
    errors = []

    for value in files:
        path = Path(value)
        detected = (
            "REVIEW"
            if _f43_is_review_file(path)
            else classify_file(path, False)
        )
        if not detected:
            continue

        try:
            canonical, records = _f43_parse_type(path, detected)
            if records:
                cache.setdefault(canonical, []).extend(records)
        except Exception as error:
            errors.append(f"{path.name}: {error}")

    for records in cache.values():
        records.sort(
            key=lambda record: (
                getattr(record, "timestamp", None) is None,
                getattr(record, "timestamp", None) or datetime.max,
                str(getattr(record, "filename", "")),
                int(getattr(record, "line_no", 0) or 0),
            )
        )
    return cache, errors


_parse_selected_zip_files_to_memory = _f43_parse_selected_files


def _f43_cache_key(value):
    normalized = str(value or "").strip().upper().replace(" ", "")
    return {
        "REVIEW.OUT": "REVIEW",
        "REVIEWOUT": "REVIEW",
        "REVIEW": "REVIEW",
        "GESYSLOG": "GESYS",
    }.get(normalized, normalized)


def _f43_cached_records(parent, source_name):
    cache = getattr(parent, "zip_import_records_by_type", None)
    if not isinstance(cache, dict):
        return []

    target = _f43_cache_key(source_name)
    for key, records in cache.items():
        if _f43_cache_key(key) == target:
            return list(records or [])
    return []


_f43_previous_source_to_records = MultiPaneLogViewer.source_to_records
def _f43_source_to_records(self, source_name, progress=None):
    parent = getattr(self, "parent_window", None)
    cached = _f43_cached_records(parent, source_name) if parent else []
    if cached:
        return cached
    return _f43_previous_source_to_records(self, source_name, progress)


MultiPaneLogViewer.source_to_records = _f43_source_to_records


def _f43_start_import_clicked(self):
    mode = (
        self.source_mode_combo.currentText()
        if hasattr(self, "source_mode_combo")
        else "Folder"
    )
    source = self.source_edit.text().strip()
    is_zip = mode == "ZIP File" or source.lower().endswith(".zip")

    if not is_zip:
        return _old_commit0017_start(self)

    if not source or not Path(source).is_file():
        QMessageBox.warning(self, APP_TITLE, "Select a valid Source ZIP file.")
        return

    holder = None
    try:
        holder, extract_root, nested = _safe_extract_zip_for_discovery(source)

        if nested:
            QMessageBox.information(
                self,
                "ZIP files inside archive",
                "Nested ZIP files were not expanded:\n\n" + "\n".join(nested),
            )

        options = self.collect_options()
        options.source_folder = str(extract_root)
        options.recursive = True
        options.include_unknown = False

        dialog = SmartFileDiscoveryDialog(self, options)
        if dialog.exec() != QDialog.Accepted:
            self.status_label.setText("ZIP import cancelled")
            return

        selected = list(dialog.selected_files or [])
        if not selected:
            QMessageBox.warning(
                self, APP_TITLE, "No recognized file was selected."
            )
            return

        cache, errors = _f43_parse_selected_files(selected)
        if not cache:
            raise ValueError("Selected files produced no Viewer records.")

        self.zip_import_records_by_type = cache
        self.zip_import_source_name = Path(source).name
        self.viewer_selected_files = []
        self.viewer_selected_types = set(cache)

        self.log_view.clear()
        self.log("Commit0043 Viewer-only ZIP import.")
        for key, records in cache.items():
            self.log(f"{key}: {len(records):,} records")
        for error in errors:
            self.log(f"Parser warning: {error}")

        if self.viewer_window is None:
            self.viewer_window = DualLogViewer(self)

        self.viewer_window.refresh_available_sources()
        self.open_dual_viewer()
        self.status_label.setText(
            f"ZIP loaded for Viewer: {len(selected)} files"
        )
    except Exception:
        QMessageBox.critical(self, APP_TITLE, traceback.format_exc())
    finally:
        if holder is not None:
            holder.cleanup()


MainWindow.start_import_clicked = _f43_start_import_clicked
MainWindow.run_clicked = _c21_disable_merge_workflow


_f43_previous_main_build_ui = MainWindow.build_ui
def _f43_main_build_ui(self):
    _f43_previous_main_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")

    if hasattr(self, "run_btn"):
        try:
            self.run_btn.clicked.disconnect()
        except Exception:
            pass
        self.run_btn.setText("▶ START")
        self.run_btn.clicked.connect(
            lambda checked=False: self.start_import_clicked()
        )

    if hasattr(self, "merge_btn") and self.merge_btn is not None:
        try:
            self.merge_btn.clicked.disconnect()
        except Exception:
            pass
        self.merge_btn.hide()
        self.merge_btn.setEnabled(False)


MainWindow.build_ui = _f43_main_build_ui


_f43_previous_refresh_sources = MultiPaneLogViewer.refresh_available_sources
def _f43_refresh_available_sources(self):
    _f43_previous_refresh_sources(self)

    parent = getattr(self, "parent_window", None)
    cache = getattr(parent, "zip_import_records_by_type", {})
    display_map = {
        "WATERSYSTEM": "WaterSystem",
        "VIMEASURE": "VIMeasure",
        "REVIEW": "Review",
    }
    available = []

    if isinstance(cache, dict) and cache:
        order = [
            "WS", "WATERSYSTEM", "VIMEASURE", "ACQUISITION",
            "CGA", "CSA", "MRSERVER", "GESYS", "LAIS", "PSC", "REVIEW",
        ]
        for key in order:
            if cache.get(key):
                available.append(display_map.get(key, key))
        for key, records in cache.items():
            if records:
                display = display_map.get(str(key).upper(), str(key))
                if display not in available:
                    available.append(display)

    if not available:
        for combo in getattr(self, "sources", []):
            for index in range(combo.count()):
                value = combo.itemText(index)
                if value.upper() not in {"MERGED", "CUSTOM FILE"}:
                    if value not in available:
                        available.append(value)

    for pane_index, combo in enumerate(getattr(self, "sources", [])):
        current = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        combo.addItems(available)
        if current in available:
            combo.setCurrentText(current)
        elif available:
            combo.setCurrentIndex(min(pane_index, len(available) - 1))
        combo.blockSignals(False)


MultiPaneLogViewer.refresh_available_sources = _f43_refresh_available_sources


_f43_previous_viewer_init = MultiPaneLogViewer.__init__
def _f43_viewer_init(self, parent_window):
    _f43_previous_viewer_init(self, parent_window)

    for index, check in enumerate(
        list(getattr(self, "pane_visible_checks", []) or [])
    ):
        check.blockSignals(True)
        check.setChecked(index < 2)
        check.blockSignals(False)
        check.setVisible(True)
        check.setEnabled(True)

    if hasattr(self, "main_splitter"):
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setHandleWidth(7)

    self.update_view_mode()


MultiPaneLogViewer.__init__ = _f43_viewer_init



# ---------------------------------------------------------------------------
# Commit0044: Log Explore with Event Viewer and Value Viewer
# ---------------------------------------------------------------------------

_F44_VALUE_TYPES = {"REVIEW", "PSC", "VIMEASURE", "ACQUISITION", "WATERSYSTEM"}


def _f44_type(value):
    normalized = str(value or "").strip().upper().replace(" ", "")
    return {
        "REVIEW.OUT": "REVIEW",
        "REVIEWOUT": "REVIEW",
        "VI MEASURE": "VIMEASURE",
    }.get(normalized, normalized)


def _f44_value_rows(source_type, records):
    """One protocol/scan/measurement becomes one horizontal row."""
    canonical = _f44_type(source_type)
    rows = []

    if canonical == "REVIEW":
        grouped = {}
        for record in records:
            key = (
                getattr(record, "filename", ""),
                getattr(record, "line_no", 0),
                getattr(record, "timestamp", None),
            )
            row = grouped.setdefault(
                key,
                {
                    "Timestamp": getattr(record, "timestamp", None),
                    "SourceType": "Review",
                    "File": getattr(record, "filename", ""),
                    "ProtocolIndex": getattr(record, "line_no", 0),
                },
            )
            parameter = str(getattr(record, "category", "") or "Value").strip()
            value = getattr(record, "message", "")
            if parameter in row:
                suffix = 2
                candidate = f"{parameter}_{suffix}"
                while candidate in row:
                    suffix += 1
                    candidate = f"{parameter}_{suffix}"
                parameter = candidate
            row[parameter] = value
        return list(grouped.values())

    for record in records:
        row = {
            "Timestamp": getattr(record, "timestamp", None),
            "SourceType": canonical,
            "File": getattr(record, "filename", ""),
            "Line": getattr(record, "line_no", 0),
        }
        raw = getattr(record, "raw", "")
        if raw:
            try:
                data = json.loads(raw)
                if isinstance(data, dict):
                    for key, value in data.items():
                        if not str(key).startswith("_"):
                            row[str(key)] = value
            except Exception:
                pass
        if getattr(record, "category", ""):
            row.setdefault("Parameter", record.category)
        if getattr(record, "message", ""):
            row.setdefault("Value", record.message)
        rows.append(row)
    return rows


class ValueTableModel(QAbstractTableModel):
    def __init__(self, rows=None, parent=None):
        super().__init__(parent)
        self.rows = list(rows or [])
        self.filtered_rows = list(self.rows)
        self.columns = self._columns()

    def _columns(self):
        preferred = [
            "Timestamp", "Protocol", "ProtocolName", "Sequence", "Series",
            "SourceType", "File", "Line", "ProtocolIndex",
        ]
        columns = [
            key for key in preferred
            if any(key in row for row in self.rows)
        ]
        for row in self.rows:
            for key in row:
                if key not in columns and not str(key).startswith("_"):
                    columns.append(key)
        return columns

    def rowCount(self, parent=QModelIndex()):
        return len(self.filtered_rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or role not in {Qt.DisplayRole, Qt.ToolTipRole}:
            return None
        value = self.filtered_rows[index.row()].get(self.columns[index.column()], "")
        if isinstance(value, datetime):
            return value.strftime("%Y/%m/%d %H:%M:%S.%f")[:-3]
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return str(value)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        return self.columns[section] if orientation == Qt.Horizontal else section + 1

    def sort(self, column, order):
        if not 0 <= column < len(self.columns):
            return
        key = self.columns[column]
        self.layoutAboutToBeChanged.emit()
        self.filtered_rows.sort(
            key=lambda row: str(row.get(key, "")).casefold(),
            reverse=order == Qt.DescendingOrder,
        )
        self.layoutChanged.emit()

    def apply_filter(self, text):
        query = str(text or "").strip().casefold()
        self.beginResetModel()
        self.filtered_rows = (
            list(self.rows)
            if not query
            else [
                row for row in self.rows
                if query in " ".join(str(value) for value in row.values()).casefold()
            ]
        )
        self.endResetModel()


class ValueViewerWidget(QWidget):
    def __init__(self, parent_window):
        super().__init__()
        self.parent_window = parent_window
        self.model = ValueTableModel([])
        self.current_type = ""

        root = QVBoxLayout(self)
        controls = QHBoxLayout()
        controls.addWidget(QLabel("Value Type:"))
        self.source_combo = QComboBox()
        self.source_combo.currentTextChanged.connect(self.load_source)
        controls.addWidget(self.source_combo)

        controls.addWidget(QLabel("Filter:"))
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText("Search protocol, parameter or value")
        self.filter_edit.textChanged.connect(self.model.apply_filter)
        controls.addWidget(self.filter_edit, 1)

        self.export_button = QPushButton("Export CSV")
        self.export_button.clicked.connect(self.export_csv)
        controls.addWidget(self.export_button)
        root.addLayout(controls)

        self.status = QLabel("No value data loaded.")
        root.addWidget(self.status)

        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.horizontalHeader().setSectionsMovable(True)
        root.addWidget(self.table, 1)

    def refresh_sources(self):
        cache = getattr(self.parent_window, "zip_import_records_by_type", {})
        available = []
        for key, records in cache.items():
            canonical = _f44_type(key)
            if canonical in _F44_VALUE_TYPES and records:
                label = "Review" if canonical == "REVIEW" else canonical
                if label not in available:
                    available.append(label)

        current = self.source_combo.currentText()
        self.source_combo.blockSignals(True)
        self.source_combo.clear()
        self.source_combo.addItems(available)
        if current in available:
            self.source_combo.setCurrentText(current)
        elif available:
            self.source_combo.setCurrentIndex(0)
        self.source_combo.blockSignals(False)

        if available:
            self.load_source(self.source_combo.currentText())
        else:
            self.model = ValueTableModel([])
            self.table.setModel(self.model)
            self.status.setText("No value-type records are available.")

    def load_source(self, source_name):
        canonical = _f44_type(source_name)
        cache = getattr(self.parent_window, "zip_import_records_by_type", {})
        records = []
        for key, values in cache.items():
            if _f44_type(key) == canonical:
                records.extend(list(values or []))
        rows = _f44_value_rows(canonical, records)
        self.current_type = canonical
        self.model = ValueTableModel(rows)
        self.table.setModel(self.model)
        self.filter_edit.textChanged.disconnect()
        self.filter_edit.textChanged.connect(self.model.apply_filter)
        self.table.resizeColumnsToContents()
        self.status.setText(
            f"{source_name}: {len(rows):,} protocol/measurement rows, "
            f"{len(self.model.columns):,} columns"
        )

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Value Viewer",
            f"{self.current_type or 'values'}.csv",
            "CSV Files (*.csv)",
        )
        if not path:
            return
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as stream:
            writer = csv.writer(stream)
            writer.writerow(self.model.columns)
            for row in self.model.filtered_rows:
                writer.writerow([row.get(column, "") for column in self.model.columns])


class LogExploreWindow(QMainWindow):
    def __init__(self, parent_window):
        super().__init__(parent_window)
        self.parent_window = parent_window
        self.setWindowTitle(f"Log Explore — {APP_VERSION}")
        self.resize(1500, 900)

        central = QWidget()
        root = QVBoxLayout(central)
        self.setCentralWidget(central)

        title_row = QHBoxLayout()
        title = QLabel("Log Explore")
        title.setStyleSheet("font-size:20px;font-weight:700;color:#005DAA;")
        title_row.addWidget(title)
        title_row.addStretch(1)
        refresh = QPushButton("Refresh Data")
        refresh.clicked.connect(self.refresh_data)
        title_row.addWidget(refresh)
        root.addLayout(title_row)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self.event_viewer = DualLogViewer(parent_window)
        self.tabs.addTab(self.event_viewer, "Event Viewer")

        self.value_viewer = ValueViewerWidget(parent_window)
        self.tabs.addTab(self.value_viewer, "Value Viewer")

        # Commit0049: advanced workspaces are lazy-loaded.
        # Startup must show Event Viewer immediately without constructing
        # Investigation or Spectrum on the GUI thread.
        self.investigation = None
        self.spectrum = None

        self._investigation_placeholder = self._make_lazy_placeholder(
            "Investigation",
            "Select this tab to initialize Investigation. "
            "Analysis starts only when Start Analysis is pressed.",
        )
        self.tabs.addTab(
            self._investigation_placeholder,
            "Investigation",
        )

        self._spectrum_placeholder = self._make_lazy_placeholder(
            "Spectrum",
            "Select this tab to initialize Spectrum Analysis. "
            "Spectrum scanning starts only when Scan Spectrum Dumps is pressed.",
        )
        self.tabs.addTab(
            self._spectrum_placeholder,
            "Spectrum",
        )

        self.tabs.currentChanged.connect(
            self._on_log_explore_tab_changed
        )

    def _make_lazy_placeholder(self, title, message):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.addStretch(1)

        heading = QLabel(title)
        heading.setAlignment(Qt.AlignCenter)
        heading.setStyleSheet(
            "font-size:20px;font-weight:700;color:#005DAA;"
        )
        layout.addWidget(heading)

        description = QLabel(message)
        description.setAlignment(Qt.AlignCenter)
        description.setWordWrap(True)
        description.setStyleSheet(
            "color:#475569;padding:12px;"
        )
        layout.addWidget(description)

        busy = QProgressBar()
        busy.setRange(0, 0)
        busy.setVisible(False)
        busy.setMaximumWidth(420)
        layout.addWidget(busy, 0, Qt.AlignHCenter)
        page.setProperty("lazy_busy_bar", busy)

        layout.addStretch(1)
        return page

    def _tab_index_by_name(self, name):
        target = str(name).strip().casefold()
        for index in range(self.tabs.count()):
            if self.tabs.tabText(index).strip().casefold() == target:
                return index
        return -1

    def _replace_named_tab(self, name, widget):
        index = self._tab_index_by_name(name)
        if index < 0:
            self.tabs.addTab(widget, name)
            return self.tabs.count() - 1

        current = self.tabs.currentIndex()
        self.tabs.removeTab(index)
        self.tabs.insertTab(index, widget, name)

        if current == index:
            self.tabs.setCurrentIndex(index)

        return index

    def _initialize_investigation(self):
        if self.investigation is not None:
            return self.investigation

        placeholder = getattr(
            self,
            "_investigation_placeholder",
            None,
        )
        busy = (
            placeholder.property("lazy_busy_bar")
            if placeholder is not None
            else None
        )
        if busy is not None:
            busy.setVisible(True)

        QApplication.processEvents()

        try:
            self.investigation = InvestigationWorkspace(
                self.event_viewer
            )
            self._replace_named_tab(
                "Investigation",
                self.investigation,
            )
            return self.investigation
        finally:
            if busy is not None:
                busy.setVisible(False)

    def _initialize_spectrum(self):
        if self.spectrum is not None:
            return self.spectrum

        investigation = self._initialize_investigation()

        placeholder = getattr(
            self,
            "_spectrum_placeholder",
            None,
        )
        busy = (
            placeholder.property("lazy_busy_bar")
            if placeholder is not None
            else None
        )
        if busy is not None:
            busy.setVisible(True)

        QApplication.processEvents()

        try:
            self.spectrum = SpectrumAnalysisWidget(
                investigation,
                standalone=False,
            )
            self._replace_named_tab(
                "Spectrum",
                self.spectrum,
            )
            return self.spectrum
        finally:
            if busy is not None:
                busy.setVisible(False)

    def _on_log_explore_tab_changed(self, index):
        if index < 0:
            return

        name = self.tabs.tabText(index).strip()

        # Defer construction until the tab switch has painted.
        if name == "Investigation" and self.investigation is None:
            QTimer.singleShot(
                0,
                self._initialize_investigation,
            )
        elif name == "Spectrum" and self.spectrum is None:
            QTimer.singleShot(
                0,
                self._initialize_spectrum,
            )

    def refresh_data(self):
        # Lightweight refresh only. Advanced analysis is never started here.
        self.event_viewer.refresh_available_sources()
        self.value_viewer.refresh_sources()

    def showEvent(self, event):
        super().showEvent(event)

        if getattr(self, "_first_show_refresh_done", False):
            return

        self._first_show_refresh_done = True
        QTimer.singleShot(0, self.refresh_data)


_f44_previous_init = MainWindow.__init__
def _f44_main_init(self, *args, **kwargs):
    _f44_previous_init(self, *args, **kwargs)
    self.log_explore_window = None


MainWindow.__init__ = _f44_main_init


_f44_previous_start = MainWindow.start_import_clicked
def _f44_start_import_clicked(self):
    result = _f44_previous_start(self)
    cache = getattr(self, "zip_import_records_by_type", {})
    if cache:
        self.open_log_explore()
    return result


MainWindow.start_import_clicked = _f44_start_import_clicked


_f44_previous_build_ui = MainWindow.build_ui
def _f44_build_ui(self):
    _f44_previous_build_ui(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")

    if hasattr(self, "run_btn"):
        try:
            self.run_btn.clicked.disconnect()
        except Exception:
            pass
        self.run_btn.setText("▶ START LOG EXPLORE")
        self.run_btn.setToolTip("Run Smart File Discovery and open Log Explore.")
        self.run_btn.clicked.connect(
            lambda checked=False: self.start_import_clicked()
        )

    for button in self.findChildren(QPushButton):
        if button.text().strip() in {
            "Open Viewer", "Log Viewer", "Open Log Viewer", "Viewer"
        }:
            button.setText("Log Explore")
            try:
                button.clicked.disconnect()
            except Exception:
                pass
            button.clicked.connect(
                lambda checked=False: self.open_log_explore()
            )


MainWindow.build_ui = _f44_build_ui



# ---------------------------------------------------------------------------
# Commit0045: Operation Intelligence adapted from FUS Investigation Platform
# ---------------------------------------------------------------------------

@dataclass
class OperationExploreEvent:
    timestamp: Optional[datetime] = None
    seconds: float = 0.0
    time_text: str = ""
    category: str = "Software"
    actor: str = "Software"
    severity: str = "Info"
    message: str = ""
    source: str = ""
    sonication: int = -1
    phase: str = "Unlinked"
    confidence: str = "Low"
    correlation_id: str = ""
    record: Any = None


_F45_OPERATION_RULES = [
    ("Error", "System", "Error", (
        "error", "exception", "failed", "failure", "fatal",
        "abort", "aborted", "crash", "invalid",
    )),
    ("Warning", "System", "Warning", (
        "warning", "warn", "timeout", "retry", "motion detected",
        "not ready", "interlock", "out of range",
    )),
    ("Treatment", "Operator", "Info", (
        "treat high", "treat low", "start sonication",
        "stop sonication", "continue treatment", "pause treatment",
        "sonication start", "sonication stop", "sonicate",
    )),
    ("Tracking", "Operator", "Info", (
        "tracking", "detect patient movement", "position detect",
        "transducer position", "xd position", "motion detection",
    )),
    ("Thermometry", "Software", "Info", (
        "tmap", "memp", "thermometry", "temperature map",
        "phase image", "temperature",
    )),
    ("MRI", "Software", "Info", (
        "loadprotocol", "load protocol", "scan start",
        "scanner=scanning", "acquisition=complete", "recon=done",
        "setrxgeometry", "series type", "reconstruction",
    )),
    ("Registration", "Operator", "Info", (
        "fusion", "registration", "ctmr", "mrmr",
        "fiducial", "align",
    )),
    ("Planning", "Operator", "Info", (
        "sdr", "calcification", "non-path", "npr",
        "target", "planning", "preplan",
    )),
    ("Acoustic", "Software", "Info", (
        "spectrum", "hydrophone", "cavitation",
        "reflection", "acoustic control",
    )),
    ("Cooling", "Software", "Info", (
        "cooling", "cool down", "post sonication", "chiller",
    )),
    ("Data", "Operator", "Info", (
        "save", "export", "import", "retrieve",
        "database update", "load patient", "open patient",
    )),
]

_F45_OPERATOR_HINTS = (
    "clicked", "button", "selected", "user", "operator",
    "menu", "pressed", "double click", "dragged", "entered",
)

_F45_MEANINGFUL_HINTS = (
    "state", "workflow", "command", "request", "response",
    "scan", "protocol", "patient", "database", "start", "stop",
    "load", "open", "close", "update", "set ", "change",
)

_F45_ID_PATTERN = re.compile(
    r"\b(?:correlation|call|request|transaction|trace|workflow|"
    r"acquisition|sonication|command|cmd|session|task)"
    r"[\s_-]*id\s*(?:=|:|<|\[)?\s*([A-Za-z0-9_.:-]+)",
    re.I,
)


def _f45_record_text(record):
    parts = [
        str(getattr(record, "category", "") or ""),
        str(getattr(record, "message", "") or ""),
        str(getattr(record, "raw", "") or ""),
    ]
    return " ".join(part for part in parts if part).strip()


def _f45_seconds(timestamp):
    if not isinstance(timestamp, datetime):
        return 0.0
    return (
        timestamp.hour * 3600
        + timestamp.minute * 60
        + timestamp.second
        + timestamp.microsecond / 1_000_000.0
    )


def _f45_categorize(message):
    low = str(message or "").lower()

    for category, actor, severity, keys in _F45_OPERATION_RULES:
        if any(key in low for key in keys):
            if any(hint in low for hint in _F45_OPERATOR_HINTS):
                actor = "Operator"
            return category, actor, severity

    if any(hint in low for hint in _F45_OPERATOR_HINTS):
        return "Software", "Operator", "Info"

    return "Software", "Software", "Info"


def _f45_phase(message):
    low = str(message or "").lower()
    if "dqa" in low or "phantom" in low:
        return "DQA"
    if any(key in low for key in ("preplan", "planning", "target")):
        return "Planning"
    if any(key in low for key in ("treat", "sonication", "thermometry")):
        return "Treatment"
    if any(key in low for key in ("cool", "post sonication")):
        return "Cooling"
    return "Workflow"


def _f45_operation_events(cache):
    all_records = []
    if isinstance(cache, dict):
        for source_type, records in cache.items():
            canonical = _f44_type(source_type)
            for record in records or []:
                all_records.append((canonical, record))

    sonication_markers = []
    for source_type, record in all_records:
        text = _f45_record_text(record).lower()
        timestamp = getattr(record, "timestamp", None)
        if isinstance(timestamp, datetime) and (
            "sonication start" in text
            or "start sonication" in text
            or "treat high" in text
            or "treat low" in text
        ):
            sonication_markers.append(timestamp)

    sonication_markers.sort()
    unique_markers = []
    for timestamp in sonication_markers:
        if (
            not unique_markers
            or abs((timestamp - unique_markers[-1]).total_seconds()) > 1.0
        ):
            unique_markers.append(timestamp)

    events = []
    for source_type, record in all_records:
        timestamp = getattr(record, "timestamp", None)
        if not isinstance(timestamp, datetime):
            continue

        message = _f45_record_text(record)
        if not message:
            continue

        category, actor, severity = _f45_categorize(message)
        low = message.lower()
        meaningful = (
            category != "Software"
            or any(key in low for key in _F45_MEANINGFUL_HINTS)
        )
        if not meaningful:
            continue

        nearest_number = -1
        nearest_gap = float("inf")
        for number, marker in enumerate(unique_markers, 1):
            gap = abs((timestamp - marker).total_seconds())
            if gap < nearest_gap:
                nearest_number = number
                nearest_gap = gap

        if nearest_gap > 180:
            nearest_number = -1

        confidence = (
            "High" if nearest_gap <= 30
            else "Medium" if nearest_gap <= 180
            else "Low"
        )

        id_match = _F45_ID_PATTERN.search(message)
        correlation_id = id_match.group(1) if id_match else ""

        events.append(
            OperationExploreEvent(
                timestamp=timestamp,
                seconds=_f45_seconds(timestamp),
                time_text=timestamp.strftime("%H:%M:%S.%f")[:-3],
                category=category,
                actor=actor,
                severity=severity,
                message=message,
                source=source_type,
                sonication=nearest_number,
                phase=_f45_phase(message),
                confidence=confidence,
                correlation_id=correlation_id,
                record=record,
            )
        )

    events.sort(
        key=lambda event: (
            event.timestamp,
            event.source,
            getattr(event.record, "line_no", 0),
        )
    )
    return events


class OperationTimelineWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.events = []
        self.setMinimumHeight(230)
        self.setMouseTracking(True)

    def set_events(self, events):
        self.events = list(events or [])
        self.update()

    def paintEvent(self, _event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor("#F8FAFC"))

        left, top, right, bottom = 125, 22, 25, 42
        plot_width = max(1, self.width() - left - right)
        plot_height = max(1, self.height() - top - bottom)

        painter.setPen(QPen(QColor("#CBD5E1"), 1))
        painter.drawRect(left, top, plot_width, plot_height)

        if not self.events:
            painter.setPen(QColor("#64748B"))
            painter.drawText(
                self.rect(),
                Qt.AlignCenter,
                "No operation events for the current filter",
            )
            return

        categories = []
        for event in self.events:
            if event.category not in categories:
                categories.append(event.category)

        start = min(event.seconds for event in self.events)
        end = max(event.seconds for event in self.events)
        span = max(1.0, end - start)
        row_height = plot_height / max(1, len(categories))

        color_map = {
            "Error": QColor("#B91C1C"),
            "Warning": QColor("#D97706"),
            "Treatment": QColor("#7C3AED"),
            "MRI": QColor("#2563EB"),
            "Thermometry": QColor("#DB2777"),
            "Acoustic": QColor("#0891B2"),
            "Tracking": QColor("#16A34A"),
            "Planning": QColor("#4F46E5"),
            "Registration": QColor("#0F766E"),
            "Cooling": QColor("#0284C7"),
            "Data": QColor("#64748B"),
            "Software": QColor("#475569"),
        }

        painter.setPen(QColor("#334155"))
        for row, category in enumerate(categories):
            y = top + (row + 0.5) * row_height
            painter.drawText(5, int(y + 4), category)
            painter.setPen(QPen(QColor("#E2E8F0"), 1))
            painter.drawLine(left, int(y), left + plot_width, int(y))
            painter.setPen(QColor("#334155"))

        for event in self.events:
            row = categories.index(event.category)
            x = left + (event.seconds - start) / span * plot_width
            y = top + (row + 0.5) * row_height
            painter.setPen(QPen(color_map.get(event.category, QColor("#475569")), 2))
            if event.actor == "Operator":
                painter.setBrush(QBrush(color_map.get(event.category, QColor("#475569"))))
                painter.drawEllipse(int(x - 4), int(y - 4), 8, 8)
            else:
                painter.drawLine(int(x), int(y - 7), int(x), int(y + 7))

        painter.setPen(QColor("#475569"))
        for tick in range(6):
            ratio = tick / 5
            seconds = start + span * ratio
            x = left + plot_width * ratio
            h = int(seconds // 3600) % 24
            m = int((seconds % 3600) // 60)
            sec = int(seconds % 60)
            painter.drawText(int(x - 28), self.height() - 14, f"{h:02d}:{m:02d}:{sec:02d}")


class OperationIntelligenceWidget(QWidget):
    def __init__(self, parent_window, event_viewer=None):
        super().__init__()
        self.parent_window = parent_window
        self.event_viewer = event_viewer
        self.all_events = []
        self.visible_events = []

        root = QVBoxLayout(self)
        controls = QHBoxLayout()
        controls.addWidget(QLabel("Category"))

        self.category = QComboBox()
        self.category.addItems([
            "All", "Operator", "Software", "Planning",
            "Registration", "Tracking", "MRI", "Treatment",
            "Thermometry", "Acoustic", "Cooling", "Data",
            "Warning", "Error",
        ])
        self.category.currentTextChanged.connect(self.refresh_view)
        controls.addWidget(self.category)

        controls.addWidget(QLabel("Search"))
        self.search = QLineEdit()
        self.search.setPlaceholderText("Operation, source, CallID / CorrelationID")
        self.search.textChanged.connect(self.refresh_view)
        controls.addWidget(self.search, 1)

        self.show_dqa = QCheckBox("Show DQA")
        self.show_dqa.setChecked(False)
        self.show_dqa.toggled.connect(self.refresh_view)
        controls.addWidget(self.show_dqa)

        reset = QPushButton("Reset")
        reset.clicked.connect(self.reset_filters)
        controls.addWidget(reset)

        rebuild = QPushButton("Rebuild Operation Analysis")
        rebuild.clicked.connect(self.rebuild)
        controls.addWidget(rebuild)
        root.addLayout(controls)

        self.summary = QLabel("No operation data")
        self.summary.setWordWrap(True)
        self.summary.setStyleSheet(
            "background:#F1F5F9;border:1px solid #CBD5E1;"
            "border-radius:4px;padding:7px;"
        )
        root.addWidget(self.summary)

        self.timeline = OperationTimelineWidget()
        root.addWidget(self.timeline, 1)

        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "Time", "Category", "Actor", "Severity", "Sonication",
            "Phase", "Confidence", "CorrelationID", "Source",
            "Operation / software event",
        ])
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(
            9, QHeaderView.Stretch
        )
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.cellDoubleClicked.connect(self.activate_row)
        root.addWidget(self.table, 2)

    def rebuild(self):
        cache = getattr(
            self.parent_window,
            "zip_import_records_by_type",
            {},
        )
        self.all_events = _f45_operation_events(cache)
        self.refresh_view()

    def reset_filters(self):
        self.category.setCurrentText("All")
        self.search.clear()
        self.show_dqa.setChecked(False)
        self.refresh_view()

    def _filter_events(self):
        events = list(self.all_events)

        if not self.show_dqa.isChecked():
            events = [event for event in events if event.phase != "DQA"]

        selected = self.category.currentText()
        if selected != "All":
            if selected in {"Operator", "Software"}:
                events = [
                    event for event in events
                    if event.actor == selected
                ]
            elif selected in {"Warning", "Error"}:
                events = [
                    event for event in events
                    if event.severity == selected
                ]
            else:
                events = [
                    event for event in events
                    if event.category == selected
                ]

        query = self.search.text().strip().casefold()
        if query:
            events = [
                event for event in events
                if query in " ".join([
                    event.message,
                    event.source,
                    event.correlation_id,
                    event.category,
                    event.actor,
                ]).casefold()
            ]

        return events

    def refresh_view(self, *_args):
        events = self._filter_events()
        self.visible_events = events
        self.timeline.set_events(events)
        self.table.setRowCount(len(events))

        for row, event in enumerate(events):
            values = [
                event.time_text,
                event.category,
                event.actor,
                event.severity,
                event.sonication if event.sonication >= 0 else "—",
                event.phase,
                event.confidence,
                event.correlation_id,
                event.source,
                event.message,
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setData(Qt.UserRole, row)
                if event.severity == "Error":
                    item.setBackground(QColor("#FEE2E2"))
                elif event.severity == "Warning":
                    item.setBackground(QColor("#FEF3C7"))
                elif event.actor == "Operator":
                    item.setBackground(QColor("#E0F2FE"))
                self.table.setItem(row, column, item)

        counts = {}
        actor_counts = {"Operator": 0, "Software": 0, "System": 0}
        linked = 0
        for event in events:
            counts[event.category] = counts.get(event.category, 0) + 1
            actor_counts[event.actor] = actor_counts.get(event.actor, 0) + 1
            if event.sonication >= 0:
                linked += 1

        if events:
            category_text = ", ".join(
                f"{key}: {value}" for key, value in sorted(counts.items())
            )
            self.summary.setText(
                f"{len(events):,} operation events | "
                f"Operator: {actor_counts.get('Operator', 0):,} | "
                f"Software: {actor_counts.get('Software', 0):,} | "
                f"System: {actor_counts.get('System', 0):,} | "
                f"Sonication linked: {linked:,} | {category_text}"
            )
        else:
            self.summary.setText(
                "No operation events for the current filter."
            )

    def activate_row(self, row, _column):
        if not 0 <= row < len(self.visible_events):
            return

        event = self.visible_events[row]
        if self.event_viewer is None:
            return

        # Select the corresponding source in the first pane and load it.
        try:
            combo = self.event_viewer.sources[0]
            display = "Review" if event.source == "REVIEW" else event.source
            index = combo.findText(display)
            if index >= 0:
                combo.setCurrentIndex(index)
                self.event_viewer.load_pane("left")
        except Exception:
            pass


# Attach Operation tab to Log Explore without replacing existing tabs.
_f46_base_log_explore_init = LogExploreWindow.__init__
def _f46_log_explore_init(self, parent_window):
    _f46_base_log_explore_init(self, parent_window)
    operation = getattr(self, "operation", None)
    if operation is None:
        operation = OperationIntelligenceWidget(
            parent_window,
            getattr(self, "event_viewer", None),
        )
        self.operation = operation
    if self.tabs.indexOf(operation) < 0:
        self.tabs.insertTab(min(2, self.tabs.count()), operation, "Operation")


LogExploreWindow.__init__ = _f46_log_explore_init


_f46_base_refresh_data = LogExploreWindow.refresh_data
def _f46_refresh_data(self):
    # Commit0049: lightweight data-source refresh only.
    # Operation analysis runs only from its Rebuild Operation Analysis button.
    _f46_base_refresh_data(self)


LogExploreWindow.refresh_data = _f46_refresh_data



# ---------------------------------------------------------------------------
# Commit0045b startup contract
# ---------------------------------------------------------------------------

def _f45b_startup_contract_check():
    required = {
        "QAbstractItemView": QAbstractItemView,
        "LogExploreWindow": LogExploreWindow,
        "ValueViewerWidget": ValueViewerWidget,
        "OperationIntelligenceWidget": OperationIntelligenceWidget,
        "OperationTimelineWidget": OperationTimelineWidget,
    }

    missing = [name for name, value in required.items() if value is None]
    if missing:
        raise RuntimeError(
            "Missing startup symbols: " + ", ".join(missing)
        )

    return True


_F45B_STARTUP_CONTRACT_OK = _f45b_startup_contract_check()



# Commit0046 canonical Log Explore open path.
def _f46_open_log_explore(self):
    window = getattr(self, "log_explore_window", None)
    if window is None:
        window = LogExploreWindow(self)
        self.log_explore_window = window
    window.show()
    window.raise_()
    window.activateWindow()
    QTimer.singleShot(0, window.refresh_data)


MainWindow.open_log_explore = _f46_open_log_explore


# ---------------------------------------------------------------------------
# Commit0047: canonical Log Explore public navigation API
# ---------------------------------------------------------------------------

def _f47_select_log_explore_tab(self, tab_name):
    self.open_log_explore()

    window = getattr(self, "log_explore_window", None)
    if window is None:
        return

    target = str(tab_name or "").strip().casefold()

    for index in range(window.tabs.count()):
        if window.tabs.tabText(index).strip().casefold() == target:
            window.tabs.setCurrentIndex(index)
            return


def _f47_open_investigation_workspace(self):
    _f47_select_log_explore_tab(self, "Investigation")


def _f47_open_standalone_spectrum_analysis(self):
    # Kept as a compatibility API, but the canonical destination is the
    # embedded Spectrum tab in Log Explore.
    _f47_select_log_explore_tab(self, "Spectrum")


MainWindow.open_investigation_workspace = (
    _f47_open_investigation_workspace
)
MainWindow.open_standalone_spectrum_analysis = (
    _f47_open_standalone_spectrum_analysis
)



# ---------------------------------------------------------------------------
# Commit0048: Embedded workspace class contract
# ---------------------------------------------------------------------------

def _f48_embedded_workspace_contract():
    required = {
        "InvestigationWorkspace": InvestigationWorkspace,
        "SpectrumAnalysisWidget": SpectrumAnalysisWidget,
        "LogExploreWindow": LogExploreWindow,
        "ValueViewerWidget": ValueViewerWidget,
        "OperationIntelligenceWidget": OperationIntelligenceWidget,
    }

    missing = [
        name for name, value in required.items()
        if value is None
    ]

    if missing:
        raise RuntimeError(
            "Missing embedded workspace classes: "
            + ", ".join(missing)
        )

    return True


_F48_EMBEDDED_WORKSPACE_CONTRACT_OK = (
    _f48_embedded_workspace_contract()
)



# ---------------------------------------------------------------------------
# Commit0050: Two-phase Log Explore bootstrap
#
# The old Event Viewer constructor can inspect selected files/cache and perform
# expensive work before Log Explore is painted. This replacement displays the
# shell first, then constructs each workspace with parent data temporarily
# hidden, reconnecting the parsed cache only after construction.
# ---------------------------------------------------------------------------

class FastLogExploreWindow(QMainWindow):
    def __init__(self, parent_window):
        super().__init__(parent_window)
        self.parent_window = parent_window
        self.setWindowTitle(f"Log Explore — {APP_VERSION}")
        self.resize(1500, 900)

        self.event_viewer = None
        self.value_viewer = None
        self.operation = None
        self.investigation = None
        self.spectrum = None

        self._bootstrap_started = False
        self._bootstrap_complete = False
        self._bootstrap_failed = False

        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(4, 4, 4, 4)
        self.setCentralWidget(central)

        header = QHBoxLayout()
        title = QLabel("Log Explore")
        title.setStyleSheet(
            "font-size:20px;font-weight:700;color:#005DAA;"
        )
        header.addWidget(title)

        self.bootstrap_status = QLabel("Opening Log Explore…")
        self.bootstrap_status.setStyleSheet("color:#475569;")
        header.addWidget(self.bootstrap_status, 1)

        self.refresh_button = QPushButton("Refresh Data")
        self.refresh_button.setEnabled(False)
        self.refresh_button.clicked.connect(self.refresh_data)
        header.addWidget(self.refresh_button)
        root.addLayout(header)

        self.bootstrap_progress = QProgressBar()
        self.bootstrap_progress.setRange(0, 4)
        self.bootstrap_progress.setValue(0)
        root.addWidget(self.bootstrap_progress)

        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(
            self._on_tab_changed
        )
        root.addWidget(self.tabs, 1)

        self._event_placeholder = self._placeholder(
            "Event Viewer",
            "Preparing the Event Viewer without loading records…",
        )
        self.tabs.addTab(
            self._event_placeholder,
            "Event Viewer",
        )

        self._value_placeholder = self._placeholder(
            "Value Viewer",
            "Value Viewer will be prepared after Event Viewer.",
        )
        self.tabs.addTab(
            self._value_placeholder,
            "Value Viewer",
        )

        self._operation_placeholder = self._placeholder(
            "Operation",
            "Select this tab to initialize Operation Intelligence.",
        )
        self.tabs.addTab(
            self._operation_placeholder,
            "Operation",
        )

        self._investigation_placeholder = self._placeholder(
            "Investigation",
            "Select this tab to initialize Investigation.",
        )
        self.tabs.addTab(
            self._investigation_placeholder,
            "Investigation",
        )

        self._spectrum_placeholder = self._placeholder(
            "Spectrum",
            "Select this tab to initialize Spectrum Analysis.",
        )
        self.tabs.addTab(
            self._spectrum_placeholder,
            "Spectrum",
        )

    def _placeholder(self, title, message):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.addStretch(1)

        heading = QLabel(title)
        heading.setAlignment(Qt.AlignCenter)
        heading.setStyleSheet(
            "font-size:22px;font-weight:700;color:#005DAA;"
        )
        layout.addWidget(heading)

        label = QLabel(message)
        label.setAlignment(Qt.AlignCenter)
        label.setWordWrap(True)
        label.setStyleSheet("color:#475569;padding:10px;")
        layout.addWidget(label)

        busy = QProgressBar()
        busy.setRange(0, 0)
        busy.setMaximumWidth(420)
        layout.addWidget(busy, 0, Qt.AlignHCenter)

        layout.addStretch(1)
        return page

    def showEvent(self, event):
        super().showEvent(event)

        if self._bootstrap_started:
            return

        self._bootstrap_started = True
        # Allow the empty shell to paint before any legacy widget constructor.
        QTimer.singleShot(100, self._bootstrap_event_viewer)

    def _replace_tab(self, name, widget):
        target = name.casefold()
        for index in range(self.tabs.count()):
            if self.tabs.tabText(index).casefold() == target:
                selected = self.tabs.currentIndex() == index
                self.tabs.removeTab(index)
                self.tabs.insertTab(index, widget, name)
                if selected:
                    self.tabs.setCurrentIndex(index)
                return index

        self.tabs.addTab(widget, name)
        return self.tabs.count() - 1

    def _temporarily_hide_parent_data(self):
        parent = self.parent_window
        saved = {
            "zip_import_records_by_type": getattr(
                parent, "zip_import_records_by_type", None
            ),
            "viewer_selected_files": getattr(
                parent, "viewer_selected_files", None
            ),
            "viewer_selected_types": getattr(
                parent, "viewer_selected_types", None
            ),
        }

        parent.zip_import_records_by_type = {}
        parent.viewer_selected_files = []
        parent.viewer_selected_types = set()
        return saved

    def _restore_parent_data(self, saved):
        parent = self.parent_window
        for name, value in saved.items():
            setattr(parent, name, value)

    def _bootstrap_event_viewer(self):
        if self.event_viewer is not None:
            return

        self.bootstrap_status.setText(
            "Preparing Event Viewer shell…"
        )
        QApplication.processEvents()

        saved = self._temporarily_hide_parent_data()

        try:
            # Construct the legacy viewer while no records/files are visible.
            self.event_viewer = DualLogViewer(
                self.parent_window
            )
        except Exception:
            self._bootstrap_failed = True
            self.bootstrap_status.setText(
                "Event Viewer initialization failed."
            )
            QMessageBox.critical(
                self,
                "Log Explore",
                traceback.format_exc(),
            )
            return
        finally:
            self._restore_parent_data(saved)

        self._replace_tab(
            "Event Viewer",
            self.event_viewer,
        )
        self.bootstrap_progress.setValue(1)
        self.bootstrap_status.setText(
            "Event Viewer ready. Preparing Value Viewer…"
        )

        # Prepare the next tab on a separate event-loop turn.
        QTimer.singleShot(50, self._bootstrap_value_viewer)

    def _bootstrap_value_viewer(self):
        if self.value_viewer is not None:
            return

        QApplication.processEvents()

        try:
            self.value_viewer = ValueViewerWidget(
                self.parent_window
            )
        except Exception:
            self._bootstrap_failed = True
            self.bootstrap_status.setText(
                "Value Viewer initialization failed."
            )
            QMessageBox.critical(
                self,
                "Log Explore",
                traceback.format_exc(),
            )
            return

        self._replace_tab(
            "Value Viewer",
            self.value_viewer,
        )
        self.bootstrap_progress.setValue(2)
        self.bootstrap_status.setText(
            "Connecting parsed data…"
        )

        QTimer.singleShot(50, self._connect_parsed_data)

    def _connect_parsed_data(self):
        try:
            if self.event_viewer is not None:
                self.event_viewer.refresh_available_sources()

            if self.value_viewer is not None:
                self.value_viewer.refresh_sources()
        except Exception:
            self._bootstrap_failed = True
            self.bootstrap_status.setText(
                "Parsed-data connection failed."
            )
            QMessageBox.critical(
                self,
                "Log Explore",
                traceback.format_exc(),
            )
            return

        self.bootstrap_progress.setValue(4)
        self.bootstrap_status.setText("Ready")
        self.refresh_button.setEnabled(True)
        self._bootstrap_complete = True

    def _initialize_operation(self):
        if self.operation is not None:
            return self.operation

        self.bootstrap_status.setText(
            "Initializing Operation…"
        )
        QApplication.processEvents()

        self.operation = OperationIntelligenceWidget(
            self.parent_window,
            self.event_viewer,
        )
        self._replace_tab("Operation", self.operation)
        self.bootstrap_status.setText("Ready")
        return self.operation

    def _initialize_investigation(self):
        if self.investigation is not None:
            return self.investigation

        self.bootstrap_status.setText(
            "Initializing Investigation…"
        )
        QApplication.processEvents()

        self.investigation = InvestigationWorkspace(
            self.event_viewer
        )
        self._replace_tab(
            "Investigation",
            self.investigation,
        )
        self.bootstrap_status.setText("Ready")
        return self.investigation

    def _initialize_spectrum(self):
        if self.spectrum is not None:
            return self.spectrum

        investigation = self._initialize_investigation()

        self.bootstrap_status.setText(
            "Initializing Spectrum…"
        )
        QApplication.processEvents()

        self.spectrum = SpectrumAnalysisWidget(
            investigation,
            standalone=False,
        )
        self._replace_tab("Spectrum", self.spectrum)
        self.bootstrap_status.setText("Ready")
        return self.spectrum

    def _on_tab_changed(self, index):
        if index < 0:
            return

        name = self.tabs.tabText(index)

        if name == "Operation" and self.operation is None:
            QTimer.singleShot(0, self._initialize_operation)
        elif (
            name == "Investigation"
            and self.investigation is None
        ):
            QTimer.singleShot(
                0,
                self._initialize_investigation,
            )
        elif name == "Spectrum" and self.spectrum is None:
            QTimer.singleShot(0, self._initialize_spectrum)

    def refresh_data(self):
        if not self._bootstrap_complete:
            return

        self.bootstrap_status.setText("Refreshing…")
        self.refresh_button.setEnabled(False)
        QApplication.processEvents()

        try:
            self.event_viewer.refresh_available_sources()
            self.value_viewer.refresh_sources()
        finally:
            self.refresh_button.setEnabled(True)
            self.bootstrap_status.setText("Ready")


# Canonical class used by MainWindow.open_log_explore.
LogExploreWindow = FastLogExploreWindow



# ---------------------------------------------------------------------------
# Commit0051: Canonical cache binding and safe Event Viewer preview
# ---------------------------------------------------------------------------

_F51_EVENT_PREVIEW_LIMIT = 5000

_F51_EVENT_ORDER = [
    "WS",
    "WATERSYSTEM",
    "CSA",
    "CGA",
    "MRSERVER",
    "GESYS",
    "LAIS",
    "PSC",
    "REVIEW",
    "VIMEASURE",
    "ACQUISITION",
]

_F51_DISPLAY_NAMES = {
    "WATERSYSTEM": "WaterSystem",
    "REVIEW": "Review",
}


def _f51_cache_key(value):
    key = str(value or "").strip().upper().replace(" ", "")
    aliases = {
        "WATER SYSTEM": "WATERSYSTEM",
        "REVIEW.OUT": "REVIEW",
        "REVIEWOUT": "REVIEW",
        "VI MEASURE": "VIMEASURE",
    }
    return aliases.get(key, key)


def _f51_display_name(key):
    canonical = _f51_cache_key(key)
    return _F51_DISPLAY_NAMES.get(canonical, canonical)


def _f51_cache(parent):
    cache = getattr(parent, "zip_import_records_by_type", {})
    return cache if isinstance(cache, dict) else {}


def _f51_available_cache_types(parent):
    cache = _f51_cache(parent)
    available = []

    for key in _F51_EVENT_ORDER:
        values = cache.get(key, [])
        if values:
            available.append(_f51_display_name(key))

    # Keep valid additional parser types without losing new formats.
    for raw_key, values in cache.items():
        if not values:
            continue
        display = _f51_display_name(raw_key)
        if display not in available:
            available.append(display)

    return available


def _f51_refresh_available_sources(self):
    parent = getattr(self, "parent_window", None)
    available = _f51_available_cache_types(parent)

    if not available:
        files = getattr(parent, "viewer_selected_files", []) or []
        found = []

        for filename in files:
            path = Path(filename)
            if is_review_file(path.name):
                display = "Review"
            else:
                detected = classify_file(path, True)
                display = _f51_display_name(detected)

            if display and display not in found:
                found.append(display)

        available = found

    if not available:
        available = [
            item for item in self.SOURCES
            if item not in {"Merged", "Custom File"}
        ]

    available.append("Custom File")

    for pane_index, combo in enumerate(
        getattr(self, "sources", [])
    ):
        current = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        combo.addItems(available)

        if current in available:
            combo.setCurrentText(current)
        else:
            default_index = min(
                pane_index,
                max(0, len(available) - 2),
            )
            combo.setCurrentIndex(default_index)

        combo.blockSignals(False)


MultiPaneLogViewer.refresh_available_sources = (
    _f51_refresh_available_sources
)


_f51_previous_source_to_records = MultiPaneLogViewer.source_to_records

def _f51_source_to_records(
    self,
    source_name,
    progress=None,
):
    parent = getattr(self, "parent_window", None)
    cache = _f51_cache(parent)

    if cache and source_name != "Custom File":
        key = _f51_cache_key(source_name)

        if key == "MERGED":
            records = [
                record
                for values in cache.values()
                for record in values
            ]
        else:
            records = list(cache.get(key, []))

        if progress is not None:
            progress.setRange(0, 1)
            progress.setValue(1)
            progress.setLabelText(
                f"Using parsed cache for {source_name}: "
                f"{len(records):,} records"
            )
            QApplication.processEvents()

        return records

    return _f51_previous_source_to_records(
        self,
        source_name,
        progress,
    )


MultiPaneLogViewer.source_to_records = _f51_source_to_records


def _f51_preview_records(records, limit=_F51_EVENT_PREVIEW_LIMIT):
    records = list(records or [])

    if len(records) <= limit:
        return records, False

    # Preserve the beginning and end of the selected log so startup remains
    # responsive while the operator can still see the complete time span.
    first_count = limit // 2
    last_count = limit - first_count
    preview = records[:first_count] + records[-last_count:]
    return preview, True


def _f51_load_preview_pane(viewer, pane_index):
    if viewer is None:
        return

    sources = getattr(viewer, "sources", [])
    if not 0 <= pane_index < len(sources):
        return

    source_name = sources[pane_index].currentText()
    parent = getattr(viewer, "parent_window", None)
    cache = _f51_cache(parent)
    key = _f51_cache_key(source_name)
    records = list(cache.get(key, []))

    if not records:
        return

    preview, limited = _f51_preview_records(records)
    rows = [record_to_viewer_row(record) for record in preview]

    viewer.all_rows[pane_index] = rows

    files = sorted({
        str(row.get("File", ""))
        for row in rows
        if row.get("File", "")
    })

    if len(files) == 1:
        file_text = files[0]
    elif files:
        file_text = f"{len(files)} files"
    else:
        file_text = source_name

    if limited:
        file_text += (
            f" — preview {len(rows):,}/{len(records):,} rows"
        )

    viewer.file_labels[pane_index].setText(
        f"File: {file_text}"
    )
    viewer.apply_view_filters(pane_index)


def _f51_bind_log_explore_data(window):
    if window is None or window.event_viewer is None:
        return

    window.bootstrap_status.setText(
        "Binding parsed records…"
    )
    QApplication.processEvents()

    window.event_viewer.refresh_available_sources()

    if window.value_viewer is not None:
        window.value_viewer.refresh_sources()

    # Automatically display lightweight previews in visible Pane 1 and Pane 2.
    visible_count = 2
    checks = getattr(
        window.event_viewer,
        "pane_visible_checks",
        [],
    )

    if checks:
        visible_count = sum(
            1 for check in checks if check.isChecked()
        )
        visible_count = max(1, min(visible_count, 2))

    for pane_index in range(visible_count):
        _f51_load_preview_pane(
            window.event_viewer,
            pane_index,
        )
        QApplication.processEvents()

    window.bootstrap_status.setText("Ready")


_f51_previous_connect_parsed_data = (
    FastLogExploreWindow._connect_parsed_data
)

def _f51_connect_parsed_data(self):
    try:
        _f51_bind_log_explore_data(self)
    except Exception:
        self._bootstrap_failed = True
        self.bootstrap_status.setText(
            "Parsed-data binding failed."
        )
        QMessageBox.critical(
            self,
            "Log Explore",
            traceback.format_exc(),
        )
        return

    self.bootstrap_progress.setValue(4)
    self.bootstrap_status.setText("Ready")
    self.refresh_button.setEnabled(True)
    self._bootstrap_complete = True


FastLogExploreWindow._connect_parsed_data = (
    _f51_connect_parsed_data
)


def _f51_refresh_data(self):
    if not self._bootstrap_complete:
        return

    self.refresh_button.setEnabled(False)

    try:
        _f51_bind_log_explore_data(self)
    finally:
        self.refresh_button.setEnabled(True)


FastLogExploreWindow.refresh_data = _f51_refresh_data


# Ensure the canonical alias points to the Commit0051 class implementation.
LogExploreWindow = FastLogExploreWindow



# ---------------------------------------------------------------------------
# Commit0051a: Reliable Event Viewer header filtering
# ---------------------------------------------------------------------------

def _f51a_normalize_column_name(value):
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(value or "").casefold(),
    )


def _f51a_resolve_row_key(row, requested_column):
    """Resolve display/header names to the actual row dictionary key."""
    if not hasattr(row, "keys"):
        return requested_column

    if requested_column in row:
        return requested_column

    requested = _f51a_normalize_column_name(requested_column)

    aliases = {
        "timestamp": ("Timestamp", "_ts", "Time", "DateTime"),
        "message": ("Message", "message", "Event", "Description"),
        "level": ("Level", "Type", "Severity"),
        "category": ("Category", "EventType", "Parameter"),
        "sourcetype": ("SourceType", "Source", "LogType"),
        "file": ("File", "Filename", "SourceFile"),
        "line": ("Line", "LineNo", "LineNumber"),
        "callid": ("CallID", "CorrelationID", "Call Id"),
        "correlationid": ("CorrelationID", "CallID"),
    }

    for candidate in aliases.get(requested, ()):
        if candidate in row:
            return candidate

    for key in row.keys():
        if _f51a_normalize_column_name(key) == requested:
            return key

    return requested_column


def _f51a_filter_cell_text(row, requested_column):
    key = _f51a_resolve_row_key(row, requested_column)
    value = row.get(key, "") if hasattr(row, "get") else ""

    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d %H:%M:%S.%f")[:-3]

    return "" if value is None else str(value)


def _f51a_manual_expression_matches(row, expression):
    mode, column, value = _c30_parse_column_filter(expression)

    if mode is None:
        return True

    search_text = str(value or "").strip().casefold()

    if mode in {"contains", "exact"}:
        cell_text = _f51a_filter_cell_text(
            row,
            column,
        ).strip().casefold()

        if mode == "contains":
            return search_text in cell_text

        return cell_text == search_text

    # Global search includes both visible values and internal structured values.
    return search_text in " ".join(
        "" if cell_value is None else str(cell_value)
        for cell_value in row.values()
    ).casefold()


# Commit0023 apply_view_filters resolves this function at runtime.
_c23_manual_expression_matches = _f51a_manual_expression_matches


def _f51a_column_source_rows(viewer, pane_index):
    """Always derive dropdown values from unfiltered Pane data."""
    all_rows = getattr(viewer, "all_rows", [])

    if 0 <= pane_index < len(all_rows):
        rows = all_rows[pane_index]
        if rows is not None:
            return list(rows)

    if 0 <= pane_index < len(getattr(viewer, "models", [])):
        model = viewer.models[pane_index]
        for name in ("rows", "_rows", "filtered_rows"):
            rows = getattr(model, name, None)
            if rows is not None:
                return list(rows)

    return []


def _f51a_column_unique_values(
    viewer,
    pane_index,
    column_name,
    limit=11,
):
    rows = _f51a_column_source_rows(
        viewer,
        pane_index,
    )

    unique = {}

    for row in rows:
        display_value = _f51a_filter_cell_text(
            row,
            column_name,
        )
        comparison_key = display_value.strip().casefold()

        if comparison_key not in unique:
            unique[comparison_key] = display_value

            if len(unique) >= limit:
                return []

    return sorted(
        unique.values(),
        key=lambda value: (
            str(value).strip() == "",
            str(value).casefold(),
        ),
    )


_c32c_column_unique_values = _f51a_column_unique_values


def _f51a_apply_selected_value(
    viewer,
    pane_index,
    column_name,
    value,
):
    edit = _c30_filter_edit(viewer, pane_index)

    if edit is None:
        return

    # Keep the exact displayed header name. The matcher resolves it to
    # the actual row key using normalized aliases.
    edit.setText(
        f"{column_name}={value}"
    )
    viewer.apply_view_filters(pane_index)


_c32c_apply_selected_value = _f51a_apply_selected_value



# ---------------------------------------------------------------------------
# Commit0052a: Quick Filter Foundation Fix
# ---------------------------------------------------------------------------

_F52A_SEVERITY_ALIASES = {
    "ERR": "ERROR", "ERROR": "ERROR", "SEVERE": "ERROR",
    "FAIL": "ERROR", "FAILED": "ERROR",
    "FATAL": "CRITICAL", "CRT": "CRITICAL", "CRITICAL": "CRITICAL",
    "WRN": "WARNING", "WARN": "WARNING", "WARNING": "WARNING",
    "INF": "INFO", "INFO": "INFO", "DBG": "INFO",
    "DEBUG": "INFO", "ENT": "INFO", "EXT": "INFO",
}


def _f52a_row_severity(row):
    candidates = [
        row.get("Level"),
        row.get("Type"),
        row.get("Severity"),
        row.get("Category"),
    ]

    text = " ".join(
        str(value or "")
        for value in candidates
    ).strip().upper()

    if not text:
        message = str(row.get("Message", "") or "").upper()
        if any(token in message for token in ("FATAL", "CRITICAL")):
            return "CRITICAL"
        if any(token in message for token in ("ERROR", "FAILED", "EXCEPTION")):
            return "ERROR"
        if any(token in message for token in ("WARNING", "WARN", "TIMEOUT")):
            return "WARNING"
        return "INFO"

    for token in re.split(r"[^A-Z]+", text):
        if token in _F52A_SEVERITY_ALIASES:
            return _F52A_SEVERITY_ALIASES[token]

    if "CRITICAL" in text or "FATAL" in text:
        return "CRITICAL"
    if "ERROR" in text or "SEVERE" in text:
        return "ERROR"
    if "WARN" in text:
        return "WARNING"
    return "INFO"


def _f52a_quick_mode_name(value):
    text = str(value or "").strip().upper()
    return {
        "ALL": "ALL",
        "ERROR": "ERROR",
        "WARNING": "WARNING",
        "WARN": "WARNING",
        "INFO": "INFO",
        "CRITICAL": "CRITICAL",
    }.get(text, "ALL")


def _f52a_manual_filter_expression(viewer, pane_index):
    if (
        hasattr(viewer, "foundation_filter_edits")
        and pane_index < len(viewer.foundation_filter_edits)
    ):
        return viewer.foundation_filter_edits[pane_index].text().strip()
    return ""


def _f52a_filtered_rows(viewer, pane_index):
    base_rows = list(viewer.all_rows[pane_index] or [])

    try:
        start, end = viewer.current_viewer_time_range()
    except Exception:
        start, end = None, None

    quick_mode = "ALL"
    if hasattr(viewer, "quick_filter_modes"):
        quick_mode = _f52a_quick_mode_name(
            viewer.quick_filter_modes[pane_index]
        )

    expression = _f52a_manual_filter_expression(viewer, pane_index)
    rows = []

    for row in base_rows:
        timestamp = row.get("_ts")

        if start or end:
            if not isinstance(timestamp, datetime):
                continue
            if start and timestamp < start:
                continue
            if end and timestamp > end:
                continue

        if quick_mode != "ALL" and _f52a_row_severity(row) != quick_mode:
            continue

        if not _f51a_manual_expression_matches(row, expression):
            continue

        rows.append(row)

    return rows, base_rows, quick_mode, expression


def _f52a_quick_counts(viewer, pane_index):
    base_rows = list(viewer.all_rows[pane_index] or [])
    counts = {
        "ALL": len(base_rows),
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "CRITICAL": 0,
    }

    for row in base_rows:
        severity = _f52a_row_severity(row)
        counts[severity] = counts.get(severity, 0) + 1

    return counts


def _f52a_update_quick_filter_labels(viewer, pane_index):
    counts = _f52a_quick_counts(viewer, pane_index)
    button_groups = getattr(viewer, "quick_filter_buttons", [])

    if not 0 <= pane_index < len(button_groups):
        return

    group = button_groups[pane_index]
    labels = {
        "ALL": f"All ({counts['ALL']:,})",
        "ERROR": f"Error ({counts['ERROR']:,})",
        "WARNING": f"Warning ({counts['WARNING']:,})",
        "INFO": f"Info ({counts['INFO']:,})",
        "CRITICAL": f"Critical ({counts['CRITICAL']:,})",
    }

    for key, text in labels.items():
        button = group.get(key)
        if button is not None:
            button.setText(text)


def _f52a_apply_view_filters(viewer, side):
    pane_index = viewer.side_index(side)
    rows, base_rows, quick_mode, expression = _f52a_filtered_rows(
        viewer,
        pane_index,
    )

    columns = viewer.pane_columns_for_rows(
        pane_index,
        rows or base_rows,
    )
    viewer.models[pane_index].set_rows(rows, columns)
    viewer.apply_table_column_widths(pane_index)

    timestamp_index = [
        (row["_ts"], index)
        for index, row in enumerate(rows)
        if isinstance(row.get("_ts"), datetime)
    ]
    timestamp_index.sort(key=lambda item: item[0])
    viewer.ts_indexes[pane_index] = timestamp_index
    viewer._sync_aliases()

    _f52a_update_quick_filter_labels(viewer, pane_index)

    file_label = viewer.file_labels[pane_index]
    existing = file_label.text().split(" • ")[0]
    file_label.setText(
        f"{existing} • Displayed {len(rows):,}/{len(base_rows):,}"
    )

    source_name = viewer.sources[pane_index].currentText()
    viewer.log(
        f"{source_name} filter — displayed "
        f"{len(rows):,}/{len(base_rows):,}; "
        f"quick={quick_mode}; contains={expression or 'none'}"
    )


MultiPaneLogViewer.apply_view_filters = _f52a_apply_view_filters


def _f52a_set_quick_filter(viewer, pane_index, mode):
    normalized = _f52a_quick_mode_name(mode)

    if not hasattr(viewer, "quick_filter_modes"):
        viewer.quick_filter_modes = ["ALL" for _ in viewer.panes]

    viewer.quick_filter_modes[pane_index] = normalized

    button_groups = getattr(viewer, "quick_filter_buttons", [])
    if 0 <= pane_index < len(button_groups):
        for key, button in button_groups[pane_index].items():
            button.setChecked(key == normalized)

    viewer.apply_view_filters(pane_index)


_c23_set_quick_filter = _f52a_set_quick_filter
_c23_row_severity = _f52a_row_severity
_c23_update_quick_filter_labels = _f52a_update_quick_filter_labels


_f52a_previous_operation_refresh = OperationIntelligenceWidget.refresh_view

def _f52a_operation_refresh_view(self, *args):
    _f52a_previous_operation_refresh(self, *args)

    events = list(getattr(self, "visible_events", []) or [])

    if not events:
        self.summary.setText(
            "Operation Summary — No operation events for the current filter."
        )
        return

    phase_counts = {}
    severity_counts = {"Error": 0, "Warning": 0, "Info": 0}
    actor_counts = {"Operator": 0, "Software": 0, "System": 0}
    linked = 0

    timestamps = []

    for event in events:
        phase = str(getattr(event, "phase", "Workflow") or "Workflow")
        phase_counts[phase] = phase_counts.get(phase, 0) + 1

        severity = str(getattr(event, "severity", "Info") or "Info")
        severity_counts[severity] = severity_counts.get(severity, 0) + 1

        actor = str(getattr(event, "actor", "Software") or "Software")
        actor_counts[actor] = actor_counts.get(actor, 0) + 1

        timestamp = getattr(event, "timestamp", None)
        if isinstance(timestamp, datetime):
            timestamps.append(timestamp)

        if getattr(event, "sonication", -1) >= 0:
            linked += 1

    phase_text = ", ".join(
        f"{phase}: {count:,}"
        for phase, count in sorted(phase_counts.items())
    )

    first_text = (
        min(timestamps).strftime("%Y/%m/%d %H:%M:%S")
        if timestamps else "—"
    )
    last_text = (
        max(timestamps).strftime("%Y/%m/%d %H:%M:%S")
        if timestamps else "—"
    )
    duration_text = "—"
    if timestamps:
        duration = max(timestamps) - min(timestamps)
        duration_text = str(duration).split(".")[0]

    self.summary.setText(
        f"Operation Summary — Events: {len(events):,} | "
        f"Operator: {actor_counts.get('Operator', 0):,} | "
        f"Software: {actor_counts.get('Software', 0):,} | "
        f"System: {actor_counts.get('System', 0):,} | "
        f"Errors: {severity_counts.get('Error', 0):,} | "
        f"Warnings: {severity_counts.get('Warning', 0):,} | "
        f"Sonication linked: {linked:,}\n"
        f"First: {first_text} | Last: {last_text} | "
        f"Duration: {duration_text}\n"
        f"Phases — {phase_text or 'None'}"
    )


OperationIntelligenceWidget.refresh_view = _f52a_operation_refresh_view



# ---------------------------------------------------------------------------
# Commit0053: Event Sync Engine
# ---------------------------------------------------------------------------

_F53_SYNC_MODES = {
    "Exact": 0.0,
    "±1 sec": 1.0,
    "±5 sec": 5.0,
    "±10 sec": 10.0,
    "±30 sec": 30.0,
    "Custom": None,
}


def _f53_normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _f53_first_value(row, names):
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return ""


def _f53_message(row):
    return _f53_normalize_text(
        _f53_first_value(
            row,
            (
                "Message",
                "Raw",
                "Status",
                "SubStatus",
                "Event",
                "Description",
            ),
        )
    )


def _f53_call_id(row):
    return _f53_normalize_text(
        _f53_first_value(
            row,
            (
                "CallID",
                "CallId",
                "Call ID",
                "Call_Id",
                "call_id",
                "CorrelationID",
                "CorrelationId",
                "Correlation ID",
            ),
        )
    )


def _f53_category(row):
    return _f53_normalize_text(
        _f53_first_value(
            row,
            (
                "Category",
                "Status",
                "SubStatus",
                "State",
                "MainState",
                "Source",
                "SourceType",
            ),
        )
    )


def _f53_severity(row):
    try:
        return _f52a_row_severity(row)
    except Exception:
        return _f53_normalize_text(
            _f53_first_value(
                row,
                ("Severity", "Level", "Type"),
            )
        ).upper()


def _f53_sync_tolerance_seconds(viewer):
    combo = getattr(viewer, "sync_mode_combo", None)
    mode = combo.currentText() if combo is not None else "±5 sec"
    value = _F53_SYNC_MODES.get(mode, 5.0)
    if value is None:
        spin = getattr(viewer, "sync_custom_spin", None)
        if spin is not None:
            return float(spin.value())
        return float(getattr(viewer, "tolerance_spin").value())
    return float(value)


def _f53_candidate_score(viewer, source_row, candidate_row, delta):
    tolerance = max(_f53_sync_tolerance_seconds(viewer), 0.001)

    # Time is always the base criterion.
    score = max(0.0, 100.0 - (delta / tolerance) * 50.0)
    if delta <= 0.0005:
        score += 100.0

    source_message = _f53_message(source_row)
    candidate_message = _f53_message(candidate_row)
    if getattr(viewer, "sync_message_chk", None) is None or viewer.sync_message_chk.isChecked():
        if source_message and candidate_message:
            if source_message == candidate_message:
                score += 140.0
            elif source_message in candidate_message or candidate_message in source_message:
                score += 70.0

    source_call_id = _f53_call_id(source_row)
    candidate_call_id = _f53_call_id(candidate_row)
    if getattr(viewer, "sync_callid_chk", None) is None or viewer.sync_callid_chk.isChecked():
        if source_call_id and candidate_call_id and source_call_id == candidate_call_id:
            score += 120.0

    source_category = _f53_category(source_row)
    candidate_category = _f53_category(candidate_row)
    if getattr(viewer, "sync_category_chk", None) is None or viewer.sync_category_chk.isChecked():
        if source_category and candidate_category and source_category == candidate_category:
            score += 45.0

    if getattr(viewer, "sync_severity_chk", None) is None or viewer.sync_severity_chk.isChecked():
        if _f53_severity(source_row) == _f53_severity(candidate_row):
            score += 20.0

    return score


def _f53_candidates_in_range(viewer, pane_index, timestamp):
    ts_index = viewer.ts_indexes[pane_index]
    if not ts_index:
        return []

    tolerance = _f53_sync_tolerance_seconds(viewer)
    lower = timestamp.timestamp() - tolerance
    upper = timestamp.timestamp() + tolerance

    # Binary-search first timestamp >= lower.
    lo, hi = 0, len(ts_index)
    while lo < hi:
        mid = (lo + hi) // 2
        if ts_index[mid][0].timestamp() < lower:
            lo = mid + 1
        else:
            hi = mid

    candidates = []
    pos = lo
    while pos < len(ts_index):
        candidate_ts, model_row = ts_index[pos]
        candidate_epoch = candidate_ts.timestamp()
        if candidate_epoch > upper:
            break

        delta = abs((candidate_ts - timestamp).total_seconds())
        if tolerance == 0.0:
            if delta <= 0.0005:
                candidates.append((candidate_ts, model_row, delta))
        elif delta <= tolerance:
            candidates.append((candidate_ts, model_row, delta))
        pos += 1

    return candidates


def _f53_select_rows(viewer, pane_index, row_numbers, primary_row):
    table = viewer.tables[pane_index]
    model = viewer.models[pane_index]
    selection_model = table.selectionModel()
    if selection_model is None:
        return

    highlight_all = (
        getattr(viewer, "sync_highlight_all_chk", None) is not None
        and viewer.sync_highlight_all_chk.isChecked()
    )

    selected_rows = row_numbers if highlight_all else [primary_row]
    selected_rows = sorted({r for r in selected_rows if 0 <= r < model.rowCount()})

    # Keep the UI responsive even when thousands of events share the same time range.
    # Build one QItemSelection and apply it once instead of emitting a selection
    # notification and repaint for every row.
    max_highlights = 250
    if len(selected_rows) > max_highlights:
        # Preserve rows nearest to the best match and always retain primary_row.
        selected_rows = sorted(selected_rows, key=lambda r: (abs(r - primary_row), r))[:max_highlights]
        selected_rows = sorted(set(selected_rows + [primary_row]))

    selection = QItemSelection()
    for row_number in selected_rows:
        left = model.index(row_number, 0)
        right = model.index(row_number, max(0, model.columnCount() - 1))
        selection.select(left, right)

    table.setUpdatesEnabled(False)
    selection_model.blockSignals(True)
    try:
        selection_model.clearSelection()
        if not selection.isEmpty():
            selection_model.select(selection, QItemSelectionModel.Select | QItemSelectionModel.Rows)
    finally:
        selection_model.blockSignals(False)
        table.setUpdatesEnabled(True)
        table.viewport().update()

    if 0 <= primary_row < model.rowCount():
        primary_index = model.index(primary_row, 0)
        selection_model.setCurrentIndex(
            primary_index,
            QItemSelectionModel.NoUpdate,
        )
        if (
            getattr(viewer, "sync_auto_scroll_chk", None) is None
            or viewer.sync_auto_scroll_chk.isChecked()
        ):
            table.scrollTo(
                primary_index,
                QTableView.PositionAtCenter,
            )


def _f53_sync_other_panes(viewer, source_idx, source_row):
    timestamp = source_row.get("_ts")
    if not isinstance(timestamp, datetime):
        viewer.log("Event Sync: selected row has no usable timestamp.")
        return

    messages = []
    viewer._syncing = True
    try:
        for pane_index in viewer.visible_indices():
            if pane_index == source_idx:
                continue

            candidates = _f53_candidates_in_range(
                viewer,
                pane_index,
                timestamp,
            )
            if not candidates:
                viewer.tables[pane_index].clearSelection()
                messages.append(
                    f"{viewer.pane_name(pane_index)}: no match"
                )
                continue

            scored = []
            for candidate_ts, model_row, delta in candidates:
                candidate_row = (
                    viewer.models[pane_index].row_at(model_row)
                    or {}
                )
                score = _f53_candidate_score(
                    viewer,
                    source_row,
                    candidate_row,
                    delta,
                )
                scored.append(
                    (
                        score,
                        -delta,
                        model_row,
                        delta,
                        candidate_row,
                    )
                )

            scored.sort(reverse=True)
            best_score, _, primary_row, best_delta, primary_data = scored[0]
            all_rows = [item[2] for item in scored]

            _f53_select_rows(
                viewer,
                pane_index,
                all_rows,
                primary_row,
            )

            message_match = bool(
                _f53_message(source_row)
                and _f53_message(source_row)
                == _f53_message(primary_data)
            )
            callid_match = bool(
                _f53_call_id(source_row)
                and _f53_call_id(source_row)
                == _f53_call_id(primary_data)
            )

            match_flags = []
            if message_match:
                match_flags.append("Message")
            if callid_match:
                match_flags.append("CallID")
            if (
                _f53_category(source_row)
                and _f53_category(source_row)
                == _f53_category(primary_data)
            ):
                match_flags.append("Category")
            if _f53_severity(source_row) == _f53_severity(primary_data):
                match_flags.append("Severity")

            messages.append(
                f"{viewer.pane_name(pane_index)}: "
                f"{len(candidates):,} event(s), "
                f"best Δ{best_delta:.3f}s"
                + (
                    " [" + ", ".join(match_flags) + "]"
                    if match_flags else ""
                )
            )
    finally:
        viewer._syncing = False

    mode = (
        viewer.sync_mode_combo.currentText()
        if hasattr(viewer, "sync_mode_combo")
        else f"±{_f53_sync_tolerance_seconds(viewer):g}s"
    )
    viewer.log(
        f"Event Sync ({mode}) — "
        + ("; ".join(messages) if messages else "no other visible panes")
    )


_old_f53_build_ui = MultiPaneLogViewer.build_ui

def _f53_build_ui(self):
    _old_f53_build_ui(self)

    # Multiple rows are required for range-based synchronized highlighting.
    for table in self.tables:
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)

    sync_bar_widget = QWidget(self)
    sync_bar = QHBoxLayout(sync_bar_widget)
    sync_bar.setContentsMargins(0, 0, 0, 0)
    sync_bar.setSpacing(6)

    sync_bar.addWidget(QLabel("Event Sync:"))

    self.sync_mode_combo = QComboBox()
    self.sync_mode_combo.addItems(
        ["Exact", "±1 sec", "±5 sec", "±10 sec", "±30 sec", "Custom"]
    )
    self.sync_mode_combo.setCurrentText("±5 sec")
    self.sync_mode_combo.setToolTip(
        "Select the timestamp range used to synchronize other visible logs."
    )
    sync_bar.addWidget(self.sync_mode_combo)

    self.sync_custom_spin = QSpinBox()
    self.sync_custom_spin.setRange(0, 3600)
    self.sync_custom_spin.setValue(5)
    self.sync_custom_spin.setSuffix(" sec")
    self.sync_custom_spin.setEnabled(False)
    self.sync_custom_spin.setToolTip(
        "Custom Event Sync range in seconds."
    )
    sync_bar.addWidget(self.sync_custom_spin)

    self.sync_highlight_all_chk = QCheckBox("Highlight all in range")
    self.sync_highlight_all_chk.setChecked(False)
    self.sync_auto_scroll_chk = QCheckBox("Auto scroll")
    self.sync_auto_scroll_chk.setChecked(True)

    self.sync_message_chk = QCheckBox("Message priority")
    self.sync_message_chk.setChecked(True)
    self.sync_callid_chk = QCheckBox("CallID priority")
    self.sync_callid_chk.setChecked(True)
    self.sync_category_chk = QCheckBox("Category priority")
    self.sync_category_chk.setChecked(True)
    self.sync_severity_chk = QCheckBox("Severity priority")
    self.sync_severity_chk.setChecked(True)

    for widget in (
        self.sync_highlight_all_chk,
        self.sync_auto_scroll_chk,
        self.sync_message_chk,
        self.sync_callid_chk,
        self.sync_category_chk,
        self.sync_severity_chk,
    ):
        sync_bar.addWidget(widget)

    sync_bar.addStretch(1)

    root_layout = self.layout()
    root_layout.insertWidget(1, sync_bar_widget)

    self.sync_mode_combo.currentTextChanged.connect(
        lambda text: self.sync_custom_spin.setEnabled(text == "Custom")
    )

    # Retain the old control only for backward compatibility. Event Sync uses
    # the new range selector and custom spin.
    self.tolerance_spin.setValue(5)
    self.tolerance_spin.setEnabled(False)
    self.tolerance_spin.setToolTip(
        "Legacy jump tolerance. Event Sync range is controlled by the Event Sync bar."
    )


MultiPaneLogViewer.build_ui = _f53_build_ui


def _f53_row_selected(self, side):
    if self._syncing:
        return

    pane_index = self.side_index(side)
    table = self.tables[pane_index]
    model = self.models[pane_index]
    selection_model = table.selectionModel()
    if selection_model is None:
        return

    current_index = selection_model.currentIndex()
    if not current_index.isValid():
        selected = selection_model.selectedRows()
        if not selected:
            return
        current_index = selected[0]

    row = model.row_at(current_index.row())
    if not row:
        return

    self.show_detail(pane_index, row)
    _f53_sync_other_panes(self, pane_index, row)


MultiPaneLogViewer.row_selected = _f53_row_selected


def _f53_jump_other_panes(self, source_idx, timestamp_or_row):
    if isinstance(timestamp_or_row, dict):
        source_row = timestamp_or_row
    else:
        source_row = {"_ts": timestamp_or_row}
    _f53_sync_other_panes(self, source_idx, source_row)


MultiPaneLogViewer.jump_other_panes = _f53_jump_other_panes


def _f53_jump_other_side(self, side, timestamp_or_row):
    self.jump_other_panes(
        self.side_index(side),
        timestamp_or_row,
    )


MultiPaneLogViewer.jump_other_side = _f53_jump_other_side



# ---------------------------------------------------------------------------
# Commit0053a: Filter Engine Stabilization
# ---------------------------------------------------------------------------

def _f53a_normalize_filter_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()


def _f53a_parse_foundation_expression(expression):
    text = str(expression or "").strip()
    if not text:
        return None
    if "~" in text:
        column, value = text.split("~", 1)
        return ("contains", column.strip() or "Message", value.strip())
    if "=" in text:
        column, value = text.split("=", 1)
        return ("exact", column.strip() or "Message", value.strip())
    return ("contains", "Message", text)


def _f53a_row_filter_value(row, column):
    if column in row:
        return row.get(column)
    wanted = _f53a_normalize_filter_text(column)
    for key, value in row.items():
        if _f53a_normalize_filter_text(key) == wanted:
            return value
    return ""


def _f53a_manual_expression_matches(row, expression):
    parsed = _f53a_parse_foundation_expression(expression)
    if parsed is None:
        return True

    mode, column, keyword = parsed
    left = _f53a_normalize_filter_text(
        _f53a_row_filter_value(row, column)
    )
    right = _f53a_normalize_filter_text(keyword)

    if mode == "exact":
        return left == right
    if not right:
        return True
    return right in left


def _f53a_filter_rows(viewer, pane_index):
    base_rows = list(viewer.all_rows[pane_index] or [])

    try:
        start, end = viewer.current_viewer_time_range()
    except Exception:
        start, end = None, None

    if not hasattr(viewer, "quick_filter_modes"):
        viewer.quick_filter_modes = ["ALL" for _ in viewer.panes]

    quick_mode = _f52a_quick_mode_name(
        viewer.quick_filter_modes[pane_index]
    )
    expression = _f52a_manual_filter_expression(viewer, pane_index)

    rows = []
    for row in base_rows:
        timestamp = row.get("_ts")

        if start or end:
            if not isinstance(timestamp, datetime):
                continue
            if start and timestamp < start:
                continue
            if end and timestamp > end:
                continue

        if quick_mode != "ALL" and _f52a_row_severity(row) != quick_mode:
            continue

        if not _f53a_manual_expression_matches(row, expression):
            continue

        rows.append(row)

    return rows, base_rows, quick_mode, expression


def _f53a_apply_view_filters(viewer, side):
    pane_index = viewer.side_index(side)
    rows, base_rows, quick_mode, expression = _f53a_filter_rows(
        viewer,
        pane_index,
    )

    columns = viewer.pane_columns_for_rows(
        pane_index,
        rows or base_rows,
    )
    viewer.models[pane_index].set_rows(rows, columns)
    viewer.apply_table_column_widths(pane_index)

    timestamp_index = [
        (row["_ts"], index)
        for index, row in enumerate(rows)
        if isinstance(row.get("_ts"), datetime)
    ]
    timestamp_index.sort(key=lambda item: item[0])
    viewer.ts_indexes[pane_index] = timestamp_index
    viewer._sync_aliases()

    _f52a_update_quick_filter_labels(viewer, pane_index)

    file_label = viewer.file_labels[pane_index]
    existing = file_label.text().split(" • ")[0]
    file_label.setText(
        f"{existing} • Displayed {len(rows):,}/{len(base_rows):,}"
    )

    table = viewer.tables[pane_index]
    table.viewport().update()
    table.update()

    source_name = viewer.sources[pane_index].currentText()
    viewer.log(
        f"{source_name} filter — displayed {len(rows):,}/{len(base_rows):,}; "
        f"quick={quick_mode}; expression={expression or 'none'}"
    )


MultiPaneLogViewer.apply_view_filters = _f53a_apply_view_filters


def _f53a_set_quick_filter(viewer, pane_index, mode):
    normalized = _f52a_quick_mode_name(mode)

    if not hasattr(viewer, "quick_filter_modes"):
        viewer.quick_filter_modes = ["ALL" for _ in viewer.panes]

    viewer.quick_filter_modes[pane_index] = normalized

    groups = getattr(viewer, "quick_filter_buttons", [])
    if 0 <= pane_index < len(groups):
        for key, button in groups[pane_index].items():
            button.setChecked(key == normalized)

    viewer.apply_view_filters(pane_index)


_c23_set_quick_filter = _f53a_set_quick_filter


def _f53a_rebind_quick_filter_buttons(viewer):
    groups = getattr(viewer, "quick_filter_buttons", [])
    for pane_index, group in enumerate(groups):
        for mode, button in group.items():
            try:
                button.clicked.disconnect()
            except Exception:
                pass
            button.clicked.connect(
                lambda checked=False, p=pane_index, m=mode:
                    _f53a_set_quick_filter(viewer, p, m)
            )


_old_f53a_build_ui = MultiPaneLogViewer.build_ui

def _f53a_build_ui(self):
    _old_f53a_build_ui(self)
    _f53a_rebind_quick_filter_buttons(self)

    light_selection_style = """
        QTableView::item:selected {
            background: #D6ECFF;
            color: #111111;
        }
        QTableView::item:selected:active {
            background: #CBE6FF;
            color: #111111;
        }
        QTableView::item:selected:!active {
            background: #E7F3FF;
            color: #111111;
        }
    """
    for table in self.tables:
        table.setStyleSheet(table.styleSheet() + light_selection_style)


MultiPaneLogViewer.build_ui = _f53a_build_ui
_f51a_manual_expression_matches = _f53a_manual_expression_matches


# ---------------------------------------------------------------------------
# Commit0060: Viewer right-click contains + Quick Filter action recovery
# ---------------------------------------------------------------------------

def _c60_normalize(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", " ", text).strip().casefold()


def _c60_resolve_column(row, requested):
    wanted = _c60_normalize(requested)
    for key in row.keys():
        if _c60_normalize(key) == wanted:
            return key
    return requested


def _c60_matches_expression(row, expression):
    text = str(expression or "").strip()
    if not text:
        return True

    if "~" in text:
        column, value = text.split("~", 1)
        key = _c60_resolve_column(row, column.strip())
        return _c60_normalize(value) in _c60_normalize(row.get(key, ""))

    if "=" in text:
        column, value = text.split("=", 1)
        key = _c60_resolve_column(row, column.strip())
        return _c60_normalize(row.get(key, "")) == _c60_normalize(value)

    needle = _c60_normalize(text)
    return needle in _c60_normalize(
        " ".join(str(value or "") for key, value in row.items() if key != "_ts")
    )


def _c60_apply_view_filters(viewer, side):
    pane_index = viewer.side_index(side)
    try:
        base_rows = list(viewer.all_rows[pane_index] or [])
        try:
            start, end = viewer.current_viewer_time_range()
        except Exception:
            start, end = None, None

        if not hasattr(viewer, "quick_filter_modes"):
            viewer.quick_filter_modes = ["ALL" for _ in viewer.panes]
        quick_mode = _f52a_quick_mode_name(viewer.quick_filter_modes[pane_index])

        edit = _c30_filter_edit(viewer, pane_index)
        expression = edit.text().strip() if edit is not None else ""

        filtered = []
        for row in base_rows:
            timestamp = row.get("_ts")
            if start or end:
                if not isinstance(timestamp, datetime):
                    continue
                if start and timestamp < start:
                    continue
                if end and timestamp > end:
                    continue
            if quick_mode != "ALL" and _f52a_row_severity(row) != quick_mode:
                continue
            if not _c60_matches_expression(row, expression):
                continue
            filtered.append(row)

        columns = viewer.pane_columns_for_rows(pane_index, filtered or base_rows)
        viewer.models[pane_index].set_rows(filtered, columns)
        viewer.apply_table_column_widths(pane_index)
        viewer.ts_indexes[pane_index] = sorted(
            (row["_ts"], index)
            for index, row in enumerate(filtered)
            if isinstance(row.get("_ts"), datetime)
        )
        viewer._sync_aliases()
        _f52a_update_quick_filter_labels(viewer, pane_index)

        file_label = viewer.file_labels[pane_index]
        base_label = file_label.text().split(" • ")[0]
        file_label.setText(f"{base_label} • Displayed {len(filtered):,}/{len(base_rows):,}")

        condition_parts = []
        if quick_mode != "ALL":
            condition_parts.append(f"Quick Filter={quick_mode.title()}")
        if expression:
            condition_parts.append(expression)
        condition_text = ", ".join(condition_parts) or "No filter"
        message = (
            f"{viewer.pane_name(pane_index)}: {condition_text} — "
            f"{len(filtered):,}/{len(base_rows):,} rows"
        )
        viewer.status.setText(message)
        viewer.log(message)
        viewer.tables[pane_index].viewport().update()
    except Exception as error:
        write_startup_log("Commit0060 filter application failed.\n\n" + traceback.format_exc())
        QMessageBox.critical(viewer, "Viewer Filter", f"The filter could not be applied.\n\n{error}")


MultiPaneLogViewer.apply_view_filters = _c60_apply_view_filters


def _c60_set_quick_filter(viewer, pane_index, mode):
    normalized = _f52a_quick_mode_name(mode)
    if not hasattr(viewer, "quick_filter_modes"):
        viewer.quick_filter_modes = ["ALL" for _ in viewer.panes]
    viewer.quick_filter_modes[pane_index] = normalized

    groups = getattr(viewer, "quick_filter_buttons", [])
    if 0 <= pane_index < len(groups):
        for key, button in groups[pane_index].items():
            button.blockSignals(True)
            button.setChecked(key == normalized)
            button.blockSignals(False)
    viewer.apply_view_filters(pane_index)


def _c60_header_menu(viewer, pane_index, position):
    if not (0 <= pane_index < len(viewer.tables)):
        return
    table = viewer.tables[pane_index]
    header = table.horizontalHeader()
    logical_index = header.logicalIndexAt(position)
    if logical_index < 0:
        return
    column_name = _c30_column_name(viewer, pane_index, logical_index)
    if not column_name:
        return

    menu = QMenu(header)
    heading = menu.addAction(f"Column: {column_name}")
    heading.setEnabled(False)
    menu.addSeparator()

    contains_action = menu.addAction("Filter contains...")
    exact_action = menu.addAction("Filter exact...")
    menu.addSeparator()
    clear_action = menu.addAction("Clear pane filter")
    edit = _c30_filter_edit(viewer, pane_index)
    clear_action.setEnabled(bool(edit and edit.text().strip()))

    selected = menu.exec(header.mapToGlobal(position))
    if selected == contains_action:
        current = ""
        if edit is not None:
            parsed = _f53a_parse_foundation_expression(edit.text())
            if parsed and _c60_normalize(parsed[1]) == _c60_normalize(column_name):
                current = str(parsed[2] or "")
        value, accepted = QInputDialog.getText(
            viewer,
            f"Filter contains: {column_name}",
            f"Show rows where {column_name} contains:",
            text=current,
        )
        if accepted and edit is not None:
            value = str(value).strip()
            edit.setText(f"{column_name}~{value}" if value else "")
            viewer.apply_view_filters(pane_index)
    elif selected == exact_action:
        value, accepted = QInputDialog.getText(
            viewer,
            f"Filter exact: {column_name}",
            f"Show rows where {column_name} exactly equals:",
        )
        if accepted and edit is not None:
            value = str(value).strip()
            edit.setText(f"{column_name}={value}" if value else "")
            viewer.apply_view_filters(pane_index)
    elif selected == clear_action and edit is not None:
        edit.clear()
        viewer.apply_view_filters(pane_index)


def _c60_bind_viewer_actions(viewer):
    # Rebind Quick Filter buttons to the final Commit0060 handler.
    for pane_index, group in enumerate(getattr(viewer, "quick_filter_buttons", [])):
        for mode, button in group.items():
            try:
                button.clicked.disconnect()
            except (TypeError, RuntimeError):
                pass
            button.clicked.connect(
                lambda checked=False, p=pane_index, m=mode: _c60_set_quick_filter(viewer, p, m)
            )

    # Replace inherited header-menu connections so only one active handler remains.
    for pane_index, table in enumerate(getattr(viewer, "tables", [])):
        header = table.horizontalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except (TypeError, RuntimeError):
            pass
        header.customContextMenuRequested.connect(
            lambda position, p=pane_index: _c60_header_menu(viewer, p, position)
        )
        header.setProperty("c60FilterMenuInstalled", True)


_old_c60_build_ui = MultiPaneLogViewer.build_ui


def _c60_build_ui(self):
    _old_c60_build_ui(self)
    _c60_bind_viewer_actions(self)


MultiPaneLogViewer.build_ui = _c60_build_ui
_c23_set_quick_filter = _c60_set_quick_filter
_c30_show_header_context_menu = _c60_header_menu


# ---------------------------------------------------------------------------
# Commit0061: definitive independent Viewer filter state
# ---------------------------------------------------------------------------

def _c61_filter_state(viewer):
    count = len(getattr(viewer, "panes", [])) or getattr(viewer, "MAX_PANES", 4)
    expressions = getattr(viewer, "_c61_filter_expressions", None)
    if not isinstance(expressions, list) or len(expressions) != count:
        expressions = [""] * count
        viewer._c61_filter_expressions = expressions
    modes = getattr(viewer, "quick_filter_modes", None)
    if not isinstance(modes, list) or len(modes) != count:
        modes = ["ALL"] * count
        viewer.quick_filter_modes = modes
    return expressions, modes


def _c61_expression_matches(row, expression):
    mode, column, value = _c30_parse_column_filter(expression)
    if mode is None:
        return True
    needle = _c60_normalize(value)
    if mode == "global":
        return needle in _c60_normalize(
            " ".join(str(v or "") for k, v in row.items() if k != "_ts")
        )
    actual_key = _c60_resolve_column(row, column)
    haystack = _c60_normalize(row.get(actual_key, ""))
    return needle in haystack if mode == "contains" else haystack == needle


def _c61_apply_view_filters(viewer, side):
    pane_index = viewer.side_index(side)
    expressions, modes = _c61_filter_state(viewer)
    base_rows = list(viewer.all_rows[pane_index] or [])
    expression = expressions[pane_index]
    quick_mode = _f52a_quick_mode_name(modes[pane_index])

    try:
        start, end = viewer.current_viewer_time_range()
    except Exception:
        start, end = None, None

    rows = []
    for row in base_rows:
        timestamp = row.get("_ts")
        if start or end:
            if not isinstance(timestamp, datetime):
                continue
            if start and timestamp < start:
                continue
            if end and timestamp > end:
                continue
        if quick_mode != "ALL" and _f52a_row_severity(row) != quick_mode:
            continue
        if not _c61_expression_matches(row, expression):
            continue
        rows.append(row)

    columns = viewer.pane_columns_for_rows(pane_index, rows or base_rows)
    viewer.models[pane_index].set_rows(rows, columns)
    viewer.apply_table_column_widths(pane_index)
    viewer.ts_indexes[pane_index] = sorted(
        (row["_ts"], index)
        for index, row in enumerate(rows)
        if isinstance(row.get("_ts"), datetime)
    )
    viewer._sync_aliases()
    _f52a_update_quick_filter_labels(viewer, pane_index)

    groups = getattr(viewer, "quick_filter_buttons", [])
    if 0 <= pane_index < len(groups):
        for key, button in groups[pane_index].items():
            button.blockSignals(True)
            button.setChecked(key == quick_mode)
            button.blockSignals(False)

    details = []
    if quick_mode != "ALL":
        details.append(f"Quick Filter={quick_mode.title()}")
    if expression:
        details.append(expression)
    status_text = ", ".join(details) or "No filter"
    message = (
        f"{viewer.pane_name(pane_index)}: {status_text} — "
        f"Displayed {len(rows):,}/{len(base_rows):,} rows"
    )
    if hasattr(viewer, "status"):
        viewer.status.setText(message)
    viewer.log(message)
    if 0 <= pane_index < len(viewer.file_labels):
        label = viewer.file_labels[pane_index]
        base = label.text().split(" • ")[0]
        label.setText(f"{base} • Displayed {len(rows):,}/{len(base_rows):,}")
    viewer.tables[pane_index].viewport().update()


def _c61_set_quick_filter(viewer, pane_index, mode):
    _, modes = _c61_filter_state(viewer)
    modes[pane_index] = _f52a_quick_mode_name(mode)
    _c61_apply_view_filters(viewer, pane_index)


def _c61_set_expression(viewer, pane_index, expression):
    expressions, _ = _c61_filter_state(viewer)
    expressions[pane_index] = str(expression or "").strip()
    # Keep a visible legacy field synchronized when an older layout provides it.
    edit = _c30_filter_edit(viewer, pane_index)
    if edit is not None:
        edit.blockSignals(True)
        edit.setText(expressions[pane_index])
        edit.blockSignals(False)
    _c61_apply_view_filters(viewer, pane_index)


def _c61_header_menu(viewer, pane_index, position):
    if not (0 <= pane_index < len(viewer.tables)):
        return
    table = viewer.tables[pane_index]
    header = table.horizontalHeader()
    logical_index = header.logicalIndexAt(position)
    if logical_index < 0:
        return
    column_name = _c30_column_name(viewer, pane_index, logical_index)
    if not column_name:
        return

    menu = QMenu(header)
    title = menu.addAction(f"Column: {column_name}")
    title.setEnabled(False)
    menu.addSeparator()
    contains_action = menu.addAction("Filter contains...")
    exact_action = menu.addAction("Filter exact...")
    menu.addSeparator()
    clear_action = menu.addAction("Clear pane filter")
    expressions, _ = _c61_filter_state(viewer)
    clear_action.setEnabled(bool(expressions[pane_index]))

    selected = menu.exec(header.mapToGlobal(position))
    if selected not in (contains_action, exact_action, clear_action):
        return
    if selected == clear_action:
        _c61_set_expression(viewer, pane_index, "")
        return

    mode = "contains" if selected == contains_action else "exact"
    _, current_column, current_value = _c30_parse_column_filter(expressions[pane_index])
    default = str(current_value or "") if _c60_normalize(current_column) == _c60_normalize(column_name) else ""
    prompt = "contains" if mode == "contains" else "exactly equals"
    value, accepted = QInputDialog.getText(
        viewer,
        f"Filter {mode}: {column_name}",
        f"Show rows where {column_name} {prompt}:",
        text=default,
    )
    if accepted:
        value = str(value).strip()
        operator = "~" if mode == "contains" else "="
        _c61_set_expression(viewer, pane_index, f"{column_name}{operator}{value}" if value else "")


def _c61_row_menu(viewer, pane_index, position):
    table = viewer.tables[pane_index]
    model = viewer.models[pane_index]
    index = table.indexAt(position)
    if not index.isValid():
        return
    row = model.row_at(index.row()) or {}
    column_name = model.columns[index.column()] if 0 <= index.column() < len(model.columns) else ""
    cell_value = model.data(index, Qt.DisplayRole)

    menu = QMenu(table)
    copy_cell = menu.addAction("Copy Cell")
    copy_row = menu.addAction("Copy Row")
    copy_timestamp = menu.addAction("Copy Timestamp")
    copy_message = menu.addAction("Copy Message / Value")
    menu.addSeparator()
    contains = menu.addAction(f"Filter {column_name} contains this value")
    exact = menu.addAction(f"Filter {column_name} equals this value")
    clear = menu.addAction("Clear Pane Filter")
    menu.addSeparator()
    copy_rule = menu.addAction("Copy Rule Text")
    noise = menu.addAction("Add to Noise Rule")

    selected = menu.exec(table.viewport().mapToGlobal(position))
    clipboard = QApplication.clipboard()
    if selected == copy_cell:
        clipboard.setText(str(cell_value or ""))
    elif selected == copy_row:
        clipboard.setText("\t".join(str(row.get(c, "")) for c in model.columns))
    elif selected == copy_timestamp:
        clipboard.setText(str(row.get("Timestamp", "") or ""))
    elif selected == copy_message:
        clipboard.setText(str(row.get("Message", row.get("Value", "")) or ""))
    elif selected == contains and column_name:
        _c61_set_expression(viewer, pane_index, f"{column_name}~{cell_value}")
    elif selected == exact and column_name:
        _c61_set_expression(viewer, pane_index, f"{column_name}={cell_value}")
    elif selected == clear:
        _c61_set_expression(viewer, pane_index, "")
    elif selected == copy_rule:
        viewer.copy_rule_text(pane_index)
    elif selected == noise:
        viewer.approve_selected_noise(pane_index)


def _c61_bind_filter_actions(viewer):
    _c61_filter_state(viewer)
    for pane_index, group in enumerate(getattr(viewer, "quick_filter_buttons", [])):
        for mode, button in group.items():
            try:
                button.clicked.disconnect()
            except (TypeError, RuntimeError):
                pass
            button.clicked.connect(
                lambda checked=False, p=pane_index, m=mode: _c61_set_quick_filter(viewer, p, m)
            )

    for pane_index, table in enumerate(getattr(viewer, "tables", [])):
        header = table.horizontalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except (TypeError, RuntimeError):
            pass
        header.customContextMenuRequested.connect(
            lambda pos, p=pane_index: _c61_header_menu(viewer, p, pos)
        )
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        try:
            table.customContextMenuRequested.disconnect()
        except (TypeError, RuntimeError):
            pass
        table.customContextMenuRequested.connect(
            lambda pos, p=pane_index: _c61_row_menu(viewer, p, pos)
        )


_old_c61_build_ui = MultiPaneLogViewer.build_ui

def _c61_build_ui(self):
    _old_c61_build_ui(self)
    _c61_bind_filter_actions(self)


_old_c61_load_pane = MultiPaneLogViewer.load_pane

def _c61_load_pane(self, side):
    pane_index = self.side_index(side)
    expressions, modes = _c61_filter_state(self)
    expressions[pane_index] = ""
    modes[pane_index] = "ALL"
    result = _old_c61_load_pane(self, side)
    # Re-apply with the final Commit0061 engine after loading.
    _c61_apply_view_filters(self, pane_index)
    return result


MultiPaneLogViewer.build_ui = _c61_build_ui
MultiPaneLogViewer.apply_view_filters = _c61_apply_view_filters
MultiPaneLogViewer.load_pane = _c61_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)
MultiPaneLogViewer.show_row_context_menu = _c61_row_menu
_c23_set_quick_filter = _c61_set_quick_filter
_c30_show_header_context_menu = _c61_header_menu


# ---------------------------------------------------------------------------
# Commit0063 - Deferred Viewer Presentation
# ---------------------------------------------------------------------------
# The viewer was previously shown before source models, filters, column widths,
# and initial pane contents had finished preparing.  A user click during that
# window could make Windows report the application as not responding.  Keep the
# viewer hidden and non-interactive until every visible pane is fully loaded.

_c63_previous_open_dual_viewer = MainWindow.open_dual_viewer
_c63_previous_start_import_clicked = MainWindow.start_import_clicked


def _c63_prepare_viewer_before_show(main_window):
    viewer = getattr(main_window, "viewer_window", None)
    if viewer is None:
        viewer = DualLogViewer(main_window)
        main_window.viewer_window = viewer

    viewer.hide()
    viewer.setEnabled(False)
    viewer.setUpdatesEnabled(False)

    progress = SoftProgressDialog(
        "Preparing Log Viewer",
        "Preparing source list...",
        None,
        0,
        100,
        main_window,
    )
    progress.setWindowModality(Qt.ApplicationModal)
    progress.setValue(5)
    progress.show()
    QApplication.processEvents()

    try:
        if hasattr(viewer, "refresh_available_sources"):
            viewer.refresh_available_sources()

        visible = list(viewer.visible_indices()) if hasattr(viewer, "visible_indices") else [0, 1]
        if not visible:
            visible = [0]

        # Preload every visible pane while the viewer is still hidden.  This
        # includes parsing, row conversion, model reset, filters, timestamp
        # indexes, column sizing, and status/row-count updates.
        total = max(1, len(visible))
        for order, pane_index in enumerate(visible, start=1):
            pane_name = viewer.pane_name(pane_index) if hasattr(viewer, "pane_name") else f"Pane {pane_index + 1}"
            source_name = ""
            if pane_index < len(getattr(viewer, "sources", [])):
                source_name = viewer.sources[pane_index].currentText()
            progress.setLabelText(
                f"Preparing {pane_name}: {source_name or 'selected log'}\n"
                f"Please wait until the Viewer is ready."
            )
            progress.setValue(5 + int(order * 88 / total))
            QApplication.processEvents()
            viewer.load_pane(pane_index)

        # Flush queued layout/model events before allowing the window to appear.
        progress.setLabelText("Finalizing Viewer layout and filters...")
        progress.setValue(96)
        QApplication.processEvents()
        viewer.updateGeometry()
        viewer.adjustSize() if viewer.width() <= 0 or viewer.height() <= 0 else None
        QApplication.processEvents()
        progress.setValue(100)
        QApplication.processEvents()
    finally:
        viewer.setUpdatesEnabled(True)
        viewer.setEnabled(True)
        progress.close()

    return viewer


def _c63_open_dual_viewer(self):
    try:
        if getattr(self, "_c63_prepare_before_show", False):
            viewer = _c63_prepare_viewer_before_show(self)
            self._c63_prepare_before_show = False
        else:
            viewer = getattr(self, "viewer_window", None)
            if viewer is None:
                viewer = DualLogViewer(self)
                self.viewer_window = viewer
            if hasattr(viewer, "refresh_available_sources"):
                viewer.refresh_available_sources()

        geo = QApplication.primaryScreen().availableGeometry()
        w = min(max(860, int(geo.width() * 0.58)), geo.width() - 80)
        h = min(max(520, int(geo.height() * 0.52)), geo.height() - 80)
        viewer.resize(w, h)
        viewer.move(
            geo.x() + max(20, (geo.width() - w) // 2),
            geo.y() + max(20, (geo.height() - h) // 2),
        )
        viewer.show()
        viewer.raise_()
        viewer.activateWindow()
        QApplication.processEvents()
        if hasattr(viewer, "status"):
            viewer.status.setText("Ready")
    except Exception:
        self._c63_prepare_before_show = False
        text = "Log Viewer preparation failed.\n\n" + traceback.format_exc()
        write_startup_log(text)
        QMessageBox.critical(self, APP_TITLE, text)


def _c63_start_import_clicked(self):
    # Preserve the original Smart File Discovery workflow.  Only change when
    # the Viewer becomes visible.
    self._c63_prepare_before_show = True
    try:
        return _c63_previous_start_import_clicked(self)
    finally:
        # Cancelled discovery or an earlier error never calls open_dual_viewer.
        if getattr(self, "viewer_window", None) is None or not self.viewer_window.isVisible():
            self._c63_prepare_before_show = False


MainWindow.open_dual_viewer = _c63_open_dual_viewer
MainWindow.start_import_clicked = _c63_start_import_clicked


# ---------------------------------------------------------------------------
# Commit0064 - Single Viewer Presentation
# ---------------------------------------------------------------------------
# Commit0044 wrapped START and automatically opened the separate Log Explore
# workspace after the normal Log Viewer had already been opened.  Commit0063
# then deferred the normal Viewer presentation, so both windows became visible
# together.  START must present one Viewer only.  Keep Log Explore available
# from its dedicated button/menu, but do not auto-open it after Smart Discovery.
_c64_start_backend = globals().get("_f44_previous_start", _c63_previous_start_import_clicked)


def _c64_open_dual_viewer(self):
    # Guard against duplicate signal delivery or nested open requests during
    # model/layout finalization.  Reuse the existing Viewer instance.
    if getattr(self, "_c64_viewer_open_in_progress", False):
        return
    self._c64_viewer_open_in_progress = True
    try:
        return _c63_open_dual_viewer(self)
    finally:
        self._c64_viewer_open_in_progress = False


def _c64_start_import_clicked(self):
    # Preserve Smart File Discovery and Commit0063's deferred presentation,
    # while bypassing the Commit0044 wrapper that opened a second window.
    self._c63_prepare_before_show = True
    try:
        return _c64_start_backend(self)
    finally:
        viewer = getattr(self, "viewer_window", None)
        if viewer is None or not viewer.isVisible():
            self._c63_prepare_before_show = False


MainWindow.open_dual_viewer = _c64_open_dual_viewer
MainWindow.start_import_clicked = _c64_start_import_clicked


# ---------------------------------------------------------------------------
# Commit0065 - Unified Viewer Preparation Progress
# ---------------------------------------------------------------------------
# Commit0063 introduced an outer "Preparing Log Viewer" dialog while the
# existing pane loader still created its own indexing dialog.  That produced
# two simultaneous modal progress windows.  Route all pane-loading progress
# into the single outer dialog and keep the nested dialog lifecycle as no-ops.

class _C65UnifiedProgressProxy:
    """QProgressDialog-compatible proxy backed by one shared outer dialog."""

    def __init__(self, viewer, outer, title, text):
        self._viewer = viewer
        self._outer = outer
        self._minimum = 0
        self._maximum = 100
        self._value = 0
        self._title = str(title or "")
        self._text = str(text or "")
        self._apply_label(self._text)

    def _apply_label(self, text):
        prefix = getattr(self._viewer, "_c65_progress_prefix", "")
        message = str(text or "").strip()
        if prefix and message:
            message = f"{prefix}\n{message}"
        elif prefix:
            message = prefix
        try:
            self._outer.setLabelText(message)
            QApplication.processEvents()
        except Exception:
            pass

    def setRange(self, minimum, maximum):
        self._minimum = int(minimum)
        self._maximum = int(maximum)

    def setMinimum(self, minimum):
        self._minimum = int(minimum)

    def setMaximum(self, maximum):
        self._maximum = int(maximum)

    def setValue(self, value):
        self._value = int(value)
        low = int(getattr(self._viewer, "_c65_progress_low", 5))
        high = int(getattr(self._viewer, "_c65_progress_high", 95))
        denominator = max(1, self._maximum - self._minimum)
        fraction = max(0.0, min(1.0, (self._value - self._minimum) / denominator))
        mapped = low + int((high - low) * fraction)
        try:
            self._outer.setValue(mapped)
            QApplication.processEvents()
        except Exception:
            pass

    def value(self):
        return self._value

    def setLabelText(self, text):
        self._text = str(text or "")
        self._apply_label(self._text)

    def labelText(self):
        return self._text

    def wasCanceled(self):
        try:
            return bool(self._outer.wasCanceled())
        except Exception:
            return False

    # The outer dialog owns visibility and lifetime.
    def show(self):
        return None

    def close(self):
        return None

    def reset(self):
        return None

    def setWindowTitle(self, _title):
        return None

    def setWindowModality(self, _modality):
        return None

    def setMinimumDuration(self, _duration):
        return None

    def setAutoClose(self, _enabled):
        return None

    def setAutoReset(self, _enabled):
        return None

    def resize(self, *_args):
        return None


_c65_previous_make_progress = MultiPaneLogViewer.make_progress


def _c65_make_progress(self, title, text):
    outer = getattr(self, "_c65_shared_progress", None)
    if outer is not None:
        return _C65UnifiedProgressProxy(self, outer, title, text)
    return _c65_previous_make_progress(self, title, text)


_c65_previous_prepare_viewer_before_show = _c63_prepare_viewer_before_show


def _c65_prepare_viewer_before_show(main_window):
    viewer = getattr(main_window, "viewer_window", None)
    if viewer is None:
        viewer = DualLogViewer(main_window)
        main_window.viewer_window = viewer

    viewer.hide()
    viewer.setEnabled(False)
    viewer.setUpdatesEnabled(False)

    progress = SoftProgressDialog(
        "Preparing Log Viewer",
        "Building viewer source list...",
        None,
        0,
        100,
        main_window,
    )
    progress.setWindowModality(Qt.ApplicationModal)
    progress.setValue(2)
    progress.show()
    QApplication.processEvents()

    viewer._c65_shared_progress = progress
    try:
        if hasattr(viewer, "refresh_available_sources"):
            viewer.refresh_available_sources()

        visible = list(viewer.visible_indices()) if hasattr(viewer, "visible_indices") else [0, 1]
        if not visible:
            visible = [0]

        total = max(1, len(visible))
        for order, pane_index in enumerate(visible, start=1):
            pane_name = viewer.pane_name(pane_index) if hasattr(viewer, "pane_name") else f"Pane {pane_index + 1}"
            source_name = ""
            if pane_index < len(getattr(viewer, "sources", [])):
                source_name = viewer.sources[pane_index].currentText()

            segment_low = 4 + int((order - 1) * 88 / total)
            segment_high = 4 + int(order * 88 / total)
            viewer._c65_progress_low = segment_low
            viewer._c65_progress_high = segment_high
            viewer._c65_progress_prefix = (
                f"Preparing {pane_name}: {source_name or 'selected log'}"
            )
            progress.setLabelText(
                f"{viewer._c65_progress_prefix}\nReading and indexing selected rows..."
            )
            progress.setValue(segment_low)
            QApplication.processEvents()
            viewer.load_pane(pane_index)

        progress.setLabelText("Finalizing Viewer layout, filters, and columns...")
        progress.setValue(95)
        QApplication.processEvents()
        viewer.updateGeometry()
        if viewer.width() <= 0 or viewer.height() <= 0:
            viewer.adjustSize()
        QApplication.processEvents()
        progress.setLabelText("Log Viewer is ready.")
        progress.setValue(100)
        QApplication.processEvents()
    finally:
        for attr in (
            "_c65_shared_progress",
            "_c65_progress_low",
            "_c65_progress_high",
            "_c65_progress_prefix",
        ):
            try:
                delattr(viewer, attr)
            except Exception:
                pass
        viewer.setUpdatesEnabled(True)
        viewer.setEnabled(True)
        progress.close()

    return viewer


# Redirect Commit0063/0064's deferred presentation to the unified implementation.
_c63_prepare_viewer_before_show = _c65_prepare_viewer_before_show
MultiPaneLogViewer.make_progress = _c65_make_progress


# ---------------------------------------------------------------------------
# Commit0066 - Viewer Shell First / Explicit LOAD LOGS
# ---------------------------------------------------------------------------
# Smart File Discovery selects the candidate files only.  Opening the Viewer
# must not parse WS, PSC, or any other log automatically.  The operator first
# chooses the number of panes and each source, then explicitly presses
# LOAD LOGS (or Load This) to begin parsing and table construction.


def _c66_reset_viewer_shell(viewer):
    """Return an existing Viewer to a clean, immediately interactive shell."""
    try:
        viewer.setEnabled(True)
        viewer.setUpdatesEnabled(True)
        for idx in range(getattr(viewer, "MAX_PANES", 4)):
            if idx < len(getattr(viewer, "all_rows", [])):
                viewer.all_rows[idx] = []
            if idx < len(getattr(viewer, "ts_indexes", [])):
                viewer.ts_indexes[idx] = []
            if idx < len(getattr(viewer, "models", [])):
                cols = ["Timestamp", "Message", "Level"]
                viewer.models[idx].set_rows([], cols)
            if idx < len(getattr(viewer, "file_labels", [])):
                viewer.file_labels[idx].setText("File: not loaded")
        if hasattr(viewer, "detail"):
            viewer.detail.clear()
        if hasattr(viewer, "status"):
            viewer.status.setText(
                "Ready — choose View mode and log sources, then press LOAD LOGS."
            )
        if hasattr(viewer, "_sync_aliases"):
            viewer._sync_aliases()
    except Exception:
        write_startup_log("Commit0066 viewer shell reset failed.\n\n" + traceback.format_exc())


def _c66_open_dual_viewer(self):
    """Show an interactive empty Viewer without loading any log data."""
    if getattr(self, "_c66_viewer_open_in_progress", False):
        return
    self._c66_viewer_open_in_progress = True
    try:
        viewer = getattr(self, "viewer_window", None)
        if viewer is None:
            viewer = DualLogViewer(self)
            self.viewer_window = viewer
        if hasattr(viewer, "refresh_available_sources"):
            viewer.refresh_available_sources()
        if getattr(self, "_c66_new_discovery_session", False):
            _c66_reset_viewer_shell(viewer)
            self._c66_new_discovery_session = False

        geo = QApplication.primaryScreen().availableGeometry()
        w = min(max(980, int(geo.width() * 0.72)), geo.width() - 60)
        h = min(max(620, int(geo.height() * 0.72)), geo.height() - 60)
        viewer.resize(w, h)
        viewer.move(
            geo.x() + max(20, (geo.width() - w) // 2),
            geo.y() + max(20, (geo.height() - h) // 2),
        )
        viewer.show()
        viewer.raise_()
        viewer.activateWindow()
        QApplication.processEvents()
    except Exception:
        text = "Log Viewer opening failed.\n\n" + traceback.format_exc()
        write_startup_log(text)
        QMessageBox.critical(self, APP_TITLE, text)
    finally:
        self._c63_prepare_before_show = False
        self._c66_viewer_open_in_progress = False


def _c66_start_import_clicked(self):
    """Run discovery, then open only the configurable Viewer shell."""
    opt = self.collect_options()
    if not opt.source_folder:
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder is required.")
        return
    if not Path(opt.source_folder).exists():
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder does not exist.")
        return
    if (
        opt.use_start and opt.use_end and opt.start_date and opt.end_date
        and opt.start_date > opt.end_date
    ):
        QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
        return

    dlg = SmartFileDiscoveryDialog(self, opt)
    if dlg.exec() != QDialog.Accepted:
        self.status_label.setText("Cancelled before opening Log Viewer")
        return

    self.viewer_selected_files = list(dlg.selected_files)
    self.viewer_selected_types = set(dlg.selected_types)
    self.save_settings()
    self.log_view.clear()
    self.log(
        f"Smart File Discovery selected {len(self.viewer_selected_files)} files. "
        "Log data is not loaded yet. Configure the Viewer and press LOAD LOGS."
    )
    self.status_label.setText("Viewer ready for configuration — press LOAD LOGS in Viewer")
    self._c66_new_discovery_session = True
    _c66_open_dual_viewer(self)


MainWindow.open_dual_viewer = _c66_open_dual_viewer
MainWindow.start_import_clicked = _c66_start_import_clicked


# ---------------------------------------------------------------------------
# Commit0067 - ZIP START discovery handoff fix
# ---------------------------------------------------------------------------
# A ZIP selected or dropped on the main window must be extracted before Smart
# File Discovery scans it.  Keep the extraction alive after discovery because
# LOAD LOGS intentionally happens later in the Viewer.


def _c67_cleanup_discovery_zip(self):
    holder = getattr(self, "_c67_zip_discovery_holder", None)
    if holder is not None:
        try:
            holder.cleanup()
        except Exception:
            pass
    self._c67_zip_discovery_holder = None
    self._c67_zip_discovery_root = None
    self._c67_zip_source_path = None


def _c67_prepare_discovery_options(self, opt):
    source = Path(opt.source_folder)
    if not (source.is_file() and source.suffix.lower() == ".zip"):
        _c67_cleanup_discovery_zip(self)
        return opt

    _c67_cleanup_discovery_zip(self)
    holder, extract_root, nested_zips = _safe_extract_zip_for_discovery(source)
    self._c67_zip_discovery_holder = holder
    self._c67_zip_discovery_root = str(extract_root)
    self._c67_zip_source_path = str(source)

    opt.source_folder = str(extract_root)
    opt.recursive = True
    opt.include_unknown = False
    try:
        self.chk_unknown.setChecked(False)
    except Exception:
        pass

    if nested_zips:
        QMessageBox.information(
            self,
            "ZIP files inside archive",
            "Nested ZIP files were not expanded. File names only:\n\n"
            + "\n".join(nested_zips),
        )
    return opt


def _c67_start_import_clicked(self):
    """Discover a folder or extracted ZIP once, then open the empty Viewer."""
    opt = self.collect_options()
    if not opt.source_folder:
        QMessageBox.warning(self, APP_TITLE, "Source Log Folder or ZIP is required.")
        return

    source = Path(opt.source_folder)
    if not source.exists():
        QMessageBox.warning(self, APP_TITLE, "Selected source does not exist.")
        return
    if source.is_file() and source.suffix.lower() != ".zip":
        QMessageBox.warning(
            self,
            APP_TITLE,
            "START accepts a folder or ZIP archive. Use file staging for ordinary files.",
        )
        return
    if (
        opt.use_start and opt.use_end and opt.start_date and opt.end_date
        and opt.start_date > opt.end_date
    ):
        QMessageBox.warning(self, APP_TITLE, "Start Date is later than End Date.")
        return

    try:
        opt = _c67_prepare_discovery_options(self, opt)
        dlg = SmartFileDiscoveryDialog(self, opt)
        if dlg.exec() != QDialog.Accepted:
            self.status_label.setText("Cancelled before opening Log Viewer")
            # A cancelled ZIP session has no future LOAD LOGS dependency.
            if source.is_file() and source.suffix.lower() == ".zip":
                _c67_cleanup_discovery_zip(self)
            return

        self.viewer_selected_files = list(dlg.selected_files)
        self.viewer_selected_types = set(dlg.selected_types)
        self.save_settings()
        self.log_view.clear()
        source_label = source.name if source.is_file() else source.name or str(source)
        self.log(
            f"Smart File Discovery selected {len(self.viewer_selected_files)} files "
            f"from {source_label}. Log data is not loaded yet. "
            "Configure the Viewer and press LOAD LOGS."
        )
        self.status_label.setText(
            "Viewer ready for configuration — press LOAD LOGS in Viewer"
        )
        self._c66_new_discovery_session = True
        _c66_open_dual_viewer(self)
    except Exception:
        if source.is_file() and source.suffix.lower() == ".zip":
            _c67_cleanup_discovery_zip(self)
        text = "Smart File Discovery failed.\n\n" + traceback.format_exc()
        write_startup_log(text)
        QMessageBox.critical(self, APP_TITLE, text)


_old_c67_close_event = getattr(MainWindow, "closeEvent", None)
def _c67_close_event(self, event):
    _c67_cleanup_discovery_zip(self)
    if _old_c67_close_event is not None:
        return _old_c67_close_event(self, event)
    event.accept()


MainWindow.start_import_clicked = _c67_start_import_clicked
MainWindow.closeEvent = _c67_close_event


# ---------------------------------------------------------------------------
# Commit0068: contextual button help, startup guide/tour, and safe exit prompt
# ---------------------------------------------------------------------------
APP_VERSION = "2.0.0-rc1-commit0069"

_C68_GUIDE_ASK_KEY = "help/ask_guide_on_startup"
_C68_GUIDE_NEXT_KEY = "help/show_guide_next_startup"

_C68_BUTTON_HELP = {
    "▶ START": "Scan the selected folder or ZIP, confirm detected files, then open Log Viewer for configuration.",
    "MERGE": "Scan selected logs and create the merged output. PSC and Review are viewer-only and are skipped from merge output.",
    "Import": "Import selected log files for viewing without creating a merged output.",
    "Log Viewer": "Open Log Viewer and configure visible panes, sources, filters, and time range.",
    "Update...": "Install a supported tool update ZIP or file-type plugin ZIP.",
    "Pause": "Temporarily pause the current merge operation.",
    "Cancel": "Cancel the current operation or close this dialog without applying changes.",
    "Browse...": "Choose a source folder, ZIP archive, file, or output folder for this field.",
    "Same as Source": "Use the selected source location as the output folder.",
    "All": "Select all items in this section.",
    "Clear": "Clear all selections in this section.",
    "All Files": "Select every detected file in the current discovery result.",
    "Clear Files": "Clear all detected-file selections.",
    "Manage Rules": "Open Smart Noise rules to review, enable, disable, or remove rules.",
    "Export Project": "Save the current project settings and selections for later reuse.",
    "Import Project": "Load a previously exported project package.",
    "Split Merge": "Split an existing merged result back into reconstructed per-source log files.",
    "Open Output": "Open the current output folder in File Explorer.",
    "Reset Defaults": "Reset folders, filters, site information, and log-type selections to defaults.",
    "LOAD LOGS": "Load the configured visible panes after selecting their log sources.",
    "Load This": "Load only this pane using its currently selected source.",
    "Find": "Jump to the next row matching the search text in this pane.",
    "Columns...": "Choose which columns are visible in this log pane.",
    "Apply Rules Now": "Reload the viewer and apply the currently enabled Smart Noise rules.",
    "Apply to Focused": "Apply the current time range to the focused log pane.",
    "Apply to Visible": "Apply the current time range to all visible log panes.",
    "Clear Time Range": "Remove the viewer time filter and show all loaded rows.",
    "Set Start from Selected": "Copy the selected row timestamp into the Start field.",
    "Set End from Selected": "Copy the selected row timestamp into the End field.",
    "Close": "Close this window.",
}


def _c68_apply_button_help(root):
    """Add concise hover popups without overwriting existing detailed help."""
    try:
        for button in root.findChildren(QPushButton):
            if button.toolTip().strip():
                continue
            text = button.text().replace("&", "").strip()
            help_text = _C68_BUTTON_HELP.get(text)
            if help_text is None:
                lowered = text.lower()
                if lowered.startswith("load"):
                    help_text = "Load the selected data for this section."
                elif lowered.startswith("open"):
                    help_text = "Open the selected view, folder, or management window."
                elif lowered.startswith("select"):
                    help_text = "Select the related items."
                elif lowered.startswith("export"):
                    help_text = "Export the selected data to a file."
                elif lowered.startswith("apply"):
                    help_text = "Apply the current settings to this section."
                elif lowered.startswith("reset"):
                    help_text = "Restore this section to its default state."
            if help_text:
                button.setToolTip(help_text)
                button.setStatusTip(help_text)
    except Exception:
        write_startup_log("Commit0068 button help failed.\n\n" + traceback.format_exc())


def _c68_show_guide(self):
    dlg = QDialog(self)
    dlg.setWindowTitle("Log Merge Tool — Quick Guide")
    dlg.setModal(True)
    dlg.resize(680, 520)
    layout = QVBoxLayout(dlg)
    title = QLabel("Quick Guide")
    title.setStyleSheet("font-size: 20pt; font-weight: 700;")
    layout.addWidget(title)
    guide = QLabel(
        "<b>1. Select Source</b><br>Drag and drop a log file, folder, or ZIP archive onto the Source area, or use <b>Browse...</b>.<br><br>"
        "<b>2. Choose log types and date range</b><br>Select only the data needed for the investigation.<br><br>"
        "<b>3. START</b><br>Run Smart File Discovery, confirm the detected files, then open Log Viewer.<br><br>"
        "<b>4. Viewer</b><br>Select pane sources, visible panes, columns, searches, and an optional time range, then press <b>LOAD LOGS</b>.<br><br>"
        "<b>Tip</b><br>Move the mouse over a button to see a short explanation popup."
    )
    guide.setWordWrap(True)
    guide.setTextInteractionFlags(Qt.TextSelectableByMouse)
    layout.addWidget(guide, 1)
    buttons = QDialogButtonBox(QDialogButtonBox.Close)
    buttons.rejected.connect(dlg.reject)
    layout.addWidget(buttons)
    dlg.exec()


def _c68_show_tour(self):
    steps = [
        ("1 / 4 — Source", "Drag and drop a log file, folder, or ZIP archive onto the Source area, or use Browse...."),
        ("2 / 4 — Filters", "Choose log types, date range, system serial, and site information before scanning."),
        ("3 / 4 — START", "START performs Smart File Discovery, lets you confirm the detected files, and opens Log Viewer."),
        ("4 / 4 — Log Viewer", "Configure panes and filters first, then press LOAD LOGS. Hover over buttons at any time for concise help."),
    ]
    for title, text in steps:
        box = QMessageBox(self)
        box.setWindowTitle("Guided Tour")
        box.setIcon(QMessageBox.Information)
        box.setText(f"<b>{title}</b>")
        box.setInformativeText(text)
        box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        box.button(QMessageBox.Ok).setText("Next" if title != steps[-1][0] else "Finish")
        box.button(QMessageBox.Cancel).setText("Stop Tour")
        if box.exec() != QMessageBox.Ok:
            break


def _c68_show_guide_and_tour(self):
    _c68_show_guide(self)
    _c68_show_tour(self)


def _c68_startup_help_flow(self):
    try:
        settings = self.settings if hasattr(self, "settings") else QSettings("LogMerge", "NoExcel")
        show_next = settings.value(_C68_GUIDE_NEXT_KEY, False, type=bool)
        if show_next:
            settings.setValue(_C68_GUIDE_NEXT_KEY, False)
            settings.sync()
            _c68_show_guide_and_tour(self)
            return
        ask = settings.value(_C68_GUIDE_ASK_KEY, True, type=bool)
        if not ask:
            return
        box = QMessageBox(self)
        box.setWindowTitle("Guide and Guided Tour")
        box.setIcon(QMessageBox.Question)
        box.setText("Do you need the quick guide and guided tour?")
        box.setInformativeText("Choose No to stop this question from appearing at future startups. You can enable a one-time guide again when closing the application.")
        yes_btn = box.addButton("Yes — Show Guide", QMessageBox.YesRole)
        no_btn = box.addButton("No — Do Not Ask Again", QMessageBox.NoRole)
        box.setDefaultButton(yes_btn)
        box.exec()
        if box.clickedButton() is yes_btn:
            _c68_show_guide_and_tour(self)
        elif box.clickedButton() is no_btn:
            settings.setValue(_C68_GUIDE_ASK_KEY, False)
            settings.sync()
    except Exception:
        write_startup_log("Commit0068 startup guide flow failed.\n\n" + traceback.format_exc())


_old_c68_main_init = MainWindow.__init__
def _c68_main_init(self):
    _old_c68_main_init(self)
    self.setWindowTitle(f"{APP_TITLE} {APP_VERSION}")
    _c68_apply_button_help(self)
    if not RELEASE_MODE:
        QTimer.singleShot(250, lambda: _c68_startup_help_flow(self))

MainWindow.__init__ = _c68_main_init
if not RELEASE_MODE:
    MainWindow.show_quick_guide = _c68_show_guide
    MainWindow.show_guided_tour = _c68_show_tour



def _c68_close_event(self, event):
    try:
        box = QMessageBox(self)
        box.setWindowTitle("Exit Log Merge Tool")
        box.setIcon(QMessageBox.Question)
        box.setText("Are you sure you want to close Log Merge Tool?")
        box.setInformativeText("Unsaved selections or an active viewer session may be lost.")
        show_next_chk = QCheckBox("Show the guide and guided tour at the next startup")
        show_next_chk.setChecked(False)
        box.setCheckBox(show_next_chk)
        exit_btn = box.addButton("Exit", QMessageBox.AcceptRole)
        cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
        box.setDefaultButton(cancel_btn)
        box.exec()
        if box.clickedButton() is not exit_btn:
            event.ignore()
            return
        settings = self.settings if hasattr(self, "settings") else QSettings("LogMerge", "NoExcel")
        if show_next_chk.isChecked():
            settings.setValue(_C68_GUIDE_NEXT_KEY, True)
        settings.sync()
        _c67_cleanup_discovery_zip(self)
        event.accept()
    except Exception:
        write_startup_log("Commit0068 close confirmation failed.\n\n" + traceback.format_exc())
        event.ignore()

if not RELEASE_MODE:
    MainWindow.closeEvent = _c68_close_event



# ---------------------------------------------------------------------------
# Commit0069: current workflow guide and daily source/output reset
# ---------------------------------------------------------------------------
_C69_LAST_LAUNCH_DATE_KEY = "session/last_launch_date"

_old_c69_main_init = MainWindow.__init__
def _c69_main_init(self):
    _old_c69_main_init(self)
    try:
        settings = self.settings if hasattr(self, "settings") else QSettings("LogMerge", "NoExcel")
        today = QDate.currentDate().toString("yyyy-MM-dd")
        last_launch_date = str(settings.value(_C69_LAST_LAUNCH_DATE_KEY, "") or "")
        if last_launch_date != today:
            # The first launch of each calendar day always starts with clean paths.
            self.source_edit.clear()
            self.output_edit.clear()
            settings.setValue("source_folder", "")
            settings.setValue("output_folder", "")
            settings.setValue(_C69_LAST_LAUNCH_DATE_KEY, today)
            settings.sync()
    except Exception:
        write_startup_log("Commit0069 daily folder reset failed.\n\n" + traceback.format_exc())

MainWindow.__init__ = _c69_main_init



def _rc79_autoload_handoff_viewer(self, ordered_types):
    """Populate and parse handoff logs without requiring the LOAD LOGS button."""
    try:
        viewer = getattr(self, "viewer_window", None)
        if viewer is None:
            return
        viewer.refresh_available_sources()
        types = [value for value in ordered_types if value]
        if not types:
            return
        # Hub handoff always starts in exactly 2 Show. Prefer WS + CSA, then
        # fill missing positions from the remaining detected log types.
        preferred = []
        for wanted in ("WS", "WaterSystem", "CSA"):
            if wanted in types and wanted not in preferred:
                preferred.append(wanted)
        for value in types:
            if value not in preferred:
                preferred.append(value)
        types = preferred[:2]
        visible_count = min(2, len(types), getattr(viewer, "MAX_PANES", 4))
        for index, checkbox in enumerate(getattr(viewer, "pane_visible_checks", [])):
            checkbox.setChecked(index < visible_count)
        if hasattr(viewer, "update_view_mode"):
            viewer.update_view_mode()
        for index, source_type in enumerate(types[:visible_count]):
            combo = viewer.sources[index]
            if combo.findText(source_type) >= 0:
                combo.setCurrentText(source_type)
            viewer.load_pane(index)
        if hasattr(viewer, "status"):
            viewer.status.setText(f"Hub handoff loaded {len(self.viewer_selected_files)} file(s) into {visible_count} pane(s).")
        self.status_label.setText(f"Hub handoff imported: {len(self.viewer_selected_files)} file(s)")
    except Exception:
        write_startup_log("RC7.9 handoff viewer auto-load failed.\n\n" + traceback.format_exc())

# RC7.2 native Service Hub handoff receiver
def _rc72_apply_handoff(self):
    try:
        handoff = _RC72_HANDOFF
        if not handoff or not handoff.auto_load:
            return
        files = [str(path) for path in handoff.input_paths() if path.is_file()]
        workspace = handoff.workspace()
        if not files and workspace:
            files = [str(path) for path in workspace.rglob("*") if path.is_file()]
        if workspace:
            self.source_edit.setText(str(workspace))
        elif files:
            self.source_edit.setText(str(Path(files[0]).parent))
        self.viewer_selected_files = files
        found = set()
        for value in files:
            path = Path(value)
            if is_review_file(path.name):
                found.add("Review")
            else:
                kind = classify_file(path, True)
                if kind:
                    found.add(kind)
        self.viewer_selected_types = found
        ordered = [name for name in ["WS", "WaterSystem", "CSA", "CGA", "MRSERVER", "GESYS", "LAIS", "PSC", "Review", "VIMeasure", "ACQUISITION"] if name in found]
        handoff.mark("log_explorer", "accepted", input_count=len(files), detected_types=sorted(found))
        self.status_label.setText(f"Hub handoff received: {len(files)} file(s); opening Viewer...")
        if files:
            QTimer.singleShot(0, self.open_dual_viewer)
            QTimer.singleShot(500, lambda values=list(ordered): _rc79_autoload_handoff_viewer(self, values))
    except Exception:
        write_startup_log("RC7.2 handoff failed.\n\n" + traceback.format_exc())

_old_rc72_main_init = MainWindow.__init__
def _rc72_main_init(self):
    _old_rc72_main_init(self)
    QTimer.singleShot(0, lambda: _rc72_apply_handoff(self))
MainWindow.__init__ = _rc72_main_init



# ---------------------------------------------------------------------------
# Commit0070: WS error/code/True quick filter + row bookmarks
# ---------------------------------------------------------------------------
_C70_WS_MODE = "WS_ERR_CODE_TRUE"


def _c70_is_ws_source_name(value):
    return str(value or "").strip().upper() in {"WS", "WATERSYSTEM"}


def _c70_ws_error_code_true(row):
    level = _c60_normalize(row.get("Level", ""))
    message = _c60_normalize(row.get("Message", ""))
    return level in {"err", "error"} and "code" in message and "true" in message


def _c70_bookmark_key(row):
    return (
        str(row.get("SourceType", "") or ""),
        str(row.get("File", "") or ""),
        str(row.get("Line", "") or ""),
        str(row.get("Timestamp", "") or ""),
        str(row.get("Message", "") or ""),
    )


def _c70_bookmark_state(viewer):
    state = getattr(viewer, "_c70_bookmarks", None)
    if not isinstance(state, set):
        state = set()
        viewer._c70_bookmarks = state
    return state


def _c70_sync_bookmark_marks(viewer, pane_index):
    bookmarks = _c70_bookmark_state(viewer)
    for row in viewer.all_rows[pane_index] or []:
        row["_bookmarked"] = _c70_bookmark_key(row) in bookmarks


def _c70_quick_mode_label(mode):
    if mode == _C70_WS_MODE:
        return "WS Err + code + True"
    return "All" if mode == "ALL" else str(mode).title()


def _c70_apply_view_filters(viewer, side):
    pane_index = viewer.side_index(side)
    expressions, modes = _c61_filter_state(viewer)
    base_rows = list(viewer.all_rows[pane_index] or [])
    expression = expressions[pane_index]
    quick_mode = str(modes[pane_index] or "ALL").upper()

    try:
        start, end = viewer.current_viewer_time_range()
    except Exception:
        start, end = None, None

    _c70_sync_bookmark_marks(viewer, pane_index)
    rows = []
    for row in base_rows:
        timestamp = row.get("_ts")
        if start or end:
            if not isinstance(timestamp, datetime):
                continue
            if start and timestamp < start:
                continue
            if end and timestamp > end:
                continue
        if quick_mode == _C70_WS_MODE:
            if not _c70_ws_error_code_true(row):
                continue
        elif quick_mode != "ALL" and _f52a_row_severity(row) != quick_mode:
            continue
        if not _c61_expression_matches(row, expression):
            continue
        rows.append(row)

    columns = viewer.pane_columns_for_rows(pane_index, rows or base_rows)
    viewer.models[pane_index].set_rows(rows, columns)
    viewer.apply_table_column_widths(pane_index)
    viewer.ts_indexes[pane_index] = sorted(
        (row["_ts"], index)
        for index, row in enumerate(rows)
        if isinstance(row.get("_ts"), datetime)
    )
    viewer._sync_aliases()

    groups = getattr(viewer, "quick_filter_buttons", [])
    if 0 <= pane_index < len(groups):
        counts = _c23_quick_filter_counts(base_rows)
        counts[_C70_WS_MODE] = sum(1 for row in base_rows if _c70_ws_error_code_true(row))
        for key, button in groups[pane_index].items():
            button.blockSignals(True)
            button.setChecked(key == quick_mode)
            button.setText(f"{_c70_quick_mode_label(key)} ({counts.get(key, 0):,})")
            button.blockSignals(False)

    details = []
    if quick_mode != "ALL":
        details.append(f"Quick Filter={_c70_quick_mode_label(quick_mode)}")
    if expression:
        details.append(expression)
    status_text = ", ".join(details) or "No filter"
    message = (
        f"{viewer.pane_name(pane_index)}: {status_text} — "
        f"Displayed {len(rows):,}/{len(base_rows):,} rows"
    )
    if hasattr(viewer, "status"):
        viewer.status.setText(message)
    viewer.log(message)
    if 0 <= pane_index < len(viewer.file_labels):
        label = viewer.file_labels[pane_index]
        base = label.text().split(" • ")[0]
        label.setText(f"{base} • Displayed {len(rows):,}/{len(base_rows):,}")
    viewer.tables[pane_index].viewport().update()


def _c70_set_quick_filter(viewer, pane_index, mode):
    _, modes = _c61_filter_state(viewer)
    modes[pane_index] = str(mode or "ALL").upper()
    _c70_apply_view_filters(viewer, pane_index)


def _c70_toggle_selected_bookmark(viewer, pane_index):
    table = viewer.tables[pane_index]
    index = table.currentIndex()
    if not index.isValid():
        QMessageBox.information(viewer, "Bookmark", "Select a row first.")
        return
    row = viewer.models[pane_index].row_at(index.row()) or {}
    key = _c70_bookmark_key(row)
    bookmarks = _c70_bookmark_state(viewer)
    if key in bookmarks:
        bookmarks.remove(key)
        action = "removed"
    else:
        bookmarks.add(key)
        action = "added"
    for p in range(len(viewer.panes)):
        _c70_sync_bookmark_marks(viewer, p)
        viewer.models[p].dataChanged.emit(
            viewer.models[p].index(0, 0),
            viewer.models[p].index(max(0, viewer.models[p].rowCount()-1), max(0, viewer.models[p].columnCount()-1)),
            [Qt.BackgroundRole, Qt.FontRole],
        )
    if hasattr(viewer, "status"):
        viewer.status.setText(f"Bookmark {action}: {row.get('Timestamp', '')} {row.get('Message', '')}")


def _c70_jump_bookmark(viewer, pane_index, direction=1):
    model = viewer.models[pane_index]
    bookmarked_rows = [i for i, row in enumerate(model.rows) if row.get("_bookmarked")]
    if not bookmarked_rows:
        QMessageBox.information(viewer, "Bookmark", "No bookmarked rows are visible in this pane.")
        return
    current = viewer.tables[pane_index].currentIndex().row()
    if direction >= 0:
        target = next((i for i in bookmarked_rows if i > current), bookmarked_rows[0])
    else:
        target = next((i for i in reversed(bookmarked_rows) if i < current), bookmarked_rows[-1])
    while model.rowCount() <= target and model.canFetchMore():
        model.fetchMore()
    idx = model.index(target, 0)
    viewer.tables[pane_index].setCurrentIndex(idx)
    viewer.tables[pane_index].scrollTo(idx, QAbstractItemView.PositionAtCenter)


def _c70_row_menu(viewer, pane_index, position):
    table = viewer.tables[pane_index]
    model = viewer.models[pane_index]
    index = table.indexAt(position)
    if not index.isValid():
        return
    table.setCurrentIndex(index)
    row = model.row_at(index.row()) or {}
    column_name = model.columns[index.column()] if 0 <= index.column() < len(model.columns) else ""
    cell_value = model.data(index, Qt.DisplayRole)

    menu = QMenu(table)
    bookmark = menu.addAction("Remove Bookmark" if row.get("_bookmarked") else "Add Bookmark to This Row")
    next_bookmark = menu.addAction("Next Bookmark")
    previous_bookmark = menu.addAction("Previous Bookmark")
    menu.addSeparator()
    copy_cell = menu.addAction("Copy Cell")
    copy_row = menu.addAction("Copy Row")
    copy_timestamp = menu.addAction("Copy Timestamp")
    copy_message = menu.addAction("Copy Message / Value")
    menu.addSeparator()
    contains = menu.addAction(f"Filter {column_name} contains this value")
    exact = menu.addAction(f"Filter {column_name} equals this value")
    clear = menu.addAction("Clear Pane Filter")
    menu.addSeparator()
    copy_rule = menu.addAction("Copy Rule Text")
    noise = menu.addAction("Add to Noise Rule")

    selected = menu.exec(table.viewport().mapToGlobal(position))
    clipboard = QApplication.clipboard()
    if selected == bookmark:
        _c70_toggle_selected_bookmark(viewer, pane_index)
    elif selected == next_bookmark:
        _c70_jump_bookmark(viewer, pane_index, 1)
    elif selected == previous_bookmark:
        _c70_jump_bookmark(viewer, pane_index, -1)
    elif selected == copy_cell:
        clipboard.setText(str(cell_value or ""))
    elif selected == copy_row:
        clipboard.setText("\t".join(str(row.get(c, "")) for c in model.columns))
    elif selected == copy_timestamp:
        clipboard.setText(str(row.get("Timestamp", "") or ""))
    elif selected == copy_message:
        clipboard.setText(str(row.get("Message", row.get("Value", "")) or ""))
    elif selected == contains and column_name:
        _c61_set_expression(viewer, pane_index, f"{column_name}~{cell_value}")
    elif selected == exact and column_name:
        _c61_set_expression(viewer, pane_index, f"{column_name}={cell_value}")
    elif selected == clear:
        _c61_set_expression(viewer, pane_index, "")
    elif selected == copy_rule:
        viewer.copy_rule_text(pane_index)
    elif selected == noise:
        viewer.approve_selected_noise(pane_index)


_old_c70_model_data = LogTableModel.data
def _c70_model_data(self, index, role=Qt.DisplayRole):
    if index.isValid() and index.row() < self._visible_count:
        row = self.rows[index.row()]
        if row.get("_bookmarked"):
            if role == Qt.BackgroundRole:
                return QBrush(QColor(255, 244, 170))
            if role == Qt.FontRole:
                from PySide6.QtGui import QFont
                font = QFont()
                font.setBold(True)
                return font
    return _old_c70_model_data(self, index, role)
LogTableModel.data = _c70_model_data


_old_c70_build_ui = MultiPaneLogViewer.build_ui
def _c70_build_ui(self):
    _old_c70_build_ui(self)
    _c70_bookmark_state(self)
    for pane_index, pane in enumerate(self.panes):
        groups = self.quick_filter_buttons[pane_index]
        if _C70_WS_MODE not in groups:
            button = QPushButton("WS Err + code + True")
            button.setCheckable(True)
            button.setAutoExclusive(True)
            button.setToolTip("WS only: Level is Err/Error and Message contains both 'code' and 'True'.")
            # Find the quick-filter bar inserted by Commit0023.
            bar_widget = pane.layout().itemAt(2).widget()
            bar_layout = bar_widget.layout() if bar_widget else None
            if bar_layout:
                bar_layout.insertWidget(max(1, bar_layout.count()-1), button)
            groups[_C70_WS_MODE] = button
        for mode, button in groups.items():
            try:
                button.clicked.disconnect()
            except (TypeError, RuntimeError):
                pass
            button.clicked.connect(lambda checked=False, p=pane_index, m=mode: _c70_set_quick_filter(self, p, m))
        table = self.tables[pane_index]
        try:
            table.customContextMenuRequested.disconnect()
        except (TypeError, RuntimeError):
            pass
        table.customContextMenuRequested.connect(lambda pos, p=pane_index: _c70_row_menu(self, p, pos))
MultiPaneLogViewer.build_ui = _c70_build_ui


_old_c70_load_pane = MultiPaneLogViewer.load_pane
def _c70_load_pane(self, side):
    pane_index = self.side_index(side)
    result = _old_c70_load_pane(self, side)
    _, modes = _c61_filter_state(self)
    source_name = self.sources[pane_index].currentText()
    modes[pane_index] = _C70_WS_MODE if _c70_is_ws_source_name(source_name) else "ALL"
    _c70_apply_view_filters(self, pane_index)
    return result

MultiPaneLogViewer.apply_view_filters = _c70_apply_view_filters
MultiPaneLogViewer.load_pane = _c70_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)
MultiPaneLogViewer.show_row_context_menu = _c70_row_menu


# ---------------------------------------------------------------------------
# Commit0071: Default Viewer layout = 2 Show (WS + CSA)
# ---------------------------------------------------------------------------
# Keep the existing explicit LOAD LOGS workflow.  Only the initial viewer
# configuration is changed: exactly two panes are visible, with WS on the left
# and CSA on the right.  Panes 3 and 4 remain available when the user changes
# the view mode or Show checkboxes.
_old_c71_build_ui = MultiPaneLogViewer.build_ui

def _c71_build_ui(self):
    _old_c71_build_ui(self)

    # Default layout: two visible panes only.
    try:
        self.mode_combo.blockSignals(True)
        self.mode_combo.setCurrentText("2 logs (Dual)")
        self.mode_combo.blockSignals(False)
    except Exception:
        pass

    for index, checkbox in enumerate(getattr(self, "pane_visible_checks", [])):
        checkbox.blockSignals(True)
        checkbox.setChecked(index < 2)
        checkbox.blockSignals(False)

    # Default sources: WS on the left and CSA on the right.
    default_sources = ("WS", "CSA")
    for index, source_name in enumerate(default_sources):
        if index >= len(getattr(self, "sources", [])):
            break
        combo = self.sources[index]
        found = combo.findText(source_name, Qt.MatchFixedString)
        if found >= 0:
            combo.setCurrentIndex(found)
        else:
            combo.setCurrentText(source_name)

    # Apply the final visibility/splitter state after all controls are set.
    self.update_view_mode()
    self.status.setText("Ready — default viewer: 2 Show (WS + CSA). Press LOAD LOGS to load data.")

MultiPaneLogViewer.build_ui = _c71_build_ui


# ---------------------------------------------------------------------------
# Commit0072: Hub 2-Show enforcement + WS-only Err/code/True Quick Filter
# ---------------------------------------------------------------------------
def _c72_sync_ws_quick_filter_button(viewer, pane_index):
    groups = getattr(viewer, "quick_filter_buttons", [])
    if not (0 <= pane_index < len(groups)):
        return
    button = groups[pane_index].get(_C70_WS_MODE)
    if button is None:
        return
    source_name = viewer.sources[pane_index].currentText() if pane_index < len(viewer.sources) else ""
    is_ws = _c70_is_ws_source_name(source_name)
    button.setVisible(is_ws)
    button.setEnabled(is_ws)
    if not is_ws:
        _, modes = _c61_filter_state(viewer)
        if modes[pane_index] == _C70_WS_MODE:
            modes[pane_index] = "ALL"

_old_c72_build_ui = MultiPaneLogViewer.build_ui
def _c72_build_ui(self):
    _old_c72_build_ui(self)
    for pane_index, combo in enumerate(self.sources):
        try:
            combo.currentTextChanged.disconnect(_c72_source_changed)
        except (TypeError, RuntimeError):
            pass
        combo.currentTextChanged.connect(lambda _text, p=pane_index: _c72_source_changed(self, p))
        _c72_sync_ws_quick_filter_button(self, pane_index)
MultiPaneLogViewer.build_ui = _c72_build_ui

def _c72_source_changed(viewer, pane_index):
    _c72_sync_ws_quick_filter_button(viewer, pane_index)
    _, modes = _c61_filter_state(viewer)
    source_name = viewer.sources[pane_index].currentText()
    modes[pane_index] = _C70_WS_MODE if _c70_is_ws_source_name(source_name) else "ALL"

_old_c72_load_pane = MultiPaneLogViewer.load_pane
def _c72_load_pane(self, side):
    result = _old_c72_load_pane(self, side)
    pane_index = self.side_index(side)
    _c72_sync_ws_quick_filter_button(self, pane_index)
    return result
MultiPaneLogViewer.load_pane = _c72_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)

APP_VERSION = "2.0.0-rc1-commit0072"

# ---------------------------------------------------------------------------
# Commit0073: CSA default Err Quick Filter + reliable LOAD LOGS workflow
# ---------------------------------------------------------------------------
# CSA panes should open on errors by default. LOAD LOGS must provide visible
# feedback, reject re-entry, keep the UI responsive between panes, and report
# partial failures instead of appearing to do nothing.

def _c73_is_csa_source(source_name):
    return str(source_name or "").strip().upper() == "CSA"


def _c73_default_quick_mode_for_source(source_name):
    if _c70_is_ws_source_name(source_name):
        return _C70_WS_MODE
    if _c73_is_csa_source(source_name):
        return "ERROR"
    return "ALL"


def _c73_apply_source_default_filter(viewer, pane_index, *, refresh=True):
    if not (0 <= pane_index < len(getattr(viewer, "sources", []))):
        return
    _, modes = _c61_filter_state(viewer)
    modes[pane_index] = _c73_default_quick_mode_for_source(
        viewer.sources[pane_index].currentText()
    )
    _c72_sync_ws_quick_filter_button(viewer, pane_index)
    _f52a_update_quick_filter_labels(viewer, pane_index)
    if refresh and pane_index < len(getattr(viewer, "all_rows", [])):
        _c70_apply_view_filters(viewer, pane_index)


_old_c73_source_changed = _c72_source_changed

def _c73_source_changed(viewer, pane_index):
    _c72_sync_ws_quick_filter_button(viewer, pane_index)
    _c73_apply_source_default_filter(viewer, pane_index, refresh=False)
    if hasattr(viewer, "status"):
        source = viewer.sources[pane_index].currentText()
        mode = _c73_default_quick_mode_for_source(source)
        viewer.status.setText(
            f"{viewer.pane_name(pane_index)} configured: {source}; "
            f"default Quick Filter={mode}. Press LOAD LOGS."
        )


_old_c73_load_pane = MultiPaneLogViewer.load_pane

def _c73_load_pane(self, side):
    pane_index = self.side_index(side)
    source_name = self.sources[pane_index].currentText()
    if hasattr(self, "status"):
        self.status.setText(f"Loading {self.pane_name(pane_index)}: {source_name} ...")
    QApplication.processEvents()
    result = _old_c73_load_pane(self, side)
    _c73_apply_source_default_filter(self, pane_index, refresh=True)
    row_count = len(self.all_rows[pane_index]) if pane_index < len(self.all_rows) else 0
    mode = _c73_default_quick_mode_for_source(source_name)
    if hasattr(self, "status"):
        self.status.setText(
            f"Loaded {self.pane_name(pane_index)}: {source_name}; "
            f"indexed {row_count:,}; Quick Filter={mode}."
        )
    QApplication.processEvents()
    return result


def _c73_load_visible(self):
    if getattr(self, "_c73_load_logs_running", False):
        if hasattr(self, "status"):
            self.status.setText("LOAD LOGS is already running. Please wait.")
        return

    indices = list(self.visible_indices())
    if not indices:
        QMessageBox.information(self, "Log Viewer", "Select at least one Show pane before LOAD LOGS.")
        return

    self._c73_load_logs_running = True
    button = getattr(self, "load_visible_btn", None)
    old_text = button.text() if button is not None else "LOAD LOGS"
    failures = []
    loaded = []
    try:
        if button is not None:
            button.setEnabled(False)
            button.setText("LOADING...")
        if hasattr(self, "status"):
            self.status.setText(f"LOAD LOGS started — {len(indices)} visible pane(s).")
        QApplication.processEvents()

        for sequence, pane_index in enumerate(indices, start=1):
            source_name = self.sources[pane_index].currentText()
            if hasattr(self, "status"):
                self.status.setText(
                    f"LOAD LOGS {sequence}/{len(indices)} — "
                    f"{self.pane_name(pane_index)}: {source_name}"
                )
            QApplication.processEvents()
            try:
                self.load_pane(pane_index)
                loaded.append(pane_index)
            except Exception:
                failures.append((pane_index, source_name, traceback.format_exc()))
                write_startup_log(
                    f"Commit0073 LOAD LOGS pane failure: {pane_index} / {source_name}\n\n"
                    + failures[-1][2]
                )
            QApplication.processEvents()

        if failures:
            summary = "\n".join(
                f"- {self.pane_name(i)} ({source})" for i, source, _tb in failures
            )
            if hasattr(self, "status"):
                self.status.setText(
                    f"LOAD LOGS completed with errors: {len(loaded)} loaded, "
                    f"{len(failures)} failed."
                )
            QMessageBox.warning(
                self,
                "LOAD LOGS",
                "Some panes could not be loaded.\n\n" + summary
                + f"\n\nDiagnostic log:\n{startup_log_path()}",
            )
        else:
            total_rows = sum(len(self.all_rows[i]) for i in loaded)
            if hasattr(self, "status"):
                self.status.setText(
                    f"LOAD LOGS complete — {len(loaded)} pane(s), "
                    f"{total_rows:,} indexed rows."
                )
    finally:
        if button is not None:
            button.setText(old_text)
            button.setEnabled(True)
        self._c73_load_logs_running = False
        QApplication.processEvents()


_old_c73_build_ui = MultiPaneLogViewer.build_ui

def _c73_build_ui(self):
    _old_c73_build_ui(self)

    # Replace every earlier LOAD LOGS connection with the final guarded handler.
    try:
        self.load_visible_btn.clicked.disconnect()
    except (TypeError, RuntimeError):
        pass
    self.load_visible_btn.clicked.connect(lambda _checked=False: _c73_load_visible(self))
    self.load_visible_btn.setToolTip(
        "Load all visible panes. The button shows progress and prevents duplicate loading. "
        "CSA uses Err as its default Quick Filter."
    )

    # Replace source-change connections so CSA receives its own default filter.
    for pane_index, combo in enumerate(self.sources):
        try:
            combo.currentTextChanged.disconnect()
        except (TypeError, RuntimeError):
            pass
        combo.currentTextChanged.connect(
            lambda _text, p=pane_index: _c73_source_changed(self, p)
        )
        _c73_apply_source_default_filter(self, pane_index, refresh=False)

    if hasattr(self, "status"):
        self.status.setText(
            "Ready — default viewer: 2 Show (WS + CSA). "
            "CSA Quick Filter defaults to Err. Press LOAD LOGS."
        )


MultiPaneLogViewer.build_ui = _c73_build_ui
MultiPaneLogViewer.load_visible = _c73_load_visible
MultiPaneLogViewer.load_both = _c73_load_visible
MultiPaneLogViewer.load_pane = _c73_load_pane
MultiPaneLogViewer.load_side = lambda self, side: self.load_pane(side)
_c72_source_changed = _c73_source_changed

APP_VERSION = "2.0.0-rc1-commit0073"

if __name__ == "__main__":
    main()
