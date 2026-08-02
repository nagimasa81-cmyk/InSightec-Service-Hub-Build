from __future__ import annotations

import csv
import json
import os
import copy
import shutil
import sys
import tempfile
import zipfile
import threading
import queue
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
import pydicom
import pyqtgraph as pg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PySide6.QtCore import Qt, Signal, QMimeData, QUrl, QTimer, QObject, QThread, Slot, QEvent, QSettings, QSize
from PySide6.QtGui import QAction, QKeySequence, QImage, QDrag, QPainter, QColor, QPen
from pydicom.uid import generate_uid
from decoder_manager import decode_dicom_pixels
from tracker_position import DirectionMeasurement, signal_peak_coordinate, solve_position
from stable_diagnostic_logger import StableDiagnosticLogger
from artifact_learning_db import ArtifactDatabase, image_features
from artifact_detection import classify_feature_vector, features_to_vector
from orientation_engine import OrientationEngine, DisplayTransform
from navigation_controller import NavigationController
from viewer_toolbar import ViewerToolbar
from core.display_state import DisplayState
from core.cache_engine import LRUCache, LRUKeySet
from core.hybrid_compensation import compensate as hybrid_compensate, detect_artifacts as hybrid_detect_artifacts
from core.auto_correct import auto_correct as run_auto_correct, auto_correct_with_retry, recalculate_with_mask
from core.roi_raw_compensation import (
    RoiCompensationDetection,
    apply_roi_background_compensation,
    build_manual_mask_detection,
    detect_roi_artifact_mask,
)
from raw_import_engine import (
    RawImportError,
    RawImportResult,
    load_raw_file_auto,
    try_render_fus_raw_exact,
)
from PySide6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDialog, QFileDialog, QFormLayout, QFrame,
    QGridLayout, QGroupBox, QHBoxLayout, QLabel, QLineEdit, QMainWindow, QMessageBox, QPushButton,
    QDoubleSpinBox, QProgressBar, QProgressDialog, QScrollArea, QSizePolicy, QSlider, QSpinBox, QSplitter, QStatusBar, QTabWidget, QTableWidget,
    QTableWidgetItem, QTextEdit, QTreeWidget, QTreeWidgetItem, QTreeWidgetItemIterator, QVBoxLayout, QWidget, QMenu, QGraphicsEllipseItem,
)

def _release_mode_enabled() -> bool:
    # An explicitly supplied runtime value always wins.  This keeps Python
    # debug launches and guide-enabled validation builds deterministic.
    raw_env = os.environ.get("INSIGHTEC_RELEASE_MODE")
    if raw_env is not None:
        return raw_env.strip().lower() in {"1", "true", "yes", "on"}

    # Packaged builds carry their build profile beside the executable.
    # The build BAT generates this file before Nuitka starts and Nuitka
    # includes it in the distribution, so release behavior does not depend
    # on an environment variable still being present when the EXE is run.
    try:
        cfg = Path(__file__).resolve().with_name("release_mode.json")
        if cfg.is_file():
            payload = json.loads(cfg.read_text(encoding="utf-8"))
            if "release_mode" in payload:
                return bool(payload["release_mode"])
            return payload.get("guide_tour_enabled_in_release") is False
    except Exception:
        pass
    return False

RELEASE_MODE = _release_mode_enabled()

APP_NAME = "MR Image Explorer"
APP_VERSION = "5.51.0 RC1 Commit0122 Quick Spike Execution and Series Open Fix"
# Compatibility markers retained for regression audits: Commit0085, Commit0084, Commit0083, Commit0072, Commit0071, Commit0069, Commit0068o.
# Legacy audit compatibility: Commit0066P1 used setObjectName("ViewerTopDisplayNavigationToolbar").

pg.setConfigOptions(imageAxisOrder="row-major", antialias=True)

BITMAP_EXTENSIONS = frozenset({
    ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"
})
SIGNAL_EXTENSIONS = frozenset({".csv", ".npy", ".npz"})
RAW_EXTENSIONS = frozenset({
    ".raw", ".bin", ".dat", ".7", ".pfile", ".img", ".kspace",
    ".cfl", ".rawdata", ".complex"
})



def median_filter_3x3_numpy(array: np.ndarray) -> np.ndarray:
    """NumPy-only 3x3 median filter used for RAW display scoring."""
    source = np.asarray(array, dtype=float)
    if source.ndim != 2 or source.size == 0:
        return source.copy()

    padded = np.pad(source, 1, mode="edge")
    try:
        windows = np.lib.stride_tricks.sliding_window_view(
            padded, (3, 3)
        )
        return np.median(windows, axis=(-2, -1))
    except Exception:
        output = np.empty_like(source, dtype=float)
        for row in range(source.shape[0]):
            for col in range(source.shape[1]):
                output[row, col] = np.median(
                    padded[row:row + 3, col:col + 3]
                )
        return output


def fft2c(image: np.ndarray) -> np.ndarray:
    return np.fft.fftshift(np.fft.fft2(np.fft.ifftshift(image)))


def ifft2c(kspace: np.ndarray) -> np.ndarray:
    return np.fft.fftshift(np.fft.ifft2(np.fft.ifftshift(kspace)))


def display_array(arr: np.ndarray, mode: str) -> np.ndarray:
    if mode == "FFT":
        return np.log1p(np.abs(arr))
    return np.abs(arr)


@dataclass
class DicomEntry:
    path: Path
    ds: pydicom.dataset.FileDataset
    image: Optional[np.ndarray]
    sort_key: tuple


class ImportWorker(QObject):
    progress = Signal(str, int, int)
    completed = Signal(object)
    failed = Signal(str)
    canceled = Signal()

    def __init__(self, paths: list[Path]):
        super().__init__()
        self.paths = [Path(path) for path in paths]
        self._cancel_requested = False
        self.temp_dirs: list[Path] = []
        self.raw_preview_cache: dict[str, RawImportResult] = {}
        self.origin_info: dict[str, dict] = {}

    @Slot()
    def cancel(self):
        self._cancel_requested = True

    def _check_cancel(self):
        if self._cancel_requested:
            raise InterruptedError("Import canceled")

    @staticmethod
    def _looks_tracker(path: Path) -> bool:
        name = path.name.lower()
        if "trackerimg" in name or "pfile" in name:
            return True
        if path.suffix == "" and path.stat().st_size > 4096:
            try:
                with path.open("rb") as handle:
                    first = handle.read(4)
                rev = np.frombuffer(first, dtype="<f4")[0]
                return bool(np.isfinite(rev) and 1 < rev < 100)
            except Exception:
                return False
        return False

    @staticmethod
    def _read_dicom_metadata(path: Path) -> DicomEntry:
        ds = pydicom.dcmread(
            str(path),
            stop_before_pixels=True,
            defer_size="1 KB",
            force=False,
        )
        rows = int(getattr(ds, "Rows", 0) or 0)
        cols = int(getattr(ds, "Columns", 0) or 0)
        if rows <= 0 or cols <= 0:
            raise ValueError("Not an image DICOM")

        instance = int(getattr(ds, "InstanceNumber", 0) or 0)
        position = getattr(ds, "ImagePositionPatient", None)
        z_value = (
            float(position[2])
            if position is not None and len(position) >= 3
            else float(instance)
        )
        series = str(getattr(ds, "SeriesInstanceUID", "") or "")
        sort_key = (series, z_value, instance, path.name.lower())
        return DicomEntry(path=path, ds=ds, image=None, sort_key=sort_key)

    def _extract_zip(self, path: Path) -> list[Path]:
        root = Path(tempfile.mkdtemp(prefix="mr_image_explorer_"))
        self.temp_dirs.append(root)

        extracted: list[Path] = []
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            total = max(len(members), 1)

            for index, member in enumerate(members, start=1):
                self._check_cancel()
                self.progress.emit(
                    f"Extracting ZIP {index:,}/{total:,}\n{member.filename}",
                    index,
                    total,
                )

                destination = (root / member.filename).resolve()
                if not str(destination).startswith(str(root.resolve())):
                    raise ValueError("Unsafe ZIP member path")

                if member.is_dir():
                    destination.mkdir(parents=True, exist_ok=True)
                    continue

                destination.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(member, "r") as source, destination.open("wb") as target:
                    while True:
                        self._check_cancel()
                        chunk = source.read(1024 * 1024)
                        if not chunk:
                            break
                        target.write(chunk)
                extracted.append(destination)
                member_path = Path(member.filename)
                self.origin_info[str(destination)] = {
                    "container": path.name,
                    "container_type": "ZIP",
                    "relative_path": member.filename,
                    "group_path": str(member_path.parent)
                    if str(member_path.parent) != "."
                    else "(ZIP root)",
                }

        return extracted

    @Slot()
    def run(self):
        try:
            supported_extensions = {
                ".dcm", ".ima", ".dicom", "",
                ".raw", ".bin", ".dat", ".7", ".pfile", ".img", ".kspace", ".cfl", ".rawdata", ".complex",
                ".csv", ".npy", ".npz",
                ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff",
            }

            files: list[Path] = []
            checked = 0

            self.progress.emit("Collecting files...", 0, 0)

            for source in self.paths:
                self._check_cancel()

                if source.is_dir():
                    for candidate in source.rglob("*"):
                        self._check_cancel()
                        checked += 1
                        if candidate.is_file():
                            files.append(candidate)
                            try:
                                relative = candidate.relative_to(source)
                            except Exception:
                                relative = Path(candidate.name)
                            self.origin_info[str(candidate)] = {
                                "container": source.name,
                                "container_type": "Folder",
                                "relative_path": str(relative),
                                "group_path": str(relative.parent)
                                if str(relative.parent) != "."
                                else "(Folder root)",
                            }
                        if checked % 100 == 0:
                            self.progress.emit(
                                f"Scanning folder...\n"
                                f"Checked: {checked:,}\n"
                                f"Candidates: {len(files):,}",
                                0,
                                0,
                            )

                elif source.is_file() and source.suffix.lower() == ".zip":
                    files.extend(self._extract_zip(source))

                elif source.is_file():
                    files.append(source)
                    self.origin_info[str(source)] = {
                        "container": source.parent.name,
                        "container_type": "File",
                        "relative_path": source.name,
                        "group_path": "(Selected files)",
                    }

            unique_files: list[Path] = []
            seen = set()
            for path in files:
                self._check_cancel()
                try:
                    key = str(path.resolve()).lower()
                except Exception:
                    key = str(path).lower()
                if key not in seen:
                    seen.add(key)
                    unique_files.append(path)

            files = unique_files
            total = len(files)

            dicoms: list[DicomEntry] = []
            trackers: list[Path] = []
            raw_files: list[Path] = []
            bitmaps: list[Path] = []
            signals: list[Path] = []
            skipped = 0

            for index, path in enumerate(files, start=1):
                self._check_cancel()
                self.progress.emit(
                    f"Indexing {index:,}/{total:,}\n{path.name}",
                    index,
                    max(total, 1),
                )

                suffix = path.suffix.lower()

                # First test every file as DICOM. This includes extensionless
                # and vendor-named image files in deep ZIP/folder structures.
                try:
                    dicoms.append(self._read_dicom_metadata(path))
                    continue
                except Exception:
                    pass

                if suffix in BITMAP_EXTENSIONS:
                    bitmaps.append(path)
                    continue
                if suffix in SIGNAL_EXTENSIONS:
                    signals.append(path)
                    continue
                if suffix in {".txt", ".json", ".xml", ".ini", ".log", ".md"}:
                    continue

                try:
                    if self._looks_tracker(path):
                        trackers.append(path)
                    elif suffix in RAW_EXTENSIONS:
                        raw_files.append(path)

                        # Build the exact FUS preview while the import worker is
                        # already indexing files. Tree selection will therefore
                        # display immediately without another decoding pass.
                        try:
                            exact_preview = try_render_fus_raw_exact(path)
                            if exact_preview is not None:
                                self.raw_preview_cache[str(path)] = exact_preview
                        except Exception:
                            pass
                    else:
                        skipped += 1
                except Exception:
                    skipped += 1

            dicoms.sort(key=lambda entry: entry.sort_key)

            self.completed.emit({
                "all_files": files,
                "dicoms": dicoms,
                "trackers": trackers,
                "raw_files": raw_files,
                "raw_preview_cache": self.raw_preview_cache,
                "origin_info": self.origin_info,
                "bitmaps": bitmaps,
                "signals": signals,
                "skipped": skipped,
                "temp_dirs": self.temp_dirs,
            })

        except InterruptedError:
            self.canceled.emit()
        except Exception as exc:
            self.failed.emit(f"{type(exc).__name__}: {exc}")


class DropBanner(QFrame):
    pathsDropped = Signal(list)
    clicked = Signal()

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.setObjectName("DropBanner")
        self.setCursor(Qt.PointingHandCursor)
        self.setToolTip(
            "Click to select files or a folder.\n"
            "You can also drag and drop files, folders, or ZIP archives here."
        )
        layout = QVBoxLayout(self)
        self.title = QLabel("Click or Drop DICOM / Raw / P File / 1D Data / Folder / ZIP Here")
        self.title.setAlignment(Qt.AlignCenter)
        self.title.setStyleSheet("font-size: 17px; font-weight: 600;")
        subtitle = QLabel("DICOM, RAW, GE TrackerImg/PFile, Siemens .dat, CSV/TXT/NPY/NPZ, images")
        subtitle.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.title)
        layout.addWidget(subtitle)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit()
            event.accept()
            return
        super().mousePressEvent(event)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setProperty("dragActive", True)
            self.style().unpolish(self); self.style().polish(self)

    def dragLeaveEvent(self, event):
        self.setProperty("dragActive", False)
        self.style().unpolish(self); self.style().polish(self)

    def dropEvent(self, event):
        self.setProperty("dragActive", False)
        self.style().unpolish(self)
        self.style().polish(self)
        paths = [
            Path(url.toLocalFile())
            for url in event.mimeData().urls()
            if url.isLocalFile()
        ]
        if paths:
            event.setDropAction(Qt.CopyAction)
            event.accept()
            self.pathsDropped.emit(paths)


class ClickableComponentLabel(QLabel):
    componentMenuRequested = Signal(object)

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.PointingHandCursor)
        self.setToolTip("Click the image component name to select Magnitude, Real, Imaginary, or Phase.")

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.componentMenuRequested.emit(event.globalPosition().toPoint())
            event.accept()
            return
        super().mousePressEvent(event)


class ImagePanel(QWidget):
    lineChanged = Signal(int, int)
    pageRequested = Signal(int)
    mouseActionChanged = Signal(str)
    viewRequested = Signal(str)
    levelWheel = Signal(float, bool)
    componentRequested = Signal(str)
    keyboardPageRequested = Signal(int)
    activated = Signal(object)
    utilityRequested = Signal(str)
    manualMaskChanged = Signal()

    def __init__(self, title: str):
        super().__init__()
        layout = QVBoxLayout(self); layout.setContentsMargins(0, 0, 0, 0)
        self.label = ClickableComponentLabel(title)
        self.label.setStyleSheet("font-weight: 600; padding: 3px;")
        self.label.componentMenuRequested.connect(self._show_component_menu)
        self.setFocusPolicy(Qt.StrongFocus)
        layout.addWidget(self.label)
        self.plot = pg.PlotWidget()
        self.plot.setFocusPolicy(Qt.StrongFocus)
        self.plot.setMenuEnabled(False)
        self.plot.getViewBox().setMenuEnabled(False)
        self.plot.scene().sigMouseClicked.connect(self._scene_mouse_clicked)
        self.plot.viewport().installEventFilter(self)
        self.mouse_action = "Standard"
        self.plot.getViewBox().setMouseMode(pg.ViewBox.PanMode)
        # Use conventional medical-image raster coordinates: NumPy row 0 is
        # displayed at the top. Without this, pyqtgraph shows the image with
        # the vertical axis reversed and anterior/posterior appear swapped.
        self.plot.getViewBox().invertY(True)
        self._wl_drag_origin = None
        self._wl_drag_start = None
        self._wl_dragged = False
        self._zoom_drag_origin = None
        self._zoom_drag_last_y = None
        self._zoom_dragged = False
        self.plot.setAspectLocked(True)
        self.plot.hideAxis("left"); self.plot.hideAxis("bottom")
        self.image_item = pg.ImageItem()
        self.plot.addItem(self.image_item)
        self.hline = pg.InfiniteLine(angle=0, movable=True, pen=pg.mkPen('#ff4949', width=1.5, style=Qt.DashLine))
        self.vline = pg.InfiniteLine(angle=90, movable=True, pen=pg.mkPen('#ff4949', width=1.5, style=Qt.DashLine))
        self.plot.addItem(self.hline); self.plot.addItem(self.vline)
        # Commit0115: crosshairs are opt-in profile cursors.  New panels,
        # comparison dialogs, candidate previews and diagnostic viewers must
        # never show them merely because an ImagePanel was constructed.
        self.hline.hide(); self.vline.hide()
        self.hline.sigPositionChanged.connect(self._emit_position)
        self.vline.sigPositionChanged.connect(self._emit_position)
        self.comp_roi = pg.RectROI([10, 10], [40, 40], pen=pg.mkPen("#35a7ff", width=2))
        self.comp_roi.addScaleHandle([1, 1], [0, 0])
        self.comp_roi.addScaleHandle([0, 0], [1, 1])
        self.comp_roi.addScaleHandle([1, 0], [0, 1])
        self.comp_roi.addScaleHandle([0, 1], [1, 0])
        self.comp_roi.hide()
        self.plot.addItem(self.comp_roi)
        self.manual_mask = None
        self.manual_mask_enabled = False
        self.manual_mask_mode = "Brush"
        self.manual_brush_size = 9
        self._manual_painting = False
        self.manual_mask_item = pg.ImageItem()
        self.manual_mask_item.setZValue(20)
        self.manual_mask_item.hide()
        self.plot.addItem(self.manual_mask_item)
        # Commit0117: image-space brush cursor.  A standard Qt cursor has a
        # fixed screen-pixel size and therefore does not represent the actual
        # mask diameter after zooming.  This ellipse lives in image coordinates
        # and always matches the current brush size.
        self.manual_brush_cursor = QGraphicsEllipseItem()
        self.manual_brush_cursor.setPen(pg.mkPen(QColor(255, 230, 80, 235), width=1.5))
        self.manual_brush_cursor.setBrush(Qt.NoBrush)
        self.manual_brush_cursor.setZValue(40)
        self.manual_brush_cursor.hide()
        self.plot.addItem(self.manual_brush_cursor)
        layout.addWidget(self.plot, 1)
        self.shape = (1, 1)
        self.current_levels = None
        self.available_components = ["Magnitude"]
        self.current_component = "Magnitude"

        self.orientation_labels = {}
        for key in ("top", "bottom", "left", "right"):
            orientation_label = QLabel("", self)
            orientation_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)
            orientation_label.setStyleSheet(
                "color:#fff4a6; font-size:17px; font-weight:900;"
                "background:rgba(0,0,0,105); padding:2px 5px; border-radius:3px;"
            )
            orientation_label.hide()
            self.orientation_labels[key] = orientation_label

        self.annotation_label = QLabel("", self)
        self.annotation_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.annotation_label.setWordWrap(True)
        self.annotation_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.annotation_label.setStyleSheet(
            "color:#f2f7fb; font-size:12px; font-weight:700;"
            "background:rgba(0,0,0,120);"
            "border:1px solid rgba(80,200,255,150);"
            "padding:5px 7px; border-radius:4px;"
        )
        self.annotation_label.hide()

        # Frequency-encoding direction marker.  It intentionally has no text:
        # only a very small paired chevron is drawn tight to the image corner.
        self.frequency_marker = pg.PlotDataItem(
            pen=pg.mkPen((255, 255, 255, 205), width=2)
        )
        self.frequency_marker.setZValue(1000)
        self.frequency_marker.hide()
        self.plot.addItem(self.frequency_marker)
        self.frequency_direction = None

    def set_frequency_direction(self, direction):
        value = str(direction or "").upper()
        self.frequency_direction = value if value in ("HORIZONTAL", "VERTICAL") else None
        self._update_frequency_marker()

    def _update_frequency_marker(self):
        if not self.frequency_direction or min(self.shape) <= 1:
            self.frequency_marker.hide()
            return
        rows, cols = self.shape
        # Keep the marker unobtrusive and almost flush with the lower-right
        # image corner.  Horizontal uses <>, vertical uses the rotated pair.
        margin = max(3.0, min(rows, cols) * 0.012)
        size = max(3.0, min(rows, cols) * 0.014)
        gap = size * 0.50
        x = cols - margin - size
        y = rows - margin - size
        if self.frequency_direction == "HORIZONTAL":
            xs = [x, x-size, x, np.nan, x+gap, x+gap+size, x+gap]
            ys = [y-size, y, y+size, np.nan, y-size, y, y+size]
        else:
            xs = [x-size, x, x+size, np.nan, x-size, x, x+size]
            ys = [y, y-size, y, np.nan, y+gap, y+gap+size, y+gap]
        self.frequency_marker.setData(xs, ys, connect="finite")
        self.frequency_marker.show()

    def set_available_components(self, components, current=None):
        ordered = [name for name in ("Magnitude", "Log Magnitude", "Real", "Imaginary", "Phase") if name in set(components)]
        self.available_components = ordered or ["Magnitude"]
        if current in self.available_components:
            self.current_component = current
        elif self.current_component not in self.available_components:
            self.current_component = self.available_components[0]

    def _show_component_menu(self, global_pos):
        menu = QMenu(self)
        actions = {}
        for component in self.available_components:
            action = menu.addAction(component)
            action.setCheckable(True)
            action.setChecked(component == self.current_component)
            actions[action] = component
        selected = menu.exec(global_pos)
        if selected in actions:
            self.current_component = actions[selected]
            self.componentRequested.emit(actions[selected])

    def mousePressEvent(self, event):
        self.setFocus(Qt.MouseFocusReason)
        super().mousePressEvent(event)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Up:
            self.keyboardPageRequested.emit(-1)
            event.accept()
            return
        if event.key() == Qt.Key_Down:
            self.keyboardPageRequested.emit(1)
            event.accept()
            return
        super().keyPressEvent(event)

    def set_orientation_labels(self, values):
        for key, label in self.orientation_labels.items():
            value = str(values.get(key, "") or "").strip()
            label.setText(value)
            label.setVisible(bool(value))
        self._position_orientation_labels()

    def clear_orientation_labels(self):
        self.set_orientation_labels({})

    def set_annotation(self, text, mode="Minimum"):
        value = str(text or "").strip()
        self.annotation_label.setText(value)
        self.annotation_label.setVisible(bool(value))
        self.annotation_label.setMaximumWidth(360 if mode == "Full" else 260)
        self._position_annotation()

    def clear_annotation(self):
        self.annotation_label.clear()
        self.annotation_label.hide()

    def _position_annotation(self):
        if not self.annotation_label.isVisible():
            return
        self.annotation_label.adjustSize()
        width = min(self.annotation_label.width(), max(180, self.width() // 2))
        self.annotation_label.resize(width, self.annotation_label.sizeHint().height())
        self.annotation_label.move(12, 38)
        self.annotation_label.raise_()

    def _position_orientation_labels(self):
        margin = 12
        rect = self.rect()
        for key, label in self.orientation_labels.items():
            if not label.isVisible():
                continue
            label.adjustSize()
            if key == "top":
                label.move((rect.width() - label.width()) // 2, 34)
            elif key == "bottom":
                label.move(
                    (rect.width() - label.width()) // 2,
                    rect.height() - label.height() - margin,
                )
            elif key == "left":
                label.move(margin, (rect.height() - label.height()) // 2)
            else:
                label.move(
                    rect.width() - label.width() - margin,
                    (rect.height() - label.height()) // 2,
                )
            label.raise_()

    def set_manual_mask_editing(self, enabled: bool, mode: str = "Brush", brush_size: int = 9):
        self.manual_mask_enabled = bool(enabled)
        self.manual_mask_mode = str(mode)
        self.manual_brush_size = max(1, int(brush_size))
        if self.manual_mask_enabled:
            if self.manual_mask is None or self.manual_mask.shape != tuple(self.shape):
                self.manual_mask = np.zeros(tuple(self.shape), dtype=bool)
            self.manual_mask_item.show()
            self._refresh_manual_mask_overlay()
            if self.manual_mask_mode in ("Brush", "Eraser"):
                self.plot.viewport().setCursor(Qt.BlankCursor)
            else:
                self.manual_brush_cursor.hide()
                self.plot.viewport().setCursor(Qt.CrossCursor)
        else:
            self._manual_painting = False
            self.manual_brush_cursor.hide()
            self.plot.viewport().unsetCursor()

    def _update_manual_brush_cursor(self, viewport_pos):
        if not self.manual_mask_enabled or self.manual_mask_mode not in ("Brush", "Eraser"):
            self.manual_brush_cursor.hide()
            return
        scene_pos = self.plot.mapToScene(viewport_pos.toPoint())
        point = self.plot.getViewBox().mapSceneToView(scene_pos)
        diameter = float(max(1, self.manual_brush_size))
        radius = diameter / 2.0
        self.manual_brush_cursor.setRect(
            float(point.x()) - radius, float(point.y()) - radius, diameter, diameter
        )
        self.manual_brush_cursor.setPen(
            pg.mkPen(QColor(255, 120, 100, 240) if self.manual_mask_mode == "Eraser" else QColor(255, 230, 80, 235), width=1.5)
        )
        self.manual_brush_cursor.show()

    def clear_manual_mask(self):
        """Clear both the logical paint mask and its cached overlay immediately."""
        shape = tuple(self.shape)
        if self.manual_mask is not None:
            shape = self.manual_mask.shape
        # Replace the array instead of only mutating it.  This prevents stale
        # references held by a previous detection/preview from repainting the
        # cleared pixels on the next refresh.
        self.manual_mask = np.zeros(shape, dtype=bool)
        transparent = np.zeros(shape + (4,), dtype=np.uint8)
        self.manual_mask_item.setImage(transparent, autoLevels=False)
        self.manual_mask_item.setRect(0, 0, shape[1], shape[0])
        self.manual_mask_item.hide()
        self.plot.update()
        self.plot.viewport().update()
        self.manualMaskChanged.emit()

    def _refresh_manual_mask_overlay(self):
        if self.manual_mask is None:
            self.manual_mask_item.hide()
            return
        rgba = np.zeros(self.manual_mask.shape + (4,), dtype=np.uint8)
        rgba[..., 0] = 255
        rgba[..., 1] = 196
        rgba[..., 2] = 0
        rgba[..., 3] = self.manual_mask.astype(np.uint8) * 120
        self.manual_mask_item.setImage(rgba, autoLevels=False)
        self.manual_mask_item.setRect(0, 0, self.manual_mask.shape[1], self.manual_mask.shape[0])
        self.manual_mask_item.setVisible(bool(np.any(self.manual_mask)))

    def _paint_manual_mask(self, viewport_pos):
        if self.manual_mask is None:
            return
        scene_pos = self.plot.mapToScene(viewport_pos.toPoint())
        point = self.plot.getViewBox().mapSceneToView(scene_pos)
        x, y = int(round(point.x())), int(round(point.y()))
        rows, cols = self.manual_mask.shape
        radius = max(0, self.manual_brush_size // 2)
        yy, xx = np.ogrid[:rows, :cols]
        region = (yy - y) ** 2 + (xx - x) ** 2 <= radius ** 2
        mode = self.manual_mask_mode
        if mode == "Eraser":
            self.manual_mask[region] = False
        elif mode == "Remove Component":
            if 0 <= y < rows and 0 <= x < cols and self.manual_mask[y, x]:
                source = self.manual_mask.copy()
                stack = [(y, x)]
                source[y, x] = False
                component = []
                while stack:
                    cy0, cx0 = stack.pop()
                    component.append((cy0, cx0))
                    for dy0, dx0 in ((-1, 0), (1, 0), (0, -1), (0, 1), (-1, -1), (-1, 1), (1, -1), (1, 1)):
                        ny, nx = cy0 + dy0, cx0 + dx0
                        if 0 <= ny < rows and 0 <= nx < cols and source[ny, nx]:
                            source[ny, nx] = False
                            stack.append((ny, nx))
                if component:
                    ys0, xs0 = zip(*component)
                    self.manual_mask[np.asarray(ys0), np.asarray(xs0)] = False
        elif mode == "Line":
            self.manual_mask[max(0, y-radius):min(rows, y+radius+1), :] = True
        elif mode == "Band":
            band = max(2, radius * 2)
            self.manual_mask[max(0, y-band):min(rows, y+band+1), :] = True
        elif mode == "Block":
            self.manual_mask[max(0, y-radius):min(rows, y+radius+1), max(0, x-radius):min(cols, x+radius+1)] = True
        elif mode == "Ring":
            inner = max(0, radius - max(1, radius // 3))
            ring = region & (((yy-y)**2 + (xx-x)**2) >= inner**2)
            self.manual_mask[ring] = True
        else:
            self.manual_mask[region] = True
        self._refresh_manual_mask_overlay()
        self.manualMaskChanged.emit()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._position_orientation_labels()
        self._position_annotation()

    def eventFilter(self, watched, event):
        if watched is self.plot.viewport():
            event_type = event.type()

            if self.manual_mask_enabled:
                if event_type == QEvent.MouseMove:
                    self._update_manual_brush_cursor(event.position())
                elif event_type == QEvent.Leave:
                    self.manual_brush_cursor.hide()
                if event_type == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                    self._manual_painting = True
                    self._paint_manual_mask(event.position())
                    event.accept()
                    return True
                if event_type == QEvent.MouseMove and self._manual_painting and event.buttons() & Qt.LeftButton:
                    self._paint_manual_mask(event.position())
                    event.accept()
                    return True
                if event_type == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
                    self._manual_painting = False
                    event.accept()
                    return True

            # Mouse controls are deliberately modifier-free:
            #   wheel rotation       = previous / next image
            #   left drag            = pan (pyqtgraph ViewBox PanMode)
            #   left double-click    = fit image to view
            #   middle-button drag   = zoom (up=in, down=out)
            #   right drag           = window / level
            #   right click          = viewer context menu
            if event_type == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                self.setFocus(Qt.MouseFocusReason)
                self.activated.emit(self)
                return False

            if event_type == QEvent.MouseButtonDblClick and event.button() == Qt.LeftButton:
                self.setFocus(Qt.MouseFocusReason)
                self.activated.emit(self)
                self.fit_to_image()
                event.accept()
                return True

            if event_type == QEvent.MouseButtonPress and event.button() == Qt.MiddleButton:
                self.setFocus(Qt.MouseFocusReason)
                self.activated.emit(self)
                self._zoom_drag_origin = event.position()
                self._zoom_drag_last_y = float(event.position().y())
                self._zoom_dragged = False
                self.plot.viewport().setCursor(Qt.SizeVerCursor)
                event.accept()
                return True

            if (
                event_type == QEvent.MouseMove
                and self._zoom_drag_origin is not None
                and event.buttons() & Qt.MiddleButton
            ):
                current_y = float(event.position().y())
                total_delta = event.position() - self._zoom_drag_origin
                if abs(float(total_delta.y())) >= 3.0:
                    self._zoom_dragged = True
                if self._zoom_dragged and self._zoom_drag_last_y is not None:
                    delta_y = current_y - self._zoom_drag_last_y
                    # Dragging upward zooms in; dragging downward zooms out.
                    factor = float(np.exp(delta_y * 0.012))
                    factor = min(1.20, max(0.83, factor))
                    self.plot.getViewBox().scaleBy((factor, factor))
                self._zoom_drag_last_y = current_y
                event.accept()
                return True

            if (
                event_type == QEvent.MouseButtonRelease
                and event.button() == Qt.MiddleButton
                and self._zoom_drag_origin is not None
            ):
                self._zoom_drag_origin = None
                self._zoom_drag_last_y = None
                self._zoom_dragged = False
                self.plot.viewport().unsetCursor()
                event.accept()
                return True

            if event_type == QEvent.MouseButtonPress and event.button() == Qt.RightButton:
                self.setFocus(Qt.MouseFocusReason)
                self.activated.emit(self)
                self._wl_drag_origin = event.position()
                self._wl_dragged = False
                window = self.window()
                if hasattr(window, "_prepare_level_controls_for_panel"):
                    window._prepare_level_controls_for_panel(self)
                self._wl_drag_start = (
                    float(window.window_level_spin.value()),
                    float(window.dynamic_range_spin.value()),
                )
                event.accept()
                return True

            if event_type == QEvent.MouseMove and self._wl_drag_origin is not None and event.buttons() & Qt.RightButton:
                delta = event.position() - self._wl_drag_origin
                if abs(float(delta.x())) + abs(float(delta.y())) >= 5.0:
                    self._wl_dragged = True
                if self._wl_dragged:
                    level, width = self._wl_drag_start
                    scale = max(width, 1.0)
                    window = self.window()
                    window.window_level_spin.setValue(level - float(delta.y()) * scale / 240.0)
                    window.dynamic_range_spin.setValue(max(1e-6, width + float(delta.x()) * scale / 170.0))
                    window._manual_levels_changed()
                event.accept()
                return True

            if event_type == QEvent.MouseButtonRelease and event.button() == Qt.RightButton and self._wl_drag_origin is not None:
                show_menu = not self._wl_dragged
                global_pos = self.plot.viewport().mapToGlobal(event.position().toPoint())
                self._wl_drag_origin = None
                self._wl_drag_start = None
                self._wl_dragged = False
                if show_menu:
                    self._show_mouse_action_menu(global_pos)
                event.accept()
                return True

            if event_type == QEvent.Wheel:
                delta = event.angleDelta().y()
                if delta:
                    self.pageRequested.emit(1 if delta < 0 else -1)
                event.accept()
                return True

        return super().eventFilter(watched, event)

    def _scene_mouse_clicked(self, event):
        if event.button() == Qt.RightButton:
            event.accept()
            self._show_mouse_action_menu(event.screenPos().toPoint())

    def _show_mouse_action_menu(self, global_pos):
        """General-purpose medical image viewer context menu."""
        menu = QMenu(self)

        navigation_menu = menu.addMenu("Navigate")
        previous_action = navigation_menu.addAction("Previous Image")
        next_action = navigation_menu.addAction("Next Image")
        previous_action.setShortcut(QKeySequence(Qt.Key_Up))
        next_action.setShortcut(QKeySequence(Qt.Key_Down))

        view_menu = menu.addMenu("View")
        fit_action = view_menu.addAction("Fit Image to View")
        actual_action = view_menu.addAction("Actual Pixels (1:1)")
        zoom_in_action = view_menu.addAction("Zoom In")
        zoom_out_action = view_menu.addAction("Zoom Out")
        crosshair_action = view_menu.addAction("Show Crosshair")
        crosshair_action.setCheckable(True)
        crosshair_action.setChecked(self.hline.isVisible() and self.vline.isVisible())

        wl_menu = menu.addMenu("Window / Level")
        auto_action = wl_menu.addAction("Auto")
        wide_action = wl_menu.addAction("Wide")
        soft_action = wl_menu.addAction("Soft Tissue")
        contrast_action = wl_menu.addAction("High Contrast")
        narrow_action = wl_menu.addAction("Narrow")

        transform_menu = menu.addMenu("Image Processing")
        fft_action = transform_menu.addAction("FFT Current Image")
        back_fft_action = None
        if getattr(self.window(), "fft_view_active", False):
            back_fft_action = transform_menu.addAction("Back from FFT")

        menu.addSeparator()
        copy_action = menu.addAction("Copy Displayed Image")
        save_action = menu.addAction("Save Displayed Image As...")

        info_menu = menu.addMenu("Image Information")
        header_action = info_menu.addAction("DICOM Header")
        orientation_action = info_menu.addAction("Orientation")
        header_action.setEnabled(getattr(self.window(), "current_ds", None) is not None)

        selected = menu.exec(global_pos)
        if selected is None:
            return
        if selected == previous_action:
            self.pageRequested.emit(-1)
        elif selected == next_action:
            self.pageRequested.emit(1)
        elif selected == fit_action:
            self.fit_to_image()
        elif selected == actual_action:
            self._show_actual_pixels()
        elif selected == zoom_in_action:
            self.plot.getViewBox().scaleBy((0.8, 0.8))
        elif selected == zoom_out_action:
            self.plot.getViewBox().scaleBy((1.25, 1.25))
        elif selected == crosshair_action:
            self._set_crosshair_visible(crosshair_action.isChecked())
        elif selected == auto_action:
            self.mouseActionChanged.emit("Auto Window/Level")
        elif selected == wide_action:
            self.mouseActionChanged.emit("Preset:Wide")
        elif selected == soft_action:
            self.mouseActionChanged.emit("Preset:Soft Tissue")
        elif selected == contrast_action:
            self.mouseActionChanged.emit("Preset:High Contrast")
        elif selected == narrow_action:
            self.mouseActionChanged.emit("Preset:Narrow")
        elif selected == fft_action:
            self.viewRequested.emit("FFT_CURRENT")
        elif back_fft_action is not None and selected == back_fft_action:
            self.viewRequested.emit("BACK_FFT")
        elif selected == copy_action:
            self._copy_displayed_image()
        elif selected == save_action:
            self._save_displayed_image()
        elif selected == header_action:
            self.utilityRequested.emit("DICOM_HEADER")
        elif selected == orientation_action:
            self.utilityRequested.emit("ORIENTATION")

    def _displayed_qimage(self):
        array = getattr(self.image_item, "image", None)
        if array is None:
            return QImage()
        image = np.asarray(array, dtype=float)
        if image.ndim > 2:
            image = np.abs(image[..., 0])
        finite = image[np.isfinite(image)]
        if finite.size == 0:
            return QImage()
        if self.current_levels is not None:
            low, high = self.current_levels
        else:
            low, high = np.percentile(finite, [1.0, 99.0])
        if high <= low:
            high = low + 1.0
        normalized = np.clip((image - low) / (high - low), 0.0, 1.0)
        pixels = np.ascontiguousarray(np.nan_to_num(normalized) * 255.0, dtype=np.uint8)
        qimage = QImage(
            pixels.data, pixels.shape[1], pixels.shape[0], pixels.strides[0],
            QImage.Format_Grayscale8
        )
        return qimage.copy()

    def _copy_displayed_image(self):
        image = self._displayed_qimage()
        if not image.isNull():
            QApplication.clipboard().setImage(image)

    def _save_displayed_image(self):
        image = self._displayed_qimage()
        if image.isNull():
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Displayed Image", "displayed_image.png",
            "PNG Image (*.png);;JPEG Image (*.jpg *.jpeg);;Bitmap (*.bmp)"
        )
        if path:
            image.save(path)

    def _set_crosshair_visible(self, visible):
        self.hline.setVisible(bool(visible))
        self.vline.setVisible(bool(visible))

    def _show_actual_pixels(self):
        rows, cols = self.shape
        view = self.plot.getViewBox()
        center_x = cols / 2.0
        center_y = rows / 2.0
        view.setRange(
            xRange=(center_x - cols / 2.0, center_x + cols / 2.0),
            yRange=(center_y - rows / 2.0, center_y + rows / 2.0),
            padding=0,
        )

    def wheelEvent(self, event):
        # Fallback for wheel events delivered to the panel rather than the
        # plot viewport. Wheel rotation is always image navigation; zoom uses
        # middle-button drag and therefore needs no keyboard modifier.
        delta = event.angleDelta().y()
        if delta:
            self.pageRequested.emit(1 if delta < 0 else -1)
        event.accept()


    def fit_to_image(self):
        """Fit the current image using a deterministic range, never autoRange()."""
        rows, cols = self.shape
        if rows <= 0 or cols <= 0:
            return
        view = self.plot.getViewBox()
        view.disableAutoRange()
        view.setRange(
            xRange=(0.0, float(cols)),
            yRange=(0.0, float(rows)),
            padding=0.0,
            update=False,
        )

    def set_image(self, image: np.ndarray, auto_levels=True, levels=None):
        """Replace the displayed frame without exposing an intermediate range.

        Commit0073: the view frame itself is locked.  The final range is set
        before ImageItem receives pixels, and all later automatic range changes
        are disabled.  This prevents the visible left/right elastic frame motion
        during slice navigation.
        """
        image = np.asarray(image, dtype=float)
        self.shape = image.shape[:2]
        self.current_levels = levels
        view = self.plot.getViewBox()
        self.setUpdatesEnabled(False)
        self.plot.viewport().setUpdatesEnabled(False)
        try:
            view.disableAutoRange()
            view.setRange(
                xRange=(0.0, float(self.shape[1])),
                yRange=(0.0, float(self.shape[0])),
                padding=0.0,
                update=False,
            )
            self.image_item.setImage(
                image,
                autoLevels=auto_levels if levels is None else False,
                levels=levels,
            )
            self.set_line_position(self.shape[0] // 2, self.shape[1] // 2)
            self._update_frequency_marker()
        finally:
            self.plot.viewport().setUpdatesEnabled(True)
            self.setUpdatesEnabled(True)
        self.plot.viewport().update()

    def set_levels(self, low: float, high: float):
        if high <= low:
            high = low + 1.0
        self.current_levels = (float(low), float(high))
        self.image_item.setLevels(self.current_levels)

    def show_comp_roi(self):
        rows, cols = self.shape
        width = max(8, cols // 5)
        height = max(8, rows // 5)
        self.comp_roi.setPos((cols - width) / 2, (rows - height) / 2)
        self.comp_roi.setSize((width, height))
        self.comp_roi.show()

    def hide_comp_roi(self):
        self.comp_roi.hide()

    def comp_roi_bounds(self):
        pos = self.comp_roi.pos()
        size = self.comp_roi.size()
        x0 = max(0, min(int(round(pos.x())), self.shape[1] - 1))
        y0 = max(0, min(int(round(pos.y())), self.shape[0] - 1))
        x1 = max(x0 + 1, min(int(round(pos.x() + size.x())), self.shape[1]))
        y1 = max(y0 + 1, min(int(round(pos.y() + size.y())), self.shape[0]))
        return y0, y1, x0, x1

    def set_line_position(self, row: int, col: int):
        row = max(0, min(int(row), self.shape[0] - 1))
        col = max(0, min(int(col), self.shape[1] - 1))
        self.hline.blockSignals(True); self.vline.blockSignals(True)
        self.hline.setPos(row); self.vline.setPos(col)
        self.hline.blockSignals(False); self.vline.blockSignals(False)

    def _emit_position(self):
        row = max(0, min(int(round(self.hline.value())), self.shape[0] - 1))
        col = max(0, min(int(round(self.vline.value())), self.shape[1] - 1))
        self.lineChanged.emit(row, col)



class CompensationComparisonDialog(QDialog):
    """Synchronized Before/After and Difference viewer for RAW compensation."""

    def __init__(self, before_image, after_image, difference_image, difference_fft, difference_phase, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Compensation Difference / Before-After")
        self.resize(1500, 900)
        self._syncing_view = False
        self._syncing_cursor = False
        self.window_level_spin = QDoubleSpinBox(self)
        self.dynamic_range_spin = QDoubleSpinBox(self)
        self.window_level_spin.setRange(-1e15, 1e15)
        self.dynamic_range_spin.setRange(1e-9, 1e15)

        root = QVBoxLayout(self)
        toolbar = QHBoxLayout()
        self.sync_view_check = QCheckBox("Synchronize zoom / pan")
        self.sync_view_check.setChecked(True)
        self.sync_cursor_check = QCheckBox("Synchronize cursor")
        self.sync_cursor_check.setChecked(False)
        self.sync_cursor_check.setToolTip("Off by default. Enable only when crosshair comparison is needed.")
        self.sync_levels_check = QCheckBox("Synchronize Before / After WW/WL")
        self.sync_levels_check.setChecked(True)
        fit_button = QPushButton("Fit All")
        fit_button.clicked.connect(self._fit_all)
        toolbar.addWidget(self.sync_view_check)
        toolbar.addWidget(self.sync_cursor_check)
        toolbar.addWidget(self.sync_levels_check)
        toolbar.addStretch(1)
        toolbar.addWidget(fit_button)
        root.addLayout(toolbar)

        top = QSplitter(Qt.Horizontal)
        bottom = QSplitter(Qt.Horizontal)
        self.before_panel = ImagePanel("Before — Reconstructed Image")
        self.after_panel = ImagePanel("After — Reconstructed Image")
        self.diff_image_panel = ImagePanel("Difference Image |After − Before|")
        self.diff_fft_panel = ImagePanel("Difference FFT |Δk-space|")
        self.diff_phase_panel = ImagePanel("Difference Phase |Δphase|")
        self.panels = [self.before_panel, self.after_panel, self.diff_image_panel, self.diff_fft_panel, self.diff_phase_panel]
        top.addWidget(self.before_panel); top.addWidget(self.after_panel)
        bottom.addWidget(self.diff_image_panel); bottom.addWidget(self.diff_fft_panel); bottom.addWidget(self.diff_phase_panel)
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(top); splitter.addWidget(bottom)
        splitter.setSizes([470, 350])
        root.addWidget(splitter, 1)

        before = np.asarray(before_image, dtype=float)
        after = np.asarray(after_image, dtype=float)
        finite = np.concatenate([before[np.isfinite(before)], after[np.isfinite(after)]])
        if finite.size:
            common_levels = tuple(float(v) for v in np.percentile(finite, [1.0, 99.5]))
            if common_levels[1] <= common_levels[0]:
                common_levels = (common_levels[0], common_levels[0] + 1.0)
        else:
            common_levels = (0.0, 1.0)
        self.window_level_spin.setValue((common_levels[0] + common_levels[1]) / 2.0)
        self.dynamic_range_spin.setValue(common_levels[1] - common_levels[0])
        self.before_panel.set_image(before, levels=common_levels)
        self.after_panel.set_image(after, levels=common_levels)
        self.diff_image_panel.set_image(np.asarray(difference_image, dtype=float), auto_levels=True)
        self.diff_fft_panel.set_image(np.log1p(np.asarray(difference_fft, dtype=float)), auto_levels=True)
        self.diff_phase_panel.set_image(np.asarray(difference_phase, dtype=float), auto_levels=True)

        for panel in self.panels:
            panel.plot.getViewBox().sigRangeChanged.connect(lambda _vb, ranges, source=panel: self._sync_range(source, ranges))
            panel.lineChanged.connect(lambda row, col, source=panel: self._sync_cursor(source, row, col))
            panel.mouseActionChanged.connect(lambda action, source=panel: self._comparison_mouse_action(source, action))
        self._fit_all()

    def _prepare_level_controls_for_panel(self, panel):
        levels = panel.current_levels
        if levels is None:
            image = np.asarray(getattr(panel.image_item, 'image', np.zeros((1, 1))), dtype=float)
            finite = image[np.isfinite(image)]
            levels = tuple(np.percentile(finite, [1.0, 99.0])) if finite.size else (0.0, 1.0)
        self.window_level_spin.setValue((float(levels[0]) + float(levels[1])) / 2.0)
        self.dynamic_range_spin.setValue(max(1e-9, float(levels[1]) - float(levels[0])))
        self._active_level_panel = panel

    def _manual_levels_changed(self):
        level = float(self.window_level_spin.value())
        width = max(1e-9, float(self.dynamic_range_spin.value()))
        levels = (level - width / 2.0, level + width / 2.0)
        target = getattr(self, '_active_level_panel', self.before_panel)
        if self.sync_levels_check.isChecked() and target in (self.before_panel, self.after_panel):
            self.before_panel.set_levels(*levels); self.after_panel.set_levels(*levels)
        else:
            target.set_levels(*levels)

    def _comparison_mouse_action(self, source, action):
        if action == "Auto Window/Level":
            image = np.asarray(getattr(source.image_item, 'image', np.zeros((1, 1))), dtype=float)
            finite = image[np.isfinite(image)]
            if finite.size:
                levels = tuple(float(v) for v in np.percentile(finite, [1.0, 99.0]))
                source.set_levels(*levels)
                if self.sync_levels_check.isChecked() and source in (self.before_panel, self.after_panel):
                    self.before_panel.set_levels(*levels); self.after_panel.set_levels(*levels)

    def _sync_range(self, source, ranges):
        if self._syncing_view or not self.sync_view_check.isChecked():
            return
        self._syncing_view = True
        try:
            x_range, y_range = ranges
            for panel in self.panels:
                if panel is source:
                    continue
                panel.plot.getViewBox().setRange(xRange=x_range, yRange=y_range, padding=0, update=False)
        finally:
            self._syncing_view = False

    def _sync_cursor(self, source, row, col):
        if self._syncing_cursor or not self.sync_cursor_check.isChecked():
            return
        self._syncing_cursor = True
        try:
            for panel in self.panels:
                if panel is not source:
                    panel.set_line_position(row, col)
        finally:
            self._syncing_cursor = False

    def _fit_all(self):
        self._syncing_view = True
        try:
            for panel in self.panels:
                panel.fit_to_image()
        finally:
            self._syncing_view = False




class ServiceProgressDialog(QFrame):
    canceled = Signal()

    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self.setObjectName("ServiceProgressOverlay")
        self.setFrameShape(QFrame.StyledPanel)
        self.setMinimumHeight(118)
        self.setMaximumHeight(142)
        self.setStyleSheet("""
            QFrame#ServiceProgressOverlay {
                background: #18212b;
                color: #f2f6fa;
                border: 2px solid #2faee5;
                border-radius: 10px;
            }
            QLabel {
                color: #f2f6fa;
                border: none;
                background: transparent;
            }
            QProgressBar {
                min-height: 28px;
                max-height: 32px;
                border: 1px solid #55758e;
                border-radius: 6px;
                text-align: center;
                background: #0e151c;
                color: white;
                font-size: 15px;
                font-weight: 700;
            }
            QProgressBar::chunk {
                background: #1ea7e8;
                border-radius: 5px;
            }
            QPushButton {
                min-height: 34px;
                max-height: 38px;
                min-width: 104px;
                font-size: 14px;
                font-weight: 700;
                background: #2e4355;
                color: white;
                border: 1px solid #6c8da6;
                border-radius: 6px;
            }
            QPushButton:hover {
                background: #3b5770;
            }
            QPushButton:disabled {
                color: #8293a1;
                background: #25333f;
            }
        """)

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 10, 18, 10)
        root.setSpacing(7)

        # Top-aligned controls remain accessible even on a short screen.
        top_row = QHBoxLayout()
        top_row.setSpacing(12)

        self.title_label = QLabel(title)
        self.title_label.setStyleSheet(
            "font-size: 18px; font-weight: 800; color: #ffffff;"
        )
        top_row.addWidget(self.title_label, 1)

        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self._cancel)
        top_row.addWidget(self.cancel_button, 0, Qt.AlignTop)

        root.addLayout(top_row)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        root.addWidget(self.progress)

        # Only one information line is displayed. There is no second stage line.
        self.message = QLabel("")
        self.message.setWordWrap(False)
        self.message.setTextInteractionFlags(Qt.NoTextInteraction)
        self.message.setStyleSheet(
            "font-size: 13px; font-weight: 600; color: #dce8f1;"
        )
        self.message.setMinimumHeight(20)
        root.addWidget(self.message)

        self._canceled = False
        self.hide()

    def _cancel(self):
        self._canceled = True
        self.cancel_button.setEnabled(False)
        self.cancel_button.setText("Canceling...")
        self.canceled.emit()

    def wasCanceled(self):
        return self._canceled

    def setMaximum(self, maximum: int):
        self.progress.setMaximum(max(1, int(maximum)))

    def setValue(self, value: int):
        self.progress.setValue(int(value))

    def setLabelText(self, text: str):
        lines = [line.strip() for line in str(text).splitlines() if line.strip()]
        detail = lines[-1] if lines else ""
        if len(detail) > 104:
            detail = "..." + detail[-101:]
        self.message.setText(detail)

    def reset_for_use(self, title: str):
        self._canceled = False
        self.title_label.setText(title)
        self.cancel_button.setText("Cancel")
        self.cancel_button.setEnabled(True)
        self.message.clear()

    def raise_(self):
        super().raise_()

    def activateWindow(self):
        window = self.window()
        if window is not None:
            window.raise_()
            window.activateWindow()



class VerticalCurtainButton(QPushButton):
    """Accordion header that rotates clockwise while collapsed."""
    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.title = str(title)
        self.vertical = False
        self.setCheckable(True)
        self.setCursor(Qt.PointingHandCursor)

    def set_vertical(self, vertical):
        self.vertical = bool(vertical)
        if self.vertical:
            self.setFixedWidth(32)
            self.setMinimumHeight(150)
            self.setMaximumHeight(16777215)
        else:
            self.setMinimumWidth(0)
            self.setMaximumWidth(16777215)
            self.setFixedHeight(30)
        self.updateGeometry()
        self.update()

    def sizeHint(self):
        if self.vertical:
            return QSize(32, max(150, super().sizeHint().width() + 28))
        return super().sizeHint()

    def paintEvent(self, event):
        if not self.vertical:
            return super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)
        hovered = self.underMouse()
        painter.setBrush(QColor("#265f8e" if hovered else "#1e4b72"))
        painter.setPen(QPen(QColor("#37698f"), 1))
        painter.drawRoundedRect(self.rect().adjusted(1, 1, -1, -1), 4, 4)
        painter.setPen(QColor("#f2f7fb"))
        font = painter.font(); font.setBold(True); painter.setFont(font)
        painter.translate(self.width(), 0)
        painter.rotate(90)
        painter.drawText(
            0, 0, self.height(), self.width(),
            Qt.AlignCenter, "▶  " + self.title
        )



class AccordionSection(QWidget):
    opened = Signal(object)
    def __init__(self, title, content, expanded=False, parent=None, vertical_when_collapsed=False):
        super().__init__(parent); self.title=title; self.content=content
        self.vertical_when_collapsed = bool(vertical_when_collapsed)
        lay=QVBoxLayout(self); lay.setContentsMargins(0,0,0,0); lay.setSpacing(3)
        self.button=(VerticalCurtainButton(title) if self.vertical_when_collapsed else QPushButton())
        self.button.setCheckable(True); self.button.clicked.connect(self._toggle)
        self.button.setStyleSheet("QPushButton{text-align:left;min-height:30px;font-weight:700;background:#1e4b72;border:1px solid #37698f;border-radius:4px;padding:4px 8px} QPushButton:hover{background:#265f8e}")
        lay.addWidget(self.button); lay.addWidget(content); self.set_expanded(expanded,False)
    def _label(self):
        if self.vertical_when_collapsed and not self.button.isChecked():
            self.button.setText(self.title)
        else:
            self.button.setText(("▼" if self.button.isChecked() else "▶")+"  "+self.title)
    def _toggle(self): self.set_expanded(self.button.isChecked(),True)
    def set_expanded(self, expanded, notify=False):
        self.button.blockSignals(True); self.button.setChecked(bool(expanded)); self.button.blockSignals(False)
        if self.vertical_when_collapsed:
            self.button.set_vertical(not bool(expanded))
        self.content.setVisible(bool(expanded)); self._label()
        if expanded and notify: self.opened.emit(self)

class DicomHeaderDialog(QDialog):
    def __init__(self, ds, source_path, parent=None):
        super().__init__(parent); self.ds=copy.deepcopy(ds) if ds is not None else None; self.source_path=Path(source_path) if source_path else None; self._editable=False
        self.setWindowTitle('DICOM Header'); self.resize(1120,720)
        root=QVBoxLayout(self); row=QHBoxLayout(); self.filter_edit=QLineEdit(); self.filter_edit.setPlaceholderText('Filter DICOM header...'); self.filter_edit.textChanged.connect(self._filter); row.addWidget(self.filter_edit,1)
        for label,cb in [('Edit',self.enable_edit),('Save',self.save_edited),('Close',self.close)]: b=QPushButton(label); b.clicked.connect(cb); row.addWidget(b)
        root.addLayout(row); self.table=QTableWidget(0,6); self.table.setHorizontalHeaderLabels(['Tag','Keyword','Name','VR','VM','Value']); self.table.horizontalHeader().setStretchLastSection(True); self.table.setSelectionMode(QTableWidget.ExtendedSelection); self.table.setContextMenuPolicy(Qt.CustomContextMenu); self.table.customContextMenuRequested.connect(self._menu); root.addWidget(self.table,1); self.status=QLabel('Read-only. Press Edit to change values.'); root.addWidget(self.status); self._populate()
    def _populate(self):
        rows=[]
        if self.ds:
            for e in self.ds.iterall():
                if e.tag.group!=0x7FE0: rows.append((str(e.tag),e.keyword,e.name,e.VR,str(e.VM),str(e.value)))
        self.table.setRowCount(len(rows))
        for r,vals in enumerate(rows):
            for c,v in enumerate(vals):
                it=QTableWidgetItem(v); it.setFlags(it.flags() & ~Qt.ItemIsEditable); self.table.setItem(r,c,it)
        self.table.resizeColumnsToContents()
    def enable_edit(self):
        self._editable=True
        for r in range(self.table.rowCount()):
            it=self.table.item(r,5)
            if it: it.setFlags(it.flags()|Qt.ItemIsEditable)
        self.status.setText('Edit mode enabled. Value column is editable.')
    def _menu(self,pos):
        m=QMenu(self); a1=m.addAction('Copy Selection'); a2=m.addAction('Copy Value'); a3=m.addAction('Copy Row'); a4=m.addAction('Select All'); a=m.exec(self.table.viewport().mapToGlobal(pos))
        if a==a4: self.table.selectAll(); return
        cur=self.table.currentItem()
        if a==a1: QApplication.clipboard().setText('\n'.join(i.text() for i in self.table.selectedItems()))
        elif a==a2 and cur: QApplication.clipboard().setText(self.table.item(cur.row(),5).text())
        elif a==a3 and cur: QApplication.clipboard().setText('\t'.join(self.table.item(cur.row(),c).text() if self.table.item(cur.row(),c) else '' for c in range(6)))
    def _filter(self,q):
        q=q.lower()
        for r in range(self.table.rowCount()): self.table.setRowHidden(r, not any(q in (self.table.item(r,c).text().lower() if self.table.item(r,c) else '') for c in range(6)))
    def _next_path(self):
        folder=(self.source_path.parent if self.source_path else Path.cwd())/'Edited'; folder.mkdir(parents=True,exist_ok=True); base=self.source_path.stem if self.source_path else 'DICOM'; out=folder/f'{base}_Edit.dcm'; n=1
        while out.exists(): out=folder/f'{base}_Edit{n}.dcm'; n+=1
        return out
    def save_edited(self):
        if not self.ds or not self._editable: QMessageBox.information(self,'DICOM Header','Press Edit before saving.'); return
        changed=0
        for r in range(self.table.rowCount()):
            try:
                g,e=self.table.item(r,0).text().strip('()').split(','); tag=(int(g,16),int(e,16)); val=self.table.item(r,5).text()
                if tag in self.ds and str(self.ds[tag].value)!=val: self.ds[tag].value=val; changed+=1
            except Exception: pass
        out=self._next_path()
        try: self.ds.save_as(str(out))
        except Exception as exc: QMessageBox.critical(self,'Save Error',str(exc)); return
        self.status.setText(f'Saved {changed} changed value(s): {out}'); QMessageBox.information(self,'DICOM Header',f'Saved:\n{out}')

class ExplorerTreeWidget(QTreeWidget):
    """Explorer tree whose Up/Down keys use the same navigation as toolbar buttons.

    Qt's default QTreeWidget behaviour moves the current row onto a collapsed
    Series header.  That bypasses NavigationController, so no destination
    series expansion or image display occurs.  Emit a navigation request
    instead and let the existing continuous navigation path perform the move.
    Left/Right and modified selection keys retain the standard tree behaviour.
    """

    continuousNavigationRequested = Signal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        # High-resolution mouse wheels and touchpads may deliver deltas smaller
        # than one traditional 120-unit wheel step.  Accumulate them so one
        # physical notch/gesture advances exactly one image.
        self._navigation_wheel_remainder = 0

    def keyPressEvent(self, event):
        if (
            event.key() in (Qt.Key_Up, Qt.Key_Down)
            and event.modifiers() == Qt.NoModifier
        ):
            self.continuousNavigationRequested.emit(
                -1 if event.key() == Qt.Key_Up else 1
            )
            event.accept()
            return
        super().keyPressEvent(event)

    def wheelEvent(self, event):
        """Route an unmodified Explorer wheel gesture through navigation.

        Wheel up is Previous and wheel down is Next, matching the toolbar and
        Up/Down keyboard behaviour.  Modified wheel gestures retain Qt's
        normal tree scrolling/selection behaviour.
        """
        if event.modifiers() != Qt.NoModifier:
            super().wheelEvent(event)
            return

        delta = int(event.angleDelta().y())
        if delta == 0:
            # Some touchpads expose only pixelDelta.  A 40-pixel threshold is
            # intentionally conservative to prevent multiple accidental moves.
            pixel_delta = int(event.pixelDelta().y())
            if pixel_delta == 0:
                event.accept()
                return
            self._navigation_wheel_remainder += pixel_delta * 3
        else:
            self._navigation_wheel_remainder += delta

        moved = False
        while abs(self._navigation_wheel_remainder) >= 120:
            direction = -1 if self._navigation_wheel_remainder > 0 else 1
            self.continuousNavigationRequested.emit(direction)
            self._navigation_wheel_remainder += -120 if self._navigation_wheel_remainder > 0 else 120
            moved = True

        # Consume all unmodified wheel events so QTreeWidget never moves to a
        # collapsed Series header or scrolls independently of the displayed image.
        event.accept()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} {APP_VERSION}")
        self.stable_diagnostics = StableDiagnosticLogger(APP_VERSION)
        self.stable_diagnostics.info("MAIN_WINDOW_INIT")
        self.resize(980, 700)
        self.setMinimumSize(820, 560)
        self.setAcceptDrops(True)
        self.temp_dirs: list[Path] = []
        self.dicom_entries: list[DicomEntry] = []
        self.current_image: Optional[np.ndarray] = None
        self.current_kspace: Optional[np.ndarray] = None
        self.current_recon: Optional[np.ndarray] = None
        self.current_ds = None
        self.current_source = ""
        self.current_raw_result: Optional[RawImportResult] = None
        self.raw_preview_cache: dict[str, RawImportResult] = {}
        self.import_origin_info: dict[str, dict] = {}
        self._active_tree_source_key = None
        self._tree_open_in_progress = False
        self.orientation_override = {
            "top": None, "bottom": None, "left": None, "right": None
        }
        self.orientation_plane_override = "Auto"
        self.orientation_engine = OrientationEngine()
        self.annotation_visible = False
        self.annotation_mode = "Minimum"
        self.raw_display_mode = "Auto"
        self.view_mode = "Both"
        self.line_orientation = "Row"
        self.line_row = 0
        self.line_col = 0
        self.line_row_ratio = 0.5
        self.line_col_ratio = 0.5
        # Independent crosshair state for Original and FFT panels.
        self.panel_crosshairs = {
            "original": {"row_ratio": 0.5, "col_ratio": 0.5},
            "fft": {"row_ratio": 0.5, "col_ratio": 0.5},
        }
        self.signals: list[dict] = []
        self.imported_paths: list[Path] = []
        self.spike_results: list[dict] = []
        self.active_spike_indices = np.array([], dtype=int)
        self.tracker_matrix: Optional[np.ndarray] = None
        self.tracker_source_path: Optional[Path] = None
        self.tracker_files: list[dict] = []
        self.tracker_file_index: int = -1
        self.tracker_line_metrics: list[dict] = []
        self.tracker_selected_lines: list[int] = []
        self.tracker_position_measurements: list[DirectionMeasurement] = []
        self.tracker_reconstructed_image: Optional[np.ndarray] = None
        self.tracker_strongest_line: Optional[np.ndarray] = None
        self.tracker_strongest_line_index: int = -1
        self.tracker_shared_name: str = ""
        self.artifact_db: Optional[ArtifactDatabase] = None
        self.artifact_db_path: Optional[Path] = None
        self.artifact_selected_indices: list[int] = []
        self.normal_reference_image: Optional[np.ndarray] = None
        self.normal_reference_path: Optional[Path] = None
        self.compensation_preview: Optional[np.ndarray] = None
        self.roi_compensation_detection: Optional[RoiCompensationDetection] = None
        self.roi_compensation_detection_bounds: Optional[tuple[int, int, int, int]] = None
        self.compensation_original: Optional[np.ndarray] = None
        self.compensation_original_kspace: Optional[np.ndarray] = None
        self.compensation_history: list[dict] = []
        self.compensation_history_index: int = -1
        self.compensation_base_kspace: Optional[np.ndarray] = None
        self.compensation_base_image: Optional[np.ndarray] = None
        self.compensation_roi_panel: Optional[ImagePanel] = None
        self.manual_compensation_mask: Optional[np.ndarray] = None
        self.compensation_difference_fft: Optional[np.ndarray] = None
        self.compensation_difference_phase: Optional[np.ndarray] = None
        self.compensation_difference_image: Optional[np.ndarray] = None
        self.compensation_metrics: dict = {}
        self.addsub_preview_result: Optional[np.ndarray] = None
        self.addsub_preview_operation: str = ""
        self.addsub_a: Optional[dict] = None
        self.addsub_b: Optional[dict] = None
        self.processed_images: dict[str, np.ndarray] = {}
        self.processed_sources: dict[str, Path] = {}
        self.window_level: Optional[float] = None
        self.dynamic_range: Optional[float] = None
        self.original_window_level: Optional[float] = None
        self.original_dynamic_range: Optional[float] = None
        self.raw_window_level: Optional[float] = None
        self.raw_dynamic_range: Optional[float] = None
        self.level_preset = "Auto"
        self.source_kind = "none"
        self.output_root_override: Optional[Path] = None
        self.last_import_output_root: Optional[Path] = None
        self.current_mouse_action = "Standard"
        self.original_display_mode = "Magnitude"
        self.fft_display_mode = "Magnitude"
        self.fft_view_active = False
        self.fft_back_state = None
        self.fft_back_image: Optional[np.ndarray] = None
        self.quick_spike_threshold = 0.0
        self.tracker_analysis_text: dict[str, str] = {}
        self.import_in_progress = False
        # Bounded pixel cache: metadata stays resident, decoded arrays are evicted.
        # The value can be tuned for large studies without changing source.
        try:
            configured_cache = int(os.environ.get("MR_IMAGE_DICOM_CACHE", "24"))
        except ValueError:
            configured_cache = 24
        self.lazy_dicom_cache_limit = max(2, min(configured_cache, 256))
        self.lazy_dicom_cache = LRUKeySet[int](
            self.lazy_dicom_cache_limit, on_evict=self._evict_dicom_pixels
        )
        # FFT arrays are expensive but deterministic for an unchanged source image.
        # Keep this cache deliberately small because complex arrays are larger than pixels.
        try:
            configured_fft_cache = int(os.environ.get("MR_IMAGE_FFT_CACHE", "8"))
        except ValueError:
            configured_fft_cache = 8
        self.fft_cache_limit = max(1, min(configured_fft_cache, 64))
        self.fft_cache: LRUCache[tuple, tuple[np.ndarray, np.ndarray]] = LRUCache(
            self.fft_cache_limit
        )
        self.performance_metrics = {
            "dicom_decode_seconds": [],
            "fft_seconds": [],
            "render_seconds": [],
        }
        # Compatibility mirror retained for diagnostics and older tests.
        self.lazy_dicom_cache_order: list[int] = []
        self.pending_tracker_paths: list[Path] = []
        self.pending_first_dicom_index: Optional[int] = None
        self.import_thread: Optional[QThread] = None
        self.import_worker: Optional[ImportWorker] = None
        self.import_python_thread: Optional[threading.Thread] = None
        self.import_event_queue: queue.Queue = queue.Queue()
        self.import_cancel_event = threading.Event()
        self.import_poll_timer = QTimer(self)
        self.import_poll_timer.setInterval(50)
        self.import_poll_timer.timeout.connect(self._poll_import_queue)
        self.import_last_event_time = 0.0
        self.import_thread_started = False
        self.import_start_time = 0.0
        self.import_history: list[dict] = []
        self._responsive_pending = False
        self._initial_fit_done = False
        self._user_main_split_sizes: list[int] = []
        self._user_left_split_sizes: list[int] = []
        self._user_vertical_split_sizes: list[int] = []
        self._user_image_split_sizes: list[int] = []
        self.developer_mode = False
        self.spike_diagnostic_records: dict[int, dict] = {}
        self._screen_fit_completed = False
        self._import_target_tab = 0
        self.viewer_settings = QSettings("InSightec", "MR_Image_Explorer")
        # Always initialize navigation state before RAW/DICOM selection handlers.
        self.slice_index = 0
        self._hub_handoff_active = False

        self._build_ui()
        self._build_menu()
        self._apply_style()
        self.tabs.currentChanged.connect(self._main_tab_changed)
        self.statusBar().showMessage("Ready — drop files or folders above")
        QTimer.singleShot(0, self._restore_viewer_layout)
        QTimer.singleShot(0, self._initial_screen_fit)
        QTimer.singleShot(180, self._initial_screen_fit)
        QTimer.singleShot(420, self._initial_screen_fit)
        # Commit0118: ask about the guide/tour after the first window layout pass.
        if not RELEASE_MODE:
            QTimer.singleShot(650, self._handle_startup_guide_flow)

    def _main_tab_changed(self, index: int):
        """Handle tab-specific activation without resetting viewer layout state."""
        self._schedule_responsive_layout()
        if index == 1:
            # Let Qt finish changing tabs before starting image processing.
            QTimer.singleShot(0, self._activate_spike_diag_from_workspace_selection)

    def _build_ui(self):
        root = QWidget()
        root_l = QVBoxLayout(root)
        root_l.setContentsMargins(8, 8, 8, 8)
        root_l.setSpacing(6)
        self.root_layout = root_l

        self.drop_banner = DropBanner()
        self.drop_banner.pathsDropped.connect(self.request_import)
        self.drop_banner.clicked.connect(self.open_import_selection)
        root_l.addWidget(self.drop_banner)

        # This panel is a normal layout widget, not an overlay.
        # It cannot be hidden behind pyqtgraph/OpenGL/native child windows.
        self.import_progress_panel = ServiceProgressDialog(
            "Importing MRI Data",
            root,
        )
        self.import_progress_panel.hide()
        self.import_progress_panel.setSizePolicy(
            QSizePolicy.Expanding,
            QSizePolicy.Fixed,
        )
        root_l.addWidget(self.import_progress_panel)

        self.tabs = QTabWidget()
        self.tabs.setMovable(False)
        self.tabs.setDocumentMode(False)
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.addTab(self._build_workspace(), "Image Workspace")
        self.tabs.addTab(self._build_spike_results(), "Spike Diag")
        self.tabs.addTab(self._build_artifact_diag(), "Artifact Diag")
        self.tabs.addTab(self._build_tracker_signal_combined(), "Tracker Signal")
        self.tabs.addTab(self._build_signal_studio(), "1D Signal Studio")

        self.global_toolbar = QWidget()
        global_toolbar_layout = QHBoxLayout(self.global_toolbar)
        global_toolbar_layout.setContentsMargins(0, 2, 0, 2)
        global_toolbar_layout.setSpacing(5)

        self.global_import_button = QPushButton("Import")
        self.global_import_button.clicked.connect(
            self.open_import_selection
        )
        global_toolbar_layout.addWidget(self.global_import_button)

        self.display_reset_button = QPushButton("Display Reset")
        self.display_reset_button.clicked.connect(self.reset_viewer_display)
        global_toolbar_layout.addWidget(self.display_reset_button)

        # These operations affect the complete workspace, so they are kept
        # outside the image-only toolbar.
        for button in (
            self.previous_import_button,
            self.clear_selected_images_button,
            self.clear_all_images_button,
            self.header_popup_button,
            self.image_orientation_button,
            self.quick_spike_button,
        ):
            global_toolbar_layout.addWidget(button)

        global_toolbar_layout.addStretch(1)
        root_l.addWidget(self.global_toolbar)
        root_l.addWidget(self.tabs, 1)

        self.setCentralWidget(root)
        self.setStatusBar(QStatusBar())



    def _build_artifact_diag(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        banner = QGroupBox("Raw Data Compensation Diagnostic Engine")
        banner_lay = QHBoxLayout(banner)
        self.artifact_auto_status = QLabel(
            "Uses the Raw Data Compensation candidate detector to classify Spike, Blob, Block, Diagonal, Line, Band and Ring proposals."
        )
        self.artifact_auto_status.setWordWrap(True)
        run_button = QPushButton("Analyze Current Image")
        run_button.setToolTip("Run one non-destructive Raw Data Compensation candidate analysis on the current k-space data.")
        run_button.clicked.connect(self.run_artifact_diag_auto_analysis)
        banner_lay.addWidget(self.artifact_auto_status, 1)
        banner_lay.addWidget(run_button)
        lay.addWidget(banner)
        inner = QTabWidget()
        inner.addTab(self._build_artifact_detection(), 'Detection')
        inner.addTab(self._build_artifact_learning(), 'Learning')
        lay.addWidget(inner, 1)
        return page

    def _raw_compensation_diagnostic(self, kspace):
        """Run the shared Raw Data Compensation detector without applying it to the workspace."""
        result = run_auto_correct(
            np.asarray(kspace), threshold_sigma=3.0, sensitivity="Conservative",
            removal=55, detail=78, protection=85,
        )
        accepted = set(result.selected_types)
        candidates = []
        for record in result.candidates:
            item = dict(record)
            item.pop("mask", None)
            item["accepted"] = bool(item.get("accepted") or item.get("type") in accepted)
            candidates.append(item)
        return {
            "mask": np.asarray(result.mask, dtype=bool),
            "selected_types": list(result.selected_types),
            "candidates": candidates,
            "quality": float(result.metrics.get("overall_quality", 0.0)),
            "coverage": float(result.metrics.get("mask_coverage", 0.0)),
            "accepted_count": int(result.metrics.get("accepted_candidates", len(result.selected_types))),
        }

    def run_artifact_diag_auto_analysis(self):
        if self.current_kspace is None:
            QMessageBox.information(self, "Artifact Diag", "Load and select an image first.")
            return
        try:
            analysis = self._raw_compensation_diagnostic(self.current_kspace)
        except Exception as exc:
            QMessageBox.warning(self, "Artifact Diag", f"Analysis failed: {type(exc).__name__}: {exc}")
            return
        accepted = ", ".join(analysis["selected_types"]) or "None"
        rejected = [str(c.get("type", "Unknown")) for c in analysis["candidates"] if not c.get("accepted")]
        self.artifact_auto_status.setText(
            f"Accepted: {accepted} | Quality: {analysis['quality']:.1f} | "
            f"Mask coverage: {analysis['coverage']:.3f}% | Rejected proposals: {len(rejected)}"
        )
        QMessageBox.information(
            self, "Artifact Diag — Raw Compensation Analysis",
            f"Accepted candidate types: {accepted}\nQuality: {analysis['quality']:.1f}\n"
            f"Mask coverage: {analysis['coverage']:.3f}%\nRejected: {', '.join(rejected) or 'None'}"
        )

    def _build_tracker_signal_combined(self):
        page=QWidget(); lay=QVBoxLayout(page); inner=QTabWidget(); inner.addTab(self._build_tracker_explorer(),'Signal Explorer'); inner.addTab(self._build_tracker_position(),'Position'); lay.addWidget(inner); return page

    def _build_workspace(self):
        page = QWidget(); layout = QHBoxLayout(page); layout.setContentsMargins(0, 6, 0, 0)
        self.main_split = QSplitter(Qt.Horizontal)
        main_split = self.main_split

        left = QWidget()
        left.setMinimumWidth(90)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        self.tree = ExplorerTreeWidget()
        self.tree.setHeaderLabel("Explorer")
        self.tree.setSelectionMode(QTreeWidget.ExtendedSelection)
        self.tree.setDragEnabled(True)
        self.tree.setAcceptDrops(False)
        self.tree.setDragDropMode(QTreeWidget.DragOnly)
        self.tree.setDefaultDropAction(Qt.CopyAction)
        self.tree.setMinimumWidth(180)
        self.tree.setTextElideMode(Qt.ElideRight)
        self.tree.setUniformRowHeights(True)
        self.tree.setIndentation(12)
        self.tree.setAnimated(False)
        self.tree.setMouseTracking(False)
        self.tree.setStyleSheet("""
            QTreeWidget::item:selected {
                background: rgba(55, 165, 255, 110);
                color: white;
            }
            QTreeWidget::item:selected:active {
                background: rgba(55, 165, 255, 145);
            }
        """)
        self.tree.itemClicked.connect(self._tree_clicked)
        self.tree.startDrag = self._start_tree_file_drag
        self.tree.itemSelectionChanged.connect(self._tree_selection_changed)
        self.tree.currentItemChanged.connect(self._tree_current_changed)
        self.tree.continuousNavigationRequested.connect(
            self._navigate_explorer_keyboard
        )
        self.left_content_splitter = QSplitter(Qt.Vertical)
        self.left_content_splitter.setChildrenCollapsible(False)
        self.explorer_container = QWidget()
        explorer_container_layout = QVBoxLayout(self.explorer_container)
        explorer_container_layout.setContentsMargins(0, 0, 0, 0)
        explorer_container_layout.setSpacing(4)

        explorer_filter = QWidget()
        explorer_filter_layout = QGridLayout(explorer_filter)
        explorer_filter_layout.setContentsMargins(0, 0, 0, 2)
        explorer_filter_layout.setHorizontalSpacing(5)
        explorer_filter_layout.setVerticalSpacing(3)

        self.series_filter_combo = QComboBox()
        self.series_filter_combo.addItem("All Series")
        self.series_filter_combo.currentTextChanged.connect(
            self._apply_explorer_filters
        )

        self.explorer_search_edit = QLineEdit()
        self.explorer_search_edit.setPlaceholderText(
            "Search file / series / protocol / instance"
        )
        self.explorer_search_edit.setClearButtonEnabled(True)
        self.explorer_search_edit.textChanged.connect(
            self._apply_explorer_filters
        )

        self.explorer_type_combo = QComboBox()
        self.explorer_type_combo.addItems(
            ["All", "DICOM", "RAW", "Bitmap", "Tracking"]
        )
        self.explorer_type_combo.currentTextChanged.connect(
            self._apply_explorer_filters
        )

        self.explorer_sort_combo = QComboBox()
        self.explorer_sort_combo.addItems(
            ["Series / Instance", "Series / Filename"]
        )
        self.explorer_sort_combo.currentTextChanged.connect(
            self._rebuild_explorer_preserving_state
        )

        clear_filter_button = QPushButton("Clear Filter")
        clear_filter_button.clicked.connect(
            self._clear_explorer_filters
        )

        explorer_filter_layout.addWidget(QLabel("Series"), 0, 0)
        explorer_filter_layout.addWidget(self.series_filter_combo, 0, 1)
        explorer_filter_layout.addWidget(QLabel("Search"), 1, 0)
        explorer_filter_layout.addWidget(self.explorer_search_edit, 1, 1)
        explorer_filter_layout.addWidget(QLabel("Type"), 2, 0)
        explorer_filter_layout.addWidget(self.explorer_type_combo, 2, 1)
        explorer_filter_layout.addWidget(QLabel("Sort"), 3, 0)
        explorer_filter_layout.addWidget(self.explorer_sort_combo, 3, 1)
        explorer_filter_layout.addWidget(clear_filter_button, 4, 0, 1, 2)

        explorer_container_layout.addWidget(explorer_filter)
        explorer_container_layout.addWidget(self.tree, 1)
        self.left_content_splitter.addWidget(self.explorer_container)

        self.annotation_panel = QGroupBox("Annotation")
        annotation_layout = QVBoxLayout(self.annotation_panel)
        annotation_layout.setContentsMargins(6, 6, 6, 6)
        annotation_controls = QHBoxLayout()
        self.annotation_toggle = QCheckBox("Show")
        self.annotation_toggle.toggled.connect(self._annotation_visibility_changed)
        self.annotation_mode_combo = QComboBox()
        self.annotation_mode_combo.addItems(["Minimum", "Full"])
        self.annotation_mode_combo.currentTextChanged.connect(self._annotation_mode_changed)
        annotation_controls.addWidget(self.annotation_toggle)
        annotation_controls.addWidget(self.annotation_mode_combo)
        annotation_layout.addLayout(annotation_controls)

        self.annotation_text = QLabel("No annotation")
        self.annotation_text.setWordWrap(True)
        self.annotation_text.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.annotation_text.setObjectName("InfoCard")
        annotation_layout.addWidget(self.annotation_text, 1)
        self.annotation_panel.hide()
        self.left_content_splitter.addWidget(self.annotation_panel)
        self.left_content_splitter.setSizes([500, 0])

        self.message_panel = QGroupBox("Message / File Info")
        message_layout = QVBoxLayout(self.message_panel)
        message_layout.setContentsMargins(4, 4, 4, 4)

        self.info = QLabel("No file loaded")
        self.info.setWordWrap(True)
        self.info.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.info.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.info.setObjectName("InfoCard")

        self.info_scroll = QScrollArea()
        self.info_scroll.setWidgetResizable(True)
        self.info_scroll.setFrameShape(QFrame.NoFrame)
        self.info_scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarAsNeeded
        )
        self.info_scroll.setVerticalScrollBarPolicy(
            Qt.ScrollBarAsNeeded
        )
        self.info_scroll.setWidget(self.info)
        message_layout.addWidget(self.info_scroll)

        self.left_content_splitter.addWidget(self.message_panel)
        self.left_content_splitter.setCollapsible(0, False)
        self.left_content_splitter.setCollapsible(1, True)
        self.left_content_splitter.setCollapsible(2, False)
        self.left_content_splitter.setSizes([500, 0, 130])
        self.left_content_splitter.splitterMoved.connect(
            lambda _pos, _idx: self._remember_left_splitter_sizes()
        )

        ll.addWidget(self.left_content_splitter, 1)
        main_split.addWidget(left)

        center = QWidget(); cl = QVBoxLayout(center); cl.setContentsMargins(0, 0, 0, 0)
        self.workspace_source_combo = QComboBox()
        self.workspace_source_combo.addItems(["Current Image"])
        self.workspace_source_combo.hide()

        self.viewer_toolbar = ViewerToolbar()
        self.viewer_toolbar.displayModeRequested.connect(self.set_view_mode)
        self.viewer_toolbar.previousRequested.connect(
            lambda: self.change_slice_continuous(-1)
        )
        self.viewer_toolbar.nextRequested.connect(
            lambda: self.change_slice_continuous(1)
        )
        # Compatibility aliases: existing display code can keep using the
        # original attribute names while the widgets now live in a dedicated
        # toolbar class.
        self.btn_fft = self.viewer_toolbar.fft_button
        self.btn_original = self.viewer_toolbar.original_button
        self.btn_both = self.viewer_toolbar.both_button
        self.prev_btn = self.viewer_toolbar.previous_button
        self.next_btn = self.viewer_toolbar.next_button
        self.slice_label = self.viewer_toolbar.slice_label
        self.display_state = DisplayState.capture(self)
        self.navigation_controller = NavigationController(self)

        self.clear_selected_images_button = QPushButton("Clear Selected")
        self.clear_all_images_button = QPushButton("Clear All Images")
        self.clear_selected_images_button.clicked.connect(
            self.clear_selected_images
        )
        self.clear_all_images_button.clicked.connect(
            self.clear_all_images
        )

        self.previous_import_button = QPushButton("Previous Import")
        self.previous_import_button.setEnabled(False)
        self.previous_import_button.setToolTip(
            "Restore the image list that existed before the latest import."
        )
        self.previous_import_button.clicked.connect(
            self.restore_previous_import
        )

        self.header_popup_button = QPushButton("DICOM Header")
        self.header_popup_button.clicked.connect(
            self.show_dicom_header_popup
        )

        self.image_orientation_button = QPushButton("Orientation")
        self.image_orientation_button.setToolTip(
            "Display DICOM orientation by default or manually edit "
            "R/L/A/P/H/F."
        )
        self.image_orientation_button.clicked.connect(
            self._open_orientation_dialog
        )
        self.image_orientation_button.setEnabled(True)

        self.quick_spike_button = QPushButton("Quick Spike Detect")
        self.quick_spike_button.clicked.connect(
            self.quick_spike_detect_prototype
        )

        self.mode_toolbar_widget = self.viewer_toolbar
        # Phase 1: the display/navigation toolbar is intentionally not added
        # to the image-column layout.  It is mounted above the complete
        # image + tools workspace after both columns are constructed.

        self.vertical_splitter = QSplitter(Qt.Vertical)
        image_container = QWidget(); il = QVBoxLayout(image_container); il.setContentsMargins(0, 0, 0, 0)
        self.image_splitter = QSplitter(Qt.Horizontal)
        # Original is left, FFT is right.
        self.primary_panel = ImagePanel("Original")
        self.secondary_panel = ImagePanel("FFT (k-space)")
        self.primary_panel.lineChanged.connect(
            lambda row, col: self._line_moved(self.primary_panel, row, col)
        )
        self.secondary_panel.lineChanged.connect(
            lambda row, col: self._line_moved(self.secondary_panel, row, col)
        )
        self.primary_panel.pageRequested.connect(self.change_slice_continuous)
        self.secondary_panel.pageRequested.connect(self.change_slice_continuous)
        self.primary_panel.utilityRequested.connect(self._image_panel_utility_requested)
        self.secondary_panel.utilityRequested.connect(self._image_panel_utility_requested)
        self.primary_panel.mouseActionChanged.connect(self._image_mouse_action_changed)
        self.secondary_panel.mouseActionChanged.connect(self._image_mouse_action_changed)
        self.primary_panel.viewRequested.connect(self._image_view_requested)
        self.secondary_panel.viewRequested.connect(self._image_view_requested)
        self.primary_panel.levelWheel.connect(
            lambda step, width: self._image_level_wheel(self.primary_panel, step, width)
        )
        self.secondary_panel.levelWheel.connect(
            lambda step, width: self._image_level_wheel(self.secondary_panel, step, width)
        )
        self.primary_panel.componentRequested.connect(
            lambda component: self._set_panel_component(self.primary_panel, component)
        )
        self.secondary_panel.componentRequested.connect(
            lambda component: self._set_panel_component(self.secondary_panel, component)
        )
        self.primary_panel.keyboardPageRequested.connect(self.change_slice_continuous)
        self.secondary_panel.keyboardPageRequested.connect(self.change_slice_continuous)
        self.primary_panel.activated.connect(self._set_active_image_panel)
        self.secondary_panel.activated.connect(self._set_active_image_panel)
        self.active_image_panel = self.primary_panel
        self.image_splitter.addWidget(self.primary_panel)
        self.image_splitter.addWidget(self.secondary_panel)
        self.image_splitter.setChildrenCollapsible(False)
        self.image_splitter.splitterMoved.connect(
            lambda _pos, _idx: self._remember_splitter_sizes()
        )
        il.addWidget(self.image_splitter, 1)

        line_bar = QHBoxLayout()
        line_bar.addWidget(QLabel("Line"))
        self.orientation_combo = QComboBox(); self.orientation_combo.addItems(["Row", "Column"])
        self.orientation_combo.currentTextChanged.connect(self._orientation_changed)
        line_bar.addWidget(self.orientation_combo)
        self.line_spin = QSpinBox(); self.line_spin.valueChanged.connect(self._spin_line_changed)
        line_bar.addWidget(self.line_spin)
        self.line_slider = QSlider(Qt.Horizontal); self.line_slider.valueChanged.connect(self._slider_line_changed)
        line_bar.addWidget(self.line_slider, 1)
        self.line_value_label = QLabel("-")
        line_bar.addWidget(self.line_value_label)
        il.addLayout(line_bar)
        self.vertical_splitter.addWidget(image_container)

        lower = QWidget(); low = QHBoxLayout(lower); low.setContentsMargins(0, 0, 0, 0)
        original_profiles = QWidget(); original_profiles_layout = QVBoxLayout(original_profiles)
        original_profiles_layout.setContentsMargins(0, 0, 0, 0)
        fft_profiles = QWidget(); fft_profiles_layout = QVBoxLayout(fft_profiles)
        fft_profiles_layout.setContentsMargins(0, 0, 0, 0)

        self.original_horizontal_profile = pg.PlotWidget(title="Original — Horizontal Crosshair Profile")
        self.original_vertical_profile = pg.PlotWidget(title="Original — Vertical Crosshair Profile")
        self.fft_horizontal_profile = pg.PlotWidget(title="FFT — Horizontal Crosshair Profile")
        self.fft_vertical_profile = pg.PlotWidget(title="FFT — Vertical Crosshair Profile")
        for plot in (
            self.original_horizontal_profile, self.original_vertical_profile,
            self.fft_horizontal_profile, self.fft_vertical_profile,
        ):
            plot.showGrid(x=True, y=True, alpha=.2)
            plot.setLabel("bottom", "Sample")
            plot.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        original_profiles_layout.addWidget(self.original_horizontal_profile, 1)
        original_profiles_layout.addWidget(self.original_vertical_profile, 1)
        fft_profiles_layout.addWidget(self.fft_horizontal_profile, 1)
        fft_profiles_layout.addWidget(self.fft_vertical_profile, 1)
        low.addWidget(original_profiles, 1)
        low.addWidget(fft_profiles, 1)
        # Backward-compatible aliases used by reset/cleanup paths.
        self.profile_plot = self.original_horizontal_profile
        self.tracker_plot = self.original_vertical_profile
        # The complete four-chart area is a curtain. It starts closed.
        self.profile_accordion = AccordionSection(
            "Crosshair Profile Charts", lower, False
        )
        self.profile_accordion.button.toggled.connect(
            self._profile_curtain_toggled
        )
        # Crosshair is visible only while Crosshair Profile Charts is expanded.
        self.primary_panel._set_crosshair_visible(False)
        self.secondary_panel._set_crosshair_visible(False)
        self.vertical_splitter.addWidget(self.profile_accordion)
        self.vertical_splitter.setCollapsible(0, False)
        self.vertical_splitter.setCollapsible(1, False)
        self.vertical_splitter.setSizes([620, 250])
        self.vertical_splitter.splitterMoved.connect(
            lambda _pos, _idx: self._remember_splitter_sizes()
        )
        cl.addWidget(self.vertical_splitter, 1)

        right = QWidget()
        right.setMinimumWidth(320)
        right.setMaximumWidth(520)
        rl = QVBoxLayout(right)
        rl.setContentsMargins(4, 0, 4, 0)
        rl.setSpacing(6)
        right.setStyleSheet("""
            QComboBox, QDoubleSpinBox, QSpinBox {
                min-height: 30px;
            }
            QPushButton {
                min-height: 34px;
                padding: 4px 8px;
            }
            QGroupBox {
                margin-top: 8px;
            }
        """)
        # Component selection now lives directly on each image title.
        # Keep a hidden compatibility selector for line-profile code paths.
        self.profile_mode = QComboBox()
        self.profile_mode.addItems(["Magnitude", "Real", "Imaginary", "Phase"])
        self.profile_mode.hide()
        self.raw_display_combo = QComboBox()
        self.raw_display_combo.addItems(
            ["Auto", "Reconstructed Image", "Direct Array", "k-space Magnitude"]
        )
        self.raw_display_combo.setToolTip(
            "RAW Auto compares direct-array and reconstructed-image views."
        )
        self.raw_display_combo.currentTextChanged.connect(
            self._change_raw_display_mode
        )
        self.raw_display_combo.setEnabled(False)
        self.raw_display_combo.hide()
        self.log_profile = QCheckBox("Log scale")
        self.log_profile.toggled.connect(self.update_line_profile)
        self.log_profile.hide()
        self.btn_send_signal = QPushButton("Send selected line to 1D Studio")
        self.btn_send_signal.clicked.connect(self.send_line_to_signal_studio)
        self.btn_auto = QPushButton("Auto Levels"); self.btn_auto.clicked.connect(self.auto_levels)
        image_tools_content = QFrame()
        image_tools_layout = QVBoxLayout(image_tools_content)
        image_tools_layout.setContentsMargins(4, 4, 4, 4)
        image_tools_layout.addWidget(self.btn_auto)
        image_tools_layout.addWidget(self.btn_send_signal)
        spike_content=QFrame(); spike_layout=QFormLayout(spike_content)
        self.spike_range_combo=QComboBox(); self.spike_range_combo.addItems(["Large","Mid","Small","Scale"]); self.spike_range_combo.currentTextChanged.connect(self._spike_range_changed); spike_layout.addRow("Range",self.spike_range_combo)
        self.spike_scale_slider=QSlider(Qt.Horizontal); self.spike_scale_slider.setRange(10,100); self.spike_scale_slider.setValue(70); self.spike_scale_slider.setVisible(False); self.spike_scale_label=QLabel("70%"); self.spike_scale_label.setVisible(False); self.spike_scale_slider.valueChanged.connect(lambda v:self.spike_scale_label.setText(f"{v}%")); sr=QHBoxLayout(); sr.addWidget(self.spike_scale_slider,1); sr.addWidget(self.spike_scale_label); spike_layout.addRow("Scale",sr)
        self.spike_level_combo=QComboBox(); self.spike_level_combo.addItems(["Wide","Mid","Fine"]); spike_layout.addRow("Level",self.spike_level_combo)
        self.btn_spike=QPushButton("Apply Spike Processing"); self.btn_spike.clicked.connect(self.apply_spike_processing); self.btn_spike.setStyleSheet("QPushButton{background:#7a2430;font-weight:700;min-height:30px} QPushButton:hover{background:#a23242}"); spike_layout.addRow(self.btn_spike)

        display_group = QGroupBox("Dynamic Range / Window Level")
        display_form = QFormLayout(display_group)
        self.level_target_combo = QComboBox()
        self.level_target_combo.addItems(["Original Image", "Raw Data"])
        self.level_target_combo.currentTextChanged.connect(self._level_target_changed)
        display_form.addRow("Adjust", self.level_target_combo)

        self.level_preset_combo = QComboBox()
        self.level_preset_combo.addItems(["Auto", "Manual", "Wide", "Soft Tissue", "High Contrast", "Narrow"])
        self.level_preset_combo.currentTextChanged.connect(self.apply_level_preset)
        display_form.addRow("Preset", self.level_preset_combo)

        self.window_level_spin = QDoubleSpinBox()
        self.window_level_spin.setRange(-1e12, 1e12)
        self.window_level_spin.setDecimals(3)
        self.window_level_spin.valueChanged.connect(self._manual_levels_changed)
        display_form.addRow("Window level", self.window_level_spin)

        self.dynamic_range_spin = QDoubleSpinBox()
        self.dynamic_range_spin.setRange(1e-9, 1e12)
        self.dynamic_range_spin.setDecimals(3)
        self.dynamic_range_spin.valueChanged.connect(self._manual_levels_changed)
        display_form.addRow("Dynamic range", self.dynamic_range_spin)

        addsub_group = QGroupBox("Image Add / Subtract")
        addsub_layout = QVBoxLayout(addsub_group)
        addsub_row1 = QHBoxLayout()
        self.set_a_button = QPushButton("Set Current as A")
        self.set_b_button = QPushButton("Set Current as B")
        self.set_a_button.clicked.connect(lambda: self.set_addsub_source("A"))
        self.set_b_button.clicked.connect(lambda: self.set_addsub_source("B"))
        addsub_row1.addWidget(self.set_a_button)
        addsub_row1.addWidget(self.set_b_button)
        addsub_layout.addLayout(addsub_row1)
        self.addsub_status = QLabel("A: -\nB: -")
        self.addsub_status.setWordWrap(True)
        addsub_layout.addWidget(self.addsub_status)
        addsub_row2 = QHBoxLayout()
        self.add_button = QPushButton("Preview A + B")
        self.subtract_button = QPushButton("Preview A - B")
        self.add_button.clicked.connect(lambda: self.preview_addsub("add"))
        self.subtract_button.clicked.connect(lambda: self.preview_addsub("subtract"))
        addsub_row2.addWidget(self.add_button)
        addsub_row2.addWidget(self.subtract_button)
        addsub_layout.addLayout(addsub_row2)
        self.save_addsub_button = QPushButton("Save Preview Result")
        self.save_addsub_button.clicked.connect(self.save_addsub_preview)
        self.save_addsub_button.setEnabled(False)
        addsub_layout.addWidget(self.save_addsub_button)
        addsub_clear_row = QHBoxLayout()
        for label, callback in [
            ("Clear A", lambda: self.clear_addsub_slot("A")),
            ("Clear B", lambda: self.clear_addsub_slot("B")),
            ("Clear Result", self.clear_addsub_result),
        ]:
            button = QPushButton(label); button.clicked.connect(callback); addsub_clear_row.addWidget(button)
        addsub_layout.addLayout(addsub_clear_row)

        comp_group = QGroupBox("ROI Raw Data Compensation")
        comp_layout = QVBoxLayout(comp_group)
        comp_layout.setContentsMargins(4, 4, 4, 4)
        self.comp_tabs = QTabWidget()
        self.comp_tabs.setDocumentMode(True)
        self.comp_tabs.setUsesScrollButtons(False)
        comp_layout.addWidget(self.comp_tabs)

        auto_tab = QWidget()
        auto_layout = QVBoxLayout(auto_tab)
        auto_layout.setContentsMargins(6, 6, 6, 6)
        auto_layout.setSpacing(6)
        paint_tab = QWidget()
        paint_layout = QVBoxLayout(paint_tab)
        paint_layout.setContentsMargins(6, 6, 6, 6)
        paint_layout.setSpacing(6)
        expert_tab = QScrollArea()
        expert_tab.setWidgetResizable(True)
        expert_tab.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        expert_content = QWidget()
        expert_layout = QVBoxLayout(expert_content)
        expert_layout.setContentsMargins(6, 6, 6, 6)
        expert_layout.setSpacing(6)
        expert_tab.setWidget(expert_content)
        self.comp_tabs.addTab(auto_tab, "Auto")
        self.comp_tabs.addTab(paint_tab, "Paint")
        self.comp_tabs.addTab(expert_tab, "Expert")
        guide_row = QHBoxLayout()
        guide_row.addStretch(1)
        self.comp_guide_button = QPushButton("? Guide Library")
        self.comp_guide_button.setToolTip("Open the guide library. Raw Data Compensation is available as an optional advanced guide.")
        if RELEASE_MODE:
            self.comp_guide_button.setVisible(False)
        else:
            self.comp_guide_button.clicked.connect(self.show_guide_library)
            guide_row.addWidget(self.comp_guide_button)
        comp_layout.addLayout(guide_row)
        comp_note = QLabel(
            "Manual ROI tool for structured RAW artifacts. "
            "Spike Noise Correction remains a separate function in Spike Diag."
        )
        comp_note.setWordWrap(True)
        paint_layout.addWidget(comp_note)
        self.select_comp_button = QPushButton("1. Start Manual Paint on Raw Data")
        self.detect_comp_button = QPushButton("2. Use Painted Mask")
        self.comp_clear_mask_button = QPushButton("Clear Paint")
        use_mask_row = QHBoxLayout()
        use_mask_row.addWidget(self.detect_comp_button, 1)
        use_mask_row.addWidget(self.comp_clear_mask_button)
        paint_row = QHBoxLayout()
        paint_row.addWidget(QLabel("Paint Tool"))
        self.comp_paint_tool_combo = QComboBox()
        self.comp_paint_tool_combo.addItems(["Brush", "Line", "Band", "Block", "Ring", "Eraser", "Remove Component"])
        paint_row.addWidget(self.comp_paint_tool_combo)
        paint_row.addWidget(QLabel("Size"))
        self.comp_brush_size_spin = QSpinBox()
        self.comp_brush_size_spin.setRange(1, 100)
        self.comp_brush_size_spin.setSingleStep(1)
        self.comp_brush_size_spin.setValue(9)
        self.comp_brush_size_spin.setKeyboardTracking(True)
        self.comp_brush_size_spin.setButtonSymbols(QSpinBox.UpDownArrows)
        self.comp_brush_size_spin.setMinimumWidth(82)
        self.comp_brush_size_spin.setStyleSheet(
            "QSpinBox { padding: 4px 26px 4px 6px; min-height: 26px; }"
            "QSpinBox::up-button { subcontrol-origin: border; subcontrol-position: top right; width: 24px; }"
            "QSpinBox::down-button { subcontrol-origin: border; subcontrol-position: bottom right; width: 24px; }"
        )
        self.comp_brush_size_down_button = QPushButton("▼")
        self.comp_brush_size_up_button = QPushButton("▲")
        for button in (self.comp_brush_size_down_button, self.comp_brush_size_up_button):
            button.setFixedSize(30, 30)
            button.setAutoRepeat(True)
            button.setAutoRepeatDelay(250)
            button.setAutoRepeatInterval(70)
        self.comp_brush_size_down_button.setToolTip("Decrease brush size")
        self.comp_brush_size_up_button.setToolTip("Increase brush size")
        self.comp_brush_size_down_button.clicked.connect(self.comp_brush_size_spin.stepDown)
        self.comp_brush_size_up_button.clicked.connect(self.comp_brush_size_spin.stepUp)
        paint_row.addWidget(self.comp_brush_size_spin)
        paint_row.addWidget(self.comp_brush_size_down_button)
        paint_row.addWidget(self.comp_brush_size_up_button)
        paint_layout.addLayout(paint_row)
        self.preview_comp_button = QPushButton("3. Preview Reconstructed Image")
        self.open_comp_comparison_button = QPushButton("Open Difference / Before-After")
        self.open_comp_comparison_button.setEnabled(False)
        self.apply_comp_button = QPushButton("4. Apply ROI Compensation")
        self.save_comp_button = QPushButton("Save Displayed Compensation")
        self.cancel_comp_button = QPushButton("Cancel Current ROI")
        self.select_comp_button.clicked.connect(self.start_compensation_roi)
        self.detect_comp_button.clicked.connect(self.detect_compensation_mask)
        self.comp_paint_tool_combo.currentTextChanged.connect(self._update_manual_paint_tool)
        self.comp_brush_size_spin.valueChanged.connect(self._update_manual_paint_tool)
        self.comp_clear_mask_button.clicked.connect(self.clear_manual_compensation_mask)
        self.comp_mask_undo_button = QPushButton("Undo Mask")
        self.comp_mask_redo_button = QPushButton("Redo Mask")
        self.comp_mask_symmetry_button = QPushButton("Symmetry Paint")
        self.comp_mask_expand_button = QPushButton("Expand Mask")
        self.comp_mask_shrink_button = QPushButton("Shrink Mask")
        self.comp_mask_fill_button = QPushButton("Fill Largest Region")
        self.comp_mask_delete_button = QPushButton("Delete Smallest Region")
        mask_edit_grid = QGridLayout()
        for index, button in enumerate((self.comp_mask_undo_button, self.comp_mask_redo_button, self.comp_mask_symmetry_button, self.comp_mask_expand_button, self.comp_mask_shrink_button, self.comp_mask_fill_button, self.comp_mask_delete_button)):
            mask_edit_grid.addWidget(button, index // 2, index % 2)
        paint_layout.addLayout(mask_edit_grid)
        self._comp_mask_undo_stack = []
        self._comp_mask_redo_stack = []
        self.comp_mask_undo_button.clicked.connect(self.undo_compensation_mask)
        self.comp_mask_redo_button.clicked.connect(self.redo_compensation_mask)
        self.comp_mask_symmetry_button.clicked.connect(self.symmetrize_compensation_mask)
        self.comp_mask_expand_button.clicked.connect(lambda: self.morph_compensation_mask(1))
        self.comp_mask_shrink_button.clicked.connect(lambda: self.morph_compensation_mask(-1))
        self.comp_mask_fill_button.clicked.connect(self.fill_largest_compensation_region)
        self.comp_mask_delete_button.clicked.connect(self.delete_smallest_compensation_region)
        self.preview_comp_button.clicked.connect(lambda: self.preview_compensation(open_comparison=True))
        self.open_comp_comparison_button.clicked.connect(self.open_compensation_comparison)
        self.apply_comp_button.clicked.connect(self.apply_compensation)
        self.save_comp_button.clicked.connect(self.save_compensation_history_state)
        self.cancel_comp_button.clicked.connect(self.clear_compensation_roi)
        paint_layout.addWidget(self.select_comp_button)
        paint_layout.addLayout(use_mask_row)

        auto_correct_group = QGroupBox("Auto Correct")
        auto_correct_layout = QVBoxLayout(auto_correct_group)
        auto_correct_row = QHBoxLayout()
        self.comp_auto_correct_button = QPushButton("Auto Correct")
        self.comp_auto_correct_button.setMinimumHeight(38)
        self.comp_auto_correct_button.setStyleSheet(
            "QPushButton{background:#1769aa;color:white;font-weight:700;font-size:14px;border-radius:4px;padding:6px;}"
            "QPushButton:hover{background:#2186d4;} QPushButton:disabled{background:#5d6670;color:#bfc5ca;}"
        )
        self.comp_show_auto_mask_check = QCheckBox("Show Generated Mask")
        self.comp_show_auto_mask_check.setChecked(True)
        self.comp_auto_more_button = QPushButton("More...")
        self.comp_auto_more_button.setVisible(False)
        self.comp_auto_more_menu = QMenu(self.comp_auto_more_button)
        self.comp_run_auto_again_action = self.comp_auto_more_menu.addAction("Run Auto Correct Again")
        self.comp_auto_more_button.setMenu(self.comp_auto_more_menu)
        auto_correct_row.addWidget(self.comp_auto_correct_button, 1)
        auto_correct_row.addWidget(self.comp_auto_more_button)
        auto_correct_row.addWidget(self.comp_show_auto_mask_check)
        auto_correct_layout.addLayout(auto_correct_row)
        self.comp_auto_quality_label = QLabel("Auto result: not calculated")
        self.comp_auto_quality_label.setWordWrap(True)
        auto_correct_layout.addWidget(self.comp_auto_quality_label)
        self.comp_session_status_label = QLabel("Session: no active mask")
        self.comp_session_status_label.setWordWrap(True)
        self.comp_session_status_label.setStyleSheet("QLabel{background:#20262d;color:#d8e6f3;border:1px solid #46525e;border-radius:3px;padding:5px;}")
        auto_correct_layout.addWidget(self.comp_session_status_label)
        quick_group = QGroupBox("Quick Adjust Controls")
        quick_form = QFormLayout(quick_group)
        quick_form.setContentsMargins(8, 4, 8, 6)
        self.comp_quick_removal = QSlider(Qt.Horizontal); self.comp_quick_removal.setRange(0,100); self.comp_quick_removal.setValue(50)
        self.comp_quick_detail = QSlider(Qt.Horizontal); self.comp_quick_detail.setRange(0,100); self.comp_quick_detail.setValue(75)
        self.comp_quick_protection = QSlider(Qt.Horizontal); self.comp_quick_protection.setRange(0,100); self.comp_quick_protection.setValue(85)
        quick_form.addRow("Artifact removal", self.comp_quick_removal)
        quick_form.addRow("Image detail", self.comp_quick_detail)
        quick_form.addRow("Protection", self.comp_quick_protection)
        quick_buttons = QGridLayout()
        self.comp_quick_recalculate_button = QPushButton("Apply Quick Adjust")
        self.comp_quick_review_button = QPushButton("Review Reconstructed Image")
        self.comp_restore_auto_button = QPushButton("Restore Auto Result")
        self.comp_expert_toggle_button = QPushButton("Expert Settings...")
        self.comp_quick_recalculate_button.setToolTip("With a mask: reconstruct once from the current mask. Without a mask: run one detection trial using the current Quick Adjust values. Auto Retry is not run.")
        self.comp_quick_review_button.setToolTip("Open the reconstructed-image review and before/after comparison.")
        quick_buttons.addWidget(self.comp_quick_recalculate_button, 0, 0)
        quick_buttons.addWidget(self.comp_quick_review_button, 0, 1)
        quick_buttons.addWidget(self.comp_restore_auto_button, 1, 0)
        quick_buttons.addWidget(self.comp_expert_toggle_button, 1, 1)
        quick_form.addRow(quick_buttons)
        self.comp_quick_accordion = AccordionSection("Quick Adjust", quick_group, False)
        auto_correct_layout.addWidget(self.comp_quick_accordion)
        auto_layout.addWidget(auto_correct_group)

        self.comp_next_step_group = QGroupBox("Next Step")
        next_step_layout = QGridLayout(self.comp_next_step_group)
        self.comp_result_ok_button = QPushButton("OK — Use This Result")
        self.comp_result_quick_button = QPushButton("Quick Adjust")
        self.comp_result_paint_button = QPushButton("Paint")
        self.comp_result_expert_button = QPushButton("Expert")
        self.comp_result_ok_button.setToolTip("Accept the current reconstructed result and proceed to the review/apply step.")
        self.comp_result_quick_button.setToolTip("Fine-tune Artifact Removal, Image Detail, and Protection.")
        self.comp_result_paint_button.setToolTip("Edit the generated mask directly with Brush, Line, Band, Block, Ring, or Eraser tools.")
        self.comp_result_expert_button.setToolTip("Research controls for detection type, thresholds, mask expansion, donor halo, pass count, strength, symmetry, and reconstruction options.")
        next_step_layout.addWidget(self.comp_result_ok_button, 0, 0)
        next_step_layout.addWidget(self.comp_result_quick_button, 0, 1)
        next_step_layout.addWidget(self.comp_result_paint_button, 1, 0)
        next_step_layout.addWidget(self.comp_result_expert_button, 1, 1)
        self.comp_next_step_group.setVisible(False)
        auto_layout.addWidget(self.comp_next_step_group)
        auto_layout.addStretch(1)
        self.comp_auto_correct_button.clicked.connect(self.auto_correct_compensation)
        self.comp_run_auto_again_action.triggered.connect(self.run_auto_correct_again)
        self.comp_quick_recalculate_button.clicked.connect(self.recalculate_quick_adjust_once)
        self.comp_quick_review_button.clicked.connect(self.review_reconstructed_image)
        self.comp_restore_auto_button.clicked.connect(self.restore_auto_compensation_result)
        self.comp_expert_toggle_button.clicked.connect(self.open_expert_after_auto_correct)
        self.comp_result_ok_button.clicked.connect(self.review_reconstructed_image)
        self.comp_result_quick_button.clicked.connect(self.open_quick_adjust)
        self.comp_result_paint_button.clicked.connect(self.open_paint_after_auto_correct)
        self.comp_result_expert_button.clicked.connect(self.open_expert_after_auto_correct)
        self.comp_show_auto_mask_check.toggled.connect(self._toggle_auto_mask_visibility)
        self.comp_quick_removal.valueChanged.connect(self._update_compensation_session_status)
        self.comp_quick_detail.valueChanged.connect(self._update_compensation_session_status)
        self.comp_quick_protection.valueChanged.connect(self._update_compensation_session_status)

        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("ROI artifact"))
        self.comp_mode_combo = QComboBox()
        # Legacy mode sequence: "Manual Only", "Auto", "Spike", "Line", "Band", "Block", "Ring"
        self.comp_mode_combo.addItems(["Manual Only", "Auto", "Spike", "Line", "Band", "Diagonal", "Block", "Ring"])
        self.comp_mode_combo.setCurrentText("Manual Only")
        self.comp_mode_combo.setToolTip(
            "Manual Only uses exactly the painted mask and never runs Auto Detection. Auto and typed modes can generate a mask that remains fully editable with Eraser or Remove Component."
        )
        self.comp_mode_combo.currentTextChanged.connect(self.invalidate_compensation_detection)
        mode_row.addWidget(self.comp_mode_combo, 1)
        expert_layout.addLayout(mode_row)

        auto_mask_row = QHBoxLayout()
        auto_mask_row.addWidget(QLabel("Auto Mask"))
        self.comp_auto_mask_merge_combo = QComboBox()
        self.comp_auto_mask_merge_combo.addItems(["Replace Paint", "Add to Paint"])
        self.comp_auto_mask_merge_combo.setToolTip(
            "Replace Paint uses only the detected mask. Add to Paint merges Auto Mask with existing paint. Generated Auto Masks remain editable with Eraser or Remove Component."
        )
        auto_mask_row.addWidget(self.comp_auto_mask_merge_combo, 1)
        expert_layout.addLayout(auto_mask_row)

        level_row = QHBoxLayout()
        level_row.addWidget(QLabel("Strength"))
        self.comp_level_combo = QComboBox()
        self.comp_level_combo.addItems(["Low", "Mid", "High", "Extreme"])
        self.comp_level_combo.setCurrentText("High")
        level_row.addWidget(self.comp_level_combo)
        expert_layout.addLayout(level_row)

        target_row = QHBoxLayout()
        target_row.addWidget(QLabel("Target background"))
        self.comp_target_spin = QDoubleSpinBox()
        self.comp_target_spin.setRange(0.50, 2.00)
        self.comp_target_spin.setSingleStep(0.05)
        self.comp_target_spin.setDecimals(2)
        self.comp_target_spin.setValue(1.00)
        self.comp_target_spin.setSuffix(" x")
        self.comp_target_spin.setToolTip(
            "Replacement magnitude relative to the robust local background level."
        )
        target_row.addWidget(self.comp_target_spin)
        expert_layout.addLayout(target_row)

        threshold_row = QHBoxLayout()
        threshold_row.addWidget(QLabel("Detection threshold"))
        self.comp_threshold_spin = QDoubleSpinBox()
        self.comp_threshold_spin.setRange(1.50, 8.00)
        self.comp_threshold_spin.setSingleStep(0.25)
        self.comp_threshold_spin.setDecimals(2)
        self.comp_threshold_spin.setValue(3.00)
        self.comp_threshold_spin.setSuffix(" σ")
        threshold_row.addWidget(self.comp_threshold_spin)
        self.comp_threshold_spin.valueChanged.connect(self.invalidate_compensation_detection)
        self.comp_target_spin.valueChanged.connect(self.invalidate_compensation_detection)
        expert_layout.addLayout(threshold_row)

        # Optional expert controls. Collapsed by default so the automatic
        # workflow remains simple, but every important compensation detail can
        # be adjusted without leaving the ROI tool.
        self.comp_advanced_group = QGroupBox("Advanced Compensation Tuning")
        # A checkable QGroupBox disables every child while unchecked.  Commit0097
        # used that pattern as a visual collapse control, which made the tuning
        # widgets appear on screen but impossible to edit.  Keep this group
        # non-checkable so all controls remain genuinely interactive.
        self.comp_advanced_group.setCheckable(False)
        self.comp_advanced_group.setEnabled(True)
        advanced_layout = QFormLayout(self.comp_advanced_group)
        self._updating_comp_preset = False
        self.comp_tuning_preset_combo = QComboBox()
        self.comp_tuning_preset_combo.addItems(["Conservative", "Balanced", "Aggressive", "Expert"])
        self.comp_tuning_preset_combo.setCurrentText("Conservative")
        self.comp_tuning_preset_combo.setToolTip("Choose a safe preset, or Expert to edit every parameter manually.")
        advanced_layout.addRow("Tuning preset", self.comp_tuning_preset_combo)
        self.comp_sensitivity_combo = QComboBox()
        self.comp_sensitivity_combo.addItems(["Conservative", "Balanced", "Sensitive"])
        self.comp_sensitivity_combo.setCurrentText("Conservative")
        self.comp_sensitivity_combo.setToolTip("Conservative rejects smooth image-forming k-space signal most strongly.")
        advanced_layout.addRow("Auto Mask sensitivity", self.comp_sensitivity_combo)
        self.comp_mask_expansion_auto = QCheckBox("Auto")
        self.comp_mask_expansion_auto.setChecked(True)
        self.comp_mask_expansion_spin = QSpinBox()
        self.comp_mask_expansion_spin.setRange(0, 10)
        self.comp_mask_expansion_spin.setValue(1)
        self.comp_mask_expansion_spin.setSingleStep(1)
        self.comp_mask_expansion_spin.setKeyboardTracking(True)
        self.comp_mask_expansion_spin.setFocusPolicy(Qt.StrongFocus)
        self.comp_mask_expansion_spin.setButtonSymbols(QSpinBox.UpDownArrows)
        self.comp_mask_expansion_spin.setMinimumWidth(96)
        mask_expansion_row = QWidget(); mask_expansion_layout = QHBoxLayout(mask_expansion_row)
        mask_expansion_layout.setContentsMargins(0, 0, 0, 0)
        mask_expansion_layout.addWidget(self.comp_mask_expansion_auto)
        mask_expansion_layout.addWidget(self.comp_mask_expansion_spin, 1)
        advanced_layout.addRow("Mask expansion", mask_expansion_row)

        self.comp_donor_halo_auto = QCheckBox("Auto")
        self.comp_donor_halo_auto.setChecked(True)
        self.comp_donor_halo_spin = QSpinBox()
        self.comp_donor_halo_spin.setRange(1, 14)
        self.comp_donor_halo_spin.setValue(3)
        self.comp_donor_halo_spin.setSingleStep(1)
        self.comp_donor_halo_spin.setKeyboardTracking(True)
        self.comp_donor_halo_spin.setFocusPolicy(Qt.StrongFocus)
        self.comp_donor_halo_spin.setButtonSymbols(QSpinBox.UpDownArrows)
        self.comp_donor_halo_spin.setMinimumWidth(96)
        donor_halo_row = QWidget(); donor_halo_layout = QHBoxLayout(donor_halo_row)
        donor_halo_layout.setContentsMargins(0, 0, 0, 0)
        donor_halo_layout.addWidget(self.comp_donor_halo_auto)
        donor_halo_layout.addWidget(self.comp_donor_halo_spin, 1)
        advanced_layout.addRow("Donor halo", donor_halo_row)

        self.comp_pass_count_auto = QCheckBox("Auto")
        self.comp_pass_count_auto.setChecked(True)
        self.comp_pass_count_spin = QSpinBox()
        self.comp_pass_count_spin.setRange(1, 6)
        self.comp_pass_count_spin.setValue(2)
        self.comp_pass_count_spin.setSingleStep(1)
        self.comp_pass_count_spin.setKeyboardTracking(True)
        self.comp_pass_count_spin.setFocusPolicy(Qt.StrongFocus)
        self.comp_pass_count_spin.setButtonSymbols(QSpinBox.UpDownArrows)
        self.comp_pass_count_spin.setMinimumWidth(96)
        pass_count_row = QWidget(); pass_count_layout = QHBoxLayout(pass_count_row)
        pass_count_layout.setContentsMargins(0, 0, 0, 0)
        pass_count_layout.addWidget(self.comp_pass_count_auto)
        pass_count_layout.addWidget(self.comp_pass_count_spin, 1)
        advanced_layout.addRow("Compensation passes", pass_count_row)

        # Commit0113: keep native increment/decrement hit areas visible and clickable.
        # The global padding previously consumed most of the narrow spin-box button area.
        expert_spin_style = """
            QSpinBox { padding: 4px 28px 4px 6px; min-height: 24px; }
            QSpinBox::up-button { subcontrol-origin: border; subcontrol-position: top right; width: 24px; }
            QSpinBox::down-button { subcontrol-origin: border; subcontrol-position: bottom right; width: 24px; }
        """
        for spin in (self.comp_mask_expansion_spin, self.comp_donor_halo_spin, self.comp_pass_count_spin):
            spin.setStyleSheet(expert_spin_style)

        self.comp_strength_override_auto = QCheckBox("Auto")
        self.comp_strength_override_auto.setChecked(True)
        self.comp_strength_override_spin = QDoubleSpinBox()
        self.comp_strength_override_spin.setRange(0.05, 1.0)
        self.comp_strength_override_spin.setSingleStep(0.05)
        self.comp_strength_override_spin.setDecimals(2)
        self.comp_strength_override_spin.setValue(0.75)
        strength_row = QWidget(); strength_layout = QHBoxLayout(strength_row)
        strength_layout.setContentsMargins(0, 0, 0, 0)
        strength_layout.addWidget(self.comp_strength_override_auto)
        strength_layout.addWidget(self.comp_strength_override_spin, 1)
        advanced_layout.addRow("Blend strength", strength_row)
        self.comp_structure_preservation_spin = QDoubleSpinBox()
        self.comp_structure_preservation_spin.setRange(0.0, 1.0)
        self.comp_structure_preservation_spin.setSingleStep(0.05)
        self.comp_structure_preservation_spin.setDecimals(2)
        self.comp_structure_preservation_spin.setValue(0.70)
        self.comp_structure_preservation_spin.setToolTip("Higher values preserve smooth image-forming k-space signal. Lower values apply stronger correction.")
        advanced_layout.addRow("Normal signal preservation", self.comp_structure_preservation_spin)
        self.comp_frequency_aware_check = QCheckBox("Frequency-aware weighting")
        self.comp_frequency_aware_check.setChecked(True)
        self.comp_poisson_check = QCheckBox("Guided Poisson reconstruction")
        self.comp_poisson_check.setChecked(True)
        self.comp_hermitian_check = QCheckBox("Hermitian symmetry")
        self.comp_hermitian_check.setChecked(True)
        advanced_layout.addRow(self.comp_frequency_aware_check)
        advanced_layout.addRow(self.comp_poisson_check)
        advanced_layout.addRow(self.comp_hermitian_check)
        self.comp_advanced_reset_button = QPushButton("Reset Advanced Settings")
        advanced_layout.addRow(self.comp_advanced_reset_button)
        self.comp_advanced_reset_button.clicked.connect(self.reset_advanced_compensation_settings)
        self.comp_tuning_preset_combo.currentTextChanged.connect(self._apply_compensation_tuning_preset)
        for control in (
            self.comp_sensitivity_combo, self.comp_mask_expansion_auto, self.comp_mask_expansion_spin,
            self.comp_donor_halo_auto, self.comp_donor_halo_spin, self.comp_pass_count_auto, self.comp_pass_count_spin,
            self.comp_strength_override_auto, self.comp_strength_override_spin, self.comp_structure_preservation_spin,
            self.comp_frequency_aware_check, self.comp_poisson_check,
            self.comp_hermitian_check,
        ):
            if isinstance(control, QComboBox):
                control.currentTextChanged.connect(self._advanced_compensation_value_changed)
            elif isinstance(control, QCheckBox):
                control.toggled.connect(self._advanced_compensation_value_changed)
            else:
                control.valueChanged.connect(self._advanced_compensation_value_changed)
        for auto_check, value_control in (
            (self.comp_mask_expansion_auto, self.comp_mask_expansion_spin),
            (self.comp_donor_halo_auto, self.comp_donor_halo_spin),
            (self.comp_pass_count_auto, self.comp_pass_count_spin),
            (self.comp_strength_override_auto, self.comp_strength_override_spin),
        ):
            # Keep the arrow controls interactive even while Auto is selected.
            # A manual arrow/keyboard edit immediately switches that item to manual mode.
            auto_check.toggled.connect(self._advanced_compensation_value_changed)
            value_control.setEnabled(True)
            value_control.valueChanged.connect(lambda _value, a=auto_check: a.setChecked(False) if a.isChecked() else None)
        self._apply_compensation_tuning_preset("Conservative")
        self.comp_advanced_group.setVisible(True)
        self.comp_expert_toggle_button.clicked.connect(self.open_expert_after_auto_correct)
        expert_layout.addWidget(self.comp_advanced_group)
        expert_action_row = QHBoxLayout()
        self.comp_expert_auto_mask_button = QPushButton("Run Auto Mask Once")
        self.comp_candidate_viewer_button = QPushButton("Candidate Viewer")
        self.comp_expert_auto_mask_button.setToolTip("Run one Auto Mask detection using the current Expert settings. Existing masks can be replaced or merged.")
        self.comp_candidate_viewer_button.setToolTip("Inspect the candidates and validation scores from the latest Auto Correct or Expert Auto Mask run.")
        self.comp_expert_auto_mask_button.clicked.connect(self.run_expert_auto_mask_once)
        self.comp_candidate_viewer_button.clicked.connect(self.show_auto_candidate_viewer)
        expert_action_row.addWidget(self.comp_expert_auto_mask_button)
        expert_action_row.addWidget(self.comp_candidate_viewer_button)
        expert_layout.addLayout(expert_action_row)
        self.comp_expert_review_button = QPushButton("Review Reconstructed Image")
        self.comp_expert_review_button.setToolTip(
            "Reconstruct directly from the current Expert mask and settings, then open the review window."
        )
        self.comp_expert_review_button.clicked.connect(self.review_current_edited_mask)
        expert_layout.addWidget(self.comp_expert_review_button)
        expert_layout.addStretch(1)

        self._install_raw_compensation_tooltips()

        self.comp_status_label = QLabel("No RAW ROI selected")
        self.comp_status_label.setWordWrap(True)
        paint_layout.addWidget(self.comp_status_label)

        comp_preview_row = QHBoxLayout()
        comp_preview_row.addWidget(self.preview_comp_button)
        comp_preview_row.addWidget(self.open_comp_comparison_button)
        comp_preview_row.addWidget(self.apply_comp_button)
        paint_layout.addLayout(comp_preview_row)

        history_row = QHBoxLayout()
        self.comp_prev_button = QPushButton("Previous")
        self.comp_next_button = QPushButton("Next")
        self.comp_prev_button.clicked.connect(lambda: self.navigate_compensation_history(-1))
        self.comp_next_button.clicked.connect(lambda: self.navigate_compensation_history(1))
        history_row.addWidget(self.comp_prev_button)
        history_row.addWidget(self.comp_next_button)
        paint_layout.addLayout(history_row)
        paint_layout.addWidget(self.save_comp_button)
        self.clear_comp_history_button = QPushButton("Clear Compensation History")
        self.clear_comp_history_button.clicked.connect(self.clear_compensation_history)
        paint_layout.addWidget(self.clear_comp_history_button)
        paint_layout.addWidget(self.cancel_comp_button)
        paint_layout.addStretch(1)

        self.detect_comp_button.setEnabled(False)
        self.preview_comp_button.setEnabled(False)
        self.apply_comp_button.setEnabled(False)
        self.save_comp_button.setEnabled(False)
        self.comp_prev_button.setEnabled(False)
        self.comp_next_button.setEnabled(False)

        self.accordion_sections=[]
        self.image_tools_accordion=AccordionSection("Image Tools",image_tools_content,False)
        self.display_accordion=AccordionSection("Dynamic Range / Window Level",display_group,False)
        self.spike_accordion=AccordionSection("Spike",spike_content,False)
        self.comp_accordion=AccordionSection("Raw Data Compensation",comp_group,False)
        self.addsub_accordion=AccordionSection("Image Add/Sub",addsub_group,False)
        self.accordion_sections=[self.image_tools_accordion,self.display_accordion,self.spike_accordion,self.comp_accordion,self.addsub_accordion]
        for section in self.accordion_sections: section.opened.connect(self._accordion_opened); rl.addWidget(section)
        rl.addStretch(1)

        output_group = QGroupBox("Output Folder")
        output_layout = QVBoxLayout(output_group)
        self.output_root_label = QLabel("Output: follows opened image folder")
        self.output_root_label.setWordWrap(True)
        output_layout.addWidget(self.output_root_label)
        self.output_root_button = QPushButton("Change Output Folder")
        self.output_root_button.clicked.connect(self.change_output_root)
        output_layout.addWidget(self.output_root_button)
        self.output_accordion = AccordionSection("Output Folder", output_group, False)
        rl.addWidget(self.output_accordion)

        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        right_scroll.setFrameShape(QFrame.NoFrame)
        right_scroll.setMinimumWidth(250)
        right_scroll.setWidget(right)

        # Phase 1 layout: toolbar spans the complete viewer workspace,
        # directly above Original / FFT and the right-side tools.  The
        # Explorer remains a separate left column.
        self.viewer_workspace = QWidget()
        viewer_workspace_layout = QVBoxLayout(self.viewer_workspace)
        viewer_workspace_layout.setContentsMargins(0, 0, 0, 0)
        viewer_workspace_layout.setSpacing(4)
        viewer_workspace_layout.addWidget(self.mode_toolbar_widget)

        self.viewer_content_splitter = QSplitter(Qt.Horizontal)
        self.viewer_content_splitter.setChildrenCollapsible(False)
        # The entire right-side tool stack is a curtain while preserving the
        # existing nested tool accordions when opened. It starts closed.
        self.right_tools_accordion = AccordionSection(
            "Right Tool Menu", right_scroll, False, vertical_when_collapsed=True
        )
        self.right_tools_accordion.setMinimumWidth(250)
        self.right_tools_accordion.setMaximumWidth(520)
        self.right_tools_accordion.button.toggled.connect(
            self._right_tools_curtain_toggled
        )

        self.viewer_content_splitter.addWidget(center)
        self.viewer_content_splitter.addWidget(self.right_tools_accordion)
        self.viewer_content_splitter.setCollapsible(0, False)
        self.viewer_content_splitter.setCollapsible(1, False)
        self.viewer_content_splitter.setSizes([840, 300])
        self.viewer_content_splitter.splitterMoved.connect(
            lambda _pos, _idx: self._remember_splitter_sizes()
        )
        viewer_workspace_layout.addWidget(self.viewer_content_splitter, 1)

        # Apply the closed geometry immediately at startup.
        self._profile_curtain_toggled(False)
        self._right_tools_curtain_toggled(False)

        main_split.addWidget(self.viewer_workspace)
        main_split.setCollapsible(0, True)
        main_split.setCollapsible(1, False)
        main_split.setSizes([210, 1140])
        main_split.splitterMoved.connect(
            lambda _pos, _idx: self._remember_splitter_sizes()
        )
        layout.addWidget(main_split)
        self.set_view_mode("Both")
        return page

    def _set_vertical_splitter_sizes(self, sizes):
        """Apply programmatic splitter sizes without recording transient values."""
        self._programmatic_splitter_change = True
        try:
            self.vertical_splitter.setSizes([int(v) for v in sizes])
        finally:
            QTimer.singleShot(0, self._clear_programmatic_splitter_change)

    def _clear_programmatic_splitter_change(self):
        self._programmatic_splitter_change = False

    def _normalized_profile_sizes(self, total=None):
        total = max(int(total or self.vertical_splitter.height() or 1), 1)
        remembered = getattr(self, "_user_vertical_split_sizes", [])
        if len(remembered) == 2 and sum(remembered) > 0 and remembered[1] >= 180:
            ratio = remembered[1] / max(1, sum(remembered))
            lower = int(total * ratio)
        else:
            lower = 250
        max_lower = max(180, total - 320)
        lower = max(180, min(max_lower, lower))
        upper = max(320, total - lower)
        return [upper, lower]

    def _profile_curtain_toggled(self, expanded):
        """Resize the profile curtain and show crosshairs only while it is open."""
        if not hasattr(self, "vertical_splitter"):
            return
        self._profile_curtain_expanded = bool(expanded)
        # Commit0096: crosshairs are profile cursors, so keep both image panels
        # clean while the profile curtain is collapsed.
        for panel in (getattr(self, "primary_panel", None), getattr(self, "secondary_panel", None)):
            if panel is not None:
                panel._set_crosshair_visible(bool(expanded))
        if expanded:
            self.profile_accordion.setMinimumHeight(180)
            self.profile_accordion.setMaximumHeight(16777215)
            self._set_vertical_splitter_sizes(self._normalized_profile_sizes())
            QTimer.singleShot(0, self._enforce_profile_curtain_geometry)
            QTimer.singleShot(120, self._enforce_profile_curtain_geometry)
        else:
            current = self.vertical_splitter.sizes()
            if len(current) == 2 and current[1] >= 180:
                self._user_vertical_split_sizes = list(current)
            header_h = self.profile_accordion.button.sizeHint().height() + 8
            self.profile_accordion.setMinimumHeight(header_h)
            self.profile_accordion.setMaximumHeight(header_h)
            self._set_vertical_splitter_sizes([
                max(1, self.vertical_splitter.height() - header_h), header_h
            ])

    def _enforce_profile_curtain_geometry(self):
        if not hasattr(self, "profile_accordion") or not self.profile_accordion.button.isChecked():
            return
        current = self.vertical_splitter.sizes()
        if len(current) != 2 or current[1] < 180:
            self._set_vertical_splitter_sizes(self._normalized_profile_sizes())

    def _right_tools_curtain_toggled(self, expanded):
        """Keep the right curtain state and width stable across image refreshes."""
        if not hasattr(self, "viewer_content_splitter"):
            return
        self._right_tools_curtain_expanded = bool(expanded)
        if expanded:
            self.right_tools_accordion.setMinimumWidth(250)
            self.right_tools_accordion.setMaximumWidth(520)
            remembered = getattr(self, "_user_viewer_content_split_sizes", [])
            if len(remembered) == 2 and sum(remembered) > 0 and remembered[1] >= 200:
                self.viewer_content_splitter.setSizes(remembered)
            else:
                total = max(1, self.viewer_content_splitter.width())
                self.viewer_content_splitter.setSizes([max(500, total - 300), 300])
        else:
            current = self.viewer_content_splitter.sizes()
            if len(current) == 2 and current[1] >= 200:
                self._user_viewer_content_split_sizes = current
            header_w = 32
            self.right_tools_accordion.setMinimumWidth(header_w)
            self.right_tools_accordion.setMaximumWidth(header_w)
            self.viewer_content_splitter.setSizes([
                max(1, self.viewer_content_splitter.width() - header_w), header_w
            ])

    def _image_panel_utility_requested(self, command):
        if command == "DICOM_HEADER":
            self.show_dicom_header_popup()
        elif command == "ORIENTATION":
            self._open_orientation_dialog()

    def _build_signal_studio(self):
        page = QWidget(); layout = QVBoxLayout(page)
        top = QHBoxLayout()
        self.back_to_workspace = QPushButton("← Back to Image Workspace")
        self.back_to_workspace.clicked.connect(lambda: self.tabs.setCurrentIndex(0))
        top.addWidget(self.back_to_workspace)
        self.signal_combo = QComboBox(); self.signal_combo.currentIndexChanged.connect(self.update_signal_plot)
        self.signal_component = QComboBox(); self.signal_component.addItems(["Magnitude", "Real", "Imaginary", "Phase"])
        self.signal_component.currentTextChanged.connect(self.update_signal_plot)
        self.signal_fft = QCheckBox("Frequency display (FFT)"); self.signal_fft.setChecked(False); self.signal_fft.toggled.connect(self.update_signal_plot)
        remove_signal = QPushButton("Remove Selected"); remove_signal.clicked.connect(self.remove_selected_signal)
        clear = QPushButton("Clear All Signals"); clear.clicked.connect(self.clear_signals)
        top.addWidget(QLabel("Signal")); top.addWidget(self.signal_combo, 1); top.addWidget(self.signal_component); top.addWidget(self.signal_fft); top.addWidget(remove_signal); top.addWidget(clear)
        layout.addLayout(top)
        self.signal_plot = pg.PlotWidget(); self.signal_plot.showGrid(x=True, y=True, alpha=.2)
        layout.addWidget(self.signal_plot, 1)
        note = QLabel("Tracker PFile/TrackerImg drops are automatically registered here as complex 1D raw signals.")
        note.setWordWrap(True); layout.addWidget(note)
        return page

    def _build_metadata(self):
        page = QWidget(); layout = QVBoxLayout(page)
        row = QHBoxLayout()
        self.header_filter = QComboBox(); self.header_filter.setEditable(True); self.header_filter.setInsertPolicy(QComboBox.NoInsert)
        self.header_filter.lineEdit().setPlaceholderText("Filter header tags...")
        self.header_filter.lineEdit().textChanged.connect(self.filter_header)
        export = QPushButton("Export Current Header to Excel"); export.clicked.connect(self.export_header_excel)
        row.addWidget(self.header_filter, 1); row.addWidget(export)
        layout.addLayout(row)
        self.header_table = QTableWidget(0, 6)
        self.header_table.setHorizontalHeaderLabels(["Tag", "Keyword", "Name", "VR", "VM", "Value"])
        self.header_table.horizontalHeader().setStretchLastSection(True)
        self.header_table.setSortingEnabled(True)
        layout.addWidget(self.header_table)
        return page

    def _build_tracker_explorer(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        controls = QHBoxLayout()
        self.tracker_prev_file_button = QPushButton("◀ Previous File")
        self.tracker_next_file_button = QPushButton("Next File ▶")
        self.tracker_prev_file_button.clicked.connect(lambda: self.navigate_tracker_file(-1))
        self.tracker_next_file_button.clicked.connect(lambda: self.navigate_tracker_file(1))
        controls.addWidget(self.tracker_prev_file_button)
        self.tracker_file_combo = QComboBox()
        self.tracker_file_combo.currentIndexChanged.connect(self.select_tracker_file_index)
        controls.addWidget(self.tracker_file_combo)
        controls.addWidget(self.tracker_next_file_button)
        self.tracker_clear_current_button = QPushButton("Clear Current Tracker")
        self.tracker_clear_all_button = QPushButton("Clear All Trackers")
        self.tracker_clear_current_button.clicked.connect(self.clear_current_tracker)
        self.tracker_clear_all_button.clicked.connect(self.clear_all_trackers)
        controls.addWidget(self.tracker_clear_current_button)
        controls.addWidget(self.tracker_clear_all_button)
        self.tracker_drop_text = QLabel(
            "Drop a Tracker PFile / TrackerImg anywhere in the window. "
            "This tab reads and ranks raw lines directly without FFT."
        )
        self.tracker_drop_text.setWordWrap(True)
        controls.addWidget(self.tracker_drop_text, 1)

        controls.addWidget(QLabel("Top lines"))
        self.tracker_top_count = QSpinBox()
        self.tracker_top_count.setRange(1, 100)
        self.tracker_top_count.setValue(1)
        self.tracker_top_count.valueChanged.connect(self._refresh_tracker_ranking)
        controls.addWidget(self.tracker_top_count)

        controls.addWidget(QLabel("Rank by"))
        self.tracker_metric_combo = QComboBox()
        self.tracker_metric_combo.addItems(["Peak", "RMS", "Energy", "Mean magnitude", "Variance"])
        self.tracker_metric_combo.currentTextChanged.connect(self._refresh_tracker_ranking)
        controls.addWidget(self.tracker_metric_combo)

        self.tracker_export_button = QPushButton("Export CSV")
        self.tracker_export_button.clicked.connect(self.export_tracker_lines_csv)
        controls.addWidget(self.tracker_export_button)
        self.tracker_to_workspace_button = QPushButton("Show in Image Workspace")
        self.tracker_to_workspace_button.clicked.connect(self.show_tracker_in_workspace)
        controls.addWidget(self.tracker_to_workspace_button)
        self.tracker_to_1d_button = QPushButton("Show Strongest Line in 1D Studio")
        self.tracker_to_1d_button.clicked.connect(self.send_tracker_to_signal_studio)
        controls.addWidget(self.tracker_to_1d_button)
        layout.addLayout(controls)

        self.tracker_summary = QLabel("No Tracker file loaded.")
        self.tracker_summary.setWordWrap(True)
        layout.addWidget(self.tracker_summary)

        splitter = QSplitter(Qt.Horizontal)

        ranking_box = QGroupBox("Strongest Tracker Line")
        ranking_layout = QVBoxLayout(ranking_box)
        self.tracker_table = QTableWidget(0, 8)
        self.tracker_table.setHorizontalHeaderLabels(
            ["Rank", "Line", "Peak", "RMS", "Energy", "Mean", "Variance", "Relative strength"]
        )
        self.tracker_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.tracker_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.tracker_table.itemSelectionChanged.connect(self._tracker_selection_changed)
        self.tracker_table.horizontalHeader().setStretchLastSection(True)
        ranking_layout.addWidget(self.tracker_table)

        viewer_box = QGroupBox("Strongest Tracker Raw Line Magnitude — FFT Disabled")
        viewer_layout = QVBoxLayout(viewer_box)
        self.tracker_raw_plot = pg.PlotWidget()
        self.tracker_raw_plot.showGrid(x=True, y=True, alpha=0.25)
        self.tracker_raw_plot.setLabel("bottom", "Sample")
        self.tracker_raw_plot.setLabel("left", "Signal")
        viewer_layout.addWidget(self.tracker_raw_plot, 1)

        viewer_controls = QHBoxLayout()
        viewer_controls.addWidget(QLabel("Component"))
        self.tracker_component_combo = QComboBox()
        self.tracker_component_combo.addItems(["Magnitude", "Real", "Imaginary", "Phase"])
        self.tracker_component_combo.currentTextChanged.connect(self._plot_tracker_selected_lines)
        viewer_controls.addWidget(self.tracker_component_combo)

        self.tracker_previous_button = QPushButton("Previous Strong Line")
        self.tracker_previous_button.clicked.connect(self._select_previous_tracker_line)
        viewer_controls.addWidget(self.tracker_previous_button)

        self.tracker_next_button = QPushButton("Next Strong Line")
        self.tracker_next_button.clicked.connect(self._select_next_tracker_line)
        viewer_controls.addWidget(self.tracker_next_button)
        viewer_controls.addStretch(1)
        viewer_layout.addLayout(viewer_controls)

        self.tracker_metrics_label = QLabel("-")
        self.tracker_metrics_label.setWordWrap(True)
        viewer_layout.addWidget(self.tracker_metrics_label)

        splitter.addWidget(ranking_box)
        splitter.addWidget(viewer_box)
        splitter.setSizes([520, 900])
        layout.addWidget(splitter, 1)

        explanation = QLabel(
            "Tracker Signal Explorer and Spike Detection are separate. "
            "This tab automatically picks high-signal raw lines and displays their original samples without FFT."
        )
        explanation.setWordWrap(True)
        layout.addWidget(explanation)
        return page

    def _build_tracker_position(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        top = QHBoxLayout()
        top.addWidget(QLabel("FFT length"))
        self.position_fft_length = QSpinBox()
        self.position_fft_length.setRange(64, 65536)
        self.position_fft_length.setSingleStep(64)
        self.position_fft_length.setValue(1024)
        top.addWidget(self.position_fft_length)

        top.addWidget(QLabel("FOV (mm)"))
        self.position_fov = QDoubleSpinBox()
        self.position_fov.setRange(1.0, 5000.0)
        self.position_fov.setDecimals(3)
        self.position_fov.setValue(500.0)
        top.addWidget(self.position_fov)

        self.position_use_top = QPushButton("Use Top 3 Tracker Lines")
        self.position_use_top.clicked.connect(self.populate_position_top_lines)
        top.addWidget(self.position_use_top)

        self.position_calculate = QPushButton("Calculate Position")
        self.position_calculate.clicked.connect(self.calculate_tracker_position)
        top.addWidget(self.position_calculate)
        top.addStretch(1)
        layout.addLayout(top)

        offset_group = QGroupBox("Center Offset and Coordinate Convention")
        offset_form = QFormLayout(offset_group)
        offset_row = QHBoxLayout()
        self.position_offset_x = QDoubleSpinBox()
        self.position_offset_y = QDoubleSpinBox()
        self.position_offset_z = QDoubleSpinBox()
        for spin in (self.position_offset_x, self.position_offset_y, self.position_offset_z):
            spin.setRange(-5000.0, 5000.0)
            spin.setDecimals(4)
            spin.setValue(0.0)
        offset_row.addWidget(QLabel("X"))
        offset_row.addWidget(self.position_offset_x)
        offset_row.addWidget(QLabel("Y"))
        offset_row.addWidget(self.position_offset_y)
        offset_row.addWidget(QLabel("Z"))
        offset_row.addWidget(self.position_offset_z)
        offset_form.addRow("CenterOffset (mm)", offset_row)

        self.position_oppose_ap = QCheckBox("Oppose AP coordinate")
        self.position_oppose_ap.setChecked(True)
        offset_form.addRow(self.position_oppose_ap)
        layout.addWidget(offset_group)

        self.position_table = QTableWidget(3, 7)
        self.position_table.setHorizontalHeaderLabels([
            "Line", "Plane", "Rotation (deg)", "Peak bin",
            "Coordinate (mm)", "Peak/Background", "Use"
        ])
        plane_names = ["Axial", "Coronal", "Sagittal"]
        for row in range(3):
            self.position_table.setItem(row, 0, QTableWidgetItem(str(row)))
            plane_combo = QComboBox()
            plane_combo.addItems(plane_names)
            plane_combo.setCurrentIndex(row)
            self.position_table.setCellWidget(row, 1, plane_combo)

            rotation = QDoubleSpinBox()
            rotation.setRange(-360.0, 360.0)
            rotation.setDecimals(4)
            rotation.setValue(0.0)
            self.position_table.setCellWidget(row, 2, rotation)

            self.position_table.setItem(row, 3, QTableWidgetItem("-"))
            self.position_table.setItem(row, 4, QTableWidgetItem("-"))
            self.position_table.setItem(row, 5, QTableWidgetItem("-"))

            use_box = QCheckBox()
            use_box.setChecked(True)
            self.position_table.setCellWidget(row, 6, use_box)

        self.position_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.position_table)

        result_group = QGroupBox("Converted Tracker Position")
        result_layout = QVBoxLayout(result_group)
        self.position_result = QLabel(
            "Load a Tracker file, choose three directional lines, set Plane / Rotation / CenterOffset, "
            "then calculate the position."
        )
        self.position_result.setWordWrap(True)
        self.position_result.setTextInteractionFlags(Qt.TextSelectableByMouse)
        result_layout.addWidget(self.position_result)
        layout.addWidget(result_group)

        note = QLabel(
            "The basic conversion matches the attached TrackerApp structure: FFT peak bin → FOV coordinate → "
            "multi-direction geometric solve → CenterOffset → optional AP sign inversion. "
            "Gradient-map non-linearity correction is not applied unless a vendor gradient adapter is added."
        )
        note.setWordWrap(True)
        layout.addWidget(note)
        layout.addStretch(1)
        return page

    def _build_spike_results(self):
        page = QWidget()
        root = QHBoxLayout(page)

        left = QWidget()
        left.setMinimumWidth(220)
        left_layout = QVBoxLayout(left)
        left_layout.addWidget(QLabel("Processed Images"))
        self.spike_result_list = QTreeWidget()
        self.spike_result_list.setHeaderHidden(True)
        self.spike_result_list.itemClicked.connect(self._spike_result_selected)
        left_layout.addWidget(self.spike_result_list, 1)
        self.spike_result_summary = QLabel("Processed: 0 | Corrected: 0")
        self.spike_result_summary.setWordWrap(True)
        left_layout.addWidget(self.spike_result_summary)
        self.save_spike_result_button = QPushButton("Save Corrected Image")
        self.save_spike_result_button.clicked.connect(self.save_current_spike_result)
        left_layout.addWidget(self.save_spike_result_button)

        self.spike_review_tabs = QTabWidget()

        # STEP 1: what data actually entered the algorithm.
        step1 = QWidget(); l1 = QVBoxLayout(step1)
        self.spike_input_note = QLabel()
        self.spike_input_note.setWordWrap(True)
        self.spike_input_note.setStyleSheet("QLabel { padding: 6px; background: #263238; color: white; }")
        l1.addWidget(self.spike_input_note)
        g1 = QWidget(); gl1 = QGridLayout(g1)
        self.spike_original_panel = ImagePanel("1A Original DICOM Image")
        self.spike_fft_input_panel = ImagePanel("1B FFT Input (spatial-domain pixels)")
        gl1.addWidget(self.spike_original_panel,0,0); gl1.addWidget(self.spike_fft_input_panel,0,1)
        l1.addWidget(g1,1)
        self.spike_review_tabs.addTab(step1,"STEP 1 Input")

        # STEP 2: transformation into k-space.
        step2 = QWidget(); gl2 = QGridLayout(step2)
        self.spike_kspace_linear_panel = ImagePanel("2A k-space Magnitude")
        self.spike_raw_before_panel = ImagePanel("2B Log k-space")
        self.spike_kspace_enhanced_panel = ImagePanel("2C Robust Enhanced k-space")
        gl2.addWidget(self.spike_kspace_linear_panel,0,0)
        gl2.addWidget(self.spike_raw_before_panel,0,1)
        gl2.addWidget(self.spike_kspace_enhanced_panel,0,2)
        self.spike_review_tabs.addTab(step2,"STEP 2 FFT")

        # STEP 3: all proposals, before final decision.
        step3 = QWidget(); gl3 = QGridLayout(step3)
        self.spike_candidate_overlay_panel = ImagePanel("3A All Candidate Proposals")
        self.spike_candidate_mask_panel = ImagePanel("3B Selected Candidate Mask")
        gl3.addWidget(self.spike_candidate_overlay_panel,0,0)
        gl3.addWidget(self.spike_candidate_mask_panel,0,1)
        self.spike_review_tabs.addTab(step3,"STEP 3 Candidates")

        # STEP 4: candidate-only inverse FFT.
        step4 = QWidget(); gl4 = QGridLayout(step4)
        self.spike_candidate_kspace_panel = ImagePanel("4A Selected Candidate in k-space")
        self.spike_candidate_ifft_panel = ImagePanel("4B Candidate-only Inverse FFT")
        self.spike_predicted_wave_panel = ImagePanel("4C Predicted Spatial Wave")
        gl4.addWidget(self.spike_candidate_kspace_panel,0,0)
        gl4.addWidget(self.spike_candidate_ifft_panel,0,1)
        gl4.addWidget(self.spike_predicted_wave_panel,0,2)
        self.spike_review_tabs.addTab(step4,"STEP 4 IFFT")

        # STEP 5: compare prediction with image residual.
        step5 = QWidget(); gl5 = QGridLayout(step5)
        self.spike_residual_panel = ImagePanel("5A Original High-frequency Residual")
        self.spike_match_panel = ImagePanel("5B Prediction × Residual Match")
        gl5.addWidget(self.spike_residual_panel,0,0)
        gl5.addWidget(self.spike_match_panel,0,1)
        self.spike_review_tabs.addTab(step5,"STEP 5 Correlation")

        # STEP 6: explicit decision and reason.
        step6 = QWidget(); l6 = QVBoxLayout(step6)
        self.spike_decision_text = QTextEdit(); self.spike_decision_text.setReadOnly(True)
        l6.addWidget(self.spike_decision_text)
        self.spike_review_tabs.addTab(step6,"STEP 6 Decision")

        # STEP 7: correction result.
        step7 = QWidget(); gl7 = QGridLayout(step7)
        self.spike_corrected_panel = ImagePanel("7A Corrected Image")
        self.spike_difference_panel = ImagePanel("7B Absolute Difference")
        self.spike_raw_after_panel = ImagePanel("7C Processed Log k-space")
        gl7.addWidget(self.spike_corrected_panel,0,0)
        gl7.addWidget(self.spike_difference_panel,0,1)
        gl7.addWidget(self.spike_raw_after_panel,0,2)
        self.spike_review_tabs.addTab(step7,"STEP 7 Compensation")

        right = QWidget(); right.setMinimumWidth(430)
        right_layout = QVBoxLayout(right)
        right_layout.addWidget(QLabel("Candidate Review — select one row to inspect every step"))
        self.spike_candidate_table = QTableWidget(0, 8)
        self.spike_candidate_table.setHorizontalHeaderLabels([
            "ID","Type","Peak Z","Energy","Angle","Period","Corr","Decision"
        ])
        self.spike_candidate_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.spike_candidate_table.setSelectionMode(QTableWidget.SingleSelection)
        self.spike_candidate_table.cellClicked.connect(self._spike_candidate_selected)
        self.spike_candidate_table.horizontalHeader().setStretchLastSection(True)
        right_layout.addWidget(self.spike_candidate_table,1)
        self.spike_candidate_detail = QTextEdit(); self.spike_candidate_detail.setReadOnly(True)
        self.spike_candidate_detail.setMaximumHeight(230)
        right_layout.addWidget(self.spike_candidate_detail)

        split = QSplitter(Qt.Horizontal)
        split.addWidget(left); split.addWidget(self.spike_review_tabs); split.addWidget(right)
        split.setCollapsible(0, True); split.setCollapsible(2, True)
        split.setSizes([230, 1050, 430])
        root.addWidget(split)
        return page


    def _build_artifact_detection(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        controls = QHBoxLayout()
        self.detect_artifact_button = QPushButton("Detect Artifact on Selected Images")
        self.detect_artifact_button.clicked.connect(self.detect_artifacts_from_db)
        controls.addWidget(self.detect_artifact_button)
        self.detect_tracker_button = QPushButton("Detect Current Tracker File")
        self.detect_tracker_button.clicked.connect(self.detect_tracker_artifact_from_db)
        controls.addWidget(self.detect_tracker_button)

        self.detect_selected_only = QCheckBox("Selected images only")
        self.detect_selected_only.setChecked(True)
        controls.addWidget(self.detect_selected_only)

        controls.addWidget(QLabel("Minimum samples / class"))
        self.detect_min_samples = QSpinBox()
        self.detect_min_samples.setRange(1, 100)
        self.detect_min_samples.setValue(2)
        controls.addWidget(self.detect_min_samples)
        controls.addStretch(1)
        layout.addLayout(controls)

        self.artifact_detection_summary = QLabel(
            "Open an Artifact DB and run detection. Classes without enough labeled samples are not predicted."
        )
        self.artifact_detection_summary.setWordWrap(True)
        layout.addWidget(self.artifact_detection_summary)

        self.artifact_detection_table = QTableWidget(0, 9)
        self.artifact_detection_table.setHorizontalHeaderLabels([
            "File", "Series", "Predicted Artifact", "Confidence",
            "Distance", "Training Support", "Status", "Alternatives", "Source Path"
        ])
        self.artifact_detection_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.artifact_detection_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.artifact_detection_table.horizontalHeader().setStretchLastSection(True)
        self.artifact_detection_table.cellDoubleClicked.connect(
            self.open_detected_artifact_image
        )
        layout.addWidget(self.artifact_detection_table, 1)

        classify_row = QHBoxLayout()
        self.send_detection_to_learning_button = QPushButton(
            "Send Selected Detection Rows to Artifact Learning"
        )
        self.send_detection_to_learning_button.clicked.connect(
            self.send_detected_rows_to_learning
        )
        classify_row.addWidget(self.send_detection_to_learning_button)
        classify_row.addStretch(1)
        layout.addLayout(classify_row)

        note = QLabel(
            "Spike is no longer a hard-coded threshold classification. "
            "Spike, Frequency, and other Artifact classes are predicted from the labeled Artifact DB. "
            "The separate Spike Diagnostic tab remains available for raw waveform review."
        )
        note.setWordWrap(True)
        layout.addWidget(note)
        return page

    def _build_artifact_learning(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        db_row = QHBoxLayout()
        self.artifact_db_label = QLabel("Artifact DB: Not opened")
        db_row.addWidget(self.artifact_db_label, 1)
        for label, callback in [
            ("Open / Create DB", self.open_artifact_database),
            ("Import DB JSON", self.import_artifact_database),
            ("Export DB JSON", self.export_artifact_database),
        ]:
            button = QPushButton(label)
            button.clicked.connect(callback)
            db_row.addWidget(button)
        layout.addLayout(db_row)

        selection = QGroupBox("Training Image Selection")
        selection_layout = QVBoxLayout(selection)
        self.artifact_selection_label = QLabel(
            "Select one or more DICOM images in Explorer, then add them as training images."
        )
        self.artifact_selection_label.setWordWrap(True)
        selection_layout.addWidget(self.artifact_selection_label)
        selection_buttons = QHBoxLayout()
        self.artifact_add_selected_button = QPushButton("Add Selected Images")
        self.artifact_add_selected_button.clicked.connect(self.collect_selected_artifact_images)
        selection_buttons.addWidget(self.artifact_add_selected_button)
        self.artifact_mark_spike_button = QPushButton("Mark Selected: Spike")
        self.artifact_mark_spike_button.clicked.connect(lambda: self._set_learning_class_quick("Spike"))
        selection_buttons.addWidget(self.artifact_mark_spike_button)
        self.artifact_mark_not_spike_button = QPushButton("Mark Selected: Not Spike")
        self.artifact_mark_not_spike_button.clicked.connect(lambda: self._set_learning_class_quick("Not Spike"))
        selection_buttons.addWidget(self.artifact_mark_not_spike_button)
        self.artifact_set_normal_button = QPushButton("Set Current as Normal Reference")
        self.artifact_set_normal_button.clicked.connect(self.set_current_normal_reference)
        selection_buttons.addWidget(self.artifact_set_normal_button)
        self.artifact_load_normal_button = QPushButton("Load Normal Reference")
        self.artifact_load_normal_button.clicked.connect(self.load_normal_reference)
        selection_buttons.addWidget(self.artifact_load_normal_button)
        self.artifact_preview_tracker_button = QPushButton("Preview Tracker Data")
        self.artifact_preview_tracker_button.clicked.connect(self.preview_tracker_in_artifact_learning)
        selection_buttons.addWidget(self.artifact_preview_tracker_button)
        self.artifact_save_tracker_button = QPushButton("Save Tracker Training Sample")
        self.artifact_save_tracker_button.clicked.connect(self.save_tracker_training_sample)
        selection_buttons.addWidget(self.artifact_save_tracker_button)
        selection_buttons.addStretch(1)
        selection_layout.addLayout(selection_buttons)
        preview_row = QHBoxLayout()
        self.artifact_preview_button = QPushButton("Preview Selected Image")
        self.artifact_preview_button.clicked.connect(self.preview_artifact_learning_image)
        preview_row.addWidget(self.artifact_preview_button)
        self.artifact_preview_prev = QPushButton("Previous Selected")
        self.artifact_preview_next = QPushButton("Next Selected")
        self.artifact_preview_prev.clicked.connect(lambda: self.navigate_artifact_preview(-1))
        self.artifact_preview_next.clicked.connect(lambda: self.navigate_artifact_preview(1))
        preview_row.addWidget(self.artifact_preview_prev)
        preview_row.addWidget(self.artifact_preview_next)
        self.artifact_clear_selected_button = QPushButton("Clear Training Selection")
        self.artifact_clear_all_button = QPushButton("Clear Training + Normal")
        self.artifact_clear_selected_button.clicked.connect(self.clear_selected_artifact_training)
        self.artifact_clear_all_button.clicked.connect(self.clear_all_artifact_training)
        preview_row.addWidget(self.artifact_clear_selected_button)
        preview_row.addWidget(self.artifact_clear_all_button)
        preview_row.addStretch(1)
        selection_layout.addLayout(preview_row)
        self.artifact_preview_panel = ImagePanel("Artifact Learning Preview")
        self.artifact_preview_panel.setMinimumHeight(260)
        selection_layout.addWidget(self.artifact_preview_panel)
        self.artifact_preview_position = 0
        self.artifact_normal_label = QLabel("Normal reference: None")
        selection_layout.addWidget(self.artifact_normal_label)
        layout.addWidget(selection)

        classify = QGroupBox("Classification")
        classify_form = QFormLayout(classify)
        type_row = QHBoxLayout()
        self.artifact_type_combo = QComboBox()
        type_row.addWidget(self.artifact_type_combo, 1)
        self.artifact_new_type = QLineEdit()
        self.artifact_new_type.setPlaceholderText("Manual new Artifact type")
        type_row.addWidget(self.artifact_new_type, 1)
        add_type = QPushButton("Add Type")
        add_type.clicked.connect(self.add_artifact_type_manual)
        type_row.addWidget(add_type)
        classify_form.addRow("Artifact", type_row)

        resolution_row = QHBoxLayout()
        self.artifact_resolution_combo = QComboBox()
        resolution_row.addWidget(self.artifact_resolution_combo, 1)
        self.artifact_new_resolution = QLineEdit()
        self.artifact_new_resolution.setPlaceholderText("Manual How to Resolved")
        resolution_row.addWidget(self.artifact_new_resolution, 1)
        add_resolution = QPushButton("Add Resolution")
        add_resolution.clicked.connect(self.add_resolution_manual)
        resolution_row.addWidget(add_resolution)
        classify_form.addRow("How to Resolved", resolution_row)

        self.artifact_notes = QTextEdit()
        self.artifact_notes.setMaximumHeight(90)
        self.artifact_notes.setPlaceholderText("Notes, conditions, service action, reproducibility...")
        classify_form.addRow("Notes", self.artifact_notes)

        save_button = QPushButton("Save Training Samples to Artifact DB")
        save_button.clicked.connect(self.save_artifact_training_samples)
        classify_form.addRow(save_button)
        layout.addWidget(classify)

        self.artifact_training_table = QTableWidget(0, 8)
        self.artifact_training_table.setHorizontalHeaderLabels([
            "ID", "File", "Series", "Instance", "Artifact",
            "How to Resolved", "Normal Compared", "Source Path"
        ])
        self.artifact_training_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.artifact_training_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.artifact_training_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.artifact_training_table, 1)

        reclassify = QPushButton("Apply Current Classification to Selected DB Rows")
        reclassify.clicked.connect(self.reclassify_selected_db_rows)
        layout.addWidget(reclassify)

        self.artifact_summary = QLabel(
            "Current detector: Spike. Artifact DB supports Frequency and future Artifact detectors."
        )
        self.artifact_summary.setWordWrap(True)
        layout.addWidget(self.artifact_summary)
        return page

    def _build_menu(self):
        file_menu = self.menuBar().addMenu("File")
        import_files = QAction("Import Files...", self)
        import_files.triggered.connect(self.open_import_selection)
        file_menu.addAction(import_files)
        import_folder = QAction("Import Folder...", self)
        import_folder.triggered.connect(self._open_folder_direct)
        file_menu.addAction(import_folder)
        file_menu.addSeparator()
        export_npz = QAction("Export Image/k-space NPZ", self); export_npz.triggered.connect(self.export_npz); file_menu.addAction(export_npz)
        export_header = QAction("Export DICOM Header Excel", self); export_header.triggered.connect(self.export_header_excel); file_menu.addAction(export_header)
        file_menu.addSeparator(); quit_action = QAction("Exit", self); quit_action.triggered.connect(self.close); file_menu.addAction(quit_action)

        diagnostic_menu = self.menuBar().addMenu("Diagnostics")
        open_logs = QAction("Open Log Folder", self)
        open_logs.triggered.connect(self.open_stable_log_folder)
        diagnostic_menu.addAction(open_logs)

        export_logs = QAction("Export Stable Diagnostic ZIP...", self)
        export_logs.triggered.connect(self.export_stable_diagnostic_zip)
        diagnostic_menu.addAction(export_logs)

        if not RELEASE_MODE:
            help_menu = self.menuBar().addMenu("Help")
            normal_guide_action = QAction("Normal Usage Guide", self)
            normal_guide_action.triggered.connect(lambda: self.show_normal_guide(offer_raw=True))
            help_menu.addAction(normal_guide_action)
            raw_guide_action = QAction("Raw Data Compensation Guide", self)
            raw_guide_action.triggered.connect(lambda: self.show_raw_compensation_guide(start_tour=True))
            help_menu.addAction(raw_guide_action)
            library_action = QAction("Guide Library", self)
            library_action.triggered.connect(self.show_guide_library)
            help_menu.addAction(library_action)



    def _stable_diagnostic_state(self):
        image_item = None
        try:
            image_item = getattr(
                self.primary_panel.image_item,
                "image",
                None,
            )
        except Exception:
            pass

        return {
            "version": APP_VERSION,
            "current_source": self.current_source,
            "source_kind": self.source_kind,
            "view_mode": self.view_mode,
            "slice_index": self.slice_index,
            "current_image": self.stable_diagnostics.array_summary(
                self.current_image
            ),
            "current_kspace": self.stable_diagnostics.array_summary(
                self.current_kspace
            ),
            "image_item": self.stable_diagnostics.array_summary(
                image_item
            ),
            "primary_visible": self.primary_panel.isVisible(),
            "secondary_visible": self.secondary_panel.isVisible(),
            "window_level": self.original_window_level,
            "dynamic_range": self.original_dynamic_range,
        }

    def open_stable_log_folder(self):
        self.stable_diagnostics.info("OPEN_LOG_FOLDER")
        try:
            os.startfile(str(self.stable_diagnostics.log_dir))
        except Exception:
            self.statusBar().showMessage(
                f"Log folder: {self.stable_diagnostics.log_dir}",
                10000,
            )

    def export_stable_diagnostic_zip(self):
        default_path = (
            self.stable_diagnostics.export_dir
            / "MR_Image_Explorer_Stable_Diagnostic.zip"
        )
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Stable Diagnostic ZIP",
            str(default_path),
            "ZIP (*.zip)",
        )
        if not filename:
            return

        output = Path(filename)
        if output.suffix.lower() != ".zip":
            output = output.with_suffix(".zip")

        try:
            self.stable_diagnostics.export_zip(
                output,
                self._stable_diagnostic_state(),
            )
            self.statusBar().showMessage(
                f"Diagnostic ZIP exported: {output}",
                10000,
            )
        except Exception as exc:
            self.stable_diagnostics.exception(
                "DIAGNOSTIC_EXPORT_FAILED",
                exc,
            )
            QMessageBox.warning(
                self,
                "Diagnostic Export",
                f"{type(exc).__name__}: {exc}",
            )

    def _apply_style(self):
        self.setStyleSheet("""
        QMainWindow, QWidget { background: #171a1f; color: #e6e9ee; }
        QMenuBar, QMenu { background: #20242b; }
        QTabWidget::pane { border: 1px solid #303946; }
        QTabBar::tab { background: #20242b; padding: 9px 18px; border: 1px solid #303946; }
        QTabBar::tab:selected { background: #1464cc; color: white; font-weight: 600; }
        QTabBar::tab:hover { background: #2b3440; }
        QPushButton, QComboBox, QSpinBox { background: #252b33; border: 1px solid #3b4654; padding: 6px; border-radius: 3px; }
        QPushButton:checked { background: #1464cc; border-color: #2c8cff; }
        QPushButton:hover { border-color: #6594c5; }
        QTreeWidget, QTableWidget { background: #111419; alternate-background-color: #191e25; border: 1px solid #303946; }
        QGroupBox { border: 1px solid #303946; border-radius: 4px; margin-top: 8px; padding-top: 8px; }
        QGroupBox::title { subcontrol-origin: margin; left: 8px; }
        #DropBanner { border: 1px dashed #5c7189; border-radius: 5px; background: #1d232b; }
        #DropBanner[dragActive="true"] { border: 2px solid #2f8cff; background: #20354f; }
        #InfoCard { border: 1px solid #303946; padding: 8px; background: #111419; }
        QSplitter::handle { background: #2f8cff; }
        QSplitter::handle:vertical { height: 4px; }
        QSplitter::handle:horizontal { width: 4px; }
        """)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls(): event.acceptProposedAction()

    def dropEvent(self, event):
        paths = [
            Path(url.toLocalFile())
            for url in event.mimeData().urls()
            if url.isLocalFile()
        ]
        if paths:
            event.setDropAction(Qt.CopyAction)
            event.accept()
            self.request_import(paths)

    def _position_progress_overlay(self, overlay=None):
        # Progress is a regular layout widget in Commit0015d.
        return


    def _make_progress(self, title: str, maximum: int):
        panel = self.import_progress_panel
        panel.reset_for_use(title)

        if maximum <= 0:
            panel.progress.setRange(0, 0)
        else:
            panel.progress.setRange(0, max(1, int(maximum)))
            panel.progress.setValue(0)

        panel.show()
        self.root_layout.invalidate()
        self.root_layout.activate()
        self.centralWidget().updateGeometry()
        self.centralWidget().repaint()
        QApplication.processEvents()
        return panel


    def _progress_step(self, progress: QProgressDialog, value: int, text: str) -> bool:
        progress.setLabelText(text)
        if progress.maximum() > 0:
            progress.setValue(min(value, progress.maximum()))
        QApplication.processEvents()
        return not progress.wasCanceled()

    @staticmethod
    def _patient_axis_label(vector):
        components = [
            (float(vector[0]), "L", "R"),
            (float(vector[1]), "P", "A"),
            (float(vector[2]), "S", "I"),
        ]
        output = []
        for value, positive, negative in sorted(
            components, key=lambda item: abs(item[0]), reverse=True
        ):
            if abs(value) >= 1e-4:
                output.append(positive if value > 0 else negative)
        return "".join(output[:2])

    def _detect_image_plane(self):
        if self.orientation_plane_override != "Auto":
            return self.orientation_plane_override

        if self.current_ds is not None:
            try:
                orientation = [float(v) for v in self.current_ds.ImageOrientationPatient]
                row = np.asarray(orientation[:3], dtype=float)
                column = np.asarray(orientation[3:6], dtype=float)
                normal = np.cross(row, column)
                axis = int(np.argmax(np.abs(normal)))
                return ("Sagittal", "Coronal", "Axial")[axis]
            except Exception:
                pass

            metadata = " ".join(
                str(getattr(self.current_ds, name, "") or "")
                for name in (
                    "SeriesDescription", "ProtocolName",
                    "StudyDescription", "ImageComments",
                )
            ).lower()
            if "cor" in metadata:
                return "Coronal"
            if "sag" in metadata:
                return "Sagittal"
            if "axi" in metadata or "tra" in metadata:
                return "Axial"

        source_text = str(self.current_source or "").lower()
        if "cor" in source_text:
            return "Coronal"
        if "sag" in source_text:
            return "Sagittal"
        if "axi" in source_text or "tra" in source_text:
            return "Axial"
        if self.source_kind == "raw_file":
            return "Coronal"
        return "Axial"

    @staticmethod
    def _plane_default_labels(plane):
        presets = {
            "Axial": {"top": "A", "bottom": "P", "left": "L", "right": "R"},
            "Coronal": {"top": "S", "bottom": "I", "left": "L", "right": "R"},
            "Sagittal": {"top": "S", "bottom": "I", "left": "P", "right": "A"},
        }
        return dict(presets.get(plane, presets["Axial"]))

    def _effective_orientation_labels(self):
        """Return safe GE/JIS display labels for the current plane.

        DICOM geometry remains available in Orientation Engine v2 for plane
        detection and diagnostics, while the normal viewer display follows the
        requested console convention (Axial A-up/L-left).
        """
        plane = self._detect_image_plane()
        defaults = self._plane_default_labels(plane)
        overrides = getattr(self, "orientation_override", {}) or {}
        return {
            key: (str(overrides.get(key) or defaults[key]).strip().upper())
            for key in ("top", "bottom", "left", "right")
        }

    @staticmethod
    def _patient_axis_label(vector):
        """Return the dominant DICOM patient-axis label for a 3D vector."""
        try:
            x, y, z = [float(value) for value in vector[:3]]
        except Exception:
            return "?"

        absolute = [abs(x), abs(y), abs(z)]
        dominant = int(np.argmax(absolute))

        if dominant == 0:
            return "L" if x > 0 else "R"
        if dominant == 1:
            return "P" if y > 0 else "A"
        return "H" if z > 0 else "F"

    @staticmethod
    def _opposite_orientation_label(label):
        return {
            "L": "R",
            "R": "L",
            "A": "P",
            "P": "A",
            "H": "F",
            "F": "H",
        }.get(label, "?")

    def _dicom_orientation_labels(self, ds=None):
        """Return untransformed DICOM edge labels.

        The image display transform is applied exactly once later in
        ``_apply_image_orientation``. Applying it here as well caused the
        orientation labels to be transformed twice and drift from the pixels.
        """
        dataset = ds if ds is not None else self.current_ds
        engine = getattr(self, "orientation_engine", None) or OrientationEngine()
        geometry = engine.parse_geometry(dataset)
        if geometry is None:
            return None
        return engine.base_labels(geometry)

    def _apply_orientation_transform_to_labels(
        self,
        labels,
        rotation_degrees=0,
        flip_horizontal=False,
        flip_vertical=False,
    ):
        """Apply display rotation/flip to orientation labels."""
        if not labels:
            return labels

        result = dict(labels)
        rotation = int(rotation_degrees) % 360

        for _ in range(rotation // 90):
            result = {
                **result,
                "top": result["left"],
                "right": result["top"],
                "bottom": result["right"],
                "left": result["bottom"],
            }

        if flip_horizontal:
            result["left"], result["right"] = (
                result["right"],
                result["left"],
            )

        if flip_vertical:
            result["top"], result["bottom"] = (
                result["bottom"],
                result["top"],
            )

        return result

    def _current_original_display_transform(self):
        """Return the transform actually applied to the displayed pixels."""
        transform = getattr(self, "console_display_transform", DisplayTransform())
        if isinstance(transform, DisplayTransform):
            return {
                "rotation_degrees": transform.rotation_degrees,
                "flip_horizontal": transform.flip_horizontal,
                "flip_vertical": transform.flip_vertical,
                "transpose": transform.transpose,
                "y_axis_up": transform.y_axis_up,
            }
        return dict(transform or {})

    def _sync_labels_to_original_display(self, labels, transform):
        """Apply the exact viewer transform through Orientation Engine v2."""
        return OrientationEngine.apply_transform(
            labels,
            DisplayTransform.from_mapping(transform),
        )

    def _resolve_console_display_transform(self, ds):
        """Find a 2-D transform that makes image pixels and labels agree.

        The target is the configured GE/JIS convention for the detected plane.
        Only lossless 90-degree rotations and flips are considered.
        """
        engine = getattr(self, "orientation_engine", None) or OrientationEngine()
        geometry = engine.parse_geometry(ds)
        if geometry is None:
            return DisplayTransform()
        base = engine.base_labels(geometry)
        target = self._plane_default_labels(self._detect_image_plane())
        candidates = []
        for transpose in (False, True):
            for rotation in (0, 90, 180, 270):
                for flip_h in (False, True):
                    for flip_v in (False, True):
                        transform = DisplayTransform(
                            rotation_degrees=rotation,
                            flip_horizontal=flip_h,
                            flip_vertical=flip_v,
                            transpose=transpose,
                            y_axis_up=False,
                        )
                        labels = engine.apply_transform(base, transform)
                        score = sum(labels.get(k) == target.get(k) for k in target)
                        complexity = int(transpose) + rotation // 90 + int(flip_h) + int(flip_v)
                        candidates.append((score, -complexity, transform))
        return max(candidates, key=lambda item: (item[0], item[1]))[2]

    @staticmethod
    def _apply_display_transform_to_array(array, transform):
        result = np.asarray(array)
        if transform.transpose:
            result = np.transpose(result)
        turns = (int(transform.rotation_degrees) // 90) % 4
        if turns:
            # OrientationEngine uses clockwise 90-degree label rotation.
            result = np.rot90(result, k=-turns)
        if transform.flip_horizontal:
            result = np.fliplr(result)
        if transform.flip_vertical:
            result = np.flipud(result)
        return np.ascontiguousarray(result)

    def _orientation_pipeline_trace(self, base_labels, final_labels, transform):
        dataset = self.current_ds
        orientation = (
            getattr(dataset, "ImageOrientationPatient", None)
            if dataset is not None
            else None
        )
        image_position = (
            getattr(dataset, "ImagePositionPatient", None)
            if dataset is not None
            else None
        )
        return {
            "patient_position": str(
                getattr(dataset, "PatientPosition", "") or ""
            ) if dataset is not None else "",
            "image_orientation_patient": (
                [float(value) for value in orientation]
                if orientation is not None else None
            ),
            "image_position_patient": (
                [float(value) for value in image_position]
                if image_position is not None else None
            ),
            "base_edges": {
                key: base_labels.get(key)
                for key in ("top", "bottom", "left", "right")
            },
            "display_transform": dict(transform),
            "final_edges": {
                key: final_labels.get(key)
                for key in ("top", "bottom", "left", "right")
            },
        }

    def _apply_image_orientation(self):
        """Synchronize DICOM labels with the displayed Original image."""
        if not hasattr(self, "primary_panel"):
            return

        empty_values = {
            "top": "",
            "bottom": "",
            "left": "",
            "right": "",
        }

        try:
            base_labels = self._dicom_orientation_labels(self.current_ds)
            if base_labels is None:
                self.primary_panel.set_orientation_labels(empty_values)
                self.secondary_panel.set_orientation_labels(empty_values)
                self.current_orientation_labels = None
                self.orientation_pipeline_trace = None
                return

            transform = self._current_original_display_transform()
            labels = self._sync_labels_to_original_display(base_labels, transform)
            # A complete manual override remains available, but partial stale
            # overrides are ignored so they cannot disagree with image pixels.
            overrides = getattr(self, "orientation_override", {}) or {}
            override_edges = ("top", "bottom", "left", "right")
            # ``None`` means "use the detected/default value". Never convert
            # it to the visible string "NONE".
            if all(
                overrides.get(edge) is not None
                and str(overrides.get(edge)).strip()
                for edge in override_edges
            ):
                labels = {
                    edge: str(overrides[edge]).strip().upper()
                    for edge in override_edges
                }

            orientation_values = {
                edge: labels[edge]
                for edge in ("top", "bottom", "left", "right")
            }

            view_mode = str(
                getattr(self, "view_mode", "Both") or "Both"
            ).lower()

            # Orientation labels belong to the spatial-domain Original panel.
            # FFT/k-space never receives patient-direction labels because the
            # Fourier-domain axes are not anatomical patient axes.
            if view_mode == "both":
                self.primary_panel.set_orientation_labels(orientation_values)
                self.secondary_panel.set_orientation_labels(empty_values)
            elif view_mode in ("fft", "k-space", "kspace"):
                self.primary_panel.set_orientation_labels(empty_values)
                self.secondary_panel.set_orientation_labels(empty_values)
            else:
                # Original-only DICOM display. RAW files reach this branch only
                # when a valid linked DICOM dataset has been assigned; otherwise
                # base_labels is None and both panels were cleared above.
                self.primary_panel.set_orientation_labels(orientation_values)
                self.secondary_panel.set_orientation_labels(empty_values)

            self.current_orientation_labels = labels
            self.orientation_pipeline_trace = (
                self._orientation_pipeline_trace(
                    base_labels,
                    labels,
                    transform,
                )
            )

            try:
                self._update_annotation_display()
            except Exception as annotation_error:
                self.statusBar().showMessage(
                    "Orientation updated; "
                    f"annotation detail unavailable: {annotation_error}",
                    8000,
                )

        except Exception as orientation_error:
            for panel in (self.primary_panel, self.secondary_panel):
                try:
                    panel.set_orientation_labels(empty_values)
                except Exception:
                    pass
            self.current_orientation_labels = None
            self.orientation_pipeline_trace = {
                "error": str(orientation_error)
            }
            self.statusBar().showMessage(
                f"Orientation unavailable: {orientation_error}",
                8000,
            )


    def _transform_current_image(self, operation):
        if self.current_image is None:
            return

        def transform(array):
            if array is None:
                return None
            if operation == "Rotate Left":
                return np.rot90(array, 1)
            if operation == "Rotate Right":
                return np.rot90(array, -1)
            if operation == "Flip Horizontal":
                return np.fliplr(array)
            if operation == "Flip Vertical":
                return np.flipud(array)
            return array

        for name in (
            "current_image", "current_kspace", "current_recon",
            "compensation_base_image", "compensation_original",
            "compensation_base_kspace", "compensation_original_kspace",
            "addsub_preview_result",
        ):
            value = getattr(self, name, None)
            if value is not None and np.asarray(value).ndim >= 2:
                setattr(self, name, transform(np.asarray(value)))

        labels = self._effective_orientation_labels()
        if operation == "Rotate Left":
            new_labels = {
                "top": labels["right"], "bottom": labels["left"],
                "left": labels["top"], "right": labels["bottom"],
            }
        elif operation == "Rotate Right":
            new_labels = {
                "top": labels["left"], "bottom": labels["right"],
                "left": labels["bottom"], "right": labels["top"],
            }
        elif operation == "Flip Horizontal":
            new_labels = {
                **labels, "left": labels["right"], "right": labels["left"]
            }
        elif operation == "Flip Vertical":
            new_labels = {
                **labels, "top": labels["bottom"], "bottom": labels["top"]
            }
        else:
            new_labels = labels

        self.orientation_override = dict(new_labels)
        self.refresh_images()
        self._configure_line_control(*self.current_image.shape)
        self.update_line_profile()
        self._apply_image_orientation()

    def _open_orientation_dialog(self):
        """Open the orientation editor reliably from the global toolbar."""
        try:
            self.statusBar().showMessage("Opening Orientation settings...", 2500)
            self.edit_image_orientation()
        except Exception as error:
            self.statusBar().showMessage(
                f"Orientation dialog could not be opened: {error}", 8000
            )
            QMessageBox.warning(
                self,
                "Orientation",
                f"Orientation settings could not be opened.\n\n{error}",
            )

    def edit_image_orientation(self):
        if self.current_image is None:
            QMessageBox.information(
                self, "Orientation", "Load an image before changing orientation."
            )
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Image Orientation")
        dialog.resize(430, 320)
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.setWindowFlag(Qt.WindowStaysOnTopHint, True)
        form = QFormLayout(dialog)

        plane_combo = QComboBox()
        plane_combo.addItems(["Auto", "Axial", "Coronal", "Sagittal"])
        plane_combo.setCurrentText(self.orientation_plane_override)
        form.addRow("Scan plane", plane_combo)

        current = self._effective_orientation_labels()
        edits = {}
        for key, title in (
            ("top", "Top"), ("bottom", "Bottom"),
            ("left", "Left"), ("right", "Right"),
        ):
            edit = QLineEdit(current[key])
            edit.setMaxLength(3)
            edits[key] = edit
            form.addRow(title, edit)

        preset_row = QHBoxLayout()
        for plane in ("Axial", "Coronal", "Sagittal"):
            button = QPushButton(plane)
            def apply_preset(_checked=False, selected=plane):
                plane_combo.setCurrentText(selected)
                for key, value in self._plane_default_labels(selected).items():
                    edits[key].setText(value)
            button.clicked.connect(apply_preset)
            preset_row.addWidget(button)
        form.addRow("Preset", preset_row)

        transform_grid = QGridLayout()
        for title, row, column in (
            ("Rotate Left", 0, 0), ("Rotate Right", 0, 1),
            ("Flip Horizontal", 1, 0), ("Flip Vertical", 1, 1),
        ):
            button = QPushButton(title)
            def apply_transform(_checked=False, operation=title):
                # Apply immediately but keep the editor open so several
                # rotations/flips can be tested without reopening it.
                self._transform_current_image(operation)
                self._apply_image_orientation()
                dialog.raise_()
                dialog.activateWindow()
            button.clicked.connect(apply_transform)
            transform_grid.addWidget(button, row, column)
        form.addRow("Image transform", transform_grid)

        buttons = QHBoxLayout()
        reset_button = QPushButton("Use detected default")
        apply_button = QPushButton("Apply labels")
        cancel_button = QPushButton("Cancel")
        buttons.addWidget(reset_button)
        buttons.addStretch(1)
        buttons.addWidget(apply_button)
        buttons.addWidget(cancel_button)
        form.addRow(buttons)

        def restore_defaults():
            selected = plane_combo.currentText()
            plane = self._detect_image_plane() if selected == "Auto" else selected
            for key, value in self._plane_default_labels(plane).items():
                edits[key].setText(value)

        def apply_labels_live():
            selected_plane = plane_combo.currentText()
            self.orientation_plane_override = selected_plane
            plane = self._detect_image_plane() if selected_plane == "Auto" else selected_plane
            defaults = self._plane_default_labels(plane)
            self.orientation_override = {
                key: (
                    None
                    if edit.text().strip().upper() == defaults.get(key, "")
                    else edit.text().strip().upper()
                )
                for key, edit in edits.items()
            }
            self.refresh_images()
            self._apply_image_orientation()
            self.image_orientation_button.setText(
                f"Orientation: {self._detect_image_plane()}"
            )
            self.statusBar().showMessage("Orientation updated", 3500)
            dialog.raise_()
            dialog.activateWindow()

        reset_button.clicked.connect(restore_defaults)
        apply_button.clicked.connect(apply_labels_live)
        cancel_button.clicked.connect(dialog.reject)

        dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        dialog.exec()

    def _annotation_lines(self, mode=None):
        mode = mode or self.annotation_mode
        plane = self._detect_image_plane()
        labels = self._effective_orientation_labels()
        lines = [
            Path(self.current_source).name if self.current_source else "No file",
            f"{plane} | {labels['top']}/{labels['bottom']} "
            f"{labels['left']}/{labels['right']}",
        ]
        if self.current_image is not None:
            lines.append(f"{self.current_image.shape[0]}×{self.current_image.shape[1]}")
        if mode == "Minimum":
            return lines

        if self.current_ds is not None:
            ds = self.current_ds
            fields = [
                ("Patient", getattr(ds, "PatientName", "")),
                ("Study", getattr(ds, "StudyDescription", "")),
                ("Series", getattr(ds, "SeriesDescription", "")),
                ("Protocol", getattr(ds, "ProtocolName", "")),
                ("Modality", getattr(ds, "Modality", "")),
                ("Series No.", getattr(ds, "SeriesNumber", "")),
                ("Instance", getattr(ds, "InstanceNumber", "")),
                ("TE/TR", f"{getattr(ds, 'EchoTime', '')}/{getattr(ds, 'RepetitionTime', '')}"),
                ("Pixel spacing", getattr(ds, "PixelSpacing", "")),
            ]
            for label, value in fields:
                if str(value).strip() and str(value) != "/":
                    lines.append(f"{label}: {value}")
        elif self.current_raw_result is not None:
            result = self.current_raw_result
            lines.extend([
                f"RAW: {result.dtype_name}, {result.endian_name}",
                f"Confidence: {result.confidence:.0%}",
                f"Image-likeness: {getattr(result, 'fus_image_likeness', 0.0):.3f}",
                f"Offset: {result.offset_bytes} bytes",
            ])
        return lines

    def _legacy__update_annotation_display(self):
        value = "\n".join(self._annotation_lines())
        self.annotation_text.setText(value or "No annotation")
        if not self.annotation_visible:
            self.primary_panel.clear_annotation()
            self.secondary_panel.clear_annotation()
            return
        if self.view_mode == "FFT":
            self.primary_panel.clear_annotation()
            self.secondary_panel.clear_annotation()
        elif self.view_mode == "Original":
            self.primary_panel.set_annotation(value, self.annotation_mode)
            self.secondary_panel.clear_annotation()
        else:
            self.primary_panel.clear_annotation()
            self.secondary_panel.set_annotation(value, self.annotation_mode)


    def _update_annotation_display(self):
        self._legacy__update_annotation_display()

        labels = getattr(
            self,
            "current_orientation_labels",
            None,
        )
        if not labels:
            labels = self._dicom_orientation_labels(
                self.current_ds
            )
        if not labels:
            return

        patient_position = labels.get(
            "patient_position",
            "-",
        )
        row_vector = labels.get("row_vector")
        column_vector = labels.get("column_vector")
        image_position = labels.get("image_position")

        orientation_summary = (
            f"Patient Position: {patient_position}\n"
            f"Display Orientation: "
            f"Top={labels.get('top', '?')}  "
            f"Bottom={labels.get('bottom', '?')}  "
            f"Left={labels.get('left', '?')}  "
            f"Right={labels.get('right', '?')}"
        )

        if row_vector is not None:
            orientation_summary += (
                "\nImageOrientationPatient:"
                f"\n  Row: "
                f"[{row_vector[0]:.4f}, "
                f"{row_vector[1]:.4f}, "
                f"{row_vector[2]:.4f}]"
                f"\n  Col: "
                f"[{column_vector[0]:.4f}, "
                f"{column_vector[1]:.4f}, "
                f"{column_vector[2]:.4f}]"
            )

        if image_position is not None:
            orientation_summary += (
                "\nImagePositionPatient: "
                f"[{image_position[0]:.3f}, "
                f"{image_position[1]:.3f}, "
                f"{image_position[2]:.3f}]"
            )

        if hasattr(self, "annotation_text"):
            current = self.annotation_text.toPlainText()
            if orientation_summary not in current:
                if current.strip():
                    current += "\n\n"
                self.annotation_text.setPlainText(
                    current + orientation_summary
                )
        elif hasattr(self, "annotation_label"):
            current = self.annotation_label.text()
            if orientation_summary not in current:
                if current.strip():
                    current += "\n\n"
                self.annotation_label.setText(
                    current + orientation_summary
                )


    def _annotation_visibility_changed(self, checked):
        self.annotation_visible = bool(checked)
        self.annotation_panel.setVisible(self.annotation_visible)
        self._resize_left_annotation_area()
        self._update_annotation_display()
        self._schedule_responsive_layout()

    def _annotation_mode_changed(self, mode):
        self.annotation_mode = str(mode)
        self._resize_left_annotation_area()
        self._update_annotation_display()

    def _resize_left_annotation_area(self):
        if not hasattr(self, "left_content_splitter"):
            return

        total = max(
            sum(self.left_content_splitter.sizes()),
            self.left_content_splitter.height(),
            360,
        )

        current = self.left_content_splitter.sizes()
        remembered = self._user_left_split_sizes

        if len(remembered) == 3 and sum(remembered) > 0:
            message_height = max(80, remembered[2])
        elif len(current) == 3 and sum(current) > 0:
            message_height = max(80, current[2])
        else:
            message_height = 120

        upper_available = max(220, total - message_height)

        if not self.annotation_visible:
            target = [upper_available, 0, message_height]
        elif self.annotation_mode == "Minimum":
            target = [
                int(upper_available * 0.67),
                int(upper_available * 0.33),
                message_height,
            ]
        else:
            target = [
                int(upper_available * 0.50),
                int(upper_available * 0.50),
                message_height,
            ]

        self.left_content_splitter.setSizes(target)


    @staticmethod
    def _raw_image_quality(image):
        array = np.abs(np.asarray(image, dtype=float))
        finite = array[np.isfinite(array)]
        if finite.size < 256:
            return -1e9
        low, high = np.percentile(finite, [1.0, 99.0])
        if high <= low:
            return -1e9
        normalized = np.clip((array - low) / (high - low), 0.0, 1.0)
        gy, gx = np.gradient(normalized)
        gradient = np.hypot(gx, gy)
        smooth_difference = np.abs(
            normalized - median_filter_3x3_numpy(normalized)
        )
        center = normalized[
            normalized.shape[0] // 4: normalized.shape[0] * 3 // 4,
            normalized.shape[1] // 4: normalized.shape[1] * 3 // 4,
        ]
        return (
            float(np.var(center)) * 7.0
            + float(np.mean(gradient > np.percentile(gradient, 85))) * 2.0
            - float(np.mean(smooth_difference)) * 4.0
            - float(np.mean((normalized < 0.01) | (normalized > 0.99))) * 2.0
        )

    @staticmethod
    def _fus_preview_normalize(image):
        array = np.asarray(image)
        if np.iscomplexobj(array):
            array = np.abs(array)
        array = np.asarray(array, dtype=float)
        finite = array[np.isfinite(array)]
        if finite.size == 0:
            return np.zeros_like(array, dtype=float)
        low, high = np.percentile(finite, [1.0, 99.0])
        if not np.isfinite(low) or not np.isfinite(high) or high <= low:
            low = float(np.min(finite)); high = float(np.max(finite))
        if high <= low:
            return np.zeros_like(array, dtype=float)
        return np.clip((array-low)/(high-low), 0.0, 1.0)

    def _raw_display_candidates(self, result):
        raw = np.asarray(result.data)
        direct = np.abs(raw) if np.iscomplexobj(raw) else np.asarray(raw, dtype=float)
        reconstructed = np.abs(ifft2c(raw))
        return {
            "Direct Array": self._fus_preview_normalize(direct),
            "Reconstructed Image": self._fus_preview_normalize(reconstructed),
            "k-space Magnitude": self._fus_preview_normalize(np.log1p(np.abs(raw))),
        }

    def _auto_raw_display(self, result):
        candidates = self._raw_display_candidates(result)
        scores = {
            key: self._raw_image_quality(value)
            for key, value in candidates.items()
        }
        recommended = getattr(result, "recommended_display", "Auto")
        fus_likeness = float(getattr(result, "fus_image_likeness", 0.0))

        if recommended == "Direct Array" and fus_likeness >= 0.45:
            selected = "Direct Array"
            scores["Direct Array"] += 100.0
            return selected, candidates[selected], scores
        elif recommended == "Reconstructed Image":
            scores["Reconstructed Image"] += 1.2
        elif result.is_kspace:
            scores["Reconstructed Image"] += 0.75
        else:
            scores["Direct Array"] += 0.20

        selected = max(scores, key=scores.get)
        return selected, candidates[selected], scores

    def _change_raw_display_mode(self, mode):
        if self.current_raw_result is None:
            return
        candidates = self._raw_display_candidates(self.current_raw_result)
        if mode == "Auto":
            selected, image, _scores = self._auto_raw_display(
                self.current_raw_result
            )
        else:
            selected = mode
            image = candidates.get(mode, candidates["Direct Array"])

        self.raw_display_mode = mode
        self.current_image = np.asarray(image, dtype=float)
        if selected == "Reconstructed Image":
            self.current_kspace = np.asarray(self.current_raw_result.data)
        else:
            self.current_kspace = fft2c(self.current_image)
        self.current_recon = ifft2c(self.current_kspace)

        self.original_window_level = None
        self.original_dynamic_range = None
        self.raw_window_level = None
        self.raw_dynamic_range = None
        self.refresh_images()
        self._configure_line_control(*self.current_image.shape)
        self.update_line_profile()
        self._apply_image_orientation()
        self.statusBar().showMessage(f"RAW display: {selected}")

    def _display_loaded_raw_result(
        self,
        path: Path,
        result: RawImportResult,
        *,
        cached: bool = False,
    ):
        self.current_raw_result = result
        self.current_ds = None
        self.current_source = str(path)
        self.source_kind = "raw_file"
        self.slice_label.setText(
            f"RAW: {result.rows}×{result.cols}"
        )

        selected_mode, selected_image, raw_scores = self._auto_raw_display(result)
        image = np.asarray(selected_image, dtype=float)

        if image.ndim != 2 or image.size == 0:
            raise RawImportError(
                f"Decoded RAW preview has invalid shape: {image.shape}"
            )

        # FUS previews are normally normalized to 0–1. Ensure a finite visible
        # image even when cached data contains NaN/Inf or a flat range.
        image = np.nan_to_num(image, nan=0.0, posinf=1.0, neginf=0.0)
        finite = image[np.isfinite(image)]
        if finite.size:
            low, high = np.percentile(finite, [1.0, 99.0])
            if high > low:
                image = np.clip((image - low) / (high - low), 0.0, 1.0)
            elif float(np.max(finite)) > float(np.min(finite)):
                minimum = float(np.min(finite))
                maximum = float(np.max(finite))
                image = np.clip(
                    (image - minimum) / (maximum - minimum),
                    0.0,
                    1.0,
                )

        self.current_image = image
        if selected_mode == "Reconstructed Image":
            self.current_kspace = np.asarray(result.data)
        else:
            self.current_kspace = fft2c(self.current_image)
        self.current_recon = ifft2c(self.current_kspace)

        self.compensation_base_kspace = np.asarray(self.current_kspace).copy()
        self.compensation_base_image = np.asarray(self.current_image).copy()
        self.compensation_original = np.asarray(self.current_image).copy()
        self.compensation_original_kspace = np.asarray(
            self.current_kspace
        ).copy()
        self.compensation_history.clear()
        self.compensation_history_index = -1

        # Clear all previous-image display state. Without this reset, a DICOM
        # manual WW/WL can make a normalized RAW image appear completely black.
        self.original_window_level = 0.5
        self.original_dynamic_range = 1.0
        self.raw_window_level = None
        self.raw_dynamic_range = None

        self.level_target_combo.blockSignals(True)
        self.level_preset_combo.blockSignals(True)
        try:
            self.level_target_combo.setCurrentText("Original Image")
            self.level_preset_combo.setCurrentText("Manual")
            self.window_level_spin.setValue(0.5)
            self.dynamic_range_spin.setValue(1.0)
        finally:
            self.level_target_combo.blockSignals(False)
            self.level_preset_combo.blockSignals(False)

        self.raw_display_mode = "Auto"
        self.raw_display_combo.blockSignals(True)
        try:
            self.raw_display_combo.setCurrentText("Auto")
            self.raw_display_combo.setEnabled(True)
        finally:
            self.raw_display_combo.blockSignals(False)

        interpretation = (
            "Complex k-space"
            if result.is_kspace
            else "FUS direct image candidate"
        )
        cache_text = (
            "Predecoded during import"
            if cached
            else "Decoded on selection"
        )
        self.info.setText(
            f"File: {path.name}\n"
            f"Type: RAW ({interpretation})\n"
            f"Display: Auto → {selected_mode}\n"
            f"Preview: Exact FUS Commit0013 compatible\n"
            f"Load mode: {cache_text}\n"
            f"Matrix: {result.rows} × {result.cols}\n"
            f"Data type: {result.dtype_name}\n"
            f"Endian: {result.endian_name}\n"
            f"Confidence: {result.confidence:.0%}\n"
            f"Status: "
            f"{'Experimental Preview' if getattr(result, 'experimental', False) else 'Recognized RAW'}\n"
            f"FUS image-likeness: "
            f"{getattr(result, 'fus_image_likeness', 0.0):.3f}\n"
            f"Candidates: "
            f"{getattr(result, 'candidate_summary', '') or 'Primary decoder match'}\n"
            f"Reason: {result.reason}"
        )

        self._update_output_root_label()

        # Preserve the user's selected FFT / Original / Both layout.  The old
        # two-stage paint briefly showed Original and then forced Both, causing
        # duplicate refreshes and visible flicker.
        if self.view_mode not in {"FFT", "Original", "Both"}:
            self.view_mode = "Both"
        for button, value in ((self.btn_fft, "FFT"), (self.btn_original, "Original"), (self.btn_both, "Both")):
            button.setChecked(value == self.view_mode)
        self._configure_line_control(*self.current_image.shape)
        self.refresh_images()
        self._apply_image_orientation()
        self._schedule_responsive_layout()

        self.statusBar().showMessage(
            f"RAW displayed: {path.name} — "
            f"{result.rows}×{result.cols}, "
            f"image-likeness "
            f"{getattr(result, 'fus_image_likeness', 0.0):.3f}"
        )
        # RAW navigation depends on the displayed file and its Explorer
        # parent folder. Refresh Previous/Next immediately after every load.
        self._update_series_navigation_ui()


    def load_raw_file(self, path: Path):
        path = Path(path)
        cache_key = str(path)

        cached_result = self.raw_preview_cache.get(cache_key)
        if cached_result is not None:
            try:
                self._display_loaded_raw_result(
                    path,
                    cached_result,
                    cached=True,
                )
                return
            except Exception:
                # A stale or incompatible cached object must not block a fresh
                # exact decode.
                self.raw_preview_cache.pop(cache_key, None)

        # The exact FUS decoder is intentionally attempted synchronously first.
        # It is bounded to three dtypes and ten matrix widths and normally
        # completes fast enough for direct list selection.
        exact_result = try_render_fus_raw_exact(path)
        if exact_result is not None:
            self.raw_preview_cache[cache_key] = exact_result
            self._display_loaded_raw_result(path, exact_result, cached=False)
            return

        # Unsupported/proprietary layouts use the extended fallback. Keep the
        # progress UI only for this uncommon path.
        progress = self._make_progress("Analyzing RAW Data", 0)
        progress.setLabelText(path.name)
        QApplication.processEvents()

        try:
            result = load_raw_file_auto(
                path,
                cache_path=self._output_root() / "raw_import_profiles.json",
            )
        except RawImportError as exc:
            QMessageBox.warning(
                self,
                "RAW Import",
                f"{path.name}\n\n{exc}",
            )
            self.statusBar().showMessage(f"RAW not recognized: {path.name}")
            return
        except Exception as exc:
            QMessageBox.critical(
                self,
                "RAW Import Error",
                f"{path.name}\n\n{type(exc).__name__}: {exc}",
            )
            return
        finally:
            progress.hide()
            self._stabilize_layout()

        self.raw_preview_cache[cache_key] = result
        self._display_loaded_raw_result(path, result, cached=False)


    def load_bitmap_image(self, path: Path):
        qimage = QImage(str(path))
        if qimage.isNull():
            raise ValueError(f"Unable to read image: {path}")
        gray = qimage.convertToFormat(QImage.Format_Grayscale8)
        width = gray.width()
        height = gray.height()
        ptr = gray.bits()
        array = np.frombuffer(ptr, dtype=np.uint8, count=gray.sizeInBytes())
        array = array.reshape((height, gray.bytesPerLine()))[:, :width].copy()

        self.current_image = array.astype(float)
        self.current_kspace = fft2c(self.current_image)
        self.current_recon = ifft2c(self.current_kspace)
        self.current_ds = None
        self.current_source = str(path)
        self.source_kind = "bitmap"
        self.view_mode = "Original"
        self.original_window_level = None
        self.original_dynamic_range = None
        self.btn_original.setChecked(True)
        self.btn_fft.setChecked(False)
        self.btn_both.setChecked(False)

        self.processed_images[str(path)] = self.current_image.copy()
        self.processed_sources[str(path)] = path
        self._add_external_image_tree_item(path)
        self._update_output_root_label()
        self.refresh_images()
        self.info.setText(
            f"File: {path.name}\nType: Bitmap image\n"
            f"Size: {width} × {height}\nOriginal Image only"
        )

    def _add_external_image_tree_item(self, path: Path):
        root = None
        for index in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(index)
            if item.text(0) == "Image Files":
                root = item
                break
        if root is None:
            root = QTreeWidgetItem(["Image Files"])
            self.tree.addTopLevelItem(root)
        child = QTreeWidgetItem([path.name])
        child.setData(0, Qt.UserRole, ("processed", str(path)))
        root.addChild(child)
        root.setExpanded(True)
        self.tree.setCurrentItem(child)

    def _open_folder_direct(self):
        if self.import_in_progress:
            self.open_import_selection()
            return
        folder = QFileDialog.getExistingDirectory(
            self,
            "Select MRI Data Folder",
            "",
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks,
        )
        if folder:
            self.request_import([Path(folder)])

    def open_import_selection(self):
        if self.import_in_progress:
            QMessageBox.information(
                self,
                "Import in progress",
                "An import is already running.\n"
                "Wait for it to finish or press Cancel in the progress window.",
            )
            if hasattr(self, "_active_import_progress") and self._active_import_progress:
                self._active_import_progress.show()
                self._active_import_progress.raise_()
                self._active_import_progress.activateWindow()
            return

        dialog = QMessageBox(self)
        dialog.setWindowTitle("Import MRI Data")
        dialog.setText("Select how to import data.")
        files_button = dialog.addButton("Files", QMessageBox.AcceptRole)
        folder_button = dialog.addButton("Folder", QMessageBox.ActionRole)
        dialog.addButton("Cancel", QMessageBox.RejectRole)
        dialog.exec()

        selected_paths = []
        if dialog.clickedButton() == files_button:
            filenames, _ = QFileDialog.getOpenFileNames(
                self,
                "Select MRI Files",
                "",
                (
                    "Supported files "
                    "(*.dcm *.ima *.dicom *.zip *.raw *.bin *.dat *.7 *.pfile "
                    "*.img *.csv *.txt *.npy *.npz *.jpg *.jpeg *.png *.bmp);;"
                    "All files (*)"
                ),
            )
            selected_paths = [Path(name) for name in filenames]

        elif dialog.clickedButton() == folder_button:
            folder = QFileDialog.getExistingDirectory(
                self,
                "Select MRI Data Folder",
                "",
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks,
            )
            if folder:
                selected_paths = [Path(folder)]

        if selected_paths:
            self.request_import(selected_paths)

    def request_import(self, paths: list[Path]):
        # Commit0121: preserve the tab where the user dropped the files.
        self._import_target_tab = int(self.tabs.currentIndex()) if hasattr(self, "tabs") else 0
        clean_paths = [Path(path) for path in paths if Path(path).exists()]
        if not clean_paths:
            return

        first_source = clean_paths[0]
        if first_source.is_dir():
            base_parent = first_source
        else:
            base_parent = first_source.parent
        self.last_import_output_root = base_parent / "MR_Image_Explorer_Output"
        self._update_output_root_label()

        if self.import_in_progress:
            self.statusBar().showMessage("Import request ignored: import is already running")
            if hasattr(self, "_active_import_progress") and self._active_import_progress:
                self._active_import_progress.show()
                self._active_import_progress.raise_()
                self._active_import_progress.activateWindow()
            return

        # Defer execution until the current drag/click event is fully completed.
        QTimer.singleShot(0, lambda p=clean_paths: self.import_paths(p))

    @Slot(str, int, int)
    def _on_import_progress(self, message: str, value: int, maximum: int):
        progress = getattr(self, "_active_import_progress", None)
        if progress is None:
            return

        lines = []
        for line in str(message).splitlines():
            if len(line) > 96:
                line = "..." + line[-93:]
            lines.append(line)
        progress.setLabelText("\n".join(lines[:3]))

        if maximum <= 0:
            progress.progress.setRange(0, 0)
        else:
            if (
                progress.progress.minimum() != 0
                or progress.progress.maximum() != maximum
            ):
                progress.progress.setRange(0, maximum)
            progress.setValue(value)

        if not progress.isVisible():
            progress.show()

        self.root_layout.invalidate()
        self.centralWidget().updateGeometry()
        progress.repaint()
        self.centralWidget().repaint()
        QApplication.processEvents()


    def _save_current_import_history(self):
        if not self.dicom_entries and not self.pending_tracker_paths:
            return
        self.import_history.append({
            "dicom_entries": list(self.dicom_entries),
            "pending_tracker_paths": list(self.pending_tracker_paths),
            "imported_paths": list(self.imported_paths),
            "slice_index": int(getattr(self, "slice_index", 0)),
        })
        self.import_history = self.import_history[-5:]
        self.previous_import_button.setEnabled(bool(self.import_history))

    def restore_previous_import(self):
        if not self.import_history:
            return
        state = self.import_history.pop()
        self.dicom_entries = list(state["dicom_entries"])
        self.pending_tracker_paths = list(state["pending_tracker_paths"])
        self.imported_paths = list(state["imported_paths"])
        self.lazy_dicom_cache.clear()
        self.lazy_dicom_cache_order.clear()
        self.populate_dicom_tree()
        self.previous_import_button.setEnabled(bool(self.import_history))
        if self.dicom_entries:
            index = min(state["slice_index"], len(self.dicom_entries) - 1)
            QTimer.singleShot(0, lambda i=index: self.show_dicom(i))
        self.statusBar().showMessage("Previous import list restored")
        self._stabilize_layout()

    def _origin_for_path(self, path):
        return self.import_origin_info.get(
            str(path),
            {
                "container": Path(path).parent.name,
                "container_type": "File",
                "relative_path": Path(path).name,
                "group_path": "(Other)",
            },
        )

    def _ensure_group_item(self, parent, cache, key, title):
        if key not in cache:
            item = QTreeWidgetItem([title])
            parent.addChild(item)
            cache[key] = item
        return cache[key]

    def _add_grouped_path_items(
        self,
        root_title,
        paths,
        data_kind,
        text_builder,
    ):
        if not paths:
            return

        root = QTreeWidgetItem([f"{root_title} ({len(paths)})"])
        self.tree.addTopLevelItem(root)
        group_cache = {}

        for index, path in enumerate(paths):
            origin = self._origin_for_path(path)
            container = origin.get("container", "Other")
            group_path = origin.get("group_path", "(Root)")
            container_type = origin.get("container_type", "File")
            key = (container_type, container, group_path)
            title = f"{container_type}: {container} / {group_path}"
            group = self._ensure_group_item(
                root,
                group_cache,
                key,
                title,
            )

            item = QTreeWidgetItem([text_builder(path, index)])
            if data_kind == "tracker_pending":
                item.setData(0, Qt.UserRole, (data_kind, index))
            else:
                item.setData(0, Qt.UserRole, (data_kind, str(path)))
            item.setToolTip(
                0,
                f"{path}\n"
                f"Source: {container_type} {container}\n"
                f"Relative: {origin.get('relative_path', path.name)}",
            )
            group.addChild(item)

        for item in group_cache.values():
            item.setExpanded(False)
        root.setExpanded(True)

    def _on_import_completed(self, result: dict):
        try:
            self._save_current_import_history()
            for temp_dir in result.get("temp_dirs", []):
                if temp_dir not in self.temp_dirs:
                    self.temp_dirs.append(temp_dir)

            files = result.get("all_files", [])
            dicoms = result.get("dicoms", [])
            trackers = result.get("trackers", [])
            raw_files = result.get("raw_files", [])
            self.raw_preview_cache.update(
                result.get("raw_preview_cache", {}) or {}
            )
            self.import_origin_info = dict(
                result.get("origin_info", {}) or {}
            )
            self._active_tree_source_key = None
            bitmaps = result.get("bitmaps", [])
            signals = result.get("signals", [])
            skipped = int(result.get("skipped", 0))

            known = set()
            for path in self.imported_paths:
                try:
                    known.add(str(path.resolve()).lower())
                except Exception:
                    pass

            for path in files:
                try:
                    key = str(path.resolve()).lower()
                except Exception:
                    key = str(path).lower()
                if key not in known:
                    self.imported_paths.append(path)
                    known.add(key)

            self.tree.blockSignals(True)
            try:
                self.tree.clear()

                if dicoms:
                    self.dicom_entries = list(dicoms)
                    self.lazy_dicom_cache.clear()
                    self.lazy_dicom_cache_order.clear()
                    self.populate_dicom_tree()

                self.pending_tracker_paths = list(trackers)

                self._add_grouped_path_items(
                    "Tracker P Files",
                    trackers,
                    "tracker_pending",
                    lambda path, index: path.name,
                )

                def raw_text(path, index):
                    preview = self.raw_preview_cache.get(str(path))
                    if preview is None:
                        return path.name
                    status = (
                        "Experimental"
                        if getattr(preview, "experimental", False)
                        else "Recognized"
                    )
                    return (
                        f"{path.name}  |  {preview.rows}×{preview.cols} "
                        f"{preview.dtype_name}  "
                        f"score {preview.fus_image_likeness:.3f}  "
                        f"{status}"
                    )

                self._add_grouped_path_items(
                    "RAW / P / k-space Images",
                    raw_files,
                    "raw_file",
                    raw_text,
                )

                self._add_grouped_path_items(
                    "Bitmap Images",
                    bitmaps,
                    "bitmap_pending",
                    lambda path, index: (
                        f"{path.name}  |  {path.suffix.lower().lstrip('.').upper()}"
                    ),
                )

                if signals:
                    root = QTreeWidgetItem([f"1D / Array Data ({len(signals)})"])
                    self.tree.addTopLevelItem(root)
                    for path in signals:
                        item = QTreeWidgetItem([path.name])
                        item.setData(0, Qt.UserRole, ("signal_pending", str(path)))
                        item.setToolTip(0, str(path))
                        root.addChild(item)
                    root.setExpanded(True)
            finally:
                self.tree.blockSignals(False)

            total_recognized = (
                len(dicoms) + len(trackers) + len(raw_files)
                + len(bitmaps) + len(signals)
            )

            target_tab = int(getattr(self, "_import_target_tab", 0))
            if trackers:
                try:
                    self.load_tracker(Path(trackers[0]))
                except Exception as exc:
                    self.statusBar().showMessage(f"Tracker import failed: {type(exc).__name__}: {exc}")
                # Hub Auto Dataset Analyzer opens the production Image Workspace after import.
                # Manual tracker-only imports keep the dedicated Tracker Signal behavior.
                if getattr(self, "_hub_handoff_active", False):
                    self.tabs.setCurrentIndex(0)
                else:
                    self.tabs.setCurrentIndex(4 if target_tab == 4 else 3)
            elif signals and not (dicoms or raw_files or bitmaps):
                self.tabs.setCurrentIndex(4)
            else:
                # Images dropped in Spike/Artifact Diag remain in that diagnostic tab.
                self.tabs.setCurrentIndex(target_tab if target_tab in (1, 2) else 0)
            self.statusBar().showMessage(
                f"Inventory completed: {len(dicoms):,} DICOM images, "
                f"{len(raw_files):,} RAW images, "
                f"{len(bitmaps):,} bitmap images, "
                f"{len(trackers):,} Tracker, {len(signals):,} signal files"
            )

            if total_recognized == 0:
                QMessageBox.information(
                    self,
                    "Import completed",
                    f"No supported MRI data was recognized.\n"
                    f"Checked: {len(files):,}\nSkipped: {skipped:,}",
                )

            # Close progress first; decode the first image afterward.
            self._close_import_progress()
            if dicoms:
                def _show_first_imported_image():
                    self.show_dicom(0)
                    target = int(getattr(self, "_import_target_tab", 0))
                    if getattr(self, "_hub_handoff_active", False):
                        self.tabs.setCurrentIndex(0)
                    elif target in (1, 2):
                        self.tabs.setCurrentIndex(target)
                        if target == 1:
                            QTimer.singleShot(0, self._activate_spike_diag_from_workspace_selection)
                QTimer.singleShot(100, _show_first_imported_image)
            elif raw_files:
                first_raw = Path(raw_files[0])
                def _show_first_raw(path=first_raw):
                    self.load_raw_file(path)
                    if getattr(self, "_hub_handoff_active", False):
                        self.tabs.setCurrentIndex(0)
                QTimer.singleShot(100, _show_first_raw)

        except Exception as exc:
            self._on_import_failed(f"{type(exc).__name__}: {exc}")


    @Slot(str)
    def _on_import_failed(self, message: str):
        self._close_import_progress()
        QMessageBox.critical(self, "Import error", message)
        self.statusBar().showMessage(f"Import failed: {message}")

    @Slot()
    def _on_import_canceled(self):
        self._close_import_progress()
        self.statusBar().showMessage("Import canceled")

    @Slot()
    def _on_import_thread_finished(self):
        self.import_thread = None
        self.import_worker = None


    def _available_screen_geometry(self):
        screen = self.screen() or QApplication.primaryScreen()
        if screen is None:
            return None
        return screen.availableGeometry()

    def _initial_screen_fit(self):
        if self._screen_fit_completed:
            self._ensure_window_inside_screen()
            return
        geometry = self._available_screen_geometry()
        if geometry is None:
            return
        self.centralWidget().ensurePolished()
        self.centralWidget().layout().activate()
        QApplication.processEvents()
        max_width = max(820, int(geometry.width() * 0.94))
        max_height = max(560, int(geometry.height() * 0.92))
        target_width = min(max_width, 1280)
        target_height = min(max_height, 820)
        self.resize(target_width, target_height)
        frame = self.frameGeometry()
        frame.moveCenter(geometry.center())
        if frame.left() < geometry.left(): frame.moveLeft(geometry.left())
        if frame.top() < geometry.top(): frame.moveTop(geometry.top())
        if frame.right() > geometry.right(): frame.moveRight(geometry.right())
        if frame.bottom() > geometry.bottom(): frame.moveBottom(geometry.bottom())
        self.move(frame.topLeft())
        self._screen_fit_completed = True
        self._initial_fit_done = True
        self._apply_responsive_layout()


    def _save_viewer_layout(self):
        try:
            self.viewer_settings.setValue(
                "window_geometry",
                self.saveGeometry(),
            )
            self.viewer_settings.setValue(
                "main_split",
                self.main_split.sizes(),
            )
            self.viewer_settings.setValue(
                "left_split",
                self.left_content_splitter.sizes(),
            )
            self.viewer_settings.setValue(
                "vertical_split",
                self.vertical_splitter.sizes(),
            )
            self.viewer_settings.setValue(
                "image_split",
                self.image_splitter.sizes(),
            )
            self.viewer_settings.setValue(
                "viewer_content_split",
                self.viewer_content_splitter.sizes(),
            )
        except Exception:
            pass

    def _restore_viewer_layout(self):
        try:
            geometry = self.viewer_settings.value("window_geometry")
            if geometry is not None:
                self.restoreGeometry(geometry)

            for key, splitter in (
                ("main_split", self.main_split),
                ("left_split", self.left_content_splitter),
                ("vertical_split", self.vertical_splitter),
                ("image_split", self.image_splitter),
                ("viewer_content_split", self.viewer_content_splitter),
            ):
                sizes = self.viewer_settings.value(key)
                if isinstance(sizes, list) and sizes:
                    splitter.setSizes(
                        [int(value) for value in sizes]
                    )

            if len(self.main_split.sizes()) == 2:
                self._user_main_split_sizes = self.main_split.sizes()
            if len(self.left_content_splitter.sizes()) == 3:
                self._user_left_split_sizes = (
                    self.left_content_splitter.sizes()
                )
            if len(self.vertical_splitter.sizes()) == 2:
                self._user_vertical_split_sizes = (
                    self.vertical_splitter.sizes()
                )
            if len(self.image_splitter.sizes()) == 2:
                self._user_image_split_sizes = (
                    self.image_splitter.sizes()
                )
            if len(self.viewer_content_splitter.sizes()) == 2:
                sizes = self.viewer_content_splitter.sizes()
                if sizes[1] >= 200:
                    self._user_viewer_content_split_sizes = sizes
        except Exception:
            pass

    def _apply_default_image_geometry(self):
        """Apply the same centered Fit layout after selection and Display Reset.

        Use actual available width instead of the historical ``[1, 1]`` sizes.
        Qt may temporarily resolve tiny splitter values asymmetrically while a
        panel is being shown, which caused the first rendered frame to look
        biased until Display Reset was pressed.
        """
        if hasattr(self, "image_splitter") and self.view_mode == "Both":
            width = max(int(self.image_splitter.width()), 2)
            half = max(width // 2, 1)
            self.image_splitter.setSizes([half, width - half])
        for panel in (self.primary_panel, self.secondary_panel):
            if panel is self.secondary_panel and self.view_mode != "Both":
                continue
            try:
                rows, cols = panel.shape
                view_box = panel.plot.getViewBox()
                view_box.disableAutoRange()
                view_box.setRange(
                    xRange=(0.0, float(cols)),
                    yRange=(0.0, float(rows)),
                    padding=0.0,
                    update=False,
                )
            except Exception:
                pass

    def _render_current_image_atomically(self):
        """Render the current image while the viewer frame is locked.

        The previous sequence still allowed the splitters and ViewBoxes to paint
        once before the final fit.  This caused the visible "biyoyon" frame
        motion on slice changes.  Commit0073 freezes the whole viewer content,
        disables responsive relayout, applies the final splitter geometry first,
        pushes both images, and then releases updates once.
        """
        targets = []
        for name in (
            "viewer_content_splitter", "vertical_splitter", "image_splitter",
            "primary_panel", "secondary_panel", "centralWidget",
        ):
            try:
                target = self.centralWidget() if name == "centralWidget" else getattr(self, name, None)
            except Exception:
                target = None
            if target is not None and target not in targets:
                targets.append(target)
        self._frame_lock_active = True
        for target in targets:
            try:
                target.setUpdatesEnabled(False)
            except Exception:
                pass
        try:
            # Establish final layout before any panel receives a new image.
            self._apply_default_image_geometry()
            self.refresh_images()
            self._apply_default_image_geometry()
            self._apply_image_orientation()
            self._apply_frequency_markers()
            for panel in (self.primary_panel, self.secondary_panel):
                if panel.isVisible():
                    panel.fit_to_image()
        finally:
            self._frame_lock_active = False
            for target in reversed(targets):
                try:
                    target.setUpdatesEnabled(True)
                    target.update()
                except Exception:
                    pass
            try:
                self.update()
            except Exception:
                pass

    def reset_viewer_display(self):
        self._user_main_split_sizes = []
        self._user_left_split_sizes = []
        self._user_vertical_split_sizes = []
        self._user_image_split_sizes = []

        self.main_split.setSizes([260, 1080])
        if hasattr(self, "viewer_content_splitter"):
            self.viewer_content_splitter.setSizes([820, 260])
        self.left_content_splitter.setSizes([500, 0, 130])
        self.vertical_splitter.setSizes([620, 220])
        self.image_splitter.setSizes([1, 1])
        self._apply_default_image_geometry()
        self._profile_curtain_toggled(
            self.profile_accordion.button.isChecked()
        )
        self._right_tools_curtain_toggled(
            self.right_tools_accordion.button.isChecked()
        )

        self.annotation_toggle.setChecked(False)
        self.annotation_mode_combo.setCurrentText("Minimum")

        self.orientation_plane_override = "Auto"
        self.orientation_override = {
            "top": None,
            "bottom": None,
            "left": None,
            "right": None,
        }

        self.original_window_level = None
        self.original_dynamic_range = None
        self.raw_window_level = None
        self.raw_dynamic_range = None

        for panel in (
            self.primary_panel,
            self.secondary_panel,
        ):
            try:
                panel.fit_to_image()
            except Exception:
                pass

        if self.current_image is not None:
            self.refresh_images()
            self._apply_image_orientation()

        self.statusBar().showMessage(
            "Display and layout reset.",
            6000,
        )

    def _remember_left_splitter_sizes(self):
        if hasattr(self, "left_content_splitter"):
            sizes = self.left_content_splitter.sizes()
            if len(sizes) == 3 and sum(sizes) > 0:
                self._user_left_split_sizes = sizes

    def _remember_splitter_sizes(self):
        if hasattr(self, "main_split"):
            self._user_main_split_sizes = self.main_split.sizes()
        if hasattr(self, "vertical_splitter"):
            sizes = self.vertical_splitter.sizes()
            expanded = bool(getattr(self, "profile_accordion", None) and self.profile_accordion.button.isChecked())
            if (not getattr(self, "_programmatic_splitter_change", False)
                    and expanded and len(sizes) == 2 and sizes[1] >= 180):
                self._user_vertical_split_sizes = list(sizes)
        if hasattr(self, "image_splitter"):
            sizes = self.image_splitter.sizes()
            if len(sizes) == 2 and sum(sizes) > 0:
                self._user_image_split_sizes = sizes
        if hasattr(self, "viewer_content_splitter"):
            sizes = self.viewer_content_splitter.sizes()
            if len(sizes) == 2 and sizes[1] >= 200:
                self._user_viewer_content_split_sizes = sizes

    def _schedule_responsive_layout(self):
        if getattr(self, "_frame_lock_active", False):
            return
        if self._responsive_pending:
            return
        self._responsive_pending = True
        def apply():
            self._responsive_pending = False
            self._apply_responsive_layout()
        QTimer.singleShot(0, apply)
        QTimer.singleShot(100, self._apply_responsive_layout)
        QTimer.singleShot(260, self._apply_responsive_layout)


    def _apply_responsive_layout(self):
        if getattr(self, "_frame_lock_active", False):
            return
        if not hasattr(self, "main_split"):
            return

        width = max(self.centralWidget().width(), 900)
        height = max(self.centralWidget().height(), 600)

        # Phase 1 layout uses two outer columns:
        # Explorer | Viewer workspace.  The viewer workspace then contains
        # Image area | Tool accordion below the shared top toolbar.
        if width < 1180:
            left_width = 150
            right_width = 250
        elif width < 1450:
            left_width = 190
            right_width = 285
        else:
            left_width = 230
            right_width = 330

        viewer_width = max(620, width - left_width - 30)
        center_width = max(420, viewer_width - right_width - 10)

        remembered = self._user_main_split_sizes
        if len(remembered) == 2 and sum(remembered) > 0:
            available = max(width - 30, 800)
            remembered_total = sum(remembered)
            scaled = [
                max(120, int(value * available / remembered_total))
                for value in remembered
            ]
            scaled[1] = max(600, scaled[1])
            self.main_split.setSizes(scaled)
        else:
            self.main_split.setSizes([left_width, viewer_width])

        if hasattr(self, "viewer_content_splitter"):
            right_expanded = self.right_tools_accordion.button.isChecked()
            if right_expanded:
                remembered = getattr(self, "_user_viewer_content_split_sizes", [])
                if len(remembered) == 2 and sum(remembered) > 0 and remembered[1] >= 200:
                    total = max(1, center_width + right_width)
                    ratio = remembered[1] / max(1, sum(remembered))
                    kept_right = max(250, min(520, int(total * ratio)))
                    self.viewer_content_splitter.setSizes([max(420, total - kept_right), kept_right])
                else:
                    self.viewer_content_splitter.setSizes([center_width, right_width])
            else:
                header_w = 32
                self.viewer_content_splitter.setSizes([max(1, viewer_width - header_w), header_w])

        self._resize_left_annotation_area()

        profile_expanded = self.profile_accordion.button.isChecked()
        if profile_expanded:
            remembered = getattr(self, "_user_vertical_split_sizes", [])
            if len(remembered) == 2 and sum(remembered) > 0 and remembered[1] > 80:
                total = max(1, self.vertical_splitter.height())
                ratio = remembered[1] / max(1, sum(remembered))
                kept_lower = max(180, min(max(180, total - 320), int(total * ratio)))
                self._set_vertical_splitter_sizes([max(320, total - kept_lower), kept_lower])
            else:
                available_tab_height = max(self.tabs.height(), height - 140)
                lower_height = 180 if available_tab_height < 820 else 220
                upper_height = max(340, available_tab_height - lower_height - 80)
                self._set_vertical_splitter_sizes([upper_height, lower_height])
        else:
            header_h = self.profile_accordion.button.sizeHint().height() + 8
            self._set_vertical_splitter_sizes([max(1, self.vertical_splitter.height() - header_h), header_h])

        if self.view_mode == "Both" and self.secondary_panel.isVisible():
            remembered = getattr(self, "_user_image_split_sizes", [])
            if len(remembered) == 2 and sum(remembered) > 0 and min(remembered) > 0:
                total = max(480, center_width)
                first = max(200, int(total * remembered[0] / sum(remembered)))
                self.image_splitter.setSizes([first, max(200, total - first)])
            else:
                half = max(240, center_width // 2)
                self.image_splitter.setSizes([half, half])
        else:
            self.image_splitter.setSizes([max(480, center_width), 0])

        self.root_layout.invalidate()
        self.root_layout.activate()
        self.centralWidget().updateGeometry()
        self.centralWidget().repaint()

    def _ensure_window_inside_screen(self):
        geometry = self._available_screen_geometry()
        if geometry is None:
            return
        frame = self.frameGeometry()
        if frame.width() > geometry.width():
            self.resize(int(geometry.width() * 0.96), self.height())
            frame = self.frameGeometry()
        if frame.height() > geometry.height():
            self.resize(self.width(), int(geometry.height() * 0.94))
            frame = self.frameGeometry()
        if not geometry.intersects(frame) or not geometry.contains(frame.center()):
            frame.moveCenter(geometry.center())
            self.move(frame.topLeft())

    def _stabilize_layout(self):
        if getattr(self, "_frame_lock_active", False):
            return
        try:
            self._ensure_window_inside_screen()
            self._apply_responsive_layout()
            self.root_layout.invalidate()
            self.root_layout.activate()
            self.centralWidget().updateGeometry()
            self.tabs.updateGeometry()
            self.image_splitter.updateGeometry()
            self.centralWidget().repaint()
            self.repaint()
        except Exception:
            pass


    def _close_import_progress(self):
        self.import_poll_timer.stop()
        self.import_in_progress = False
        self.import_thread_started = False
        self.import_python_thread = None
        self.import_cancel_event.clear()

        self.drop_banner.setEnabled(True)
        self.drop_banner.title.setText(
            "Click or Drop DICOM / Raw / P File / 1D Data / Folder / ZIP Here"
        )

        progress = getattr(self, "_active_import_progress", None)
        self._active_import_progress = None
        if progress is not None:
            progress.cancel_button.setEnabled(True)
            progress.hide()

        self._stabilize_layout()
        QTimer.singleShot(0, self._stabilize_layout)
        QTimer.singleShot(120, self._stabilize_layout)
        QTimer.singleShot(300, self._schedule_responsive_layout)
        QApplication.processEvents()


    def _queue_import_event(self, kind: str, payload=None):
        self.import_event_queue.put((kind, payload, time.monotonic()))

    def _import_thread_main(self, paths: list[Path]):
        temp_dirs = []
        try:
            supported = {
                ".dcm", ".ima", ".dicom", "",
                ".raw", ".bin", ".dat", ".7", ".pfile", ".img", ".kspace", ".cfl", ".rawdata", ".complex",
                ".csv", ".npy", ".npz",
                ".jpg", ".jpeg", ".png", ".bmp",
            }
            files = []
            checked = 0
            self._queue_import_event("progress", ("Collecting files...", 0, 0))

            for source in paths:
                if self.import_cancel_event.is_set():
                    self._queue_import_event("canceled")
                    return
                source = Path(source)
                if source.is_dir():
                    for candidate in source.rglob("*"):
                        if self.import_cancel_event.is_set():
                            self._queue_import_event("canceled")
                            return
                        checked += 1
                        if candidate.is_file() and candidate.suffix.lower() in supported:
                            files.append(candidate)
                        if checked % 100 == 0:
                            self._queue_import_event(
                                "progress",
                                (f"Scanning folder...\nChecked: {checked:,}\nCandidates: {len(files):,}", 0, 0),
                            )
                elif source.is_file() and source.suffix.lower() == ".zip":
                    root = Path(tempfile.mkdtemp(prefix="mr_image_explorer_"))
                    temp_dirs.append(root)
                    with zipfile.ZipFile(source) as archive:
                        members = archive.infolist()
                        total_members = max(len(members), 1)
                        for member_index, member in enumerate(members, start=1):
                            if self.import_cancel_event.is_set():
                                self._queue_import_event("canceled")
                                return
                            self._queue_import_event(
                                "progress",
                                (f"Extracting ZIP {member_index:,}/{total_members:,}\n{member.filename}", member_index, total_members),
                            )
                            target = (root / member.filename).resolve()
                            if not str(target).startswith(str(root.resolve())):
                                raise ValueError("Unsafe ZIP member path")
                            if member.is_dir():
                                target.mkdir(parents=True, exist_ok=True)
                                continue
                            target.parent.mkdir(parents=True, exist_ok=True)
                            with archive.open(member) as src, target.open("wb") as dst:
                                while True:
                                    if self.import_cancel_event.is_set():
                                        self._queue_import_event("canceled")
                                        return
                                    chunk = src.read(1024 * 1024)
                                    if not chunk:
                                        break
                                    dst.write(chunk)
                                    self._queue_import_event(
                                        "progress",
                                        (
                                            f"Extracting ZIP {member_index:,}/{total_members:,}\n"
                                            f"{member.filename}",
                                            member_index,
                                            total_members,
                                        ),
                                    )
                            if target.suffix.lower() in supported:
                                files.append(target)
                elif source.is_file():
                    files.append(source)

            unique=[]; seen=set()
            for path in files:
                try: key=str(path.resolve()).lower()
                except Exception: key=str(path).lower()
                if key not in seen:
                    seen.add(key); unique.append(path)
            files=unique
            total=max(len(files),1)
            dicoms=[]; trackers=[]; raws=[]; bitmaps=[]; signals=[]; skipped=0

            for idx,path in enumerate(files, start=1):
                if self.import_cancel_event.is_set():
                    self._queue_import_event("canceled"); return
                self._queue_import_event("progress", (f"Indexing {idx:,}/{len(files):,}\n{path.name}", idx, total))
                suffix=path.suffix.lower()
                if suffix==".txt":
                    continue
                if suffix in BITMAP_EXTENSIONS:
                    bitmaps.append(path); continue
                if suffix in SIGNAL_EXTENSIONS:
                    signals.append(path); continue
                try:
                    ds=pydicom.dcmread(str(path), stop_before_pixels=True, defer_size="1 KB", force=False)
                    rows=int(getattr(ds,"Rows",0) or 0); cols=int(getattr(ds,"Columns",0) or 0)
                    if rows<=0 or cols<=0: raise ValueError("not image")
                    instance=int(getattr(ds,"InstanceNumber",0) or 0)
                    pos=getattr(ds,"ImagePositionPatient",None)
                    z=float(pos[2]) if pos is not None and len(pos)>=3 else float(instance)
                    series=str(getattr(ds,"SeriesInstanceUID","") or "")
                    dicoms.append(DicomEntry(path=path, ds=ds, image=None, sort_key=(series,z,instance,path.name.lower())))
                    continue
                except Exception:
                    pass
                try:
                    name=path.name.lower(); is_tracker=("trackerimg" in name or "pfile" in name)
                    if not is_tracker and path.suffix=="" and path.stat().st_size>4096:
                        with path.open("rb") as fh: first=fh.read(4)
                        if len(first)==4:
                            rev=np.frombuffer(first,dtype="<f4")[0]
                            is_tracker=bool(np.isfinite(rev) and 1<rev<100)
                    if is_tracker: trackers.append(path)
                    elif suffix in {".raw",".bin",".dat",".7",".pfile",".img",".kspace",".cfl",".rawdata",".complex"}: raws.append(path)
                    else: skipped+=1
                except Exception:
                    skipped+=1

            dicoms.sort(key=lambda entry: entry.sort_key)
            self._queue_import_event("completed", {
                "all_files": files, "dicoms": dicoms, "trackers": trackers,
                "raw_files": raws, "bitmaps": bitmaps, "signals": signals,
                "skipped": skipped, "temp_dirs": temp_dirs,
            })
        except Exception as exc:
            self._queue_import_event("failed", f"{type(exc).__name__}: {exc}")

    def _poll_import_queue(self):
        handled = False

        while True:
            try:
                kind, payload, event_time = self.import_event_queue.get_nowait()
            except queue.Empty:
                break

            handled = True
            self.import_last_event_time = event_time

            if kind == "progress":
                message, value, maximum = payload
                self._on_import_progress(message, value, maximum)

            elif kind == "completed":
                self.import_poll_timer.stop()
                self._on_import_completed(payload)
                return

            elif kind == "failed":
                self.import_poll_timer.stop()
                self._on_import_failed(payload)
                return

            elif kind == "canceled":
                self.import_poll_timer.stop()
                self._on_import_canceled()
                return

        thread = self.import_python_thread
        if not self.import_in_progress or thread is None:
            return

        # Never treat a worker as stopped before thread.start() has completed.
        if not self.import_thread_started:
            return

        # Give the worker a startup grace period before declaring abnormal exit.
        elapsed = time.monotonic() - self.import_start_time
        if elapsed < 1.5:
            return

        if not thread.is_alive() and not handled:
            self.import_poll_timer.stop()

            # Drain once more in case the worker exited immediately after
            # placing its final event in the queue.
            try:
                kind, payload, event_time = self.import_event_queue.get_nowait()
            except queue.Empty:
                self._on_import_failed(
                    "Import worker stopped without returning a result."
                )
                return

            self.import_last_event_time = event_time
            if kind == "completed":
                self._on_import_completed(payload)
            elif kind == "failed":
                self._on_import_failed(payload)
            elif kind == "canceled":
                self._on_import_canceled()
            elif kind == "progress":
                message, value, maximum = payload
                self._on_import_progress(message, value, maximum)
                self._on_import_failed(
                    "Import worker ended after its final progress update "
                    "without returning a completion result."
                )


    def _cancel_python_import(self):
        self.import_cancel_event.set()
        progress = getattr(self, "_active_import_progress", None)
        if progress is not None:
            progress.setLabelText("Canceling import...")
            progress.cancel_button.setEnabled(False)

    def import_paths(self, paths: list[Path]):
        if self.import_in_progress:
            progress = getattr(self, "_active_import_progress", None)
            if progress is not None:
                progress.show()
                self.root_layout.invalidate()
                self.centralWidget().repaint()
                QApplication.processEvents()
            return

        clean_paths = [Path(path) for path in paths if Path(path).exists()]
        if not clean_paths:
            QMessageBox.information(
                self,
                "Import MRI Data",
                "The selected path no longer exists.",
            )
            return

        self.import_in_progress = True
        self.import_thread_started = False
        self.import_start_time = time.monotonic()
        self.import_cancel_event.clear()

        while True:
            try:
                self.import_event_queue.get_nowait()
            except queue.Empty:
                break

        self.drop_banner.setEnabled(False)
        self.drop_banner.title.setText("Importing MRI Data...")
        self.drop_banner.repaint()
        QApplication.processEvents()

        progress = self._make_progress("Importing MRI Data", 1)
        progress.progress.setRange(0, 0)
        progress.setLabelText("Preparing import...")
        self._active_import_progress = progress
        progress.canceled.connect(self._cancel_python_import)

        progress.show()
        progress.raise_()
        QApplication.processEvents()

        thread = threading.Thread(
            target=self._import_thread_main,
            args=(clean_paths,),
            daemon=True,
            name="MRImageExplorerImport",
        )
        self.import_python_thread = thread

        # Start the worker first. Only then start the watchdog/poll timer.
        try:
            thread.start()
            self.import_thread_started = True
        except Exception as exc:
            self._on_import_failed(
                f"Unable to start Import worker: {type(exc).__name__}: {exc}"
            )
            return

        self.import_last_event_time = time.monotonic()
        self.import_poll_timer.start()






    def extract_zip(self, path: Path, parent_progress: Optional[QProgressDialog] = None) -> list[Path]:
        root = Path(tempfile.mkdtemp(prefix='mri_fft_'))
        self.temp_dirs.append(root)
        with zipfile.ZipFile(path) as z:
            members = z.infolist()
            for index, m in enumerate(members, start=1):
                if parent_progress is not None:
                    parent_progress.setLabelText(f"Extracting {path.name}: {index}/{len(members)}")
                    QApplication.processEvents()
                    if parent_progress.wasCanceled():
                        return []
                target = (root / m.filename).resolve()
                if not str(target).startswith(str(root.resolve())):
                    raise ValueError("Unsafe ZIP member")
                z.extract(m, root)
        return [p for p in root.rglob('*') if p.is_file()]

    def read_dicom_metadata(self, path: Path) -> DicomEntry:
        # Header-only read. Pixel Data is not decompressed during initial import.
        ds = pydicom.dcmread(
            str(path),
            stop_before_pixels=True,
            defer_size="1 KB",
            force=False,
        )
        rows = int(getattr(ds, "Rows", 0) or 0)
        cols = int(getattr(ds, "Columns", 0) or 0)
        if rows <= 0 or cols <= 0:
            raise ValueError("Not an image DICOM")

        instance = int(getattr(ds, "InstanceNumber", 0) or 0)
        pos = getattr(ds, "ImagePositionPatient", None)
        z = float(pos[2]) if pos is not None and len(pos) >= 3 else float(instance)
        series = str(getattr(ds, "SeriesInstanceUID", "") or "")
        sort_key = (series, z, instance, path.name.lower())
        return DicomEntry(path=path, ds=ds, image=None, sort_key=sort_key)

    def _evict_dicom_pixels(self, index: int) -> None:
        """Release only decoded pixels; header metadata remains available."""
        if 0 <= int(index) < len(self.dicom_entries):
            self.dicom_entries[int(index)].image = None

    def _touch_dicom_cache(self, index: int) -> None:
        self.lazy_dicom_cache.touch(int(index))
        self.lazy_dicom_cache_order = list(self.lazy_dicom_cache.snapshot())

    def _ensure_dicom_image(self, index: int) -> DicomEntry:
        if not (0 <= index < len(self.dicom_entries)):
            raise IndexError(index)
        entry = self.dicom_entries[index]
        if entry.image is None:
            full_entry = self.read_dicom(entry.path)
            entry.ds = full_entry.ds
            entry.image = full_entry.image
            entry.sort_key = full_entry.sort_key

        # A cache hit must also refresh recency. Header metadata is never evicted.
        self._touch_dicom_cache(index)
        return entry

    def read_dicom(self, path: Path) -> DicomEntry:
        ds = pydicom.dcmread(str(path), force=False)
        if not hasattr(ds, 'PixelData'): raise ValueError('No pixel data')
        decode_result = decode_dicom_pixels(ds)
        arr = decode_result.array
        if arr.ndim == 3:
            arr = arr[0]
        if arr.ndim != 2:
            raise ValueError('Not 2D')
        arr = arr.astype(float) * float(getattr(ds, 'RescaleSlope', 1)) + float(getattr(ds, 'RescaleIntercept', 0))
        instance = int(getattr(ds, 'InstanceNumber', 0) or 0)
        pos = getattr(ds, 'ImagePositionPatient', None)
        z = float(pos[2]) if pos is not None and len(pos) >= 3 else float(instance)
        series = str(getattr(ds, 'SeriesInstanceUID', '') or '')
        # Keep the same stable key used by the metadata-only import path.
        sort_key = (series, z, instance, path.name.lower())
        return DicomEntry(path, ds, arr, sort_key)

    @staticmethod
    def _explorer_series_purpose(ds):
        value = " ".join(
            str(getattr(ds, name, "") or "")
            for name in (
                "SeriesDescription",
                "ProtocolName",
                "SequenceName",
                "ImageType",
            )
        ).upper()
        if "TMAP" in value or "TEMP MAP" in value:
            return "Temperature Map"
        if "MEMP" in value:
            return "Thermometry MEMP"
        if any(token in value for token in ("LOCALIZER", "SCOUT", "3-PLANE")):
            return "Localizer"
        if any(token in value for token in ("PLAN", "PLANNING", "TARGET")):
            return "Planning"
        if any(token in value for token in ("DWI", "DIFFUSION", "ADC")):
            return "Diffusion"
        if "T2" in value:
            return "T2"
        if "T1" in value:
            return "T1"
        return "Other MRI"

    @staticmethod
    def _safe_series_number(ds):
        try:
            return int(getattr(ds, "SeriesNumber", 0) or 0)
        except Exception:
            return 0

    @staticmethod
    def _safe_instance_number(ds, fallback):
        try:
            return int(getattr(ds, "InstanceNumber", fallback) or fallback)
        except Exception:
            return int(fallback)

    def _capture_explorer_state(self):
        expanded = set()
        selected = []
        iterator = QTreeWidgetItemIterator(self.tree)
        while iterator.value():
            item = iterator.value()
            data = item.data(0, Qt.UserRole)
            if data:
                key = tuple(data)
                if item.isExpanded():
                    expanded.add(key)
                if item.isSelected():
                    selected.append(key)
            iterator += 1
        return {
            "expanded": expanded,
            "selected": selected,
            "scroll": self.tree.verticalScrollBar().value(),
        }

    def _restore_explorer_state(self, state):
        selected = set(state.get("selected", []))
        first_selected = None
        iterator = QTreeWidgetItemIterator(self.tree)
        while iterator.value():
            item = iterator.value()
            data = item.data(0, Qt.UserRole)
            key = tuple(data) if data else None
            if key in state.get("expanded", set()):
                item.setExpanded(True)
            if key in selected:
                item.setSelected(True)
                first_selected = first_selected or item
            iterator += 1

        if first_selected is not None:
            self.tree.setCurrentItem(first_selected)

        QTimer.singleShot(
            0,
            lambda: self.tree.verticalScrollBar().setValue(
                state.get("scroll", 0)
            ),
        )

    def _clear_explorer_filters(self):
        self.series_filter_combo.blockSignals(True)
        self.explorer_search_edit.blockSignals(True)
        self.explorer_type_combo.blockSignals(True)
        try:
            self.series_filter_combo.setCurrentText("All Series")
            self.explorer_search_edit.clear()
            self.explorer_type_combo.setCurrentText("All")
        finally:
            self.series_filter_combo.blockSignals(False)
            self.explorer_search_edit.blockSignals(False)
            self.explorer_type_combo.blockSignals(False)
        self._apply_explorer_filters()

    def _rebuild_explorer_preserving_state(self, *_):
        state = self._capture_explorer_state()
        self.populate_dicom_tree()
        self._restore_explorer_state(state)

    def _apply_explorer_filters(self, *_):
        series_filter = self.series_filter_combo.currentText()
        search_text = self.explorer_search_edit.text().strip().lower()
        type_filter = self.explorer_type_combo.currentText()

        def filter_item(item):
            children_visible = False
            for child_index in range(item.childCount()):
                if filter_item(item.child(child_index)):
                    children_visible = True

            data = item.data(0, Qt.UserRole)
            own_visible = True
            if data:
                kind = data[0]
                if kind == "series" and series_filter != "All Series":
                    own_visible = item.text(0) == series_filter
                elif kind == "dicom":
                    own_visible = type_filter in ("All", "DICOM")
                elif kind == "raw_file":
                    own_visible = type_filter in ("All", "RAW")
                elif kind == "bitmap_pending":
                    own_visible = type_filter in ("All", "Bitmap")
                elif str(kind).startswith("tracker"):
                    own_visible = type_filter in ("All", "Tracking")

            if search_text:
                own_visible = (
                    own_visible
                    and search_text in item.text(0).lower()
                )

            visible = own_visible or children_visible
            item.setHidden(not visible)
            return visible

        root = self.tree.invisibleRootItem()
        for index in range(root.childCount()):
            filter_item(root.child(index))

    def populate_dicom_tree(self):
        state = (
            self._capture_explorer_state()
            if self.tree.topLevelItemCount()
            else {"expanded": set(), "selected": [], "scroll": 0}
        )

        self.tree.blockSignals(True)
        try:
            self.tree.clear()
            root = QTreeWidgetItem(
                [f"DICOM Images ({len(self.dicom_entries)})"]
            )
            root.setData(0, Qt.UserRole, ("dicom_root", "all"))
            self.tree.addTopLevelItem(root)

            patients = {}
            for index, entry in enumerate(self.dicom_entries):
                ds = entry.ds
                patient_name = str(
                    getattr(ds, "PatientName", "")
                    or getattr(ds, "PatientID", "")
                    or "Unknown Patient"
                )
                patient_id = str(
                    getattr(ds, "PatientID", "") or patient_name
                )
                study_date = str(
                    getattr(ds, "StudyDate", "") or "Unknown Date"
                )
                study_description = str(
                    getattr(ds, "StudyDescription", "") or "MRI Exam"
                )
                study_uid = str(
                    getattr(ds, "StudyInstanceUID", "")
                    or f"{patient_id}:{study_date}:{study_description}"
                )
                series_number = self._safe_series_number(ds)
                series_description = str(
                    getattr(ds, "SeriesDescription", "")
                    or getattr(ds, "ProtocolName", "")
                    or "Unnamed Series"
                )
                series_uid = str(
                    getattr(ds, "SeriesInstanceUID", "")
                    or f"{study_uid}:{series_number}:{series_description}"
                )

                patients.setdefault(
                    (patient_name, patient_id), {}
                ).setdefault(
                    (study_date, study_description, study_uid), {}
                ).setdefault(
                    (series_number, series_description, series_uid), []
                ).append((index, entry))

            series_filter_labels = []

            for (patient_name, patient_id), exams in sorted(
                patients.items(),
                key=lambda item: (
                    item[0][0].lower(),
                    item[0][1].lower(),
                ),
            ):
                patient_item = QTreeWidgetItem(
                    [f"Patient: {patient_name}"]
                )
                patient_item.setData(
                    0, Qt.UserRole, ("patient", patient_id)
                )
                root.addChild(patient_item)

                for (
                    study_date,
                    study_description,
                    study_uid,
                ), series_map in sorted(
                    exams.items(),
                    key=lambda item: (
                        item[0][0],
                        item[0][1].lower(),
                    ),
                ):
                    exam_item = QTreeWidgetItem(
                        [f"Exam: {study_date} | {study_description}"]
                    )
                    exam_item.setData(
                        0, Qt.UserRole, ("exam", study_uid)
                    )
                    patient_item.addChild(exam_item)

                    for (
                        series_number,
                        series_description,
                        series_uid,
                    ), entries in sorted(
                        series_map.items(),
                        key=lambda item: (
                            item[0][0],
                            item[0][1].lower(),
                        ),
                    ):
                        prefix = (
                            f"Series {series_number:02d}"
                            if series_number
                            else "Series —"
                        )
                        label = (
                            f"{prefix} | {series_description} | "
                            f"{len(entries)} images"
                        )
                        series_filter_labels.append(label)

                        series_item = QTreeWidgetItem([label])
                        series_item.setData(
                            0, Qt.UserRole, ("series", series_uid)
                        )
                        exam_item.addChild(series_item)

                        entries.sort(
                            key=lambda pair: (
                                self._safe_instance_number(
                                    pair[1].ds,
                                    pair[0] + 1,
                                ),
                                pair[1].path.name.lower(),
                            )
                        )

                        for index, entry in entries:
                            instance = self._safe_instance_number(
                                entry.ds,
                                index + 1,
                            )
                            image_item = QTreeWidgetItem(
                                [f"{entry.path.name}    Inst {instance}"]
                            )
                            image_item.setData(
                                0, Qt.UserRole, ("dicom", index)
                            )
                            image_item.setFlags(
                                image_item.flags()
                                | Qt.ItemIsUserCheckable
                                | Qt.ItemIsSelectable
                                | Qt.ItemIsEnabled
                            )
                            image_item.setCheckState(0, Qt.Unchecked)
                            series_item.addChild(image_item)

                        series_item.setExpanded(False)

                    exam_item.setExpanded(True)

                patient_item.setExpanded(True)

            root.setExpanded(True)

            previous = self.series_filter_combo.currentText()
            self.series_filter_combo.blockSignals(True)
            self.series_filter_combo.clear()
            self.series_filter_combo.addItem("All Series")
            self.series_filter_combo.addItems(series_filter_labels)
            self.series_filter_combo.setCurrentText(
                previous
                if previous in series_filter_labels
                else "All Series"
            )
            self.series_filter_combo.blockSignals(False)

        finally:
            self.tree.blockSignals(False)

        self._restore_explorer_state(state)
        self._apply_explorer_filters()


    def _record_performance_metric(self, name: str, seconds: float, **context):
        values = self.performance_metrics.setdefault(name, [])
        values.append(float(seconds))
        if len(values) > 200:
            del values[:-200]
        if os.environ.get("MR_IMAGE_PERFORMANCE_LOG", "0") == "1":
            stats = self.fft_cache.stats()
            self.stable_diagnostics.info(
                "PERFORMANCE_METRIC",
                metric=name,
                seconds=round(float(seconds), 6),
                fft_cache_hits=stats.hits,
                fft_cache_misses=stats.misses,
                fft_cache_hit_rate=round(stats.hit_rate, 4),
                **context,
            )

    def _fft_cache_key_for_current_image(self, source_key: str = "") -> tuple:
        image = np.asarray(self.current_image)
        transform = getattr(self, "console_display_transform", None)
        return (
            str(source_key or self.current_source or "memory"),
            tuple(image.shape),
            str(image.dtype),
            repr(transform),
        )

    def _fft_for_current_image(self, source_key: str = "") -> tuple[np.ndarray, np.ndarray, bool]:
        key = self._fft_cache_key_for_current_image(source_key)
        cached = self.fft_cache.get(key)
        if cached is not None:
            kspace, recon = cached
            return np.asarray(kspace), np.asarray(recon), True
        started = time.perf_counter()
        kspace = fft2c(np.asarray(self.current_image, dtype=float))
        recon = ifft2c(kspace)
        self.fft_cache.put(key, (kspace, recon))
        self._record_performance_metric(
            "fft_seconds", time.perf_counter() - started, source=str(source_key)
        )
        return kspace, recon, False

    def show_dicom(self, index: int):
        self.stable_diagnostics.info(
            "SHOW_DICOM_BEGIN",
            index=int(index),
            entries=len(self.dicom_entries),
        )
        if not self.dicom_entries:
            return

        index = max(0, min(int(index), len(self.dicom_entries) - 1))

        decode_started = time.perf_counter()
        try:
            entry = self._ensure_dicom_image(index)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "DICOM Load Error",
                f"Unable to load image:\n{self.dicom_entries[index].path}\n\n{exc}",
            )
            return
        self._record_performance_metric(
            "dicom_decode_seconds",
            time.perf_counter() - decode_started,
            index=int(index),
            source=str(entry.path),
        )

        self.current_ds = entry.ds
        self.console_display_transform = self._resolve_console_display_transform(self.current_ds)
        self.current_image = self._apply_display_transform_to_array(
            np.asarray(entry.image).copy(), self.console_display_transform
        )
        self.current_source = str(entry.path)
        self.current_kspace, self.current_recon, fft_cache_hit = self._fft_for_current_image(
            self.current_source
        )
        self.source_kind = "dicom"
        self.raw_display_combo.setEnabled(False)
        self.slice_index = index

        self.compensation_base_image = np.asarray(self.current_image).copy()
        self.compensation_base_kspace = np.asarray(self.current_kspace).copy()
        self.compensation_history = []
        self.compensation_history_index = -1
        self.compensation_preview = None
        self._update_compensation_history_buttons()

        self.original_window_level = None
        self.original_dynamic_range = None
        if self.raw_window_level is None or self.raw_dynamic_range is None:
            self._set_levels_for_target("Raw Data", np.log1p(np.abs(self.current_kspace)))

        # Preserve the user's current Single/FFT/Both layout while paging.
        # Loading a new slice must never force the viewer back to two panels.
        if self.source_kind == "bitmap":
            self.view_mode = "Original"
        for button, value in [
            (self.btn_fft, "FFT"),
            (self.btn_original, "Original"),
            (self.btn_both, "Both"),
        ]:
            button.setChecked(value == self.view_mode)

        self._update_series_navigation_ui()
        self.info.setText(self.dicom_info(entry.ds, entry.path, index))
        self._select_tree_dicom_index(index)
        self._update_output_root_label()
        render_started = time.perf_counter()
        self._render_current_image_atomically()
        self._record_performance_metric(
            "render_seconds",
            time.perf_counter() - render_started,
            index=int(index),
            fft_cache_hit=bool(fft_cache_hit),
        )
        # Geometry is finalized inside the atomic render.  Do not queue a
        # delayed stabilizer here; it is the source of the visible frame
        # rebound during image navigation.
        self.stable_diagnostics.info(
            "SHOW_DICOM_END",
            state=self._stable_diagnostic_state(),
        )


    def dicom_info(self, ds, path, index):
        spacing = getattr(ds, 'PixelSpacing', ['-', '-'])
        return (f"File: {path.name}\nType: DICOM\nSize: {getattr(ds,'Columns','-')} × {getattr(ds,'Rows','-')}\n"
                f"Bit Depth: {getattr(ds,'BitsAllocated','-')} bit\nPixel Spacing: {spacing}\n"
                f"Series: {getattr(ds,'SeriesDescription','-')}\nInstance: {index+1}/{len(self.dicom_entries)}\n"
                f"TE / TR: {getattr(ds,'EchoTime','-')} / {getattr(ds,'RepetitionTime','-')} ms")

    @staticmethod
    def _entry_series_key(entry):
        """Return a conservative key that cannot merge unrelated series.

        SeriesInstanceUID is authoritative when present.  Older/converted
        datasets sometimes omit it, so the fallback includes study, sequence,
        geometry, image type and source folder instead of collapsing every
        missing-UID image into one global list.
        """
        ds = getattr(entry, "ds", None)
        if ds is None:
            return ("path", str(getattr(entry, "path", "")))
        uid = str(getattr(ds, "SeriesInstanceUID", "") or "").strip()
        study = str(getattr(ds, "StudyInstanceUID", "") or "").strip()
        def text(name):
            return str(getattr(ds, name, "") or "").strip()
        # Keep the same boundaries shown by the explorer tree even when a
        # vendor/exporter incorrectly reuses a SeriesInstanceUID.
        if uid:
            return (
                "uid", study, uid, text("SeriesNumber"),
                text("SeriesDescription"), text("ProtocolName"),
                text("SequenceName"), text("AcquisitionNumber"),
                text("EchoNumbers"), text("TemporalPositionIdentifier"),
            )
        orientation = tuple(round(float(v), 5) for v in (getattr(ds, "ImageOrientationPatient", None) or []))
        image_type = tuple(str(v).strip() for v in (getattr(ds, "ImageType", None) or []))
        path = Path(getattr(entry, "path", ""))
        return (
            "fallback", study, text("SeriesNumber"), text("SeriesDescription"),
            text("ProtocolName"), text("SequenceName"), text("AcquisitionNumber"),
            text("EchoNumbers"), text("TemporalPositionIdentifier"), image_type,
            int(getattr(ds, "Rows", 0) or 0), int(getattr(ds, "Columns", 0) or 0),
            orientation, str(path.parent).lower(),
        )

    @staticmethod
    def _entry_study_key(entry):
        ds = getattr(entry, "ds", None)
        if ds is None:
            return ("path", str(Path(getattr(entry, "path", "")).parent).lower())
        study_uid = str(getattr(ds, "StudyInstanceUID", "") or "").strip()
        if study_uid:
            return ("uid", study_uid)
        patient_id = str(getattr(ds, "PatientID", "") or "").strip()
        study_date = str(getattr(ds, "StudyDate", "") or "").strip()
        study_description = str(getattr(ds, "StudyDescription", "") or "").strip()
        return ("fallback", patient_id, study_date, study_description)

    def _entry_navigation_series_key(self, entry):
        """Return exactly the same series boundary used by the Explorer tree.

        Navigation must not use acquisition/echo/temporal fields here because
        those values may vary image-by-image inside one displayed series.
        """
        ds = getattr(entry, "ds", None)
        if ds is None:
            return (self._entry_study_key(entry), "path", str(Path(getattr(entry, "path", "")).parent).lower())
        study_key = self._entry_study_key(entry)
        series_uid = str(getattr(ds, "SeriesInstanceUID", "") or "").strip()
        series_number = self._safe_series_number(ds)
        series_description = str(
            getattr(ds, "SeriesDescription", "")
            or getattr(ds, "ProtocolName", "")
            or "Unnamed Series"
        ).strip()
        if series_uid:
            return (study_key, series_number, series_description, series_uid)
        return (
            study_key,
            series_number,
            series_description,
            str(Path(getattr(entry, "path", "")).parent).lower(),
        )

    def _current_series_indices(self):
        if not self.dicom_entries:
            return []
        current_index = max(
            0, min(int(getattr(self, "slice_index", 0)), len(self.dicom_entries) - 1)
        )
        key = self._entry_navigation_series_key(self.dicom_entries[current_index])
        indices = [
            index for index, entry in enumerate(self.dicom_entries)
            if self._entry_navigation_series_key(entry) == key
        ]
        indices.sort(key=lambda idx: (
            self._safe_instance_number(self.dicom_entries[idx].ds, idx + 1),
            self.dicom_entries[idx].path.name.lower(),
        ))
        return indices or [current_index]

    def _ensure_navigation_controller(self):
        controller = getattr(self, "navigation_controller", None)
        if controller is None:
            controller = NavigationController(self)
            self.navigation_controller = controller
        return controller

    def _update_series_navigation_ui(self):
        self._ensure_navigation_controller().update_ui()

    def change_slice(self, delta):
        """Navigate inside the current DICOM series."""
        return self._ensure_navigation_controller().navigate_series(delta)

    def _ordered_series_groups(self):
        """Return all series in the current Study in Explorer display order."""
        if not self.dicom_entries:
            return []
        current_index = max(
            0, min(int(getattr(self, "slice_index", 0)), len(self.dicom_entries) - 1)
        )
        current_study = self._entry_study_key(self.dicom_entries[current_index])
        grouped = {}
        for index, entry in enumerate(self.dicom_entries):
            if self._entry_study_key(entry) != current_study:
                continue
            key = self._entry_navigation_series_key(entry)
            grouped.setdefault(key, []).append(index)

        groups = []
        for key, indices in grouped.items():
            indices.sort(key=lambda idx: (
                self._safe_instance_number(self.dicom_entries[idx].ds, idx + 1),
                self.dicom_entries[idx].path.name.lower(),
            ))
            first_ds = self.dicom_entries[indices[0]].ds
            groups.append((
                key,
                indices,
                self._safe_series_number(first_ds),
                str(
                    getattr(first_ds, "SeriesDescription", "")
                    or getattr(first_ds, "ProtocolName", "")
                    or "Unnamed Series"
                ).lower(),
            ))
        groups.sort(key=lambda value: (value[2], value[3], value[1][0]))
        return [(key, indices) for key, indices, _number, _description in groups]

    def _find_tree_dicom_item(self, index: int):
        """Resolve a DICOM leaf from the current, possibly rebuilt Explorer tree."""
        iterator = QTreeWidgetItemIterator(self.tree)
        while iterator.value():
            item = iterator.value()
            data = item.data(0, Qt.UserRole)
            if data and data[0] == "dicom" and int(data[1]) == int(index):
                return item
            iterator += 1
        return None

    def _set_tree_series_expansion_for_index(self, index: int):
        """Reliably expose the destination DICOM series and image.

        Explorer filters may rebuild the QTreeWidget, invalidating every item
        reference obtained before the rebuild.  Therefore filters are cleared
        first, then the destination item is resolved again from the live tree.
        The target series is expanded, all sibling series in the same Exam are
        collapsed, ancestors are expanded, and the destination image is made
        current and visible.
        """
        # Clear every filter that can hide or rebuild the destination branch.
        filters_changed = False
        if hasattr(self, "series_filter_combo") and self.series_filter_combo.currentText() != "All Series":
            self.series_filter_combo.blockSignals(True)
            self.series_filter_combo.setCurrentText("All Series")
            self.series_filter_combo.blockSignals(False)
            filters_changed = True
        if hasattr(self, "explorer_search_edit") and self.explorer_search_edit.text():
            self.explorer_search_edit.blockSignals(True)
            self.explorer_search_edit.clear()
            self.explorer_search_edit.blockSignals(False)
            filters_changed = True
        if hasattr(self, "explorer_type_combo") and self.explorer_type_combo.currentText() not in {"All", "All Types"}:
            self.explorer_type_combo.blockSignals(True)
            preferred = "All Types" if self.explorer_type_combo.findText("All Types") >= 0 else "All"
            self.explorer_type_combo.setCurrentText(preferred)
            self.explorer_type_combo.blockSignals(False)
            filters_changed = True
        if filters_changed:
            self._apply_explorer_filters()

        # Resolve only after any possible tree rebuild.
        target_item = self._find_tree_dicom_item(index)
        if target_item is None:
            return False
        target_series = target_item.parent()
        target_exam = target_series.parent() if target_series is not None else None

        self.tree.blockSignals(True)
        try:
            if target_exam is not None:
                for position in range(target_exam.childCount()):
                    sibling = target_exam.child(position)
                    sibling_data = sibling.data(0, Qt.UserRole)
                    if sibling_data and sibling_data[0] == "series":
                        sibling.setExpanded(sibling is target_series)
            elif target_series is not None:
                target_series.setExpanded(True)

            parent = target_series
            while parent is not None:
                parent.setExpanded(True)
                parent = parent.parent()

            self.tree.setCurrentItem(target_item)
            target_item.setSelected(True)
        finally:
            self.tree.blockSignals(False)

        # Ensure Qt has laid out the expanded branch before scrolling.
        self.tree.scrollToItem(target_item)
        self.tree.viewport().update()
        return bool(target_series is None or target_series.isExpanded())

    def _navigate_tree_source_continuous(self, delta):
        """Navigate RAW/bitmap sibling items in explorer order.

        RAW files do not have DICOM SeriesInstanceUIDs, so series expansion is
        represented by their explorer parent folder/group.
        """
        current = self.tree.currentItem()
        if current is None:
            return False
        candidates = []
        iterator = QTreeWidgetItemIterator(self.tree)
        while iterator.value():
            item = iterator.value()
            data = item.data(0, Qt.UserRole)
            if data and data[0] in {"raw_file", "bitmap_pending", "tracker_file", "tracker_pending"}:
                candidates.append(item)
            iterator += 1
        if current not in candidates or not candidates:
            return False
        pos = candidates.index(current)
        target_pos = pos + (-1 if int(delta) < 0 else 1)
        if not (0 <= target_pos < len(candidates)):
            return False
        target = candidates[target_pos]
        old_parent, new_parent = current.parent(), target.parent()
        if old_parent is not None and old_parent is not new_parent:
            old_parent.setExpanded(False)
        parent = new_parent
        while parent is not None:
            parent.setExpanded(True)
            parent = parent.parent()
        self.tree.setCurrentItem(target)
        self.tree.scrollToItem(target)
        self._open_tree_item(target, force=True)
        return True

    def change_slice_continuous(self, delta):
        """Navigate continuously across DICOM series or explorer groups."""
        return self._ensure_navigation_controller().navigate_continuous(delta)

    def looks_tracker(self, p: Path) -> bool:
        n = p.name.lower()
        if 'trackerimg' in n or 'pfile' in n: return True
        if p.suffix == '' and p.stat().st_size > 4096:
            try:
                rev = np.frombuffer(p.read_bytes()[:4], dtype='<f4')[0]
                return np.isfinite(rev) and 1 < rev < 100
            except Exception: return False
        return False

    def show_tracker_in_workspace(self):
        if self.tracker_matrix is None:
            QMessageBox.information(self, "Tracker", "Load a Tracker file first.")
            return
        self.workspace_source_combo.setCurrentText("Tracker Raw Magnitude")
        self.current_kspace = np.asarray(self.tracker_matrix).copy()
        self.current_recon = ifft2c(self.current_kspace)
        self.current_image = np.abs(self.current_recon)
        self.current_source = str(self.tracker_source_path or "Tracker")
        self.tabs.setCurrentIndex(0)
        self.refresh_images()

    def _capture_tracker_state(self, path: Path, matrix: np.ndarray, matrix_size: int, offset: int):
        magnitude = np.abs(matrix).astype(np.float64)
        metrics=[]
        for i,line in enumerate(magnitude):
            metrics.append({"line":i,"peak":float(np.max(line)),"rms":float(np.sqrt(np.mean(line**2))),"energy":float(np.sum(line**2)),"mean":float(np.mean(line)),"variance":float(np.var(line)),"relative":0.0})
        rms=np.array([m["rms"] for m in metrics],float); lo=float(np.min(rms)); hi=float(np.max(rms)); span=max(hi-lo,np.finfo(float).eps)
        for m in metrics: m["relative"]=100.0*(m["rms"]-lo)/span
        strongest=max(metrics,key=lambda m:m["rms"]); idx=int(strongest["line"])
        return {"path":path,"matrix":np.asarray(matrix).copy(),"matrix_size":matrix_size,"offset":offset,"line_metrics":metrics,"strongest_line_index":idx,"strongest_line":np.asarray(matrix[idx]).copy(),"reconstructed_image":np.abs(ifft2c(matrix))}

    def _apply_tracker_state(self, index: int):
        if not (0 <= index < len(self.tracker_files)): return
        self.tracker_file_index=index; s=self.tracker_files[index]
        self.tracker_source_path=s["path"]; self.tracker_matrix=np.asarray(s["matrix"]).copy(); self.tracker_line_metrics=list(s["line_metrics"])
        self.tracker_strongest_line_index=int(s["strongest_line_index"]); self.tracker_strongest_line=np.asarray(s["strongest_line"]).copy(); self.tracker_reconstructed_image=np.asarray(s["reconstructed_image"]).copy(); self.tracker_shared_name=s["path"].name; self.tracker_selected_lines=[self.tracker_strongest_line_index]
        self.tracker_file_combo.blockSignals(True); self.tracker_file_combo.clear(); self.tracker_file_combo.addItems([x["path"].name for x in self.tracker_files]); self.tracker_file_combo.setCurrentIndex(index); self.tracker_file_combo.blockSignals(False)
        self._refresh_tracker_ranking(); self._plot_tracker_selected_lines(); self._update_tracker_navigation_buttons()
        self.tracker_summary.setText(f"{s['path'].name} | {index+1}/{len(self.tracker_files)} | Matrix {s['matrix_size']} × {s['matrix_size']} | Strongest line {self.tracker_strongest_line_index}")
        signal_name=f"{s['path'].name} | Strongest line {self.tracker_strongest_line_index}"
        self.signals=[x for x in self.signals if not str(x.get('name','')).startswith(f"{s['path'].name} | Strongest line")]
        self.signals.append({"name":signal_name,"data":self.tracker_strongest_line.copy()}); self.signal_combo.clear(); self.signal_combo.addItems([x["name"] for x in self.signals])
        if self.workspace_source_combo.currentText().startswith("Tracker"): self.change_workspace_source(self.workspace_source_combo.currentText())

    def _update_tracker_navigation_buttons(self):
        has=bool(self.tracker_files)
        self.tracker_prev_file_button.setEnabled(has and self.tracker_file_index>0); self.tracker_next_file_button.setEnabled(has and self.tracker_file_index<len(self.tracker_files)-1); self.tracker_clear_current_button.setEnabled(has); self.tracker_clear_all_button.setEnabled(has)

    def select_tracker_file_index(self, index:int):
        if 0 <= index < len(self.tracker_files): self._apply_tracker_state(index)

    def navigate_tracker_file(self, step:int):
        target=self.tracker_file_index+int(step)
        if 0 <= target < len(self.tracker_files): self._apply_tracker_state(target)

    def clear_current_tracker(self):
        if not self.tracker_files: return
        if QMessageBox.question(self,"Clear Tracker",f"Remove current Tracker file?\n{self.tracker_files[self.tracker_file_index]['path'].name}") != QMessageBox.Yes: return
        self.tracker_files.pop(self.tracker_file_index)
        if self.tracker_files: self._apply_tracker_state(min(self.tracker_file_index,len(self.tracker_files)-1))
        else: self._reset_tracker_state()

    def clear_all_trackers(self):
        if not self.tracker_files: return
        if QMessageBox.question(self,"Clear All Trackers","Clear all imported Tracker files?") != QMessageBox.Yes: return
        self.tracker_files.clear(); self._reset_tracker_state()

    def _reset_tracker_state(self):
        self.tracker_file_index=-1; self.tracker_matrix=None; self.tracker_source_path=None; self.tracker_line_metrics=[]; self.tracker_selected_lines=[]; self.tracker_strongest_line=None; self.tracker_strongest_line_index=-1; self.tracker_reconstructed_image=None; self.tracker_shared_name=""
        self.tracker_file_combo.clear(); self.tracker_table.setRowCount(0); self.tracker_raw_plot.clear(); self.tracker_summary.setText("No Tracker file loaded."); self.tracker_metrics_label.setText("-"); self._update_tracker_navigation_buttons()

    def keyPressEvent(self, event):
        if self.tabs.currentIndex() in (3,4,5,6,7):
            if event.key()==Qt.Key_Left: self.navigate_tracker_file(-1); event.accept(); return
            if event.key()==Qt.Key_Right: self.navigate_tracker_file(1); event.accept(); return
        super().keyPressEvent(event)

    def load_tracker(self, path: Path):
        raw = path.read_bytes()
        best = None

        for matrix in (512, 384, 320, 256, 192, 160, 128, 96, 64):
            need = matrix * matrix * 4
            if len(raw) < need:
                continue

            offset = len(raw) - need
            vals = np.frombuffer(
                raw,
                dtype="<i2",
                count=matrix * matrix * 2,
                offset=offset,
            )
            complex_data = (
                vals[0::2].astype(float)
                + 1j * vals[1::2].astype(float)
            )
            raw_matrix = complex_data.reshape(matrix, matrix)

            line_rms = np.sqrt(np.mean(np.abs(raw_matrix) ** 2, axis=1))
            score = float(
                np.percentile(line_rms, 99)
                / (np.median(line_rms) + 1e-9)
            )
            if best is None or score > best[0]:
                best = (score, matrix, offset, raw_matrix)

        if best is None:
            raise ValueError("No Tracker matrix candidate found")

        _, matrix, offset, raw_matrix = best
        state = self._capture_tracker_state(
            path, raw_matrix, matrix, offset
        )

        txt_path = path.with_suffix(".txt")
        if txt_path.exists():
            state["analysis_text"] = txt_path.read_text(
                encoding="utf-8", errors="replace"
            )
        else:
            state["analysis_text"] = ""

        existing = next(
            (
                i for i, item in enumerate(self.tracker_files)
                if item["path"] == path
            ),
            None,
        )
        if existing is None:
            self.tracker_files.append(state)
            index = len(self.tracker_files) - 1
        else:
            self.tracker_files[existing] = state
            index = existing

        self._apply_tracker_state(index)
        self.statusBar().showMessage(
            f"Loaded Tracker {index + 1}/{len(self.tracker_files)}: {path.name}"
        )


    def load_signal_file(self, path: Path):
        if path.suffix.lower() == '.npy': arr = np.load(path, allow_pickle=False)
        elif path.suffix.lower() == '.npz':
            z = np.load(path, allow_pickle=False); arr = z[list(z.keys())[0]]
        else:
            rows=[]
            with open(path, encoding='utf-8-sig', errors='replace') as f:
                for line in f:
                    nums=[]
                    for token in line.strip().replace(';', ',').split(','):
                        try: nums.append(float(token.strip()))
                        except ValueError: pass
                    if nums: rows.append(nums)
            if not rows: return
            arr = np.array([r[0]+1j*r[1] if len(r)>1 else r[0] for r in rows])
        arr=np.asarray(arr).squeeze()
        if arr.ndim == 1: self.add_signal(arr, path.name)
        elif arr.ndim == 2:
            self.current_image = np.abs(arr); self.current_kspace = fft2c(arr); self.current_recon = ifft2c(self.current_kspace); self.refresh_images()

    def add_signal(self, signal, name):
        signal=np.asarray(signal).squeeze()
        if signal.ndim != 1 or signal.size < 2: return
        self.signals.append({'name': name, 'data': signal})
        self.active_spike_indices = np.array([], dtype=int)
        self.signal_combo.addItem(name)
        self.signal_combo.setCurrentIndex(len(self.signals)-1)
        self.update_signal_plot(); self.update_tracker_preview(signal)

    def clear_signals(self):
        self.signals.clear(); self.signal_combo.clear(); self.signal_plot.clear(); self.tracker_plot.clear()

    def update_signal_plot(self):
        self.signal_plot.clear(); idx=self.signal_combo.currentIndex()
        if idx < 0 or idx >= len(self.signals): return
        data=np.asarray(self.signals[idx]['data']); component=self.signal_component.currentText()
        if self.signal_fft.isChecked(): data=np.fft.fftshift(np.fft.fft(data)); self.signal_plot.setLabel('bottom','Frequency bin')
        else: self.signal_plot.setLabel('bottom','Sample')
        y=self.component(data, component)
        self.signal_plot.plot(y)
        if self.active_spike_indices.size:
            valid = self.active_spike_indices[self.active_spike_indices < len(y)]
            if valid.size:
                scatter = pg.ScatterPlotItem(
                    x=valid,
                    y=np.asarray(y)[valid],
                    pen=pg.mkPen("#ff3b3b"),
                    brush=pg.mkBrush("#ff3b3b"),
                    size=9,
                    symbol="x",
                )
                self.signal_plot.addItem(scatter)

    def update_tracker_preview(self, signal):
        self.tracker_plot.clear(); self.tracker_plot.plot(np.abs(np.asarray(signal)))

    def _analyze_tracker_lines(self):
        self.tracker_line_metrics = []
        if self.tracker_matrix is None:
            return

        magnitude = np.abs(self.tracker_matrix).astype(np.float64)
        for line_index, line in enumerate(magnitude):
            self.tracker_line_metrics.append({
                "line": line_index,
                "peak": float(np.max(line)),
                "rms": float(np.sqrt(np.mean(line ** 2))),
                "energy": float(np.sum(line ** 2)),
                "mean": float(np.mean(line)),
                "variance": float(np.var(line)),
                "relative": 0.0,
            })

        rms_values = np.array([item["rms"] for item in self.tracker_line_metrics], dtype=float)
        low = float(np.min(rms_values))
        high = float(np.max(rms_values))
        span = max(high - low, np.finfo(float).eps)
        for item in self.tracker_line_metrics:
            item["relative"] = 100.0 * (item["rms"] - low) / span

    def _tracker_metric_key(self) -> str:
        return {
            "Peak": "peak",
            "RMS": "rms",
            "Energy": "energy",
            "Mean magnitude": "mean",
            "Variance": "variance",
        }.get(self.tracker_metric_combo.currentText(), "peak")

    def _refresh_tracker_ranking(self):
        if not hasattr(self, "tracker_table"):
            return
        if not self.tracker_line_metrics:
            self.tracker_table.setRowCount(0)
            return

        key = self._tracker_metric_key()
        ranked = sorted(self.tracker_line_metrics, key=lambda item: item[key], reverse=True)
        ranked = ranked[: self.tracker_top_count.value()]

        self.tracker_table.blockSignals(True)
        self.tracker_table.setRowCount(len(ranked))
        for row, item in enumerate(ranked):
            values = [
                row + 1,
                item["line"],
                f'{item["peak"]:.6g}',
                f'{item["rms"]:.6g}',
                f'{item["energy"]:.6g}',
                f'{item["mean"]:.6g}',
                f'{item["variance"]:.6g}',
                f'{item["relative"]:.1f}%',
            ]
            for col, value in enumerate(values):
                table_item = QTableWidgetItem(str(value))
                table_item.setData(Qt.UserRole, item["line"])
                self.tracker_table.setItem(row, col, table_item)
        self.tracker_table.resizeColumnsToContents()
        self.tracker_table.blockSignals(False)

        if ranked:
            self.tracker_table.selectRow(0)
            self._tracker_selection_changed()

    def _tracker_selection_changed(self):
        rows = sorted({item.row() for item in self.tracker_table.selectedItems()})
        selected = []
        for row in rows:
            item = self.tracker_table.item(row, 1)
            if item is not None:
                selected.append(int(item.data(Qt.UserRole)))
        self.tracker_selected_lines = selected[:8]
        self._plot_tracker_selected_lines()

    def _tracker_component(self, line: np.ndarray) -> np.ndarray:
        mode = self.tracker_component_combo.currentText()
        if mode == "Real":
            return np.real(line)
        if mode == "Imaginary":
            return np.imag(line)
        if mode == "Phase":
            return np.angle(line)
        return np.abs(line)

    def _plot_tracker_selected_lines(self):
        if not hasattr(self, "tracker_raw_plot"):
            return
        self.tracker_raw_plot.clear()
        if self.tracker_matrix is None or not self.tracker_selected_lines:
            self.tracker_metrics_label.setText("-")
            return

        summaries = []
        for line_index in self.tracker_selected_lines:
            raw_line = self.tracker_matrix[line_index]
            values = np.asarray(self._tracker_component(raw_line), dtype=float)
            self.tracker_raw_plot.plot(values)

            metric = next(
                (item for item in self.tracker_line_metrics if item["line"] == line_index),
                None,
            )
            if metric:
                summaries.append(
                    f'Line {line_index}: Peak {metric["peak"]:.6g}, '
                    f'RMS {metric["rms"]:.6g}, Energy {metric["energy"]:.6g}'
                )
        self.tracker_metrics_label.setText(" | ".join(summaries))

    def _select_next_tracker_line(self):
        row = self.tracker_table.currentRow()
        if 0 <= row < self.tracker_table.rowCount() - 1:
            self.tracker_table.clearSelection()
            self.tracker_table.selectRow(row + 1)

    def _select_previous_tracker_line(self):
        row = self.tracker_table.currentRow()
        if row > 0:
            self.tracker_table.clearSelection()
            self.tracker_table.selectRow(row - 1)

    def export_tracker_lines_csv(self):
        if not self.tracker_line_metrics:
            QMessageBox.information(self, "Tracker Export", "Load a Tracker file first.")
            return
        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Tracker Line Metrics", "tracker_line_metrics.csv", "CSV files (*.csv)"
        )
        if not filename:
            return
        if not filename.lower().endswith(".csv"):
            filename += ".csv"

        import csv
        with open(filename, "w", newline="", encoding="utf-8-sig") as stream:
            writer = csv.writer(stream)
            writer.writerow(["Line", "Peak", "RMS", "Energy", "MeanMagnitude", "Variance", "RelativeStrength"])
            for item in sorted(self.tracker_line_metrics, key=lambda value: value["line"]):
                writer.writerow([
                    item["line"], item["peak"], item["rms"], item["energy"],
                    item["mean"], item["variance"], item["relative"]
                ])
        self.statusBar().showMessage(f"Tracker metrics exported: {filename}")

    def _raw_candidate_paths(self) -> list[Path]:
        """RAW image/signal candidates; Tracking PFiles are excluded."""
        candidates = []
        seen = set()
        for path in self.imported_paths:
            if not path.exists() or not path.is_file():
                continue

            # Tracking PFile belongs only to Tracker analysis.
            if self.looks_tracker(path):
                continue

            suffix = path.suffix.lower()
            if suffix in {
                ".raw",
                ".bin",
                ".dat",
                ".pfile",
            }:
                key = str(path.resolve())
                if key not in seen:
                    candidates.append(path)
                    seen.add(key)
        return candidates


    def _robust_z(values: np.ndarray) -> np.ndarray:
        values = np.asarray(values, dtype=np.float64)
        if values.size == 0:
            return values
        median = np.nanmedian(values)
        mad = np.nanmedian(np.abs(values - median))
        scale = max(1.4826 * mad, np.finfo(float).eps)
        return np.abs(values - median) / scale

    @staticmethod
    def _group_spike_indices(indices: np.ndarray) -> list[np.ndarray]:
        indices = np.asarray(indices, dtype=int)
        if indices.size == 0:
            return []
        cuts = np.where(np.diff(indices) > 1)[0] + 1
        return [group for group in np.split(indices, cuts) if group.size]

    def _extract_tracker_signal(self, path: Path) -> tuple[np.ndarray, str]:
        raw = path.read_bytes()
        best = None
        for matrix in (512, 384, 320, 256, 192, 160, 128, 96, 64):
            need = matrix * matrix * 4
            if len(raw) < need:
                continue
            offset = len(raw) - need
            vals = np.frombuffer(raw, dtype="<i2", count=matrix * matrix * 2, offset=offset)
            signal = vals[0::2].astype(np.float64) + 1j * vals[1::2].astype(np.float64)
            image = ifft2c(signal.reshape(matrix, matrix))
            contrast = np.percentile(np.abs(image), 99) / (np.median(np.abs(image)) + 1e-9)
            if best is None or contrast > best[0]:
                best = (contrast, signal, f"Tracker complex int16 {matrix}x{matrix}, offset {offset}")
        if best is None:
            raise ValueError("No valid Tracker complex matrix candidate")
        return best[1], best[2]

    def _extract_generic_raw_signal(self, path: Path) -> tuple[np.ndarray, str]:
        # Limit analysis memory while retaining enough data for spike screening.
        max_bytes = 32 * 1024 * 1024
        size = path.stat().st_size
        offset = max(0, size - max_bytes)
        if offset % 2:
            offset += 1
        with open(path, "rb") as stream:
            stream.seek(offset)
            raw = np.fromfile(stream, dtype="<i2")
        if raw.size < 8:
            raise ValueError("Not enough int16 samples")
        # Treat even/odd pairs as complex when possible. Magnitude analysis is
        # resistant to phase rotation and works for many MRI raw layouts.
        if raw.size >= 16 and raw.size % 2 == 0:
            signal = raw[0::2].astype(np.float64) + 1j * raw[1::2].astype(np.float64)
            description = f"Generic complex int16 tail, byte offset {offset}"
        else:
            signal = raw.astype(np.float64)
            description = f"Generic real int16 tail, byte offset {offset}"
        return signal, description

    def _analyze_spike_signal(self, signal: np.ndarray, threshold: float) -> dict:
        signal = np.asarray(signal).squeeze()
        magnitude = np.abs(signal).astype(np.float64)
        finite = np.isfinite(magnitude)
        if magnitude.size < 8 or not np.any(finite):
            raise ValueError("Signal has insufficient valid samples")
        replacement = np.nanmedian(magnitude[finite])
        magnitude = np.where(finite, magnitude, replacement)

        amplitude_z = self._robust_z(magnitude)
        first_difference = np.abs(np.diff(magnitude, prepend=magnitude[0]))
        difference_z = self._robust_z(first_difference)

        combined = np.maximum(amplitude_z, difference_z)
        raw_indices = np.flatnonzero(combined >= threshold)
        groups = self._group_spike_indices(raw_indices)

        # Represent each consecutive group by its strongest sample.
        representatives = []
        for group in groups:
            representatives.append(int(group[np.argmax(combined[group])]))
        representatives = np.asarray(representatives, dtype=int)

        strongest = int(np.argmax(combined))
        score = float(combined[strongest])
        return {
            "score": score,
            "flagged": bool(representatives.size and score >= threshold),
            "indices": representatives,
            "groups": len(groups),
            "strongest": strongest,
            "samples": int(magnitude.size),
        }

    @staticmethod
    def _dicom_descendant_indices(item):
        """Return DICOM leaves below an Explorer node in visible tree order."""
        indices = []
        data = item.data(0, Qt.UserRole)
        if data and data[0] == "dicom":
            indices.append(int(data[1]))
        for child_index in range(item.childCount()):
            indices.extend(
                MainWindow._dicom_descendant_indices(item.child(child_index))
            )
        return indices

    def _selected_dicom_indices(self):
        """Resolve selected image/series items in Explorer display order.

        Selecting a Series includes every DICOM leaf below it. Explicit image or
        Series selection takes precedence over checked leaves. The currently
        displayed image is the final fallback.
        """
        ordered = []
        seen = set()

        def append_index(value):
            value = int(value)
            if value not in seen and 0 <= value < len(self.dicom_entries):
                seen.add(value)
                ordered.append(value)

        def selected_image_or_series_ancestor(item):
            node = item
            while node is not None:
                data = node.data(0, Qt.UserRole)
                if (
                    node.isSelected()
                    and data
                    and data[0] in ("dicom", "series")
                ):
                    return True
                node = node.parent()
            return False

        # QTreeWidgetItemIterator preserves the Explorer's top-to-bottom order,
        # which defines the first result displayed in Spike Diag.
        iterator = QTreeWidgetItemIterator(self.tree)
        while iterator.value():
            item = iterator.value()
            data = item.data(0, Qt.UserRole)
            if (
                data
                and data[0] == "dicom"
                and selected_image_or_series_ancestor(item)
            ):
                append_index(data[1])
            iterator += 1

        if not ordered:
            iterator = QTreeWidgetItemIterator(self.tree)
            while iterator.value():
                item = iterator.value()
                data = item.data(0, Qt.UserRole)
                if (
                    data
                    and data[0] == "dicom"
                    and item.checkState(0) == Qt.Checked
                ):
                    append_index(data[1])
                iterator += 1

        if not ordered and self.dicom_entries:
            append_index(getattr(self, "slice_index", 0))
        return ordered



    def _spike_range_percent(self):
        mapping = {"Large": 100, "Mid": 70, "Small": 40}
        if self.spike_range_combo.currentText() == "Scale":
            return int(self.spike_scale_slider.value())
        return mapping.get(self.spike_range_combo.currentText(), 100)

    @staticmethod
    def _box_blur_2d(array: np.ndarray, radius: int) -> np.ndarray:
        """Fast reflected-edge box blur used by the DoG-like blob detector."""
        src=np.asarray(array,dtype=float)
        radius=max(1,int(radius)); pad=radius
        padded=np.pad(src,((pad,pad),(pad,pad)),mode="reflect")
        integral=np.pad(padded,((1,0),(1,0)),mode="constant").cumsum(0).cumsum(1)
        size=2*radius+1
        total=(integral[size:,size:]-integral[:-size,size:]-integral[size:,:-size]+integral[:-size,:-size])
        return total/float(size*size)

    @classmethod
    def _derived_kspace_cluster_candidates(cls, magnitude: np.ndarray, valid: np.ndarray) -> list[dict]:
        """Detect compact local anomalies, not complete k-space lines.

        A two-scale DoG-like residual removes the smooth radial falloff and the
        normal DC cross. Connected components are retained only when compact,
        locally prominent, and not line-like across most of the matrix.
        """
        mag=np.asarray(magnitude,dtype=float); rows,cols=mag.shape; cy,cx=rows//2,cols//2
        logmag=np.log1p(mag)
        small=cls._box_blur_2d(logmag,max(1,int(min(rows,cols)*0.006)))
        large=cls._box_blur_2d(logmag,max(3,int(min(rows,cols)*0.035)))
        dog=small-large
        vals=dog[valid]; med=float(np.median(vals)) if vals.size else 0.0
        mad=float(np.median(np.abs(vals-med))) if vals.size else 1.0
        scale=max(1.4826*mad,1e-9); z=(dog-med)/scale
        hot=valid & (z>=4.5)
        yy,xx=np.indices((rows,cols)); cw=max(2,int(min(rows,cols)*0.012))
        hot &= ~(((np.abs(yy-cy)<=cw)|(np.abs(xx-cx)<=cw)) & (np.hypot(yy-cy,xx-cx)<min(rows,cols)*0.20))
        seen=np.zeros_like(hot,bool); out=[]
        for sy,sx in zip(*np.where(hot)):
            if seen[sy,sx]: continue
            stack=[(int(sy),int(sx))]; seen[sy,sx]=True; pts=[]
            while stack:
                y,x=stack.pop(); pts.append((y,x))
                for dy in (-1,0,1):
                    for dx in (-1,0,1):
                        if not (dy or dx): continue
                        ny,nx=y+dy,x+dx
                        if 0<=ny<rows and 0<=nx<cols and hot[ny,nx] and not seen[ny,nx]:
                            seen[ny,nx]=True; stack.append((ny,nx))
            if len(pts)<2: continue
            ys=np.array([q[0] for q in pts]); xs=np.array([q[1] for q in pts])
            h=int(ys.max()-ys.min()+1); w=int(xs.max()-xs.min()+1)
            if h>rows*0.35 or w>cols*0.35: continue
            fill=len(pts)/max(h*w,1); elong=max(h,w)/max(min(h,w),1)
            if elong>10 and fill<0.45: continue
            mask=np.zeros_like(hot); mask[ys,xs]=True
            # small dilation captures the whole local island, not a full row/column
            dil=mask.copy()
            for dy in (-1,0,1):
                for dx in (-1,0,1):
                    dil |= np.roll(np.roll(mask,dy,0),dx,1)
            cys=(2*cy-ys)%rows; cxs=(2*cx-xs)%cols; dil[cys,cxs]=True
            edge=min(xs.min(),cols-1-xs.max(),ys.min(),rows-1-ys.max())
            out.append({"type":"cluster","mask":dil,"z":float(np.max(z[ys,xs])),
                        "points":[(int(np.mean(ys)),int(np.mean(xs)))],
                        "bbox":(int(ys.min()),int(ys.max()),int(xs.min()),int(xs.max())),
                        "cluster_size":int(len(pts)),"edge_distance":int(edge),
                        "compactness":float(fill),"elongation":float(elong)})
        return sorted(out,key=lambda d:(d["z"],d["cluster_size"]),reverse=True)[:32]

    def _suppress_image_spikes(self, image: np.ndarray):
        """Detect spike noise using k-space physics and image-space validation.

        Candidates are not limited to row/column intersections.  Isolated
        points, complete/partial rows, columns, adjacent bands and oblique
        k-space lines are each converted independently with inverse FFT.  A
        candidate is accepted only when its predicted spatial wave has enough
        energy and agrees with structure present in the original image.
        """
        source = np.asarray(image, dtype=float)
        kspace = fft2c(source)
        magnitude = np.abs(kspace)
        rows, cols = magnitude.shape
        cy, cx = rows // 2, cols // 2

        area_fraction = max(min(self._spike_range_percent() / 100.0, 1.0), 0.10)
        side_fraction = np.sqrt(area_fraction)
        half_h = max(4, int(rows * side_fraction / 2.0))
        half_w = max(4, int(cols * side_fraction / 2.0))
        y0, y1 = max(1, cy-half_h), min(rows-1, cy+half_h)
        x0, x1 = max(1, cx-half_w), min(cols-1, cx+half_w)
        valid = np.zeros_like(magnitude, dtype=bool)
        valid[y0:y1, x0:x1] = True
        yy, xx = np.ogrid[:rows, :cols]
        center_radius = max(3, int(min(rows, cols) * 0.018))
        valid &= ((yy-cy)**2 + (xx-cx)**2) > center_radius**2

        raw = self._raw_spike_features(magnitude, valid)
        logmag = np.log1p(magnitude)

        # Remove slow anatomical intensity variation before correlation.  The
        # residual retains periodic stripe/wave components caused by spikes.
        smooth = median_filter_3x3_numpy(median_filter_3x3_numpy(source))
        residual = source - smooth
        residual_std = float(np.std(residual)) + 1e-9
        source_std = float(np.std(source)) + 1e-9

        level = self.spike_level_combo.currentText()
        point_z_limit = {"Wide": 8.0, "Mid": 6.5, "Fine": 5.0}.get(level, 6.0)
        line_z_limit = 999.0  # retained in report for compatibility; line detector disabled
        score_limit = {"Wide": 0.40, "Mid": 0.32, "Fine": 0.26}.get(level, 0.32)

        proposals = []

        # Localized DICOM-derived k-space islands are primary candidates.
        # They remain candidates even near the outer edges and even when a
        # conjugate partner is visible.
        proposals.extend(self._derived_kspace_cluster_candidates(magnitude, valid))

        # Isolated point candidates.  Each point is paired with its conjugate,
        # because together they create the real sinusoidal wave in image space.
        coords = raw["candidate_coords"]
        if coords.size:
            scores = np.array([raw["local_z"][y, x] + 0.35*raw["global_z"][y, x] for y, x in coords])
            for idx in np.argsort(scores)[::-1][:20]:
                y, x = map(int, coords[idx])
                if scores[idx] < point_z_limit:
                    break
                mask = np.zeros((rows, cols), dtype=bool)
                mask[y, x] = True
                sy, sx = (2*cy-y) % rows, (2*cx-x) % cols
                mask[sy, sx] = True
                proposals.append({"type":"point", "mask":mask, "z":float(scores[idx]), "points":[(y,x),(int(sy),int(sx))]})

        # Commit0076: complete row/column and oblique line proposals are intentionally
        # excluded. They represent normal derived-k-space structure too often.

        accepted = []
        reviewed = []
        accepted_mask = np.zeros((rows, cols), dtype=bool)
        for proposal_index, proposal in enumerate(proposals, start=1):
            mask = proposal["mask"]
            component_k = np.where(mask, kspace, 0)
            wave_complex = ifft2c(component_k)
            wave = np.real(wave_complex)
            wave_std = float(np.std(wave))
            energy_ratio = wave_std / source_std
            corr = 0.0
            projection_strength = 0.0
            period = float("inf")
            wave_angle = 0.0
            kx = ky = 0.0
            score = 0.0
            reasons = []
            if wave_std > 1e-12:
                corr = abs(float(np.mean((wave-wave.mean())*(residual-residual.mean()))) / (wave_std*residual_std))
                weights = magnitude * mask
                total = float(weights.sum()) + 1e-9
                ky = float(((ygrid-cy)*weights).sum()/total)
                kx = float(((xgrid-cx)*weights).sum()/total)
                radial = float(np.hypot(kx, ky))
                period = float(min(rows, cols)/radial) if radial > 1e-6 else float('inf')
                wave_angle = float((np.degrees(np.arctan2(ky, kx)) + 90.0) % 180.0)
                theta = np.deg2rad(wave_angle - 90.0)
                coordinate = xgrid*np.cos(theta) + ygrid*np.sin(theta)
                projection_bins = np.rint(coordinate - coordinate.min()).astype(int)
                bin_count = int(projection_bins.max()) + 1
                projection = np.bincount(
                    projection_bins.ravel(), weights=residual.ravel(), minlength=bin_count
                ) / np.maximum(np.bincount(projection_bins.ravel(), minlength=bin_count), 1)
                projection = projection - np.median(projection)
                projection_strength = float(np.std(projection)) / residual_std
                z_norm = min(1.0, proposal["z"] / 10.0)
                energy_norm = min(1.0, energy_ratio / 0.055)
                correlation_norm = min(1.0, corr / 0.22)
                projection_norm = min(1.0, projection_strength / 0.42)
                score = 0.34*z_norm + 0.26*energy_norm + 0.18*correlation_norm + 0.22*projection_norm

            minimum_energy = 0.0008 if proposal["type"] == "cluster" else 0.0007
            normal_structure = False
            compact = float(proposal.get("compactness",1.0))
            elong = float(proposal.get("elongation",1.0))
            cluster_pass = (proposal["type"] == "cluster" and proposal.get("cluster_size",0) >= 2
                            and proposal.get("z",0) >= 4.5 and compact >= 0.08 and elong <= 10.0
                            and energy_ratio >= minimum_energy
                            and (corr >= 0.015 or projection_strength >= 0.06 or score >= score_limit*0.62))
            point_pass = (proposal["type"] == "point" and proposal.get("z",0) >= point_z_limit
                          and energy_ratio >= minimum_energy and score >= score_limit*0.72)
            is_accepted = bool(wave_std > 1e-12 and (cluster_pass or point_pass))
            if wave_std <= 1e-12: reasons.append("candidate-only IFFT has no measurable energy")
            if energy_ratio < minimum_energy: reasons.append(f"energy {energy_ratio:.5f} below {minimum_energy:.5f}")
            if normal_structure: reasons.append("rejected: full-span derived-k-space anatomical structure")
            if not cluster_pass and not point_pass: reasons.append("localized anomaly did not pass image-space validation")
            if is_accepted:
                reasons.append("accepted: physical wave evidence passed")
            proposal.update({
                "id": proposal_index, "score":score, "correlation":corr,
                "energy_ratio":energy_ratio, "predicted_period":period,
                "predicted_wave_angle":wave_angle, "projection_strength":projection_strength,
                "centroid_k":(kx,ky), "decision":"PASS" if is_accepted else "REJECT",
                "reason":"; ".join(reasons), "component_wave":wave,
            })
            reviewed.append(proposal)
            if is_accepted:
                accepted.append(proposal)
                accepted_mask |= mask

        # Merge accepted components and repair only those samples.  Axis-aligned
        # bands are interpolated from adjacent lines; all other shapes use a
        # local complex median.  This avoids the old behaviour of marking the
        # normal DC cross at regularly spaced points.
        corrected = kspace.copy()
        for proposal in accepted:
            ptype = proposal["type"]
            if True:
                ys, xs = np.where(proposal["mask"])
                for y, x in zip(ys, xs):
                    ya,yb=max(0,y-2),min(rows,y+3); xa,xb=max(0,x-2),min(cols,x+3)
                    patch=corrected[ya:yb,xa:xb]
                    local_mask=~proposal["mask"][ya:yb,xa:xb]
                    vals=patch[local_mask]
                    if vals.size:
                        corrected[y,x]=np.median(np.real(vals))+1j*np.median(np.imag(vals))

        result = np.abs(ifft2c(corrected)) if accepted else source.copy()
        candidate_points=[]
        for proposal in accepted:
            candidate_points.extend(proposal.get("points", []))
        candidate_points=list(dict.fromkeys((int(y),int(x)) for y,x in candidate_points))
        analysis={"regions":accepted, "reviewed_regions":reviewed, "accepted_mask":accepted_mask,
                  "proposal_count":len(proposals), "accepted_count":len(accepted),
                  "thresholds":{"point_z":point_z_limit,"line_z":line_z_limit,"score":score_limit}}
        return result, kspace, corrected, candidate_points, analysis


    def _activate_spike_diag_from_workspace_selection(self):
        """Auto-process the Image Workspace selection when Spike Diag opens."""
        if self.tabs.currentIndex() != 1:
            return
        indices = self._selected_dicom_indices()
        if not indices:
            if self.current_image is None:
                self.spike_image_results = []
                self._refresh_spike_diag(select_first=True)
                self.statusBar().showMessage("Spike Diag: select an image in Image Workspace")
                return
            source = np.asarray(self.current_image).copy()
            stripe_info = self._extract_stripe_only(source)
            actual_raw = np.asarray(self.current_kspace).copy() if self.current_kspace is not None else fft2c(source.astype(float))
            raw_mapping = self._map_stripe_to_raw(stripe_info, actual_raw)
            raw_comp_analysis = self._raw_compensation_diagnostic(actual_raw)
            corrected, raw_before, raw_after, candidates, spike_analysis = self._suppress_image_spikes(source)
            self.spike_image_results = [{
                "index": None, "path": Path(self.current_source or "Workspace_Current_Image"),
                "output": None, "original": source, "corrected": corrected,
                "raw_before": raw_before, "raw_after": raw_after, "candidates": candidates,
                "stripe_only": stripe_info["stripe_only"], "raw_candidate": raw_mapping["candidate_image"],
                "stripe_groups": stripe_info["groups"], "raw_agreement": float(raw_mapping["agreement"]),
                "matched_raw_count": int(len(raw_mapping["matched_coords"])),
                "detected": bool(spike_analysis.get("accepted_count") or raw_comp_analysis.get("accepted_count")),
                "spike_analysis": spike_analysis, "raw_comp_analysis": raw_comp_analysis,
            }]
            self._refresh_spike_diag(select_first=True)
            self.statusBar().showMessage("Spike Diag loaded the current Image Workspace image")
            return

        progress = self._make_progress("Spike Detection", len(indices))
        results = []
        try:
            for number, index in enumerate(indices, start=1):
                if progress.wasCanceled():
                    return
                entry = self._ensure_dicom_image(index)
                progress.setLabelText(
                    f"Analyzing spike noise {number}/{len(indices)}: {entry.path.name}"
                )
                QApplication.processEvents()

                stripe_info = self._extract_stripe_only(entry.image)
                actual_raw = fft2c(np.asarray(entry.image, dtype=float))
                raw_mapping = self._map_stripe_to_raw(stripe_info, actual_raw)
                raw_comp_analysis = self._raw_compensation_diagnostic(actual_raw)
                corrected, raw_before, raw_after, candidates, spike_analysis = (
                    self._suppress_image_spikes(entry.image)
                )
                detected = bool(spike_analysis.get("accepted_count") or raw_comp_analysis.get("accepted_count"))
                results.append({
                    "index": index,
                    "path": entry.path,
                    "output": None,
                    "original": np.asarray(entry.image).copy(),
                    "corrected": corrected,
                    "raw_before": raw_before,
                    "raw_after": raw_after,
                    "candidates": candidates,
                    "stripe_only": stripe_info["stripe_only"],
                    "raw_candidate": raw_mapping["candidate_image"],
                    "stripe_groups": stripe_info["groups"],
                    "raw_agreement": float(raw_mapping["agreement"]),
                    "matched_raw_count": int(len(raw_mapping["matched_coords"])),
                    "detected": detected,
                    "spike_analysis": spike_analysis,
                    "raw_comp_analysis": raw_comp_analysis,
                })
                progress.setValue(number)
        finally:
            progress.close()

        self.spike_image_results = results
        self._refresh_spike_diag(select_first=True)
        detected_count = sum(bool(item.get("detected")) for item in results)
        self.statusBar().showMessage(
            f"Spike Diag processed {len(results)} image(s): "
            f"{detected_count} spike candidate(s)"
        )

    def apply_spike_processing(self):
        dicom_indices = self._selected_dicom_indices()
        if not dicom_indices:
            # No original image selection: use existing RAW/PFile diagnostic.
            self.detect_spikes()
            return

        progress = self._make_progress("Spike Processing", len(dicom_indices))
        processed = []
        try:
            for number, index in enumerate(dicom_indices, start=1):
                if progress.wasCanceled():
                    return
                entry = self._ensure_dicom_image(index)
                progress.setLabelText(
                    f"Detecting repeated stripes and k-space spikes {number}/{len(dicom_indices)}: {entry.path.name}"
                )
                QApplication.processEvents()

                stripe_info = self._extract_stripe_only(entry.image)
                actual_raw = fft2c(np.asarray(entry.image, dtype=float))
                raw_mapping = self._map_stripe_to_raw(stripe_info, actual_raw)
                corrected, raw_before, raw_after, candidates, spike_analysis = self._suppress_image_spikes(entry.image)
                if spike_analysis.get("accepted_count"):

                    self.current_ds = entry.ds
                    self.current_source = str(entry.path)
                    output = self._save_processed_image(
                        corrected,
                        f"{entry.path.stem}_NS",
                        "_NS",
                    )
                    processed.append({
                        "index": index,
                        "path": entry.path,
                        "output": output,
                        "original": np.asarray(entry.image).copy(),
                        "corrected": corrected,
                        "raw_before": raw_before,
                        "raw_after": raw_after,
                        "candidates": candidates,
                        "stripe_only": stripe_info["stripe_only"],
                        "raw_candidate": raw_mapping["candidate_image"],
                        "stripe_groups": stripe_info["groups"],
                        "raw_agreement": float(raw_mapping["agreement"]),
                        "matched_raw_count": int(len(raw_mapping["matched_coords"])),
                        "spike_analysis": spike_analysis,
                        "detected": bool(spike_analysis.get("accepted_count")),
                    })
                progress.setValue(number)
        finally:
            progress.close()

        if not processed:
            QMessageBox.information(self, "Spike Processing", "No Spike")
            return

        self.spike_image_results = processed
        last = processed[-1]
        self._display_processed_result(last["corrected"], last["output"])
        self._refresh_spike_diag(select_first=False)
        self.tabs.setCurrentIndex(1)
        self._schedule_responsive_layout()
        QMessageBox.information(
            self,
            "Spike Processing",
            f"Processed {len(dicom_indices)} image(s).\n"
            f"Spike corrected: {len(processed)} image(s).",
        )

    def _refresh_spike_diag(self, select_first: bool = False):
        self.spike_result_list.clear()
        results = getattr(self, "spike_image_results", [])
        detected_count = 0
        for index, result in enumerate(results):
            detected = bool(result.get("detected", result.get("candidates")))
            detected_count += int(detected)
            status = "Spike" if detected else "No Spike"
            item = QTreeWidgetItem([f'{result["path"].name}  [{status}]'])
            item.setData(0, Qt.UserRole, index)
            item.setToolTip(0, str(result["path"]))
            self.spike_result_list.addTopLevelItem(item)
        self.spike_result_summary.setText(
            f"Processed: {len(results)} | Spike detected: {detected_count}"
        )
        if results:
            index = 0 if select_first else len(results) - 1
            item = self.spike_result_list.topLevelItem(index)
            self.spike_result_list.setCurrentItem(item)
            self._show_spike_result(index)

    def _spike_result_selected(self, item, column):
        index = item.data(0, Qt.UserRole)
        if index is not None:
            self._show_spike_result(int(index))

    def _show_spike_result(self, index: int):
        results = getattr(self, "spike_image_results", [])
        if not (0 <= index < len(results)):
            return
        result = results[index]
        self.current_spike_result_index = index
        original = np.asarray(result["original"], dtype=float)
        raw_before = np.asarray(result["raw_before"])
        raw_after = np.asarray(result["raw_after"])
        analysis = result.get("spike_analysis", {})
        reviewed = analysis.get("reviewed_regions", analysis.get("regions", []))

        self.spike_input_note.setText(
            "DICOM input: acquired scanner RAW samples are not present in this file. "
            "The detector intentionally supports DICOM-derived k-space because native scanner RAW is usually unavailable. "
            "1B is the exact spatial-domain source and STEP 2 is its derived FFT. Candidates are accepted only when their inverse-FFT contribution explains periodic image artifact and improves a compensation preview."
        )
        self.spike_original_panel.set_image(original)
        self.spike_fft_input_panel.set_image(original)
        mag = np.abs(raw_before)
        self.spike_kspace_linear_panel.set_image(mag)
        self.spike_raw_before_panel.set_image(np.log1p(mag))
        logmag = np.log1p(mag)
        enhanced = self._robust_zscore(logmag)
        self.spike_kspace_enhanced_panel.set_image(np.clip(enhanced, -2, 12))
        self.spike_candidate_overlay_panel.set_image(np.log1p(mag))
        self._draw_spike_proposal_overlays(self.spike_candidate_overlay_panel, reviewed)
        self.spike_corrected_panel.set_image(result["corrected"])
        self.spike_difference_panel.set_image(np.abs(original - np.asarray(result["corrected"], dtype=float)))
        self.spike_raw_after_panel.set_image(np.log1p(np.abs(raw_after)))
        smooth = median_filter_3x3_numpy(median_filter_3x3_numpy(original))
        self._current_spike_residual = original - smooth
        self.spike_residual_panel.set_image(self._current_spike_residual)

        table = self.spike_candidate_table
        table.setRowCount(len(reviewed))
        for row, region in enumerate(reviewed):
            values = [
                str(region.get("id", row+1)), region.get("type", ""), f'{region.get("z",0):.2f}',
                f'{region.get("energy_ratio",0):.5f}', f'{region.get("predicted_wave_angle",0):.1f}',
                "∞" if not np.isfinite(region.get("predicted_period", float("inf"))) else f'{region.get("predicted_period",0):.2f}',
                f'{region.get("correlation",0):.3f}', region.get("decision", "")
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setData(Qt.UserRole, row)
                table.setItem(row, col, item)
        table.resizeColumnsToContents()
        if reviewed:
            table.selectRow(0)
            self._show_spike_candidate(0)
        else:
            self.spike_candidate_mask_panel.set_image(np.zeros_like(original))
            self.spike_candidate_kspace_panel.set_image(np.zeros_like(mag))
            self.spike_candidate_ifft_panel.set_image(np.zeros_like(original))
            self.spike_predicted_wave_panel.set_image(np.zeros_like(original))
            self.spike_match_panel.set_image(np.zeros_like(original))
            self.spike_candidate_detail.setPlainText("No candidate proposals were generated. Review STEP 2 and candidate thresholds.")
            self.spike_decision_text.setPlainText("NO PROPOSALS\nThe extraction stage did not produce any point, line, band, or oblique candidate.")

    def _draw_spike_proposal_overlays(self, panel, regions):
        rows = panel.image_item.image.shape[0] if panel.image_item.image is not None else 0
        cols = panel.image_item.image.shape[1] if panel.image_item.image is not None else 0
        for region in regions:
            passed = region.get("decision") == "PASS"
            pen = pg.mkPen("g" if passed else "r", width=2)
            kind = region.get("type")
            if kind in ("point", "cluster"):
                for y, x in region.get("points", [])[:2]:
                    panel.plot.addItem(pg.ScatterPlotItem([x],[y],symbol="o",size=8,pen=pen,brush=None))
            elif kind in ("row","col","band") and "range" in region:
                lo, hi = region["range"]
                if region.get("axis") == "row":
                    for y in (lo,hi): panel.plot.addItem(pg.PlotDataItem([0,cols-1],[y,y],pen=pen))
                else:
                    for x in (lo,hi): panel.plot.addItem(pg.PlotDataItem([x,x],[0,rows-1],pen=pen))
            elif kind == "oblique":
                angle=np.deg2rad(region.get("angle",0.0)); d=region.get("distance",0.0); points=[]
                for x in (-cols/2, cols/2):
                    if abs(np.sin(angle))>1e-6:
                        y=(d-x*np.cos(angle))/np.sin(angle)
                        if -rows/2<=y<=rows/2: points.append((x+cols/2,y+rows/2))
                for y in (-rows/2, rows/2):
                    if abs(np.cos(angle))>1e-6:
                        x=(d-y*np.sin(angle))/np.cos(angle)
                        if -cols/2<=x<=cols/2: points.append((x+cols/2,y+rows/2))
                if len(points)>=2: panel.plot.addItem(pg.PlotDataItem([points[0][0],points[1][0]],[points[0][1],points[1][1]],pen=pen))

    def _spike_candidate_selected(self, row, column):
        self._show_spike_candidate(int(row))

    def _show_spike_candidate(self, candidate_index: int):
        results = getattr(self, "spike_image_results", [])
        result_index = getattr(self, "current_spike_result_index", -1)
        if not (0 <= result_index < len(results)):
            return
        result = results[result_index]
        reviewed = result.get("spike_analysis", {}).get("reviewed_regions", result.get("spike_analysis", {}).get("regions", []))
        if not (0 <= candidate_index < len(reviewed)):
            return
        region = reviewed[candidate_index]
        raw = np.asarray(result["raw_before"])
        mask = np.asarray(region.get("mask"), dtype=bool)
        component = np.where(mask, raw, 0)
        wave = np.asarray(region.get("component_wave", np.real(ifft2c(component))), dtype=float)
        residual = getattr(self, "_current_spike_residual", np.zeros_like(wave))
        self.spike_candidate_mask_panel.set_image(mask.astype(float))
        self.spike_candidate_kspace_panel.set_image(np.log1p(np.abs(component)))
        self.spike_candidate_ifft_panel.set_image(wave)
        self.spike_predicted_wave_panel.set_image(wave)
        denom = (np.std(wave)*np.std(residual)) + 1e-9
        match = ((wave-wave.mean())*(residual-residual.mean())) / denom
        self.spike_match_panel.set_image(match)
        thresholds = result.get("spike_analysis", {}).get("thresholds", {})
        detail = [
            f'Candidate ID: {region.get("id", candidate_index+1)}',
            f'Type: {region.get("type")}',
            f'Peak robust Z: {region.get("z",0):.3f}',
            f'Candidate-only IFFT energy ratio: {region.get("energy_ratio",0):.6f}',
            f'Predicted wave angle: {region.get("predicted_wave_angle",0):.2f}°',
            f'Predicted period: {region.get("predicted_period",float("inf")):.3f} px',
            f'Whole-image correlation: {region.get("correlation",0):.4f}',
            f'Directional projection strength: {region.get("projection_strength",0):.4f}',
            f'Combined score: {region.get("score",0):.4f}',
            f'Decision: {region.get("decision")}',
            f'Reason: {region.get("reason","")}',
        ]
        self.spike_candidate_detail.setPlainText("\n".join(detail))
        self.spike_decision_text.setPlainText(
            f'{region.get("decision")} — Candidate {region.get("id", candidate_index+1)}\n\n'
            f'{region.get("reason", "No reason recorded")}\n\n'
            f'Active thresholds\nPoint Z: {thresholds.get("point_z",0):.2f}\n'
            f'Line Z: {thresholds.get("line_z",0):.2f}\nScore: {thresholds.get("score",0):.3f}\n\n'
            'Review order: STEP 2 source anomaly → STEP 3 extraction → STEP 4 candidate-only IFFT → '
            'STEP 5 spatial match → this decision → STEP 7 correction.'
        )


    def save_current_spike_result(self):
        results = getattr(self, "spike_image_results", [])
        index = getattr(self, "current_spike_result_index", -1)
        if not (0 <= index < len(results)):
            return
        result = results[index]
        self.current_ds = self.dicom_entries[result["index"]].ds
        self.current_source = str(result["path"])
        output = self._save_processed_image(
            result["corrected"], f"{result['path'].stem}_NS", "_NS"
        )
        QMessageBox.information(self, "Spike Diag", f"Saved:\n{output}")

    def detect_spikes(self):
        candidates = self._raw_candidate_paths()
        if not candidates:
            QMessageBox.information(
                self,
                "Spike Detection",
                "No supported RAW, BIN, DAT, or non-tracking PFile is imported. "
                "Drop a file or folder first.",
            )
            return

        threshold = float(self.spike_threshold.value())
        progress = self._make_progress("Detecting Spike Noise", len(candidates))
        results = []
        try:
            for index, path in enumerate(candidates, start=1):
                if not self._progress_step(
                    progress,
                    index - 1,
                    f"Analyzing {index}/{len(candidates)}: {path.name}",
                ):
                    self.statusBar().showMessage("Spike detection canceled")
                    return
                try:
                    if self.looks_tracker(path):
                        signal, source_type = self._extract_tracker_signal(path)
                    else:
                        signal, source_type = self._extract_generic_raw_signal(path)
                    analysis = self._analyze_spike_signal(signal, threshold)
                    analysis.update({
                        "path": path,
                        "name": path.name,
                        "type": source_type,
                        "signal": signal,
                        "error": "",
                    })
                except Exception as exc:
                    analysis = {
                        "path": path,
                        "name": path.name,
                        "type": "Unreadable",
                        "signal": np.array([], dtype=float),
                        "score": 0.0,
                        "flagged": False,
                        "indices": np.array([], dtype=int),
                        "groups": 0,
                        "strongest": -1,
                        "samples": 0,
                        "error": str(exc),
                    }
                results.append(analysis)
                progress.setValue(index)
        finally:
            progress.close()

        self.spike_results = sorted(results, key=lambda item: item["score"], reverse=True)
        self._refresh_spike_table()
        flagged = sum(1 for item in results if item["flagged"])
        errors = sum(1 for item in results if item["error"])
        self.spike_summary.setText(
            f"Scanned {len(results)} raw file(s) at robust threshold {threshold:.1f}. "
            f"Spike candidates: {flagged}. Read errors: {errors}."
        )
        self.tabs.setCurrentIndex(1)
        self.statusBar().showMessage(f"Spike detection complete: {flagged}/{len(results)} flagged")

    def _refresh_spike_table(self):
        if not hasattr(self, "spike_table"):
            return
        show_flagged = self.spike_only_flagged.isChecked()
        rows = [
            result for result in self.spike_results
            if (result["flagged"] or not show_flagged)
        ]
        self.spike_table.setSortingEnabled(False)
        self.spike_table.setRowCount(len(rows))
        for row, result in enumerate(rows):
            status = "SPIKE" if result["flagged"] else ("ERROR" if result["error"] else "OK")
            values = [
                status,
                result["name"],
                result["type"],
                f'{result["score"]:.2f}',
                str(result["groups"]),
                str(result["strongest"]) if result["strongest"] >= 0 else "-",
                str(result["samples"]),
                str(result["path"]),
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setData(Qt.UserRole, result)
                if status == "SPIKE":
                    item.setBackground(pg.mkColor("#7a2430"))
                elif status == "ERROR":
                    item.setBackground(pg.mkColor("#6a5520"))
                self.spike_table.setItem(row, col, item)
        self.spike_table.setSortingEnabled(True)
        self.spike_table.resizeColumnsToContents()

    def _open_spike_result(self, row: int, column: int):
        item = self.spike_table.item(row, 0)
        if item is None:
            return
        result = item.data(Qt.UserRole)
        if not result or result["signal"].size < 2:
            if result and result["error"]:
                QMessageBox.warning(self, "Spike Result", result["error"])
            return

        name = f'Spike scan: {result["name"]}'
        self.signals.append({"name": name, "data": result["signal"]})
        self.signal_combo.addItem(name)
        self.signal_combo.setCurrentIndex(len(self.signals) - 1)
        self.active_spike_indices = np.asarray(result["indices"], dtype=int)
        self.signal_fft.setChecked(False)
        self.signal_component.setCurrentText("Magnitude")
        self.update_signal_plot()
        self.update_tracker_preview(result["signal"])
        self.tabs.setCurrentIndex(1)
        self.statusBar().showMessage(
            f'Opened {result["name"]}: {len(self.active_spike_indices)} spike group(s)'
        )

    @staticmethod
    def _position_plane_index(name: str) -> int:
        return {"Axial": 0, "Coronal": 1, "Sagittal": 2}.get(name, 0)

    def populate_position_top_lines(self):
        if self.tracker_matrix is None or not self.tracker_line_metrics:
            QMessageBox.information(self, "Tracker Position", "Load a Tracker PFile / TrackerImg first.")
            return

        ranked = sorted(self.tracker_line_metrics, key=lambda item: item["rms"], reverse=True)[:3]
        for row, metric in enumerate(ranked):
            self.position_table.setItem(row, 0, QTableWidgetItem(str(metric["line"])))
        self.tabs.setCurrentIndex(4)
        self.statusBar().showMessage("Top three Tracker lines assigned to position conversion.")

    def calculate_tracker_position(self):
        if self.tracker_matrix is None:
            QMessageBox.information(self, "Tracker Position", "Load a Tracker file first.")
            return

        fft_length = int(self.position_fft_length.value())
        fov = float(self.position_fov.value())
        measurements = []

        try:
            for row in range(self.position_table.rowCount()):
                use_widget = self.position_table.cellWidget(row, 6)
                if use_widget is not None and not use_widget.isChecked():
                    continue

                line_item = self.position_table.item(row, 0)
                if line_item is None:
                    continue
                line_index = int(line_item.text())
                if line_index < 0 or line_index >= self.tracker_matrix.shape[0]:
                    raise ValueError(f"Line {line_index} is outside the available range.")

                plane_widget = self.position_table.cellWidget(row, 1)
                rotation_widget = self.position_table.cellWidget(row, 2)
                plane = self._position_plane_index(plane_widget.currentText())
                rotation_deg = float(rotation_widget.value())

                peak_bin, coordinate_mm, snr_like, spectrum = signal_peak_coordinate(
                    self.tracker_matrix[line_index],
                    fft_length=fft_length,
                    fov_mm=fov,
                )
                measurement = DirectionMeasurement(
                    line_index=line_index,
                    plane=plane,
                    rotation_deg=rotation_deg,
                    peak_bin=peak_bin,
                    coordinate_mm=coordinate_mm,
                    snr_like=snr_like,
                )
                measurements.append(measurement)

                self.position_table.setItem(row, 3, QTableWidgetItem(f"{peak_bin:.4f}"))
                self.position_table.setItem(row, 4, QTableWidgetItem(f"{coordinate_mm:.4f}"))
                self.position_table.setItem(row, 5, QTableWidgetItem(f"{snr_like:.3f}"))

            offset = np.array([
                self.position_offset_x.value(),
                self.position_offset_y.value(),
                self.position_offset_z.value(),
            ], dtype=float)

            scanner_xyz, pqr, rms_error = solve_position(
                measurements,
                center_offset=offset,
                oppose_ap=self.position_oppose_ap.isChecked(),
            )
            self.tracker_position_measurements = measurements
            self.position_result.setText(
                f"Coordinates MR / Scanner (mm):\n"
                f"  X = {scanner_xyz[0]:.4f}\n"
                f"  Y = {scanner_xyz[1]:.4f}\n"
                f"  Z = {scanner_xyz[2]:.4f}\n\n"
                f"Coordinates PQR (mm):\n"
                f"  P = {pqr[0]:.4f}\n"
                f"  Q = {pqr[1]:.4f}\n"
                f"  R = {pqr[2]:.4f}\n\n"
                f"Projection RMS residual = {rms_error:.5f} mm\n"
                f"Gradient-map correction = Not applied"
            )
            self.statusBar().showMessage("Tracker position conversion completed.")
        except Exception as exc:
            QMessageBox.critical(self, "Tracker Position Error", str(exc))

    def component(self, arr, mode):
        if mode == 'Real': return np.real(arr)
        if mode == 'Imaginary': return np.imag(arr)
        if mode == 'Phase': return np.angle(arr)
        return np.abs(arr)

    def change_workspace_source(self, source_name: str):
        if source_name == "Tracker Raw Magnitude":
            if self.tracker_matrix is None:
                return
            self.current_kspace = np.asarray(self.tracker_matrix).copy()
            self.current_recon = ifft2c(self.current_kspace)
            self.current_image = np.abs(self.current_recon)
            self.current_source = str(self.tracker_source_path or "Tracker")
        elif source_name == "Tracker Reconstructed":
            if self.tracker_reconstructed_image is None:
                return
            self.current_image = np.asarray(self.tracker_reconstructed_image).copy()
            self.current_kspace = fft2c(self.current_image)
            self.current_recon = ifft2c(self.current_kspace)
            self.current_source = str(self.tracker_source_path or "Tracker")
        self.refresh_images()

    def send_tracker_to_signal_studio(self):
        if self.tracker_strongest_line is None:
            QMessageBox.information(self, "Tracker", "Load a Tracker file first.")
            return
        name = f"{self.tracker_shared_name} | Strongest line {self.tracker_strongest_line_index}"
        existing = next((i for i, item in enumerate(self.signals) if item.get("name") == name), None)
        if existing is None:
            self.add_signal(np.asarray(self.tracker_strongest_line).copy(), name)
            existing = len(self.signals) - 1
        self.signal_combo.setCurrentIndex(existing)
        self.signal_component.setCurrentText("Magnitude")
        self.signal_fft.setChecked(False)
        self.tabs.setCurrentIndex(1)
        self.update_signal_plot()

    def preview_tracker_in_artifact_learning(self):
        if self.tracker_matrix is None:
            QMessageBox.information(self, "Artifact Learning", "Load a Tracker file first.")
            return
        image = np.log1p(np.abs(self.tracker_matrix))
        self.artifact_preview_panel.label.setText(
            f"Tracker Raw Magnitude: {self.tracker_shared_name}"
        )
        self.artifact_preview_panel.set_image(image)
        self.artifact_selection_label.setText(
            f"Tracker data ready for learning: {self.tracker_shared_name}"
        )

    def save_tracker_training_sample(self):
        if self.tracker_matrix is None or self.tracker_source_path is None:
            QMessageBox.information(self, "Artifact Learning", "Load a Tracker file first.")
            return
        if not self._ensure_artifact_database():
            return
        artifact_type = self.artifact_type_combo.currentText().strip() or "Not Artifact"
        resolution = self.artifact_resolution_combo.currentText().strip()
        notes = self.artifact_notes.toPlainText().strip()
        tracker_image = np.log1p(np.abs(self.tracker_matrix))
        features = image_features(tracker_image, self.normal_reference_image)
        features.update({
            "data_type": "Tracker Raw",
            "strongest_line": self.tracker_strongest_line_index,
            "line_peak": float(np.max(np.abs(self.tracker_strongest_line))) if self.tracker_strongest_line is not None else 0.0,
            "line_rms": float(np.sqrt(np.mean(np.abs(self.tracker_strongest_line) ** 2))) if self.tracker_strongest_line is not None else 0.0,
        })
        self.artifact_db.add_sample(
            source_path=str(self.tracker_source_path),
            source_name=self.tracker_source_path.name,
            series_uid="TRACKER",
            series_description="Tracker Raw Data",
            instance_number=self.tracker_strongest_line_index,
            artifact_type=artifact_type,
            resolution=resolution,
            is_normal_reference=False,
            normal_reference_path=str(self.normal_reference_path or ""),
            features=features,
            notes=notes,
        )
        self.refresh_artifact_training_table()
        self.artifact_summary.setText(
            f"Saved Tracker training sample as '{artifact_type}'."
        )

    def detect_tracker_artifact_from_db(self):
        if self.tracker_matrix is None:
            QMessageBox.information(self, "Artifact Detection", "Load a Tracker file first.")
            return
        if not self._ensure_artifact_database():
            return
        keys, training, labels, metadata = self.artifact_db.training_feature_vectors()
        tracker_image = np.log1p(np.abs(self.tracker_matrix))
        features = image_features(tracker_image, self.normal_reference_image)
        features.update({
            "data_type": "Tracker Raw",
            "strongest_line": self.tracker_strongest_line_index,
        })
        vector = features_to_vector(features, keys)
        result = classify_feature_vector(
            vector, training, labels,
            minimum_samples_per_class=self.detect_min_samples.value(),
        )
        self.artifact_detection_table.setRowCount(1)
        alternatives = ", ".join(
            f"{label} {confidence * 100:.1f}%"
            for label, confidence, distance, support in result.alternatives[1:4]
        )
        values = [
            self.tracker_source_path.name if self.tracker_source_path else "Tracker",
            "Tracker Raw Data",
            result.label,
            f"{result.confidence * 100:.1f}%",
            f"{result.distance:.4f}" if np.isfinite(result.distance) else "-",
            result.support,
            result.status,
            alternatives,
            str(self.tracker_source_path or ""),
        ]
        for column, value in enumerate(values):
            self.artifact_detection_table.setItem(0, column, QTableWidgetItem(str(value)))
        self.artifact_detection_summary.setText(
            f"Tracker prediction: {result.label} | Confidence {result.confidence * 100:.1f}%"
        )
        self.tabs.setCurrentIndex(2)

    def _image_mouse_action_changed(self, action: str):
        if action == "Auto Window/Level":
            self.auto_levels()
            self.statusBar().showMessage("Window/Level recalculated automatically")
            return
        if action.startswith("Preset:"):
            preset = action.split(":", 1)[1]
            self.level_preset_combo.setCurrentText(preset)
            self.apply_level_preset(preset)
            self.statusBar().showMessage(f"Window/Level preset applied: {preset}")
            return
        # Mouse modes are no longer switchable. Keep a single predictable
        # standard mapping across Original, FFT and Both layouts.
        action = "Standard"
        self.current_mouse_action = action
        self.primary_panel.mouse_action = action
        self.secondary_panel.mouse_action = action
        self.primary_panel.plot.getViewBox().setMouseMode(pg.ViewBox.PanMode)
        self.secondary_panel.plot.getViewBox().setMouseMode(pg.ViewBox.PanMode)
        self.mouse_action_label.setText("Mouse: Standard")
        tooltips = {
            "Standard": "Wheel: Previous / Next image\nLeft-drag: Pan\nRight-drag: Window/Level\nCtrl+Wheel: Zoom",
        }
        self.mouse_action_label.setToolTip(tooltips.get(action, action))
        if action == "ROI":
            if self.current_kspace is not None and self.source_kind != "bitmap":
                self.primary_panel.show_comp_roi()
            else:
                self.statusBar().showMessage("ROI requires raw/k-space data")
        self.statusBar().showMessage(f"Mouse action selected: {action}")


    def _image_view_requested(self, mode: str):
        if mode == "FFT_CURRENT":
            if self.current_image is None:
                return
            # Preserve the complete display state. Back must restore the original
            # k-space and its independent WW/WL, not create a new FFT state.
            self.fft_back_state = {
                "image": np.asarray(self.current_image).copy(),
                "kspace": None if self.current_kspace is None else np.asarray(self.current_kspace).copy(),
                "recon": None if self.current_recon is None else np.asarray(self.current_recon).copy(),
                "view_mode": str(self.view_mode),
                "original_window_level": self.original_window_level,
                "original_dynamic_range": self.original_dynamic_range,
                "raw_window_level": self.raw_window_level,
                "raw_dynamic_range": self.raw_dynamic_range,
                "original_display_mode": self.original_display_mode,
                "fft_display_mode": self.fft_display_mode,
                "level_target": self.level_target_combo.currentText(),
                "level_preset": self.level_preset_combo.currentText(),
                "fft_level_signature": getattr(self, "_fft_level_signature", None),
            }
            self.fft_back_image = np.asarray(self.current_image).copy()
            self.current_kspace, self.current_recon, _cache_hit = self._fft_for_current_image(
                self.current_source
            )
            self.raw_window_level = None
            self.raw_dynamic_range = None
            self._fft_level_signature = None
            self.fft_view_active = True
            self.set_view_mode("FFT")
            self.statusBar().showMessage("FFT calculated and displayed in the current panel")
            return
        if mode == "BACK_FFT":
            state = self.fft_back_state
            if state:
                self.current_image = np.asarray(state["image"]).copy()
                self.current_kspace = None if state["kspace"] is None else np.asarray(state["kspace"]).copy()
                self.current_recon = None if state["recon"] is None else np.asarray(state["recon"]).copy()
                self.original_window_level = state["original_window_level"]
                self.original_dynamic_range = state["original_dynamic_range"]
                self.raw_window_level = state["raw_window_level"]
                self.raw_dynamic_range = state["raw_dynamic_range"]
                self.original_display_mode = state["original_display_mode"]
                self.fft_display_mode = state["fft_display_mode"]
                self._fft_level_signature = state["fft_level_signature"]
                self.level_target_combo.blockSignals(True)
                self.level_preset_combo.blockSignals(True)
                self.level_target_combo.setCurrentText(state["level_target"])
                self.level_preset_combo.setCurrentText(state["level_preset"])
                self.level_target_combo.blockSignals(False)
                self.level_preset_combo.blockSignals(False)
            elif self.fft_back_image is not None:
                self.current_image = np.asarray(self.fft_back_image).copy()
                self.current_kspace = fft2c(self.current_image)
                self.current_recon = ifft2c(self.current_kspace)
            self.fft_view_active = False
            self.fft_back_image = None
            self.fft_back_state = None
            self.set_view_mode("Original")
            self._sync_level_controls()
            self.statusBar().showMessage("Returned to the image before FFT")
            return
        if mode == "Default":
            self.profile_mode.setCurrentText("Magnitude")
            self.level_target_combo.setCurrentText("Original Image")
            self.original_window_level = None
            self.original_dynamic_range = None
            self.raw_window_level = None
            self.raw_dynamic_range = None
            self.fft_view_active = False
            self.fft_back_image = None
            self.fft_back_state = None
            self.primary_panel.fit_to_image()
            self.secondary_panel.fit_to_image()
            self.set_view_mode("Both" if self.source_kind != "bitmap" else "Original")
            return
        self.fft_view_active = False
        self.fft_back_image = None
        self.fft_back_state = None
        self.set_view_mode(mode)


    def _prepare_level_controls_for_panel(self, panel):
        """Point the shared WW/WL controls at the clicked panel's real role."""
        target = "Raw Data" if self._panel_role(panel) == "fft" else "Original Image"
        self.level_target_combo.blockSignals(True)
        self.level_target_combo.setCurrentText(target)
        self.level_target_combo.blockSignals(False)
        self._sync_level_controls()

    def _image_level_wheel(self, panel, step: float, adjust_width: bool):
        target = "Raw Data" if self._panel_role(panel) == "fft" else "Original Image"
        self.level_target_combo.setCurrentText(target)

        if target == "Raw Data":
            center = float(self.raw_window_level or 0.0)
            width = max(float(self.raw_dynamic_range or 1.0), 1e-9)
        else:
            center = float(self.original_window_level or 0.0)
            width = max(float(self.original_dynamic_range or 1.0), 1e-9)

        increment = max(width * 0.04, 1e-6)
        if adjust_width:
            width = max(width + step * increment, 1e-9)
        else:
            center += step * increment

        if target == "Raw Data":
            self.raw_window_level = center
            self.raw_dynamic_range = width
        else:
            self.original_window_level = center
            self.original_dynamic_range = width

        self._sync_level_controls()
        self.refresh_images()

    def _accordion_opened(self, opened):
        for section in self.accordion_sections:
            if section is not opened: section.set_expanded(False,False)

    def _spike_range_changed(self,value):
        visible=value=="Scale"; self.spike_scale_slider.setVisible(visible); self.spike_scale_label.setVisible(visible)

    def show_dicom_header_popup(self):
        if self.current_ds is None: QMessageBox.information(self,"DICOM Header","Load a DICOM image first."); return
        DicomHeaderDialog(self.current_ds,self.current_source,self).exec()

    def _toggle_developer_mode(self, enabled: bool):
        self.developer_mode = bool(enabled)
        if hasattr(self, "spike_developer_text"):
            self.spike_developer_text.setVisible(self.developer_mode)
        if self.developer_mode and hasattr(self, "current_spike_result_index"):
            self._show_spike_result(self.current_spike_result_index)

    @classmethod
    def _extract_stripe_only(cls, image: np.ndarray) -> dict:
        source = np.asarray(image, dtype=float)
        source = source - np.median(source)
        rows, cols = source.shape
        spectrum = fft2c(source)
        magnitude = np.abs(spectrum)
        cy, cx = rows // 2, cols // 2
        yy, xx = np.indices((rows, cols))
        dy = yy - cy
        dx = xx - cx
        radius = np.hypot(dy, dx)
        angle = np.arctan2(dy, dx)
        center_radius = max(5, int(min(rows, cols) * 0.04))
        cross_width = max(2, int(min(rows, cols) * 0.015))
        valid = (radius > center_radius) & (np.abs(dy) > cross_width) & (np.abs(dx) > cross_width)
        logmag = np.log1p(magnitude)
        zmap = np.zeros_like(logmag)
        zmap[valid] = cls._robust_zscore(logmag[valid])

        # Direction/scale groups. Multi-spike images may contain many groups.
        groups=[]
        candidate_mask=np.zeros_like(valid)
        angle_bins=np.linspace(-np.pi, np.pi, 37)
        radial_bins=[center_radius, max(center_radius+1,int(min(rows,cols)*0.12)), int(min(rows,cols)*0.25), int(min(rows,cols)*0.50)]
        for a0,a1 in zip(angle_bins[:-1],angle_bins[1:]):
            amask=valid & (angle>=a0) & (angle<a1)
            for ri,(r0,r1) in enumerate(zip(radial_bins[:-1],radial_bins[1:])):
                mask=amask & (radius>=r0) & (radius<r1)
                if not np.any(mask): continue
                peak=float(np.percentile(zmap[mask],99.7))
                count=int(np.sum(zmap[mask]>5.5))
                if peak>=6.0 and count>=1:
                    group_mask=mask & (zmap>=max(5.5,peak*0.62))
                    candidate_mask |= group_mask
                    groups.append({"angle":float(np.rad2deg((a0+a1)/2.0)),"scale":["Wide","Mid","Fine"][min(ri,2)],"peak_z":peak,"points":int(np.sum(group_mask))})
        # Symmetric completion and thin sparse support avoid ring/band inclusion.
        candidate_mask |= np.flip(np.flip(candidate_mask,axis=0),axis=1)
        candidate_spectrum=np.where(candidate_mask,spectrum,0)
        stripe_only=np.real(ifft2c(candidate_spectrum))
        return {"stripe_only":stripe_only,"candidate_spectrum":candidate_spectrum,"candidate_mask":candidate_mask,"groups":groups,"valid_mask":valid,"z_map":zmap}

    @classmethod
    def _map_stripe_to_raw(cls, stripe_info: dict, actual_kspace: np.ndarray) -> dict:
        predicted=np.abs(stripe_info["candidate_spectrum"])
        actual=np.abs(actual_kspace)
        mask=np.asarray(stripe_info["candidate_mask"],dtype=bool)
        if not np.any(mask):
            return {"agreement":0.0,"matched_mask":mask,"candidate_image":predicted,"matched_coords":np.empty((0,2),dtype=int)}
        local=np.stack([np.roll(actual,(dy,dx),axis=(0,1)) for dy,dx in ((-1,0),(1,0),(0,-1),(0,1),(-1,-1),(1,1))])
        med=np.median(local,axis=0)
        mad=np.median(np.abs(local-med[None,...]),axis=0)
        z=(actual-med)/np.maximum(1.4826*mad,1e-9)
        matched=mask & (z>6.0) & (actual>np.maximum(med*2.5,1e-9))
        predicted_points=max(int(np.sum(mask)),1)
        agreement=float(np.sum(matched)/predicted_points)
        return {"agreement":agreement,"matched_mask":matched,"candidate_image":np.log1p(predicted),"matched_coords":np.argwhere(matched),"raw_z":z}

    @staticmethod
    def _robust_zscore(values: np.ndarray) -> np.ndarray:
        array = np.asarray(values, dtype=float)
        if array.size == 0:
            return np.asarray([], dtype=float)
        median = float(np.median(array))
        mad = float(np.median(np.abs(array - median)))
        scale = max(1.4826 * mad, 1e-9)
        return (array - median) / scale

    @classmethod
    def _image_stripe_features(cls, image: np.ndarray) -> dict:
        source = np.asarray(image, dtype=float)
        source = source - np.median(source)
        rows, cols = source.shape

        def projection_score(projection):
            spectrum = np.abs(np.fft.rfft(projection))
            if spectrum.size < 12:
                return 0.0, 0
            z = cls._robust_zscore(spectrum[2:])
            return float(np.mean(np.sort(z)[-8:])), int(np.sum(z > 5.5))

        row_strength, row_count = projection_score(np.mean(source, axis=1))
        col_strength, col_count = projection_score(np.mean(source, axis=0))

        yy, xx = np.indices((rows, cols))
        angle_results = []
        for angle in (-75, -60, -45, -30, -15, 15, 30, 45, 60, 75):
            theta = np.deg2rad(angle)
            coordinate = xx * np.cos(theta) + yy * np.sin(theta)
            bins = np.floor(coordinate - coordinate.min()).astype(int)
            sums = np.bincount(bins.ravel(), weights=source.ravel())
            counts = np.bincount(bins.ravel())
            projection = sums / np.maximum(counts, 1)
            strength, count = projection_score(projection)
            angle_results.append((strength, count, angle))
        angle_strength, angle_count, best_angle = max(angle_results, key=lambda item: item[0])

        frequency = np.abs(fft2c(source))
        cy, cx = rows // 2, cols // 2
        radius = np.hypot(yy - cy, xx - cx)
        valid = (
            (radius > max(5, int(min(rows, cols) * 0.04)))
            & (np.abs(yy - cy) > max(2, int(min(rows, cols) * 0.018)))
            & (np.abs(xx - cx) > max(2, int(min(rows, cols) * 0.018)))
        )
        z_map = np.zeros_like(frequency)
        z_map[valid] = cls._robust_zscore(np.log1p(frequency[valid]))
        diagonal_strength = float(np.percentile(z_map[valid], 99.9)) if np.any(valid) else 0.0

        gx = np.diff(source, axis=1, prepend=source[:, :1])
        gy = np.diff(source, axis=0, prepend=source[:1, :])
        gradient = np.hypot(gx, gy)
        edge_mask = gradient >= np.percentile(gradient, 92)
        line_persistence = max(
            float(np.max(np.mean(edge_mask, axis=1))),
            float(np.max(np.mean(edge_mask, axis=0))),
        )

        return {
            "stripe_strength": max(row_strength, col_strength, angle_strength),
            "periodic_count": max(row_count, col_count, angle_count),
            "diagonal_strength": diagonal_strength,
            "direction_concentration": min(1.0, angle_strength / 12.0),
            "line_persistence": line_persistence,
            "best_stripe_angle": float(best_angle),
            "frequency": frequency,
            "z_map": z_map,
            "valid_mask": valid,
        }


    @classmethod
    def _directional_kspace_line_features(cls, frequency: np.ndarray) -> dict:
        """Find narrow abnormal k-space lines without discarding the centre axes.

        The previous detector masked the complete central row and column.  Real
        spike-contaminated images in the supplied DICOM set often produce a
        narrow line that crosses the k-space centre, so that mask removed the
        strongest evidence.  Here only a small DC box is excluded from each
        projection and line energy is compared with its local neighbours.
        """
        magnitude = np.log1p(np.maximum(np.asarray(frequency, dtype=float), 0.0))
        rows, cols = magnitude.shape
        cy, cx = rows // 2, cols // 2
        dc_y = max(3, int(rows * 0.012))
        dc_x = max(3, int(cols * 0.012))

        row_data = magnitude.copy()
        row_data[:, max(0, cx-dc_x):min(cols, cx+dc_x+1)] = np.nan
        col_data = magnitude.copy()
        col_data[max(0, cy-dc_y):min(rows, cy+dc_y+1), :] = np.nan
        row_energy = np.nanpercentile(row_data, 92.0, axis=1)
        col_energy = np.nanpercentile(col_data, 92.0, axis=0)

        def local_score(values):
            values = np.nan_to_num(values, nan=float(np.nanmedian(values)))
            neighbours = np.stack([
                np.roll(values, shift) for shift in (-4, -3, -2, -1, 1, 2, 3, 4)
            ])
            median = np.median(neighbours, axis=0)
            mad = np.median(np.abs(neighbours - median[None, :]), axis=0)
            return (values - median) / np.maximum(1.4826 * mad, 1e-6)

        row_score = np.maximum(local_score(row_energy), cls._robust_zscore(row_energy))
        col_score = np.maximum(local_score(col_energy), cls._robust_zscore(col_energy))
        row_flags = np.flatnonzero(row_score >= 7.0)
        col_flags = np.flatnonzero(col_score >= 7.0)
        return {
            "row_score": row_score,
            "col_score": col_score,
            "rows": row_flags.tolist(),
            "cols": col_flags.tolist(),
            "maximum": float(max(np.max(row_score), np.max(col_score))),
        }

    @classmethod
    def _raw_spike_features(cls, frequency: np.ndarray, valid_mask: np.ndarray) -> dict:
        """Detect isolated points and abnormal complete/partial k-space lines.

        MRI spike noise is not always a single bright sample.  Some acquisitions
        contain a short cluster or a complete phase-encode line, which the old
        point-only detector missed.  This detector combines local 2-D MAD with
        robust row/column energy and peak projections.
        """
        magnitude = np.asarray(frequency, dtype=float)
        rows, cols = magnitude.shape
        cy, cx = rows // 2, cols // 2
        safe = np.log1p(np.maximum(magnitude, 0.0))

        neighbor_stack = np.stack([
            np.roll(safe, (dy, dx), axis=(0, 1))
            for dy, dx in (
                (-2, 0), (-1, -1), (-1, 0), (-1, 1),
                (0, -2), (0, -1), (0, 1), (0, 2),
                (1, -1), (1, 0), (1, 1), (2, 0),
            )
        ])
        local_median = np.median(neighbor_stack, axis=0)
        local_mad = np.median(np.abs(neighbor_stack - local_median[None, ...]), axis=0)
        local_z = (safe - local_median) / np.maximum(1.4826 * local_mad, 1e-6)

        global_z = np.zeros_like(safe)
        global_z[valid_mask] = cls._robust_zscore(safe[valid_mask])
        point_mask = valid_mask & (local_z > 6.0) & (global_z > 4.5)

        # Projection statistics reveal bright horizontal/vertical k-space lines.
        masked = np.where(valid_mask, safe, np.nan)
        row_energy = np.nanmedian(masked, axis=1)
        col_energy = np.nanmedian(masked, axis=0)
        row_peak = np.nanpercentile(masked, 97.5, axis=1)
        col_peak = np.nanpercentile(masked, 97.5, axis=0)
        row_score = np.maximum(cls._robust_zscore(np.nan_to_num(row_energy, nan=np.nanmedian(row_energy))),
                               cls._robust_zscore(np.nan_to_num(row_peak, nan=np.nanmedian(row_peak))))
        col_score = np.maximum(cls._robust_zscore(np.nan_to_num(col_energy, nan=np.nanmedian(col_energy))),
                               cls._robust_zscore(np.nan_to_num(col_peak, nan=np.nanmedian(col_peak))))

        center_guard_y = max(3, int(rows * 0.025))
        center_guard_x = max(3, int(cols * 0.025))
        row_flags = np.flatnonzero((row_score > 5.0) & (np.abs(np.arange(rows) - cy) > center_guard_y))
        col_flags = np.flatnonzero((col_score > 5.0) & (np.abs(np.arange(cols) - cx) > center_guard_x))

        # Merge adjacent lines into clusters and retain the strongest member plus
        # immediate neighbors so correction does not leave a split bright band.
        def expanded_lines(indices, score, limit):
            chosen = set()
            for group in cls._group_spike_indices(indices):
                strongest = int(group[np.argmax(score[group])])
                for value in range(max(0, int(group.min()) - 1), min(limit, int(group.max()) + 2)):
                    chosen.add(value)
                chosen.add(strongest)
            return sorted(chosen)

        abnormal_rows = expanded_lines(row_flags, row_score, rows)
        abnormal_cols = expanded_lines(col_flags, col_score, cols)
        line_mask = np.zeros_like(valid_mask, dtype=bool)
        for y in abnormal_rows:
            line_mask[y, :] = valid_mask[y, :]
        for x in abnormal_cols:
            line_mask[:, x] = valid_mask[:, x]

        candidate_mask = point_mask | line_mask
        coords = []
        # Isolated points are all retained.
        coords.extend(map(tuple, np.argwhere(point_mask)))
        # For line anomalies, correct only locally extreme samples, capped to keep
        # processing responsive.  If a complete line is uniformly abnormal, take
        # regularly spaced samples across that line.
        for y in abnormal_rows:
            valid_x = np.flatnonzero(valid_mask[y])
            if valid_x.size:
                ranked = valid_x[np.argsort(local_z[y, valid_x])[-min(48, valid_x.size):]]
                coords.extend((int(y), int(x)) for x in ranked if local_z[y, x] > 2.5)
                if not any(c[0] == y for c in coords):
                    coords.extend((int(y), int(x)) for x in valid_x[::max(1, valid_x.size // 24)])
        for x in abnormal_cols:
            valid_y = np.flatnonzero(valid_mask[:, x])
            if valid_y.size:
                ranked = valid_y[np.argsort(local_z[valid_y, x])[-min(48, valid_y.size):]]
                coords.extend((int(y), int(x)) for y in ranked if local_z[y, x] > 2.5)
                if not any(c[1] == x for c in coords):
                    coords.extend((int(y), int(x)) for y in valid_y[::max(1, valid_y.size // 24)])
        coords = np.asarray(sorted(set(coords)), dtype=int) if coords else np.empty((0, 2), dtype=int)

        pair_hits = 0
        for y, x in coords[:1000]:
            sy = (2 * cy - int(y)) % rows
            sx = (2 * cx - int(x)) % cols
            if candidate_mask[sy, sx] or local_z[sy, sx] > 4.0:
                pair_hits += 1

        return {
            "isolated_count": int(np.sum(point_mask)),
            "max_local_z": float(np.max(local_z[candidate_mask])) if np.any(candidate_mask) else 0.0,
            "pair_hits": pair_hits,
            "random_spread": 0.0,
            "candidate_coords": coords,
            "local_z": local_z,
            "global_z": global_z,
            "abnormal_rows": abnormal_rows,
            "abnormal_cols": abnormal_cols,
            "row_score": row_score,
            "col_score": col_score,
            "line_mask": line_mask,
        }

    @classmethod
    def _quick_spike_fast_raw_support(cls, actual_raw: np.ndarray) -> dict:
        """Run the real lightweight k-space spike screen used by Quick Spike.

        This deliberately avoids the heavy Blob/Band/Ring candidate engine, but it
        still performs robust local-MAD point detection, row/column anomaly checks,
        and Hermitian-pair validation.  Returning stage metadata makes it possible
        to distinguish "analysis completed with no candidates" from a skipped run.
        """
        raw = np.asarray(actual_raw)
        if raw.ndim != 2 or min(raw.shape) < 8:
            raise ValueError(f"Quick Spike requires a 2-D image/raw plane, got {raw.shape}")
        magnitude = np.abs(raw).astype(float, copy=False)
        rows, cols = magnitude.shape
        cy, cx = rows // 2, cols // 2
        yy, xx = np.indices(magnitude.shape)
        radius = np.hypot(yy - cy, xx - cx)
        valid = radius > max(5, int(min(rows, cols) * 0.045))
        valid &= np.abs(yy - cy) > max(1, int(rows * 0.008))
        valid &= np.abs(xx - cx) > max(1, int(cols * 0.008))
        if not np.any(valid):
            return {
                "analysis_completed": True,
                "stages_completed": 3,
                "accepted_count": 0,
                "selected_types": [],
                "quality": 0.0,
                "peak_count": 0,
                "pair_hits": 0,
                "abnormal_line_count": 0,
            }

        features = cls._raw_spike_features(magnitude, valid)
        coords = np.asarray(features.get("candidate_coords", np.empty((0, 2), dtype=int)))
        isolated = int(features.get("isolated_count", 0))
        pair_hits = int(features.get("pair_hits", 0))
        abnormal_lines = len(features.get("abnormal_rows", [])) + len(features.get("abnormal_cols", []))
        max_local_z = float(features.get("max_local_z", 0.0))
        peak_count = int(len(coords))
        quality = min(100.0, isolated * 2.0 + pair_hits * 3.0 + abnormal_lines * 5.0 + max_local_z * 1.5)
        accepted = int(
            (isolated >= 1 and (pair_hits >= 1 or max_local_z >= 9.0))
            or abnormal_lines >= 1
        )
        return {
            "analysis_completed": True,
            "stages_completed": 3,
            "accepted_count": accepted,
            "selected_types": ["Spike"] if accepted else [],
            "quality": float(quality),
            "peak_count": peak_count,
            "pair_hits": pair_hits,
            "abnormal_line_count": int(abnormal_lines),
            "max_local_z": max_local_z,
        }

    def quick_spike_detect_prototype(self):
        if not self.dicom_entries:
            QMessageBox.information(self, "Quick Spike Detect", "Load DICOM images first.")
            return
        indices = self._selected_dicom_indices() or list(range(len(self.dicom_entries)))
        # Fast screening limits very large studies while preserving first/last and even coverage.
        if len(indices) > 160:
            sample_pos = np.linspace(0, len(indices)-1, 160, dtype=int)
            indices = [indices[int(i)] for i in sample_pos]
        progress=self._make_progress("Quick Spike Detect — Fast", max(1, len(indices) * 4))
        records=[]; errors=0; completed_images=0; completed_stages=0
        started_at = time.perf_counter()
        try:
            for position, index in enumerate(indices):
                if progress.wasCanceled(): break
                try:
                    entry=self._ensure_dicom_image(index)
                    base = position * 4
                    progress.setLabelText(f"1/4 Preparing image {position+1}/{len(indices)}\n{entry.path.name}")
                    progress.setValue(base); QApplication.processEvents()
                    image=np.asarray(entry.image,dtype=float)
                    if image.ndim != 2 or image.size == 0:
                        raise ValueError(f"Unsupported image plane: {image.shape}")
                    progress.setLabelText(f"2/4 Building FFT {position+1}/{len(indices)}\n{entry.path.name}")
                    actual_raw=fft2c(image)
                    progress.setValue(base+1); QApplication.processEvents()
                    progress.setLabelText(f"3/4 Detecting stripe evidence {position+1}/{len(indices)}\n{entry.path.name}")
                    stripe=self._extract_stripe_only(image)
                    mapping=self._map_stripe_to_raw(stripe,actual_raw)
                    progress.setValue(base+2); QApplication.processEvents()
                    progress.setLabelText(f"4/4 Validating k-space spikes {position+1}/{len(indices)}\n{entry.path.name}")
                    raw_comp_analysis=self._quick_spike_fast_raw_support(actual_raw)
                    if not raw_comp_analysis.get("analysis_completed", False):
                        raise RuntimeError("Quick Spike validation did not complete")
                    progress.setValue(base+3); QApplication.processEvents()
                    completed_images += 1
                    completed_stages += int(raw_comp_analysis.get("stages_completed", 0)) + 1
                    groups=stripe["groups"]
                    stripe_energy=float(np.std(stripe["stripe_only"])/(np.std(image)+1e-9))
                    group_score=sum(min(g["peak_z"],15.0) for g in groups)
                    multi_direction=len({round(g["angle"]/15)*15 for g in groups})
                    raw_agreement=float(mapping["agreement"])
                    # Raw agreement is mandatory. Strong curved/ring energy without sparse matched peaks is not Spike.
                    auto_types=set(raw_comp_analysis.get("selected_types", []))
                    auto_spike_support=1.0 if auto_types.intersection({"Spike","Blob","Block","Diagonal"}) else 0.0
                    score=(min(group_score,60)*0.08 + min(len(groups),16)*0.35 + min(multi_direction,8)*0.4 + min(stripe_energy,1.0)*5.0 + raw_agreement*18.0 + auto_spike_support*8.0 + min(raw_comp_analysis.get("quality",0.0),100.0)*0.04)
                    series=str(getattr(entry.ds,"SeriesInstanceUID","") or getattr(entry.ds,"SeriesNumber","") or "UnknownSeries")
                    records.append({"index":index,"series":series,"score":score,"groups":groups,"group_count":len(groups),"directions":multi_direction,"stripe_energy":stripe_energy,"raw_agreement":raw_agreement,"matched_count":int(len(mapping["matched_coords"])),"stripe_only":stripe["stripe_only"],"raw_candidate":mapping["candidate_image"],"raw_comp_analysis":raw_comp_analysis,"auto_spike_support":auto_spike_support})
                except Exception as exc:
                    errors+=1
        finally:
            progress.setValue(max(1, len(indices) * 4))
            progress.hide(); self._stabilize_layout()
        elapsed = time.perf_counter() - started_at
        by_series={}
        for r in records: by_series.setdefault(r["series"],[]).append(r)
        high=[]; review=[]; self.quick_spike_details={}
        for group in by_series.values():
            vals=np.array([r["score"] for r in group],dtype=float)
            zvals=self._robust_zscore(vals) if len(vals)>=4 else np.zeros(len(vals))
            for r,z in zip(group,zvals):
                # Mandatory image stripe + actual raw match. Multiple directions/scales are allowed and rewarded.
                raw_ok=(r["raw_agreement"]>=0.015 and r["matched_count"]>=1) or bool(r.get("auto_spike_support"))
                stripe_ok=r["group_count"]>=1 and r["stripe_energy"]>=0.006
                multi_ok=r["directions"]>=2 or r["group_count"]>=3
                confidence="No Spike"
                if raw_ok and stripe_ok and (multi_ok or z>=3.5) and r["score"]>=5.0:
                    high.append(r["index"]); confidence="High Confidence"
                elif raw_ok and stripe_ok and r["score"]>=3.0:
                    review.append(r["index"]); confidence="Review"
                r["series_z"]=float(z); r["confidence"]=confidence
                self.quick_spike_details[r["index"]]=r
        self.quick_spike_indices=sorted(set(high)); self.quick_spike_review_indices=sorted(set(review))
        self._highlight_quick_spike_candidates(self.quick_spike_indices,self.quick_spike_review_indices)
        dialog=QMessageBox(self); dialog.setWindowTitle("Quick Spike Detect")
        execution_state = "Completed" if completed_images else "Not completed"
        dialog.setText(
            f"Analysis: {execution_state}\n"
            f"High Confidence: {len(self.quick_spike_indices)}\n"
            f"Review: {len(self.quick_spike_review_indices)}\n"
            f"Read Errors: {errors}\n"
            f"Analyzed Images: {completed_images} / {len(indices)}\n"
            f"Detection Stages Completed: {completed_stages}\n"
            f"Elapsed: {elapsed:.3f} seconds\n\n"
            "Logic: FFT generation + stripe extraction + robust local-MAD/line detection + symmetry validation. "
            "Full Blob/Band/Ring analysis remains in Artifact Diag."
        )
        dialog.addButton("Close",QMessageBox.RejectRole); check=dialog.addButton("Check High Confidence",QMessageBox.AcceptRole); dialog.exec()
        if dialog.clickedButton()==check:
            hs=set(self.quick_spike_indices); it=QTreeWidgetItemIterator(self.tree)
            while it.value():
                item=it.value(); data=item.data(0,Qt.UserRole)
                if data and data[0]=="dicom": item.setCheckState(0,Qt.Checked if int(data[1]) in hs else Qt.Unchecked)
                it+=1
            if hs: self.show_dicom(min(hs))


    def _highlight_quick_spike_candidates(self, high_indices, review_indices=None):
        high={int(v) for v in high_indices}; review={int(v) for v in (review_indices or [])}
        it=QTreeWidgetItemIterator(self.tree)
        while it.value():
            item=it.value(); data=item.data(0,Qt.UserRole)
            if data and data[0]=="dicom":
                i=int(data[1])
                if i in high: item.setForeground(0,pg.mkBrush("#ff4f4f")); item.setToolTip(0,"Quick Spike: High Confidence")
                elif i in review: item.setForeground(0,pg.mkBrush("#ffb347")); item.setToolTip(0,"Quick Spike: Review")
                else: item.setForeground(0,pg.mkBrush("#d8e5ef")); item.setToolTip(0,"")
            it+=1


    def _select_quick_spike_candidates(self, indices):
        target = {int(value) for value in indices}
        self.tree.blockSignals(True)
        try:
            self.tree.clearSelection()
            iterator = QTreeWidgetItemIterator(self.tree)
            first_item = None
            while iterator.value():
                item = iterator.value()
                data = item.data(0, Qt.UserRole)
                if data and data[0] == "dicom" and int(data[1]) in target:
                    item.setSelected(True)
                    if first_item is None:
                        first_item = item
                iterator += 1
            if first_item is not None:
                self.tree.setCurrentItem(first_item)
        finally:
            self.tree.blockSignals(False)

        if target:
            self.show_dicom(min(target))


    def on_display_type_changed(self, value: str):
        self.refresh_images()
        self.update_line_profile()

    @staticmethod
    def _display_component(array: np.ndarray, mode: str) -> np.ndarray:
        data = np.asarray(array)
        if mode == "Log Magnitude":
            return np.log1p(np.abs(data))
        if mode == "Real":
            return np.real(data)
        if mode == "Imaginary":
            return np.imag(data)
        if mode == "Phase":
            return np.angle(data)
        return np.abs(data)

    @staticmethod
    def _available_components_for_array(array: np.ndarray, *, fft: bool = False):
        if array is None:
            return ["Magnitude"]
        data = np.asarray(array)
        if fft:
            return ["Magnitude", "Log Magnitude", "Real", "Imaginary", "Phase"]
        if np.iscomplexobj(data) and np.any(np.abs(np.imag(data)) > 0):
            return ["Magnitude", "Real", "Imaginary", "Phase"]
        return ["Magnitude"]

    def set_view_mode(self, mode):
        if self.source_kind == "bitmap" and mode != "Original":
            mode = "Original"
        self.view_mode = mode
        for button, value in [
            (self.btn_fft, "FFT"),
            (self.btn_original, "Original"),
            (self.btn_both, "Both"),
        ]:
            button.setChecked(value == mode)
        self.refresh_images()
        self._schedule_responsive_layout()



    def _panel_role(self, panel):
        if self.view_mode == "Both":
            return "original" if panel is self.primary_panel else "fft"
        if self.view_mode == "FFT":
            return "fft"
        return "original"

    def _set_panel_component(self, panel, component: str):
        role = self._panel_role(panel)
        available = self._available_components_for_array(
            self.current_kspace if role == "fft" else self.current_image,
            fft=(role == "fft"),
        )
        if component not in available:
            component = available[0]
        if role == "fft":
            self.fft_display_mode = component
            self.raw_window_level = None
            self.raw_dynamic_range = None
        else:
            self.original_display_mode = component
            self.original_window_level = None
            self.original_dynamic_range = None
        if component in [self.profile_mode.itemText(i) for i in range(self.profile_mode.count())]:
            self.profile_mode.setCurrentText(component)
        self.refresh_images()

    def _frequency_encoding_direction(self):
        """Return display-axis direction perpendicular to DICOM phase encoding."""
        ds = getattr(self, "current_ds", None)
        if ds is None:
            return None
        value = str(getattr(ds, "InPlanePhaseEncodingDirection", "") or "").strip().upper()
        if not value:
            try:
                element = ds.get((0x0018, 0x1312))
                value = str(getattr(element, "value", element) or "").strip().upper()
            except Exception:
                value = ""
        # DICOM ROW means phase varies along image rows, therefore frequency is
        # along columns (screen-horizontal). COL is the perpendicular case.
        if value == "ROW":
            direction = "HORIZONTAL"
        elif value == "COL":
            direction = "VERTICAL"
        else:
            return None
        # Follow the current display transform when a 90/270 degree rotation is used.
        transform = getattr(self, "console_display_transform", None)
        rotation = int(getattr(transform, "rotation_degrees", 0) or 0) % 360
        if rotation in (90, 270):
            direction = "VERTICAL" if direction == "HORIZONTAL" else "HORIZONTAL"
        return direction

    def _apply_frequency_markers(self):
        direction = self._frequency_encoding_direction()
        for panel in (self.primary_panel, self.secondary_panel):
            panel.set_frequency_direction(direction)

    def refresh_images(self):
        self.stable_diagnostics.info(
            "REFRESH_IMAGES_BEGIN",
            view_mode=self.view_mode,
            current_image=self.stable_diagnostics.array_summary(
                self.current_image
            ),
            current_kspace=self.stable_diagnostics.array_summary(
                self.current_kspace
            ),
        )
        if self.current_image is None:
            return
        original_available = self._available_components_for_array(self.current_image, fft=False)
        fft_available = self._available_components_for_array(self.current_kspace, fft=True)
        original_mode = getattr(self, "original_display_mode", "Magnitude")
        fft_mode = getattr(self, "fft_display_mode", "Magnitude")
        if original_mode not in original_available:
            original_mode = original_available[0]
            self.original_display_mode = original_mode
        if fft_mode not in fft_available:
            fft_mode = fft_available[0]
            self.fft_display_mode = fft_mode
        orig = self._display_component(self.current_image, original_mode)

        if self.source_kind == "bitmap":
            self.view_mode = "Original"
            auto_original = (self.level_target_combo.currentText() == "Original Image" and self.level_preset_combo.currentText() == "Auto")
            original_levels = self._levels_for_target("Original Image", orig, auto_original=auto_original)
            self.primary_panel.show()
            self.secondary_panel.hide()
            self.primary_panel.set_available_components(original_available, original_mode)
            self.primary_panel.label.setText(f"Original Image — {original_mode} ▼")
            self.primary_panel.set_image(orig, levels=original_levels)
            self.primary_panel.set_frequency_direction(None)
            self.secondary_panel.set_frequency_direction(None)
            self._sync_level_controls()
            rows, cols = orig.shape
            self._configure_line_control(rows, cols)
            self._sync_lines()
            self.update_line_profile()
            return

        if self.current_kspace is None:
            return
        fft_img = self._display_component(self.current_kspace, fft_mode)
        fft_signature = (tuple(np.asarray(self.current_kspace).shape), float(np.nanmax(np.abs(self.current_kspace))) if np.asarray(self.current_kspace).size else 0.0)
        if getattr(self, "_fft_level_signature", None) != fft_signature:
            self._fft_level_signature = fft_signature
            self.raw_window_level = None
            self.raw_dynamic_range = None
        auto_original = (self.level_target_combo.currentText() == "Original Image" and self.level_preset_combo.currentText() == "Auto")
        original_levels = self._levels_for_target("Original Image", orig, auto_original=auto_original)
        raw_levels = self._levels_for_target("Raw Data", fft_img, auto_original=False)

        if self.view_mode == "FFT":
            self.primary_panel.show()
            self.secondary_panel.hide()
            self.primary_panel.set_available_components(fft_available, fft_mode)
            self.primary_panel.label.setText(f"FFT (k-space) — {fft_mode} ▼")
            self.primary_panel.set_image(fft_img, levels=raw_levels)
        elif self.view_mode == "Original":
            self.primary_panel.show()
            self.secondary_panel.hide()
            self.primary_panel.set_available_components(original_available, original_mode)
            self.primary_panel.label.setText(f"Original Image — {original_mode} ▼")
            self.primary_panel.set_image(orig, levels=original_levels)
        else:
            self.primary_panel.show()
            self.secondary_panel.show()
            self.primary_panel.set_available_components(original_available, original_mode)
            self.secondary_panel.set_available_components(fft_available, fft_mode)
            self.primary_panel.label.setText(f"Original Image — {original_mode} ▼")
            self.secondary_panel.label.setText(f"FFT (k-space) — {fft_mode} ▼")
            self.primary_panel.set_image(orig, levels=original_levels)
            self.secondary_panel.set_image(fft_img, levels=raw_levels)

        rows, cols = orig.shape
        self.line_row = min(max(int(self.line_row_ratio * max(rows - 1, 1)), 0), rows - 1)
        self.line_col = min(max(int(self.line_col_ratio * max(cols - 1, 1)), 0), cols - 1)
        self._configure_line_control(rows, cols)
        self._sync_lines()
        self._sync_level_controls()
        self.update_line_profile()
        self._apply_image_orientation()
        self._apply_frequency_markers()
        self.stable_diagnostics.info(
            "REFRESH_IMAGES_END",
            state=self._stable_diagnostic_state(),
        )


    def _configure_line_control(self, rows, cols):
        maximum = rows-1 if self.line_orientation=='Row' else cols-1
        self.line_spin.blockSignals(True); self.line_slider.blockSignals(True)
        self.line_spin.setRange(0, maximum); self.line_slider.setRange(0, maximum)
        value=self.line_row if self.line_orientation=='Row' else self.line_col
        self.line_spin.setValue(value); self.line_slider.setValue(value)
        self.line_spin.blockSignals(False); self.line_slider.blockSignals(False)

    def _orientation_changed(self, text):
        self.line_orientation=text
        if self.current_image is not None: self._configure_line_control(*self.current_image.shape); self.update_line_profile()

    def _spin_line_changed(self, value):
        self._set_selected_line(value)

    def _slider_line_changed(self, value):
        self._set_selected_line(value)

    def _set_selected_line(self, value):
        if self.current_image is None: return
        panel, role, arr = self._active_panel_role_and_array()
        row, col = self._panel_crosshair_position(role, arr.shape)
        if self.line_orientation == 'Row': row = int(value)
        else: col = int(value)
        state = self.panel_crosshairs.setdefault(role, {})
        state['row_ratio'] = row / max(arr.shape[0] - 1, 1)
        state['col_ratio'] = col / max(arr.shape[1] - 1, 1)
        self.line_row, self.line_col = row, col
        self.line_spin.blockSignals(True); self.line_slider.blockSignals(True); self.line_spin.setValue(value); self.line_slider.setValue(value); self.line_spin.blockSignals(False); self.line_slider.blockSignals(False)
        panel.set_line_position(row, col)
        self.update_line_profile()

    def _panel_crosshair_position(self, role, shape):
        rows, cols = shape
        state = self.panel_crosshairs.setdefault(role, {"row_ratio": 0.5, "col_ratio": 0.5})
        row = min(max(int(state["row_ratio"] * max(rows - 1, 1)), 0), rows - 1)
        col = min(max(int(state["col_ratio"] * max(cols - 1, 1)), 0), cols - 1)
        return row, col

    def _line_moved(self, panel, row: int, col: int):
        role = self._panel_role(panel)
        array = self.current_kspace if role == "fft" else self.current_image
        if array is None:
            return
        rows, cols = np.asarray(array).shape
        state = self.panel_crosshairs.setdefault(role, {})
        state["row_ratio"] = int(row) / max(rows - 1, 1)
        state["col_ratio"] = int(col) / max(cols - 1, 1)
        if panel is getattr(self, "active_image_panel", None):
            self.line_row, self.line_col = int(row), int(col)
            self.line_row_ratio = state["row_ratio"]
            self.line_col_ratio = state["col_ratio"]
            self._configure_line_control(rows, cols)
        self.update_line_profile()

    def _sync_lines(self):
        if self.current_image is not None:
            row, col = self._panel_crosshair_position("original", np.asarray(self.current_image).shape)
            panel = self.primary_panel if self._panel_role(self.primary_panel) == "original" else self.secondary_panel
            panel.set_line_position(row, col)
        if self.current_kspace is not None:
            row, col = self._panel_crosshair_position("fft", np.asarray(self.current_kspace).shape)
            panel = self.primary_panel if self._panel_role(self.primary_panel) == "fft" else self.secondary_panel
            panel.set_line_position(row, col)

    def active_display_array(self):
        _panel, _role, array = self._active_panel_role_and_array()
        return array

    def _active_panel_role_and_array(self):
        panel = getattr(self, "active_image_panel", None)
        if panel is None or not panel.isVisible():
            panel = self.primary_panel if self.primary_panel.isVisible() else self.secondary_panel
        role = self._panel_role(panel)
        if role == "fft":
            mode = getattr(self, "fft_display_mode", "Magnitude")
            array = self._display_component(self.current_kspace, mode)
        else:
            mode = getattr(self, "original_display_mode", "Magnitude")
            array = self._display_component(self.current_image, mode)
        return panel, role, np.asarray(array)

    def _set_active_image_panel(self, panel):
        self.active_image_panel = panel
        for candidate in (self.primary_panel, self.secondary_panel):
            active = candidate is panel
            candidate.setStyleSheet(
                "ImagePanel { border: 2px solid #35a7ff; border-radius: 3px; }" if active
                else "ImagePanel { border: 1px solid transparent; }"
            )
        role = self._panel_role(panel)
        array = self.current_kspace if role == "fft" else self.current_image
        if array is not None:
            self.line_row, self.line_col = self._panel_crosshair_position(role, np.asarray(array).shape)
            self._configure_line_control(*np.asarray(array).shape)
        if hasattr(self, "level_target_combo"):
            self.level_target_combo.blockSignals(True)
            self.level_target_combo.setCurrentText("Raw Data" if role == "fft" else "Original Image")
            self.level_target_combo.blockSignals(False)
            self._sync_level_controls()
        self.update_line_profile()

    def _plot_panel_profiles(self, role, array, horizontal_plot, vertical_plot):
        if array is None:
            horizontal_plot.clear(); vertical_plot.clear(); return
        arr = np.asarray(array)
        if arr.ndim != 2 or arr.size == 0:
            horizontal_plot.clear(); vertical_plot.clear(); return
        row, col = self._panel_crosshair_position(role, arr.shape)
        horizontal = np.asarray(arr[row, :], dtype=float)
        vertical = np.asarray(arr[:, col], dtype=float)
        if self.log_profile.isChecked():
            horizontal = np.log10(np.maximum(np.abs(horizontal), 1e-12))
            vertical = np.log10(np.maximum(np.abs(vertical), 1e-12))
        horizontal_plot.clear(); horizontal_plot.plot(horizontal)
        vertical_plot.clear(); vertical_plot.plot(vertical)

    def update_line_profile(self):
        if self.current_image is None:
            return
        original = self._display_component(self.current_image, getattr(self, "original_display_mode", "Magnitude"))
        fft = None if self.current_kspace is None else self._display_component(self.current_kspace, getattr(self, "fft_display_mode", "Magnitude"))
        self._plot_panel_profiles("original", original, self.original_horizontal_profile, self.original_vertical_profile)
        self._plot_panel_profiles("fft", fft, self.fft_horizontal_profile, self.fft_vertical_profile)
        panel, role, arr = self._active_panel_role_and_array()
        row, col = self._panel_crosshair_position(role, arr.shape)
        self.line_value_label.setText(f"{role.title()} Row {row} / Col {col}")

    def send_line_to_signal_studio(self):
        if self.current_image is None: return
        arr=self.active_display_array(); line=arr[self.line_row,:] if self.line_orientation=='Row' else arr[:,self.line_col]
        self.add_signal(line, f"{self.line_orientation} {self.line_row if self.line_orientation=='Row' else self.line_col} — {self.view_mode}")
        self.tabs.setCurrentIndex(1)

    def auto_levels(self):
        if self.current_image is None:
            return
        panel, role, array = self._active_panel_role_and_array()
        target = "Raw Data" if role == "fft" else "Original Image"
        self._set_levels_for_target(target, array)
        if hasattr(self, "level_target_combo"):
            self.level_target_combo.blockSignals(True)
            self.level_target_combo.setCurrentText(target)
            self.level_target_combo.blockSignals(False)
        self.level_preset_combo.blockSignals(True)
        self.level_preset_combo.setCurrentText("Auto")
        self.level_preset_combo.blockSignals(False)
        self._sync_level_controls()
        self.refresh_images()
        self.statusBar().showMessage(f"Auto Levels applied to {target}")


    def _navigate_explorer_keyboard(self, delta: int):
        """Route Explorer Up/Down through Previous/Next continuous navigation."""
        moved = self.change_slice_continuous(delta)
        if not moved:
            self._update_series_navigation_ui()


    def _tree_current_changed(self, current, previous):
        try:
            self._open_tree_item(current, force=False)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Explorer Selection Error",
                f"{type(exc).__name__}: {exc}",
            )


    def _tree_original_paths(self):
        paths, seen = [], set()
        def add_path(value):
            try: path = Path(value)
            except Exception: return
            if path.exists() and path.is_file():
                key = str(path.resolve()).lower()
                if key not in seen:
                    seen.add(key); paths.append(path)
                if path.suffix.lower() in {".dat", ".pfile", ".7", ".img"}:
                    txt = path.with_suffix(".txt")
                    if txt.exists():
                        k = str(txt.resolve()).lower()
                        if k not in seen: seen.add(k); paths.append(txt)
        for item in self.tree.selectedItems():
            data=item.data(0,Qt.UserRole)
            if not data: continue
            kind=data[0]
            if kind=="dicom":
                i=int(data[1])
                if 0<=i<len(self.dicom_entries): add_path(self.dicom_entries[i].path)
            elif kind=="tracker_file":
                i=int(data[1])
                if 0<=i<len(self.tracker_files): add_path(self.tracker_files[i]["path"])
            elif kind in {"processed","tracker_workspace","raw_file","bitmap_pending","signal_pending"}:
                add_path(data[1])
            elif kind=="series":
                for j in range(item.childCount()):
                    d=item.child(j).data(0,Qt.UserRole)
                    if d and d[0]=="dicom":
                        i=int(d[1])
                        if 0<=i<len(self.dicom_entries): add_path(self.dicom_entries[i].path)
        return paths



    def _start_tree_file_drag(self, supported_actions):
        paths=self._tree_original_paths()
        if not paths:
            self.statusBar().showMessage("No original files are available for drag and drop")
            return
        mime=QMimeData(); mime.setUrls([QUrl.fromLocalFile(str(p)) for p in paths])
        drag=QDrag(self.tree); drag.setMimeData(mime)
        self.statusBar().showMessage(f"Dragging {len(paths)} original file(s)")
        drag.exec(Qt.CopyAction)

    def _open_tree_item(self, item, *, force=False):
        # Defensive initialization for sessions restored from an older build.
        if not hasattr(self, "_tree_open_in_progress"):
            self._tree_open_in_progress = False
        if not hasattr(self, "_active_tree_source_key"):
            self._active_tree_source_key = None

        if item is None or self._tree_open_in_progress:
            return

        data = item.data(0, Qt.UserRole)
        if not data:
            return
        self.stable_diagnostics.info(
            "TREE_ITEM_OPEN",
            text=item.text(0),
            data=data,
            force=bool(force),
        )

        source_type = str(data[0])
        source_value = data[1] if len(data) > 1 else None
        source_key = (source_type, str(source_value))

        if not force and source_key == self._active_tree_source_key:
            return

        self._tree_open_in_progress = True
        try:
            if source_type == "series":
                # Clicking a collapsed series expands it and immediately opens the
                # first image, matching the behavior users expect from a file tree.
                item.setExpanded(True)
                first_leaf = None
                stack = [item.child(i) for i in range(item.childCount())]
                while stack and first_leaf is None:
                    node = stack.pop(0)
                    node_data = node.data(0, Qt.UserRole)
                    if node_data and node_data[0] == "dicom":
                        first_leaf = node
                        break
                    stack[0:0] = [node.child(i) for i in range(node.childCount())]
                if first_leaf is not None:
                    self.tree.setCurrentItem(first_leaf)
                    self.tree.scrollToItem(first_leaf)
                    first_data = first_leaf.data(0, Qt.UserRole)
                    self.show_dicom(int(first_data[1]))
                    self._active_tree_source_key = ("dicom", str(first_data[1]))
                else:
                    self.statusBar().showMessage("The selected series contains no loadable images")

            elif source_type == "dicom":
                self.show_dicom(int(source_value))
                self.tabs.setCurrentIndex(0)

            elif source_type == "signal":
                self.tabs.setCurrentIndex(4)
                self.signal_combo.setCurrentIndex(int(source_value))
                self.update_signal_plot()

            elif source_type == "processed":
                self.load_processed_image(source_value)
                self.refresh_images()
                self.tabs.setCurrentIndex(0)

            elif source_type == "tracker_workspace":
                self.show_tracker_in_workspace()
                self.tabs.setCurrentIndex(0)

            elif source_type == "tracker_signal":
                self.send_tracker_to_signal_studio()

            elif source_type == "tracker_file":
                self._apply_tracker_state(int(source_value))
                self.tabs.setCurrentIndex(3)

            elif source_type == "tracker_pending":
                index = int(source_value)
                if 0 <= index < len(self.pending_tracker_paths):
                    path = self.pending_tracker_paths[index]
                    self.statusBar().showMessage(
                        f"Loading Tracker: {path.name}"
                    )
                    QApplication.processEvents()
                    self.load_tracker(path)
                    self.tabs.setCurrentIndex(3)

            elif source_type == "bitmap_pending":
                self.load_bitmap_image(Path(source_value))
                self.tabs.setCurrentIndex(0)

            elif source_type == "signal_pending":
                self.load_signal_file(Path(source_value))
                self.tabs.setCurrentIndex(4)

            elif source_type == "raw_file":
                path = Path(source_value)
                self.statusBar().showMessage(
                    f"Opening RAW: {path.name}"
                )
                QApplication.processEvents()
                self.load_raw_file(path)
                self.tabs.setCurrentIndex(0)

            self._active_tree_source_key = source_key

        except Exception as exc:
            self._active_tree_source_key = None
            QMessageBox.critical(
                self,
                "File Display Error",
                f"{source_type}: {source_value}\n\n"
                f"{type(exc).__name__}: {exc}",
            )
            self.statusBar().showMessage(
                f"Unable to display {source_type}: {exc}"
            )
        finally:
            self._tree_open_in_progress = False


    def _tree_clicked(self, item, column):
        try:
            self._open_tree_item(item, force=True)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Explorer Selection Error",
                f"{type(exc).__name__}: {exc}",
            )


    def _select_tree_dicom_index(self, index: int):
        """Select a DICOM image without losing its series expansion state."""
        item = self._find_tree_dicom_item(index)
        if item is None:
            return False
        parent = item.parent()
        ancestor = parent
        while ancestor is not None:
            ancestor.setExpanded(True)
            ancestor = ancestor.parent()
        self.tree.blockSignals(True)
        try:
            self.tree.setCurrentItem(item)
            item.setSelected(True)
        finally:
            self.tree.blockSignals(False)
        self.tree.scrollToItem(item)
        return bool(parent is None or parent.isExpanded())

    def _level_target_changed(self, target: str):
        self._sync_level_controls()
        self.refresh_images()

    def _set_levels_for_target(self, target: str, array: np.ndarray):
        finite = np.asarray(array, dtype=float)
        finite = finite[np.isfinite(finite)]
        if finite.size == 0:
            center, width = 0.5, 1.0
        else:
            low, high = np.percentile(finite, [1.0, 99.5])
            if high <= low:
                high = low + 1.0
            center = float((low + high) / 2.0)
            width = float(high - low)

        if target == "Raw Data":
            self.raw_window_level = center
            self.raw_dynamic_range = width
        else:
            self.original_window_level = center
            self.original_dynamic_range = width

    def _levels_for_target(self, target: str, array: np.ndarray, auto_original: bool = False):
        if target == "Original Image":
            if auto_original or self.original_window_level is None or self.original_dynamic_range is None:
                self._set_levels_for_target(target, array)
            center = float(self.original_window_level)
            width = max(float(self.original_dynamic_range), 1e-9)
        else:
            if self.raw_window_level is None or self.raw_dynamic_range is None:
                self._set_levels_for_target(target, array)
            center = float(self.raw_window_level)
            width = max(float(self.raw_dynamic_range), 1e-9)
        return center - width / 2.0, center + width / 2.0

    def _active_level_array(self) -> np.ndarray:
        if self.current_image is None or self.current_kspace is None:
            return np.array([0.0, 1.0])
        if self.view_mode == "Original":
            return np.abs(self.current_image)
        return np.log1p(np.abs(self.current_kspace))

    def _set_auto_levels_from_array(self, array: np.ndarray):
        finite = np.asarray(array, dtype=float)
        finite = finite[np.isfinite(finite)]
        if finite.size == 0:
            center, width = 0.5, 1.0
        else:
            low, high = np.percentile(finite, [1.0, 99.5])
            if high <= low:
                high = low + 1.0
            center = (low + high) / 2.0
            width = high - low
        self.window_level = float(center)
        self.dynamic_range = float(width)
        self._sync_level_controls()

    def _levels_for_array(self, array: np.ndarray):
        finite = np.asarray(array, dtype=float)
        finite = finite[np.isfinite(finite)]
        if finite.size == 0:
            return (0.0, 1.0)
        low, high = np.percentile(finite, [1.0, 99.5])
        if high <= low:
            high = low + 1.0
        return (float(low), float(high))

    def _current_levels(self):
        if self.window_level is None or self.dynamic_range is None:
            self._set_auto_levels_from_array(self._active_level_array())
        half = max(float(self.dynamic_range) / 2.0, 1e-9)
        return (float(self.window_level) - half, float(self.window_level) + half)

    def _sync_level_controls(self):
        if not hasattr(self, "window_level_spin"):
            return
        target = self.level_target_combo.currentText() if hasattr(self, "level_target_combo") else "Original Image"
        if target == "Raw Data":
            center = self.raw_window_level
            width = self.raw_dynamic_range
        else:
            center = self.original_window_level
            width = self.original_dynamic_range

        self.window_level_spin.blockSignals(True)
        self.dynamic_range_spin.blockSignals(True)
        self.window_level_spin.setValue(float(center or 0.0))
        self.dynamic_range_spin.setValue(max(float(width or 1.0), 1e-9))
        self.window_level_spin.blockSignals(False)
        self.dynamic_range_spin.blockSignals(False)


    def _manual_levels_changed(self):
        target = self.level_target_combo.currentText() if hasattr(self, "level_target_combo") else "Original Image"
        center = float(self.window_level_spin.value())
        width = max(float(self.dynamic_range_spin.value()), 1e-9)
        if target == "Raw Data":
            self.raw_window_level = center
            self.raw_dynamic_range = width
        else:
            self.original_window_level = center
            self.original_dynamic_range = width
        self.level_preset_combo.blockSignals(True)
        self.level_preset_combo.setCurrentText("Manual")
        self.level_preset_combo.blockSignals(False)
        self.refresh_images()


    def apply_level_preset(self, preset: str):
        if self.current_image is None:
            return
        array = self._active_level_array()
        finite = np.asarray(array, dtype=float)
        finite = finite[np.isfinite(finite)]
        if finite.size == 0:
            return

        percentiles = {
            "Auto": (1.0, 99.5),
            "Wide": (0.0, 100.0),
            "Soft Tissue": (5.0, 95.0),
            "High Contrast": (10.0, 90.0),
            "Narrow": (25.0, 75.0),
        }
        low_p, high_p = percentiles.get(preset, (1.0, 99.5))
        low, high = np.percentile(finite, [low_p, high_p])
        if high <= low:
            high = low + 1.0
        self.level_preset = preset
        self.window_level = float((low + high) / 2.0)
        self.dynamic_range = float(high - low)
        self._sync_level_controls()
        levels = self._current_levels()
        self.primary_panel.set_levels(*levels)
        if self.secondary_panel.isVisible():
            self.secondary_panel.set_levels(*levels)

    def _current_image_record(self):
        if self.current_image is None:
            return None
        name = Path(self.current_source).name if self.current_source else "current_image"
        return {
            "name": name,
            "path": Path(self.current_source) if self.current_source else None,
            "image": np.asarray(self.current_image, dtype=float).copy(),
            "ds": copy.deepcopy(self.current_ds) if self.current_ds is not None else None,
        }

    def set_addsub_source(self, slot: str):
        record = self._current_image_record()
        if record is None:
            QMessageBox.information(self, "Add/Subtract", "Load an image first.")
            return
        if slot == "A":
            self.addsub_a = record
        else:
            self.addsub_b = record
        self.addsub_status.setText(
            f"A: {self.addsub_a['name'] if self.addsub_a else '-'}\\n"
            f"B: {self.addsub_b['name'] if self.addsub_b else '-'}"
        )

    def run_addsub(self, operation: str):
        # Backward-compatible alias.
        self.preview_addsub(operation)

    def preview_addsub(self, operation: str):
        if self.addsub_a is None or self.addsub_b is None:
            QMessageBox.information(self, "Add/Subtract", "Set both image A and image B first.")
            return
        a = np.asarray(self.addsub_a["image"], dtype=float)
        b = np.asarray(self.addsub_b["image"], dtype=float)
        if a.shape != b.shape:
            QMessageBox.warning(self, "Add/Subtract", f"Image sizes do not match: {a.shape} and {b.shape}")
            return

        result = a + b if operation == "add" else a - b
        self.addsub_preview_result = result
        self.addsub_preview_operation = operation
        self.save_addsub_button.setEnabled(True)

        dialog = QDialog(self)
        dialog.setWindowTitle("Add / Subtract Preview")
        dialog.resize(1450, 620)
        layout = QVBoxLayout(dialog)
        panels = QSplitter(Qt.Horizontal)
        panel_a = ImagePanel("Image A")
        panel_b = ImagePanel("Image B")
        panel_result = ImagePanel("Result: A + B" if operation == "add" else "Result: A - B")
        panels.addWidget(panel_a)
        panels.addWidget(panel_b)
        panels.addWidget(panel_result)
        panel_a.set_image(a)
        panel_b.set_image(b)
        panel_result.set_image(result)
        layout.addWidget(panels, 1)

        buttons = QHBoxLayout()
        save_button = QPushButton("Save Result")
        close_button = QPushButton("Close Preview")
        save_button.clicked.connect(lambda: (self.save_addsub_preview(), dialog.accept()))
        close_button.clicked.connect(dialog.reject)
        buttons.addStretch(1)
        buttons.addWidget(save_button)
        buttons.addWidget(close_button)
        layout.addLayout(buttons)
        dialog.exec()

    def save_addsub_preview(self):
        if self.addsub_preview_result is None or not self.addsub_preview_operation:
            QMessageBox.information(
                self,
                "Add/Subtract",
                "Preview an Add/Subtract result first.",
            )
            return

        operation = self.addsub_preview_operation
        tag = "add" if operation == "add" else "sub"
        base = Path(self.addsub_a["name"]).stem if self.addsub_a else "image"

        previous_ds = self.current_ds
        previous_source = self.current_source
        try:
            if self.addsub_a is not None:
                self.current_ds = copy.deepcopy(self.addsub_a.get("ds"))
                source_path = self.addsub_a.get("path")
                if source_path:
                    self.current_source = str(source_path)

            output_path = self._save_processed_image(
                self.addsub_preview_result,
                f"{base}_{tag}_addsub",
                "_addsub",
            )

            if not output_path.exists():
                raise IOError(f"Saved output file was not created: {output_path}")

            npy_path = output_path.with_suffix(".npy")
            if output_path.suffix.lower() == ".dcm" and not npy_path.exists():
                raise IOError(f"Add/Subtract array file was not created: {npy_path}")

            self._display_processed_result(
                self.addsub_preview_result,
                output_path,
            )
            self.addsub_status.setText(
                f"A: {self.addsub_a['name'] if self.addsub_a else '-'}\n"
                f"B: {self.addsub_b['name'] if self.addsub_b else '-'}\n"
                f"Saved: {output_path}"
            )
            self.save_addsub_button.setEnabled(False)
            self.statusBar().showMessage(
                f"Add/Subtract saved: {output_path}"
            )
            QMessageBox.information(
                self,
                "Add/Subtract Saved",
                f"Saved successfully:\n{output_path}",
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Add/Subtract Save Error",
                str(exc),
            )
        finally:
            if self.current_ds is None:
                self.current_ds = previous_ds
            if not self.current_source:
                self.current_source = previous_source


    def _raw_compensation_panel(self):
        """Return the panel that is currently displaying FFT/raw k-space data.

        In Both mode the FFT is shown in ``secondary_panel``; in FFT-only mode it
        is shown in ``primary_panel``. Compensation ROI coordinates belong to
        k-space, so the ROI must never be attached to the Original Image panel.
        """
        for panel in (self.primary_panel, self.secondary_panel):
            if panel.isVisible() and self._panel_role(panel) == "fft":
                return panel
        return None

    def start_compensation_roi(self):
        if self.current_kspace is None:
            QMessageBox.information(self, "Compensation", "Load raw/k-space data first.")
            return

        if not self.compensation_history:
            self.compensation_history = [{
                "kspace": np.asarray(self.current_kspace).copy(),
                "image": np.asarray(self.current_image).copy(),
                "label": "Original",
                "roi": None,
                "level": None,
            }]
            self.compensation_history_index = 0

        self.compensation_original_kspace = np.asarray(self.current_kspace).copy()
        self.compensation_original = np.asarray(self.current_image).copy()
        self.compensation_preview = None
        self.roi_compensation_detection = None
        self.roi_compensation_detection_bounds = None

        # Compensation is performed in k-space. If the user is currently in
        # Original-only mode, reveal the FFT panel first instead of drawing the
        # ROI over the anatomical image.
        if self.view_mode == "Original":
            self.set_view_mode("Both")

        self.compensation_roi_panel = self._raw_compensation_panel()
        if self.compensation_roi_panel is None:
            QMessageBox.warning(
                self,
                "Compensation",
                "The FFT/raw-data panel is not available.",
            )
            return
        self.compensation_roi_panel.hide_comp_roi()
        self.manual_compensation_mask = np.zeros(np.asarray(self.current_kspace).shape, dtype=bool)
        self.compensation_roi_panel.manual_mask = self.manual_compensation_mask
        self.compensation_roi_panel.manualMaskChanged.connect(self.invalidate_compensation_detection, Qt.UniqueConnection)
        self.compensation_roi_panel.set_manual_mask_editing(
            True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value()
        )
        self._set_active_image_panel(self.compensation_roi_panel)
        self.detect_comp_button.setEnabled(True)
        self.preview_comp_button.setEnabled(False)
        self.apply_comp_button.setEnabled(False)
        self.comp_status_label.setText(
            "Manual RAW paint active. Paint only the area that should be compensated."
        )
        self.statusBar().showMessage("Paint the RAW-data area to compensate. Use Eraser to remove mask pixels.")

    def clear_compensation_roi(self):
        if self.compensation_roi_panel is not None:
            self.compensation_roi_panel.hide_comp_roi()
            self.compensation_roi_panel.set_manual_mask_editing(False)
            self.compensation_roi_panel.clear_manual_mask()
        if self.compensation_original_kspace is not None:
            self.current_kspace = self.compensation_original_kspace.copy()
            self.current_recon = ifft2c(self.current_kspace)
            self.current_image = np.abs(self.current_recon)
        self.compensation_preview = None
        self.roi_compensation_detection = None
        self.roi_compensation_detection_bounds = None
        self.compensation_original = None
        self.compensation_original_kspace = None
        self.compensation_roi_panel = None
        self.manual_compensation_mask = None
        self.detect_comp_button.setEnabled(False)
        self.preview_comp_button.setEnabled(False)
        self.apply_comp_button.setEnabled(False)
        self.comp_status_label.setText("No RAW ROI selected")
        self.refresh_images()

    def _update_compensation_session_status(self, *args, source_label=None):
        """Show the authoritative mask, parameters, and preview state in one place."""
        if not hasattr(self, "comp_session_status_label"):
            return
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        mask = None if panel is None else getattr(panel, "manual_mask", None)
        pixels = int(np.count_nonzero(mask)) if mask is not None else 0
        regions = 0
        if mask is not None and pixels:
            a = np.asarray(mask, dtype=bool)
            seen = np.zeros(a.shape, dtype=bool)
            h, w = a.shape
            for y in range(h):
                for x in range(w):
                    if not a[y, x] or seen[y, x]:
                        continue
                    regions += 1
                    stack = [(y, x)]; seen[y, x] = True
                    while stack:
                        py, px = stack.pop()
                        for ny in range(max(0, py-1), min(h, py+2)):
                            for nx in range(max(0, px-1), min(w, px+2)):
                                if a[ny, nx] and not seen[ny, nx]:
                                    seen[ny, nx] = True; stack.append((ny, nx))
        if source_label is None:
            if getattr(self, "compensation_preview", None) is not None:
                source_label = "Current reconstructed result"
            elif getattr(self, "comp_auto_result", None) is not None:
                source_label = "Auto Correct mask / original RAW"
            else:
                source_label = "Manual session"
        preview_state = "current" if getattr(self, "compensation_preview", None) is not None else "needs recalculation"
        self.comp_session_status_label.setText(
            f"Session: {source_label} | Mask {regions} region(s), {pixels} pixels | "
            f"Removal {self.comp_quick_removal.value()} / Detail {self.comp_quick_detail.value()} / "
            f"Protection {self.comp_quick_protection.value()} | Preview {preview_state}"
        )

    def _toggle_auto_mask_visibility(self, visible):
        panel = self.compensation_roi_panel
        if panel is None:
            return
        if visible:
            panel._refresh_manual_mask_overlay()
        elif hasattr(panel, "manual_mask_overlay") and panel.manual_mask_overlay is not None:
            panel.manual_mask_overlay.setVisible(False)

    def open_quick_adjust(self):
        self.comp_tabs.setCurrentIndex(0)
        if hasattr(self.comp_quick_accordion, "set_expanded"):
            self.comp_quick_accordion.set_expanded(True, True)
        elif hasattr(self.comp_quick_accordion, "button") and not self.comp_quick_accordion.content.isVisible():
            self.comp_quick_accordion.button.click()
        self.comp_quick_recalculate_button.setFocus()

    def _enter_manual_edit_from_auto_result(self, tab_index):
        """Enter the same editable state as Start Manual while retaining Auto Correct's mask.

        The corrected preview is replaced by the original RAW/k-space display so the
        generated mask can be inspected and edited in context.  The existing mask is
        deliberately preserved instead of being cleared as Start Manual normally does.
        """
        if self.current_kspace is None:
            QMessageBox.information(self, "Compensation", "Load raw/k-space data first.")
            return False
        source = (
            np.asarray(self.compensation_original_kspace).copy()
            if self.compensation_original_kspace is not None
            else np.asarray(self.current_kspace).copy()
        )
        existing_mask = None
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        if panel is not None and getattr(panel, "manual_mask", None) is not None:
            candidate = np.asarray(panel.manual_mask, dtype=bool)
            if candidate.shape == source.shape:
                existing_mask = candidate.copy()
        if existing_mask is None and getattr(self, "comp_auto_result", None) is not None:
            candidate = np.asarray(self.comp_auto_result.mask, dtype=bool)
            if candidate.shape == source.shape:
                existing_mask = candidate.copy()
        if existing_mask is None:
            existing_mask = np.zeros(source.shape, dtype=bool)

        if not self.compensation_history:
            original_image = np.abs(ifft2c(source))
            self.compensation_history = [{
                "kspace": source.copy(), "image": original_image.copy(),
                "label": "Original", "roi": None, "level": None,
            }]
            self.compensation_history_index = 0

        self.compensation_original_kspace = source.copy()
        self.compensation_original = np.abs(ifft2c(source))
        self.current_kspace = source.copy()
        self.current_recon = ifft2c(self.current_kspace)
        self.current_image = np.abs(self.current_recon)
        self.compensation_preview = None
        self.roi_compensation_detection = None
        self.roi_compensation_detection_bounds = None

        if self.view_mode == "Original":
            self.set_view_mode("Both")
        self.refresh_images()
        panel = self._raw_compensation_panel() or panel
        if panel is None:
            QMessageBox.warning(self, "Compensation", "The FFT/raw-data panel is not available.")
            return False
        self.compensation_roi_panel = panel
        panel.hide_comp_roi()
        panel.manual_mask = existing_mask
        self.manual_compensation_mask = panel.manual_mask
        try:
            panel.manualMaskChanged.connect(self.invalidate_compensation_detection, Qt.UniqueConnection)
        except (TypeError, RuntimeError):
            pass
        panel._refresh_manual_mask_overlay()
        self._toggle_auto_mask_visibility(True)
        panel.set_manual_mask_editing(
            True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value()
        )
        self.comp_show_auto_mask_check.setChecked(True)
        self._set_active_image_panel(panel)
        has_mask = bool(np.any(existing_mask))
        self.detect_comp_button.setEnabled(True)
        self.preview_comp_button.setEnabled(has_mask)
        self.apply_comp_button.setEnabled(False)
        self.comp_tabs.setCurrentIndex(tab_index)
        self._update_compensation_session_status(source_label="Editable original RAW with retained mask")
        return True

    def open_paint_after_auto_correct(self):
        if self._enter_manual_edit_from_auto_result(1):
            self.comp_status_label.setText(
                "Manual RAW paint active with the Auto Correct mask. Edit the existing mask, then use Painted Mask or Preview Reconstructed Image."
            )
            self.statusBar().showMessage("Auto Correct mask retained. Paint or erase the RAW-data mask.")

    def open_expert_after_auto_correct(self):
        if self._enter_manual_edit_from_auto_result(2):
            self.comp_status_label.setText(
                "Expert mode active with the Auto Correct mask. Edit the existing mask and advanced settings, then preview the reconstructed image."
            )
            self.statusBar().showMessage("Auto Correct mask retained. Expert settings apply to this editable RAW-data mask.")

    def review_current_edited_mask(self):
        """Reconstruct and review directly from the current Paint/Expert state.

        Preview is an output, not a prerequisite. Every click rebuilds exactly once
        from the current manual mask and the currently visible Expert settings, then
        opens the before/after comparison without requiring a trip through Paint.
        """
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        has_mask = bool(
            panel is not None
            and getattr(panel, "manual_mask", None) is not None
            and np.any(panel.manual_mask)
        )
        if not has_mask:
            QMessageBox.information(
                self,
                "Review Reconstructed Image",
                "Paint or retain at least one RAW-data mask area first.",
            )
            return
        self.preview_compensation(open_comparison=False)
        if self.compensation_preview is None:
            return
        self.review_reconstructed_image()

    def review_reconstructed_image(self):
        if self.compensation_after_image is None or self.compensation_before_image is None:
            QMessageBox.information(self, "Review Reconstructed Image", "Run Auto Correct or Recalculate first.")
            return
        self.open_comp_comparison_button.setEnabled(True)
        self.open_compensation_comparison()
        self.comp_status_label.setText("Review the reconstructed image. Apply the result when the image and residual artifact are acceptable.")

    def _install_auto_result(self, result, source, status_text, *, store_as_auto_result=False):
        """Install a reconstructed result while preserving the original RAW baseline.

        Auto Correct stores its result as the restore point. Quick Adjust installs only
        the newly calculated preview and must not replace that Auto Correct restore point.
        Every later calculation therefore starts from the same unmodified RAW/k-space.
        """
        source = np.asarray(source).copy()
        if self.compensation_original_kspace is None:
            self.compensation_original_kspace = source.copy()
            self.compensation_original = np.abs(ifft2c(source))
        if store_as_auto_result:
            self.comp_auto_result = result
            self.comp_auto_original_mask = np.asarray(result.mask, dtype=bool).copy()
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        if panel is not None:
            self.compensation_roi_panel = panel
            panel.manual_mask = result.mask.copy()
            self.manual_compensation_mask = panel.manual_mask
            panel._refresh_manual_mask_overlay()
            self._toggle_auto_mask_visibility(self.comp_show_auto_mask_check.isChecked())
            panel.set_manual_mask_editing(True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value())
        self.compensation_before_kspace = np.asarray(source).copy()
        self.compensation_before_image = np.abs(ifft2c(source))
        self.compensation_after_image = result.image.copy()
        # Commit0121: Auto Correct results are immediately reviewable even when
        # the closed Quick Action panel was the previous route to Review.
        before_complex = ifft2c(source)
        after_complex = ifft2c(result.kspace)
        self.compensation_difference_image = np.abs(after_complex) - np.abs(before_complex)
        self.compensation_difference_fft = np.log1p(np.abs(result.kspace)) - np.log1p(np.abs(source))
        self.compensation_difference_phase = np.angle(after_complex) - np.angle(before_complex)
        self.compensation_preview = result.kspace.copy()
        self.current_kspace = result.kspace.copy()
        self.current_recon = ifft2c(self.current_kspace)
        self.current_image = np.abs(self.current_recon)
        self.refresh_images()
        has_mask = bool(np.any(result.mask))
        self.preview_comp_button.setEnabled(has_mask)
        self.apply_comp_button.setEnabled(has_mask)
        self.open_comp_comparison_button.setEnabled(True)
        self._set_auto_result_mode(True)
        self.comp_status_label.setText(status_text)
        self._update_compensation_session_status(
            source_label="Auto Correct result" if store_as_auto_result else "Quick Adjust result"
        )

    def _set_auto_result_mode(self, completed):
        """Switch between initial Auto Correct mode and post-result actions."""
        completed = bool(completed)
        self.comp_auto_correct_button.setVisible(not completed)
        self.comp_auto_more_button.setVisible(completed)
        if completed:
            self.comp_next_step_group.setVisible(True)
        else:
            self.comp_next_step_group.setVisible(False)

    def run_auto_correct_again(self):
        """Return to the initial state and explicitly rerun Auto Correct."""
        self._set_auto_result_mode(False)
        self.comp_auto_quality_label.setText("Auto result: recalculating...")
        self.auto_correct_compensation()

    def recalculate_quick_adjust_once(self):
        """Apply Quick Adjust exactly once.

        With an existing mask, reconstruct from Current Mask + current Quick Adjust
        values.  When the mask is empty, run one candidate-detection trial using
        those same values, install any detected mask, and reconstruct once.  The
        six-trial Auto Retry / Best Candidate workflow is never used here.
        """
        if self.current_kspace is None:
            QMessageBox.information(self, "Quick Adjust", "Load raw/k-space data first.")
            return
        source = (
            np.asarray(self.compensation_original_kspace).copy()
            if self.compensation_original_kspace is not None
            else np.asarray(self.current_kspace).copy()
        )
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        active_mask = None if panel is None else getattr(panel, "manual_mask", None)
        has_mask = bool(
            active_mask is not None
            and np.asarray(active_mask).shape == np.asarray(source).shape
            and np.any(active_mask)
        )
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            if has_mask:
                result = recalculate_with_mask(
                    source, np.asarray(active_mask, dtype=bool),
                    artifact_type=self.comp_mode_combo.currentText(),
                    threshold_sigma=self.comp_threshold_spin.value(),
                    sensitivity=self.comp_sensitivity_combo.currentText(),
                    removal=self.comp_quick_removal.value(),
                    detail=self.comp_quick_detail.value(),
                    protection=self.comp_quick_protection.value(),
                    target_ratio=self.comp_target_spin.value(),
                )
                mode_text = "Current Mask"
                status_text = (
                    "Quick Adjust recalculated once from the current mask and selected values. "
                    "Review the reconstructed image."
                )
            else:
                # Single trial only: use the visible Quick Adjust values to detect,
                # validate, compensate, and evaluate once.  Do not call Auto Retry.
                result = run_auto_correct(
                    source,
                    threshold_sigma=self.comp_threshold_spin.value(),
                    sensitivity=self.comp_sensitivity_combo.currentText(),
                    removal=self.comp_quick_removal.value(),
                    detail=self.comp_quick_detail.value(),
                    protection=self.comp_quick_protection.value(),
                    target_ratio=self.comp_target_spin.value(),
                )
                if not np.any(result.mask):
                    self.comp_auto_quality_label.setText(
                        "Quick Adjust single detection found no acceptable candidate. "
                        "Change the sliders and try again, or use Paint / Expert."
                    )
                    self.comp_status_label.setText(
                        "No mask was detected with the current Quick Adjust values. "
                        "Quick Adjust remains available for another single trial."
                    )
                    self._update_compensation_session_status(
                        source_label="Quick Adjust single trial: no candidate"
                    )
                    QMessageBox.information(
                        self,
                        "Quick Adjust",
                        "No reliable artifact mask was detected with the current settings.\n\n"
                        "Adjust Artifact Removal, Image Detail, or Protection and try again, "
                        "or continue with Paint / Expert.",
                    )
                    return
                mode_text = "Single Detection"
                status_text = (
                    "Quick Adjust ran one detection trial with the selected values, "
                    "created a mask, and reconstructed the image."
                )

            self._install_auto_result(
                result, source, status_text, store_as_auto_result=False,
            )
            m = result.metrics
            self.comp_auto_quality_label.setText(
                f"Quick Adjust {mode_text} | Mask pixels {int(m.get('mask_pixels',0))} | "
                f"Quality {m.get('overall_quality',0):.1f} | "
                f"Artifact reduction {m.get('artifact_reduction',0):.1f}% | "
                f"Detail {m.get('detail_preservation',100):.1f}% | "
                f"Image change {m.get('outside_image_change',0):.1f}%"
            )
            self.review_reconstructed_image()
        except Exception as exc:
            QMessageBox.critical(self, "Quick Adjust Error", str(exc))
        finally:
            QApplication.restoreOverrideCursor()

    def restore_auto_compensation_result(self):
        if getattr(self, "comp_auto_result", None) is None:
            QMessageBox.information(self, "Auto Correct", "Run Auto Correct first.")
            return
        result = self.comp_auto_result
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        restore_mask = np.asarray(
            getattr(self, "comp_auto_original_mask", result.mask), dtype=bool
        ).copy()
        if panel is not None:
            panel.manual_mask = restore_mask
            self.manual_compensation_mask = panel.manual_mask
            panel._refresh_manual_mask_overlay()
        self.current_kspace = result.kspace.copy()
        self.current_recon = ifft2c(self.current_kspace)
        self.current_image = np.abs(self.current_recon)
        self.refresh_images()
        self.preview_comp_button.setEnabled(bool(np.any(result.mask)))
        self.apply_comp_button.setEnabled(bool(np.any(result.mask)))

    def _show_auto_correct_unreliable_dialog(self):
        box = QMessageBox(self)
        box.setWindowTitle("Auto Correct")
        box.setIcon(QMessageBox.Warning)
        box.setText("Auto Correct could not detect\na reliable artifact.")
        box.setInformativeText(
            "Possible reasons\n\n"
            "• Image is already clean\n"
            "• Artifact is too weak\n"
            "• Unsupported artifact"
        )
        manual_button = box.addButton("Manual Paint", QMessageBox.ActionRole)
        expert_button = box.addButton("Expert Settings", QMessageBox.ActionRole)
        box.addButton("Close", QMessageBox.RejectRole)
        box.exec()
        clicked = box.clickedButton()
        if clicked is manual_button:
            self.comp_tabs.setCurrentIndex(1)
            self.start_compensation_roi()
        elif clicked is expert_button:
            self.comp_tabs.setCurrentIndex(2)

    def auto_correct_compensation(self):
        if self.current_kspace is None:
            QMessageBox.information(self, "Auto Correct", "Load raw/k-space data first.")
            return
        progress = QProgressDialog("Searching Artifacts...", "Cancel", 0, 6, self)
        progress.setWindowTitle("Auto Correct Progress")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setMinimumWidth(520)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()
        try:
            source = self.compensation_original_kspace if self.compensation_original_kspace is not None else np.asarray(self.current_kspace)

            def update_progress(trial, total, phase, settings, record):
                if phase == "searching":
                    progress.setLabelText(
                        f"Searching Artifacts...\n\nTrial {trial} / {total}\n"
                        f"Artifact Removal {settings['removal']} | Image Detail {settings['detail']} | Protection {settings['protection']}"
                    )
                    progress.setValue(trial - 1)
                else:
                    quality = float((record or {}).get("quality", 0.0))
                    progress.setLabelText(
                        f"Evaluating Quality...\n\nTrial {trial} / {total}\nQuality {quality:.1f}"
                    )
                    progress.setValue(trial)
                QApplication.processEvents()

            result, trial_log = auto_correct_with_retry(
                source,
                threshold_sigma=self.comp_threshold_spin.value(),
                sensitivity=self.comp_sensitivity_combo.currentText(),
                target_ratio=self.comp_target_spin.value(),
                quality_threshold=60.0,
                progress_callback=update_progress,
                cancel_callback=progress.wasCanceled,
            )
            if progress.wasCanceled():
                progress.close()
                return
            self.comp_last_candidate_result = result
            self._install_auto_result(
                result,
                source,
                "Auto Correct completed. Review the reconstructed image, or choose Quick Adjust, Paint, or Expert.",
                store_as_auto_result=True,
            )
            accepted = ", ".join(result.selected_types) if result.selected_types else "none (original retained)"
            m = result.metrics
            best_trial = next((x for x in trial_log if x.get("quality") == m.get("overall_quality")), None)
            trial_text = f"Trial {best_trial['trial']} / 6 | " if best_trial else ""
            self.comp_auto_quality_label.setText(
                f"{trial_text}Selected: {accepted} | Quality {m.get('overall_quality',0):.1f} | "
                f"Artifact reduction {m.get('artifact_reduction',0):.1f}% | "
                f"Detail {m.get('detail_preservation',100):.1f}% | "
                f"Image change {m.get('outside_image_change',0):.1f}%"
            )
            reliable = bool(result.selected_types and np.any(result.mask) and m.get("overall_quality", 0.0) >= 60.0)
            if reliable:
                self.comp_status_label.setText("Auto Correct completed. Review the reconstructed image, or choose Quick Adjust, Paint, or Expert.")
            else:
                self.comp_status_label.setText("Auto Correct did not find a reliable result. Paint or Expert may be used.")
            progress.setLabelText("Auto Correct completed")
            progress.setValue(6)
            QApplication.processEvents()
            progress.close()
            if reliable:
                self.review_reconstructed_image()
            else:
                self._show_auto_correct_unreliable_dialog()
        except Exception as exc:
            progress.close()
            QMessageBox.critical(self, "Auto Correct Error", str(exc))

    def invalidate_compensation_detection(self):
        """Invalidate a prior preview while keeping an edited mask reconstructable.

        After "2. Use Painted Mask", Brush/Eraser/Remove Component emits the mask
        changed signal.  The previous implementation disabled Preview entirely,
        leaving no route to reconstruct the edited mask.  Preview now remains
        available whenever the current painted mask still contains pixels.
        """
        self.roi_compensation_detection = None
        self.roi_compensation_detection_bounds = None
        self.compensation_preview = None
        panel = self.compensation_roi_panel
        has_mask = bool(panel is not None and panel.manual_mask is not None and np.any(panel.manual_mask))
        if hasattr(self, "preview_comp_button"):
            self.preview_comp_button.setEnabled(has_mask)
        if hasattr(self, "apply_comp_button"):
            self.apply_comp_button.setEnabled(False)
        if hasattr(self, "detect_comp_button"):
            self.detect_comp_button.setEnabled(panel is not None)
        if has_mask and hasattr(self, "comp_status_label"):
            self.comp_status_label.setText(
                "Mask edited. Preview Reconstructed Image will rebuild from the current painted mask."
            )
        self._update_compensation_session_status(source_label="Edited current mask")

    def _update_manual_paint_tool(self, *args):
        if self.compensation_roi_panel is not None:
            self.compensation_roi_panel.set_manual_mask_editing(
                True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value()
            )

    def _current_compensation_mask(self):
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        if panel is None or getattr(panel, "manual_mask", None) is None:
            return panel, None
        return panel, np.asarray(panel.manual_mask, dtype=bool)

    def _set_compensation_mask(self, mask, *, remember=True, source_label="Edited current mask"):
        panel, current = self._current_compensation_mask()
        if panel is None:
            return False
        if remember and current is not None:
            self._comp_mask_undo_stack.append(current.copy())
            self._comp_mask_undo_stack = self._comp_mask_undo_stack[-30:]
            self._comp_mask_redo_stack.clear()
        panel.manual_mask = np.asarray(mask, dtype=bool).copy()
        self.manual_compensation_mask = panel.manual_mask
        panel._refresh_manual_mask_overlay()
        panel.set_manual_mask_editing(True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value())
        self.invalidate_compensation_detection()
        self._update_compensation_session_status(source_label=source_label)
        return True

    def undo_compensation_mask(self):
        panel, current = self._current_compensation_mask()
        if panel is None or not self._comp_mask_undo_stack:
            return
        if current is not None:
            self._comp_mask_redo_stack.append(current.copy())
        self._set_compensation_mask(self._comp_mask_undo_stack.pop(), remember=False, source_label="Mask undo")

    def redo_compensation_mask(self):
        panel, current = self._current_compensation_mask()
        if panel is None or not self._comp_mask_redo_stack:
            return
        if current is not None:
            self._comp_mask_undo_stack.append(current.copy())
        self._set_compensation_mask(self._comp_mask_redo_stack.pop(), remember=False, source_label="Mask redo")

    @staticmethod
    def _binary_dilate(mask):
        m = np.asarray(mask, dtype=bool)
        out = m.copy()
        for dy in (-1, 0, 1):
            for dx in (-1, 0, 1):
                out |= np.roll(np.roll(m, dy, axis=0), dx, axis=1)
        return out

    @staticmethod
    def _binary_erode(mask):
        m = np.asarray(mask, dtype=bool)
        out = m.copy()
        for dy in (-1, 0, 1):
            for dx in (-1, 0, 1):
                out &= np.roll(np.roll(m, dy, axis=0), dx, axis=1)
        return out

    def morph_compensation_mask(self, direction):
        panel, mask = self._current_compensation_mask()
        if mask is None or not np.any(mask):
            return
        new_mask = self._binary_dilate(mask) if direction > 0 else self._binary_erode(mask)
        self._set_compensation_mask(new_mask, source_label="Mask expanded" if direction > 0 else "Mask shrunk")

    def symmetrize_compensation_mask(self):
        panel, mask = self._current_compensation_mask()
        if mask is None or not np.any(mask):
            return
        symmetric = mask | np.flip(np.flip(mask, axis=0), axis=1)
        self._set_compensation_mask(symmetric, source_label="Hermitian symmetry paint")

    @staticmethod
    def _mask_components(mask):
        m = np.asarray(mask, dtype=bool)
        seen = np.zeros_like(m, dtype=bool)
        components = []
        h, w = m.shape
        for y, x in np.argwhere(m):
            y, x = int(y), int(x)
            if seen[y, x]:
                continue
            stack = [(y, x)]; seen[y, x] = True; pts = []
            while stack:
                cy, cx = stack.pop(); pts.append((cy, cx))
                for dy, dx in ((-1,0),(1,0),(0,-1),(0,1)):
                    ny, nx = cy+dy, cx+dx
                    if 0 <= ny < h and 0 <= nx < w and m[ny, nx] and not seen[ny, nx]:
                        seen[ny, nx] = True; stack.append((ny, nx))
            components.append(pts)
        return components

    def fill_largest_compensation_region(self):
        panel, mask = self._current_compensation_mask()
        if mask is None or not np.any(mask):
            return
        comps = self._mask_components(mask)
        if not comps:
            return
        pts = max(comps, key=len)
        ys = [p[0] for p in pts]; xs = [p[1] for p in pts]
        new_mask = mask.copy(); new_mask[min(ys):max(ys)+1, min(xs):max(xs)+1] = True
        self._set_compensation_mask(new_mask, source_label="Largest region filled")

    def delete_smallest_compensation_region(self):
        panel, mask = self._current_compensation_mask()
        if mask is None or not np.any(mask):
            return
        comps = self._mask_components(mask)
        if not comps:
            return
        new_mask = mask.copy()
        for y, x in min(comps, key=len):
            new_mask[y, x] = False
        self._set_compensation_mask(new_mask, source_label="Smallest region deleted")

    def run_expert_auto_mask_once(self):
        if self.current_kspace is None:
            QMessageBox.information(self, "Expert Auto Mask", "Load raw/k-space data first.")
            return
        source = np.asarray(self.compensation_original_kspace if self.compensation_original_kspace is not None else self.current_kspace).copy()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            result = run_auto_correct(
                source, threshold_sigma=self.comp_threshold_spin.value(),
                sensitivity=self.comp_sensitivity_combo.currentText(),
                removal=self.comp_quick_removal.value(), detail=self.comp_quick_detail.value(),
                protection=self.comp_quick_protection.value(), target_ratio=self.comp_target_spin.value(),
            )
        finally:
            QApplication.restoreOverrideCursor()
        self.comp_last_candidate_result = result
        if not np.any(result.mask):
            QMessageBox.information(self, "Expert Auto Mask", "No reliable mask was detected with the current Expert settings.")
            self.show_auto_candidate_viewer()
            return
        panel, current = self._current_compensation_mask()
        merge = False
        if current is not None and np.any(current):
            box = QMessageBox(self)
            box.setWindowTitle("Expert Auto Mask")
            box.setText("A current mask already exists. Replace it or merge the new Auto Mask?")
            replace_btn = box.addButton("Replace Current Mask", QMessageBox.AcceptRole)
            merge_btn = box.addButton("Merge With Current Mask", QMessageBox.ActionRole)
            box.addButton(QMessageBox.Cancel)
            box.exec()
            if box.clickedButton() is merge_btn:
                merge = True
            elif box.clickedButton() is not replace_btn:
                return
        new_mask = result.mask if not merge or current is None else (current | result.mask)
        self._set_compensation_mask(new_mask, source_label="Expert Auto Mask: merged" if merge else "Expert Auto Mask: replaced")
        self.comp_auto_quality_label.setText(f"Expert Auto Mask | {len(result.selected_types)} accepted types | {int(np.count_nonzero(new_mask))} pixels | Quality {result.metrics.get('overall_quality',0):.1f}")

    def show_auto_candidate_viewer(self):
        """Interactive product-level candidate viewer (Commit0113)."""
        result = getattr(self, "comp_last_candidate_result", None) or getattr(self, "comp_auto_result", None)
        if result is None:
            QMessageBox.information(self, "Candidate Viewer", "Run Auto Correct or Expert Auto Mask first.")
            return
        candidates = list(getattr(result, "candidates", []) or [])
        if not candidates:
            QMessageBox.information(self, "Candidate Viewer", "No candidates were produced.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Auto Correct Candidate Viewer")
        dialog.resize(1120, 720)
        root = QVBoxLayout(dialog)
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Show"))
        filter_combo = QComboBox()
        filter_combo.addItems(["All", "Accepted only", "Rejected only"])
        filter_row.addWidget(filter_combo)
        filter_row.addStretch(1)
        root.addLayout(filter_row)

        split = QSplitter(Qt.Horizontal)
        table = QTableWidget(0, 7)
        table.setHorizontalHeaderLabels(["#", "Type", "Status", "Score", "Confidence", "Coverage", "Reason"])
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setMinimumWidth(560)
        split.addWidget(table)

        preview = QWidget()
        pv = QVBoxLayout(preview)
        title = QLabel("Select a candidate")
        title.setStyleSheet("font-size:16px;font-weight:700")
        pv.addWidget(title)
        image_split = QSplitter(Qt.Horizontal)
        fft_view = pg.ImageView(); fft_view.ui.roiBtn.hide(); fft_view.ui.menuBtn.hide()
        mask_view = pg.ImageView(); mask_view.ui.roiBtn.hide(); mask_view.ui.menuBtn.hide()
        image_split.addWidget(fft_view); image_split.addWidget(mask_view)
        pv.addWidget(image_split, 1)
        details = QTextEdit(); details.setReadOnly(True); details.setMaximumHeight(170)
        pv.addWidget(details)
        actions = QHBoxLayout()
        apply_btn = QPushButton("Apply This Candidate")
        paint_btn = QPushButton("Apply and Continue in Paint")
        close_btn = QPushButton("Close")
        actions.addWidget(apply_btn); actions.addWidget(paint_btn); actions.addStretch(1); actions.addWidget(close_btn)
        pv.addLayout(actions)
        split.addWidget(preview); split.setStretchFactor(1, 1)
        root.addWidget(split, 1)

        shown = []
        def reason_text(c):
            reason = c.get("reason")
            if reason and reason != "-": return str(reason)
            if c.get("accepted", False): return "Accepted by validation"
            reasons=[]
            if c.get("quality_gain", 0) < 0: reasons.append("quality did not improve")
            if c.get("artifact_reduction", 0) <= 4: reasons.append("artifact reduction too small")
            if c.get("outside_image_change", 0) > 8: reasons.append("image change too large")
            return ", ".join(reasons) or "validation threshold not met"

        def rebuild():
            mode=filter_combo.currentText(); table.setRowCount(0); shown.clear()
            for idx,c in enumerate(candidates):
                accepted=bool(c.get("accepted",False))
                if mode=="Accepted only" and not accepted: continue
                if mode=="Rejected only" and accepted: continue
                row=table.rowCount(); table.insertRow(row); shown.append((idx,c))
                vals=[str(idx+1), str(c.get("type","Unknown")), "Accepted" if accepted else "Rejected",
                      f"{c.get('candidate_score',c.get('quality_gain',0)):.1f}",
                      f"{100*c.get('confidence',0):.1f}%", f"{100*c.get('coverage',0):.3f}%", reason_text(c)]
                for col,val in enumerate(vals): table.setItem(row,col,QTableWidgetItem(val))
            table.resizeColumnsToContents()
            if table.rowCount(): table.selectRow(0); update_preview()

        def selected_candidate():
            row=table.currentRow()
            return shown[row][1] if 0 <= row < len(shown) else None

        def update_preview():
            c=selected_candidate()
            if c is None: return
            title.setText(f"Candidate {candidates.index(c)+1:02d} — {c.get('type','Unknown')} — {'Accepted' if c.get('accepted') else 'Rejected'}")
            source = self.compensation_original_kspace if self.compensation_original_kspace is not None else self.current_kspace
            if source is not None: fft_view.setImage(np.log1p(np.abs(np.asarray(source))), autoLevels=True)
            mask=np.asarray(c.get("mask", np.zeros(np.asarray(source).shape if source is not None else (2,2))),dtype=float)
            mask_view.setImage(mask, autoLevels=True)
            details.setPlainText(
                f"Score: {c.get('candidate_score',c.get('quality_gain',0)):.2f}\n"
                f"Confidence: {100*c.get('confidence',0):.2f}%\nCoverage: {100*c.get('coverage',0):.4f}%\n"
                f"Artifact reduction: {c.get('artifact_reduction',0):.2f}%\n"
                f"Detail preservation: {c.get('detail_preservation',0):.2f}%\n"
                f"Outside image change: {c.get('outside_image_change',0):.2f}%\n"
                f"Edge support: {100*c.get('edge_fraction',0):.1f}%\n"
                f"Hermitian pair support: {100*c.get('pair_support',0):.1f}%\n"
                f"Centre-axis occupancy: {100*c.get('axis_fraction',0):.2f}%\n"
                f"Decision: {reason_text(c)}")

        def apply_candidate(go_paint=False):
            c=selected_candidate()
            if c is None: return
            mask=np.asarray(c.get("mask", []),dtype=bool)
            if mask.size == 0 or not np.any(mask):
                QMessageBox.information(dialog,"Candidate Viewer","This candidate has no usable mask."); return
            self._set_compensation_mask(mask, source_label=f"Candidate Viewer: {c.get('type','Unknown')}")
            self.comp_last_candidate_result = result
            if go_paint:
                dialog.accept(); self.open_paint_after_auto_correct()
            else:
                QMessageBox.information(dialog,"Candidate Viewer","The candidate mask is now the current mask. You can Review, Paint, or Expert-adjust it.")

        filter_combo.currentTextChanged.connect(rebuild)
        table.itemSelectionChanged.connect(update_preview)
        apply_btn.clicked.connect(lambda: apply_candidate(False))
        paint_btn.clicked.connect(lambda: apply_candidate(True))
        close_btn.clicked.connect(dialog.accept)
        rebuild()
        dialog.exec()

    def clear_manual_compensation_mask(self):
        panel = self.compensation_roi_panel
        if panel is not None:
            panel.clear_manual_mask()
            self.manual_compensation_mask = panel.manual_mask
        else:
            self.manual_compensation_mask = None
        self.invalidate_compensation_detection()
        self.roi_compensation_detection = None
        self.roi_compensation_detection_bounds = None
        self.compensation_preview = None
        self.compensation_difference_fft = None
        self.compensation_difference_phase = None
        self.compensation_difference_image = None
        self.compensation_metrics = {}
        self.open_comp_comparison_button.setEnabled(False)
        self.preview_comp_button.setEnabled(False)
        self.apply_comp_button.setEnabled(False)
        self.detect_comp_button.setEnabled(panel is not None)
        self.comp_status_label.setText("Manual RAW mask cleared. Paint a new area, then select Use Painted Mask.")
        self.statusBar().showMessage("RAW paint cleared")

    def _rebuild_compensation_detection_from_current_mask(self):
        """Build the compensation model from the mask exactly as currently edited."""
        if self.current_kspace is None:
            raise ValueError("Load raw/k-space data first.")
        source = (
            self.compensation_original_kspace
            if self.compensation_original_kspace is not None
            else np.asarray(self.current_kspace)
        )
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        if panel is None or panel.manual_mask is None or not np.any(panel.manual_mask):
            raise ValueError("Paint or generate at least one RAW mask pixel first.")
        profile = self._manual_compensation_profile()
        selected_mask_expand = profile["mask_expand"] if self.comp_mask_expansion_auto.isChecked() else self.comp_mask_expansion_spin.value()
        selected_donor_halo = profile["donor_halo"] if self.comp_donor_halo_auto.isChecked() else self.comp_donor_halo_spin.value()
        selected_passes = profile["model_passes"] if self.comp_pass_count_auto.isChecked() else self.comp_pass_count_spin.value()
        detection, bounds = build_manual_mask_detection(
            source,
            np.asarray(panel.manual_mask, dtype=bool),
            target_ratio=self.comp_target_spin.value(),
            mask_expand=selected_mask_expand,
            donor_halo=selected_donor_halo,
            model_passes=selected_passes,
            stripe_suppression=profile["stripe_suppression"],
            edge_blend=profile["edge_blend"],
        )
        self.manual_compensation_mask = panel.manual_mask
        self.roi_compensation_detection = detection
        self.roi_compensation_detection_bounds = bounds
        self.compensation_preview = None
        self.preview_comp_button.setEnabled(detection.stats.get("changed", 0) > 0)
        self.apply_comp_button.setEnabled(False)
        return detection, bounds

    def detect_compensation_mask(self):
        if self.current_kspace is None:
            QMessageBox.information(self, "ROI Raw Compensation", "Load raw/k-space data first.")
            return
        try:
            source = (
                self.compensation_original_kspace
                if self.compensation_original_kspace is not None
                else np.asarray(self.current_kspace)
            )
            panel = self.compensation_roi_panel or self._raw_compensation_panel()
            if panel is None or panel.manual_mask is None:
                raise ValueError("Start Manual Paint and paint a RAW-data area first.")
            requested_mode = self.comp_mode_combo.currentText()
            auto_detection = None
            if requested_mode == "Manual Only":
                if not np.any(panel.manual_mask):
                    raise ValueError(
                        "Manual Only mode requires a painted mask. Use Brush, Line, Band, Block or Ring first."
                    )
            elif requested_mode == "Auto":
                auto_detection = hybrid_detect_artifacts(
                    source, "Auto", self.comp_threshold_spin.value(),
                    self.comp_sensitivity_combo.currentText(),
                )
                if not np.any(auto_detection.mask):
                    raise ValueError(
                        "Auto Detection did not find a sufficiently confident artifact outside the protected MRI centre. "
                        "Paint the target manually or lower the threshold."
                    )
                existing = np.asarray(panel.manual_mask, dtype=bool)
                merge_mode = self.comp_auto_mask_merge_combo.currentText()
                if merge_mode == "Add to Paint" and np.any(existing):
                    panel.manual_mask = existing | auto_detection.mask
                else:
                    panel.manual_mask = auto_detection.mask.copy()
                self.manual_compensation_mask = panel.manual_mask
                panel._refresh_manual_mask_overlay()
                panel.set_manual_mask_editing(
                    True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value()
                )
            elif not np.any(panel.manual_mask):
                typed_detection = hybrid_detect_artifacts(
                    source, requested_mode, self.comp_threshold_spin.value(),
                    self.comp_sensitivity_combo.currentText(),
                )
                if not np.any(typed_detection.mask):
                    raise ValueError(f"No {requested_mode} candidate was detected outside the protected MRI centre. Paint the target manually or lower the threshold.")
                panel.manual_mask = typed_detection.mask.copy()
                self.manual_compensation_mask = panel.manual_mask
                panel._refresh_manual_mask_overlay()
                panel.set_manual_mask_editing(
                    True, self.comp_paint_tool_combo.currentText(), self.comp_brush_size_spin.value()
                )
                auto_detection = typed_detection
            detection, bounds = self._rebuild_compensation_detection_from_current_mask()
            y0, y1, x0, x1 = bounds
            detection_summary = ""
            if auto_detection is not None:
                confidence_parts = auto_detection.stats.get("confidences", {})
                ranked = sorted(confidence_parts.items(), key=lambda item: item[1], reverse=True)[:3]
                confidence_text = ", ".join(f"{name.title()} {value*100:.0f}%" for name, value in ranked)
                detection_summary = (
                    f" | Auto={auto_detection.artifact_type} | confidence={auto_detection.confidence*100:.0f}%"
                    f" | direction={auto_detection.direction} | candidates: {confidence_text} | editable with Eraser/Remove Component"
                )
            self.comp_status_label.setText(
                f"Mask ready: {detection.mode} | pixels={detection.stats.get('changed', 0)} | "
                f"background={detection.background:.4g} | threshold={self.comp_threshold_spin.value():.2f} σ"
                f"{detection_summary}."
            )
        except Exception as exc:
            QMessageBox.critical(self, "ROI Mask Detection Error", str(exc))


    def _set_advanced_controls_enabled(self, enabled=True):
        """Keep expert controls editable; presets populate values but never lock them."""
        for control in (
            self.comp_sensitivity_combo, self.comp_mask_expansion_auto, self.comp_mask_expansion_spin,
            self.comp_donor_halo_auto, self.comp_donor_halo_spin, self.comp_pass_count_auto, self.comp_pass_count_spin,
            self.comp_strength_override_auto, self.comp_strength_override_spin, self.comp_structure_preservation_spin,
            self.comp_frequency_aware_check, self.comp_poisson_check,
            self.comp_hermitian_check,
        ):
            control.setEnabled(bool(enabled))

    def _apply_compensation_tuning_preset(self, preset):
        """Populate a tested preset. Expert preserves the current values."""
        if self._updating_comp_preset:
            return
        self._updating_comp_preset = True
        try:
            self._set_advanced_controls_enabled(True)
            values = {
                "Conservative": ("Conservative", True, 1, True, 3, True, 2, True, 0.75, 0.88, True, True, True),
                "Balanced": ("Balanced", False, 1, False, 3, False, 2, False, 0.75, 0.72, True, True, True),
                "Aggressive": ("Sensitive", False, 3, False, 5, False, 3, False, 0.95, 0.50, True, True, True),
            }.get(str(preset))
            if values is not None:
                controls = (
                    self.comp_sensitivity_combo,
                    self.comp_mask_expansion_auto, self.comp_mask_expansion_spin,
                    self.comp_donor_halo_auto, self.comp_donor_halo_spin,
                    self.comp_pass_count_auto, self.comp_pass_count_spin,
                    self.comp_strength_override_auto, self.comp_strength_override_spin,
                    self.comp_structure_preservation_spin,
                    self.comp_frequency_aware_check, self.comp_poisson_check,
                    self.comp_hermitian_check,
                )
                for control in controls:
                    control.blockSignals(True)
                try:
                    sensitivity, exp_auto, expansion, halo_auto, halo, passes_auto, passes, strength_auto, strength, preserve, freq, poisson, hermitian = values
                    self.comp_sensitivity_combo.setCurrentText(sensitivity)
                    self.comp_mask_expansion_auto.setChecked(exp_auto)
                    self.comp_mask_expansion_spin.setValue(expansion)
                    self.comp_donor_halo_auto.setChecked(halo_auto)
                    self.comp_donor_halo_spin.setValue(halo)
                    self.comp_pass_count_auto.setChecked(passes_auto)
                    self.comp_pass_count_spin.setValue(passes)
                    self.comp_strength_override_auto.setChecked(strength_auto)
                    self.comp_strength_override_spin.setValue(strength)
                    self.comp_structure_preservation_spin.setValue(preserve)
                    self.comp_frequency_aware_check.setChecked(freq)
                    self.comp_poisson_check.setChecked(poisson)
                    self.comp_hermitian_check.setChecked(hermitian)
                    self.comp_mask_expansion_spin.setEnabled(not exp_auto)
                    self.comp_donor_halo_spin.setEnabled(not halo_auto)
                    self.comp_pass_count_spin.setEnabled(not passes_auto)
                    self.comp_strength_override_spin.setEnabled(not strength_auto)
                finally:
                    for control in controls:
                        control.blockSignals(False)
            self.invalidate_compensation_detection()
            if hasattr(self, "comp_status_label"):
                self.comp_status_label.setText(
                    f"Advanced tuning preset: {preset}. Preview uses the current painted mask and these settings."
                )
        finally:
            self._updating_comp_preset = False

    def _advanced_compensation_value_changed(self, *args):
        """Switch to Expert and invalidate only the old preview, not the painted mask."""
        if self._updating_comp_preset:
            return
        self._updating_comp_preset = True
        try:
            self.comp_tuning_preset_combo.blockSignals(True)
            self.comp_tuning_preset_combo.setCurrentText("Expert")
            self.comp_tuning_preset_combo.blockSignals(False)
        finally:
            self._updating_comp_preset = False
        self.invalidate_compensation_detection()
        if hasattr(self, "comp_status_label"):
            self.comp_status_label.setText(
                "Advanced setting changed. Preview Reconstructed Image applies it to the current mask."
            )

    def reset_advanced_compensation_settings(self):
        """Restore the safe editable defaults for compensation controls."""
        self.comp_tuning_preset_combo.setCurrentText("Conservative")
        self._apply_compensation_tuning_preset("Conservative")

    def _manual_compensation_profile(self):
        # High is intentionally aggressive: it expands beyond the painted core
        # and takes donor values from farther away so residual line/blob energy
        # immediately surrounding the brush is also removed.
        return {
            "Low": {"strength": 0.40, "mask_expand": 0, "donor_halo": 1, "model_passes": 1, "stripe_suppression": 0.0, "edge_blend": False},
            "Mid": {"strength": 0.75, "mask_expand": 1, "donor_halo": 2, "model_passes": 2, "stripe_suppression": 0.20, "edge_blend": True},
            # Commit0115: High/Extreme previously over-expanded broad edge-blob
            # masks and could reinforce a centre line.  Keep stronger blending,
            # but limit geometric expansion and stripe suppression.
            "High": {"strength": 0.92, "mask_expand": 2, "donor_halo": 4, "model_passes": 2, "stripe_suppression": 0.35, "edge_blend": True},
            "Extreme": {"strength": 0.98, "mask_expand": 3, "donor_halo": 5, "model_passes": 3, "stripe_suppression": 0.55, "edge_blend": True},
        }.get(
            self.comp_level_combo.currentText(),
            {"strength": 0.75, "mask_expand": 1, "donor_halo": 2, "model_passes": 2, "stripe_suppression": 0.20, "edge_blend": True},
        )

    def _compensation_strength(self):
        return float(self._manual_compensation_profile()["strength"])

    def _compensation_bounds(self):
        if self.roi_compensation_detection_bounds is not None:
            return self.roi_compensation_detection_bounds
        panel = self.compensation_roi_panel or self._raw_compensation_panel()
        if panel is None or panel.manual_mask is None or not np.any(panel.manual_mask):
            raise ValueError("Paint at least one RAW-data pixel first.")
        ys, xs = np.where(panel.manual_mask)
        return max(1, int(ys.min()) - 1), min(panel.manual_mask.shape[0] - 1, int(ys.max()) + 2), max(1, int(xs.min()) - 1), min(panel.manual_mask.shape[1] - 1, int(xs.max()) + 2)

    def preview_compensation(self, open_comparison=True):
        if self.current_kspace is None:
            QMessageBox.information(self, "Compensation", "Load raw/k-space data first.")
            return
        try:
            source_kspace = (
                self.compensation_original_kspace
                if self.compensation_original_kspace is not None
                else np.asarray(self.current_kspace)
            )
            # Rebuild directly from the currently edited painted mask.  Do not
            # rerun Auto Detection here: doing so replaced additions/erasures
            # made after "2. Use Painted Mask" and could prevent reconstruction.
            self._rebuild_compensation_detection_from_current_mask()
            if self.roi_compensation_detection is None:
                return
            y0, y1, x0, x1 = self.roi_compensation_detection_bounds
            panel = self.compensation_roi_panel or self._raw_compensation_panel()
            active_mask = None if panel is None else panel.manual_mask
            hybrid = hybrid_compensate(
                source_kspace,
                active_mask,
                artifact_type=self.comp_mode_combo.currentText(),
                level=self.comp_level_combo.currentText(),
                threshold_sigma=self.comp_threshold_spin.value(),
                target_ratio=self.comp_target_spin.value(),
                adaptive_direction=True,
                frequency_aware=self.comp_frequency_aware_check.isChecked(),
                harmonic_poisson=self.comp_poisson_check.isChecked(),
                multi_pass=True,
                hermitian_symmetry=self.comp_hermitian_check.isChecked(),
                mask_expansion=None if self.comp_mask_expansion_auto.isChecked() else self.comp_mask_expansion_spin.value(),
                donor_halo=None if self.comp_donor_halo_auto.isChecked() else self.comp_donor_halo_spin.value(),
                pass_count=None if self.comp_pass_count_auto.isChecked() else self.comp_pass_count_spin.value(),
                strength_override=None if self.comp_strength_override_auto.isChecked() else self.comp_strength_override_spin.value(),
                structure_preservation=self.comp_structure_preservation_spin.value(),
                detection_sensitivity=self.comp_sensitivity_combo.currentText(),
            )
            compensated_kspace = hybrid.kspace
            compensation_stats = dict(hybrid.metrics)
            compensation_stats.update({"background": self.roi_compensation_detection.background, "changed": int(hybrid.metrics.get("changed_pixels", 0)), "max_abs_delta": float(np.max(hybrid.difference_fft)) if hybrid.difference_fft.size else 0.0})
            reconstructed = ifft2c(compensated_kspace)
            self.compensation_before_kspace = np.asarray(source_kspace).copy()
            self.compensation_before_image = np.abs(ifft2c(source_kspace))
            self.compensation_after_image = np.abs(reconstructed)
            self.compensation_difference_fft = hybrid.difference_fft
            self.compensation_difference_phase = hybrid.difference_phase
            self.compensation_difference_image = hybrid.difference_image
            self.compensation_metrics = dict(hybrid.metrics)

            self.compensation_preview = compensated_kspace
            previous_raw_window_level = self.raw_window_level
            previous_raw_dynamic_range = self.raw_dynamic_range
            self.current_kspace = compensated_kspace
            self.current_recon = reconstructed
            self.current_image = np.abs(reconstructed)

            # Keep the same RAW display scale so Before/After intensity changes
            # remain visually apparent instead of being hidden by auto-rescaling.
            self._fft_level_signature = (tuple(np.asarray(self.current_kspace).shape), float(np.nanmax(np.abs(self.current_kspace))) if np.asarray(self.current_kspace).size else 0.0)
            self.raw_window_level = previous_raw_window_level
            self.raw_dynamic_range = previous_raw_dynamic_range
            self.refresh_images()
            self.apply_comp_button.setEnabled(True)
            self.open_comp_comparison_button.setEnabled(True)
            if open_comparison:
                self.open_compensation_comparison()

            original_roi = np.abs(source_kspace[y0:y1, x0:x1])
            new_roi = np.abs(compensated_kspace[y0:y1, x0:x1])
            change = np.abs(new_roi - original_roi)
            self.comp_status_label.setText(
                f"Preview: {self.comp_mode_combo.currentText()} / "
                f"{self.comp_level_combo.currentText()} | "
                f"changed={compensation_stats['changed']} | "
                                f"background={compensation_stats['background']:.4g} | "
                f"max {compensation_stats.get('before_max', 0.0):.4g} → "
                f"{compensation_stats.get('after_max', 0.0):.4g} | "
                f"painted mean {compensation_stats.get('before_mean', 0.0):.4g} → "
                f"{compensation_stats.get('after_mean', 0.0):.4g} | "
                f"max |Δraw|={compensation_stats.get('max_abs_delta', 0.0):.4g} | "
                f"Artifact Reduction Score={compensation_stats.get('artifact_reduction_score', 0.0):.1f}/100 | "
                f"Preserve={self.comp_structure_preservation_spin.value():.2f} | "
                f"Sensitivity={self.comp_sensitivity_combo.currentText()}"
            )
        except Exception as exc:
            QMessageBox.critical(self, "Compensation Preview Error", str(exc))

    def open_compensation_comparison(self):
        required = (
            getattr(self, "compensation_before_image", None),
            getattr(self, "compensation_after_image", None),
            self.compensation_difference_image,
            self.compensation_difference_fft,
            self.compensation_difference_phase,
        )
        if any(value is None for value in required):
            QMessageBox.information(self, "Compensation Comparison", "Preview compensation first.")
            return
        existing = getattr(self, "compensation_comparison_dialog", None)
        if existing is not None:
            try:
                existing.close()
                existing.deleteLater()
            except Exception:
                pass
        dialog = CompensationComparisonDialog(*required, parent=self)
        self.compensation_comparison_dialog = dialog
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        dialog.destroyed.connect(lambda: setattr(self, "compensation_comparison_dialog", None))
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def apply_compensation(self):
        try:
            y0, y1, x0, x1 = self._compensation_bounds()
            if self.compensation_preview is None:
                self.preview_compensation()
            if self.compensation_preview is None:
                return

            state = {
                "kspace": np.asarray(self.current_kspace).copy(),
                "image": np.asarray(self.current_image).copy(),
                "label": (
                    f"{self.comp_mode_combo.currentText()} / "
                    f"{self.comp_level_combo.currentText()} ROI {x0}:{x1},{y0}:{y1}"
                ),
                "roi": (y0, y1, x0, x1),
                "level": self.comp_level_combo.currentText(),
                "mode": self.comp_mode_combo.currentText(),
                "threshold_sigma": self.comp_threshold_spin.value(),
                "target_ratio": self.comp_target_spin.value(),
                "metrics": dict(self.compensation_metrics),
                "difference_fft": None if self.compensation_difference_fft is None else self.compensation_difference_fft.copy(),
                "difference_phase": None if self.compensation_difference_phase is None else self.compensation_difference_phase.copy(),
                "difference_image": None if self.compensation_difference_image is None else self.compensation_difference_image.copy(),
            }

            if self.compensation_history_index < len(self.compensation_history) - 1:
                self.compensation_history = self.compensation_history[
                    : self.compensation_history_index + 1
                ]
            self.compensation_history.append(state)
            self.compensation_history_index = len(self.compensation_history) - 1

            if self.compensation_roi_panel is not None:
                self.compensation_roi_panel.hide_comp_roi()
                self.compensation_roi_panel.set_manual_mask_editing(False)
            self.compensation_roi_panel = None
            self.compensation_preview = None
            self.roi_compensation_detection = None
            self.roi_compensation_detection_bounds = None
            self.compensation_original_kspace = None
            self.compensation_original = None
            self.detect_comp_button.setEnabled(False)
            self.preview_comp_button.setEnabled(False)
            self.apply_comp_button.setEnabled(False)
            self.save_comp_button.setEnabled(True)
            self._update_compensation_history_buttons()
            self.comp_status_label.setText(
                f"Committed {self.compensation_history_index}/"
                f"{len(self.compensation_history)-1}: {state['label']}. "
                f"Select another ROI to continue."
            )
            self.refresh_images()
        except Exception as exc:
            QMessageBox.critical(self, "Compensation Error", str(exc))

    def _update_compensation_history_buttons(self):
        self.comp_prev_button.setEnabled(self.compensation_history_index > 0)
        self.comp_next_button.setEnabled(
            0 <= self.compensation_history_index < len(self.compensation_history) - 1
        )
        self.save_comp_button.setEnabled(self.compensation_history_index >= 0)

    def navigate_compensation_history(self, step: int):
        target = self.compensation_history_index + int(step)
        if target < 0 or target >= len(self.compensation_history):
            return
        self.compensation_history_index = target
        state = self.compensation_history[target]
        self.current_kspace = np.asarray(state["kspace"]).copy()
        self.current_recon = ifft2c(self.current_kspace)
        self.current_image = np.abs(self.current_recon)
        self.compensation_preview = None
        self.compensation_original_kspace = None
        if self.compensation_roi_panel is not None:
            self.compensation_roi_panel.hide_comp_roi()
        self.compensation_roi_panel = None
        self.refresh_images()
        self._update_compensation_history_buttons()
        self.comp_status_label.setText(
            f"History {target}/{len(self.compensation_history)-1}: {state['label']}"
        )

    def save_compensation_history_state(self):
        if not self.compensation_history or self.compensation_history_index < 0:
            QMessageBox.information(self, "Compensation", "No compensation state is available.")
            return
        state = self.compensation_history[self.compensation_history_index]
        base = Path(self.current_source).stem if self.current_source else "raw"
        output_path = self._save_processed_image(
            np.asarray(state["image"]),
            f"{base}_comp_h{self.compensation_history_index}",
            "_comp",
        )
        # Save raw/k-space separately for exact restoration.
        np.save(
            output_path.parent / f"{Path(output_path).stem}_raw.npy",
            np.asarray(state["kspace"]),
        )
        self.comp_status_label.setText(
            f"Saved displayed state {self.compensation_history_index}: {output_path.name}"
        )
        self.statusBar().showMessage(f"Compensation state saved: {output_path}")

    def change_output_root(self):
        initial = str(self._output_root())
        selected = QFileDialog.getExistingDirectory(self, "Select Output Top Folder", initial)
        if not selected:
            return
        self.output_root_override = Path(selected)
        self._update_output_root_label()

    def _output_root(self) -> Path:
        if self.output_root_override is not None:
            return self.output_root_override
        if self.last_import_output_root is not None:
            return self.last_import_output_root
        source = Path(self.current_source) if self.current_source else Path.cwd() / "image"
        parent = source.parent if source.parent.exists() else Path.cwd()
        return parent / "MR_Image_Explorer_Output"


    def _update_output_root_label(self):
        if hasattr(self, "output_root_label"):
            self.output_root_label.setText(f"Output: {self._output_root()}")

    def _work_folder(self, category: str = "Processed") -> Path:
        folder = self._output_root() / category
        folder.mkdir(parents=True, exist_ok=True)
        self._update_output_root_label()
        return folder


    def _save_processed_image(self, image: np.ndarray, stem: str, suffix: str) -> Path:
        suffix_lower = suffix.lower()
        if "addsub" in suffix_lower:
            category = "AddSub"
        elif "comp" in suffix_lower:
            category = "Compensated"
        elif "ns" in suffix_lower or "spike" in suffix_lower:
            category = "NoSpike"
        else:
            category = "Processed"

        folder = self._work_folder(category)
        safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem)
        candidate_stem = safe_stem
        counter = 1
        while (
            (folder / f"{candidate_stem}.npy").exists()
            or (folder / f"{candidate_stem}.dcm").exists()
        ):
            candidate_stem = f"{safe_stem}_{counter}"
            counter += 1
        safe_stem = candidate_stem

        npy_path = folder / f"{safe_stem}.npy"
        np.save(npy_path, np.asarray(image, dtype=np.float32))

        if self.current_ds is not None:
            ds = copy.deepcopy(self.current_ds)
            arr = np.asarray(image, dtype=float)
            finite = arr[np.isfinite(arr)]
            low = float(np.min(finite)) if finite.size else 0.0
            high = float(np.max(finite)) if finite.size else 1.0
            if high <= low:
                high = low + 1.0
            scaled = np.clip((arr - low) / (high - low) * 65535.0, 0, 65535).astype(np.uint16)
            ds.PixelData = scaled.tobytes()
            ds.Rows, ds.Columns = scaled.shape
            ds.BitsAllocated = 16
            ds.BitsStored = 16
            ds.HighBit = 15
            ds.PixelRepresentation = 0
            ds.RescaleSlope = (high - low) / 65535.0
            ds.RescaleIntercept = low
            ds.SOPInstanceUID = generate_uid()
            ds.SeriesInstanceUID = generate_uid()
            ds.ImageType = ["DERIVED", "SECONDARY"]
            original_description = str(getattr(ds, "SeriesDescription", "MR"))
            ds.SeriesDescription = f"{original_description} {suffix}"
            dcm_path = folder / f"{safe_stem}.dcm"
            ds.save_as(str(dcm_path), write_like_original=False)
            primary_path = dcm_path
        else:
            primary_path = npy_path

        self.processed_images[str(primary_path)] = np.asarray(image, dtype=float).copy()
        self.processed_sources[str(primary_path)] = primary_path
        self._add_processed_tree_item(primary_path)
        self._update_output_root_label()
        return primary_path


    def _add_processed_tree_item(self, path: Path):
        root = None
        for i in range(self.tree.topLevelItemCount()):
            candidate = self.tree.topLevelItem(i)
            if candidate.text(0) == "Processed Files":
                root = candidate
                break
        if root is None:
            root = QTreeWidgetItem(["Processed Files"])
            self.tree.addTopLevelItem(root)
        item = QTreeWidgetItem([path.name])
        item.setData(0, Qt.UserRole, ("processed", str(path)))
        root.addChild(item)
        root.setExpanded(True)
        self.tree.setCurrentItem(item)

    def _display_processed_result(self, image: np.ndarray, path: Path):
        self.current_image = np.asarray(image, dtype=float)
        self.current_kspace = fft2c(self.current_image)
        self.current_recon = ifft2c(self.current_kspace)
        self.current_source = str(path)
        self.current_ds = None
        self.source_kind = "processed"
        self.original_window_level = None
        self.original_dynamic_range = None
        self.raw_window_level = None
        self.raw_dynamic_range = None
        self.view_mode = "Both"
        self._update_output_root_label()
        self.refresh_images()


    def load_processed_image(self, path_value: str):
        path = Path(path_value)
        if path.suffix.lower() in BITMAP_EXTENSIONS:
            self.load_bitmap_image(path)
            return

        if str(path) in self.processed_images:
            image = self.processed_images[str(path)]
        elif path.suffix.lower() == ".npy":
            image = np.load(path, allow_pickle=False)
        elif path.suffix.lower() == ".dcm":
            entry = self.read_dicom(path)
            image = entry.image
            self.current_ds = entry.ds
        else:
            return

        self.current_image = np.asarray(image, dtype=float)
        self.current_kspace = fft2c(self.current_image)
        self.current_recon = ifft2c(self.current_kspace)
        self.current_source = str(path)
        self.source_kind = "processed"
        self.original_window_level = None
        self.original_dynamic_range = None
        self.raw_window_level = None
        self.raw_dynamic_range = None
        self.view_mode = "Both"
        self._update_output_root_label()
        self.refresh_images()


    def _artifact_detection_indices(self):
        self._tree_selection_changed()
        if self.detect_selected_only.isChecked() and self.artifact_selected_indices:
            return list(self.artifact_selected_indices)
        return list(range(len(self.dicom_entries)))

    def detect_artifacts_from_db(self):
        if not self._ensure_artifact_database():
            return

        keys, training, labels, metadata = self.artifact_db.training_feature_vectors()
        indices = self._artifact_detection_indices()
        if not indices:
            QMessageBox.information(self, "Artifact Detection", "No DICOM images are available.")
            return

        progress = self._make_progress("Artifact Detection", len(indices))
        results = []
        try:
            for count, index in enumerate(indices, start=1):
                if progress.wasCanceled():
                    break
                entry = self.dicom_entries[index]
                progress.setLabelText(f"Classifying image: {entry.path.name}")
                features = image_features(entry.image, self.normal_reference_image)
                vector = features_to_vector(features, keys)
                result = classify_feature_vector(
                    vector,
                    training,
                    labels,
                    minimum_samples_per_class=self.detect_min_samples.value(),
                )
                results.append({
                    "index": index,
                    "entry": entry,
                    "features": features,
                    "result": result,
                })
                progress.setValue(count)
        finally:
            progress.close()

        self._artifact_detection_results = results
        self.artifact_detection_table.setRowCount(len(results))
        predicted_counts = {}
        for row, item in enumerate(results):
            entry = item["entry"]
            result = item["result"]
            predicted_counts[result.label] = predicted_counts.get(result.label, 0) + 1
            alternatives = ", ".join(
                f"{label} {confidence * 100:.1f}%"
                for label, confidence, distance, support in result.alternatives[1:4]
            )
            values = [
                entry.path.name,
                str(getattr(entry.ds, "SeriesDescription", "") or getattr(entry.ds, "ProtocolName", "")),
                result.label,
                f"{result.confidence * 100:.1f}%",
                f"{result.distance:.4f}" if np.isfinite(result.distance) else "-",
                result.support,
                result.status,
                alternatives,
                str(entry.path),
            ]
            for column, value in enumerate(values):
                table_item = QTableWidgetItem(str(value))
                table_item.setData(Qt.UserRole, row)
                self.artifact_detection_table.setItem(row, column, table_item)

        self.artifact_detection_table.resizeColumnsToContents()
        summary = ", ".join(f"{label}: {count}" for label, count in sorted(predicted_counts.items()))
        self.artifact_detection_summary.setText(
            f"Detected {len(results)} image(s). {summary or 'No prediction results.'}"
        )

    def open_detected_artifact_image(self, row: int, column: int):
        if not hasattr(self, "_artifact_detection_results"):
            return
        if row < 0 or row >= len(self._artifact_detection_results):
            return
        result = self._artifact_detection_results[row]
        self.show_dicom(int(result["index"]))
        self.tabs.setCurrentIndex(0)

    def send_detected_rows_to_learning(self):
        if not hasattr(self, "_artifact_detection_results"):
            return
        selected_rows = sorted(
            {item.row() for item in self.artifact_detection_table.selectedItems()}
        )
        if not selected_rows:
            return
        indices = []
        predicted_labels = []
        for row in selected_rows:
            result = self._artifact_detection_results[row]
            indices.append(int(result["index"]))
            predicted_labels.append(result["result"].label)
        self.artifact_selected_indices = sorted(set(indices))
        if predicted_labels and len(set(predicted_labels)) == 1:
            self.artifact_type_combo.setCurrentText(predicted_labels[0])
        self.artifact_selection_label.setText(
            f"Training images ready from detection: {len(self.artifact_selected_indices)}"
        )
        self.tabs.setCurrentIndex(2)

    def clear_selected_images(self):
        selected=[]
        for item in self.tree.selectedItems():
            data=item.data(0,Qt.UserRole)
            if data and data[0]=="dicom": selected.append(int(data[1]))
        selected=sorted(set(selected),reverse=True)
        if not selected: return
        if QMessageBox.question(self,"Clear Selected Images",f"Remove {len(selected)} selected image(s)?") != QMessageBox.Yes: return
        for i in selected:
            if 0 <= i < len(self.dicom_entries): self.dicom_entries.pop(i)
        if self.dicom_entries: self.populate_dicom_tree(); self.show_dicom(min(getattr(self,"slice_index",0),len(self.dicom_entries)-1))
        else: self._reset_image_state()

    def clear_all_images(self):
        if QMessageBox.question(self,"Clear All Images","Clear all imported images and image editing state?") != QMessageBox.Yes: return
        self.dicom_entries.clear(); self.lazy_dicom_cache.clear(); self.lazy_dicom_cache_order = []; self._reset_image_state()

    def _reset_image_state(self):
        self.current_image=None; self.current_kspace=None; self.current_recon=None; self.current_ds=None; self.current_source=""; self.tree.clear(); self.primary_panel.image_item.clear(); self.secondary_panel.image_item.clear(); self.profile_plot.clear(); self.slice_label.setText("Slice: -"); self.info.setText("No file loaded")
        self.compensation_history=[]; self.compensation_history_index=-1; self.addsub_a=None; self.addsub_b=None; self.addsub_preview_result=None; self.addsub_status.setText("A: -\nB: -")

    def remove_selected_signal(self):
        i=self.signal_combo.currentIndex()
        if 0 <= i < len(self.signals): self.signals.pop(i); self.signal_combo.clear(); self.signal_combo.addItems([x["name"] for x in self.signals]); self.update_signal_plot()

    def clear_addsub_slot(self,slot:str):
        if slot=="A": self.addsub_a=None
        else: self.addsub_b=None
        self.addsub_status.setText(f"A: {self.addsub_a['name'] if self.addsub_a else '-'}\nB: {self.addsub_b['name'] if self.addsub_b else '-'}")

    def clear_addsub_result(self):
        self.addsub_preview_result=None; self.addsub_preview_operation=""; self.save_addsub_button.setEnabled(False)

    def clear_compensation_history(self):
        if self.compensation_base_kspace is not None:
            self.current_kspace = np.asarray(self.compensation_base_kspace).copy()
            self.current_recon = ifft2c(self.current_kspace)
            if self.compensation_base_image is not None:
                self.current_image = np.asarray(self.compensation_base_image).copy()
            else:
                self.current_image = np.abs(self.current_recon)

        self.compensation_history = []
        self.compensation_history_index = -1
        self.compensation_preview = None
        self.compensation_original = None
        self.compensation_original_kspace = None
        if self.compensation_roi_panel is not None:
            self.compensation_roi_panel.hide_comp_roi()
        self.compensation_roi_panel = None
        self._update_compensation_history_buttons()
        self.comp_status_label.setText("Compensation history cleared — original restored")
        self.refresh_images()


    def clear_selected_artifact_training(self):
        self.artifact_selected_indices=[]; self.artifact_preview_position=0; self.artifact_selection_label.setText("No training images selected."); self.artifact_preview_panel.image_item.clear()

    def clear_all_artifact_training(self):
        self.clear_selected_artifact_training(); self.normal_reference_image=None; self.normal_reference_path=None; self.artifact_normal_label.setText("Normal reference: None")

    def _tree_selection_changed(self):
        selected_items = self.tree.selectedItems()

        indices = []
        for item in selected_items:
            data = item.data(0, Qt.UserRole)
            if data and data[0] == "dicom":
                indices.append(int(data[1]))

        self.artifact_selected_indices = sorted(set(indices))
        if hasattr(self, "artifact_selection_label"):
            self.artifact_selection_label.setText(
                f"Selected DICOM images in Explorer: "
                f"{len(self.artifact_selected_indices)}"
            )

        # For a single RAW/bitmap/tracker/signal selection, ensure the item is
        # opened even on Windows configurations where itemClicked is not
        # delivered after a drag or focus change.
        if len(selected_items) == 1:
            item = selected_items[0]
            data = item.data(0, Qt.UserRole)
            if data and data[0] != "dicom":
                QTimer.singleShot(
                    0,
                    lambda selected=item: self._open_tree_item(
                        selected,
                        force=False,
                    ),
                )



    def _ensure_artifact_database(self):
        if self.artifact_db is None:
            self.open_artifact_database()
        return self.artifact_db is not None

    def open_artifact_database(self):
        choice = QMessageBox(self)
        choice.setWindowTitle("Artifact Database")
        choice.setText("Open an existing database or create a new database.")
        open_button = choice.addButton("Open Existing", QMessageBox.AcceptRole)
        create_button = choice.addButton("Create New", QMessageBox.ActionRole)
        choice.addButton("Cancel", QMessageBox.RejectRole)
        choice.exec()

        clicked = choice.clickedButton()
        if clicked == open_button:
            filename, _ = QFileDialog.getOpenFileName(
                self, "Open Artifact Database", "",
                "SQLite database (*.sqlite *.db);;All files (*)"
            )
        elif clicked == create_button:
            filename, _ = QFileDialog.getSaveFileName(
                self, "Create Artifact Database", "MR_Artifact_Learning.sqlite",
                "SQLite database (*.sqlite *.db)"
            )
        else:
            return

        if not filename:
            return
        path = Path(filename)
        if path.suffix.lower() not in {".sqlite", ".db"}:
            path = path.with_suffix(".sqlite")
        if self.artifact_db is not None:
            self.artifact_db.close()
        self.artifact_db = ArtifactDatabase(path)
        self.artifact_db_path = path
        self.artifact_db_label.setText(f"Artifact DB: {path}")
        self.refresh_artifact_lists()
        self.refresh_artifact_training_table()


    def refresh_artifact_lists(self):
        if self.artifact_db is None:
            return
        self.artifact_type_combo.clear()
        self.artifact_type_combo.addItems(self.artifact_db.artifact_types())
        self.artifact_resolution_combo.clear()
        self.artifact_resolution_combo.addItems(self.artifact_db.resolutions())

    def _set_learning_class_quick(self, class_name: str):
        if hasattr(self, "artifact_type_combo"):
            index=self.artifact_type_combo.findText(class_name)
            if index<0:
                self.artifact_new_type.setText(class_name)
                self.add_artifact_type_manual()
                index=self.artifact_type_combo.findText(class_name)
            if index>=0: self.artifact_type_combo.setCurrentIndex(index)
        self.collect_selected_artifact_images()
        self.statusBar().showMessage(f"Learning class prepared: {class_name}")

    def add_artifact_type_manual(self):
        if not self._ensure_artifact_database():
            return
        value = self.artifact_new_type.text().strip()
        if value:
            self.artifact_db.add_artifact_type(value)
            self.refresh_artifact_lists()
            self.artifact_type_combo.setCurrentText(value)
            self.artifact_new_type.clear()

    def add_resolution_manual(self):
        if not self._ensure_artifact_database():
            return
        value = self.artifact_new_resolution.text().strip()
        if value:
            self.artifact_db.add_resolution(value)
            self.refresh_artifact_lists()
            self.artifact_resolution_combo.setCurrentText(value)
            self.artifact_new_resolution.clear()

    def collect_selected_artifact_images(self):
        self._tree_selection_changed()
        if not self.artifact_selected_indices and self.dicom_entries:
            self.artifact_selected_indices = [getattr(self, "slice_index", 0)]
        self.artifact_selection_label.setText(
            f"Training images ready: {len(self.artifact_selected_indices)}"
        )
        self.tabs.setCurrentIndex(2)

    def preview_artifact_learning_image(self):
        self._tree_selection_changed()
        if not self.artifact_selected_indices:
            QMessageBox.information(
                self, "Artifact Learning Preview", "Select one or more DICOM images in Explorer."
            )
            return
        self.artifact_preview_position = max(
            0, min(self.artifact_preview_position, len(self.artifact_selected_indices) - 1)
        )
        index = self.artifact_selected_indices[self.artifact_preview_position]
        if not (0 <= index < len(self.dicom_entries)):
            return
        entry = self.dicom_entries[index]
        self.artifact_preview_panel.label.setText(
            f"Artifact Learning Preview: {entry.path.name} "
            f"({self.artifact_preview_position + 1}/{len(self.artifact_selected_indices)})"
        )
        self.artifact_preview_panel.set_image(np.asarray(entry.image, dtype=float))
        self.artifact_selection_label.setText(
            f"Training images ready: {len(self.artifact_selected_indices)} | "
            f"Previewing {entry.path.name}"
        )

    def navigate_artifact_preview(self, step: int):
        if not self.artifact_selected_indices:
            self._tree_selection_changed()
        if not self.artifact_selected_indices:
            return
        self.artifact_preview_position = (
            self.artifact_preview_position + int(step)
        ) % len(self.artifact_selected_indices)
        self.preview_artifact_learning_image()

    def set_current_normal_reference(self):
        if self.current_image is None:
            QMessageBox.information(self, "Normal Reference", "Load an image first.")
            return
        self.normal_reference_image = np.asarray(self.current_image, dtype=float).copy()
        self.normal_reference_path = Path(self.current_source) if self.current_source else None
        self.artifact_normal_label.setText(
            f"Normal reference: {self.normal_reference_path or 'Current image'}"
        )

    def load_normal_reference(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Load Normal Reference", "",
            "DICOM or NumPy (*.dcm *.ima *.npy *.npz);;All files (*)"
        )
        if not filename:
            return
        path = Path(filename)
        try:
            if path.suffix.lower() == ".npy":
                image = np.load(path, allow_pickle=False)
            elif path.suffix.lower() == ".npz":
                archive = np.load(path, allow_pickle=False)
                image = archive[list(archive.keys())[0]]
            else:
                image = self.read_dicom(path).image
            image = np.asarray(image).squeeze()
            if image.ndim != 2:
                raise ValueError(f"Normal reference must be 2D: {image.shape}")
            self.normal_reference_image = image.astype(float)
            self.normal_reference_path = path
            self.artifact_normal_label.setText(f"Normal reference: {path}")
        except Exception as exc:
            QMessageBox.critical(self, "Normal Reference Error", str(exc))

    def save_artifact_training_samples(self):
        if not self._ensure_artifact_database():
            return
        indices = [
            i for i in self.artifact_selected_indices
            if 0 <= i < len(self.dicom_entries)
        ]
        if not indices:
            QMessageBox.information(self, "Artifact Learning", "Select DICOM images first.")
            return
        artifact_type = self.artifact_type_combo.currentText().strip() or "Not Artifact"
        resolution = self.artifact_resolution_combo.currentText().strip()
        notes = self.artifact_notes.toPlainText().strip()
        progress = self._make_progress("Saving Artifact Training Samples", len(indices))
        saved = 0
        try:
            for count, index in enumerate(indices, start=1):
                if progress.wasCanceled():
                    break
                entry = self.dicom_entries[index]
                progress.setLabelText(f"Analyzing image: {entry.path.name}")
                features = image_features(entry.image, self.normal_reference_image)
                ds = entry.ds
                self.artifact_db.add_sample(
                    source_path=str(entry.path),
                    source_name=entry.path.name,
                    series_uid=str(getattr(ds, "SeriesInstanceUID", "")),
                    series_description=str(
                        getattr(ds, "SeriesDescription", "")
                        or getattr(ds, "ProtocolName", "")
                    ),
                    instance_number=int(getattr(ds, "InstanceNumber", index) or index),
                    artifact_type=artifact_type,
                    resolution=resolution,
                    is_normal_reference=False,
                    normal_reference_path=str(self.normal_reference_path or ""),
                    features=features,
                    notes=notes,
                )
                saved += 1
                progress.setValue(count)
        finally:
            progress.close()
        self.refresh_artifact_training_table()
        self.artifact_summary.setText(
            f"Saved {saved} sample(s) as {artifact_type}. "
            f"Normal comparison: {'Yes' if self.normal_reference_image is not None else 'No'}"
        )

    def refresh_artifact_training_table(self):
        if self.artifact_db is None:
            self.artifact_training_table.setRowCount(0)
            return
        rows = self.artifact_db.samples()
        self.artifact_training_table.setRowCount(len(rows))
        for r, sample in enumerate(rows):
            try:
                features = json.loads(sample.get("features_json", "{}"))
            except Exception:
                features = {}
            values = [
                sample["id"], sample["source_name"], sample["series_description"],
                sample["instance_number"], sample["artifact_type"],
                sample["resolution"] or "",
                "Yes" if features.get("normal_comparison") else "No",
                sample["source_path"] or "",
            ]
            for c, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setData(Qt.UserRole, int(sample["id"]))
                self.artifact_training_table.setItem(r, c, item)
        self.artifact_training_table.resizeColumnsToContents()

    def reclassify_selected_db_rows(self):
        if self.artifact_db is None:
            return
        rows = sorted({i.row() for i in self.artifact_training_table.selectedItems()})
        artifact_type = self.artifact_type_combo.currentText().strip() or "Not Artifact"
        resolution = self.artifact_resolution_combo.currentText().strip()
        notes = self.artifact_notes.toPlainText().strip()
        for row in rows:
            item = self.artifact_training_table.item(row, 0)
            if item is not None:
                self.artifact_db.update_sample_classification(
                    int(item.data(Qt.UserRole)), artifact_type, resolution, notes
                )
        self.refresh_artifact_training_table()

    def import_artifact_database(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Import Artifact DB JSON", "",
            "Artifact DB JSON (*.json);;All files (*)"
        )
        if not filename:
            return
        if self.artifact_db is None:
            self.open_artifact_database()
        if self.artifact_db is None:
            return
        self.artifact_db.import_json(Path(filename))
        self.refresh_artifact_lists()
        self.refresh_artifact_training_table()


    def export_artifact_database(self):
        if self.artifact_db is None:
            QMessageBox.information(self, "Artifact DB", "Open an Artifact DB first.")
            return
        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Artifact DB JSON", "MRI_Artifact_DB.json",
            "Artifact DB JSON (*.json)"
        )
        if filename:
            if not filename.lower().endswith(".json"):
                filename += ".json"
            self.artifact_db.export_json(Path(filename))

    def fill_header(self, ds):
        return


    def filter_header(self, value):
        return


    def export_header_excel(self):
        if self.current_ds is None: QMessageBox.information(self,'No DICOM','Load a DICOM file first.'); return
        filename,_=QFileDialog.getSaveFileName(self,'Export DICOM Header','DICOM_Header.xlsx','Excel (*.xlsx)')
        if not filename: return
        if not filename.lower().endswith('.xlsx'): filename += '.xlsx'
        wb=Workbook(); ws=wb.active; ws.title='DICOM Header'; headers=['Tag','Keyword','Name','VR','VM','Value']; ws.append(headers)
        for cell in ws[1]: cell.font=Font(bold=True); cell.fill=PatternFill('solid', fgColor='D9EAF7')
        for elem in self.current_ds.iterall():
            if elem.tag.group != 0x7FE0: ws.append([str(elem.tag),elem.keyword,elem.name,elem.VR,str(elem.VM),str(elem.value)])
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        for col,width in {'A':16,'B':28,'C':38,'D':8,'E':8,'F':80}.items(): ws.column_dimensions[col].width=width
        for row in ws.iter_rows():
            for cell in row: cell.alignment=Alignment(vertical='top', wrap_text=True)
        wb.save(filename); self.statusBar().showMessage(f'Exported {filename}')

    def export_npz(self):
        if self.current_image is None: return
        filename,_=QFileDialog.getSaveFileName(self,'Export NPZ','MRI_FFT_Data.npz','NPZ (*.npz)')
        if filename:
            if not filename.lower().endswith('.npz'): filename += '.npz'
            np.savez_compressed(filename,image=self.current_image,kspace=self.current_kspace,reconstruction=self.current_recon,source=self.current_source)

    def _install_raw_compensation_tooltips(self):
        """Install concise help on the compensation controls."""
        tips = {
            "select_comp_button": "Start manual mask editing on the original RAW/k-space data.",
            "detect_comp_button": "Convert the current painted regions into the active compensation mask.",
            "comp_clear_mask_button": "Remove the current painted mask.",
            "comp_brush_size_spin": "Brush diameter in image pixels. The circular cursor shows the active size.",
            "comp_brush_size_down_button": "Decrease brush size by one pixel.",
            "comp_brush_size_up_button": "Increase brush size by one pixel.",
            "preview_comp_button": "Reconstruct once from the current mask and settings, then open Before/After review.",
            "comp_auto_correct_button": "Run automatic candidate detection, validation and best-result selection.",
            "comp_auto_more_button": "Open additional Auto Correct actions, including running Auto Correct again.",
            "comp_quick_recalculate_button": "Apply the current Quick Adjust values. With no mask, run one detection trial.",
            "comp_quick_review_button": "Review the latest reconstructed result without changing the current mask.",
            "comp_restore_auto_button": "Restore the saved Auto Correct mask, settings and reconstructed result.",
            "comp_expert_toggle_button": "Open advanced detection and reconstruction controls.",
            "comp_result_ok_button": "Accept the current reconstructed result.",
            "comp_result_quick_button": "Fine tune the result with three simplified controls.",
            "comp_result_paint_button": "Continue from the Auto Correct mask in manual paint mode.",
            "comp_result_expert_button": "Continue from the Auto Correct mask in Expert mode.",
            "comp_expert_auto_mask_button": "Run one automatic mask trial using the current Expert settings.",
            "comp_candidate_viewer_button": "Inspect accepted and rejected candidates and apply a selected candidate.",
            "comp_expert_review_button": "Reconstruct directly from the current Expert mask and settings.",
            "apply_comp_button": "Apply the reviewed compensation to the current RAW data.",
            "save_comp_button": "Save the compensated result.",
            "cancel_comp_button": "Cancel the current compensation workflow and restore the source state.",
            "comp_mask_undo_button": "Undo the most recent mask edit.",
            "comp_mask_redo_button": "Redo the most recently undone mask edit.",
            "comp_mask_symmetry_button": "Mirror the current mask using Hermitian symmetry.",
            "comp_mask_expand_button": "Expand the mask by one pixel around its boundary.",
            "comp_mask_shrink_button": "Shrink the mask by one pixel around its boundary.",
            "comp_mask_fill_button": "Keep and fill the largest connected mask region.",
            "comp_mask_delete_button": "Delete the smallest connected mask region.",
        }
        for name, tip in tips.items():
            widget = getattr(self, name, None)
            if widget is not None:
                widget.setToolTip(tip)

    def _guide_pages(self):
        """Commit0119 guide content, separated into normal and advanced phases."""
        normal = [
            ("Welcome", "MR Image Explorer supports DICOM, RAW, folders and ZIP packages. Image Workspace is the current production-ready workflow. Other analysis tabs are still Work in Progress and their results must be reviewed carefully."),
            ("Import Images", "Use File > Import Files, Import Folder, or drag and drop supported data. Wait for the import progress to complete before starting analysis."),
            ("Image Workspace", "Production-ready area: use Original, FFT and combined views to inspect the current image. Select a series or image in the tree to update the workspace."),
            ("Navigation", "Use Previous and Next, the keyboard arrows, or the mouse wheel to move through images and series."),
            ("Display Controls", "Window/Level changes contrast. Zoom and Pan inspect detail. Reset returns to the default display state."),
            ("FFT Basics", "FFT shows the frequency-domain RAW data. Bright isolated spots, bands or rings can indicate structured artifacts."),
            ("Quick Spike Detect — Work in Progress", "Quick Spike Detect now shares the Raw Data Compensation candidate engine, but this feature is still Work in Progress. Treat its confidence result as diagnostic support, not a final decision."),
            ("Spike and Artifact Diagnosis — Work in Progress", "Spike Diag and Artifact Diag use the Raw Data Compensation candidate engine for broader detection. Both tabs are still Work in Progress and require visual confirmation in Image Workspace."),
            ("Other Tabs — Work in Progress", "Tracker Signal, 1D Signal Studio, Spike Diag and Artifact Diag are still Work in Progress. Image Workspace is the supported primary workflow. Save or export only after visual verification and keep the original data unchanged."),
        ]
        raw = [
            ("Raw Data Compensation", "Advanced feature: remove structured RAW/k-space artifacts while protecting normal image-forming signal."),
            ("Beginner Workflow", "1. Auto Correct\n2. Review Reconstructed Image\n3. Apply Compensation\n4. Save the result"),
            ("Quick Adjust", "Use the current mask and current simplified settings for one recalculation. With no mask, one detection trial is performed; there is no retry loop."),
            ("Paint", "Start Manual Paint, refine the inherited Auto Correct mask with Brush or Eraser, then use the painted mask for reconstruction."),
            ("Expert", "Auto Mask Once performs one trial. Mask Expansion controls mask size, Donor Halo controls nearby donor signal, and Compensation Passes controls iteration count."),
            ("Review and Apply", "Compare Before, After and Difference. Apply only when artifact reduction is clear and normal detail remains acceptable."),
        ]
        return normal, raw

    def _show_guide_pages(self, title_text, pages, progress_key=None, finish_callback=None):
        dialog = QDialog(self)
        dialog.setWindowTitle(title_text)
        dialog.setModal(True)
        dialog.resize(680, 440)
        layout = QVBoxLayout(dialog)
        title = QLabel(); title.setStyleSheet("font-size: 19px; font-weight: bold;")
        body = QLabel(); body.setWordWrap(True); body.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        body.setStyleSheet("font-size: 13px; padding: 8px;")
        counter = QLabel(); counter.setAlignment(Qt.AlignRight)
        layout.addWidget(title); layout.addWidget(body, 1); layout.addWidget(counter)
        row = QHBoxLayout()
        previous = QPushButton("Previous")
        next_button = QPushButton("Next")
        close_button = QPushButton("Close")
        row.addWidget(previous); row.addWidget(next_button); row.addStretch(1); row.addWidget(close_button)
        layout.addLayout(row)
        state = {"index": 0}
        def render():
            index = state["index"]
            title.setText(pages[index][0]); body.setText(pages[index][1])
            counter.setText(f"{index + 1} / {len(pages)}")
            previous.setEnabled(index > 0)
            next_button.setText("Finish" if index == len(pages) - 1 else "Next")
            if progress_key:
                self.viewer_settings.setValue(progress_key, index)
        def go_previous():
            state["index"] = max(0, state["index"] - 1); render()
        def go_next():
            if state["index"] >= len(pages) - 1:
                if progress_key:
                    self.viewer_settings.setValue(progress_key + "/completed", True)
                dialog.accept()
            else:
                state["index"] += 1; render()
        previous.clicked.connect(go_previous); next_button.clicked.connect(go_next); close_button.clicked.connect(dialog.reject)
        render()
        accepted = dialog.exec() == QDialog.Accepted
        if accepted and finish_callback:
            finish_callback()
        return accepted

    def _handle_startup_guide_flow(self):
        if RELEASE_MODE:
            return
        """Phase 1: startup choice and one-time forced guide settings."""
        force_normal = self.viewer_settings.value("guide/force_normal_next_start", False, type=bool)
        force_raw = self.viewer_settings.value("guide/force_raw_after_normal_next_start", False, type=bool)
        if force_normal or force_raw:
            self.viewer_settings.setValue("guide/force_normal_next_start", False)
            self.viewer_settings.setValue("guide/force_raw_after_normal_next_start", False)
            if force_normal:
                self.show_normal_guide(offer_raw=force_raw)
            elif force_raw:
                self.show_raw_compensation_guide(start_tour=True)
            return
        ask_enabled = self.viewer_settings.value("guide/ask_at_startup", True, type=bool)
        if not ask_enabled:
            return
        box = QMessageBox(self)
        box.setWindowTitle("Welcome to MR Image Explorer")
        box.setIcon(QMessageBox.Question)
        box.setText("Would you like to open the normal usage guide and guided tour?")
        box.setInformativeText("Raw Data Compensation is offered only at the end of the normal guide. Choose No to stop this startup question from appearing again.")
        yes_button = box.addButton("Yes — Start Guide", QMessageBox.YesRole)
        no_button = box.addButton("No — Do Not Ask Again", QMessageBox.NoRole)
        later_button = box.addButton("Not Now", QMessageBox.RejectRole)
        box.setDefaultButton(yes_button)
        box.exec()
        if box.clickedButton() is yes_button:
            self.show_normal_guide(offer_raw=True)
        elif box.clickedButton() is no_button:
            self.viewer_settings.setValue("guide/ask_at_startup", False)

    def show_normal_guide(self, offer_raw=True):
        normal, _ = self._guide_pages()
        def after_normal():
            if not offer_raw:
                return
            box = QMessageBox(self)
            box.setWindowTitle("Optional Advanced Guide")
            box.setIcon(QMessageBox.Information)
            box.setText("Would you like to learn Raw Data Compensation now?")
            box.setInformativeText("This is an advanced feature and is not required for normal image viewing and diagnosis.")
            open_button = box.addButton("Open Advanced Guide", QMessageBox.AcceptRole)
            box.addButton("Later", QMessageBox.RejectRole)
            box.exec()
            if box.clickedButton() is open_button:
                self.show_raw_compensation_guide(start_tour=True)
        self._show_guide_pages("MR Image Explorer — Normal Usage Guide", normal, "guide/normal_progress", after_normal)

    def show_raw_compensation_guide(self, start_tour: bool = True):
        _, raw = self._guide_pages()
        self._show_guide_pages("Raw Data Compensation — Optional Advanced Guide", raw, "guide/raw_progress")

    def show_guide_library(self):
        """Phase 3/4: persistent guide library and topic access."""
        normal, raw = self._guide_pages()
        dialog = QDialog(self); dialog.setWindowTitle("Guide Library"); dialog.resize(760, 500)
        layout = QHBoxLayout(dialog)
        tree = QTreeWidget(); tree.setHeaderHidden(True)
        normal_root = QTreeWidgetItem(["Normal Usage Guide"])
        for i, (title, _) in enumerate(normal):
            item = QTreeWidgetItem([title]); item.setData(0, Qt.UserRole, ("normal", i)); normal_root.addChild(item)
        raw_root = QTreeWidgetItem(["Raw Data Compensation — Advanced"])
        for i, (title, _) in enumerate(raw):
            item = QTreeWidgetItem([title]); item.setData(0, Qt.UserRole, ("raw", i)); raw_root.addChild(item)
        tree.addTopLevelItem(normal_root); tree.addTopLevelItem(raw_root); tree.expandAll()
        right = QVBoxLayout(); heading = QLabel("Select a guide topic"); heading.setStyleSheet("font-size:18px;font-weight:bold;")
        text = QLabel("Normal usage is listed first. Raw Data Compensation is an optional advanced section."); text.setWordWrap(True); text.setAlignment(Qt.AlignTop)
        open_normal = QPushButton("Start Normal Guide")
        open_raw = QPushButton("Open Raw Data Compensation Guide")
        reset_ask = QPushButton("Ask About Guide at Next Startup")
        close_button = QPushButton("Close")
        right.addWidget(heading); right.addWidget(text,1); right.addWidget(open_normal); right.addWidget(open_raw); right.addWidget(reset_ask); right.addWidget(close_button)
        layout.addWidget(tree,1); layout.addLayout(right,2)
        def select_item(item):
            data=item.data(0, Qt.UserRole)
            if not data: return
            group,index=data; pages=normal if group=="normal" else raw
            heading.setText(pages[index][0]); text.setText(pages[index][1])
        tree.currentItemChanged.connect(lambda current, previous: select_item(current) if current else None)
        open_normal.clicked.connect(lambda: self.show_normal_guide(offer_raw=True))
        open_raw.clicked.connect(lambda: self.show_raw_compensation_guide(start_tour=True))
        reset_ask.clicked.connect(lambda: (self.viewer_settings.setValue("guide/ask_at_startup", True), QMessageBox.information(dialog, "Guide", "The startup guide question is enabled.")))
        close_button.clicked.connect(dialog.accept)
        dialog.exec()

    def closeEvent(self, event):
        """Phase 1: exit confirmation with independent next-start guide options."""
        dialog = QDialog(self); dialog.setWindowTitle("Close MR Image Explorer"); dialog.setModal(True); dialog.resize(520, 250)
        layout = QVBoxLayout(dialog)
        title = QLabel("Close MR Image Explorer?"); title.setStyleSheet("font-size:18px;font-weight:bold;")
        note = QLabel("Unsaved work may be lost."); note.setWordWrap(True)
        normal_check = QCheckBox("Show the normal usage guide at the next startup")
        raw_check = QCheckBox("Offer the Raw Data Compensation guide after the normal guide")
        normal_check.setObjectName("ExitGuideCheck"); raw_check.setObjectName("ExitGuideCheck")
        normal_check.setStyleSheet("QCheckBox{spacing:10px;color:#f2f4f8;} QCheckBox::indicator{width:20px;height:20px;border:2px solid #d7dde8;background:#ffffff;border-radius:3px;} QCheckBox::indicator:checked{background:#2f8cff;border-color:#9bc8ff;} QCheckBox::indicator:disabled{background:#777;border-color:#aaa;}")
        raw_check.setStyleSheet(normal_check.styleSheet())
        normal_check.setChecked(False); raw_check.setChecked(False)
        raw_check.toggled.connect(lambda checked: normal_check.setChecked(True) if checked else None)
        buttons = QHBoxLayout(); buttons.addStretch(1)
        exit_button = QPushButton("Exit"); cancel_button = QPushButton("Cancel"); cancel_button.setDefault(True)
        buttons.addWidget(exit_button); buttons.addWidget(cancel_button)
        layout.addWidget(title); layout.addWidget(note); layout.addSpacing(8); layout.addWidget(normal_check); layout.addWidget(raw_check); layout.addStretch(1); layout.addLayout(buttons)
        exit_button.clicked.connect(dialog.accept); cancel_button.clicked.connect(dialog.reject)
        if dialog.exec() != QDialog.Accepted:
            event.ignore(); return
        if normal_check.isChecked():
            self.viewer_settings.setValue("guide/force_normal_next_start", True)
            self.viewer_settings.setValue("guide/ask_at_startup", True)
        if raw_check.isChecked():
            self.viewer_settings.setValue("guide/force_raw_after_normal_next_start", True)
        self._save_viewer_layout()
        self.import_cancel_event.set()
        if self.import_worker is not None:
            try: self.import_worker.cancel()
            except Exception: pass
        for directory in self.temp_dirs:
            try: shutil.rmtree(directory, ignore_errors=True)
            except Exception: pass
        event.accept()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._schedule_responsive_layout()


    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(0, self._initial_screen_fit)
        QTimer.singleShot(180, self._initial_screen_fit)

def validate_startup_dependencies():
    required_names = [
        "QApplication",
        "QMainWindow",
        "QGridLayout",
        "QProgressBar",
        "QSizePolicy",
        "QTreeWidget",
        "QMimeData",
        "QUrl",
        "QDrag",
        "MainWindow",
    ]
    missing = [name for name in required_names if name not in globals()]
    if missing:
        raise RuntimeError(
            "Missing startup dependency/import: " + ", ".join(missing)
        )


def write_startup_error(error: BaseException) -> Path:
    log_dir = Path.home() / "AppData" / "Local" / "MR_Image_Explorer"
    try:
        log_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        log_dir = Path.cwd()

    log_path = log_dir / "startup_error.log"
    try:
        import traceback
        log_path.write_text(
            "MR Image Explorer startup failed.\n\n"
            + "".join(
                traceback.format_exception(
                    type(error),
                    error,
                    error.__traceback__,
                )
            ),
            encoding="utf-8",
        )
    except Exception:
        pass
    return log_path


def main():
    try:
        validate_startup_dependencies()
        from insightec_handoff import load_handoff
        handoff = load_handoff("fus_image_explore")

        app = QApplication(sys.argv)
        app.setApplicationName(APP_NAME)
        app.setOrganizationName("InSightec")

        window = MainWindow()
        window._hub_handoff_active = bool(handoff and handoff.auto_load)
        window.show()
        window.raise_()
        window.activateWindow()

        if handoff and handoff.auto_load:
            inputs = handoff.input_paths()
            if not inputs and handoff.workspace():
                inputs = [handoff.workspace()]
            handoff.mark("fus_image_explore", "accepted", input_count=len(inputs))
            if inputs:
                QTimer.singleShot(0, lambda values=list(inputs): window.import_paths(values))

        # Used by GitHub Actions to verify that both Python and packaged EXE
        # actually create the application window and enter the Qt event loop.
        if os.environ.get("MRIE_STARTUP_SMOKE_TEST") == "1":
            QTimer.singleShot(1200, app.quit)

        return app.exec()

    except BaseException as error:
        log_path = write_startup_error(error)

        try:
            app = QApplication.instance()
            if app is None:
                app = QApplication(sys.argv)
            QMessageBox.critical(
                None,
                "MR Image Explorer - Startup Error",
                f"{type(error).__name__}: {error}\n\n"
                f"Log:\n{log_path}",
            )
        except Exception:
            pass

        return 1


if __name__ == "__main__":
    raise SystemExit(main())
