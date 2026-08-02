from __future__ import annotations

import json
import math
import os
import re
import sys
import traceback
import shutil
import tempfile
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

RELEASE_MODE = os.environ.get("INSIGHTEC_RELEASE_MODE", "0").strip().lower() in {"1", "true", "yes", "on"}

APP_NAME = "TrackerSNR_CLEAN_EXE"
SPEC_SNR = 100.0
DATA_HEADERS = [
    "File Index", "File Name", "Time", "Tracker No", "Scan No", "Block No",
    "Signal", "Noise", "SNR", "Peak Loc", "Peak COM", "Peak Center", "Source Type", "Full Path"
]
SUMMARY_HEADERS = [
    "File Index", "File Name", "Full Path", "Rows", "Trackers", "Signal/Noise", "Blocks", "Max SNR", "Min SNR"
]
MONTHS = {m.lower(): i for i, m in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent




def settings_path() -> Path:
    """Return a per-user writable settings path that also works for one-file EXEs."""
    base = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA")
    if base:
        folder = Path(base) / "InSightec" / "TrackerSNRAnalyzer"
    else:
        folder = Path.home() / ".insightec" / "TrackerSNRAnalyzer"
    folder.mkdir(parents=True, exist_ok=True)
    return folder / "settings.json"


def load_user_settings() -> dict[str, Any]:
    defaults = {"ask_startup_guide": True, "show_guide_once": False}
    try:
        data = json.loads(settings_path().read_text(encoding="utf-8"))
        if isinstance(data, dict):
            defaults.update({k: data[k] for k in defaults if k in data})
    except FileNotFoundError:
        pass
    except Exception as exc:
        write_log(f"Could not read settings: {exc}")
    return defaults


def save_user_settings(data: dict[str, Any]) -> None:
    try:
        settings_path().write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        write_log(f"Could not save settings: {exc}")


def log_path() -> Path:
    # Prefer the EXE/source folder, but fall back to TEMP if the EXE folder is not writable.
    primary = app_dir() / "TrackerSNR_runtime.log"
    try:
        primary.parent.mkdir(parents=True, exist_ok=True)
        with primary.open("a", encoding="utf-8"):
            pass
        return primary
    except Exception:
        return Path(os.environ.get("TEMP", ".")) / "TrackerSNR_runtime.log"


def normalize_input_path(path: str | Path) -> str:
    # QFileDialog can return native Windows paths, forward-slash paths, or quoted paths.
    # Do not resolve(strict=True), because PyInstaller one-file and network paths can fail there.
    text = str(path).strip().strip('"')
    if not text:
        return ""
    return os.path.normpath(os.path.abspath(text))


def write_log(message: str) -> None:
    try:
        with log_path().open("a", encoding="utf-8") as f:
            f.write(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {message}\n")
    except Exception:
        pass


def handle_exception(context: str, exc: BaseException) -> None:
    text = f"{context}: {type(exc).__name__}: {exc}\n{traceback.format_exc()}"
    write_log(text)


def file_name_only(path: str | Path) -> str:
    return Path(path).name


def safe_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", name)


def parse_log_stamp_from_name(path: str | Path) -> Optional[datetime]:
    m = re.search(r"(\d{4})_([A-Za-z]{3})_(\d{1,2})_(\d{1,2})_(\d{1,2})_(\d{1,2})", Path(path).name)
    if not m:
        return None
    try:
        y = int(m.group(1)); mo = MONTHS.get(m.group(2).lower(), 0); d = int(m.group(3))
        hh = int(m.group(4)); mm = int(m.group(5)); ss = int(m.group(6))
        if mo < 1:
            return None
        return datetime(y, mo, d, hh, mm, ss)
    except Exception:
        return None


def sort_paths_newest_first(paths: Iterable[str | Path]) -> list[str]:
    def key(p: str | Path) -> tuple[float, str]:
        pp = Path(p)
        dt = parse_log_stamp_from_name(pp)
        if dt:
            ts = dt.timestamp()
        else:
            try:
                ts = pp.stat().st_mtime
            except OSError:
                ts = 0.0
        return (ts, str(pp).lower())
    return [str(p) for p in sorted(paths, key=key, reverse=True)]


def is_target_log_file(path: str | Path) -> bool:
    return Path(path).suffix.lower() in {".log", ".txt"}


def is_target_folder_log_file(path: str | Path) -> bool:
    if not is_target_log_file(path):
        return False
    return re.search(r"^\d{4}_[A-Za-z]{3}_\d{1,2}_\d{1,2}_\d{1,2}_\d{1,2}\.(log|txt)$", Path(path).name, re.I) is not None


def sort_paths_oldest_first(paths: Iterable[str | Path]) -> list[str]:
    return list(reversed(sort_paths_newest_first(paths)))

def collect_target_logs(folder: str | Path, max_depth: int = 5) -> list[str]:
    root=Path(normalize_input_path(folder)); found=[]
    if not root.exists(): return []
    base=len(root.parts)
    for current,dirs,files in os.walk(root):
        depth=len(Path(current).parts)-base
        if depth>=max_depth: dirs[:]=[]
        for name in files:
            p=Path(current)/name
            if is_target_log_file(p): found.append(str(p))
    return sort_paths_oldest_first(found)

def collect_inputs(inputs: Iterable[str | Path], temp_roots: list[str] | None=None, max_depth: int=5) -> list[str]:
    temp_roots=temp_roots if temp_roots is not None else []; found=[]
    for raw in inputs:
        p=Path(normalize_input_path(raw))
        if p.is_dir(): found.extend(collect_target_logs(p,max_depth))
        elif p.is_file() and p.suffix.lower()=='.zip':
            tmp=tempfile.mkdtemp(prefix='TrackerSNR_'); temp_roots.append(tmp)
            try:
                with zipfile.ZipFile(p) as z:
                    for info in z.infolist():
                        if info.is_dir() or Path(info.filename).suffix.lower() not in {'.log','.txt'}: continue
                        dst=(Path(tmp)/info.filename).resolve()
                        if not str(dst).startswith(str(Path(tmp).resolve())): continue
                        dst.parent.mkdir(parents=True,exist_ok=True)
                        with z.open(info) as src,dst.open('wb') as out: shutil.copyfileobj(src,out)
                found.extend(collect_target_logs(tmp,max_depth))
            except Exception as exc: write_log(f'ZIP extraction failed: {p}: {exc}')
        elif p.is_file() and is_target_log_file(p): found.append(str(p))
    return sort_paths_oldest_first(dict.fromkeys(found))

def read_text_lines(path: str | Path) -> list[str]:
    path = normalize_input_path(path)
    encodings = ["utf-8-sig", "utf-8", "cp932", "shift_jis", "latin-1"]
    last_exc: Optional[Exception] = None
    for enc in encodings:
        try:
            return Path(path).read_text(encoding=enc, errors="strict").splitlines()
        except UnicodeDecodeError as exc:
            last_exc = exc
        except Exception as exc:
            last_exc = exc
            break
    try:
        return Path(path).read_text(encoding="latin-1", errors="ignore").splitlines()
    except Exception as exc:
        if last_exc:
            raise last_exc
        raise exc


def extract_time_text(s: str) -> str:
    m = re.search(r"(?:^|\s)(\d{1,2}:\d{2}:\d{2}[:.]\d{1,3})", s)
    return m.group(1).replace(".", ":") if m else ""


def extract_tracker_number(s: str) -> Optional[int]:
    patterns = [
        r"\btracker\s*#?\s*[:=]?\s*(\d+)",
        r"\btrack\s*#?\s*[:=]?\s*(\d+)",
        r"\btrader[s]?\s*#?\s*[:=]?\s*(\d+)",
        r"\btracker\D{0,20}(\d+)",
        r"\btrack\D{0,20}(\d+)",
        r"\btrader[s]?\D{0,20}(\d+)",
    ]
    for pat in patterns:
        m = re.search(pat, s, re.I)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                return None
    return None


def extract_named_number(s: str, name: str) -> Optional[float]:
    m = re.search(rf"\b{re.escape(name)}\b\s*[:=]?\s*(-?\d+(?:\.\d+)?)", s, re.I)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def extract_last_named_number(s: str, name: str) -> Optional[float]:
    matches = re.findall(rf"\b{re.escape(name)}\b\s*[:=]?\s*(-?\d+(?:\.\d+)?)", s, re.I)
    if not matches:
        return None
    try:
        return float(matches[-1])
    except ValueError:
        return None


def extract_first_of(s: str, names: list[str]) -> Optional[float]:
    for name in names:
        v = extract_named_number(s, name)
        if v is not None:
            return v
    return None


def is_signal_noise_line(s: str) -> bool:
    lower = s.lower()
    if "signal" not in lower or "noise" not in lower or "snr" not in lower:
        return False
    if "tracker" not in lower and "track" not in lower and "trader" not in lower:
        return False
    if extract_tracker_number(s) is None:
        return False
    return extract_named_number(s, "signal") is not None and extract_named_number(s, "noise") is not None and extract_named_number(s, "snr") is not None


def is_peak_snr_line(s: str) -> bool:
    lower = s.lower()
    if "signal" in lower and "noise" in lower:
        return False
    if "tracker peak log" in lower:
        return extract_last_named_number(s, "snr") is not None
    if "peak" not in lower or "snr" not in lower:
        return False
    if extract_tracker_number(s) is None:
        return False
    return extract_last_named_number(s, "snr") is not None


def positive(v: Optional[float]) -> bool:
    return v is not None and math.isfinite(v) and v > 0


@dataclass
class DataRow:
    file_index: int
    file_name: str
    time: str
    tracker_no: int
    scan_no: int | None
    block_no: int | None
    signal: float | None
    noise: float | None
    snr: float
    peak_loc: float | None
    peak_com: float | None
    peak_center: float | None
    source_type: str
    full_path: str

    def as_excel_row(self) -> list[Any]:
        return [
            self.file_index, self.file_name, self.time, self.tracker_no, self.scan_no, self.block_no,
            self.signal, self.noise, self.snr, self.peak_loc, self.peak_com, self.peak_center, self.source_type, self.full_path,
        ]


@dataclass
class FileSummary:
    file_index: int
    file_name: str
    full_path: str
    rows: int
    trackers: int
    has_signal_noise: bool
    blocks: int
    max_snr: float | None
    min_snr: float | None

    def as_excel_row(self) -> list[Any]:
        return [
            self.file_index, self.file_name, self.full_path, self.rows, self.trackers,
            "Detected" if self.has_signal_noise else "Not detected", self.blocks,
            round(self.max_snr, 1) if self.max_snr is not None else None,
            round(self.min_snr, 1) if self.min_snr is not None else None,
        ]


@dataclass
class Dataset:
    rows: list[DataRow]
    summaries: list[FileSummary]
    skipped: list[tuple[str, str]]

    def to_project_dict(self) -> dict[str, Any]:
        return {
            "version": 1,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "rows": [asdict(r) for r in self.rows],
            "summaries": [asdict(s) for s in self.summaries],
            "skipped": self.skipped,
        }

    @staticmethod
    def from_project_dict(d: dict[str, Any]) -> "Dataset":
        rows = [DataRow(**r) for r in d.get("rows", [])]
        summaries = [FileSummary(**s) for s in d.get("summaries", [])]
        skipped = [tuple(x) for x in d.get("skipped", [])]
        return Dataset(rows, summaries, skipped)


def detect_tracker_count(trackers: set[int]) -> int:
    if not trackers:
        return 4
    max_tracker = max(trackers)
    if max_tracker >= 5:
        return 6
    if max_tracker >= 3:
        return 4
    return len(trackers)


def parse_one_file(path: str | Path, file_index: int) -> tuple[list[DataRow], FileSummary | None, str | None]:
    path = Path(path)
    try:
        lines = read_text_lines(path)
    except Exception as exc:
        return [], None, f"Read failed: {exc}"

    rows: list[DataRow] = []
    trackers: set[int] = set()
    has_signal_noise = False
    file_name = path.name
    full_path = str(path)

    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        time_text = extract_time_text(line)
        if is_signal_noise_line(line):
            tracker = extract_tracker_number(line)
            signal = extract_named_number(line, "signal")
            noise = extract_named_number(line, "noise")
            snr = extract_last_named_number(line, "snr")
            if tracker is not None and tracker >= 0 and positive(signal) and positive(noise) and positive(snr):
                has_signal_noise = True
                trackers.add(tracker)
                rows.append(DataRow(file_index, file_name, time_text, tracker, None, None, signal, noise, float(snr), None, None, None, "SIGNAL", full_path))
        elif is_peak_snr_line(line):
            tracker = extract_tracker_number(line)
            snr = extract_last_named_number(line, "snr")
            if tracker is not None and tracker >= 0 and positive(snr):
                trackers.add(tracker)
                peak_loc = extract_first_of(line, ["peakLocation", "peakLoc"])
                peak_com = extract_first_of(line, ["peakCOM", "peakCom"])
                peak_center = extract_first_of(line, ["peakCenterNoWeight", "peakCenter"])
                rows.append(DataRow(file_index, file_name, time_text, tracker, None, None, None, None, float(snr), peak_loc, peak_com, peak_center, "PEAK", full_path))

    if not rows:
        return [], None, "No valid SNR data"

    tracker_count = detect_tracker_count(trackers)
    # Assign Scan0..Scan3 and Block per source type and tracker.
    counters: dict[tuple[str, int], int] = {}
    max_block = 1
    for row in rows:
        key = (row.source_type.upper(), row.tracker_no)
        counters[key] = counters.get(key, 0) + 1
        occurrence = counters[key]
        row.scan_no = (occurrence - 1) % 4
        row.block_no = ((occurrence - 1) // 4) + 1
        max_block = max(max_block, row.block_no)

    snrs = [r.snr for r in rows if positive(r.snr)]
    summary = FileSummary(
        file_index=file_index,
        file_name=file_name,
        full_path=full_path,
        rows=len(rows),
        trackers=tracker_count,
        has_signal_noise=has_signal_noise,
        blocks=max_block,
        max_snr=max(snrs) if snrs else None,
        min_snr=min(snrs) if snrs else None,
    )
    return rows, summary, None


def parse_files(paths: list[str | Path], progress_cb=None) -> Dataset:
    valid_rows: list[DataRow] = []
    summaries: list[FileSummary] = []
    skipped: list[tuple[str, str]] = []
    sorted_paths = sort_paths_oldest_first([p for p in paths if is_target_log_file(p)])
    next_index = 1
    for pos, p in enumerate(sorted_paths, 1):
        if progress_cb:
            progress_cb(f"Processing {pos}/{len(sorted_paths)}: {Path(p).name}")
        rows, summary, reason = parse_one_file(p, next_index)
        if rows and summary:
            # Re-index only valid files so UI file list has no gaps.
            for r in rows:
                r.file_index = next_index
            summary.file_index = next_index
            valid_rows.extend(rows)
            summaries.append(summary)
            next_index += 1
        else:
            skipped.append((str(p), reason or "No valid data"))
    if progress_cb:
        progress_cb(f"Loaded {len(summaries)} valid file(s), skipped {len(skipped)} file(s).")
    return Dataset(valid_rows, summaries, skipped)


def get_block_value(dataset: Dataset, file_index: int, block_no: int, tracker_no: int, scan_no: int, metric: str) -> Optional[float]:
    metric_u = metric.upper()
    if metric_u == "SIGNAL":
        sources = ["SIGNAL"]
        attr = "signal"
    elif metric_u == "NOISE":
        sources = ["SIGNAL"]
        attr = "noise"
    else:
        sources = ["SIGNAL", "PEAK"]
        attr = "snr"
    for source in sources:
        for row in dataset.rows:
            if (row.file_index == file_index and row.block_no == block_no and row.tracker_no == tracker_no
                    and row.scan_no == scan_no and row.source_type.upper() == source):
                v = getattr(row, attr)
                if positive(v):
                    return float(v)
    return None


def matrix_for(dataset: Dataset, file_index: int, block_no: int, tracker_count: int) -> list[list[Optional[float]]]:
    return [[get_block_value(dataset, file_index, block_no, t, s, "SNR") for s in range(4)] for t in range(tracker_count)]


def save_excel(dataset: Dataset, out_path: str | Path, default_file_index: int = 1, default_block_no: int = 1) -> Path:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_summary = wb.create_sheet("Summary")
    ws_data = wb.create_sheet("Tracker_SNR")
    ws_lists = wb.create_sheet("_Lists")

    # Data
    ws_data.append(DATA_HEADERS)
    for row in dataset.rows:
        ws_data.append(row.as_excel_row())
    ws_data.auto_filter.ref = f"A1:{get_column_letter(len(DATA_HEADERS))}{ws_data.max_row}"
    for cell in ws_data[1]:
        cell.font = Font(bold=True)

    # Summary
    ws_summary.append(SUMMARY_HEADERS)
    for summary in dataset.summaries:
        ws_summary.append(summary.as_excel_row())
    ws_summary.auto_filter.ref = f"A1:I{ws_summary.max_row}"
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    # Lists
    for i, summary in enumerate(dataset.summaries, 1):
        ws_lists.cell(i, 1).value = summary.file_name
    for i, name in enumerate(["SNR", "Signal", "Noise"], 1):
        ws_lists.cell(i, 2).value = name
    blocks = 1
    tracker_count = 4
    summary = next((s for s in dataset.summaries if s.file_index == default_file_index), dataset.summaries[0] if dataset.summaries else None)
    if summary:
        blocks = max(1, summary.blocks)
        tracker_count = max(1, summary.trackers)
    for i in range(1, blocks + 1):
        ws_lists.cell(i, 3).value = f"Block {i}"

    # Dashboard equivalent
    ws_dash["A1"] = "Tracker SNR Report"
    ws_dash["A1"].font = Font(bold=True, size=16)
    info = [
        ("A3", "File", "B3", summary.file_name if summary else ""),
        ("A4", "Graph data", "B4", "SNR"),
        ("A5", "Block", "B5", f"Block {default_block_no}"),
        ("E2", "Source", "F2", summary.file_name if summary else ""),
        ("E3", "Trackers", "F3", tracker_count),
        ("E4", "Signal/Noise", "F4", "Detected" if summary and summary.has_signal_noise else "Not detected"),
        ("E5", "Blocks", "F5", blocks),
        ("E6", "Matrix spec", "F6", "Blue: >=100, Red: <100"),
    ]
    for lc, label, vc, value in info:
        ws_dash[lc] = label
        ws_dash[lc].font = Font(bold=True)
        ws_dash[vc] = value
    ws_dash["E8"] = "Selected Block SNR Matrix"
    ws_dash["E8"].font = Font(bold=True)

    start_row = 9
    start_col = 5
    ws_dash.cell(start_row, start_col).value = "Tracker / Scan"
    for s in range(4):
        ws_dash.cell(start_row, start_col + 1 + s).value = f"Scan{s}"
    if summary:
        mat = matrix_for(dataset, summary.file_index, default_block_no, tracker_count)
        for t, values in enumerate(mat):
            ws_dash.cell(start_row + 1 + t, start_col).value = f"T{t}"
            for s, value in enumerate(values):
                cell = ws_dash.cell(start_row + 1 + t, start_col + 1 + s)
                if value is not None:
                    cell.value = round(value, 1)
                    cell.font = Font(color="0046B4" if value >= SPEC_SNR else "BE0000")
    thin = Side(style="thin", color="999999")
    for r in range(start_row, start_row + tracker_count + 1):
        for c in range(start_col, start_col + 5):
            cell = ws_dash.cell(r, c)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if r == start_row or c == start_col:
                cell.font = Font(bold=True, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None)
                cell.fill = PatternFill("solid", fgColor="EBF2FA")

    # Chart data on Dashboard near M columns.
    chart_start_col = 13
    chart_start_row = 2
    ws_dash.cell(chart_start_row, chart_start_col).value = "Scan"
    selected_trackers = list(range(tracker_count))
    for idx, t in enumerate(selected_trackers, 1):
        ws_dash.cell(chart_start_row, chart_start_col + idx).value = f"T{t}"
    for s in range(4):
        ws_dash.cell(chart_start_row + 1 + s, chart_start_col).value = f"Scan{s}"
        for idx, t in enumerate(selected_trackers, 1):
            value = get_block_value(dataset, default_file_index, default_block_no, t, s, "SNR") if summary else None
            ws_dash.cell(chart_start_row + 1 + s, chart_start_col + idx).value = value
    if summary:
        chart = LineChart()
        chart.title = f"SNR by Tracker - Block {default_block_no}"
        chart.y_axis.title = "SNR"
        chart.x_axis.title = "Scan"
        data = Reference(ws_dash, min_col=chart_start_col + 1, max_col=chart_start_col + tracker_count,
                         min_row=chart_start_row, max_row=chart_start_row + 4)
        cats = Reference(ws_dash, min_col=chart_start_col, min_row=chart_start_row + 1, max_row=chart_start_row + 4)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws_dash.add_chart(chart, "A16")

    # sizes
    for ws in [ws_dash, ws_summary, ws_data, ws_lists]:
        for col in range(1, min(ws.max_column, 20) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
    ws_lists.sheet_state = "hidden"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def load_excel_dataset(xlsx_path: str | Path) -> Dataset:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if "Tracker_SNR" not in wb.sheetnames:
        raise ValueError("Tracker_SNR sheet not found")
    ws = wb["Tracker_SNR"]
    rows: list[DataRow] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values or values[0] is None:
            continue
        vals = list(values) + [None] * (len(DATA_HEADERS) - len(values))
        rows.append(DataRow(
            file_index=int(vals[0]), file_name=str(vals[1] or ""), time=str(vals[2] or ""),
            tracker_no=int(vals[3]), scan_no=int(vals[4]) if vals[4] is not None else None,
            block_no=int(vals[5]) if vals[5] is not None else None,
            signal=float(vals[6]) if vals[6] is not None and vals[6] != "" else None,
            noise=float(vals[7]) if vals[7] is not None and vals[7] != "" else None,
            snr=float(vals[8]),
            peak_loc=float(vals[9]) if vals[9] is not None and vals[9] != "" else None,
            peak_com=float(vals[10]) if vals[10] is not None and vals[10] != "" else None,
            peak_center=float(vals[11]) if vals[11] is not None and vals[11] != "" else None,
            source_type=str(vals[12] or ""), full_path=str(vals[13] or ""),
        ))
    summaries: list[FileSummary] = []
    for file_index in sorted({r.file_index for r in rows}):
        fr = [r for r in rows if r.file_index == file_index]
        trackers = detect_tracker_count({r.tracker_no for r in fr})
        blocks = max([r.block_no or 1 for r in fr], default=1)
        snrs = [r.snr for r in fr if positive(r.snr)]
        summaries.append(FileSummary(
            file_index=file_index, file_name=fr[0].file_name, full_path=fr[0].full_path,
            rows=len(fr), trackers=trackers, has_signal_noise=any(r.source_type.upper() == "SIGNAL" for r in fr),
            blocks=blocks, max_snr=max(snrs) if snrs else None, min_snr=min(snrs) if snrs else None,
        ))
    return Dataset(rows, summaries, [])


# ---------------- GUI ----------------

class SimpleLineChartWidget:
    def __new__(cls, *args, **kwargs):
        from PySide6.QtWidgets import QWidget, QToolTip
        from PySide6.QtCore import Qt, QPointF, QRectF
        from PySide6.QtGui import QPainter, QPen, QColor, QBrush
        class _Chart(QWidget):
            def __init__(self):
                super().__init__(); self.setMinimumHeight(190); self.setMouseTracking(True); self.title='No data'; self.metric_label='SNR'; self.series=[]; self.x_labels=[]; self.y_min=0.; self.y_max=1.; self._hit_points=[]
            def set_empty(self,title,metric='SNR'):
                self.title=title; self.metric_label=metric; self.series=[]; self.x_labels=[]; self._hit_points=[]; self.update()
            def set_chart(self,title,metric,series,x_labels=None):
                self.title=title; self.metric_label=metric; self.series=series; self.x_labels=x_labels or []
                vals=[v for ser in series for _,v in ser.get('points',[]) if v is not None]
                if vals:
                    lo,hi=min(vals),max(vals); m=max((hi-lo)*.1,hi*.03,1); self.y_min=max(0,lo-m); self.y_max=hi+m
                else: self.y_min,self.y_max=0,1
                self.update()
            def mouseMoveEvent(self, event):
                if not self._hit_points:
                    QToolTip.hideText(); return
                pos = event.position()
                nearest = min(self._hit_points, key=lambda item: (item[0].x()-pos.x())**2 + (item[0].y()-pos.y())**2)
                dist2 = (nearest[0].x()-pos.x())**2 + (nearest[0].y()-pos.y())**2
                if dist2 <= 144:
                    QToolTip.showText(event.globalPosition().toPoint(), nearest[1], self)
                else:
                    QToolTip.hideText()
            def leaveEvent(self, event):
                QToolTip.hideText(); super().leaveEvent(event)
            def paintEvent(self,event):
                self._hit_points=[]; p=QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing,True); w,h=self.width(),self.height(); p.fillRect(self.rect(),QColor('#f7f9fb'))
                l,r,t,b=58,88,32,40; pw,ph=max(50,w-l-r),max(50,h-t-b); n=max(1,len(self.x_labels))
                p.setPen(QPen(QColor('#17365d'))); f=p.font(); f.setPointSize(10); f.setBold(True); p.setFont(f); p.drawText(QRectF(0,3,w,23),Qt.AlignmentFlag.AlignCenter,self.title)
                p.setPen(QPen(QColor('#d7e0ea')))
                for i in range(5): y=t+ph*i/4; p.drawLine(l,int(y),l+pw,int(y))
                p.setPen(QPen(QColor('#4b5d70'))); p.drawLine(l,t,l,t+ph); p.drawLine(l,t+ph,l+pw,t+ph)
                f.setPointSize(8); f.setBold(False); p.setFont(f)
                for i in range(5):
                    val=self.y_max-(self.y_max-self.y_min)*i/4; y=t+ph*i/4; p.drawText(2,int(y)-8,l-8,16,Qt.AlignmentFlag.AlignRight,f'{val:.0f}')
                ticks=min(6,n)
                for i in range(ticks):
                    idx=round((n-1)*i/max(1,ticks-1)); x=l+pw*idx/max(1,n-1); label=self.x_labels[idx] if self.x_labels else ''
                    p.drawText(int(x)-42,t+ph+5,84,28,Qt.AlignmentFlag.AlignHCenter|Qt.AlignmentFlag.AlignTop,label)
                def mp(xv,val): return QPointF(l+pw*xv/max(1,n-1),t+ph*(1-(val-self.y_min)/(self.y_max-self.y_min)))
                for si,ser in enumerate(self.series):
                    c=QColor(ser.get('color','#1565c0')); p.setPen(QPen(c,2)); pts=[mp(x,v) for x,v in ser.get('points',[])]
                    for a,z in zip(pts,pts[1:]): p.drawLine(a,z)
                    p.setBrush(QBrush(c))
                    for (xv, val), pt in zip(ser.get('points',[]), pts):
                        p.drawEllipse(pt,3,3)
                        meta = ser.get('point_meta',{}).get(xv)
                        if meta: self._hit_points.append((pt, meta))
                    p.drawLine(w-r+5,t+si*18+6,w-r+22,t+si*18+6); p.setPen(QPen(QColor('#263746'))); p.drawText(w-r+26,t+si*18,58,16,Qt.AlignmentFlag.AlignLeft,ser.get('name',''))
                if not self.series: p.setPen(QPen(QColor('#7a8793'))); p.drawText(QRectF(l,t,pw,ph),Qt.AlignmentFlag.AlignCenter,self.title)
        return _Chart(*args,**kwargs)

try:
    from PySide6.QtCore import QObject
except ImportError:
    class QObject:
        """Fallback used only for non-GUI self-test environments."""
        def __init__(self, *args, **kwargs):
            super().__init__()


class GuidedTourOverlay:
    """Modal-like spotlight tour drawn over the Tracker SNR main window."""
    def __init__(self, owner, steps):
        from PySide6.QtCore import QObject, QPoint, QRect, Qt, QEvent
        from PySide6.QtGui import QColor, QPainter, QPainterPath, QPen
        from PySide6.QtWidgets import QWidget, QFrame, QVBoxLayout, QHBoxLayout, QLabel, QPushButton

        self.owner = owner
        self.steps = steps
        self.index = 0
        self.QPoint, self.QRect, self.Qt, self.QEvent = QPoint, QRect, Qt, QEvent
        self.QColor, self.QPainter, self.QPainterPath, self.QPen = QColor, QPainter, QPainterPath, QPen

        class OverlayWidget(QWidget):
            def paintEvent(inner, event):
                self._paint(inner)
            def keyPressEvent(inner, event):
                if event.key() == Qt.Key.Key_Escape:
                    self.finish()
                else:
                    super(OverlayWidget, inner).keyPressEvent(event)
            def mousePressEvent(inner, event):
                event.accept()

        self.overlay = OverlayWidget(owner.window)
        self.overlay.setObjectName('guidedTourOverlay')
        self.overlay.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.overlay.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.overlay.setGeometry(owner.window.rect())

        self.card = QFrame(self.overlay)
        self.card.setObjectName('tourCard')
        self.card.setStyleSheet(
            'QFrame#tourCard{background:white;border:1px solid #9fb3c7;border-radius:10px;}'
            'QLabel#tourStep{color:#2d5d88;font-weight:700;font-size:12px;}'
            'QLabel#tourTitle{color:#17365d;font-weight:700;font-size:17px;}'
            'QLabel#tourBody{color:#263746;font-size:13px;}'
            'QPushButton{background:#2d5d88;color:white;border:0;border-radius:4px;padding:6px 12px;}'
            'QPushButton:hover{background:#3e75a6;}'
        )
        cv=QVBoxLayout(self.card); cv.setContentsMargins(18,16,18,14); cv.setSpacing(8)
        self.step_label=QLabel(); self.step_label.setObjectName('tourStep'); cv.addWidget(self.step_label)
        self.title_label=QLabel(); self.title_label.setObjectName('tourTitle'); self.title_label.setWordWrap(True); cv.addWidget(self.title_label)
        self.body_label=QLabel(); self.body_label.setObjectName('tourBody'); self.body_label.setWordWrap(True); cv.addWidget(self.body_label)
        nav=QHBoxLayout(); self.skip=QPushButton('Skip Tour'); self.back=QPushButton('Back'); self.next=QPushButton('Next')
        nav.addWidget(self.skip); nav.addStretch(); nav.addWidget(self.back); nav.addWidget(self.next); cv.addLayout(nav)
        self.skip.clicked.connect(self.finish); self.back.clicked.connect(self.previous); self.next.clicked.connect(self.advance)
        self.card.resize(390,220)
        owner.window.installEventFilter(self.overlay)
        self.show_step()
        self.overlay.show(); self.overlay.raise_(); self.overlay.setFocus()

    def _target_rect(self):
        step=self.steps[self.index]
        target=step.get('target')
        if target is None or not target.isVisible():
            return self.QRect()
        pos=target.mapTo(self.owner.window, self.QPoint(0,0))
        rect=self.QRect(pos, target.size()).adjusted(-7,-7,7,7)
        return rect.intersected(self.overlay.rect().adjusted(5,5,-5,-5))

    def _paint(self, widget):
        painter=self.QPainter(widget)
        painter.setRenderHint(self.QPainter.RenderHint.Antialiasing, True)
        painter.fillRect(widget.rect(), self.QColor(8,24,41,185))
        rect=self._target_rect()
        if not rect.isNull():
            painter.setCompositionMode(self.QPainter.CompositionMode.CompositionMode_Clear)
            painter.drawRoundedRect(rect, 9, 9)
            painter.setCompositionMode(self.QPainter.CompositionMode.CompositionMode_SourceOver)
            painter.setPen(self.QPen(self.QColor(108,190,255), 3))
            painter.drawRoundedRect(rect, 9, 9)

    def _position_card(self):
        rect=self._target_rect(); ow=self.overlay.width(); oh=self.overlay.height()
        cw=min(410,max(330,ow//3)); ch=230
        self.card.resize(cw,ch)
        margin=18
        if rect.isNull():
            x=(ow-cw)//2; y=(oh-ch)//2
        else:
            below=rect.bottom()+18
            above=rect.top()-ch-18
            if below+ch <= oh-margin:
                y=below
            elif above >= margin:
                y=above
            else:
                y=max(margin,min(oh-ch-margin,rect.center().y()-ch//2))
            x=max(margin,min(ow-cw-margin,rect.center().x()-cw//2))
            if self.QRect(x,y,cw,ch).intersects(rect):
                x=margin if rect.center().x()>ow//2 else ow-cw-margin
        self.card.move(x,y)

    def show_step(self):
        if not self.steps:
            self.finish(); return
        step=self.steps[self.index]
        self.step_label.setText(f"STEP {self.index+1} OF {len(self.steps)}")
        self.title_label.setText(step['title'])
        self.body_label.setText(step['body'])
        self.back.setEnabled(self.index>0)
        self.next.setText('Finish' if self.index==len(self.steps)-1 else 'Next')
        self._position_card(); self.overlay.update(); self.card.raise_()

    def advance(self):
        if self.index>=len(self.steps)-1:
            self.finish(); return
        self.index+=1; self.show_step()

    def previous(self):
        if self.index>0:
            self.index-=1; self.show_step()

    def finish(self):
        try:
            self.owner.window.removeEventFilter(self.overlay)
        except Exception:
            pass
        self.overlay.close(); self.overlay.deleteLater()
        self.owner._active_tour=None


class TrackerApp(QObject):
    MODE_SINGLE = 'Single File'
    MODE_COMPARE = 'Comparison'

    def __init__(self):
        super().__init__()
        from PySide6.QtCore import Qt
        from PySide6.QtWidgets import (
            QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
            QComboBox, QTableWidget, QTableWidgetItem, QSplitter, QTextEdit,
            QMessageBox, QFileDialog, QGridLayout, QAbstractItemView, QFrame,
            QStackedWidget, QDialog, QDialogButtonBox, QCheckBox, QTextBrowser
        )
        from PySide6.QtGui import QFont
        self.Qt, self.QMessageBox, self.QFileDialog, self.QTableWidgetItem = Qt, QMessageBox, QFileDialog, QTableWidgetItem
        self.QDialog, self.QDialogButtonBox, self.QCheckBox, self.QTextBrowser = QDialog, QDialogButtonBox, QCheckBox, QTextBrowser
        self.user_settings = load_user_settings()
        self._allow_close = False
        self._active_tour = None

        owner = self
        class GuideAwareMainWindow(QMainWindow):
            def closeEvent(inner_self, event):
                owner.handle_close_event(event)

        self.window = GuideAwareMainWindow()
        self.window.setWindowTitle('Tracker SNR Analyzer - FUS')
        self.window.resize(1500, 940)
        self.window.setAcceptDrops(True)
        self.window.installEventFilter(self)
        self.dataset = Dataset([], [], [])
        self.current_excel_path = None
        self.temp_roots = []

        central = QWidget(); central.setAcceptDrops(True); central.installEventFilter(self)
        self.window.setCentralWidget(central)
        main = QVBoxLayout(central); main.setContentsMargins(8, 8, 8, 8)

        hdr = QFrame(); hdr.setObjectName('header'); h = QHBoxLayout(hdr)
        title = QLabel('Tracker SNR Analysis'); title.setObjectName('title'); h.addWidget(title)
        h.addStretch(); hint = QLabel('Drop files / folders / ZIP here'); hint.setObjectName('hint'); h.addWidget(hint)
        main.addWidget(hdr)

        top = QHBoxLayout(); main.addLayout(top)
        self.btn_import_files = QPushButton('Import Files / ZIP')
        self.btn_import_folder = QPushButton('Import Folder')
        self.btn_reset_data = QPushButton('Reset')
        self.btn_save_excel = QPushButton('Save Excel')
        self.btn_save_project = QPushButton('Save Project')
        self.btn_open_project = QPushButton('Open Project')
        self.btn_help = QPushButton('Guide / Tour')
        for b in [self.btn_import_files, self.btn_import_folder, self.btn_reset_data, self.btn_save_excel, self.btn_save_project, self.btn_open_project, self.btn_help]:
            top.addWidget(b)
        top.addStretch()
        top.addWidget(QLabel('Display Mode'))
        self.combo_mode = QComboBox(); self.combo_mode.addItems([self.MODE_SINGLE, self.MODE_COMPARE]); self.combo_mode.setEnabled(False); top.addWidget(self.combo_mode)
        top.addWidget(QLabel('Metric'))
        self.combo_metric = QComboBox(); self.combo_metric.addItems(['SNR', 'Signal', 'Noise']); top.addWidget(self.combo_metric)

        self.info_label = QLabel('No data loaded.'); main.addWidget(self.info_label)
        self.stack = QStackedWidget(); main.addWidget(self.stack, 1)

        # Legacy/single-file view
        self.single_page = QWidget(); single_layout = QVBoxLayout(self.single_page); single_layout.setContentsMargins(0,0,0,0)
        single_controls = QHBoxLayout(); single_layout.addLayout(single_controls)
        self.btn_prev = QPushButton('P'); self.btn_prev.setFixedWidth(38); single_controls.addWidget(self.btn_prev)
        single_controls.addWidget(QLabel('File'))
        self.combo_file = QComboBox(); self.combo_file.setMinimumWidth(330); single_controls.addWidget(self.combo_file, 1)
        self.btn_next = QPushButton('N'); self.btn_next.setFixedWidth(38); single_controls.addWidget(self.btn_next)
        single_controls.addSpacing(12); single_controls.addWidget(QLabel('Block'))
        self.combo_block = QComboBox(); self.combo_block.setMinimumWidth(110); single_controls.addWidget(self.combo_block)
        single_controls.addStretch()
        self.single_chart = SimpleLineChartWidget(); self.single_chart.setObjectName('chart'); single_layout.addWidget(self.single_chart, 1)
        single_lower = QSplitter(Qt.Orientation.Horizontal)
        self.single_matrix = QTableWidget(0, 4); self.single_matrix.setHorizontalHeaderLabels([f'Scan{i}' for i in range(4)]); single_lower.addWidget(self.single_matrix)
        self.single_index = QTableWidget(0, len(SUMMARY_HEADERS)); self.single_index.setHorizontalHeaderLabels(SUMMARY_HEADERS); self.single_index.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows); single_lower.addWidget(self.single_index)
        single_lower.setSizes([360, 1050]); single_layout.addWidget(single_lower, 0)
        self.stack.addWidget(self.single_page)

        # Multi-file comparison view
        self.compare_page = QWidget(); compare_layout = QVBoxLayout(self.compare_page); compare_layout.setContentsMargins(0,0,0,0)
        compare_controls = QHBoxLayout(); compare_layout.addLayout(compare_controls)
        compare_controls.addWidget(QLabel('Block Display'))
        self.combo_compare_blocks = QComboBox(); self.combo_compare_blocks.addItems(['All Blocks', 'File Block Median']); compare_controls.addWidget(self.combo_compare_blocks)
        compare_controls.addStretch()
        split = QSplitter(Qt.Orientation.Vertical); compare_layout.addWidget(split, 1)
        charts = QWidget(); grid = QGridLayout(charts); grid.setContentsMargins(0,0,0,0); self.plots = []
        for t in range(4):
            c = SimpleLineChartWidget(); c.setObjectName('chart'); self.plots.append(c); grid.addWidget(c, t//2, t%2)
        split.addWidget(charts)
        lower = QSplitter(Qt.Orientation.Horizontal)
        self.matrix = QTableWidget(0,4); self.matrix.setHorizontalHeaderLabels([f'Scan{i}' for i in range(4)]); self.matrix.setVisible(False); lower.addWidget(self.matrix)
        self.table = QTableWidget(0,len(SUMMARY_HEADERS)); self.table.setHorizontalHeaderLabels(SUMMARY_HEADERS); self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows); lower.addWidget(self.table)
        lower.setSizes([0,1440]); split.addWidget(lower); split.setSizes([680,210])
        self.stack.addWidget(self.compare_page)

        self.btn_toggle_log = QPushButton('Show Detail Log'); self.btn_toggle_log.setCheckable(True); main.addWidget(self.btn_toggle_log)
        self.log = QTextEdit(); self.log.setReadOnly(True); self.log.setVisible(False); self.log.setMaximumHeight(90); self.log.setFont(QFont('Consolas',8)); main.addWidget(self.log)

        self.window.setStyleSheet("QMainWindow,QWidget{background:#eef2f5;color:#23384d} QFrame#header{background:#17365d;border-radius:5px} QLabel#title{color:white;font-size:16pt;font-weight:600} QLabel#hint{color:#dce8f5} QPushButton{background:#2d5d88;color:white;border:0;border-radius:4px;padding:6px 11px} QPushButton:hover{background:#3e75a6} QComboBox{background:white;border:1px solid #9fb3c7;padding:4px} QTableWidget{background:white;gridline-color:#d8e1ea} QWidget#chart{border:1px solid #b9c9d8;border-radius:4px}")

        self.btn_import_files.clicked.connect(self.import_files)
        self.btn_import_folder.clicked.connect(self.import_folder)
        self.btn_reset_data.clicked.connect(self.reset_data)
        self.btn_save_excel.clicked.connect(self.save_excel_dialog)
        self.btn_save_project.clicked.connect(self.save_project_dialog)
        self.btn_open_project.clicked.connect(self.open_project_dialog)
        if RELEASE_MODE:
            self.btn_help.setVisible(False)
        else:
            self.btn_help.clicked.connect(self.show_guide_and_tour)
        self.combo_metric.currentIndexChanged.connect(self.refresh_all)
        self.combo_mode.currentIndexChanged.connect(self.on_mode_changed)
        self.combo_compare_blocks.currentIndexChanged.connect(self.refresh_compare)
        self.combo_file.currentIndexChanged.connect(self.on_single_file_changed)
        self.combo_block.currentIndexChanged.connect(self.refresh_single)
        self.btn_prev.clicked.connect(lambda: self.move_file(-1))
        self.btn_next.clicked.connect(lambda: self.move_file(1))
        self.btn_toggle_log.toggled.connect(self.toggle_detail_log)
        self.single_index.itemSelectionChanged.connect(self.on_single_index_selection)

    def eventFilter(self, obj, event):
        from PySide6.QtCore import QEvent
        if event.type() == QEvent.Type.DragEnter and event.mimeData().hasUrls():
            event.acceptProposedAction(); return True
        if event.type() == QEvent.Type.Drop:
            self.load_inputs([u.toLocalFile() for u in event.mimeData().urls() if u.toLocalFile()]); event.acceptProposedAction(); return True
        return False

    def show(self):
        self.window.show()
        # Run after the window is visible so guide dialogs are correctly parented.
        from PySide6.QtCore import QTimer
        if not RELEASE_MODE:
            QTimer.singleShot(0, self.handle_startup_guide)

    def handle_startup_guide(self):
        if RELEASE_MODE:
            return
        if self.user_settings.get("show_guide_once", False):
            self.user_settings["show_guide_once"] = False
            save_user_settings(self.user_settings)
            self.show_guide_and_tour()
            return
        if not self.user_settings.get("ask_startup_guide", True):
            return

        box = self.QMessageBox(self.window)
        box.setWindowTitle("Tracker SNR Quick Guide")
        box.setIcon(self.QMessageBox.Icon.Question)
        box.setText("Would you like to view the Quick Guide and Guided Tour?")
        box.setInformativeText("You can request the guide for the next startup from the exit confirmation window.")
        yes = box.addButton("Yes — Show Guide", self.QMessageBox.ButtonRole.AcceptRole)
        no = box.addButton("No — Do Not Ask Again", self.QMessageBox.ButtonRole.DestructiveRole)
        box.setDefaultButton(yes)
        box.exec()
        if box.clickedButton() is yes:
            self.show_guide_and_tour()
        elif box.clickedButton() is no:
            self.user_settings["ask_startup_guide"] = False
            save_user_settings(self.user_settings)

    def show_guide_and_tour(self):
        if RELEASE_MODE:
            return
        from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                                       QWidget, QStackedWidget, QFrame, QProgressBar)
        from PySide6.QtCore import QTimer, Qt
        dialog = QDialog(self.window)
        dialog.setWindowTitle("Tracker SNR Analyzer — Quick Guide")
        dialog.resize(900, 650)
        dialog.setMinimumSize(760, 540)
        dialog.setStyleSheet(
            "QDialog{background:#eef2f5;color:#263746;}"
            "QFrame#guideHeader{background:#17365d;border-radius:7px;}"
            "QLabel#guideTitle{color:white;font-size:20px;font-weight:700;}"
            "QLabel#guideSubtitle{color:#dce8f5;font-size:12px;}"
            "QFrame#guideCard{background:white;border:1px solid #c6d3df;border-radius:8px;}"
            "QLabel#pageHeading{font-size:19px;font-weight:700;color:#17365d;}"
            "QLabel#pageBody{font-size:13px;color:#263746;}"
            "QPushButton{background:#2d5d88;color:white;border:0;border-radius:4px;padding:7px 14px;}"
            "QPushButton:hover{background:#3e75a6;}"
        )
        root=QVBoxLayout(dialog); root.setContentsMargins(14,14,14,14); root.setSpacing(10)
        header=QFrame(); header.setObjectName('guideHeader'); hv=QVBoxLayout(header); hv.setContentsMargins(18,13,18,13)
        ht=QLabel('Tracker SNR Analyzer Quick Guide'); ht.setObjectName('guideTitle'); hv.addWidget(ht)
        hs=QLabel('Import, inspect, compare, and export Tracker Signal / Noise / SNR data.'); hs.setObjectName('guideSubtitle'); hv.addWidget(hs)
        root.addWidget(header)
        pages=QStackedWidget(); root.addWidget(pages,1)

        guide_pages=[
            ('Welcome', 'Tracker SNR Analyzer extracts Tracker, Scan, Block, Signal, Noise, and SNR values from supported log files.',
             '<b>Use this guide to learn the complete workflow.</b><br><br>Single File mode preserves detailed block-level inspection. Comparison mode aligns multiple files chronologically and separates Tracker 0–3 into four charts.'),
            ('Import Data', 'Load a log file, folder, or ZIP package.',
             '<b>Import Files / ZIP</b> accepts individual .log/.txt files and ZIP packages.<br><b>Import Folder</b> searches recursively up to five levels.<br><b>Drag & Drop</b> works anywhere on the main window.<br><br>One valid file opens Single File mode. Multiple files enable both Single File and Comparison modes.'),
            ('Single File Mode', 'Inspect one file and one block in detail.',
             'Use <b>P</b> and <b>N</b>, or the File list, to move between files. Select a Block from the block list. Change Metric between <b>SNR</b>, <b>Signal</b>, and <b>Noise</b>.<br><br>The lower-left matrix summarizes Scan 0–3 for each Tracker. The file index remains visible on the right.'),
            ('Comparison Mode', 'Compare multiple files on one chronological view.',
             'Tracker 0, 1, 2, and 3 are displayed as four independent charts. Each chart contains Scan 0–3 curves.<br><br><b>All Blocks</b> shows every detected block grouped by acquisition date. <b>File Block Median</b> shows one median point per file, Tracker, and Scan.'),
            ('Read the Charts', 'Use the metric selector and hover tooltips.',
             'Select <b>SNR</b>, <b>Signal</b>, or <b>Noise</b> from the Metric list. Move the pointer over a plotted point to display file name, block or median status, Tracker, Scan, current metric, SNR, Signal, and Noise.<br><br>Unavailable Signal or Noise fields are omitted instead of showing misleading zeros.'),
            ('Understand Blocks and Dates', 'Keep repeated measurements in the correct acquisition group.',
             'In All Blocks mode, blocks remain grouped under the date parsed from the source file name. Blocks are ordered chronologically within that date.<br><br>This prevents blocks from different acquisition dates from being mixed together during comparison.'),
            ('Save Results', 'Export analysis or preserve a working project.',
             '<b>Save Excel</b> exports detailed rows and file summaries. <b>Save Project</b> stores the current Tracker SNR dataset for later reopening. <b>Open Project</b> restores a saved project without re-parsing the original logs.'),
            ('Ready to Start', 'You now know the complete Tracker SNR workflow.',
             'Select <b>Start Guided Tour</b> to highlight the actual controls in the application. The tour can be skipped at any time with <b>Skip Tour</b> or the Esc key.<br><br>You can reopen this guide later using <b>Guide / Tour</b> on the main screen.')
        ]
        for idx,(heading,lead,body) in enumerate(guide_pages,1):
            card=QFrame(); card.setObjectName('guideCard'); cv=QVBoxLayout(card); cv.setContentsMargins(24,22,24,20); cv.setSpacing(13)
            ph=QLabel(f'{idx}. {heading}'); ph.setObjectName('pageHeading'); cv.addWidget(ph)
            lead_lbl=QLabel(lead); lead_lbl.setStyleSheet('font-size:15px;font-weight:600;color:#2d5d88;'); lead_lbl.setWordWrap(True); cv.addWidget(lead_lbl)
            body_lbl=QLabel(body); body_lbl.setObjectName('pageBody'); body_lbl.setTextFormat(Qt.TextFormat.RichText); body_lbl.setWordWrap(True); body_lbl.setAlignment(Qt.AlignmentFlag.AlignTop); cv.addWidget(body_lbl,1)
            pages.addWidget(card)

        footer=QVBoxLayout(); progress=QProgressBar(); progress.setRange(1,len(guide_pages)); progress.setTextVisible(True); footer.addWidget(progress)
        nav=QHBoxLayout(); back=QPushButton('Back'); skip=QPushButton('Close Guide'); nxt=QPushButton('Next'); tour=QPushButton('Start Guided Tour')
        nav.addWidget(back); nav.addWidget(skip); nav.addStretch(); nav.addWidget(nxt); nav.addWidget(tour); footer.addLayout(nav); root.addLayout(footer)
        start_tour={'value':False}
        def update_nav():
            i=pages.currentIndex(); progress.setValue(i+1); progress.setFormat(f'Page {i+1} of {pages.count()}')
            back.setEnabled(i>0); nxt.setVisible(i<pages.count()-1); tour.setVisible(i==pages.count()-1)
        back.clicked.connect(lambda: (pages.setCurrentIndex(max(0,pages.currentIndex()-1)),update_nav()))
        nxt.clicked.connect(lambda: (pages.setCurrentIndex(min(pages.count()-1,pages.currentIndex()+1)),update_nav()))
        skip.clicked.connect(dialog.reject)
        def accept_tour(): start_tour['value']=True; dialog.accept()
        tour.clicked.connect(accept_tour)
        update_nav()
        result=dialog.exec()
        if result==QDialog.DialogCode.Accepted and start_tour['value']:
            QTimer.singleShot(80,self.start_guided_tour)

    def start_guided_tour(self):
        if RELEASE_MODE:
            return
        if self._active_tour is not None:
            return
        # Show the comparison workspace so every spotlight target is visible.
        self.stack.setCurrentWidget(self.compare_page)
        if self.dataset.summaries and len(self.dataset.summaries)>1:
            self.combo_mode.setEnabled(True)
            self.combo_mode.blockSignals(True); self.combo_mode.setCurrentText(self.MODE_COMPARE); self.combo_mode.blockSignals(False)
        steps=[
            {'target':self.btn_import_files,'title':'Import Tracker Data','body':'Choose individual Tracker logs or a ZIP package. You can also drag files directly onto the application.'},
            {'target':self.btn_import_folder,'title':'Search a Folder','body':'Select a folder to search supported logs recursively, up to five directory levels.'},
            {'target':self.combo_mode,'title':'Choose the Display Mode','body':'Single File provides detailed block inspection. Comparison aligns multiple files chronologically. Comparison becomes available after loading multiple files.'},
            {'target':self.combo_metric,'title':'Select the Metric','body':'Switch every active chart between SNR, Signal, and Noise without reloading the data.'},
            {'target':self.combo_compare_blocks,'title':'Choose the Block Method','body':'In Comparison mode, show every block or calculate one median point for each file, Tracker, and Scan.'},
            {'target':self.plots[0],'title':'Inspect Comparison Charts','body':'Tracker 0–3 use separate charts. Hover over any plotted point to see the source file, block, Scan, SNR, Signal, and Noise.'},
            {'target':self.btn_save_excel,'title':'Export the Results','body':'Save the extracted rows and file summaries to Excel. Save Project preserves the current dataset for later use.'},
            {'target':self.btn_help,'title':'Open Help Anytime','body':'Use Guide / Tour whenever you need to review this guide or replay the spotlight tour.'}
        ]
        self._active_tour=GuidedTourOverlay(self,steps)

    def handle_close_event(self, event):
        if RELEASE_MODE:
            event.accept()
            return
        if self._allow_close:
            event.accept(); return
        dialog = self.QDialog(self.window)
        dialog.setWindowTitle("Exit Tracker SNR Analyzer")
        dialog.setModal(True)
        layout = __import__('PySide6.QtWidgets', fromlist=['QVBoxLayout']).QVBoxLayout(dialog)
        label = __import__('PySide6.QtWidgets', fromlist=['QLabel']).QLabel("Exit Tracker SNR Analyzer?")
        label.setStyleSheet("font-size:15px;font-weight:600")
        layout.addWidget(label)
        check = self.QCheckBox("Show the guide and guided tour at the next startup")
        check.setChecked(False); layout.addWidget(check)
        buttons = self.QDialogButtonBox()
        cancel_button = buttons.addButton("Cancel", self.QDialogButtonBox.ButtonRole.RejectRole)
        exit_button = buttons.addButton("Exit", self.QDialogButtonBox.ButtonRole.AcceptRole)
        cancel_button.clicked.connect(dialog.reject)
        exit_button.clicked.connect(dialog.accept)
        layout.addWidget(buttons)
        if dialog.exec() == self.QDialog.DialogCode.Accepted:
            if check.isChecked():
                self.user_settings["show_guide_once"] = True
                save_user_settings(self.user_settings)
            self._allow_close = True
            for root in self.temp_roots:
                shutil.rmtree(root, ignore_errors=True)
            event.accept()
        else:
            event.ignore()

    def append_log(self, msg):
        try: print(msg, flush=True)
        except Exception: pass
        self.log.append(msg); write_log(msg)
    def message(self, title, text): self.QMessageBox.information(self.window, title, text)
    def process_events(self):
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

    def import_files(self):
        paths, _ = self.QFileDialog.getOpenFileNames(self.window, 'Select Tracker logs or ZIP', '', 'Tracker inputs (*.log *.txt *.zip);;All files (*.*)')
        if paths: self.load_inputs(paths)
    def import_folder(self):
        folder = self.QFileDialog.getExistingDirectory(self.window, 'Select folder (searches up to 5 levels)')
        if folder: self.load_inputs([folder])

    def load_inputs(self, inputs):
        self.info_label.setText('Discovering Tracker SNR files...'); self.process_events()
        paths = collect_inputs(inputs, self.temp_roots, 5)
        self.append_log(f'Discovered {len(paths)} target log(s).')
        if not paths:
            self.message('Tracker SNR', 'No .log or .txt files were found.'); return
        self.dataset = parse_files(paths, self.append_log)
        self.update_ui_from_dataset()
        count = len(self.dataset.summaries)
        if count <= 1:
            self.combo_mode.blockSignals(True); self.combo_mode.setCurrentText(self.MODE_SINGLE); self.combo_mode.blockSignals(False)
            self.combo_mode.setEnabled(False)
        else:
            self.combo_mode.setEnabled(True)
            # Multiple inputs default to comparison, while still allowing manual switching.
            self.combo_mode.blockSignals(True); self.combo_mode.setCurrentText(self.MODE_COMPARE); self.combo_mode.blockSignals(False)
        self.on_mode_changed()
        mode_text = 'Single-file view' if count <= 1 else 'Comparison view; Single File and Comparison modes are available.'
        self.info_label.setText(f'Loaded {count} files / {len(self.dataset.rows)} rows / skipped {len(self.dataset.skipped)}. {mode_text}')

    def reset_data(self):
        self.dataset = Dataset([], [], [])
        for table in [self.table, self.matrix, self.single_matrix, self.single_index]: table.setRowCount(0)
        self.combo_file.clear(); self.combo_block.clear(); self.combo_mode.setEnabled(False)
        for p in self.plots: p.set_empty('No data')
        self.single_chart.set_empty('No data')
        for root in self.temp_roots: shutil.rmtree(root, ignore_errors=True)
        self.temp_roots.clear(); self.info_label.setText('No data loaded.')

    def update_ui_from_dataset(self):
        for table in [self.table, self.single_index]:
            table.setRowCount(len(self.dataset.summaries))
            for r, summary in enumerate(self.dataset.summaries):
                for c, value in enumerate(summary.as_excel_row()):
                    table.setItem(r, c, self.QTableWidgetItem('' if value is None else str(value)))
            table.resizeColumnsToContents()
        self.combo_file.blockSignals(True); self.combo_file.clear()
        for summary in self.dataset.summaries: self.combo_file.addItem(summary.file_name, summary.file_index)
        if self.combo_file.count(): self.combo_file.setCurrentIndex(self.combo_file.count()-1)
        self.combo_file.blockSignals(False)
        self.on_single_file_changed()
        self.update_matrix(); self.refresh_compare()

    def on_mode_changed(self, *args):
        compare = self.combo_mode.currentText() == self.MODE_COMPARE and len(self.dataset.summaries) > 1
        self.stack.setCurrentWidget(self.compare_page if compare else self.single_page)
        if compare: self.refresh_compare()
        else: self.refresh_single()

    def on_single_file_changed(self, *args):
        file_index = self.combo_file.currentData()
        summary = next((s for s in self.dataset.summaries if s.file_index == file_index), None)
        self.combo_block.blockSignals(True); self.combo_block.clear()
        if summary:
            for block in range(1, max(1, summary.blocks)+1): self.combo_block.addItem(f'Block {block}', block)
            self.combo_block.setCurrentIndex(max(0, self.combo_block.count()-1))
            row = self.combo_file.currentIndex()
            if 0 <= row < self.single_index.rowCount(): self.single_index.selectRow(row)
        self.combo_block.blockSignals(False)
        self.refresh_single()

    def move_file(self, delta):
        if not self.combo_file.count(): return
        self.combo_file.setCurrentIndex(max(0, min(self.combo_file.count()-1, self.combo_file.currentIndex()+delta)))

    def on_single_index_selection(self):
        rows = self.single_index.selectionModel().selectedRows()
        if rows and rows[0].row() != self.combo_file.currentIndex(): self.combo_file.setCurrentIndex(rows[0].row())

    def update_matrix(self):
        self.matrix.setRowCount(4); self.matrix.setVerticalHeaderLabels([f'T{i}' for i in range(4)])
        if not self.dataset.summaries: return
        latest = self.dataset.summaries[-1]
        self.fill_matrix(self.matrix, latest.file_index, max(1, latest.blocks))

    def fill_matrix(self, table, file_index, block):
        from PySide6.QtGui import QColor
        table.setRowCount(4); table.setVerticalHeaderLabels([f'T{i}' for i in range(4)])
        for t in range(4):
            for sc in range(4):
                value = get_block_value(self.dataset, file_index, block, t, sc, 'SNR')
                item = self.QTableWidgetItem('' if value is None else f'{value:.1f}')
                if value is not None: item.setForeground(QColor('#1455a0') if value >= SPEC_SNR else QColor('#c62828'))
                table.setItem(t, sc, item)
        table.resizeColumnsToContents()

    def refresh_all(self, *args):
        if self.combo_mode.currentText() == self.MODE_COMPARE and len(self.dataset.summaries) > 1: self.refresh_compare()
        else: self.refresh_single()

    def refresh_single(self, *args):
        metric = self.combo_metric.currentText() or 'SNR'
        file_index = self.combo_file.currentData(); block = self.combo_block.currentData()
        if file_index is None or block is None:
            self.single_chart.set_empty('No data', metric); self.single_matrix.setRowCount(0); return
        colors = ['#1565c0','#ef6c00','#2e7d32','#c62828','#6a1b9a','#00838f']
        series = []
        trackers = sorted({r.tracker_no for r in self.dataset.rows if r.file_index == file_index and r.block_no == block})
        for tracker in trackers:
            points = []
            for scan in range(4):
                value = get_block_value(self.dataset, file_index, block, tracker, scan, metric)
                if value is not None: points.append((scan, value))
            if points: series.append({'name': f'T{tracker}', 'color': colors[tracker % len(colors)], 'points': points})
        summary = next((s for s in self.dataset.summaries if s.file_index == file_index), None)
        title = f'{summary.file_name if summary else "Selected file"} - Block {block} - {metric}'
        self.single_chart.set_chart(title, metric, series, [f'Scan {i}' for i in range(4)])
        self.fill_matrix(self.single_matrix, file_index, block)

    def build_comparison_points(self):
        """Build chronological comparison samples.

        All Blocks: one x point for every block, grouped by acquisition date.
        File Block Median: one x point per file; each Tracker/Scan value is the
        median of all valid block values in that file.
        """
        grouped = {}
        undated = []
        for summary in self.dataset.summaries:
            dt = parse_log_stamp_from_name(summary.file_name)
            if dt is None: undated.append((dt, summary))
            else: grouped.setdefault(dt.date(), []).append((dt, summary))
        use_median = self.combo_compare_blocks.currentText() == 'File Block Median'
        samples = []
        for day in sorted(grouped):
            first_in_day = True
            for dt, summary in sorted(grouped[day], key=lambda item: item[0]):
                if use_median:
                    prefix = dt.strftime('%m/%d') if first_in_day else dt.strftime('%H:%M')
                    samples.append({'summary':summary,'block':None,'label':prefix,'day':day,'median':True})
                    first_in_day = False
                else:
                    for block in range(1, max(1, summary.blocks)+1):
                        prefix = dt.strftime('%m/%d') if first_in_day else ''
                        label = f'{prefix}\nB{block}' if prefix else f'B{block}'
                        samples.append({'summary':summary,'block':block,'label':label,'day':day,'median':False})
                        first_in_day = False
        for _, summary in undated:
            if use_median:
                samples.append({'summary':summary,'block':None,'label':Path(summary.file_name).stem[-8:],'day':None,'median':True})
            else:
                for block in range(1, max(1, summary.blocks)+1):
                    samples.append({'summary':summary,'block':block,'label':f'B{block}','day':None,'median':False})
        return samples

    def comparison_values(self, sample, tracker, scan):
        summary = sample['summary']
        if not sample['median']:
            block = sample['block']
            return {m:get_block_value(self.dataset, summary.file_index, block, tracker, scan, m) for m in ('SNR','Signal','Noise')}
        import statistics
        result = {}
        for metric in ('SNR','Signal','Noise'):
            vals=[]
            for block in range(1, max(1, summary.blocks)+1):
                v=get_block_value(self.dataset, summary.file_index, block, tracker, scan, metric)
                if v is not None: vals.append(v)
            result[metric]=statistics.median(vals) if vals else None
        return result

    def refresh_compare(self, *args):
        metric = self.combo_metric.currentText() or 'SNR'
        colors = ['#1565c0','#ef6c00','#2e7d32','#c62828']
        samples = self.build_comparison_points()
        labels = [sample['label'] for sample in samples]
        date_count = len({sample['day'] for sample in samples if sample['day'] is not None})
        for tracker, plot in enumerate(self.plots):
            series=[]
            for scan in range(4):
                points=[]; point_meta={}
                for x, sample in enumerate(samples):
                    vals=self.comparison_values(sample, tracker, scan)
                    value=vals.get(metric)
                    if value is None: continue
                    points.append((x,value))
                    summary=sample['summary']
                    scope='Median of file blocks' if sample['median'] else f"Block {sample['block']}"
                    lines=[summary.file_name, scope, f'Tracker {tracker} / Scan {scan}', f'{metric}: {value:.3f}']
                    if metric != 'SNR' and vals.get('SNR') is not None: lines.append(f"SNR: {vals['SNR']:.3f}")
                    if vals.get('Signal') is not None: lines.append(f"Signal: {vals['Signal']:.3f}")
                    if vals.get('Noise') is not None: lines.append(f"Noise: {vals['Noise']:.3f}")
                    point_meta[x]='\n'.join(lines)
                if points: series.append({'name':f'Scan {scan}','color':colors[scan],'points':points,'point_meta':point_meta})
            mode=self.combo_compare_blocks.currentText()
            suffix=f' - {date_count} date group(s) - {mode}' if date_count else f' - {mode}'
            plot.set_chart(f'Tracker {tracker} - {metric}{suffix}', metric, series, labels)

    def toggle_detail_log(self, checked): self.log.setVisible(checked); self.btn_toggle_log.setText('Hide Detail Log' if checked else 'Show Detail Log')
    def default_excel_path(self):
        folder = Path(self.dataset.summaries[0].full_path).parent if self.dataset.summaries else app_dir()
        stem = 'Tracker_SNR_Comparison' if len(self.dataset.summaries) > 1 else 'Tracker_SNR'
        return folder / f'{stem}_{datetime.now():%Y%m%d_%H%M%S}.xlsx'
    def save_excel_dialog(self):
        if not self.dataset.rows: return
        path, _ = self.QFileDialog.getSaveFileName(self.window, 'Save Excel', str(self.default_excel_path()), 'Excel Workbook (*.xlsx)')
        if path: save_excel(self.dataset, path if path.lower().endswith('.xlsx') else path+'.xlsx')
    def save_project_dialog(self):
        if not self.dataset.rows: return
        path, _ = self.QFileDialog.getSaveFileName(self.window, 'Save project', 'TrackerSNR_Project.tsnrproj', 'Tracker SNR Project (*.tsnrproj)')
        if path: Path(path if path.endswith('.tsnrproj') else path+'.tsnrproj').write_text(json.dumps(self.dataset.to_project_dict(), ensure_ascii=False, indent=2), encoding='utf-8')
    def open_project_dialog(self):
        path, _ = self.QFileDialog.getOpenFileName(self.window, 'Open project', '', 'Tracker SNR Project (*.tsnrproj)')
        if path:
            self.dataset = Dataset.from_project_dict(json.loads(Path(path).read_text(encoding='utf-8')))
            self.update_ui_from_dataset()
            count = len(self.dataset.summaries); self.combo_mode.setEnabled(count > 1)
            self.combo_mode.setCurrentText(self.MODE_COMPARE if count > 1 else self.MODE_SINGLE)
            self.on_mode_changed()


def run_gui() -> int:
    try:
        from PySide6.QtWidgets import QApplication, QMessageBox
    except Exception as exc:
        handle_exception("PySide6 import failed", exc)
        print("PySide6 import failed. Install requirements first: pip install -r requirements.txt")
        print(exc)
        return 1
    from PySide6.QtCore import QTimer
    from insightec_handoff import load_handoff
    handoff = load_handoff("tracker_snr")
    app = QApplication(sys.argv)
    try:
        tracker_app = TrackerApp()
        tracker_app.show()
        if handoff and handoff.auto_load:
            inputs = [str(path) for path in handoff.input_paths()]
            if not inputs and handoff.workspace():
                inputs = [str(handoff.workspace())]
            handoff.mark("tracker_snr", "accepted", input_count=len(inputs))
            if inputs:
                QTimer.singleShot(0, lambda values=inputs: tracker_app.load_inputs(values))
        return app.exec()
    except Exception as exc:
        handle_exception("GUI startup failed", exc)
        try:
            QMessageBox.critical(None, "Tracker SNR startup error", f"Startup failed. See TrackerSNR_runtime.log.\n\n{exc}")
        except Exception:
            print(traceback.format_exc())
        return 1


def run_self_test() -> int:
    base = app_dir() / "test_logs"
    base.mkdir(exist_ok=True)
    log1 = base / "2025_Jan_22_22_05_31.Log"
    log2 = base / "2025_Jan_23_22_05_31.Log"
    nodata = base / "2025_Jan_24_22_05_31.Log"
    def make_log(offset: int) -> str:
        lines = []
        for block in range(2):
            for tracker in range(6):
                for scan in range(4):
                    signal = 1000 + offset + block * 100 + tracker * 10 + scan
                    noise = 10 + tracker + scan / 10
                    snr = signal / noise
                    lines.append(f"12:00:{tracker}{scan}.000 SNR: tracker {tracker} signal: {signal:.2f} noise: {noise:.2f} SNR: {snr:.2f}")
                    lines.append(f"12:01:{tracker}{scan}.000 TRACKER PEAK LOG tracker {tracker} peakLocation: {scan} peakCOM: {scan+0.1:.1f} peakCenter: {scan+0.2:.1f} SNR: {snr+5:.2f}")
        return "\n".join(lines)
    log1.write_text(make_log(0), encoding="utf-8")
    log2.write_text(make_log(50), encoding="utf-8")
    nodata.write_text("this file has no target SNR rows\n", encoding="utf-8")
    paths = collect_target_logs(base)
    ds = parse_files(paths, print)
    assert len(ds.summaries) == 2, f"expected 2 valid summaries, got {len(ds.summaries)}"
    assert len(ds.skipped) == 1, f"expected 1 skipped, got {len(ds.skipped)}"
    assert ds.summaries[0].trackers == 6, ds.summaries[0]
    assert ds.summaries[0].blocks >= 2, ds.summaries[0]
    assert get_block_value(ds, ds.summaries[0].file_index, 1, 0, 0, "SNR") is not None
    out = base / "self_test_output.xlsx"
    save_excel(ds, out)
    ds2 = load_excel_dataset(out)
    assert len(ds2.rows) == len(ds.rows), "Excel reload row count mismatch"
    proj = base / "self_test_project.tsnrproj"
    proj.write_text(json.dumps(ds.to_project_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    ds3 = Dataset.from_project_dict(json.loads(proj.read_text(encoding="utf-8")))
    assert len(ds3.summaries) == len(ds.summaries), "Project reload mismatch"
    print("SELF TEST OK")
    print(f"Valid files: {len(ds.summaries)}, rows: {len(ds.rows)}, skipped: {len(ds.skipped)}")
    print(f"Excel: {out}")
    return 0


def main() -> int:
    write_log(f"Starting {APP_NAME}; argv={sys.argv}")
    if "--self-test" in sys.argv:
        return run_self_test()
    return run_gui()


if __name__ == "__main__":
    raise SystemExit(main())
